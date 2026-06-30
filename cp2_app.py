"""
CP2 Maintenance System - SAL DE ORO · POSCO Argentina
"""
import streamlit as st
import sqlite3, hashlib, pandas as pd, plotly.express as px
from pathlib import Path
from datetime import datetime, date, timedelta
import random, string as sm

DB_PATH = Path(__file__).parent / "cp2_planta.db"

st.set_page_config(
    page_title="CP2 · SAL DE ORO",
    page_icon="⚙",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif !important}
#MainMenu,footer,header,.stDeployButton,[data-testid="stToolbar"],
[data-testid="stSidebar"],[data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"]{display:none !important}

.stApp,[data-testid="stAppViewContainer"],[data-testid="stMain"]{background:#0f0f10 !important}
.block-container{padding:0 0 2rem 0 !important;max-width:100% !important}
/* Contenido centrado con margen */
section[data-testid="stMain"] > div > div > div > div {max-width:1400px;margin:0 auto}

input,textarea,[data-baseweb="input"] input,[data-baseweb="textarea"] textarea{
    background:#1a1a1e !important;border-color:#333 !important;
    color:#ffffff !important;caret-color:#fff !important}
input::placeholder,textarea::placeholder{color:#555 !important}
.stTextInput label,.stTextArea label,.stSelectbox label,
.stNumberInput label,.stMultiSelect label{color:#999 !important;font-size:12px !important}
[data-baseweb="select"] div{background:#1a1a1e !important;color:#fff !important;border-color:#333 !important}
[data-baseweb="option"]{background:#1a1a1e !important;color:#fff !important}
[data-baseweb="option"]:hover{background:#222 !important}
[data-baseweb="tag"]{background:#222 !important;border:1px solid #444 !important;
    border-radius:4px !important;max-width:none !important}
[data-baseweb="tag"] span{color:#fff !important;font-size:12px !important;white-space:nowrap !important}

[data-testid="metric-container"]{
    background:#111113 !important;border:1px solid #222 !important;
    border-radius:8px !important;padding:16px !important}
[data-testid="metric-container"] label{color:#aaa !important;font-size:11px !important;text-transform:uppercase;letter-spacing:0.05em !important}
[data-testid="stMetricValue"]{color:#fff !important;font-size:24px !important;font-weight:600 !important}
[data-testid="stMetricDelta"]{color:#aaa !important;font-size:11px !important}

.stButton>button{background:#1a1a1e !important;border:1px solid #333 !important;
    color:#e2e2e5 !important;border-radius:6px !important;font-size:13px !important}
.stButton>button:hover{background:#222 !important;border-color:#555 !important;color:#fff !important}
.stButton>button[kind="primary"]{background:#2563eb !important;border-color:#2563eb !important;color:#fff !important}
.stButton>button[kind="primary"]:hover{background:#1d4ed8 !important}

.streamlit-expanderHeader{background:#111113 !important;border:1px solid #222 !important;
    border-radius:6px !important;color:#e2e2e5 !important;font-size:13px !important}
.streamlit-expanderContent{background:#0f0f10 !important;border:1px solid #222 !important;border-top:none !important}

.stTabs [data-baseweb="tab-list"]{background:transparent !important;border-bottom:1px solid #222 !important}
.stTabs [data-baseweb="tab"]{background:transparent !important;color:#555 !important;
    font-size:13px !important;border:none !important;border-bottom:2px solid transparent !important}
.stTabs [aria-selected="true"]{color:#fff !important;border-bottom-color:#2563eb !important}
.stTabs [data-baseweb="tab-panel"]{background:transparent !important;padding-top:16px !important}

[data-testid="stDataFrame"]{border:1px solid #222 !important;border-radius:6px !important}
hr{border-color:#222 !important}
p,span,.stMarkdown p{color:#cccccc !important}
h1,h2,h3,h4{color:#ffffff !important}
code{background:#1a1a1e !important;color:#7dd3fc !important;border:1px solid #222 !important}
[data-testid="stVerticalBlockBorderWrapper"]{border:none !important;background:transparent !important}
</style>
""", unsafe_allow_html=True)

def hp(p): return hashlib.sha256(p.encode()).hexdigest()

# ── DATABASE: Supabase (cloud) o SQLite (local) ──────────────────────────────
import os as _os

_SUPABASE_URL = _os.environ.get("SUPABASE_URL", "")
_SUPABASE_KEY = _os.environ.get("SUPABASE_KEY", "")
_SUPABASE_DB_PASS = _os.environ.get("SUPABASE_DB_PASSWORD", "")
_USE_SUPABASE = bool(_SUPABASE_URL and _SUPABASE_DB_PASS)

if _USE_SUPABASE:
    import psycopg2 as _psycopg2

    def _pg_dsn():
        project_id = _SUPABASE_URL.replace("https://","").replace(".supabase.co","")
        return (f"postgresql://postgres:{_SUPABASE_DB_PASS}"
                f"@db.{project_id}.supabase.co:5432/postgres"
                f"?sslmode=require")

    def get_conn():
        return _psycopg2.connect(_pg_dsn())

    def qdf(sql, params=()):
        pg_sql = sql.replace("?", "%s")
        with _psycopg2.connect(_pg_dsn()) as conn:
            return pd.read_sql_query(pg_sql, conn, params=list(params))

    def run(sql, params=()):
        pg_sql = sql.replace("?", "%s")
        with _psycopg2.connect(_pg_dsn()) as conn:
            with conn.cursor() as cur:
                cur.execute(pg_sql, list(params))
            conn.commit()

else:
    def get_conn():
        if not DB_PATH.exists():
            st.error(f"Base de datos no encontrada: {DB_PATH}")
            st.stop()
        c = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        c.row_factory = sqlite3.Row
        return c

    def qdf(sql, params=()):
        return pd.read_sql_query(sql, get_conn(), params=params)

    def run(sql, params=()):
        c = get_conn(); c.execute(sql, params); c.commit()

def migrate():
    if _USE_SUPABASE:
        return  # PostgreSQL - tables already created via schema.sql
    cur = get_conn().cursor()
    cols = [r[1] for r in cur.execute("PRAGMA table_info(usuarios)").fetchall()]
    for col,typ in [("dni","TEXT"),("telefono","TEXT"),("puesto","TEXT"),("area_trabajo","TEXT")]:
        if col not in cols:
            cur.execute(f"ALTER TABLE usuarios ADD COLUMN {col} {typ}")
    get_conn().commit()

def seed_supabase():
    """Carga datos iniciales en Supabase si está vacío"""
    if not _USE_SUPABASE:
        return
    try:
        # Check if already seeded
        check = qdf("SELECT COUNT(*) as n FROM equipos")
        if check.iloc[0]["n"] > 0:
            return  # Already has data
        
        # Load seed data
        import gzip, json
        from pathlib import Path
        seed_file = Path(__file__).parent / "seed_data.json.gz"
        if not seed_file.exists():
            st.warning("seed_data.json.gz no encontrado — la DB estará vacía")
            return
        
        with gzip.open(seed_file, 'rt', encoding='utf-8') as f:
            all_data = json.load(f)
        
        progress = st.progress(0, text="Cargando datos iniciales en la base de datos...")
        tables = list(all_data.keys())
        
        for i, table in enumerate(tables):
            cols = all_data[table]["cols"]
            rows = all_data[table]["rows"]
            if not rows:
                continue
            
            # Skip auto-increment id column for insertion
            # Insert in batches of 50
            cols_str = ", ".join(cols)
            placeholders = ", ".join(["%s"] * len(cols))
            sql = f"INSERT INTO {table} ({cols_str}) VALUES ({placeholders}) ON CONFLICT DO NOTHING"
            
            conn = get_conn()
            with conn.cursor() as cur:
                for j in range(0, len(rows), 50):
                    batch = rows[j:j+50]
                    cur.executemany(sql, batch)
            conn.commit()
            conn.close()
            
            pct = (i+1)/len(tables)
            progress.progress(pct, text=f"Cargando {table} ({len(rows)} registros)...")
        
        progress.progress(1.0, text="✅ Datos cargados correctamente")
        import time; time.sleep(1)
        progress.empty()
        st.rerun()
        
    except Exception as e:
        st.warning(f"Error cargando datos iniciales: {e}")

migrate()
if _USE_SUPABASE:
    seed_supabase()

def safe_str(val, default=""):
    if val is None: return default
    s = str(val)
    return default if s in ("nan","None","NaT","") else s

def get_config(clave, default=""):
    try:
        r = qdf("SELECT valor FROM configuracion WHERE clave=?", (clave,))
        return r.iloc[0]["valor"] if len(r)>0 else default
    except:
        return default

def enviar_alerta_email(tag, mensaje, prioridad, tipo):
    """Envia email de alerta si hay configuracion SMTP"""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        host   = get_config("smtp_host")
        port   = int(get_config("smtp_port","587"))
        user   = get_config("smtp_user")
        pwd    = get_config("smtp_pass")
        dest   = get_config("email_alertas")

        if not all([host,user,pwd,dest]):
            return False

        msg = MIMEMultipart()
        msg["From"]    = user
        msg["To"]      = dest
        msg["Subject"] = f"[CP2 ALERTA {prioridad.upper()}] {tag}"
        body = f"""
        <h3>Alerta de Mantenimiento CP2 - SAL DE ORO</h3>
        <p><b>Equipo:</b> {tag}</p>
        <p><b>Tipo:</b> {tipo}</p>
        <p><b>Prioridad:</b> {prioridad.upper()}</p>
        <p><b>Mensaje:</b> {mensaje}</p>
        <p><b>Fecha:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
        <hr>
        <small>CP2 Maintenance System - POSCO Argentina</small>
        """
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(host, port) as server:
            server.starttls()
            server.login(user, pwd)
            server.send_message(msg)
        return True
    except:
        return False

def generar_html_imprimible(tipo, datos):
    """Genera HTML imprimible para OT, Plan o Parada"""
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    filas = "".join(f"<tr><td><b>{k}</b></td><td>{v}</td></tr>"
                    for k,v in datos.items() if v and str(v).strip())
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>CP2 - {tipo}</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:12px;margin:20px}}
  h2{{color:#003087;border-bottom:2px solid #003087;padding-bottom:6px}}
  table{{width:100%;border-collapse:collapse;margin-top:12px}}
  td{{padding:6px 10px;border:1px solid #ddd;vertical-align:top}}
  td:first-child{{width:30%;background:#f5f5f5;font-weight:bold}}
  .header{{display:flex;justify-content:space-between;align-items:center;margin-bottom:16px}}
  .logo{{font-size:18px;font-weight:bold;color:#003087}}
  .fecha{{font-size:11px;color:#666}}
  @media print{{button{{display:none}}}}
</style>
</head><body>
<div class="header">
  <div class="logo">POSCO Argentina · CP2 Maintenance System</div>
  <div class="fecha">SAL DE ORO · {ahora}</div>
</div>
<h2>{tipo}</h2>
<table>{filas}</table>
<br><br>
<table style="border:none">
  <tr>
    <td style="border:none;width:40%;border-top:1px solid #000;text-align:center">Firma Técnico</td>
    <td style="border:none;width:20%"></td>
    <td style="border:none;width:40%;border-top:1px solid #000;text-align:center">Firma Supervisor</td>
  </tr>
</table>
<script>window.print()</script>
</body></html>"""

migrate()

# Restaurar sesion desde query params
if not st.session_state.get("logged_in"):
    uid = st.query_params.get("uid")
    if uid:
        try:
            u = qdf("SELECT * FROM usuarios WHERE id=? AND activo=1",(int(uid),))
            if len(u)>0:
                r = u.iloc[0]
                st.session_state.update({
                    "logged_in":True,"user_id":int(r["id"]),
                    "username":r["username"],"nombre":r["nombre"],
                    "apellido":r["apellido"],"rol":r["rol"],
                    "lang":r.get("idioma","es") or "es",
                })
        except: pass

T = {
    "dashboard":       {"es":"Dashboard",             "en":"Dashboard",           "ko":"대시보드"},
    "equipos":         {"es":"Equipos",               "en":"Equipment",           "ko":"장비"},
    "ots":             {"es":"Órdenes de trabajo",    "en":"Work orders",         "ko":"작업 지시"},
    "plan":            {"es":"Plan preventivo",       "en":"Maintenance plan",    "ko":"예방 계획"},
    "alertas":         {"es":"Alertas",               "en":"Alerts",              "ko":"알림"},
    "paradas":         {"es":"Paradas",               "en":"Shutdowns",           "ko":"정지"},
    "historial":       {"es":"Historial",             "en":"History",             "ko":"이력"},
    "usuarios":        {"es":"Usuarios",              "en":"Users",               "ko":"사용자"},
    "salir":           {"es":"Salir",                 "en":"Sign out",            "ko":"로그아웃"},
    "buscar":          {"es":"Buscar equipo por tag, nombre, fabricante, área...",
                        "en":"Search equipment by tag, name, manufacturer, area...",
                        "ko":"태그, 이름, 제조사, 구역으로 장비 검색..."},
    "sin_resultados":  {"es":"Sin resultados",        "en":"No results",          "ko":"결과 없음"},
    "resultados":      {"es":"resultados",            "en":"results",             "ko":"결과"},
    "todas_areas":     {"es":"Todas las áreas",       "en":"All areas",           "ko":"모든 구역"},
    "todos_tipos":     {"es":"Todos los tipos",       "en":"All types",           "ko":"모든 유형"},
    "todos":           {"es":"Todos",                 "en":"All",                 "ko":"전체"},
    "con_motor":       {"es":"Con motor",             "en":"With motor",          "ko":"모터 있음"},
    "con_specs":       {"es":"Con specs",             "en":"With specs",          "ko":"사양 있음"},
    "con_instr":       {"es":"Con instrumentos",      "en":"With instruments",    "ko":"계기 있음"},
    "con_ds":          {"es":"Con datasheet",         "en":"With datasheet",      "ko":"데이터시트 있음"},
    "con_fab":         {"es":"Con fabricante",        "en":"With manufacturer",   "ko":"제조사 있음"},
    "selecciona_eq":   {"es":"Seleccioná un equipo",  "en":"Select equipment",    "ko":"장비를 선택하세요"},
    "crear_ot":        {"es":"Crear orden de trabajo","en":"Create work order",   "ko":"작업 지시 생성"},
    "nueva_ot":        {"es":"Nueva orden de trabajo","en":"New work order","ko":"새 작업 지시"},
    "tag_eq":          {"es":"Tag del equipo *","en":"Equipment tag *","ko":"장비 태그 *"},
    "titulo":          {"es":"Título *","en":"Title *","ko":"제목 *"},
    "tipo":            {"es":"Tipo",                  "en":"Type",                "ko":"유형"},
    "prioridad":       {"es":"Prioridad",             "en":"Priority",            "ko":"우선순위"},
    "descripcion":     {"es":"Descripción",           "en":"Description",         "ko":"설명"},
    "horas_est":       {"es":"Horas estimadas",       "en":"Estimated hours",     "ko":"예상 시간"},
    "req_parada":      {"es":"Requiere parada de planta","en":"Requires plant shutdown","ko":"플랜트 정지 필요"},
    "crear":           {"es":"Crear OT",              "en":"Create WO",           "ko":"작업 지시 생성"},
    "actualizar":      {"es":"Actualizar",            "en":"Update",              "ko":"업데이트"},
    "estado":          {"es":"Estado",                "en":"Status",              "ko":"상태"},
    "horas_reales":    {"es":"Horas reales",          "en":"Actual hours",        "ko":"실제 시간"},
    "pendiente":       {"es":"Pendiente",             "en":"Pending",             "ko":"대기중"},
    "en_curso":        {"es":"En curso",              "en":"In progress",         "ko":"진행중"},
    "completada":      {"es":"Completada",            "en":"Completed",           "ko":"완료"},
    "cancelada":       {"es":"Cancelada",             "en":"Cancelled",           "ko":"취소"},
    "critica":         {"es":"Crítica",               "en":"Critical",            "ko":"긴급"},
    "alta":            {"es":"Alta",                  "en":"High",                "ko":"높음"},
    "normal":          {"es":"Normal",                "en":"Normal",              "ko":"보통"},
    "baja":            {"es":"Baja",                  "en":"Low",                 "ko":"낮음"},
    "correctivo":      {"es":"Correctivo",            "en":"Corrective",          "ko":"사후 정비"},
    "preventivo":      {"es":"Preventivo",            "en":"Preventive",          "ko":"예방 정비"},
    "inspeccion":      {"es":"Inspección",            "en":"Inspection",          "ko":"점검"},
    "lubricacion":     {"es":"Lubricación",           "en":"Lubrication",         "ko":"윤활"},
    "calibracion":     {"es":"Calibración",           "en":"Calibration",         "ko":"교정"},
    "especialidad":    {"es":"Especialidad",          "en":"Specialty",           "ko":"전문분야"},
    "mecanico":        {"es":"Mecánico",              "en":"Mechanical",          "ko":"기계"},
    "electrico":       {"es":"Eléctrico",             "en":"Electrical",          "ko":"전기"},
    "instrumentacion": {"es":"Instrumentación",       "en":"Instrumentation",     "ko":"계장"},
    "otro":            {"es":"Otro",                  "en":"Other",               "ko":"기타"},
    "solo_parada":     {"es":"Solo con parada requerida","en":"Only with shutdown required","ko":"정지 필요만"},
    "reg_hoy":         {"es":"Registrar ejecución hoy","en":"Register execution today","ko":"오늘 실행 등록"},
    "registrado":      {"es":"Registrado",            "en":"Registered",          "ko":"등록됨"},
    "proxima":         {"es":"Próxima",               "en":"Next",                "ko":"다음"},
    "ultima":          {"es":"Última ejecución",      "en":"Last execution",      "ko":"마지막 실행"},
    "frecuencia":      {"es":"Frecuencia",            "en":"Frequency",           "ko":"주기"},
    "horas":           {"es":"Horas",                 "en":"Hours",               "ko":"시간"},
    "nombre":          {"es":"Nombre",                "en":"Name",                "ko":"이름"},
    "fabricante":      {"es":"Fabricante",            "en":"Manufacturer",        "ko":"제조사"},
    "modelo":          {"es":"Modelo",                "en":"Model",               "ko":"모델"},
    "capacidad":       {"es":"Capacidad",             "en":"Capacity",            "ko":"용량"},
    "funcion":         {"es":"Función",               "en":"Function",            "ko":"기능"},
    "material":        {"es":"Material",              "en":"Material",            "ko":"재질"},
    "presion":         {"es":"Presión",               "en":"Pressure",            "ko":"압력"},
    "temperatura":     {"es":"Temperatura",           "en":"Temperature",         "ko":"온도"},
    "norma":           {"es":"Norma",                 "en":"Standard",            "ko":"규격"},
    "potencia":        {"es":"Potencia",              "en":"Power",               "ko":"출력"},
    "voltaje":         {"es":"Voltaje",               "en":"Voltage",             "ko":"전압"},
    "corriente":       {"es":"Corriente FLA",         "en":"FLA Current",         "ko":"정격 전류"},
    "arranque":        {"es":"Arranque",              "en":"Starting method",     "ko":"기동 방식"},
    "tablero":         {"es":"Tablero",               "en":"Panel",               "ko":"배전반"},
    "resumen":         {"es":"Resumen",               "en":"Summary",             "ko":"요약"},
    "repuestos":       {"es":"Repuestos Q",           "en":"Q Spare parts",       "ko":"Q 예비품"},
    "sin_repuestos":   {"es":"Los códigos Q se incorporan en la próxima etapa",
                        "en":"Q codes will be added in the next stage",
                        "ko":"Q 코드는 다음 단계에서 추가됩니다"},
    "programar_parada":{"es":"Programar nueva parada","en":"Schedule new shutdown","ko":"새 정지 예약"},
    "programar":       {"es":"Programar",             "en":"Schedule",            "ko":"예약"},
    "area":            {"es":"Área",                  "en":"Area",                "ko":"구역"},
    "fecha_inicio":    {"es":"Fecha inicio",          "en":"Start date",          "ko":"시작일"},
    "fecha_fin":       {"es":"Fecha fin",             "en":"End date",            "ko":"종료일"},
    "ots_recientes":   {"es":"Órdenes de trabajo recientes","en":"Recent work orders","ko":"최근 작업 지시"},
    "alertas_activas": {"es":"Alertas activas",       "en":"Active alerts",       "ko":"활성 알림"},
    "sin_alertas":     {"es":"Sin alertas activas",   "en":"No active alerts",    "ko":"활성 알림 없음"},
    "prox_mant":       {"es":"Próximos mantenimientos","en":"Upcoming maintenance","ko":"예정 유지보수"},
    "cobertura":       {"es":"Cobertura por área",    "en":"Coverage by area",    "ko":"구역별 커버리지"},
    "equipos_reg":     {"es":"Equipos registrados",   "en":"Registered equipment","ko":"등록 장비"},
    "con_motor_label": {"es":"Con datos de motor",    "en":"With motor data",     "ko":"모터 데이터"},
    "pot_instalada":   {"es":"Potencia instalada",    "en":"Installed power",     "ko":"설치 용량"},
    "ots_activas":     {"es":"OTs activas",           "en":"Active WOs",          "ko":"진행 작업"},
    "alertas_n":       {"es":"Alertas activas",       "en":"Active alerts",       "ko":"활성 알림"},
    "con_ds_label":    {"es":"Con datasheet",         "en":"With datasheet",      "ko":"데이터시트"},
    "solicitudes":     {"es":"Solicitudes pendientes","en":"Pending requests",    "ko":"승인 대기"},
    "usuarios_activos":{"es":"Usuarios activos",      "en":"Active users",        "ko":"활성 사용자"},
    "aprobar":         {"es":"Aprobar",               "en":"Approve",             "ko":"승인"},
    "rechazar":        {"es":"Rechazar",              "en":"Reject",              "ko":"거부"},
    "guardar":         {"es":"Guardar",               "en":"Save",                "ko":"저장"},
    "desactivar":      {"es":"Desactivar",            "en":"Deactivate",          "ko":"비활성화"},
    "agregar_usuario":  {"es":"Agregar usuario",        "en":"Add user",             "ko":"사용자 추가"},
    "repuestos_nav":   {"es":"Repuestos Q",            "en":"Spare Parts Q",        "ko":"Q 예비품"},
    "mapa":            {"es":"Mapa de planta",          "en":"Plant map",            "ko":"플랜트 지도"},
    # Login
    "usuario_lbl":     {"es":"Usuario",                "en":"Username",             "ko":"사용자명"},
    "contrasena_lbl":  {"es":"Contraseña",             "en":"Password",             "ko":"비밀번호"},
    "ingresar_btn":    {"es":"Ingresar",               "en":"Sign in",              "ko":"로그인"},
    "crear_cuenta_tab":{"es":"Crear cuenta",           "en":"Create account",       "ko":"계정 생성"},
    "ingresar_tab":    {"es":"Ingresar",               "en":"Sign in",              "ko":"로그인"},
    # Dashboard
    "tend_ots":        {"es":"Tendencia de órdenes de trabajo","en":"Work order trends","ko":"작업 지시 추이"},
    "dist_tipo":       {"es":"Distribución por tipo",  "en":"Distribution by type", "ko":"유형별 분포"},
    # Equipos
    "todos_filtro":    {"es":"Todos",                  "en":"All",                  "ko":"전체"},
    "con_motor_f":     {"es":"Con motor",              "en":"With motor",           "ko":"모터 있음"},
    "con_specs_f":     {"es":"Con specs",              "en":"With specs",           "ko":"사양 있음"},
    "con_instr_f":     {"es":"Con instrumentos",       "en":"With instruments",     "ko":"계기 있음"},
    "con_ds_f":        {"es":"Con datasheet",          "en":"With datasheet",       "ko":"데이터시트 있음"},
    "con_fab_f":       {"es":"Con fabricante",         "en":"With manufacturer",    "ko":"제조사 있음"},
    # Plan
    "solo_parada":     {"es":"Solo con parada requerida","en":"Only with shutdown required","ko":"정지 필요만 표시"},
    # Alertas
    "estado_lbl":      {"es":"Estado",                 "en":"Status",               "ko":"상태"},
    "prioridad_lbl":   {"es":"Prioridad",              "en":"Priority",             "ko":"우선순위"},
    "activa":          {"es":"Activa",                 "en":"Active",               "ko":"활성"},
    "vista":           {"es":"Vista",                  "en":"Seen",                 "ko":"확인됨"},
    "resuelta":        {"es":"Resuelta",               "en":"Resolved",             "ko":"해결됨"},
    # Historial
    "ots_completadas": {"es":"OTs completadas",        "en":"Completed WOs",        "ko":"완료된 작업"},
    "horas_totales":   {"es":"Horas totales",          "en":"Total hours",          "ko":"총 시간"},
    "correctivos_n":   {"es":"Correctivos",            "en":"Corrective",           "ko":"사후 정비"},
    "preventivos_n":   {"es":"Preventivos",            "en":"Preventive",           "ko":"예방 정비"},
    # Q codes
    "total_q_lbl":     {"es":"Total Q codes",          "en":"Total Q codes",        "ko":"Q 코드 합계"},
    "activos_lbl":     {"es":"Activos",                "en":"Active",               "ko":"활성"},
    "categorias_lbl":  {"es":"Categorías",             "en":"Categories",           "ko":"카테고리"},
    "vinculos_lbl":    {"es":"Vínculos a equipos",     "en":"Equipment links",      "ko":"장비 연결"},
    # Mapa
    "area_lbl":        {"es":"Área",                   "en":"Area",                 "ko":"구역"},
    "equipos_lbl":     {"es":"Equipos",                "en":"Equipment",            "ko":"장비"},
    "con_motor_m":     {"es":"Con motor",              "en":"With motor",           "ko":"모터 있음"},
    "con_specs_m":     {"es":"Con specs",              "en":"With specs",           "ko":"사양 있음"},
    "con_fab_m":       {"es":"Con fabricante",         "en":"With manufacturer",    "ko":"제조사 있음"},
    "sel_area":        {"es":"Seleccioná un área",     "en":"Select an area",       "ko":"구역을 선택하세요"},
    # Email config
    "config_email":    {"es":"Configuración de alertas por email","en":"Email alert configuration","ko":"이메일 알림 설정"},
    "email_rem":       {"es":"Email remitente",        "en":"Sender email",         "ko":"발신 이메일"},
    "smtp_server":     {"es":"Servidor SMTP",          "en":"SMTP server",          "ko":"SMTP 서버"},
    "smtp_port_lbl":   {"es":"Puerto",                  "en":"Port",                 "ko":"포트"},
    "email_dest_lbl":  {"es":"Emails destino",         "en":"Destination emails",   "ko":"수신 이메일"},
    "guardar_config":  {"es":"Guardar configuración",  "en":"Save configuration",   "ko":"설정 저장"},
    "vibraciones":     {"es":"Vibraciones",            "en":"Vibrations",           "ko":"진동"},
    "mediciones":      {"es":"Mediciones",             "en":"Measurements",         "ko":"측정"},
    "med_mec":         {"es":"Mecánicas",              "en":"Mechanical",           "ko":"기계"},
    "med_elec":        {"es":"Eléctricas",             "en":"Electrical",           "ko":"전기"},
    "plc_diario":      {"es":"Inspección PLC diaria",  "en":"Daily PLC inspection", "ko":"일일 PLC 점검"},
    "insp_semanal":    {"es":"Inspección semanal",     "en":"Weekly inspection",    "ko":"주간 점검"},
    "act_archivo":     {"es":"Actualizar desde archivo","en":"Update from file",    "ko":"파일로 업데이트"},
    "arrastra":        {"es":"Arrastrá el archivo Excel aquí",
                        "en":"Drag Excel file here",
                        "ko":"Excel 파일을 여기로 드래그"},
    "actualizado":     {"es":"Datos actualizados correctamente",
                        "en":"Data updated successfully",
                        "ko":"데이터가 성공적으로 업데이트되었습니다"},
    "comisionado":     {"es":"Comisionado",             "en":"Commissioning",        "ko":"시운전"},
    "pr_nav":          {"es":"Pedidos (PR)",             "en":"Purchase Req.",        "ko":"구매 요청"},
    # Registro
    "reg_nombre":      {"es":"Nombre *",               "en":"First name *",         "ko":"이름 *"},
    "reg_apellido":    {"es":"Apellido *",              "en":"Last name *",          "ko":"성 *"},
    "reg_email":       {"es":"Email *",                 "en":"Email *",              "ko":"이메일 *"},
    "reg_dni":         {"es":"DNI *",                   "en":"ID number *",          "ko":"신분증 번호 *"},
    "reg_puesto":      {"es":"Puesto de trabajo *",     "en":"Job title *",          "ko":"직책 *"},
    "reg_area":        {"es":"Área de trabajo *",       "en":"Work area *",          "ko":"작업 구역 *"},
    "reg_tel":         {"es":"Teléfono",                "en":"Phone",                "ko":"전화번호"},
    "reg_usuario":     {"es":"Usuario deseado *",       "en":"Desired username *",   "ko":"사용자명 *"},
    "reg_pass":        {"es":"Contraseña *",            "en":"Password *",           "ko":"비밀번호 *"},
    "reg_pass2":       {"es":"Confirmar contraseña *",  "en":"Confirm password *",   "ko":"비밀번호 확인 *"},
    "reg_submit":      {"es":"Solicitar acceso",        "en":"Request access",       "ko":"접근 요청"},
    "reg_pending":     {"es":"Tu cuenta quedará pendiente de aprobación.",
                        "en":"Your account will be pending administrator approval.",
                        "ko":"계정은 관리자 승인 대기 중입니다."},
    "reg_ok":          {"es":"Solicitud enviada. El administrador te habilitará el acceso.",
                        "en":"Request sent. The administrator will activate your access.",
                        "ko":"요청이 전송되었습니다. 관리자가 접근을 활성화합니다."},
    "reg_err_campos":  {"es":"Completá todos los campos obligatorios",
                        "en":"Please complete all required fields",
                        "ko":"모든 필수 항목을 입력하세요"},
    "reg_err_pass":    {"es":"Las contraseñas no coinciden",
                        "en":"Passwords do not match",
                        "ko":"비밀번호가 일치하지 않습니다"},
    "reg_err_min":     {"es":"Mínimo 6 caracteres",
                        "en":"Minimum 6 characters",
                        "ko":"최소 6자 이상"},
    "reg_err_existe":  {"es":"Ese nombre de usuario ya existe",
                        "en":"That username already exists",
                        "ko":"이미 존재하는 사용자명입니다"},
    "config_guardada": {"es":"Configuración guardada", "en":"Configuration saved",  "ko":"설정 저장됨"},
    "imprimir_ot":     {"es":"Imprimir OT",            "en":"Print WO",             "ko":"작업 지시 인쇄"},
    "imprimir_tarea":  {"es":"Imprimir tarea",         "en":"Print task",           "ko":"작업 인쇄"},
    "imprimir_parada": {"es":"Imprimir parada",        "en":"Print shutdown",       "ko":"정지 인쇄"},
    "seleccionar_area":{"es":"Seleccioná un área:",    "en":"Select an area:",      "ko":"구역을 선택하세요:"},
    "req_parada_label":{"es":"Requiere parada",       "en":"Requires shutdown",   "ko":"정지 필요"},
    "marcar_vista":    {"es":"Marcar vista",          "en":"Mark as seen",        "ko":"확인 표시"},
    "resolver":        {"es":"Resolver",              "en":"Resolve",             "ko":"해결"},
    "ingresar":        {"es":"Ingresar",              "en":"Sign in",             "ko":"로그인"},
    "crear_cuenta":    {"es":"Crear cuenta",          "en":"Create account",      "ko":"계정 생성"},
    "solicitar_acceso":{"es":"Solicitar acceso",      "en":"Request access",      "ko":"접근 요청"},
    "nueva_tarea_prev":{"es":"Nueva tarea preventiva", "en":"New preventive task",  "ko":"새 예방 작업"},
    "reg_nueva_med":   {"es":"+ Registrar nueva medición","en":"+ Register new measurement","ko":"+ 새 측정 등록"},
    "buscar_equipo":   {"es":"Buscar equipo",          "en":"Search equipment",     "ko":"장비 검색"},
    "causa_sintoma":   {"es":"Causa / Síntoma",        "en":"Cause / Symptom",      "ko":"원인 / 증상"},
    "buscar_equipos":  {"es":"Buscar equipos a intervenir","en":"Search equipment", "ko":"장비 검색"},
    "nombre_parada":   {"es":"Nombre de la parada *",  "en":"Shutdown name *",      "ko":"정지 이름 *"},
    "area_codigo":     {"es":"Área (código)",          "en":"Area (code)",          "ko":"구역 (코드)"},
    "sel_pedido":      {"es":"Seleccionar pedido",     "en":"Select request",       "ko":"요청 선택"},
    "agregar_pedido":  {"es":"Agregar pedido",         "en":"Add request",          "ko":"요청 추가"},
    "editar_ped":      {"es":"Editar / Actualizar",    "en":"Edit / Update",        "ko":"편집 / 업데이트"},
    "tipo_med":        {"es":"Tipo de medición",       "en":"Measurement type",     "ko":"측정 유형"},
    "actividad_lbl":   {"es":"Actividad",              "en":"Activity",             "ko":"활동"},
    "buscar_arriba":   {"es":"Buscá un equipo arriba para crear la OT","en":"Search above to create WO","ko":"위에서 장비 검색"},
    "buscar_plan":     {"es":"Buscá un equipo para agregar tarea","en":"Search equipment for task","ko":"작업용 장비 검색"},
    "filtrar_pr":      {"es":"Filtrar por título, solicitante, N° PR...","en":"Filter by title, requester, PR no...","ko":"제목, 요청자, PR 번호로 필터링..."},
    "solicitante_lbl": {"es":"Solicitante",            "en":"Requester",            "ko":"요청자"},
    "n_solicitud_lbl": {"es":"N° Solicitud",           "en":"Request No.",          "ko":"요청 번호"},
    "f_solicitud_lbl": {"es":"F. Solicitud",           "en":"Req. Date",            "ko":"요청일"},
    "f_aprobacion_lbl":{"es":"F. Aprobación",          "en":"Appr. Date",           "ko":"승인일"},
    "f_pr_lbl":        {"es":"F. PR",                  "en":"PR Date",              "ko":"PR 날짜"},
    "n_pr_lbl":        {"es":"N° PR",                  "en":"PR No.",               "ko":"PR 번호"},
    "comprador_lbl":   {"es":"Comprador",              "en":"Buyer",                "ko":"구매자"},
    "ultima_cons_lbl": {"es":"Última consulta",        "en":"Last inquiry",         "ko":"최근 문의"},
    "obs_lbl":         {"es":"Observaciones",          "en":"Observations",         "ko":"비고"},
    "pag_label":       {"es":"Páginas",                "en":"Pages",                "ko":"페이지"},
    "cond_inst":       {"es":"Condiciones de instalación","en":"Installation conditions","ko":"설치 조건"},
    "vib_critica_tit": {"es":"Tendencia de vibración — equipos críticos","en":"Vibration trend — critical equipment","ko":"진동 추세 — 임계 장비"},
    # Tooltips / explicaciones
    "tip_desbalance":  {
        "es":"Desbalance entre CCTs: diferencia entre el CCT con mayor y menor amperaje dividida por el promedio. Verde <10%, Amarillo <20%, Rojo ≥20%.",
        "en":"CCT imbalance: difference between the highest and lowest CCT amperage divided by the average. Green <10%, Yellow <20%, Red ≥20%.",
        "ko":"CCT 불균형: 최대-최소 전류의 차이를 평균으로 나눈 값. 녹색 <10%, 노랑 <20%, 빨강 ≥20%."},
    "tip_desvio_ccts": {
        "es":"Desvío estándar de los amperajes de todos los CCTs del panel para ese día. Un desvío alto indica que algunos circuitos consumen mucho más o menos que el resto.",
        "en":"Standard deviation of all CCT amperages in the panel for that day. A high deviation means some circuits draw significantly more or less than others.",
        "ko":"해당 날짜의 모든 CCT 전류 표준편차. 편차가 크면 일부 회로가 나머지보다 훨씬 많거나 적은 전류를 소비함을 의미합니다."},
    "tip_mc_estado":   {
        "es":"Estado del contactor principal (MC = Main Contactor). 'On' = circuito energizado y en operación normal.",
        "en":"Main Contactor status. 'On' = circuit energized and operating normally.",
        "ko":"주 접촉기(MC) 상태. 'On' = 회로 통전 및 정상 운전 중."},
    "tip_cct":         {
        "es":"CCT = Circuit (Circuito). Cada CCT corresponde a un cable de heat trace independiente dentro del panel.",
        "en":"CCT = Circuit. Each CCT corresponds to an independent heat trace cable within the panel.",
        "ko":"CCT = 회로(Circuit). 각 CCT는 패널 내의 독립적인 히트 트레이스 케이블에 해당합니다."},
    "tip_media":       {
        "es":"Media aritmética: suma de todos los valores dividida por la cantidad de mediciones.",
        "en":"Arithmetic mean: sum of all values divided by the number of measurements.",
        "ko":"산술 평균: 모든 값의 합계를 측정 횟수로 나눈 값."},
    "tip_mediana":     {
        "es":"Mediana: valor central cuando los datos se ordenan de menor a mayor. Menos sensible a valores extremos que la media.",
        "en":"Median: middle value when data is sorted. Less sensitive to extreme values than the mean.",
        "ko":"중앙값: 데이터를 정렬했을 때의 중간 값. 평균보다 극단값에 덜 민감합니다."},
    "tip_moda":        {
        "es":"Moda: valor que aparece con mayor frecuencia en las mediciones.",
        "en":"Mode: value that appears most frequently in the measurements.",
        "ko":"최빈값: 측정값 중 가장 자주 나타나는 값."},
    "tip_desvio":      {
        "es":"Desvío estándar (σ): mide cuánto se alejan los valores de la media. Un σ bajo indica mediciones estables y consistentes.",
        "en":"Standard deviation (σ): measures how far values deviate from the mean. Low σ = stable, consistent measurements.",
        "ko":"표준편차(σ): 값이 평균에서 얼마나 벗어나는지를 측정. 낮은 σ = 안정적이고 일관된 측정값."},
    "tip_varianza":    {
        "es":"Varianza: cuadrado del desvío estándar. Amplifica las diferencias respecto a la media.",
        "en":"Variance: square of the standard deviation. Amplifies differences from the mean.",
        "ko":"분산: 표준편차의 제곱. 평균으로부터의 차이를 증폭합니다."},
    "tip_sigma":       {
        "es":"Puntos a más de 1σ (naranja) o 2σ (rojo) de la media son valores anómalos. Se usan para detectar fallas o comportamientos inusuales.",
        "en":"Points more than 1σ (orange) or 2σ (red) from the mean are anomalous. Used to detect faults or unusual behavior.",
        "ko":"평균에서 1σ(주황) 또는 2σ(빨강) 이상 떨어진 점은 이상값입니다. 결함이나 비정상적인 동작을 감지하는 데 사용됩니다."},
    "tip_run":         {
        "es":"RUN: indica que el PLC está en ejecución normal. Si no está marcado = anomalía crítica, el PLC puede estar parado o en falla.",
        "en":"RUN: indicates the PLC is running normally. If not marked = critical anomaly, the PLC may be stopped or faulted.",
        "ko":"RUN: PLC가 정상 실행 중임을 나타냅니다. 표시 없음 = 심각한 이상, PLC가 정지되거나 고장 상태일 수 있습니다."},
    "tip_baf":         {
        "es":"BAF (Battery Alarm Fault): falla en la batería de respaldo del PLC. Debe reemplazarse antes de que agote.",
        "en":"BAF (Battery Alarm Fault): backup battery fault in the PLC. Should be replaced before it depletes.",
        "ko":"BAF: PLC 백업 배터리 결함. 방전되기 전에 교체해야 합니다."},
    "tip_batt":        {
        "es":"BATT1F / BATT2F: falla en batería 1 o 2 del módulo de CPU. Indica que la batería de memoria está agotada.",
        "en":"BATT1F / BATT2F: battery 1 or 2 fault in the CPU module. Indicates the memory backup battery is depleted.",
        "ko":"BATT1F / BATT2F: CPU 모듈의 배터리 1 또는 2 결함. 메모리 백업 배터리가 소진되었음을 나타냅니다."},
    "tip_q_code":      {
        "es":"Código Q: identificador único de un repuesto o material en el sistema PosAppia de POSCO. Necesario para crear una Purchase Request (PR).",
        "en":"Q code: unique identifier for a spare part or material in POSCO's PosAppia system. Required to create a Purchase Request (PR).",
        "ko":"Q 코드: POSCO PosAppia 시스템에서 예비품 또는 자재의 고유 식별자. 구매 요청(PR) 생성에 필요합니다."},
    "tip_pr":          {
        "es":"PR (Purchase Request): solicitud formal de compra generada en el sistema POSPIA/GoWorks. Tiene un número único como 'PR 2606042'.",
        "en":"PR (Purchase Request): formal purchase request generated in the POSPIA/GoWorks system. Has a unique number like 'PR 2606042'.",
        "ko":"PR (구매 요청): POSPIA/GoWorks 시스템에서 생성된 공식 구매 요청. 'PR 2606042'와 같은 고유 번호를 갖습니다."},
    "tip_goworks":     {
        "es":"GoWorks: sistema interno de POSCO para aprobación de solicitudes de compra antes de generar la PR en POSPIA.",
        "en":"GoWorks: POSCO's internal system for purchase request approval before generating the PR in POSPIA.",
        "ko":"GoWorks: POSPIA에서 PR을 생성하기 전 구매 요청 승인을 위한 POSCO 내부 시스템."},
    "tip_mm_s":        {
        "es":"mm/s: milímetros por segundo. Unidad de medida de velocidad de vibración. El límite normal para rodamientos es <4.5 mm/s según norma ISO 10816.",
        "en":"mm/s: millimeters per second. Vibration velocity measurement unit. Normal limit for bearings is <4.5 mm/s per ISO 10816 standard.",
        "ko":"mm/s: 초당 밀리미터. 진동 속도 측정 단위. ISO 10816 기준에 따라 베어링의 정상 한계는 <4.5 mm/s입니다."},
    "tip_con_motor":   {
        "es":"Equipos que tienen datos de motor cargados en el sistema: potencia (kW), voltaje, corriente FLA, fabricante, etc. Extraídos de los datasheets de los fabricantes y la planilla eléctrica de la planta.",
        "en":"Equipment with motor data loaded: power (kW), voltage, FLA current, manufacturer, etc. Extracted from manufacturer datasheets and the plant electrical schedule.",
        "ko":"모터 데이터가 등록된 장비: 출력(kW), 전압, 정격전류(FLA), 제조사 등. 제조사 데이터시트 및 플랜트 전기 일람표에서 추출."},
    "tip_potencia":    {
        "es":"Potencia total instalada: suma de la potencia nominal (kW) de todos los motores registrados en el sistema. No incluye equipos sin datos de motor aún cargados.",
        "en":"Total installed power: sum of the nominal power (kW) of all motors registered in the system. Does not include equipment without motor data yet loaded.",
        "ko":"총 설치 용량: 시스템에 등록된 모든 모터의 정격 출력(kW) 합계. 아직 모터 데이터가 없는 장비는 포함되지 않습니다."},
    "tip_ots_activas": {
        "es":"Órdenes de trabajo con estado Pendiente o En Curso en este momento. Una OT se crea cuando hay una tarea de mantenimiento correctivo, preventivo, inspección, lubricación o calibración.",
        "en":"Work orders currently with Pending or In Progress status. A WO is created when there is a corrective, preventive, inspection, lubrication or calibration maintenance task.",
        "ko":"현재 대기 중 또는 진행 중 상태의 작업 지시. 사후 정비, 예방 정비, 점검, 윤활 또는 교정 작업 시 작업 지시가 생성됩니다."},
    "tip_con_ds":      {
        "es":"Equipos que tienen el número de datasheet técnico del fabricante vinculado en el sistema. El datasheet contiene todas las especificaciones técnicas del equipo.",
        "en":"Equipment with the manufacturer's technical datasheet number linked in the system. The datasheet contains all technical specifications of the equipment.",
        "ko":"시스템에 제조사 기술 데이터시트 번호가 연결된 장비. 데이터시트에는 장비의 모든 기술 사양이 포함됩니다."},
    "tip_estado_alim": {
        "es":"Estado de alimentación eléctrica del instrumento. 'ON' = energizado. 'OFF' o 'AN' (anormal) = sin tensión o con problema.",
        "en":"Instrument power supply status. 'ON' = energized. 'OFF' or 'AN' (abnormal) = no power or issue detected.",
        "ko":"계기 전원 공급 상태. 'ON' = 통전. 'OFF' 또는 'AN'(비정상) = 전원 없음 또는 문제 감지."},
    "error_login_msg": {"es":"Ingresá usuario y contraseña","en":"Enter username and password","ko":"사용자명과 비밀번호를 입력하세요"},
    "error_login_bad": {"es":"Usuario o contraseña incorrectos","en":"Incorrect username or password","ko":"잘못된 사용자명 또는 비밀번호"},
}
def t(k):
    lg = st.session_state.get("lang","es")
    return T.get(k,{}).get(lg, T.get(k,{}).get("es",k))

def tip(k):
    """Muestra un tooltip con explicacion segun idioma"""
    msg = t(k)
    if msg and msg != k:
        st.caption(f"ℹ️ {msg}")

# ── LOGIN ─────────────────────────────────────────────────────────────────────
if not st.session_state.get("logged_in"):
    st.markdown("<div style='padding:24px'>",unsafe_allow_html=True)
    c1,c2,c3,_ = st.columns([1,1,1,7])
    with c1:
        if st.button("🇦🇷 Español"):
            st.session_state["lang"]="es"; st.rerun()
    with c2:
        if st.button("🇬🇧 English"):
            st.session_state["lang"]="en"; st.rerun()
    with c3:
        if st.button("🇰🇷 한국어"):
            st.session_state["lang"]="ko"; st.rerun()

    st.markdown("<br>",unsafe_allow_html=True)
    _,col,_ = st.columns([1,1,1])
    with col:
        st.markdown("""
        <div style="background:#111113;border:1px solid #222;border-radius:12px;
                    padding:32px 28px;text-align:center;margin-bottom:20px">
            <div style="font-size:26px;font-weight:700;color:#fff;letter-spacing:-0.02em">POSCO</div>
            <div style="font-size:10px;color:#2563eb;letter-spacing:0.2em;text-transform:uppercase">Argentina</div>
            <div style="height:1px;background:#222;margin:18px 0"></div>
            <div style="font-size:15px;font-weight:600;color:#e2e2e5">CP2 Maintenance System</div>
            <div style="font-size:10px;color:#444;margin-top:4px;text-transform:uppercase;letter-spacing:0.08em">
                SAL DE ORO · Salar del Hombre Muerto
            </div>
        </div>
        """,unsafe_allow_html=True)

        tab_in,tab_reg = st.tabs([t("ingresar_tab"),t("crear_cuenta_tab")])
        with tab_in:
            with st.form("login"):
                usr = st.text_input(t("usuario_lbl"), placeholder=t("usuario_lbl").lower())
                pwd = st.text_input(t("contrasena_lbl"), placeholder="••••••••", type="password")
                if st.form_submit_button(t("ingresar_btn"), type="primary", use_container_width=True):
                    if not usr or not pwd:
                        st.error(t("error_login"))
                    else:
                        u = qdf("SELECT * FROM usuarios WHERE username=? AND activo=1",(usr.strip(),))
                        if len(u)>0 and u.iloc[0]["password_hash"]==hp(pwd):
                            r = u.iloc[0]
                            st.session_state.update({
                                "logged_in":True,"user_id":int(r["id"]),
                                "username":r["username"],"nombre":r["nombre"],
                                "apellido":r["apellido"],"rol":r["rol"],
                                "lang":r.get("idioma","es") or "es",
                            })
                            run("UPDATE usuarios SET last_login=? WHERE id=?",
                                (datetime.now().isoformat(),int(r["id"])))
                            st.query_params["uid"] = str(int(r["id"]))
                            st.rerun()
                        else:
                            st.error(t("error_login"))

        with tab_reg:
            with st.form("registro"):
                st.caption(t("reg_pending"))
                r_nom  = st.text_input(t("reg_nombre"))
                r_ape  = st.text_input(t("reg_apellido"))
                r_mail = st.text_input(t("reg_email"))
                r_dni  = st.text_input(t("reg_dni"))
                r_pue  = st.text_input(t("reg_puesto"))
                r_area = st.text_input(t("reg_area"))
                r_tel  = st.text_input(t("reg_tel"))
                r_usr  = st.text_input(t("reg_usuario"))
                r_pwd  = st.text_input(t("reg_pass"), type="password")
                r_pwd2 = st.text_input(t("reg_pass2"), type="password")
                if st.form_submit_button(t("reg_submit"), type="primary", use_container_width=True):
                    if not all([r_nom,r_ape,r_mail,r_dni,r_pue,r_area,r_usr,r_pwd]):
                        st.error(t("reg_err_campos"))
                    elif r_pwd != r_pwd2: st.error(t("reg_err_pass"))
                    elif len(r_pwd) < 6:  st.error(t("reg_err_min"))
                    else:
                        try:
                            run("""INSERT INTO usuarios
                                (username,password_hash,nombre,apellido,email,
                                 dni,telefono,puesto,area_trabajo,rol,idioma,activo)
                                VALUES (?,?,?,?,?,?,?,?,?,'invitado',?,0)""",
                                (r_usr.strip(),hp(r_pwd),r_nom,r_ape,r_mail,
                                 r_dni,r_tel or None,r_pue,r_area,
                                 st.session_state.get("lang","es")))
                            st.success(t("reg_ok"))
                        except: st.error(t("reg_err_existe"))
    st.stop()

# ── SESION ACTIVA ─────────────────────────────────────────────────────────────
ROL          = st.session_state.get("rol","invitado")
NOMBRE       = st.session_state.get("nombre","")
APELLIDO     = st.session_state.get("apellido","")
INICIALES    = (NOMBRE[:1]+APELLIDO[:1]).upper()
ES_ADMIN     = ROL == "admin"
PUEDE_EDITAR = ROL in ("admin","supervisor","tecnico")
LANG         = st.session_state.get("lang","es")

ROL_ES = {"admin":"Administrador","supervisor":"Supervisor","tecnico":"Técnico",
          "directivo":"Directivo","invitado":"Invitado"}

ESTADO_IC = {"pendiente":"🟡","en_curso":"🔵","completada":"✅","cancelada":"⚫"}
PRIO_IC   = {"critica":"🔴","alta":"🟠","normal":"🟡","baja":"🟢"}

# ── NAVEGACION HORIZONTAL ────────────────────────────────────────────────────
if "pagina" not in st.session_state:
    st.session_state["pagina"] = "dashboard"

nav_items = ["dashboard","equipos","ots","plan","alertas","paradas","repuestos","mapa","mediciones","pr","historial"]
if ES_ADMIN: nav_items.append("usuarios")

# Topbar
# Topbar con idiomas al lado del usuario
tb1, tb2, tb3, tb4, tb5 = st.columns([6, 1, 1, 1, 4])
with tb1:
    st.markdown(f"""
    <div style="background:#111113;border-bottom:1px solid #222;padding:0 16px;
                display:flex;align-items:center;height:44px">
        <div style="font-size:14px;font-weight:700;color:#fff;letter-spacing:-0.01em">
            CP2 Maintenance System
        </div>
        <div style="font-size:11px;color:#333;margin-left:12px">
            SAL DE ORO · POSCO Argentina
        </div>
    </div>
    """, unsafe_allow_html=True)
with tb2:
    if st.button("🇦🇷", use_container_width=True,
                 type="primary" if LANG=="es" else "secondary", key="lang_es"):
        st.session_state["lang"] = "es"; st.rerun()
with tb3:
    if st.button("🇬🇧", use_container_width=True,
                 type="primary" if LANG=="en" else "secondary", key="lang_en"):
        st.session_state["lang"] = "en"; st.rerun()
with tb4:
    if st.button("🇰🇷", use_container_width=True,
                 type="primary" if LANG=="ko" else "secondary", key="lang_ko"):
        st.session_state["lang"] = "ko"; st.rerun()
with tb5:
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:8px;justify-content:flex-end;
                padding-right:8px;height:44px">
        <div style="width:26px;height:26px;border-radius:50%;background:#2563eb;
                    display:flex;align-items:center;justify-content:center;
                    font-size:10px;font-weight:600;color:#fff">{INICIALES}</div>
        <span style="font-size:12px;color:#888">{NOMBRE} {APELLIDO}</span>
        <span style="font-size:11px;color:#444;border:1px solid #222;
                     border-radius:4px;padding:1px 6px">{ROL_ES.get(ROL,ROL)}</span>
    </div>
    """, unsafe_allow_html=True)

# Barra de navegacion + idioma en esquina derecha
NAV_LABELS = {
    "dashboard": t("dashboard"),
    "equipos":   t("equipos"),
    "ots":       t("ots"),
    "plan":      t("plan"),
    "alertas":   t("alertas"),
    "paradas":   t("paradas"),
    "historial":     t("historial"),
    "repuestos":     t("repuestos_nav"),
    "mapa":          t("mapa"),
    "mediciones":    t("mediciones"),
    "pr":            t("pr_nav"),
    "usuarios":      t("usuarios"),
}

# Navegacion en fila, idiomas a la derecha separados
nav_cols = st.columns(len(nav_items))
for i, item in enumerate(nav_items):
    with nav_cols[i]:
        is_active = st.session_state["pagina"] == item
        if st.button(NAV_LABELS[item], key=f"nav_{item}",
                     use_container_width=True,
                     type="primary" if is_active else "secondary"):
            st.session_state["pagina"] = item
            st.rerun()



# Cerrar sesion en topbar
with st.container():
    col_logout = st.columns([10,1])
    with col_logout[1]:
        if st.button(t("salir"), type="secondary"):
            for k in ["logged_in","user_id","username","nombre","apellido","rol"]:
                st.session_state.pop(k,None)
            st.query_params.clear()
            st.rerun()

pagina = st.session_state["pagina"]
st.markdown("<div style='padding:24px 40px 40px 40px;max-width:1400px;margin:0 auto'>",
            unsafe_allow_html=True)

# ════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════
if pagina == "dashboard":
    st.markdown(f"## {t('dashboard')}")

    # Buscador con sugerencias en tiempo real
    busq = st.text_input("",
        placeholder=t("buscar"),
        label_visibility="collapsed",
        key="busq_dash")

    if busq and len(busq) >= 1:
        eq_all = qdf("SELECT * FROM equipos")
        m = (eq_all.tag.str.lower().str.contains(busq.lower(),na=False)|
             eq_all.spec_nombre_equipo.fillna("").str.lower().str.contains(busq.lower())|
             eq_all.motor_descripcion.fillna("").str.lower().str.contains(busq.lower())|
             eq_all.area_descripcion.fillna("").str.lower().str.contains(busq.lower())|
             eq_all.mec_fabricante.fillna("").str.lower().str.contains(busq.lower())|
             eq_all.tipo_descripcion.fillna("").str.lower().str.contains(busq.lower()))
        resultados = eq_all[m].head(15)
        if len(resultados) > 0:
            st.markdown(f"<div style='font-size:12px;color:#555;margin-bottom:8px'>{len(resultados)} resultados</div>",
                        unsafe_allow_html=True)
            for _,row in resultados.iterrows():
                nm  = row.get("spec_nombre_equipo") or row.get("motor_descripcion") or ""
                fab = row.get("mec_fabricante") or ""
                area= row.get("area_descripcion","")
                c1,c2 = st.columns([1,4])
                with c1:
                    if st.button(row["tag"], key=f"sq_{row['tag']}",
                                 use_container_width=True, type="primary"):
                        st.session_state["eq_sel"] = row["tag"]
                        st.session_state["pagina"] = "equipos"
                        st.rerun()
                with c2:
                    fab_html = f"<span style='color:#555'>  &middot;  {fab}</span>" if fab else ""
                    st.markdown(
                        f"<div style='padding:7px 0;font-size:13px'>"
                        f"<span style='color:#ccc'>{nm}</span>{fab_html}"
                        f"<span style='color:#333;font-size:11px'>  &middot;  {area}</span></div>",
                        unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='color:#555;font-size:13px'>Sin resultados para '{busq}'</div>",
                        unsafe_allow_html=True)
        st.markdown("---")

    eq     = qdf("SELECT * FROM equipos")
    ots_df = qdf("SELECT * FROM ordenes_trabajo ORDER BY fecha_creacion DESC")
    al_df  = qdf("SELECT * FROM alertas WHERE estado='activa' ORDER BY prioridad DESC")
    pl_df  = qdf("SELECT * FROM plan_mantenimiento WHERE activo=1 ORDER BY proxima_fecha")

    kw = pd.to_numeric(eq["motor_kw"], errors="coerce").fillna(0).sum()
    eq["motor_kw"] = pd.to_numeric(eq["motor_kw"], errors="coerce")
    cm = eq["motor_kw"].notna().sum()
    oa = len(ots_df[ots_df.estado.isin(["pendiente","en_curso"])])

    c1,c2,c3,c4,c5,c6 = st.columns(6)
    with c1: st.metric(t("equipos_reg"), f"{len(eq):,}")
    with c2:
        st.metric(t("con_motor_label"), cm, f"{round(cm/max(len(eq),1)*100)}%")
        st.caption(f"ℹ️ {t('tip_con_motor')}")
    with c3:
        st.metric(t("pot_instalada"), f"{round(kw):,} kW")
        st.caption(f"ℹ️ {t('tip_potencia')}")
    with c4:
        st.metric(t("ots_activas"), oa)
        st.caption(f"ℹ️ {t('tip_ots_activas')}")
    with c5: st.metric(t("alertas_n"), len(al_df))
    with c6:
        st.metric(t("con_ds_label"), eq["mec_doc_numero"].notna().sum())
        st.caption(f"ℹ️ {t('tip_con_ds')}")

    # ── PANEL DE VIBRACIONES CRÍTICAS ────────────────────────────────────────
    try:
        vib_criticos = qdf("""
            SELECT tag, equipo_desc, MAX(valor) val, limite, tipo_medicion
            FROM vibraciones WHERE estado='critico'
            GROUP BY tag ORDER BY MAX(valor) DESC
        """)

        if len(vib_criticos) > 0:
            import plotly.graph_objects as go
            import json
            import streamlit.components.v1 as components

            LANG_D = st.session_state.get("lang","es")
            titulo_alerta = {
                "es": f"🚨 {len(vib_criticos)} equipos con vibración CRÍTICA",
                "en": f"🚨 {len(vib_criticos)} equipment with CRITICAL vibration",
                "ko": f"🚨 {len(vib_criticos)}개 장비 임계 진동 감지"
            }.get(LANG_D,"")

            # Lista de badges
            badges = "".join([
                f"<span style='background:#1a0505;border:1px solid #7f1d1d;"
                f"border-radius:5px;padding:4px 10px;font-family:monospace;"
                f"font-size:11px;color:#fca5a5;white-space:nowrap;display:inline-block'>"
                f"<b>{r['tag']}</b>&nbsp;&nbsp;{r['val']} mm/s</span>"
                for _,r in vib_criticos.iterrows()
            ])

            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#1a0505,#2a0808);
                        border:1.5px solid #ef4444;border-radius:10px;
                        padding:14px 18px;margin-bottom:16px;
                        box-shadow:0 0 20px rgba(239,68,68,0.15)">
                <div style="font-size:14px;font-weight:700;color:#ef4444;
                            letter-spacing:0.02em;margin-bottom:10px">{titulo_alerta}</div>
                <div style="display:flex;flex-wrap:wrap;gap:6px">{badges}</div>
            </div>
            """, unsafe_allow_html=True)

            # Graficos rotativos
            lbl_graf = t("vib_critica_tit")
            st.markdown(f"#### {lbl_graf}")

            tags_crit = vib_criticos["tag"].tolist()
            tags_sql  = ",".join([f"'{t2}'" for t2 in tags_crit])
            vib_trend = qdf(f"""
                SELECT tag, equipo_desc, fecha, MAX(valor) valor, limite
                FROM vibraciones
                WHERE tag IN ({tags_sql}) AND tipo_medicion LIKE '%Vib%'
                GROUP BY tag, fecha ORDER BY tag, fecha
            """)
            vib_trend["fecha"] = pd.to_datetime(vib_trend["fecha"], errors="coerce")

            traces_data = []
            for tag2 in tags_crit:
                df_t = vib_trend[vib_trend.tag==tag2].sort_values("fecha")
                if len(df_t) == 0: continue
                nombre = df_t["equipo_desc"].iloc[0]
                lim    = float(df_t["limite"].iloc[0])
                traces_data.append({
                    "tag":    tag2,
                    "nombre": str(nombre)[:35],
                    "fechas": df_t["fecha"].dt.strftime("%d/%m").tolist(),
                    "vals":   [round(float(v),2) for v in df_t["valor"].tolist()],
                    "limite": lim,
                    "max_v":  round(float(df_t["valor"].max()),2),
                })

            charts_json = json.dumps(traces_data)
            html_chart = f"""<!DOCTYPE html><html><head>
            <script src="https://cdn.plot.ly/plotly-2.26.0.min.js"></script>
            <style>
            body{{margin:0;padding:8px;background:transparent;font-family:Inter,Arial,sans-serif}}
            #chart{{width:100%;height:240px}}
            #nav{{display:flex;justify-content:center;gap:6px;margin-top:6px;flex-wrap:wrap}}
            .dot{{width:8px;height:8px;border-radius:50%;background:#333;cursor:pointer;
                  transition:all 0.2s;border:1px solid #444}}
            .dot.active{{background:#ef4444;border-color:#ef4444;transform:scale(1.3)}}
            #tag_label{{text-align:center;font-size:11px;color:#666;margin-top:3px;font-family:monospace}}
            </style></head><body>
            <div id="chart"></div>
            <div id="tag_label"></div>
            <div id="nav"></div>
            <script>
            var data={charts_json};
            var current=0, timer=null;
            function render(idx){{
              var d=data[idx];
              var colors=d.vals.map(v=>v>d.limite?'#ef4444':v>d.limite*0.8?'#f59e0b':'#22c55e');
              Plotly.newPlot('chart',[
                {{x:d.fechas,y:d.vals,type:'scatter',mode:'lines+markers',
                  line:{{color:'#6366f1',width:2}},
                  marker:{{color:colors,size:5,line:{{width:0}}}},
                  name:d.tag,hovertemplate:'%{{x}}: <b>%{{y}}</b> mm/s<extra></extra>'}},
                {{x:[d.fechas[0],d.fechas[d.fechas.length-1]],y:[d.limite,d.limite],
                  type:'scatter',mode:'lines',
                  line:{{color:'#ef4444',width:1.5,dash:'dash'}},
                  name:'Límite '+d.limite,hoverinfo:'skip'}}
              ],{{
                paper_bgcolor:'rgba(0,0,0,0)',plot_bgcolor:'rgba(0,0,0,0)',
                margin:{{l:45,r:15,t:28,b:30}},
                title:{{text:d.nombre,font:{{color:'#aaa',size:12,family:'Inter'}},x:0.5}},
                xaxis:{{color:'#444',tickfont:{{size:9,color:'#555'}},gridcolor:'#1e1e22',nticks:8}},
                yaxis:{{color:'#444',tickfont:{{size:9}},gridcolor:'#1e1e22',
                        title:{{text:'mm/s',font:{{color:'#555',size:10}}}}}},
                legend:{{font:{{color:'#555',size:9}},bgcolor:'rgba(0,0,0,0)',
                         orientation:'h',y:-0.18,x:0.5,xanchor:'center'}},
                showlegend:true
              }},{{displayModeBar:false,responsive:true}});
              document.getElementById('tag_label').innerText=
                d.tag+'  ·  max: '+Math.max(...d.vals).toFixed(1)+' mm/s  ·  límite: '+d.limite;
              document.querySelectorAll('.dot').forEach((dot,i)=>{{
                dot.classList.toggle('active',i===idx);
              }});
              current=idx;
            }}
            var nav=document.getElementById('nav');
            data.forEach((d,i)=>{{
              var dot=document.createElement('div');
              dot.className='dot'+(i===0?' active':'');
              dot.title=d.tag+' (max: '+d.max_v+' mm/s)';
              dot.onclick=function(){{clearInterval(timer);render(i);timer=setInterval(next,15000);}};
              nav.appendChild(dot);
            }});
            function next(){{render((current+1)%data.length);}}
            if(data.length>0){{render(0);timer=setInterval(next,15000);}}
            </script></body></html>"""

            components.html(html_chart, height=330, scrolling=False)
    except:
        pass
    st.markdown("---")
    col_a, col_b = st.columns([3,2])

    with col_a:
        st.markdown(f"#### {t('ots_recientes')}")
        if len(ots_df)==0:
            st.info(t("sin_resultados"))
        for _,ot in ots_df.head(6).iterrows():
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;padding:8px 0;
                        border-bottom:1px solid #1a1a1a;font-size:13px">
                <span style="font-family:monospace;color:#444;min-width:90px">{ot["numero_ot"]}</span>
                <span style="font-family:monospace;color:#3b82f6;min-width:130px">{ot["tag_equipo"]}</span>
                <span style="flex:1;color:#ccc">{str(ot["titulo"])[:40]}</span>
                <span>{PRIO_IC.get(ot["prioridad"],"")}</span>
                <span>{ESTADO_IC.get(ot["estado"],"")}</span>
            </div>
            """,unsafe_allow_html=True)

    with col_b:
        st.markdown(f"#### {t('alertas_activas')}")
        if len(al_df)==0:
            st.success(t("sin_alertas"))
        for _,al in al_df.iterrows():
            msg = al.get(f"mensaje_{LANG}") or al.get("mensaje_es","")
            color = {"critica":"#ef4444","alta":"#f59e0b"}.get(al["prioridad"],"#22c55e")
            st.markdown(f"""
            <div style="border-left:3px solid {color};padding:10px 14px;
                        background:#111113;border-radius:0 6px 6px 0;margin-bottom:8px">
                <div style="font-family:monospace;font-size:11px;color:{color};font-weight:600">{al.get("tag_equipo","")}</div>
                <div style="font-size:13px;color:#ccc;margin-top:2px">{msg}</div>
                <div style="font-size:11px;color:#333;margin-top:2px">{al.get("tipo_alerta","")} · {str(al.get("fecha_generada",""))[:10]}</div>
            </div>
            """,unsafe_allow_html=True)

    st.markdown("---")
    col_c,col_d = st.columns([2,3])

    with col_c:
        st.markdown(f"#### {t('prox_mant')}")
        plan_prox = pl_df[pl_df.proxima_fecha.notna()].head(8)
        if len(plan_prox) == 0:
            st.info("Sin tareas programadas")
        for _,p in plan_prox.iterrows():
            par = " · parada" if p.get("requiere_parada") else ""
            prox = str(p.get("proxima_fecha",""))[:10]
            try:
                from datetime import date as _date
                dias = (_date.fromisoformat(prox) - _date.today()).days
                color_dot = "#ef4444" if dias<=7 else "#f59e0b" if dias<=30 else "#3b82f6"
            except:
                color_dot = "#3b82f6"
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:8px;padding:7px 0;
                        border-bottom:1px solid #1a1a1a;font-size:12px">
                <div style="width:6px;height:6px;border-radius:50%;background:{color_dot};flex-shrink:0"></div>
                <span style="font-family:monospace;color:#3b82f6;min-width:110px">{p["tag_equipo"]}</span>
                <span style="flex:1;color:#666">{str(p["descripcion"])[:25]}{par}</span>
                <span style="color:#444;font-size:11px">{prox}</span>
            </div>
            """,unsafe_allow_html=True)

    with col_d:
        st.markdown(f"#### {t('cobertura')}")
        stats = qdf("""SELECT area_codigo,COUNT(*) total,
            SUM(CASE WHEN motor_kw IS NOT NULL THEN 1 ELSE 0 END) motor,
            SUM(CASE WHEN spec_nombre_equipo IS NOT NULL AND spec_nombre_equipo!='' THEN 1 ELSE 0 END) spec,
            SUM(CASE WHEN mec_fabricante IS NOT NULL AND mec_fabricante!='' THEN 1 ELSE 0 END) fab
            FROM equipos GROUP BY area_codigo ORDER BY total DESC LIMIT 14""")
        stats["pct"] = ((stats["motor"]+stats["spec"]+stats["fab"])/(stats["total"]*3)*100).round(0)
        fig = px.bar(stats,x="pct",y="area_codigo",orientation="h",
                     color="pct",
                     color_continuous_scale=["#1e293b","#334155","#0ea5e9","#6366f1"],
                     range_color=[0,100],labels={"pct":"","area_codigo":""},height=320)
        fig.update_layout(margin=dict(l=0,r=20,t=0,b=0),coloraxis_showscale=False,
                          plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                          font=dict(family="Inter",size=11,color="#888"),
                          xaxis=dict(showgrid=True,gridcolor="#1e1e22",ticksuffix="%",color="#555"),
                          yaxis=dict(showgrid=False,color="#888"))
        st.plotly_chart(fig,use_container_width=True)

    st.markdown("---")
    col_e,col_f = st.columns([3,2])
    with col_e:
        st.markdown(f"#### {t('tend_ots')}")
        ots_hist = qdf("""
            SELECT strftime('%Y-%m',fecha_inicio) mes, tipo_tarea, COUNT(*) n
            FROM ordenes_trabajo WHERE fecha_inicio IS NOT NULL
            GROUP BY mes,tipo_tarea ORDER BY mes
        """)
        if len(ots_hist) > 0:
            fig2 = px.bar(ots_hist,x="mes",y="n",color="tipo_tarea",height=220,barmode="stack",
                          color_discrete_map={
                              "correctivo":"#6366f1","preventivo":"#0ea5e9",
                              "inspeccion":"#14b8a6","lubricacion":"#8b5cf6","calibracion":"#64748b"})
            fig2.update_layout(margin=dict(l=0,r=0,t=0,b=0),
                               plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                               font=dict(family="Inter",size=11,color="#888"),
                               xaxis=dict(color="#444",gridcolor="#1e1e22",showgrid=True),
                               yaxis=dict(color="#444",gridcolor="#1e1e22",showgrid=True),
                               legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)",
                                           orientation="h",yanchor="bottom",y=1.02,xanchor="left",x=0))
            st.plotly_chart(fig2,use_container_width=True)
        else:
            st.markdown('<div style="height:100px;display:flex;align-items:center;justify-content:center;color:#333;font-size:13px">Las OTs aparecerán acá a medida que se registren</div>',unsafe_allow_html=True)
    with col_f:
        st.markdown(f"#### {t('dist_tipo')}")
        tipos_dist = qdf("""SELECT tipo_codigo,COUNT(*) n FROM equipos
            WHERE tipo_codigo IS NOT NULL AND tipo_codigo!=''
            GROUP BY tipo_codigo ORDER BY n DESC LIMIT 10""")
        fig3 = px.pie(tipos_dist,values="n",names="tipo_codigo",height=220,
                      color_discrete_sequence=[
                          "#6366f1","#0ea5e9","#14b8a6","#8b5cf6","#64748b",
                          "#3b82f6","#06b6d4","#7c3aed","#0369a1","#0891b2"])
        fig3.update_layout(margin=dict(l=0,r=0,t=0,b=0),
                           plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                           font=dict(family="Inter",size=11,color="#888"),
                           legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)"))
        fig3.update_traces(textposition="inside",textinfo="percent+label",
                           textfont=dict(color="white",size=9))
        st.plotly_chart(fig3,use_container_width=True)

# ════════════════════════════════════
# EQUIPOS
# ════════════════════════════════════
elif pagina == "equipos":
    st.markdown(f"## {t('equipos')}")
    eq = qdf("SELECT * FROM equipos ORDER BY area_codigo,tipo_codigo,tag")

    c1,c2,c3,c4 = st.columns([3,2,1,2])
    with c1: busq=st.text_input("",placeholder=t("buscar"),label_visibility="collapsed")
    with c2:
        areas=[""]+sorted(eq.area_codigo.dropna().unique().tolist())
        area_s=st.selectbox("",areas,
            format_func=lambda x:t("todas_areas") if x=="" else
            f"{x}  {eq[eq.area_codigo==x].area_descripcion.iloc[0] if len(eq[eq.area_codigo==x])>0 else ''}",
            label_visibility="collapsed")
    with c3:
        tipos=[""]+sorted(eq.tipo_codigo.dropna().unique().tolist())
        tipo_s=st.selectbox("",tipos,label_visibility="collapsed")
    with c4:
        filtro=st.selectbox("",[t("todos_filtro"),t("con_motor_f"),t("con_specs_f"),
                                t("con_instr_f"),t("con_ds_f"),t("con_fab_f")],
                             label_visibility="collapsed")

    res=eq.copy()
    if area_s: res=res[res.area_codigo==area_s]
    if tipo_s: res=res[res.tipo_codigo==tipo_s]
    if filtro==t("con_motor_f"): res=res[pd.to_numeric(res.motor_kw,errors="coerce").notna()]
    if filtro==t("con_specs_f"):  res=res[res.spec_nombre_equipo.notna()|res.mec_doc_numero.notna()]
    if filtro==t("con_instr_f"):  res=res[res.instr_tags.notna()&(res.instr_tags!="")]
    if filtro==t("con_ds_f"):     res=res[res.mec_doc_numero.notna()&(res.mec_doc_numero!="")]
    if filtro==t("con_fab_f"):    res=res[res.mec_fabricante.notna()&(res.mec_fabricante!="")]
    if busq:
        m=(res.tag.str.lower().str.contains(busq.lower(),na=False)|
           res.spec_nombre_equipo.fillna("").str.lower().str.contains(busq.lower())|
           res.motor_descripcion.fillna("").str.lower().str.contains(busq.lower())|
           res.area_descripcion.fillna("").str.lower().str.contains(busq.lower())|
           res.mec_fabricante.fillna("").str.lower().str.contains(busq.lower()))
        res=res[m]

    st.caption(f"{len(res)} resultados")
    col_list,col_ficha = st.columns([1,2])

    with col_list:
        cont=st.container(height=560)
        with cont:
            for _,row in res.head(150).iterrows():
                is_sel = st.session_state.get("eq_sel")==row["tag"]
                nm = row.get("spec_nombre_equipo") or row.get("motor_descripcion") or ""
                nm_str = str(nm) if nm and str(nm) not in ("nan","None","") else ""
                label = f"{row['tag']}  {nm_str[:20]}" if nm_str else row["tag"]
                if st.button(label,key=f"eq_{row['tag']}",use_container_width=True,
                             type="primary" if is_sel else "secondary"):
                    st.session_state["eq_sel"]=row["tag"]; st.rerun()

    with col_ficha:
        tag_s = st.session_state.get("eq_sel")
        if not tag_s:
            st.markdown("""
            <div style="height:300px;display:flex;flex-direction:column;align-items:center;
                        justify-content:center;border:1px solid #1e1e22;border-radius:8px;color:#333">
                <div style="font-size:32px;margin-bottom:8px">⚙</div>
                <div style="font-size:13px">Seleccioná un equipo</div>
            </div>""",unsafe_allow_html=True)
        else:
            rows=eq[eq.tag==tag_s]
            if len(rows)==0: st.info("No encontrado")
            else:
                r=rows.iloc[0]
                nm  = safe_str(r.get("spec_nombre_equipo") or r.get("motor_descripcion") or r.get("tipo_descripcion"))
                fab = safe_str(r.get("mec_fabricante") or r.get("motor_proveedor"))

                # Header de la ficha
                badges = ""
                if r.get("pfd_documento"): badges += "<span style='background:#1a2744;border:1px solid #1e3a6e;border-radius:4px;padding:2px 8px;font-size:11px;color:#60a5fa;margin-right:4px'>PFD</span>"
                if r.get("spec_nombre_equipo") or r.get("mec_doc_numero"): badges += "<span style='background:#0f2a1e;border:1px solid #166534;border-radius:4px;padding:2px 8px;font-size:11px;color:#4ade80;margin-right:4px'>Mecánico</span>"
                if pd.notna(r.get("motor_kw")): badges += "<span style='background:#2a1f0a;border:1px solid #92400e;border-radius:4px;padding:2px 8px;font-size:11px;color:#fbbf24;margin-right:4px'>Motor</span>"
                if r.get("elec_tipo_doc"): badges += "<span style='background:#1e1a2a;border:1px solid #5b21b6;border-radius:4px;padding:2px 8px;font-size:11px;color:#a78bfa;margin-right:4px'>Eléctrico</span>"
                if r.get("instr_tags"): badges += "<span style='background:#2a0f0f;border:1px solid #991b1b;border-radius:4px;padding:2px 8px;font-size:11px;color:#f87171;margin-right:4px'>Instrumentos</span>"

                st.markdown(f"""
                <div style="background:#111113;border:1px solid #222;border-radius:8px;
                            padding:16px 18px;margin-bottom:12px">
                    <div style="font-family:monospace;font-size:20px;font-weight:700;color:#fff">{r["tag"]}</div>
                    <div style="font-size:13px;color:#666;margin-top:2px">{nm}</div>
                    {"<div style='font-size:12px;color:#444;margin-top:1px'>" + fab + "</div>" if fab else ""}
                    <div style="display:flex;gap:6px;margin-top:10px;flex-wrap:wrap">
                        <span style="background:#1a1a1e;border:1px solid #222;border-radius:4px;padding:2px 8px;font-size:11px;color:#888">{r.get("tipo_descripcion",r.get("tipo_codigo",""))}</span>
                        <span style="background:#1a1a1e;border:1px solid #222;border-radius:4px;padding:2px 8px;font-size:11px;color:#888">{r.get("area_descripcion",r.get("area_codigo",""))}</span>
                        {badges}
                    </div>
                </div>
                """,unsafe_allow_html=True)

                # TODAS las especificaciones en tabs
                tab_labels = ["Resumen"]
                if r.get("pfd_documento"): tab_labels.append("PFD")
                if r.get("spec_nombre_equipo") or r.get("mec_doc_numero") or r.get("mec_fabricante"): tab_labels.append(t("mecanico"))
                if pd.notna(r.get("motor_kw")): tab_labels.append(t("potencia"))
                if r.get("elec_tipo_doc") or r.get("elec_voltaje_kv"): tab_labels.append(t("electrico"))
                if r.get("instr_tags"): tab_labels.append(t("instrumentacion"))
                tab_labels.append(t("repuestos"))

                tabs=st.tabs(tab_labels); ti=0

                # RESUMEN - todos los datos no vacios
                with tabs[ti]:
                    todos = {
                        "Tag":            r.get("tag"),
                        t("nombre"):         r.get("spec_nombre_equipo") or r.get("motor_descripcion"),
                        "Tipo":           r.get("tipo_descripcion") or r.get("tipo_codigo"),
                        "Área":           r.get("area_descripcion") or r.get("area_codigo"),
                        t("fabricante"):     r.get("mec_fabricante") or r.get("motor_proveedor"),
                        t("modelo"):         r.get("mec_modelo") or r.get("motor_modelo"),
                        t("capacidad"):      r.get("spec_capacidad") or r.get("mec_capacidad"),
                        t("funcion"):        r.get("spec_funcion"),
                        "Input desde":    r.get("spec_input_desde"),
                        "Output hacia":   r.get("spec_output_hacia"),
                        "Flujo de masa":  r.get("spec_flujo_masa"),
                        t("material"):       r.get("mec_material"),
                        t("presion"):        r.get("mec_presion_bar"),
                        t("temperatura"):    r.get("mec_temperatura"),
                        "Peso":           f"{r.get('mec_peso_kg')} kg" if r.get("mec_peso_kg") else None,
                        t("norma"):          r.get("mec_norma"),
                        t("potencia"): f"{r['motor_kw']} kW" if pd.notna(r.get("motor_kw")) else None,
                        "Fab. motor":     r.get("motor_fabricante"),
                        "Modelo motor":   r.get("motor_modelo"),
                        t("voltaje"):        f"{r.get('motor_volt')} V" if r.get("motor_volt") else None,
                        t("corriente"):  f"{r['motor_fla_a']} A" if pd.notna(r.get("motor_fla_a")) else None,
                        t("arranque"):       r.get("motor_arranque"),
                        t("tablero"):        r.get("motor_feeder"),
                        "RPM":            r.get("motor_rpm"),
                        "Polos":          r.get("motor_polos"),
                        "Factor potencia":f"{r['motor_pf_pct']}%" if pd.notna(r.get("motor_pf_pct")) else None,
                        "Eficiencia":     f"{r['motor_eff_pct']}%" if pd.notna(r.get("motor_eff_pct")) else None,
                        "Diseño motor":   r.get("motor_diseno"),
                        "Tipo eléctrico": r.get("elec_tipo_doc"),
                        "Voltaje kV":     r.get("elec_voltaje_kv"),
                        "Fab. eléctrico": r.get("elec_fabricante"),
                        "Doc. eléctrico": r.get("elec_doc_num"),
                        "N° Datasheet":   r.get("mec_doc_numero"),
                        "PFD":            r.get("pfd_documento"),
                        "Páginas PFD":    r.get("paginas_pdf"),
                        "Instrumentos":   r.get("instr_tags"),
                        "Condiciones":    r.get("elec_params"),
                    }
                    datos = [(k,str(v)) for k,v in todos.items() if v and str(v).strip() not in ("","nan","None","0")]
                    c1,c2 = st.columns(2)
                    for i,(k,v) in enumerate(datos):
                        with (c1 if i%2==0 else c2):
                            st.markdown(f"<div style='padding:4px 0;border-bottom:1px solid #1a1a1a'><span style='font-size:11px;color:#444'>{k}</span><br><span style='font-size:13px;color:#e2e2e5'>{v}</span></div>",unsafe_allow_html=True)
                ti+=1

                if r.get("pfd_documento"):
                    with tabs[ti]:
                        c1,c2=st.columns(2)
                        with c1: st.caption("Documento"); st.code(r.get("pfd_documento",""),language=None)
                        with c2: t("pag_label"); st.write(r.get("paginas_pdf","—"))
                    ti+=1

                if r.get("spec_nombre_equipo") or r.get("mec_doc_numero") or r.get("mec_fabricante"):
                    with tabs[ti]:
                        campos=[
                            ("Nombre",r.get("spec_nombre_equipo") or r.get("motor_descripcion")),
                            ("Fabricante",r.get("mec_fabricante") or r.get("motor_proveedor")),
                            ("Modelo",r.get("mec_modelo")),
                            ("Capacidad",r.get("spec_capacidad") or r.get("mec_capacidad")),
                            ("Función",r.get("spec_funcion")),
                            ("Input desde",r.get("spec_input_desde")),
                            ("Output hacia",r.get("spec_output_hacia")),
                            ("Flujo de masa",r.get("spec_flujo_masa")),
                            ("Material",r.get("mec_material")),
                            ("Presión",r.get("mec_presion_bar")),
                            ("Temperatura",r.get("mec_temperatura")),
                            ("Peso",f"{r.get('mec_peso_kg')} kg" if r.get("mec_peso_kg") else None),
                            ("Norma",r.get("mec_norma")),
                            ("N° Datasheet",r.get("mec_doc_numero")),
                        ]
                        datos=[(k,v) for k,v in campos if v and str(v) not in ("","nan","None")]
                        c1,c2=st.columns(2)
                        for i,(k,v) in enumerate(datos):
                            with (c1 if i%2==0 else c2): st.caption(k); st.write(str(v))
                    ti+=1

                if pd.notna(r.get("motor_kw")):
                    with tabs[ti]:
                        mc=[
                            ("Fabricante",r.get("motor_fabricante")),
                            ("Modelo",r.get("motor_modelo")),
                            ("Potencia",f"{r['motor_kw']} kW" if pd.notna(r.get("motor_kw")) else None),
                            ("Voltaje",f"{r.get('motor_volt')} V" if r.get("motor_volt") else None),
                            ("Corriente FLA",f"{r['motor_fla_a']} A" if pd.notna(r.get("motor_fla_a")) else None),
                            ("Arranque",r.get("motor_arranque")),
                            ("Tablero",r.get("motor_feeder")),
                            ("RPM",r.get("motor_rpm")),
                            ("Polos",r.get("motor_polos")),
                            ("Factor potencia",f"{r['motor_pf_pct']}%" if pd.notna(r.get("motor_pf_pct")) else None),
                            ("Eficiencia",f"{r['motor_eff_pct']}%" if pd.notna(r.get("motor_eff_pct")) else None),
                            ("Diseño",r.get("motor_diseno")),
                            ("Proveedor",r.get("motor_proveedor")),
                        ]
                        datos=[(k,v) for k,v in mc if v and str(v) not in ("","nan","None")]
                        c1,c2,c3=st.columns(3)
                        for i,(k,v) in enumerate(datos):
                            with [c1,c2,c3][i%3]: st.caption(k); st.write(str(v))
                    ti+=1

                if r.get("elec_tipo_doc") or r.get("elec_voltaje_kv"):
                    with tabs[ti]:
                        ec=[
                            ("Tipo",r.get("elec_tipo_doc")),
                            ("Voltaje",r.get("elec_voltaje_v") or r.get("elec_voltaje_kv")),
                            ("Corriente",r.get("elec_corriente_a")),
                            ("Potencia",r.get("elec_potencia_kw") or r.get("elec_potencia_kva")),
                            ("Fabricante",r.get("elec_fabricante")),
                            ("N° Documento",r.get("elec_doc_num")),
                            ("IP",r.get("elec_ip")),
                        ]
                        datos=[(k,v) for k,v in ec if v and str(v) not in ("","nan","None")]
                        c1,c2=st.columns(2)
                        for i,(k,v) in enumerate(datos):
                            with (c1 if i%2==0 else c2): st.caption(k); st.write(str(v))
                        if r.get("elec_params"):
                            st.caption(t('cond_inst')); st.write(r["elec_params"])
                    ti+=1

                if r.get("instr_tags"):
                    with tabs[ti]:
                        st.caption(f"Instrumentos asociados ({r.get('instr_cantidad','')})")
                        chips=[x.strip() for x in str(r["instr_tags"]).split(",")]
                        cols=st.columns(min(len(chips),4))
                        for i,chip in enumerate(chips):
                            with cols[i%4]: st.code(chip,language=None)
                    ti+=1

                with tabs[ti]:
                    tipo_eq = r.get("tipo_codigo","")
                    try:
                        rep = qdf("""
                            SELECT r.codigo_q, r.leaf_class, r.descripcion,
                                   r.unidad, r.lead_time, r.status
                            FROM repuestos_q r
                            JOIN repuesto_tipo_equipo rte ON r.codigo_q=rte.codigo_q
                            WHERE rte.tipo_equipo=?
                            ORDER BY r.leaf_class, r.codigo_q
                        """, (tipo_eq,))
                    except:
                        rep = qdf("SELECT codigo_q,leaf_class,descripcion,unidad FROM repuestos_q LIMIT 0")

                    if len(rep) > 0:
                        st.caption(f"{len(rep)} repuestos disponibles para tipo {tipo_eq} — PosAppia Argentina")
                        cats = ["Todas"] + sorted(rep["leaf_class"].dropna().unique().tolist())
                        cat_sel = st.selectbox("Categoria",cats,label_visibility="collapsed",key=f"rc_{tag_s}")
                        if cat_sel != "Todas":
                            rep = rep[rep["leaf_class"]==cat_sel]
                        busq_r = st.text_input("",placeholder="Buscar repuesto...",label_visibility="collapsed",key=f"rb_{tag_s}")
                        if busq_r:
                            rep = rep[rep.descripcion.fillna("").str.lower().str.contains(busq_r.lower())|
                                      rep.leaf_class.fillna("").str.lower().str.contains(busq_r.lower())|
                                      rep.codigo_q.fillna("").str.lower().str.contains(busq_r.lower())]
                        st.dataframe(
                            rep[["codigo_q","leaf_class","descripcion","unidad","lead_time","status"]].rename(columns={
                                "codigo_q":"Código Q","leaf_class":"Tipo",
                                "descripcion":"Especificación","unidad":"Ud",
                                "lead_time":"Lead Time","status":"Estado"}),
                            use_container_width=True,hide_index=True,height=280)
                    else:
                        st.info(t("sin_repuestos"))

                if PUEDE_EDITAR:
                    if st.button(t("crear_ot"),type="primary"):
                        st.session_state["nueva_ot_tag"]=tag_s
                        st.session_state["pagina"]="ots"
                        st.rerun()

# ════════════════════════════════════
# ORDENES DE TRABAJO
# ════════════════════════════════════
elif pagina == "ots":
    st.markdown(f"## {t('ots')}")
    ots_df=qdf("SELECT * FROM ordenes_trabajo ORDER BY fecha_creacion DESC")

    c1,c2,c3=st.columns(3)
    with c1: f_est=st.multiselect(t("estado"),[t("pendiente"),t("en_curso"),t("completada"),t("cancelada")],default=[t("pendiente"),t("en_curso")])
    with c2: f_pri=st.multiselect(t("prioridad"),[t("critica"),t("alta"),t("normal"),t("baja")])
    with c3: f_tip=st.multiselect(t("tipo"),[t("correctivo"),t("preventivo"),t("inspeccion"),t("lubricacion"),t("calibracion")])

    res_ot=ots_df.copy()
    if f_est: res_ot=res_ot[res_ot.estado.isin(f_est)]
    if f_pri: res_ot=res_ot[res_ot.prioridad.isin(f_pri)]
    if f_tip: res_ot=res_ot[res_ot.tipo_tarea.isin(f_tip)]

    if PUEDE_EDITAR:
        with st.expander(t("nueva_ot")):
            busq_ot = st.text_input(t("buscar_equipo"),
                placeholder=t("buscar"),
                value=st.session_state.get("nueva_ot_busq",""),
                key="busq_ot_live")
            st.session_state["nueva_ot_busq"] = busq_ot

            if busq_ot and len(busq_ot) >= 1:
                eq_all = qdf("SELECT tag,spec_nombre_equipo,motor_descripcion,area_descripcion FROM equipos")
                mask = (eq_all.tag.str.lower().str.contains(busq_ot.lower(),na=False)|
                        eq_all.spec_nombre_equipo.fillna("").str.lower().str.contains(busq_ot.lower())|
                        eq_all.motor_descripcion.fillna("").str.lower().str.contains(busq_ot.lower())|
                        eq_all.area_descripcion.fillna("").str.lower().str.contains(busq_ot.lower()))
                sugs = eq_all[mask].head(6)
                for _,row in sugs.iterrows():
                    nm = row.get("spec_nombre_equipo") or row.get("motor_descripcion") or ""
                    if st.button(f"{row['tag']}  —  {nm[:45]}  ·  {row.get('area_descripcion','')}",
                                 key=f"pick_ot_{row['tag']}"):
                        st.session_state["nueva_ot_tag"] = row["tag"]
                        st.session_state["nueva_ot_busq"] = row["tag"]
                        st.rerun()

            eq_sel = st.session_state.get("nueva_ot_tag","")
            if eq_sel:
                eq_info = qdf("SELECT * FROM equipos WHERE tag=?", (eq_sel,))
                if len(eq_info) > 0:
                    r_eq = eq_info.iloc[0]
                    nm_eq = r_eq.get("spec_nombre_equipo") or r_eq.get("motor_descripcion") or r_eq.get("tipo_descripcion") or ""
                    st.markdown(f"""
                    <div style="background:#0f2a1e;border:1px solid #166534;border-radius:6px;
                                padding:10px 14px;margin:6px 0 12px">
                        <span style="font-family:monospace;color:#4ade80;font-weight:600">{eq_sel}</span>
                        <span style="color:#888;margin-left:10px">{nm_eq}</span>
                        <span style="color:#444;margin-left:8px;font-size:11px">· {r_eq.get("area_descripcion","")}</span>
                    </div>
                    """, unsafe_allow_html=True)

                    with st.form("form_ot"):
                        c1,c2 = st.columns(2)
                        with c1:
                            tipo_ot = st.selectbox(t("tipo"),
                                ["correctivo","preventivo","inspeccion","lubricacion","calibracion"],
                                format_func=lambda x: t(x))
                            prio_ot = st.selectbox(t("prioridad"),
                                ["normal","baja","alta","critica"],
                                format_func=lambda x: t(x))
                            horas   = st.number_input(t("horas_est"),min_value=0.0,step=0.5)
                            parada  = st.checkbox(t("req_parada"))
                        with c2:
                            tit_ot  = st.text_input(t("titulo"),
                                value=f"{t(tipo_ot) if tipo_ot else ''} — {nm_eq[:35]}" if nm_eq else "")
                            desc_ot = st.text_area(t("descripcion"),height=80,
                                placeholder="Describí qué hay que hacer, síntoma, causa probable...")
                            causa   = st.text_input(t("causa_sintoma"),
                                placeholder="Ej: fuga en sello, vibración anormal...")

                        if st.form_submit_button(t("crear"),type="primary",use_container_width=True):
                            if tit_ot:
                                num = "OT-"+datetime.now().strftime("%Y%m")+"-"+"".join(random.choices(sm.digits,k=3))
                                run("""INSERT INTO ordenes_trabajo
                                    (numero_ot,tag_equipo,titulo,descripcion,causa_falla,
                                     tipo_tarea,prioridad,estado,requiere_parada,
                                     horas_estimadas,creado_por,fecha_inicio)
                                    VALUES (?,?,?,?,?,?,?,'pendiente',?,?,?,?)""",
                                    (num,eq_sel,tit_ot,desc_ot,causa or None,
                                     tipo_ot,prio_ot,int(parada),horas or None,
                                     st.session_state.get("user_id",1),
                                     datetime.now().strftime("%Y-%m-%d")))
                                st.success(f"OT {num} creada para {eq_sel}")
                                for k in ["nueva_ot_tag","nueva_ot_busq"]:
                                    st.session_state.pop(k,None)
                                st.rerun()
                            else:
                                st.error("El título es obligatorio")
            elif not busq_ot:
                st.caption(t('buscar_arriba'))

    st.caption(f"{len(res_ot)} órdenes")
    for _,ot in res_ot.iterrows():
        with st.expander(f"{PRIO_IC.get(ot['prioridad'],'')}  {ot['numero_ot']}  ·  {ot['tag_equipo']}  ·  {ot['titulo']}"):
            c1,c2,c3=st.columns(3)
            with c1:
                st.markdown(f"**Estado:** {ot['estado']}")
                st.markdown(f"**Tipo:** {ot['tipo_tarea']}")
                st.markdown(f"**Prioridad:** {ot['prioridad']}")
            with c2:
                st.markdown(f"**Inicio:** {str(ot.get('fecha_inicio','—') or '—')[:10]}")
                st.markdown(f"**Horas:** {ot.get('horas_estimadas','—')} est. / {ot.get('horas_reales','—')} reales")
            with c3:
                if ot.get("requiere_parada"): st.warning("Requiere parada")
                if ot.get("descripcion"): st.write(ot["descripcion"])
            # Boton imprimir OT
            if st.button(t("imprimir_ot"), key=f"print_ot_{ot['id']}"):
                html = generar_html_imprimible("Orden de Trabajo", {
                    "N OT":          ot["numero_ot"],
                    "Equipo Tag":    ot["tag_equipo"],
                    "Titulo":        ot["titulo"],
                    "Tipo":          ot["tipo_tarea"],
                    "Prioridad":     ot["prioridad"],
                    "Estado":        ot["estado"],
                    "Fecha inicio":  str(ot.get("fecha_inicio","") or ""),
                    "Horas estimadas":str(ot.get("horas_estimadas","") or ""),
                    "Requiere parada":"Si" if ot.get("requiere_parada") else "No",
                    "Descripcion":   str(ot.get("descripcion","") or ""),
                })
                st.components.v1.html(html, height=0, scrolling=False)

            if PUEDE_EDITAR:
                st.markdown("---")
                cu1,cu2=st.columns(2)
                with cu1:
                    n_est=st.selectbox(t("estado"),[t("pendiente"),t("en_curso"),t("completada"),t("cancelada")],
                        index=["pendiente","en_curso","completada","cancelada"].index(ot["estado"]),
                        key=f"est_{ot['id']}")
                with cu2:
                    h_real=st.number_input("Horas reales",min_value=0.0,step=0.5,
                                            value=float(ot.get("horas_reales") or 0),key=f"hr_{ot['id']}")
                if st.button(t("actualizar"),key=f"upd_{ot['id']}"):
                    ffin=datetime.now().strftime("%Y-%m-%d") if n_est=="completada" else ot.get("fecha_fin")
                    run("UPDATE ordenes_trabajo SET estado=?,horas_reales=?,fecha_fin=?,updated_at=? WHERE id=?",
                        (n_est,h_real or None,ffin,datetime.now().strftime("%Y-%m-%d %H:%M"),ot["id"]))
                    st.success("Actualizado"); st.rerun()

# ════════════════════════════════════
# PLAN
# ════════════════════════════════════
elif pagina == "plan":
    st.markdown(f"## {t('plan')}")
    plan_df=qdf("SELECT * FROM plan_mantenimiento WHERE activo=1 ORDER BY proxima_fecha")

    c1,c2=st.columns(2)
    with c1: f_esp=st.multiselect(t("especialidad"),[t("mecanico"),t("electrico"),t("instrumentacion"),t("otro")])
    with c2: f_par=st.checkbox(t("solo_parada"))

    res_p=plan_df.copy()
    if f_esp: res_p=res_p[res_p.especialidad.isin(f_esp)]
    if f_par: res_p=res_p[res_p.requiere_parada==1]

    # Nueva tarea preventiva
    if PUEDE_EDITAR:
        with st.expander(t("nueva_tarea_prev")):
            busq_plan = st.text_input(t("buscar_equipo"),
                placeholder=t("buscar"),
                key="busq_plan_live")

            if busq_plan and len(busq_plan) >= 1:
                eq_all = qdf("SELECT tag,spec_nombre_equipo,motor_descripcion,area_descripcion FROM equipos")
                mask = (eq_all.tag.str.lower().str.contains(busq_plan.lower(),na=False)|
                        eq_all.spec_nombre_equipo.fillna("").str.lower().str.contains(busq_plan.lower())|
                        eq_all.area_descripcion.fillna("").str.lower().str.contains(busq_plan.lower()))
                sugs = eq_all[mask].head(6)
                for _,row in sugs.iterrows():
                    nm = row.get("spec_nombre_equipo") or row.get("motor_descripcion") or ""
                    if st.button(f"{row['tag']}  —  {nm[:45]}  ·  {row.get('area_descripcion','')}",
                                 key=f"pick_plan_{row['tag']}"):
                        st.session_state["plan_eq_sel"] = row["tag"]
                        st.rerun()

            plan_eq = st.session_state.get("plan_eq_sel","")
            if plan_eq:
                eq_info = qdf("SELECT * FROM equipos WHERE tag=?", (plan_eq,))
                if len(eq_info) > 0:
                    r_eq = eq_info.iloc[0]
                    nm_eq = r_eq.get("spec_nombre_equipo") or r_eq.get("motor_descripcion") or ""
                    st.markdown(f"""
                    <div style="background:#0f2a1e;border:1px solid #166534;border-radius:6px;
                                padding:10px 14px;margin:6px 0 12px">
                        <span style="font-family:monospace;color:#4ade80;font-weight:600">{plan_eq}</span>
                        <span style="color:#888;margin-left:10px">{nm_eq}</span>
                    </div>""", unsafe_allow_html=True)

                    with st.form("form_plan_nuevo"):
                        c1,c2 = st.columns(2)
                        with c1:
                            p_desc  = st.text_input("Descripción de la tarea *",
                                placeholder="Ej: Cambio de sello mecánico, lubricación rodamientos...")
                            p_frec  = st.selectbox("Frecuencia",["dias","semanas","meses","años"],index=2)
                            p_val   = st.number_input("Cada cuánto",min_value=1,value=1)
                            p_esp   = st.selectbox(t("especialidad"),
                                ["mecanico","electrico","instrumentacion","otro"],
                                format_func=lambda x: t(x))
                        with c2:
                            p_horas = st.number_input("Horas estimadas",min_value=0.0,step=0.5,value=2.0)
                            p_par   = st.checkbox(t("req_parada"))
                            p_hpar  = st.number_input("Duración parada (h)",min_value=0.0,step=0.5) if p_par else 0
                            p_proc  = st.text_area("Procedimiento",height=80,
                                placeholder="Pasos a seguir, herramientas necesarias...")

                        if st.form_submit_button("Agregar al plan",type="primary",use_container_width=True):
                            if p_desc:
                                from datetime import date, timedelta
                                dm = {"dias":1,"semanas":7,"meses":30,"años":365}
                                prox_n = (date.today()+timedelta(days=p_val*dm[p_frec])).isoformat()
                                run("""INSERT INTO plan_mantenimiento
                                    (tag_equipo,tipo_tarea,descripcion,frecuencia_tipo,frecuencia_valor,
                                     proxima_fecha,horas_estimadas,requiere_parada,duracion_parada_h,
                                     especialidad,procedimiento,activo,created_by)
                                    VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?)""",
                                    (plan_eq,"preventivo",p_desc,p_frec,p_val,prox_n,
                                     p_horas,int(p_par),p_hpar or None,p_esp,p_proc or None,
                                     st.session_state.get("user_id",1)))
                                st.success(f"Tarea agregada para {plan_eq}")
                                st.session_state.pop("plan_eq_sel",None)
                                st.rerun()
                            else:
                                st.error("La descripción es obligatoria")
            elif not busq_plan:
                st.caption(t('buscar_plan'))

    st.markdown("---")
    for _,p in res_p.iterrows():
        prox=str(p.get("proxima_fecha","—") or "—")[:10]
        par=" · parada" if p.get("requiere_parada") else ""
        with st.expander(f"{p['tag_equipo']}  ·  {p['descripcion']}  ·  {prox}{par}"):
            c1,c2=st.columns(2)
            with c1:
                st.markdown(f"**Frecuencia:** cada {p['frecuencia_valor']} {p['frecuencia_tipo']}")
                st.markdown(f"**Especialidad:** {p.get('especialidad','—')}")
                st.markdown(f"**Horas:** {p.get('horas_estimadas','—')} h")
            with c2:
                st.markdown(f"**Última:** {str(p.get('ultima_ejecucion','—') or '—')[:10]}")
                st.markdown(f"**Próxima:** {prox}")
            cp1, cp2 = st.columns(2)
            with cp1:
                if st.button(t("imprimir_tarea"), key=f"pprint_{p['id']}"):
                    html = generar_html_imprimible("Tarea de Mantenimiento Preventivo", {
                        "Equipo (Tag)": p["tag_equipo"],
                        "Descripción":  p["descripcion"],
                        "Especialidad": p.get("especialidad",""),
                        "Frecuencia":   f"Cada {p['frecuencia_valor']} {p['frecuencia_tipo']}",
                        "Próxima fecha":str(p.get("proxima_fecha","") or ""),
                        "Horas estimadas": str(p.get("horas_estimadas","") or ""),
                        "Requiere parada": "Sí" if p.get("requiere_parada") else "No",
                        "Procedimiento":p.get("procedimiento","") or "",
                    })
                    st.components.v1.html(html, height=0, scrolling=False)
            with cp2:
                pass
            if PUEDE_EDITAR and st.button(t("reg_hoy"),key=f"exec_{p['id']}"):
                hoy=date.today().isoformat()
                dm={"dias":1,"semanas":7,"meses":30,"años":365}
                prox_n=(date.today()+timedelta(days=p["frecuencia_valor"]*dm.get(p["frecuencia_tipo"],30))).isoformat()
                run("UPDATE plan_mantenimiento SET ultima_ejecucion=?,proxima_fecha=? WHERE id=?",(hoy,prox_n,p["id"]))
                num="OT-"+datetime.now().strftime("%Y%m")+"-"+"".join(random.choices(sm.digits,k=3))
                run("INSERT INTO ordenes_trabajo (numero_ot,tag_equipo,titulo,tipo_tarea,prioridad,estado,fecha_inicio,fecha_fin,horas_reales,creado_por) VALUES (?,?,?,?,?,'completada',?,?,?,?)",
                    (num,p["tag_equipo"],p["descripcion"],"preventivo","normal",hoy,hoy,p.get("horas_estimadas"),st.session_state.get("user_id",1)))
                st.success(f"Registrado · {num} · Próxima: {prox_n}"); st.rerun()

# ════════════════════════════════════
# ALERTAS
# ════════════════════════════════════
elif pagina == "alertas":
    st.markdown(f"## {t('alertas')}")
    al_df=qdf("SELECT * FROM alertas ORDER BY prioridad DESC,fecha_generada DESC")

    c1,c2=st.columns(2)
    with c1: f_est=st.multiselect(t("estado_lbl"),[t("activa"),t("vista"),t("resuelta")],default=[t("activa")])
    with c2: f_pri=st.multiselect(t("prioridad_lbl"),[t("critica"),t("alta"),t("normal")])

    res_al=al_df.copy()
    f_est_raw = [x for x in ["activa","vista","resuelta"] if t(x) in f_est]
    if f_est_raw: res_al=res_al[res_al.estado.isin(f_est_raw)]
    f_pri_raw = [x for x in ["critica","alta","normal","baja"] if t(x) in f_pri]
    if f_pri_raw: res_al=res_al[res_al.prioridad.isin(f_pri_raw)]

    for _,al in res_al.iterrows():
        msg=al.get(f"mensaje_{LANG}") or al.get("mensaje_es","")
        color={"critica":"#ef4444","alta":"#f59e0b"}.get(al["prioridad"],"#22c55e")
        st.markdown(f"""
        <div style="border-left:3px solid {color};padding:12px 16px;background:#111113;
                    border-radius:0 6px 6px 0;margin-bottom:8px">
            <div style="display:flex;justify-content:space-between">
                <span style="font-family:monospace;font-size:11px;font-weight:600;color:{color}">{al.get("tag_equipo","")}</span>
                <span style="font-size:11px;color:#333">{str(al.get("fecha_generada",""))[:10]}</span>
            </div>
            <div style="font-size:13px;color:#ccc;margin-top:3px">{msg}</div>
            <div style="font-size:11px;color:#333;margin-top:3px">{al.get("tipo_alerta","")} · {al.get("prioridad","")}</div>
        </div>
        """,unsafe_allow_html=True)
        if PUEDE_EDITAR:
            cb1,cb2,_=st.columns([1,1,5])
            with cb1:
                if al["estado"]=="activa":
                    if st.button("Vista",key=f"v_{al['id']}"):
                        run("UPDATE alertas SET estado='vista' WHERE id=?",(al["id"],)); st.rerun()
            with cb2:
                if al["estado"] in ["activa","vista"]:
                    if st.button(t("resolver"),key=f"r_{al['id']}"):
                        run("UPDATE alertas SET estado='resuelta' WHERE id=?",(al["id"],)); st.rerun()

# ════════════════════════════════════
# PARADAS
# ════════════════════════════════════
elif pagina == "paradas":
    st.markdown(f"## {t('paradas')}")
    par_df=qdf("SELECT * FROM paradas_planta ORDER BY fecha_inicio DESC")

    if PUEDE_EDITAR:
        with st.expander(t("programar_parada")):
            busq_par = st.text_input(t("buscar_equipos"),
                placeholder="Ej: 3200...", key="busq_par_live")
            equipos_par = st.session_state.get("parada_equipos",[])
            if busq_par and len(busq_par) >= 1:
                eq_all = qdf("SELECT tag,spec_nombre_equipo,motor_descripcion,area_descripcion FROM equipos")
                mask = (eq_all.tag.str.lower().str.contains(busq_par.lower(),na=False)|
                        eq_all.spec_nombre_equipo.fillna("").str.lower().str.contains(busq_par.lower())|
                        eq_all.area_descripcion.fillna("").str.lower().str.contains(busq_par.lower()))
                for _,row in eq_all[mask].head(5).iterrows():
                    nm = row.get("spec_nombre_equipo") or row.get("motor_descripcion") or ""
                    if row["tag"] not in equipos_par:
                        if st.button(f"+ {row['tag']}  —  {nm[:40]}", key=f"pick_par_{row['tag']}"):
                            equipos_par.append(row["tag"])
                            st.session_state["parada_equipos"] = equipos_par
                            st.rerun()
            if equipos_par:
                st.markdown("**Equipos:** " + "  ·  ".join([f"`{e}`" for e in equipos_par]))
                if st.button("Limpiar", key="limpiar_par"):
                    st.session_state.pop("parada_equipos",None); st.rerun()
            with st.form("form_par"):
                c1,c2=st.columns(2)
                with c1:
                    n_par=st.text_input("Nombre de la parada *", placeholder="Ej: Parada área 3200")
                    t_par=st.selectbox("Tipo",["programada","revision_anual","emergencia"])
                    a_par=st.text_input("Área (código)",placeholder="3200")
                with c2:
                    fi=st.date_input("Fecha inicio",value=date.today())
                    ff=st.date_input("Fecha fin",value=date.today())
                    d_par=st.text_area("Descripción",height=70)
                if st.form_submit_button(t("programar"),type="primary",use_container_width=True):
                    if n_par:
                        equipos_txt = ", ".join(st.session_state.get("parada_equipos",[]))
                        desc_final = f"{d_par}\nEquipos: {equipos_txt}" if equipos_txt else d_par
                        run("INSERT INTO paradas_planta (nombre,tipo,area_codigo,fecha_inicio,fecha_fin,duracion_h,descripcion,estado,responsable) VALUES (?,?,?,?,?,?,?,'planificada',?)",
                            (n_par,t_par,a_par or None,fi.isoformat(),ff.isoformat(),(ff-fi).days*24,desc_final,st.session_state.get("user_id",1)))
                        st.session_state.pop("parada_equipos",None)
                        st.success("Parada programada"); st.rerun()
                    else:
                        st.error("El nombre es obligatorio")

    for _,p in par_df.iterrows():
        with st.expander(f"{p['nombre']}  ·  {str(p['fecha_inicio'])[:10]}  ·  {p['estado']}"):
            c1,c2=st.columns(2)
            with c1:
                st.markdown(f"**Tipo:** {p['tipo']}")
                st.markdown(f"**Área:** {p.get('area_codigo','—') or '—'}")
                st.markdown(f"**Estado:** {p['estado']}")
            with c2:
                st.markdown(f"**Inicio:** {str(p['fecha_inicio'])[:10]}")
                st.markdown(f"**Fin:** {str(p['fecha_fin'])[:10]}")
                st.markdown(f"**Duración:** {p.get('duracion_h','?') } h")
            if p.get("descripcion"): st.write(p["descripcion"])
            if st.button(t("imprimir_parada"), key=f"parprint_{p['id']}"):
                html = generar_html_imprimible("Parada de Planta", {
                    "Nombre":        p["nombre"],
                    "Tipo":          p["tipo"],
                    "Área":          str(p.get("area_codigo","") or ""),
                    "Estado":        p["estado"],
                    "Fecha inicio":  str(p["fecha_inicio"]),
                    "Fecha fin":     str(p["fecha_fin"]),
                    "Duración":      f"{p.get('duracion_h','?')} horas",
                    "Descripción":   str(p.get("descripcion","") or ""),
                })
                st.components.v1.html(html, height=0, scrolling=False)
            if PUEDE_EDITAR:
                n_est=st.selectbox("Estado",["planificada","en_curso","completada","cancelada"],
                    index=["planificada","en_curso","completada","cancelada"].index(p["estado"]),
                    key=f"pest_{p['id']}")
                if st.button(t("actualizar"),key=f"pup_{p['id']}"):
                    run("UPDATE paradas_planta SET estado=? WHERE id=?",(n_est,p["id"])); st.rerun()

# ════════════════════════════════════
# PURCHASE REQUESTS
# ════════════════════════════════════
elif pagina == "pr":
    LANG_PR = st.session_state.get("lang","es")
    tit_pr = {"es":"Seguimiento de Pedidos — Mantenimiento CP2",
              "en":"Purchase Request Tracking — CP2 Maintenance",
              "ko":"구매 요청 추적 — CP2 유지보수"}
    st.markdown(f"## {tit_pr.get(LANG_PR,'PR')}")

    # ── Actualizar desde Excel ───────────────────────────────────────────────
    with st.expander(f"⬆ {t('act_archivo')} — Purchase Requests"):
        st.caption(t("arrastra"))
        pr_file = st.file_uploader("", type=["xlsx"], key="uploader_pr",
            label_visibility="collapsed")
        if pr_file is not None:
            import io as _io_pr, openpyxl as _opx_pr
            try:
                wb_pr = _opx_pr.load_workbook(_io_pr.BytesIO(pr_file.read()), data_only=True)
                ws_pr = wb_pr.active
                def _fmtd(v):
                    if v is None: return ''
                    if hasattr(v,'strftime'): return v.strftime('%Y-%m-%d')
                    s=str(v).strip()
                    if '/' in s:
                        try: p=s.split('/'); return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
                        except: pass
                    return '' if s in ('None','nan','','-') else s[:10]
                def _cln(v):
                    if v is None: return ''
                    s=str(v).strip(); return '' if s in ('None','nan','','-') else s
                run("DELETE FROM purchase_requests")
                est_map={'pending':'Pending','in progress':'In progress','in transit':'In transit',
                         'delivered':'Delivered','cancelled':'Cancelled'}
                cnt=0
                for row in range(9, ws_pr.max_row+1):
                    sol=_cln(ws_pr.cell(row,1).value)
                    if not sol: continue
                    tit=_cln(ws_pr.cell(row,3).value)
                    if not tit: continue
                    try: q_n=int(float(str(ws_pr.cell(row,6).value or ws_pr.cell(row,5).value or 0)))
                    except: q_n=0
                    est=_cln(ws_pr.cell(row,18).value)
                    est_n=est_map.get(est.lower().strip(),est) if est else 'In progress'
                    run("""INSERT INTO purchase_requests
                        (solicitante,solicitud,titulo,q_codes_n,fecha_solicitud,
                         fecha_aprobacion,fecha_pr,numero_pr,comprador,estado,
                         ultima_consulta,observaciones)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (sol,_cln(ws_pr.cell(row,2).value),tit,q_n,
                         _fmtd(ws_pr.cell(row,7).value),_fmtd(ws_pr.cell(row,8).value),
                         _fmtd(ws_pr.cell(row,10).value),_cln(ws_pr.cell(row,11).value),
                         _cln(ws_pr.cell(row,14).value),est_n,
                         _fmtd(ws_pr.cell(row,20).value),_cln(ws_pr.cell(row,21).value)))
                    cnt+=1
                st.success(f"{t('actualizado')} — {cnt} PRs cargadas"); st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")

    try:
        pr_df = qdf("SELECT * FROM purchase_requests ORDER BY id")
    except:
        run("""CREATE TABLE IF NOT EXISTS purchase_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitud TEXT, titulo TEXT, solicitante TEXT,
            fecha_solicitud TEXT, fecha_cotizacion TEXT, fecha_aprobacion TEXT,
            fecha_pr TEXT, numero_pr TEXT, comprador TEXT, estado TEXT,
            q_codes_n INTEGER DEFAULT 0, observaciones TEXT, ultima_consulta TEXT,
            created_at TEXT DEFAULT (datetime('now')))""")
        pr_df = qdf("SELECT * FROM purchase_requests ORDER BY id")

    # Metricas resumen
    if len(pr_df) > 0:
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: st.metric("Total", len(pr_df))
        with c2: st.metric("🔵 In progress", len(pr_df[pr_df.estado=="In progress"]))
        with c3: st.metric("🟡 Pending", len(pr_df[pr_df.estado=="Pending"]))
        with c4: st.metric("🟠 In transit", len(pr_df[pr_df.estado=="In transit"]))
        with c5: st.metric("✅ Delivered", len(pr_df[pr_df.estado=="Delivered"]))

    st.markdown("---")

    # ── TABLA ESTILO EXCEL ────────────────────────────────────────────────────
    # Encabezado de la tabla


    def est_badge(e):
        css = {"In progress":"est-prog","Pending":"est-pend",
               "In transit":"est-tran","Delivered":"est-delv","Cancelled":"est-canc"}
        return f'<span class="{css.get(e,"est-canc")}">{e}</span>'

    def fmt_d(d):
        return str(d or "")[:10] if d and str(d).strip() not in ("","None","nan") else ""

    if len(pr_df) > 0:
        # Filtro rapido
        busq_pr = st.text_input("",
            placeholder=t("filtrar_pr"),
            label_visibility="collapsed", key="busq_pr_tbl")

        res_pr = pr_df.copy()
        if busq_pr:
            res_pr = res_pr[
                res_pr.titulo.fillna("").str.lower().str.contains(busq_pr.lower()) |
                res_pr.solicitante.fillna("").str.lower().str.contains(busq_pr.lower()) |
                res_pr.numero_pr.fillna("").str.lower().str.contains(busq_pr.lower()) |
                res_pr.solicitud.fillna("").str.lower().str.contains(busq_pr.lower())]

        # Construir tabla HTML
        filas_html = ""
        for _, pr in res_pr.iterrows():
            def s(col, default=''):
                v = pr[col] if col in pr.index else default
                return str(v) if v and str(v) not in ('nan','None') else default
            filas_html += f"""
            <tr>
                <td style="white-space:nowrap">{s('solicitante')}</td>
                <td style="white-space:nowrap;color:#888">{s('solicitud')}</td>
                <td style="min-width:160px">{s('titulo')}</td>
                <td style="text-align:center">{s('q_codes_n','0')}</td>
                <td style="white-space:nowrap;color:#555">{fmt_d(s('fecha_solicitud'))}</td>
                <td style="white-space:nowrap;color:#555">{fmt_d(s('fecha_aprobacion'))}</td>
                <td style="white-space:nowrap;color:#555">{fmt_d(s('fecha_pr'))}</td>
                <td style="white-space:nowrap;color:#60a5fa;font-weight:500">{s('numero_pr')}</td>
                <td style="white-space:nowrap">{s('comprador')}</td>
                <td>{est_badge(s('estado'))}</td>
                <td style="white-space:nowrap;color:#555">{fmt_d(s('ultima_consulta'))}</td>
                <td style="min-width:180px;color:#777;font-size:10px">{s('observaciones')}</td>
            </tr>"""

        tabla_html = f"""
        <!DOCTYPE html><html><head><style>
        body{{margin:0;padding:0;background:transparent;font-family:Inter,Arial,sans-serif}}
        .pr-table{{width:100%;border-collapse:collapse;font-size:11px}}
        .pr-table th{{background:#1a1a1e;color:#888;padding:7px 9px;border:1px solid #222;
                     text-align:left;font-weight:500;white-space:nowrap}}
        .pr-table td{{padding:5px 9px;border:1px solid #1e1e22;color:#ccc;vertical-align:top}}
        .pr-table tr:nth-child(even) td{{background:#111113}}
        .pr-table tr:nth-child(odd) td{{background:#0f0f10}}
        .pr-table tr:hover td{{background:#161618}}
        .est-prog{{background:#1a2744;color:#60a5fa;padding:2px 6px;border-radius:3px;font-size:10px;white-space:nowrap}}
        .est-pend{{background:#2a1f0a;color:#fbbf24;padding:2px 6px;border-radius:3px;font-size:10px;white-space:nowrap}}
        .est-tran{{background:#1a1506;color:#f97316;padding:2px 6px;border-radius:3px;font-size:10px;white-space:nowrap}}
        .est-delv{{background:#0f2a1e;color:#4ade80;padding:2px 6px;border-radius:3px;font-size:10px;white-space:nowrap}}
        .est-canc{{background:#1a1a1e;color:#555;padding:2px 6px;border-radius:3px;font-size:10px;white-space:nowrap}}
        </style></head><body>
        <div style="overflow-x:auto">
        <table class="pr-table">
            <thead><tr>
                <th>{t("solicitante_lbl")}</th><th>{t("n_solicitud_lbl")}</th><th>Título</th>
                <th>Q codes</th><th>{t("f_solicitud_lbl")}</th><th>{t("f_aprobacion_lbl")}</th>
                <th>{t("f_pr_lbl")}</th><th>{t("n_pr_lbl")}</th><th>{t("comprador_lbl")}</th>
                <th>{t("estado_lbl")}</th><th>{t("ultima_cons_lbl")}</th><th>{t("obs_lbl")}</th>
            </tr></thead>
            <tbody>{filas_html}</tbody>
        </table>
        </div>
        </body></html>"""

        import streamlit.components.v1 as components
        height = max(200, min(len(res_pr) * 30 + 60, 600))
        components.html(tabla_html, height=height, scrolling=True)
        st.caption(f"{len(res_pr)} pedidos")
        with st.expander(f"ℹ️ Glosario / Glossary"):
            st.caption(f"**Q code:** {t('tip_q_code')}")
            st.caption(f"**PR:** {t('tip_pr')}")
            st.caption(f"**GoWorks:** {t('tip_goworks')}")

    # ── EDITAR / AGREGAR ──────────────────────────────────────────────────────
    st.markdown("---")
    tab_add, tab_edit = st.tabs([t("agregar_pedido"), t("editar_ped")])

    with tab_add:
        with st.form("form_pr_add"):
            c1,c2,c3 = st.columns(3)
            with c1:
                a_req  = st.text_input("Solicitante *")
                a_sol  = st.text_input(t("n_solicitud"))
                a_tit  = st.text_input("Título *")
                a_q    = st.number_input("Q codes encontrados", min_value=0, value=0)
            with c2:
                a_fec  = st.text_input("Fecha solicitud", placeholder="2026-06-XX")
                a_fap  = st.text_input("Fecha aprobación", placeholder="2026-06-XX")
                a_fpr  = st.text_input("Fecha PR", placeholder="2026-06-XX")
                a_npr  = st.text_input(t("n_pr"), placeholder="PR 2606XXX")
            with c3:
                a_comp = st.text_input(t("comprador"))
                a_est  = st.selectbox("Estado",
                    ["In progress","Pending","In transit","Delivered","Cancelled"])
                a_ult  = st.text_input(t("ultima_consulta"), placeholder="2026-06-XX")
                a_obs  = st.text_area(t("observaciones"), height=68)

            if st.form_submit_button("Agregar pedido", type="primary", use_container_width=True):
                if a_tit and a_req:
                    run("""INSERT INTO purchase_requests
                        (solicitud,titulo,solicitante,fecha_solicitud,fecha_aprobacion,
                         fecha_pr,numero_pr,comprador,estado,q_codes_n,
                         observaciones,ultima_consulta)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (a_sol,a_tit,a_req,a_fec,a_fap,a_fpr,a_npr,
                         a_comp,a_est,a_q,a_obs,a_ult))
                    st.success("Pedido agregado"); st.rerun()
                else:
                    st.error("Solicitante y título son obligatorios")

    with tab_edit:
        if len(pr_df) > 0:
            # Selector de PR a editar
            pr_opts = [f"{r['solicitud']} — {r['titulo'][:40]}" for _,r in pr_df.iterrows()]
            pr_sel_idx = st.selectbox(t("sel_pedido"),
                range(len(pr_opts)), format_func=lambda i: pr_opts[i],
                label_visibility="collapsed", key="pr_edit_sel")

            pr_sel = pr_df.iloc[pr_sel_idx]
            with st.form("form_pr_edit"):
                c1,c2,c3 = st.columns(3)
                with c1:
                    e_comp = st.text_input(t("comprador"), value=pr_sel.get("comprador","") or "")
                    e_npr  = st.text_input(t("n_pr"), value=pr_sel.get("numero_pr","") or "")
                    e_fpr  = st.text_input("Fecha PR", value=fmt_d(pr_sel.get("fecha_pr")))
                with c2:
                    e_est  = st.selectbox("Estado",
                        ["In progress","Pending","In transit","Delivered","Cancelled"],
                        index=["In progress","Pending","In transit","Delivered","Cancelled"].index(
                            pr_sel["estado"]) if pr_sel["estado"] in
                            ["In progress","Pending","In transit","Delivered","Cancelled"] else 0)
                    e_ult  = st.text_input(t("ultima_consulta"),
                        value=fmt_d(pr_sel.get("ultima_consulta")))
                with c3:
                    e_obs  = st.text_area(t("observaciones"),
                        value=pr_sel.get("observaciones","") or "", height=100)

                cc1,cc2 = st.columns([2,1])
                with cc1:
                    col_sv,col_del = st.columns([3,1])
                with col_sv:
                    if st.form_submit_button("Guardar cambios", type="primary",
                                             use_container_width=True):
                        run("""UPDATE purchase_requests
                            SET comprador=?,numero_pr=?,fecha_pr=?,estado=?,
                                ultima_consulta=?,observaciones=?
                            WHERE id=?""",
                            (e_comp,e_npr,e_fpr,e_est,e_ult,e_obs,pr_sel["id"]))
                        st.success("Guardado"); st.rerun()
                with cc2:
                    if st.form_submit_button("Eliminar pedido"):
                        run("DELETE FROM purchase_requests WHERE id=?", (pr_sel["id"],))
                        st.rerun()


# ════════════════════════════════════
# HISTORIAL
# ════════════════════════════════════
elif pagina == "historial":
    st.markdown(f"## {t('historial')}")
    ots_c=qdf("SELECT * FROM ordenes_trabajo WHERE estado='completada' ORDER BY fecha_fin DESC")

    if len(ots_c)==0:
        st.info(t("sin_resultados"))
    else:
        c1,c2,c3,c4=st.columns(4)
        with c1: st.metric(t("ots_completadas"),len(ots_c))
        with c2: st.metric(t("horas_totales"),f"{ots_c['horas_reales'].fillna(0).sum():.0f} h")
        with c3: st.metric(t("correctivos_n"),len(ots_c[ots_c.tipo_tarea=="correctivo"]))
        with c4: st.metric(t("preventivos_n"),len(ots_c[ots_c.tipo_tarea=="preventivo"]))

        c_g1,c_g2=st.columns(2)
        with c_g1:
            pt=ots_c["tipo_tarea"].value_counts().reset_index()
            pt.columns=["tipo","cantidad"]
            fig1=px.pie(pt,values="cantidad",names="tipo",height=260,
                        color_discrete_sequence=["#3b82f6","#22c55e","#f59e0b","#ef4444","#a78bfa"])
            fig1.update_layout(plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                               font=dict(family="Inter",size=11,color="#888"),
                               margin=dict(l=0,r=0,t=10,b=0))
            st.plotly_chart(fig1,use_container_width=True)
        with c_g2:
            pa=ots_c.merge(qdf("SELECT DISTINCT tag,area_codigo FROM equipos"),
                           left_on="tag_equipo",right_on="tag",how="left")
            pa=pa["area_codigo"].value_counts().head(8).reset_index()
            pa.columns=["area","cantidad"]
            fig2=px.bar(pa,x="area",y="cantidad",height=260,color_discrete_sequence=["#3b82f6"])
            fig2.update_layout(plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                               font=dict(family="Inter",size=11,color="#888"),
                               margin=dict(l=0,r=0,t=10,b=0),
                               xaxis=dict(color="#444"),yaxis=dict(color="#444",gridcolor="#1e1e22"))
            st.plotly_chart(fig2,use_container_width=True)
        st.dataframe(
            ots_c[["numero_ot","tag_equipo","titulo","tipo_tarea","fecha_inicio","fecha_fin","horas_reales"]],
            use_container_width=True,hide_index=True)

# ════════════════════════════════════
# USUARIOS
# ════════════════════════════════════
elif pagina == "usuarios" and ES_ADMIN:
    st.markdown(f"## {t('usuarios')}")
    users_df=qdf("SELECT * FROM usuarios ORDER BY activo DESC,rol,nombre")

    with st.expander(t("agregar_usuario")):
        with st.form("form_user"):
            c1,c2=st.columns(2)
            with c1:
                u_usr=st.text_input("Usuario *")
                u_nom=st.text_input("Nombre *")
                u_ape=st.text_input("Apellido *")
                u_mail=st.text_input("Email *")
                u_dni=st.text_input("DNI *")
            with c2:
                u_pue=st.text_input("Puesto *")
                u_area=st.text_input("Área de trabajo")
                u_tel=st.text_input("Teléfono")
                u_rol=st.selectbox("Rol",["tecnico","supervisor","directivo","invitado","admin"])
                u_lang=st.selectbox("Idioma",["es","en","ko"],format_func=lambda x:{"es":"Español","en":"English","ko":"한국어"}[x])
                u_pwd=st.text_input("Contraseña *",type="password")
                u_pwd2=st.text_input("Confirmar *",type="password")
            if st.form_submit_button("Crear usuario",type="primary"):
                if not all([u_usr,u_nom,u_ape,u_mail,u_dni,u_pue,u_pwd]):
                    st.error("Completá todos los campos obligatorios")
                elif u_pwd!=u_pwd2: st.error("Las contraseñas no coinciden")
                elif len(u_pwd)<6: st.error("Mínimo 6 caracteres")
                else:
                    try:
                        run("INSERT INTO usuarios (username,password_hash,nombre,apellido,email,dni,telefono,puesto,area_trabajo,rol,idioma,activo) VALUES (?,?,?,?,?,?,?,?,?,?,?,1)",
                            (u_usr.strip(),hp(u_pwd),u_nom,u_ape,u_mail,u_dni,u_tel or None,u_pue,u_area or None,u_rol,u_lang))
                        st.success(f"Usuario '{u_usr}' creado"); st.rerun()
                    except: st.error("Ese usuario ya existe")

    pendientes=users_df[users_df.activo==0]
    if len(pendientes)>0:
        st.markdown(f"#### {t('solicitudes')}")
        for _,u in pendientes.iterrows():
            with st.expander(f"{u['nombre']} {u['apellido']}  ·  {u['username']}"):
                st.markdown(f"**Email:** {u.get('email','—')}  |  **DNI:** {u.get('dni','—')}  |  **Puesto:** {u.get('puesto','—')}  |  **Área:** {u.get('area_trabajo','—')}")
                c1,c2,c3=st.columns(3)
                with c1: nr=st.selectbox("Rol",["tecnico","supervisor","directivo","invitado"],key=f"nr_{u['id']}")
                with c2:
                    if st.button(t("aprobar"),key=f"apr_{u['id']}",type="primary"):
                        run("UPDATE usuarios SET activo=1,rol=? WHERE id=?",(nr,u["id"])); st.success("Aprobado"); st.rerun()
                with c3:
                    if st.button(t("rechazar"),key=f"rej_{u['id']}"):
                        run("DELETE FROM usuarios WHERE id=?",(u["id"],)); st.rerun()

    # Configuracion de email
    if ES_ADMIN:
        st.markdown("---")
        st.markdown(f"#### {t('config_email')}")
        with st.expander("Configurar servidor de correo POSCO Argentina"):
            with st.form("form_email_config"):
                st.caption(t("config_email"))
                c1,c2 = st.columns(2)
                with c1:
                    smtp_user = st.text_input(t("email_rem"),
                        value=get_config("smtp_user",""),
                        placeholder="sistema@poscoargentina.com")
                    smtp_host = st.text_input(t("smtp_server"),
                        value=get_config("smtp_host","smtp.office365.com"))
                    smtp_port = st.text_input(t("smtp_port_lbl"),
                        value=get_config("smtp_port","587"))
                with c2:
                    smtp_pass = st.text_input("Contraseña",
                        type="password", placeholder="••••••••")
                    email_dest = st.text_input(t("email_dest_lbl"),
                        value=get_config("email_alertas",""),
                        placeholder="ian.herrera@poscoargentina.com, supervisor@poscoargentina.com")
                    st.caption("Las alertas criticas y altas se enviaran automaticamente a estos emails")

                if st.form_submit_button(t("guardar_config"), type="primary"):
                    run("INSERT OR REPLACE INTO configuracion VALUES (?,?,?)",
                        ("smtp_user", smtp_user, "Email remitente"))
                    run("INSERT OR REPLACE INTO configuracion VALUES (?,?,?)",
                        ("smtp_host", smtp_host, t("smtp_server")))
                    run("INSERT OR REPLACE INTO configuracion VALUES (?,?,?)",
                        ("smtp_port", smtp_port, "Puerto SMTP"))
                    run("INSERT OR REPLACE INTO configuracion VALUES (?,?,?)",
                        ("email_alertas", email_dest, "Emails destino alertas"))
                    if smtp_pass:
                        run("INSERT OR REPLACE INTO configuracion VALUES (?,?,?)",
                            ("smtp_pass", smtp_pass, "Password SMTP"))
                    st.success(t("config_guardada"))

                    # Test de conexion
                    if st.session_state.get("test_smtp"):
                        ok = enviar_alerta_email(
                            "TEST", "Test de conexion CP2 Maintenance System",
                            "normal", "Test")
                        st.success("Email enviado correctamente") if ok else st.error("Error enviando email - verificar credenciales")

    st.markdown(f"#### {t('usuarios_activos')} ({len(users_df[users_df.activo==1])})")
    for _,u in users_df[users_df.activo==1].iterrows():
        with st.expander(f"{u['nombre']} {u['apellido']}  ·  {u['username']}  ·  {u['rol']}"):
            st.markdown(f"**Email:** {u.get('email','—')}  |  **DNI:** {u.get('dni','—')}  |  **Puesto:** {u.get('puesto','—')}  |  **Área:** {u.get('area_trabajo','—')}")
            st.markdown(f"**Último acceso:** {str(u.get('last_login','Nunca') or 'Nunca')[:16]}")
            if u["username"]!=st.session_state.get("username"):
                cu1,cu2,cu3=st.columns(3)
                with cu1:
                    nr=st.selectbox("Rol",["tecnico","supervisor","directivo","invitado","admin"],
                        index=["tecnico","supervisor","directivo","invitado","admin"].index(u["rol"]) if u["rol"] in ["tecnico","supervisor","directivo","invitado","admin"] else 0,
                        key=f"rol_{u['id']}")
                with cu2: np_=st.text_input("Nueva contraseña",type="password",key=f"np_{u['id']}")
                with cu3:
                    if st.button(t("guardar"),key=f"sv_{u['id']}"):
                        run("UPDATE usuarios SET rol=? WHERE id=?",(nr,u["id"]))
                        if np_ and len(np_)>=6: run("UPDATE usuarios SET password_hash=? WHERE id=?",(hp(np_),u["id"]))
                        st.success("Guardado"); st.rerun()
                if st.button(t("desactivar"),key=f"dis_{u['id']}"):
                    run("UPDATE usuarios SET activo=0 WHERE id=?",(u["id"],)); st.rerun()

# ════════════════════════════════════
# MEDICIONES
# ════════════════════════════════════
elif pagina == "mediciones":
    import re as _re
    import io as _io
    LANG_M2 = st.session_state.get("lang","es")
    st.markdown(f"## {t('mediciones')}")

    tab_mec, tab_elec = st.tabs([f"⚙ {t('med_mec')}", f"⚡ {t('med_elec')}"])

    # ═════════════════ TAB MECÁNICAS ═════════════════
    with tab_mec:
        st.markdown(f"### {t('vibraciones')}")

        with st.expander(f"⬆ {t('act_archivo')} — Vibraciones"):
            st.caption(t("arrastra"))
            vib_file = st.file_uploader("", type=["xlsx"],
                key="uploader_vib", label_visibility="collapsed")
            if vib_file is not None:
                try:
                    wb_up = __import__('openpyxl').load_workbook(_io.BytesIO(vib_file.read()), data_only=True)
                    ws_up = wb_up['Sheet1']
                    from datetime import date as _date
                    TODAY2 = _date.today()
                    col_fechas2 = {}
                    mes_actual2 = 4
                    prev_dia2 = None
                    for col in range(7, ws_up.max_column+1):
                        v = ws_up.cell(4,col).value
                        if v is None: break
                        try: dia = int(float(str(v).strip()))
                        except: break
                        if prev_dia2 is not None and dia < prev_dia2:
                            mes_actual2 += 1
                        try:
                            f = _date(2026, mes_actual2, dia)
                            if f <= TODAY2: col_fechas2[col] = f.isoformat()
                            prev_dia2 = dia
                        except: pass
                        prev_dia2 = dia

                    TIPOS_NUM2 = ['Bearing Vib.','Agitator 1 Vib','Agitator 2 Vib','Bearing Temp.']
                    area2=tag2=nombre2=""
                    nuevos = 0
                    fechas_exist = set(r[0] for r in get_conn().execute("SELECT DISTINCT fecha FROM vibraciones").fetchall())

                    for row in range(5, ws_up.max_row+1):
                        v0=ws_up.cell(row,1).value; v1=ws_up.cell(row,2).value
                        v2=ws_up.cell(row,3).value; v3=ws_up.cell(row,4).value
                        if v0 and str(v0).strip() != 'None': area2=str(v0).strip()
                        if v1 and str(v1).strip() != 'None': nombre2=str(v1).strip()
                        if v3 and str(v3).strip():
                            raw=str(v3).strip()
                            m2=_re.match(r'^(\d{4})([A-Z]{2,3})(\d+.*)$',raw)
                            tag2=f"{m2.group(1)}-{m2.group(2)}-{m2.group(3)}" if m2 else raw
                        if not v2: continue
                        tipo2=str(v2).strip()
                        if tipo2 not in TIPOS_NUM2: continue
                        try: lim2=float(str(ws_up.cell(row,5).value or '4.5').replace('<','').replace('>','').strip())
                        except: lim2=4.5
                        unid2=str(ws_up.cell(row,6).value or '').strip()
                        for col,fecha in col_fechas2.items():
                            if fecha in fechas_exist: continue
                            val=ws_up.cell(row,col).value
                            if val is None: continue
                            val_s=str(val).strip().replace('\u3000','').replace(',','.')
                            if val_s in ('','None','Normal','Stand By','STAND BY','Non Taken','-','N/A','0'): continue
                            try:
                                val_n=float(val_s)
                                if val_n<=0: continue
                            except: continue
                            est2='critico' if val_n>lim2 else 'alerta' if val_n>lim2*0.8 else 'normal'
                            run("INSERT INTO vibraciones (tag,equipo_desc,tipo_medicion,valor,unidad,limite,semana,fecha,estado,area) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                (tag2,nombre2,tipo2,val_n,unid2,lim2,0,fecha,est2,area2))
                            nuevos += 1

                    if nuevos > 0:
                        st.success(f"{t('actualizado')} — {nuevos} mediciones nuevas")
                    else:
                        st.info("Sin datos nuevos — ya estaba actualizado")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # Cargar datos
        try:
            vib_df = qdf("SELECT tag,equipo_desc,tipo_medicion,valor,unidad,limite,fecha,semana,estado,area FROM vibraciones ORDER BY tag,tipo_medicion,fecha")
            vib_df["fecha"] = pd.to_datetime(vib_df["fecha"], errors="coerce")
        except:
            vib_df = pd.DataFrame()

        if len(vib_df) == 0:
            st.info("Sin datos de vibración")
        else:
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.metric("Equipos", vib_df["tag"].nunique())
            with c2: st.metric("Mediciones", len(vib_df))
            with c3: st.metric("🔴 Críticos", len(vib_df[vib_df.estado=="critico"]["tag"].unique()))
            with c4:
                mn = vib_df["fecha"].min()
                mx = vib_df["fecha"].max()
                if pd.notna(mn) and pd.notna(mx):
                    st.metric("Período", f"{mn.strftime('%d/%m')} → {mx.strftime('%d/%m')}")

            st.caption(f"ℹ️ {t('tip_mm_s')}")
            st.markdown("---")
            c1,c2,c3,c4 = st.columns([2,2,1,1])
            with c1:
                busq_vib = st.text_input("",placeholder=t("buscar"),label_visibility="collapsed",key="busq_vib")
            with c2:
                tags_vib = ["Todos"]+sorted(vib_df["tag"].unique().tolist())
                tag_sel  = st.selectbox("",tags_vib,label_visibility="collapsed",key="vib_tag_sel")
            with c3:
                tipos_vib= ["Todos"]+sorted(vib_df["tipo_medicion"].unique().tolist())
                tipo_sel = st.selectbox("",tipos_vib,label_visibility="collapsed",key="vib_tipo_sel")
            with c4:
                est_sel  = st.selectbox("",["Todos","critico","alerta","normal"],label_visibility="collapsed",key="vib_est_sel")

            res_vib = vib_df.copy()
            if busq_vib: res_vib=res_vib[res_vib.tag.str.lower().str.contains(busq_vib.lower(),na=False)|res_vib.equipo_desc.fillna("").str.lower().str.contains(busq_vib.lower())]
            if tag_sel !="Todos": res_vib=res_vib[res_vib.tag==tag_sel]
            if tipo_sel!="Todos": res_vib=res_vib[res_vib.tipo_medicion==tipo_sel]
            if est_sel !="Todos": res_vib=res_vib[res_vib.estado==est_sel]

            import plotly.express as px
            if tag_sel != "Todos":
                vib_eq=res_vib[res_vib.tipo_medicion.str.contains('Vib',na=False)].copy()
                if len(vib_eq)>0:
                    fig=px.line(vib_eq.sort_values("fecha"),x="fecha",y="valor",color="tipo_medicion",
                        markers=True,height=280,color_discrete_sequence=["#6366f1","#0ea5e9","#14b8a6"],
                        labels={"fecha":"Fecha","valor":"mm/s","tipo_medicion":"Punto"})
                    if vib_eq["limite"].notna().any():
                        lv=float(vib_eq["limite"].dropna().iloc[0])
                        fig.add_hline(y=lv,line_dash="dash",line_color="#ef4444",
                            annotation_text=f"Límite {lv} mm/s",annotation_font_color="#ef4444")
                    fig.update_layout(margin=dict(l=0,r=0,t=0,b=0),
                        plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Inter",size=11,color="#888"),
                        xaxis=dict(color="#444",gridcolor="#1e1e22"),
                        yaxis=dict(color="#444",gridcolor="#1e1e22"),
                        legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)"))
                    st.plotly_chart(fig,use_container_width=True)
                if st.button(f"Ver ficha de {tag_sel}",type="primary"):
                    st.session_state["eq_sel"]=tag_sel
                    st.session_state["pagina"]="equipos"
                    st.rerun()
            else:
                vib_fil=vib_df[vib_df.tipo_medicion.str.contains('Vib',na=False)].copy()
                if len(vib_fil)>0:
                    ultimo=vib_fil.loc[vib_fil.groupby("tag")["fecha"].idxmax()].copy()
                    ultimo=ultimo.sort_values("valor",ascending=False).head(20).reset_index(drop=True)
                    ultimo["label"]=ultimo["tag"]+" — "+ultimo["equipo_desc"].str[:18]
                    fig2=px.bar(ultimo,x="label",y="valor",color="estado",height=300,
                        color_discrete_map={"critico":"#ef4444","alerta":"#f59e0b","normal":"#22c55e"},
                        labels={"label":"","valor":"mm/s"})
                    fig2.add_hline(y=4.5,line_dash="dash",line_color="#ef4444",annotation_text="4.5 mm/s")
                    fig2.update_layout(margin=dict(l=0,r=0,t=10,b=80),
                        plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Inter",size=10,color="#888"),
                        xaxis=dict(color="#444",tickangle=45,tickfont=dict(size=9)),
                        yaxis=dict(color="#444",gridcolor="#1e1e22",title="mm/s"),
                        legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)",orientation="h",y=1.1))
                    st.plotly_chart(fig2,use_container_width=True)

            show_vib=res_vib[["tag","equipo_desc","tipo_medicion","valor","unidad","limite","fecha","estado"]].copy()
            show_vib["fecha"]=pd.to_datetime(show_vib["fecha"],errors="coerce").dt.strftime("%d/%m/%Y")
            st.dataframe(show_vib.rename(columns={"tag":"Tag","equipo_desc":"Equipo",
                "tipo_medicion":"Medición","valor":"Valor","unidad":"Ud",
                "limite":"Límite","fecha":"Fecha","estado":"Estado"}),
                use_container_width=True,hide_index=True,height=280)

            with st.expander(t("reg_nueva_med")):
                with st.form("form_vib"):
                    c1,c2,c3=st.columns(3)
                    with c1:
                        eq_all=qdf("SELECT tag FROM equipos ORDER BY tag")
                        new_tag=st.selectbox("Equipo *",eq_all["tag"].tolist(),key="new_vib_tag")
                        new_tipo=st.selectbox(t("tipo_med"),["Bearing Vib.","Agitator 1 Vib","Agitator 2 Vib","Bearing Temp."])
                    with c2:
                        new_val=st.number_input("Valor *",min_value=0.0,step=0.1)
                        new_unit=st.selectbox("Unidad",["mm/s","℃","-"])
                        new_lim=st.number_input("Límite",min_value=0.0,step=0.1,value=4.5)
                    with c3:
                        new_fecha=st.date_input("Fecha",value=date.today())
                        new_area=st.text_input("Área",placeholder="3100, 3200...")
                    if st.form_submit_button("Registrar",type="primary"):
                        est_n='critico' if new_lim>0 and new_val>new_lim else 'alerta' if new_lim>0 and new_val>new_lim*0.8 else 'normal'
                        run("INSERT INTO vibraciones (tag,equipo_desc,tipo_medicion,valor,unidad,limite,semana,fecha,estado,area) VALUES (?,?,?,?,?,?,?,?,?,?)",
                            (new_tag,new_tag,new_tipo,new_val,new_unit,new_lim,0,new_fecha.isoformat(),est_n,new_area))
                        st.success(f"Registrado — {new_tag} {new_fecha}: {new_val} {new_unit}")
                        st.rerun()

    # ═════════════════ TAB ELÉCTRICAS ═════════════════
    with tab_elec:
        st.markdown(f"### {t('med_elec')}")

        # Botón actualizar
        with st.expander(f"⬆ {t('act_archivo')} — Eléctrico"):
            st.caption(t("arrastra"))
            elec_tipo = st.radio("Tipo", ["Heat Trace","PLC diario","Inspección semanal"],
                horizontal=True, key="elec_tipo")
            elec_file = st.file_uploader("", type=["xlsx"],
                key="uploader_elec", label_visibility="collapsed")
            if elec_file is not None:
                import io as _io2
                import openpyxl as _opx
                import re as _re2
                try:
                    wb_up = _opx.load_workbook(_io2.BytesIO(elec_file.read()), data_only=True)
                    nuevos = 0

                    if elec_tipo == "Heat Trace":
                        PANELES_UP = {1:("0000-EHT-001","Liming Plant"),4:("0000-EHT-002","LC Plant"),
                            7:("0000-EHT-003","LC Plant"),10:("0000-EHT-004","Water Plant"),
                            13:("0000-EHT-005","Truck Shop"),16:("0000-EHT-006","6700 Plant")}
                        fechas_ex = set(r[0] for r in get_conn().execute("SELECT DISTINCT fecha FROM mediciones_heat_trace").fetchall())
                        for sh2 in wb_up.sheetnames:
                            if sh2 == 'Data 위치': continue
                            try: p2=sh2.strip().split('.'); fecha2=f"20{p2[0]}-{p2[1]}-{p2[2]}"
                            except: continue
                            if fecha2 in fechas_ex: continue
                            ws2 = wb_up[sh2]
                            for col_i,(panel,ubic) in PANELES_UP.items():
                                volt_f=None; corr_f=None
                                try: volt_f=float(ws2.cell(4,col_i+1).value)
                                except: pass
                                try: corr_f=float(ws2.cell(5,col_i+1).value)
                                except: pass
                                mc=str(ws2.cell(6,col_i+1).value or '')
                                hora=str(ws2.cell(7,col_i+1).value or '')
                                for row2 in range(10,40):
                                    cct_l=ws2.cell(row2,col_i).value
                                    cct_v=ws2.cell(row2,col_i+1).value
                                    if not cct_l: continue
                                    amp_f=None
                                    try: amp_f=float(cct_v)
                                    except: pass
                                    run("INSERT INTO mediciones_heat_trace (fecha,panel,ubicacion,voltaje_v,corriente_a,mc_estado,hora_insp,cct,amperaje_a) VALUES (?,?,?,?,?,?,?,?,?)",
                                        (fecha2,panel,ubic,volt_f,corr_f,mc,hora,str(cct_l).strip(),amp_f))
                                    nuevos+=1

                    elif elec_tipo == "PLC diario":
                        fechas_ex = set(r[0] for r in get_conn().execute("SELECT DISTINCT fecha FROM mediciones_plc").fetchall())
                        ANOM_ON={'MAINT','BAF','BATT1F','BATT2F','INTF','BUS1F','BUS5F','BUS8F','STOP'}
                        ANOM_OFF={'RUN','DC5V','DC24V','FMR','Power On'}
                        BLOQUES_UP=[
                            (7,24,[("2100-PCS7-001 (A)",1,2),("2100-PCS7-001 (A)",3,4),
                                   ("2100-PCS7-001 (B)",6,7),("2100-PCS7-001 (B)",8,9),
                                   ("3000-PCS7-001 (A)",11,12),("3000-PCS7-001 (A)",13,14),
                                   ("3000-PCS7-001 (B)",16,17),("3000-PCS7-001 (B)",18,19)]),
                            (30,44,[("4000-PCS7-001 (A)",1,2),("4000-PCS7-001 (A)",3,4),
                                    ("4000-PCS7-001 (B)",6,7),("4000-PCS7-001 (B)",8,9),
                                    ("6000-PCS7-001 (A)",11,12),("6000-PCS7-001 (A)",13,14),
                                    ("6000-PCS7-001 (B)",16,17),("6000-PCS7-001 (B)",18,19)])]
                        for sh2 in wb_up.sheetnames:
                            if sh2=='Data 위치': continue
                            try: p2=sh2.strip().split('.'); fecha2=f"20{p2[0]}-{p2[1]}-{p2[2]}"
                            except: continue
                            if fecha2 in fechas_ex: continue
                            ws2=wb_up[sh2]
                            mods={}
                            for rh in [4,27]:
                                for c in range(1,20):
                                    v=ws2.cell(rh,c).value
                                    if v and str(v).strip() in ('Power Module','CPU Module'):
                                        mods[c]=str(v).strip()
                            for fi,ff,plcs2 in BLOQUES_UP:
                                for pid,cn,ce in plcs2:
                                    mod=mods.get(cn,'Unknown')
                                    for row2 in range(fi,ff+1):
                                        lamp=ws2.cell(row2,cn).value; est=ws2.cell(row2,ce).value
                                        if not lamp or str(lamp).strip() in ('Name','Lamp','None',''): continue
                                        ls=str(lamp).strip(); es=str(est).strip() if est else ''
                                        anom=1 if (ls in ANOM_ON and es in ('O','ON','on','On','o')) else                                              1 if (ls in ANOM_OFF and es not in ('O','ON','on','On','o')) else 0
                                        run("INSERT INTO mediciones_plc (fecha,plc_id,modulo,lampara,estado,es_anomalia) VALUES (?,?,?,?,?,?)",
                                            (fecha2,pid,mod,ls,es,anom)); nuevos+=1
                    else:
                        fechas_ex=set(r[0] for r in get_conn().execute("SELECT DISTINCT fecha FROM mediciones_semanal").fetchall())
                        for sh2 in wb_up.sheetnames:
                            if sh2 in ('Rangos',): continue
                            try: p2=sh2.strip().split('.'); fecha_h=f"20{p2[0]}-{p2[1]}-{p2[2]}"
                            except: continue
                            ws2=wb_up[sh2]
                            for row2 in range(4,ws2.max_row+1):
                                tag_e=ws2.cell(row2,2).value
                                if not tag_e: continue
                                fec_e=ws2.cell(row2,4).value
                                fecha_r=fec_e.strftime('%Y-%m-%d') if hasattr(fec_e,'strftime') else fecha_h
                                if fecha_r in fechas_ex: continue
                                val_e=ws2.cell(row2,6).value; val_s=str(val_e).strip() if val_e else None
                                val_n=None
                                if val_e is not None:
                                    try: val_n=float(val_e)
                                    except:
                                        m2=_re2.search(r'[-+]?\d*\.?\d+',str(val_e))
                                        if m2:
                                            try: val_n=float(m2.group())
                                            except: pass
                                run("INSERT INTO mediciones_semanal (fecha,tag,descripcion,estado_alim,valor,valor_num,observacion,tipo) VALUES (?,?,?,?,?,?,?,?)",
                                    (fecha_r,str(tag_e).strip(),
                                     str(ws2.cell(row2,3).value or '').strip(),
                                     str(ws2.cell(row2,5).value or '').strip() if ws2.cell(row2,5).value else None,
                                     val_s,val_n,
                                     str(ws2.cell(row2,7).value or '').strip() if ws2.cell(row2,7).value else None,
                                     str(ws2.cell(row2,9).value or '').strip() if ws2.cell(row2,9).value else None))
                                nuevos+=1

                    st.success(f"{t('actualizado')} — {nuevos} registros nuevos") if nuevos>0 else st.info("Sin datos nuevos")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

        # Sub-tabs
        sub_ht, sub_plc, sub_sem = st.tabs([
            "🔌 Heat Trace", f"🖥 {t('plc_diario')}", f"📋 {t('insp_semanal')}"])

        # ─── HEAT TRACE ───────────────────────────────────────────────────────
        with sub_ht:
            try:
                ht_df = qdf("SELECT fecha,panel,ubicacion,voltaje_v,corriente_a,mc_estado,hora_insp,cct,amperaje_a FROM mediciones_heat_trace ORDER BY fecha,panel,cct")
                if len(ht_df)==0:
                    st.info("Sin datos de Heat Trace")
                else:
                    c1,c2,c3,c4 = st.columns(4)
                    with c1: st.metric("Días registrados", ht_df["fecha"].nunique())
                    with c2: st.metric("Paneles", ht_df["panel"].nunique())
                    with c3: st.metric("CCTs por panel", ht_df[ht_df.cct.str.match(r"CCT.*\d", na=False)]["cct"].nunique())
                    with c4: st.metric("Período", f"{ht_df['fecha'].min()} → {ht_df['fecha'].max()}"[:21])

                    st.markdown("---")
                    st.caption(f"ℹ️ {t('tip_cct')}  ·  {t('tip_desbalance')}")

                    c1,c2 = st.columns(2)
                    with c1:
                        paneles = sorted(ht_df["panel"].unique().tolist())
                        panel_sel = st.selectbox("Panel", paneles, key="ht_panel")
                    with c2:
                        fechas_ht2 = ["Última"]+sorted(ht_df["fecha"].unique().tolist(), reverse=True)
                        fecha_ht_sel = st.selectbox("Fecha", fechas_ht2, key="ht_fecha")

                    df_panel = ht_df[ht_df.panel==panel_sel]
                    fecha_use = df_panel["fecha"].max() if fecha_ht_sel=="Última" else fecha_ht_sel
                    panel_day = df_panel[df_panel.fecha==fecha_use]
                    es_trifasico = panel_sel != "0000-EHT-005"

                    if len(panel_day)>0:
                        row_panel = panel_day.iloc[0]
                        import numpy as _np_pb
                        ccts_amp = panel_day.dropna(subset=["amperaje_a"])["amperaje_a"].values
                        if len(ccts_amp)>0:
                            amp_min = float(_np_pb.min(ccts_amp))
                            amp_max = float(_np_pb.max(ccts_amp))
                            desbalance_pct = float((amp_max-amp_min)/amp_min*100) if amp_min>0 else 0
                            varianza_ccts  = float(_np_pb.var(ccts_amp))
                        else:
                            desbalance_pct = varianza_ccts = 0

                        col_desbal = "#4ade80" if desbalance_pct<10 else "#f59e0b" if desbalance_pct<20 else "#ef4444"
                        mc_col = "#4ade80" if str(row_panel.get("mc_estado","")).lower() in ("on","o") else "#ef4444"
                        tipo_panel = "Trifásico" if es_trifasico else "Monofásico"

                        cc = st.columns(2)
                        with cc[0]:
                            st.markdown(f"""<div style="background:#111113;border:1px solid #222;border-radius:6px;padding:10px">
                                <div style="font-size:10px;color:#555">{tipo_panel}</div>
                                <div style="font-family:monospace;font-size:12px;font-weight:600;color:#fff">{panel_sel}</div>
                                <div style="font-size:10px;color:#444">{row_panel.get('ubicacion','')}</div>
                            </div>""", unsafe_allow_html=True)
                        with cc[1]: st.metric("Voltaje", f"{row_panel.get('voltaje_v','-')} V")

                    # CCTs sorted 1-25
                    import re as _re_ht
                    def cct_key(c):
                        m = _re_ht.search(r"\d+", str(c))
                        return int(m.group()) if m else 999
                    ccts_raw = panel_day["cct"].dropna().unique().tolist()
                    ccts = sorted(ccts_raw, key=cct_key)
                    cct_sel = st.selectbox("CCT", ccts, key="ht_cct")

                    df_cct = df_panel[df_panel.cct==cct_sel].dropna(subset=["amperaje_a"]).copy()
                    df_cct["fecha"] = pd.to_datetime(df_cct["fecha"])

                    if len(df_cct)>0:
                        import plotly.express as px
                        fig_ht = px.line(df_cct.sort_values("fecha"), x="fecha", y="amperaje_a",
                            markers=True, height=240, color_discrete_sequence=["#6366f1"],
                            labels={"fecha":"Fecha","amperaje_a":"Amperaje (A)"})
                        fig_ht.update_layout(title=dict(text=f"{panel_sel} — {cct_sel}",
                            font=dict(color="#aaa",size=12),x=0.5),
                            margin=dict(l=0,r=0,t=30,b=0),
                            plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                            font=dict(family="Inter",size=11,color="#888"),
                            xaxis=dict(color="#444",gridcolor="#1e1e22"),
                            yaxis=dict(color="#444",gridcolor="#1e1e22",title="A"))
                        st.plotly_chart(fig_ht, use_container_width=True)

                        vals = df_cct["amperaje_a"].dropna()
                        if len(vals)>0:
                            media_v = float(vals.mean())
                            mediana_v = float(vals.median())
                            varianza_v = float(vals.var())
                            c1,c2,c3,c4 = st.columns(4)
                            with c1: st.metric("Media",f"{media_v:.2f} A"); st.caption(f"ℹ️ {t('tip_media')}")
                            with c2: st.metric("Mediana",f"{mediana_v:.2f} A"); st.caption(f"ℹ️ {t('tip_mediana')}")
                            with c3: st.metric("Varianza",f"{varianza_v:.3f}"); st.caption(f"ℹ️ {t('tip_varianza')}")
                            with c4: st.metric("Rango",f"{vals.min():.1f}–{vals.max():.1f} A")
                            import numpy as _np_v
                            outliers = df_cct[abs(df_cct["amperaje_a"]-media_v)>2*_np_v.sqrt(varianza_v)]
                            if len(outliers)>0:
                                st.warning(f"⚠️ {len(outliers)} anomalías por varianza (>2σ)")

                    st.markdown("---")
                    st.markdown(f"#### Todos los CCTs — {panel_sel} — {fecha_use}")
                    day_ccts = panel_day[["cct","amperaje_a"]].copy().dropna(subset=["cct"])
                    day_ccts["_s"] = day_ccts["cct"].apply(lambda c: cct_key(str(c)))
                    day_ccts = day_ccts.sort_values("_s").drop("_s",axis=1)
                    day_ccts["amperaje_a"] = day_ccts["amperaje_a"].apply(lambda x: f"{x:.2f} A" if pd.notna(x) else "—")
                    st.dataframe(day_ccts.rename(columns={"cct":"CCT","amperaje_a":"Amperaje"}), use_container_width=True, hide_index=True)

            except Exception as e:
                st.error(f"Error Heat Trace: {e}")

        # ─── PLC DIARIO ───────────────────────────────────────────────────────
        with sub_plc:
            try:
                plc_df = qdf("SELECT fecha,plc_id,modulo,lampara,estado,es_anomalia FROM mediciones_plc ORDER BY fecha DESC,plc_id,lampara")
                if len(plc_df)==0:
                    st.info("Sin datos PLC")
                else:
                    c1,c2,c3,c4 = st.columns(4)
                    with c1: st.metric("Días registrados", plc_df["fecha"].nunique())
                    with c2: st.metric("PLCs", plc_df["plc_id"].nunique())
                    with c3: st.metric("🔴 Anomalías totales", len(plc_df[plc_df.es_anomalia==1]))
                    with c4: st.metric("Días con anomalía", plc_df[plc_df.es_anomalia==1]["fecha"].nunique())

                    with st.expander("ℹ️ Anomalías: MAINT, BAF, BATT1F, BATT2F"):
                        st.caption(f"**BAF:** {t('tip_baf')}")
                        st.caption(f"**BATT1F / BATT2F:** {t('tip_batt')}")

                    st.markdown("---")
                    st.markdown("#### Estado de PLCs")
                    ultima_fecha = plc_df["fecha"].max()
                    plcs_list_all = sorted(plc_df["plc_id"].unique().tolist())
                    ec = st.columns(min(len(plcs_list_all),4))
                    for i,plc_id in enumerate(plcs_list_all):
                        with ec[i%4]:
                            anom_plc = plc_df[(plc_df.plc_id==plc_id)&(plc_df.es_anomalia==1)]
                            anom_hoy = anom_plc[anom_plc.fecha==ultima_fecha]
                            color = "#ef4444" if len(anom_hoy)>0 else "#22c55e"
                            estado_txt = "⚠️ ANOMALÍA" if len(anom_hoy)>0 else "✅ OK"
                            lamps_txt = ", ".join(anom_hoy["lampara"].unique().tolist()) if len(anom_hoy)>0 else ""
                            detalle = f"<br><span style='font-size:10px;color:#fca5a5'>{lamps_txt}</span>" if lamps_txt else ""
                            dias_anom_plc = int(anom_plc["fecha"].nunique())
                            dias_txt = f"{dias_anom_plc} " + ("día" if dias_anom_plc==1 else "días") + " con anomalía"
                            html_card = (
                                f'<div style="background:#111113;border:1.5px solid {color};border-radius:8px;'
                                f'padding:10px 14px;margin-bottom:8px;text-align:center">'
                                f'<div style="font-family:monospace;font-size:10px;color:#888">{plc_id}</div>'
                                f'<div style="font-size:13px;font-weight:600;color:{color};margin-top:3px">{estado_txt}</div>'
                                f'{detalle}'
                                f'<div style="font-size:10px;color:#444;margin-top:3px">{dias_txt}</div>'
                                f'</div>'
                            )
                            st.markdown(html_card, unsafe_allow_html=True)

                    st.markdown("---")
                    c1,c2,c3 = st.columns(3)
                    with c1:
                        plc_sel = st.selectbox("PLC",["Todos"]+plcs_list_all,key="plc_sel2",label_visibility="collapsed")
                    with c2:
                        lamp_sel = st.selectbox("Lámpara",["Todas"]+sorted(plc_df["lampara"].unique().tolist()),key="lamp_sel",label_visibility="collapsed")
                    with c3:
                        solo_anom = st.checkbox("Solo anomalías",value=True,key="solo_anom")

                    res_plc = plc_df.copy()
                    if plc_sel!="Todos": res_plc=res_plc[res_plc.plc_id==plc_sel]
                    if lamp_sel!="Todas": res_plc=res_plc[res_plc.lampara==lamp_sel]
                    if solo_anom: res_plc=res_plc[res_plc.es_anomalia==1]

                    if len(res_plc)==0:
                        st.success("✅ Sin anomalías con los filtros seleccionados")
                    else:
                        import plotly.express as px
                        anom_by_day=res_plc.groupby(["fecha","plc_id"]).size().reset_index(name="n")
                        fig_plc=px.bar(anom_by_day,x="fecha",y="n",color="plc_id",height=220,barmode="stack",
                            color_discrete_sequence=["#6366f1","#0ea5e9","#14b8a6","#8b5cf6","#ef4444","#f59e0b","#22c55e","#64748b"],
                            labels={"fecha":"","n":"Anomalías","plc_id":"PLC"})
                        fig_plc.update_layout(margin=dict(l=0,r=0,t=0,b=0),
                            plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                            font=dict(family="Inter",size=10,color="#888"),
                            xaxis=dict(color="#444",tickangle=45,tickfont=dict(size=8)),
                            yaxis=dict(color="#444",gridcolor="#1e1e22"),
                            legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)"))
                        st.plotly_chart(fig_plc,use_container_width=True)
                        st.dataframe(res_plc.rename(columns={"fecha":"Fecha","plc_id":"PLC","modulo":"Módulo","lampara":"Lámpara","estado":"Estado","es_anomalia":"Anomalía"}),
                            use_container_width=True,hide_index=True,height=280)
            except Exception as e:
                st.error(f"Error PLC: {e}")

        # ─── INSPECCIÓN SEMANAL ───────────────────────────────────────────────
        with sub_sem:
            try:
                sem_df = qdf("SELECT fecha,tag,descripcion,estado_alim,valor,valor_num,observacion,tipo,grupo_sensor,tipo_sensor,fecha_hoja FROM mediciones_semanal ORDER BY fecha_hoja DESC,tag")
                if len(sem_df)==0:
                    st.info("Sin inspecciones semanales")
                else:
                    fechas_hojas = sorted(sem_df["fecha_hoja"].dropna().unique().tolist(), reverse=True)
                    total_instr = sem_df["tag"].nunique()
                    # Expected per inspection = total unique instruments
                    # Realized = rows with valor not null
                    con_dato = sem_df["valor"].notna().sum()
                    total_registros = len(sem_df)
                    pct_total = round(con_dato / max(total_registros,1) * 100)

                    c1,c2,c3,c4 = st.columns(4)
                    with c1: st.metric("Semanas de inspección", len(fechas_hojas))
                    with c2: st.metric("Instrumentos registrados", total_instr)
                    with c3:
                        st.metric("% Mediciones realizadas", f"{pct_total}%")
                        st.caption("Mediciones con dato / total de filas")
                    with c4:
                        off_n = len(sem_df[sem_df.estado_alim=="off"])
                        st.metric("🔴 Sensores OFF", off_n)

                    st.markdown("---")

                    # ── % COMPLETITUD POR SEMANA ──────────────────────────────
                    st.markdown("#### % Mediciones realizadas por semana")
                    completitud = []
                    for fh in sorted(fechas_hojas):
                        df_fh = sem_df[sem_df.fecha_hoja==fh]
                        con = df_fh["valor"].notna().sum()
                        tot = len(df_fh)
                        pct = round(con/max(tot,1)*100)
                        completitud.append({"Semana":fh, "Realizadas":con, "Total":tot, "Pct":pct})

                    import plotly.express as px
                    comp_df = pd.DataFrame(completitud)
                    fig_comp = px.bar(comp_df, x="Semana", y="Pct", height=200,
                        color="Pct",
                        color_continuous_scale=["#ef4444","#f59e0b","#22c55e"],
                        range_color=[0,100],
                        text="Pct",
                        labels={"Semana":"","Pct":"%"},
                        hover_data={"Realizadas":True,"Total":True})
                    fig_comp.update_traces(texttemplate="%{text}%", textposition="outside",
                                          textfont=dict(color="#aaa",size=9))
                    fig_comp.update_layout(margin=dict(l=0,r=0,t=0,b=0),
                        coloraxis_showscale=False,
                        plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Inter",size=10,color="#888"),
                        xaxis=dict(color="#444",tickangle=45),
                        yaxis=dict(color="#444",gridcolor="#1e1e22",range=[0,115]))
                    st.plotly_chart(fig_comp, use_container_width=True)

                    # ── ESTADO ON/OFF/AN última semana ────────────────────────
                    ultima_insp = fechas_hojas[0]
                    st.markdown(f"#### Estado de sensores — {ultima_insp}")
                    df_ult = sem_df[sem_df.fecha_hoja==ultima_insp]
                    on_n   = len(df_ult[df_ult.estado_alim=="on"])
                    an_n   = len(df_ult[df_ult.estado_alim=="an"])
                    off_n2 = len(df_ult[df_ult.estado_alim=="off"])
                    sin_n  = len(df_ult[~df_ult.estado_alim.isin(["on","an","off"])])

                    c1,c2 = st.columns(2)
                    with c1:
                        pie_d = pd.DataFrame({
                            "Estado":["ON — digital encendido","AN — analógico","OFF — digital apagado","Sin dato"],
                            "n":[on_n,an_n,off_n2,sin_n]})
                        pie_d = pie_d[pie_d.n>0]
                        fig_pie = px.pie(pie_d, values="n", names="Estado", height=220,
                            color_discrete_map={
                                "ON — digital encendido":"#22c55e",
                                "AN — analógico":"#3b82f6",
                                "OFF — digital apagado":"#ef4444",
                                "Sin dato":"#374151"})
                        fig_pie.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                            plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                            font=dict(family="Inter",size=10,color="#888"),
                            legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)"))
                        fig_pie.update_traces(textposition="inside",
                            textinfo="percent",textfont=dict(size=9,color="white"))
                        st.plotly_chart(fig_pie, use_container_width=True)
                        st.caption("ON=digital encendido  ·  AN=analógico  ·  OFF=digital apagado")
                    with c2:
                        df_off = df_ult[df_ult.estado_alim=="off"][["tag","descripcion","observacion"]].copy()
                        if len(df_off)>0:
                            st.markdown("**Sensores OFF:**")
                            st.dataframe(df_off.rename(columns={
                                "tag":"Tag","descripcion":"Descripción","observacion":"Motivo"}),
                                use_container_width=True,hide_index=True)
                        else:
                            st.success("✅ Sin sensores apagados")

                    st.markdown("---")

                    # ── FILTROS ───────────────────────────────────────────────
                    c1,c2,c3,c4 = st.columns(4)
                    with c1:
                        fs_sel = st.selectbox("",
                            ["Todas las semanas"]+fechas_hojas,
                            label_visibility="collapsed", key="sem_f")
                    with c2:
                        tipos_list = ["Todos"]+sorted(sem_df["tipo_sensor"].dropna().unique().tolist())
                        ts_sel = st.selectbox("",
                            tipos_list,
                            label_visibility="collapsed", key="sem_ts")
                    with c3:
                        alim_opts = ["Todos","on","an","off"]
                        alim_sel = st.selectbox("",
                            alim_opts,
                            label_visibility="collapsed", key="sem_alim",
                            format_func=lambda x: {"Todos":"Todos","on":"ON","an":"AN","off":"OFF"}.get(x,x))
                    with c4:
                        busq_s = st.text_input("",
                            placeholder="Buscar por tag (ej: 3100)",
                            label_visibility="collapsed", key="sem_b")

                    # Apply filters
                    res_sem = sem_df.copy()
                    if fs_sel != "Todas las semanas":
                        res_sem = res_sem[res_sem.fecha_hoja==fs_sel]
                    if ts_sel != "Todos":
                        res_sem = res_sem[res_sem.tipo_sensor==ts_sel]
                    if alim_sel != "Todos":
                        res_sem = res_sem[res_sem.estado_alim==alim_sel]
                    if busq_s and busq_s.strip():
                        b = busq_s.strip().lower()
                        res_sem = res_sem[
                            res_sem.tag.str.lower().str.contains(b, na=False) |
                            res_sem.descripcion.fillna("").str.lower().str.contains(b)]

                    # Completitud del filtro actual
                    con_f = res_sem["valor"].notna().sum()
                    tot_f = len(res_sem)
                    pct_f = round(con_f/max(tot_f,1)*100)
                    st.caption(f"📊 {con_f}/{tot_f} mediciones con dato ({pct_f}%)")
                    st.caption(f"ℹ️ {t('tip_estado_alim')}")

                    # ── GRÁFICO TENDENCIA (vinculado a búsqueda) ──────────────
                    tags_filtrados = sorted(res_sem["tag"].unique().tolist())
                    tag_opts = ["— Seleccionar instrumento —"] + tags_filtrados
                    tag_sem_sel = st.selectbox("",
                        tag_opts, label_visibility="collapsed", key="sem_tag_sel")

                    if tag_sem_sel != "— Seleccionar instrumento —":
                        # Use full history for selected tag (not filtered)
                        df_tag = sem_df[sem_df.tag==tag_sem_sel].dropna(subset=["valor_num"]).copy()
                        df_tag["fecha_plot"] = pd.to_datetime(df_tag["fecha_hoja"])
                        df_tag = df_tag.sort_values("fecha_plot")

                        if len(df_tag) > 1:
                            import plotly.graph_objects as go
                            import numpy as _np2
                            vals2 = df_tag["valor_num"]
                            media2 = vals2.mean()
                            desvio2 = vals2.std()

                            def cls2(v):
                                d = abs(v-media2)
                                return "🔴 >2σ" if d>2*desvio2 else "🟠 >1σ" if d>desvio2 else "🟢 Normal"
                            df_tag["cls"] = df_tag["valor_num"].apply(cls2)

                            desc_tag = df_tag["descripcion"].iloc[0]
                            tipo_tag = df_tag["tipo_sensor"].iloc[0]

                            fig_s = go.Figure()
                            fig_s.add_trace(go.Scatter(
                                x=df_tag["fecha_plot"], y=df_tag["valor_num"],
                                mode="lines", line=dict(color="#0ea5e9",width=1.5),
                                showlegend=False))
                            for cls,clr in [("🟢 Normal","#22c55e"),("🟠 >1σ","#f59e0b"),("🔴 >2σ","#ef4444")]:
                                m = df_tag["cls"]==cls
                                if m.any():
                                    fig_s.add_trace(go.Scatter(
                                        x=df_tag.loc[m,"fecha_plot"],
                                        y=df_tag.loc[m,"valor_num"],
                                        mode="markers",
                                        marker=dict(color=clr,size=8,
                                                    line=dict(color="white",width=1)),
                                        name=cls,
                                        hovertemplate="%{x|%d/%m}<br><b>%{y}</b><extra>"+cls+"</extra>"))
                            fr = [df_tag["fecha_plot"].min(), df_tag["fecha_plot"].max()]
                            fig_s.add_trace(go.Scatter(x=fr,y=[media2,media2],mode="lines",
                                line=dict(color="#60a5fa",width=1,dash="dash"),
                                name=f"Media: {media2:.4f}"))
                            fig_s.add_trace(go.Scatter(x=fr,y=[media2+desvio2,media2+desvio2],
                                mode="lines",line=dict(color="#f59e0b",width=1,dash="dot"),
                                name=f"+1σ: {(media2+desvio2):.4f}"))
                            fig_s.add_trace(go.Scatter(x=fr,y=[media2-desvio2,media2-desvio2],
                                mode="lines",line=dict(color="#f59e0b",width=1,dash="dot"),
                                showlegend=False))
                            fig_s.add_trace(go.Scatter(x=fr,y=[media2+2*desvio2,media2+2*desvio2],
                                mode="lines",line=dict(color="#ef4444",width=1,dash="dot"),
                                name=f"+2σ: {(media2+2*desvio2):.4f}"))
                            fig_s.add_trace(go.Scatter(x=fr,y=[media2-2*desvio2,media2-2*desvio2],
                                mode="lines",line=dict(color="#ef4444",width=1,dash="dot"),
                                showlegend=False))

                            fig_s.update_layout(
                                title=dict(text=f"{tag_sem_sel} ({tipo_tag}) — {str(desc_tag)[:50]}",
                                    font=dict(color="#aaa",size=11),x=0.5),
                                height=260,margin=dict(l=0,r=0,t=30,b=0),
                                plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                                font=dict(family="Inter",size=10,color="#888"),
                                xaxis=dict(color="#444",gridcolor="#1e1e22"),
                                yaxis=dict(color="#444",gridcolor="#1e1e22"),
                                legend=dict(font=dict(color="#888",size=9),
                                           bgcolor="rgba(0,0,0,0)",
                                           orientation="h",y=-0.3,x=0))
                            st.plotly_chart(fig_s, use_container_width=True)

                            # Stats
                            c1,c2,c3,c4 = st.columns(4)
                            with c1: st.metric("Media",f"{media2:.4f}"); st.caption(f"ℹ️ {t('tip_media')}")
                            with c2: st.metric("Mediana",f"{vals2.median():.4f}"); st.caption(f"ℹ️ {t('tip_mediana')}")
                            with c3: st.metric("Varianza",f"{vals2.var():.6f}"); st.caption(f"ℹ️ {t('tip_varianza')}")
                            with c4: st.metric("Desvío σ",f"{desvio2:.4f}"); st.caption(f"ℹ️ {t('tip_desvio')}")
                            st.caption(f"ℹ️ {t('tip_sigma')}")

                            out2 = df_tag[df_tag["cls"]!="🟢 Normal"]
                            if len(out2)>0:
                                st.warning(f"⚠️ {len(out2)} medición(es) fuera de 1σ — {len(df_tag[df_tag['cls']=='🔴 >2σ'])} fuera de 2σ")
                        elif len(df_tag)==1:
                            st.info("Solo 1 medición registrada — se necesitan al menos 2 para el gráfico")
                        else:
                            st.info("Sin valores numéricos para este instrumento")

                    st.markdown("---")

                    # ── TABLA (vinculada a filtros) ───────────────────────────
                    st.caption(f"{len(res_sem)} registros")
                    st.dataframe(
                        res_sem[["tag","descripcion","fecha_hoja","tipo_sensor",
                                  "estado_alim","valor","observacion"]].rename(columns={
                            "tag":"Tag",
                            "descripcion":"Descripción",
                            "fecha_hoja":"Semana",
                            "tipo_sensor":"Tipo",
                            "estado_alim":"Alim.",
                            "valor":"Valor",
                            "observacion":"Observación"}),
                        use_container_width=True,
                        hide_index=True,
                        height=350)

            except Exception as e:
                st.error(f"Error semanal: {e}")
                import traceback; st.code(traceback.format_exc())


elif pagina == "comisionado":
    LANG_C = st.session_state.get("lang","es")
    tit_c = {"es":"Comisionado LC Plant","en":"LC Plant Commissioning","ko":"LC 플랜트 시운전"}
    st.markdown(f"## {tit_c.get(LANG_C,'Comisionado')}")

    com_df = qdf("SELECT * FROM comisionado ORDER BY actividad, tag")

    if len(com_df) == 0:
        st.info("Sin datos de comisionado cargados")
    else:
        actividades = com_df["actividad"].unique().tolist()
        cols_met = st.columns(len(actividades))
        for i, act in enumerate(actividades):
            df_act = com_df[com_df.actividad==act]
            ok_n  = len(df_act[df_act.estado=="OK"])
            tot_a = len(df_act)
            pct   = round(ok_n/tot_a*100) if tot_a>0 else 0
            color = "#22c55e" if pct>=80 else "#f59e0b" if pct>=50 else "#ef4444"
            with cols_met[i]:
                st.markdown(f"""
                <div style="background:#111113;border:1px solid #222;border-radius:8px;
                            padding:14px;text-align:center">
                    <div style="font-size:10px;color:#555;text-transform:uppercase;
                                letter-spacing:0.05em;margin-bottom:4px">{act}</div>
                    <div style="font-size:22px;font-weight:600;color:{color}">{pct}%</div>
                    <div style="font-size:11px;color:#444">{ok_n}/{tot_a}</div>
                    <div style="height:4px;background:#1e1e22;border-radius:2px;margin-top:6px">
                        <div style="height:4px;width:{pct}%;background:{color};border-radius:2px"></div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("---")

        import plotly.express as px
        import pandas as pd

        # Filtro actividad
        act_sel = st.selectbox("Actividad", ["Todas"] + actividades,
                               label_visibility="collapsed", key="com_act_sel")
        com_show = com_df if act_sel=="Todas" else com_df[com_df.actividad==act_sel]

        # Gantt
        gantt = []
        for _, r in com_show[com_show.fecha_plan.notna()].iterrows():
            fin = r.get("fecha_real") or r.get("fecha_plan")
            if fin:
                gantt.append({"Tag":r["tag"],"Actividad":r["actividad"],
                              "Estado":r["estado"],"Inicio":r["fecha_plan"],"Fin":fin})
        if gantt:
            gdf = pd.DataFrame(gantt)
            gdf["Inicio"] = pd.to_datetime(gdf["Inicio"],errors="coerce")
            gdf["Fin"]    = pd.to_datetime(gdf["Fin"],errors="coerce")+pd.Timedelta(days=1)
            gdf = gdf.dropna(subset=["Inicio","Fin"])
            if len(gdf)>0:
                h = max(280, min(len(gdf)*14, 600))
                fig_g = px.timeline(gdf, x_start="Inicio", x_end="Fin",
                    y="Tag", color="Estado", height=h,
                    color_discrete_map={"OK":"#22c55e","NO OK":"#ef4444",
                                        "CHECK":"#f59e0b","Pending":"#334155"},
                    hover_data=["Actividad"])
                fig_g.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                    plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter",size=10,color="#888"),
                    xaxis=dict(color="#444",gridcolor="#1e1e22"),
                    yaxis=dict(color="#888",tickfont=dict(size=9)),
                    legend=dict(font=dict(color="#888"),bgcolor="rgba(0,0,0,0)",
                               orientation="h",y=1.05))
                st.plotly_chart(fig_g, use_container_width=True)

        st.markdown("---")

        # Tabla
        show_c = [c for c in ["tag","equipo_desc","actividad","estado",
                               "fecha_plan","fecha_real","observacion"]
                  if c in com_show.columns]
        st.dataframe(
            com_show[show_c].rename(columns={
                "tag":"Tag","equipo_desc":"Equipo","actividad":"Actividad",
                "estado":"Estado","fecha_plan":"Fecha plan",
                "fecha_real":"Fecha real","observacion":"Mediciones"}),
            use_container_width=True, hide_index=True, height=400)


# ════════════════════════════════════
# REPUESTOS Q - BUSCADOR GLOBAL
# ════════════════════════════════════
elif pagina == "repuestos":
    LANG_R = st.session_state.get("lang","es")
    tit_r = {"es":"Repuestos Q","en":"Q Spare Parts","ko":"Q 예비품"}
    st.markdown(f"## {tit_r.get(LANG_R,'Repuestos Q')}")

    # Metricas rapidas
    total_q   = qdf("SELECT COUNT(*) n FROM repuestos_q").iloc[0]["n"]
    try:
        total_act = qdf("SELECT COUNT(*) n FROM repuestos_q WHERE status='ACTIVE'").iloc[0]["n"]
    except:
        total_act = total_q
    try:
        total_vin = qdf("SELECT COUNT(*) n FROM repuesto_tipo_equipo").iloc[0]["n"]
    except:
        total_vin = 0
    cats_raw  = qdf("SELECT DISTINCT leaf_class FROM repuestos_q WHERE leaf_class IS NOT NULL ORDER BY leaf_class")
    cats_list = sorted(cats_raw["leaf_class"].dropna().tolist())
    total_cat = len(cats_list)

    c1,c2,c3,c4 = st.columns(4)
    with c1: st.metric(t("total_q_lbl"), f"{total_q:,}")
    with c2: st.metric(t("activos_lbl"), f"{total_act:,}")
    with c3: st.metric(t("categorias_lbl"), total_cat)
    with c4: st.metric(t("vinculos_lbl"), f"{total_vin:,}")

    st.markdown("---")

    # Filtros
    c1,c2,c3 = st.columns([3,2,2])
    with c1:
        busq_q = st.text_input("",
            placeholder="Buscar codigo Q, descripcion, tipo...",
            label_visibility="collapsed", key="busq_q_global")
    with c2:
        cats_all = ["Todas"] + cats_list
        cat_q = st.selectbox("", cats_all,
            label_visibility="collapsed", key="cat_q_global")
    with c3:
        status_q = st.selectbox("", ["Todos","ACTIVE","INACTIVE"],
            label_visibility="collapsed", key="status_q_global")

    # Construir query con parametros seguros
    where = "WHERE 1=1"
    params = []
    if busq_q and busq_q.strip():
        where += " AND (codigo_q LIKE ? OR descripcion LIKE ? OR leaf_class LIKE ?)"
        v = f"%{busq_q.strip()}%"
        params += [v, v, v]
    if cat_q and cat_q != "Todas":
        where += " AND leaf_class=?"
        params.append(cat_q)
    if status_q and status_q != "Todos":
        try:
            # Only add status filter if column exists
            test = qdf("SELECT status FROM repuestos_q LIMIT 1")
            where += " AND status=?"
            params.append(status_q)
        except:
            pass

    sql = f"SELECT codigo_q,leaf_class,descripcion,unidad,lead_time,status,categoria_compra FROM repuestos_q {where} ORDER BY leaf_class,codigo_q LIMIT 500"
    rep_df = qdf(sql, tuple(params))

    st.caption(f"{len(rep_df)} resultados")

    if len(rep_df) > 0:
        st.dataframe(
            rep_df.rename(columns={
                "codigo_q":      "Codigo Q",
                "leaf_class":    "Tipo",
                "descripcion":   "Especificacion",
                "unidad":        "Ud",
                "lead_time":     "Lead Time",
                "status":        "Estado",
                "categoria_compra": "Categoria"
            }),
            use_container_width=True,
            hide_index=True,
            height=520)
    else:
        st.info("Sin resultados — proba con otro termino")


# ════════════════════════════════════
# MAPA DE PLANTA
# ════════════════════════════════════
elif pagina == "mapa":
    LANG_M = st.session_state.get("lang","es")
    tit = {"es":"Mapa de planta","en":"Plant map","ko":"플랜트 지도"}
    st.markdown(f"## {tit.get(LANG_M,'Mapa')}")

    AREAS_MAPA = [
        (37, 10, "5500", "Soda Ash",           "Soda Ash Plant",     "소다회"),
        (47, 22, "4600", "Secado/Packaging",   "Drying & Packaging", "건조/포장"),
        (43, 25, "4300", "LC Plant",           "LC Plant",           "LC 플랜트"),
        (40, 28, "4400", "Filtrado/Lavado",    "Filter & Wash",      "여과/세척"),
        (38, 32, "3200", "Clarificador",       "Clarifier",          "정화기"),
        (37, 28, "3100", "Filtro Prensa",      "Filter Press",       "필터프레스"),
        (36, 22, "3300", "Cal Viva",           "Lime",               "석회"),
        (43, 36, "4500", "Licor Madre",        "Mother Liquor",      "모액"),
        (46, 36, "5400", "TK HCl",            "HCl Tank",           "HCl 탱크"),
        (49, 36, "5300", "Soda Caustica",      "Caustic Soda",       "가성소다"),
        (52, 36, "5100", "TK H2SO4",          "H2SO4 Tank",         "황산탱크"),
        (55, 32, "6300", "Calderas",           "Steam Boilers",      "보일러"),
        (59, 32, "6500", "Compresores",        "Compressors",        "압축기"),
        (65, 28, "6100", "Water Plant",        "Water Plant",        "용수"),
        (35, 42, "4700", "Descarte LC",        "LC Discards",        "폐수"),
        (20, 28, "2300", "Post Liming",        "Post Liming Pond",   "후처리"),
        (80, 45, "0000", "Planta Gas",         "Gas Plant",          "가스"),
        (90, 45, "0000", "Power Plant",        "Power Plant",        "발전소"),
    ]

    idx_n = {"es":3,"en":4,"ko":5}.get(LANG_M,3)
    area_sel = st.session_state.get("mapa_area_sel","")

    eq_por_area = {}
    for row in qdf("SELECT area_codigo, COUNT(*) n FROM equipos WHERE area_codigo IS NOT NULL GROUP BY area_codigo").itertuples(index=False):
        eq_por_area[row[0]] = row[1]

    # Imagen embebida en base64 - funciona en cualquier PC
    PLANTA_IMG = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAK5BLEDASIAAhEBAxEB/8QAHQAAAQUBAQEBAAAAAAAAAAAABAIDBQYHAQAICf/EAF4QAAIBAgQEAwYCBQcHCQQEDwECAwQRAAUSIQYTMUEiUWEHFDJxgZEjoRVCscHRCBYkM1Lh8DRTYnKS0vEXNUNUc4KTorIlJmPCNkRVZHSDhLPilJXTGCdFpDeFo//EABwBAAIDAQEBAQAAAAAAAAAAAAIDAAEEBQYHCP/EAEwRAAEDAgQDAwcIBwUIAgMBAAEAAhEDIQQFEjETQVEGImEHFDJxkaHRFUJSgZKxwfAWFyMzYnLhNENEU6IkVIKywtLi8SU1JmOTZP/aAAwDAQACEQMRAD8AZqIhHAlTy5JDqADOTcL5YCgR62oFHDYyabhgTpUeZHfr0wVl1fU1I9wpVTRpIkdk8KXN7D1vfEvSUNLlNO05iCsBc92bbpjr7pektuU/ktHHQUMdPouASZGNgWbpqsPpgqon5cRaQEx97eWIlMwkkUkJpRzsp7D54brah5I2DfCOh23/ALsUWqGoITNXUmqckEoC1gPPAMkccUmlnsSTckdP78cdjIoSNT0GwPXBdHl0ghWRWTUQQS9zYf474px5ImOsmInhkc6XJKgWAHw+mHKmOMzRvHCpJHW3UY7MuqWGJYr8ok7DY/lhyVXjCoWAO5NxuMKNkyUOukTDlzNpXfS67fa2HJZeTGXRy5J2ttf0/bgeSoAqDURyFk2Grr6X9MeDcwXp2jY9d+/nitVkQEXSBXIWJK6H2B8WwJOHHmDJqhIPW2564FqzvpRASo1E31W/jgZpZ0DG1y26gncfTCS66IGbqRYyCQtpZSF3Kk44XV5CpYkCwKgb/XDcLvJGZHjJ1LuLbf3YdJjWPwr4mte+5G2KlRdktGpZSStunS2OTSKsFyGEncKRe5x1ZIdWhn8RAJsoO3bDdToVQdWpmNx18+mCCPkuxylQ7sStlve4uDhqAs12Ckg7nC4ZUijVZmCMPFpff6232wFPnVLUuIKVjV7N40IUEi1wPM2PT54Mx1QSeikZTHLELEA9BvsPnhVI0asdTCwG7nY/niuvPM8jQTMNSvoKwnVpYfLt3v3vhmGF1pFlnk5pdiZFabQNVuwtfckdT5YQ6oAEwMJKnqvNaVJdEdSssjMFAPdjY2uepsR0wDV1FU9aHRmMkDBtDMQjgqR2G4tq77EDAsJGsqWldja2kKApsTcDqN/O5w6aZH5RMMPMF+pJ0ggXA22+FOx6HGN9Uk2WplEEXSDC7gHko6tZWSV7FiQS1+oJXbre9j6YWEhDssiJIXUqyrva4tfpv2tfyx6hoppJeZLLIF0qDG/jVrHfr8rX2v1tg+NY4w9kUlSBc9D9MJ1E7pjQGmQhaajJB1yB41GixtqCADTe97+G9/U3wTBCpWcVkUckrENI63A63W3yFgCLHa/XCnnYALDZt7kg2GIrMPxYJ4WZ194BDuhsTcHYfTa//DAkppNpUqzQNEyFFVANIB3G3a/frgcAJIEkaQizFSvTp/j54q+RVmZ0jxZZVrrVkX3edydUi3ta29jt1J2uNu2J7LabMJ5AYBUOo0vK4HgChVtv0sSxOw3t2tuWhx2QCq3mnKh3kZPG0elbEhQUG4J67E2NgR698Ns0dLG1VUXVVLu+ptJcndR5nvg5aDlsoZ2a1m1Fun0J/diPzb3KoqJUpmjqJYtQe/jCuVsF0+e+48sMFB5KU7EMiyr7Vud5u8VJklHKsYuhqai6IFa1rb7/AGPXF44U4ZyTJNeZ5lHDXZixZpZ5fhW48QC9LbnqMN5TTQZeYMty2llrcyq7Fgg1TTsNizH9VRf0CjFhgyGCGtiqM5kgzWaLdKNB/RYn/wBI/wDTEW72Xr12xsZSACyOeTuj8iro8wPvdHTLR5YrkiqWAAzmxGmIdTv+uRp/dPtP7yoYx6IR8Ea/O+/re+IySZmcTVMpfSAACbKg8gPL0x6szikoaJqytYQxbKiDeWZidlRerMewG+GNDRdAXEhTAlJDG6krub+Gw8yf3Yi6F844nqaiDh9Up6CJik2bVC3QMCQVgU/1hFtzcKPMnbFflizfOpETiKCbJsquGky6GUioqVZdllcfAN1uovfcX7YPzbiKSTKVhpnjyrKoEaOGCIBVUICCo6Cw6b7YviA90IAOZU9VZZwdkSmNUTNczK+OvrW50lx/ZJ2ToNlAGEUfEVDDQz1LV6zNCrIFjlNlYXuOp3GMupcwSrPPy+J1V7NFV1ADu41WPLWwFrBvEfSwYb4BzaSqesky+Kolp1FS00AVdrNE7M3+tqex2F9I74WQLmUbahEW5rV+Aa1K6hFdV3eVaiSJSCWIUlSL9+5O+K9mFTb2g5daFRFNNM0pt+qEO1+24Hzwv2O0rwcO1sUdUJpBXtrdwSWdVW/22H0wzJTPVcR0uZMwX3eOUMCtwQ/94GM7zYHqtzLko72izZdqgy1VRxKxcuwVfDpsQA21zr7+eKtSVEKTxUbLSvGu51Q+AhjfUVG42uAQLXB6bDC5q+qmzt6mrhp1huY0bdhp1JbbTsdSg7+WGRG8dXICZIUTWsOuM3b4Tc9CBcEX8/MEYcXWkpMXXM7pGiEfu5vGxKRLG2kLc36bA9B9++GKSlFPEwksGZRe3RRbcfK/7sH0YmkppairqFR0Asm4G5vYC3iIUEHe/Tzw1NVhWKoRICRuV/xbCnASHNCrVIIKZQrIovYdrY68i2sR8sCyzJCrPK6oqsSdrAb4ZfMKVwn9Kj8Xwi+GhwCTKkAVaygX+eHoiplCsoAJwFl06VqsacMQCfFpIHXtgyOFgoOn1Nv24LUDzViUQUUqWAWwwpJARpJUgdLHt+zAMtU1KvMtKUJsqCIlmJNtvPDqkjbQynUSfAbXwOoHYqOmERIit4lNrC/S177du2FO2l4gkbEHqQDYfMnphLao1JZOoFrfTBMekoCG22uATtfvsMWXRdUASE0yjxEIpF7kjphACMSGjUkEdhgoTRLJo5i7bEW/K3XC6unErI8bNHZwW0Eb/cHbFcQIw0oEKqi6oBbsceOkoAUXVfyw/LEokOliVO5sPh+mGtIC6viubelsNBlBsbpMgFlUjxNa9sER6d7FQxGwPTA5Lc27Wta/TD6L3sBte5GIVZSbIGcMq7dcdQi52AHnjoVNJaI3BvsvQH1OOKbAA3setlxQCqV0aR4dK6evzx66AEEbD4rj8sNO1iQAethdbYUt2YfPrg9grkFPw2O5UX9BjpAF3tv6jHAbDw2A9MLUXUWG2AJUKbAGq2lb28sODSu2kAdscIAYkjfytjtwSRfxdcWChKWNBAuoHqceYAi2kW7G2EMpIBubX64UoJJ8RAJxcqLwtcqAPlbC0EauPAv1GOLcAsbG/rjmkE6mToO5wSm65eO99G+Fqikg2XDKDqLgi3nhxLMCNrWxEsp1FCqTpOFwkCUEBdhYahfbyw11tfYYeBubWF7WOKCiX+GH1C2kjuMJbSH1IgAHyx0h9IGmwud/XCoREHUufgOojcg/b5+Y6YsGEQTZCXUqgvfyGCWIEa2UFrDsAcdVEK6gDtcjSwt9N798JvMoco2m9gG6Eef1xCbI2pd4E1NJIkejdgTuL2t0G3XvjsRpWZbHUrD52PrYD9+BVWUqoEgVLknl3G5vuSdyfUjp0thzVpQIqLde5t+218LuVZICeJCsCqqSDa2kWI+t8PKY7FhGQVNyAQLflvganVn0xtfUTcg7A4JZDpPLsi2sXc9fkMSVUrq8hdR0s1zv4tj+WFsWYhhEq26HY/c2wyGjhQjrtsx2/LDBqJJCF7ebG1vpi5ULuqOkqFMSxlQxFwbjw/L1+uGveC5TRrYg2D2sF9BhNFTCeqUySkLchmI1AfS9/rg+pjhSNZFkVpEJ1amtqXoD87f8MUXBWAXXSSTyw8jgaiASvgIJ87Dc/XHHjjuqxOdVz1PkTfqfP0w5SiGend5NjYGJgwYEH179twMeikflvAj7mQsQ5up36/PCXFHpsmYJFlUs/iQSaQ5B7Hrv0vbBVXqFIkgCDtp0gYBlgXU0rzK2kWshsuo3BPn0/j1tjtS7GTSzAKOmkXAwvcpvEACTOyyLpbRqsN1H7sMhVBOlSzdSfI4WtlFlIJB2Nuv9+EF9ClnIB7gXwzSkyE6iqikkIQfXDTFL3su297dMNLUx3srEXboRjzSgNa6gX3xTkQunXcaSoANxa4wxJCEPhN1t0OOc6KzEN1PnjslSlwL7+q3Fvn54oOI2VOAQjMFdoyguP1rdcNqkQLFgpB8x1wdEsTs5YqbDe+2/bAk0aNIui4UXDEG4GHNeClAXXI1uCBGtyR074fEgXwlRt3GGBeJDpKtY2AO2ADnEp0qmUVcoP64KAfYtf8sNBBCo2Uw2hlNxYG/WwxwJHb4CVvtbAivKWVtQVWA8NxfrvhrMEqSRHT1To225mUA773AiJtbe/wDxxFOSmqCaIkhoiFAIIvvftjzVkaHklNAAu9k7+uIeZVQ6lqJ3ZLeHX4G+wBwTDXU95IYgNcAALM1vERewJ+eIEs7ouURGUsg1A77Gx6YTUiPQrm2oW69sJhVuqrrtudJvYnCKhjKtyCHtYnBNU2KFdFZ5IwBpI2sbYAsNSyW8YIJNr3xJzIyqrXsQRuT2wJLHufW/Q4jk2mn6SUCFZAyizlQtjcAE9z12tvidppaaVYiqhWte3WxtviuUM1zJCxW0YBQXt999+mCoSVAfqC1reuINlbt1ZNCOvLbStv1r3wGIyZmUx2K/2h5YRTPMZb/rFSdgevywUHE6rqXxkWJI7+mBiFAUHMIyusNd73AA6HClj5jG2mymxIP7sEpTRLKNYY6Rax3ufXCl0hS5jYdrL0vipVghCzoLqIwbjYki1u2EmlQ05ZidR+A264M0lWudLlTcknCXcFXEjiyeIg3vcjtbElU5A8uMxLIVUBTcgjr5WwiseI2ZQRdgCQRuMNPMNJLK/iN9XrvsB5YSGVpQWDtEDsOhI+2KQuICclK0+oyRgyOAdTC4AI2t64ZupYsxUkdSe2F1mppCQCAwGm/lhhmJTTEq3HQk/FguSWDdOSPGL3YbjfAsjLUMylQQBa47bYI5AbUS4dibHbzwPyzzNCsWZjup6fQ4sFGE7AYaWHmNrdmOlVXcdOt8dLguljqFuqi4++GW1TVbIuoKg0KR0Y+e32wqRL+MFlVOvliSiUhDIGEjNGt+2kXw+iREFlB1jzGIeKUpNZNXit1N7bYLSY9SQL9x0xRvsqRanbS/hHmw64VI4BSyWGwJXCHqNIuVWS+OlhIWKpaw2sLg+e+IonIRGzErZCE3+pP7sdrQiPoBBAYkXPbpb8sM2AJQhgRYWBwiQM0gJc9Oh2xEU8k/BNHpKhSQOpthxpkJuUttbcDA0S77i4OwYnqcOqiszMr9VtYb2xRVDdBTRq4fQPhN/ngAgKN1sTvuMTJAVyhXtsbYFq4QyXKi9zv5YrdUVHxITHq0C/lfr6YVF4Xs6kAYTIWjQspuqHrj0mtiSAH6AlTf88WoE80in8NQoUdThMkcbRBkjDfTDaI6Hx7nzvgh2vHp02uOt8FuovGURxKIgZFU2KuBZSe4H3+wwlbLY7OoO63tvjyxxMbum52tfrj0gYMXABB7C1h9MTZCVyJWUPpHj8vXHhIVk8eqPYghRex9b748quE6gC+21yT17YS7yxqzLG7MLFCCPCfS+LCqQLLwaJmIETC43FtsIuS5Kqum43IufX6488XLIWAmRtO3itdr/n064bKur6lkjV+joG8/lbvfFjooWoqMqJCyxi7LubWBPn9LY4xQHYD7WuceQ6AC56d/PCnANm204shD60lQjJ8ABvdbdvnjkSKXFluAfuccQLpNgSR9gcJZmVCwF9/O18CrAlKJOpyqWvqCsbWuDhuptFSyO234d/y+WFUoVgHMLq1jqsbi98LqSWpZdBUNoNr74pQWK+peaf7Q++PYY5r+S49jHJXQgdF84RQ0+UU2oRguW2UgXue9xgFnmmnV6w3W99Pl5YkZ5En0LOn4lgbE4ja1TrCICt+o32x2GrkvMpUwWVSYQAvQAG+Aqp1Eaxq7HVubAG2PTyzQQf0cWA6sNye2G4V5ciysSzHey9vTFuMIQJXYo2ijMosxI8Vhe3oB54NllhVEYsBrTZb9MJVpjeZJAka/q+ne+EyrDYCSIxuBuUYEMMKJTmtCXQ61Iu1gpvqIv9djj2aBStlaTxdGTt8zgasqlp40URvZrJaM3HzN8BI9QNQlZeSN0e+/1wtxRwkRyVURKVAuD0cjc+mEySlpRCjBUUXYA/EfLBNVoaMgm6tYW8v44DgRIJAsyjrqHkBhZ6JgAhdhW6gqqjzAHU4WINdZpdQscQBLHvft649PIoLLEw3eym3YHpftfp9cSE19xGLILKoJuT/fiNaCqcYTbRyAkxsNJN2W/UY4rM4sy2B6m+3pjyPpGh0KG4JF7W+mI3M86jVJlo15tTGraAT4ddtgem/pgXFrVYBOyLqzFSxGWZ7Jb4mO1vIYjTmaGQRUUJZOXq5ziygXA287b36YiKqpjZZKvNZ+dpnVTDuFjUKNRt12LC/z6WwVPWBYEW3KMEsJmR4gSV3Ggb2G/ltt64z1atu6VppUid09LEFqBHUxl2cBC99yd26HsPFY367WwASGr3kWLSyG8bSOLo1l1C2wZiARc9NwMJMxj5Rkji1qPxSQQEOxkJC/rXIAA+nfDsiPOI1Mbx6YxZSAVQ73tud9yAbdsZuI7qn6G9El5FinY0aq87uGeQ9ASACbdb2Fj6ffCaeCV21yNIClmV3F+Ze52FjYXt8V+mDIKcLOziJFGlSoA8V7738+vS3rgpUC2cxJ4uoZtXp06DbAuqOjdGKbQZhM0MUcpDiIiQuSS3Q72/aDsLDB5EUDB9i9reEeeGDKoAsoT6bffCVmijQEo9rdVFx1wuSSriEQs7BHOqyWJ3/jgWVyUYgl2I8Ntv24bkmZoAItQUsdxsx+RtbA0dlDvMzFiwuWNzfawJHzwekKWKLZmWUMpGnYMD38unlhwksI5DHfmEG4IuR877WGBqankd11IA4cFirEW/P5/bEhTpGiMRHKUG5tcg38u9sDF7KwSFF5pla5rEKOSR4GU84SSuSUIva1j0vf/AGDeDuKJ6Clq6OshaKeMlJDIbdP11sd/n6+uBMuzmNJ6yoeikmgEggSVYCRawNgQbk3P7MCTZHmOe1NPmKUUtJAhKjmktLMovYmMbdth5Em/bG6jLW2XNrOl6lpnzHi6qpXmphFlyJrZnuxk/sqOlhbv64lcmopJzNQ8PUtPBFTsBU1ky6YYL2JF+rvY3CLcnuQN8Tp4aSiYUWZ1swhazTU1NJaWUi/hdhtEDsfCSeo264OlVHjSFYKelpYb8qngXRHGOp2+ffqcP1aboNPRNZVTU+W0b0eUpyY5R/Sap956pupLHsCdwg2GHpZY6eG3L3B+DufmcBTZkI3YRMQF6v0t8v44jYlzDPaY1VFKKPLmk0nMJE1awOphTrIe2o2S/c9MWfFUXBF1NRVVeZJQ0NN75mMovFArWEaf5yRv1IwSLt62AJ2w5CKbIJv0nV19LXZsieKvcWhpVIF0gDfDv1kPiPmOmEe8UuXxz0HDtDLLNUeKWZpLzTmxOuokPRb/qgWHYYjauZ8rrIZM4WmzOtk0mny+jiZolIIbnOz+TJYbEdSAb7C4qgJuiHrYGy1syzCaSloHIcGXUks6ny2uAdrHqe3Y4rtQz586y1sccGWQBfc6C1gEXZRJ52sPB0He/ZNW1ZXZgcwzio59XcGNFJ5cPX4QerW21H8sc5zoSH69QFHX54qVYAKMgFpNdiWUAgW2xHZ7ORn8fjhSKalV5nGlSpjfwWJ3/6RhYbm4HTBskweIoGIOnY9/liPqVpUr8tzBo3nnhkKrHq0hiUYJ4r+EF9IJ8ifTBN9JXV9BX/2eSw02W5sEexNS8vU/wBlb/niImnmGa2JtEKc/VtQ2+18NZVWTUcM0DMqB1tywpBsSeoPe2m/rfDGZPLrM9OyGZRqiDmyg2I3PbqRjHXIBA6LZRkjUOaU0jK4YuqE2PTqL2w60hiiKyR6gd9SjxD0v1tiv1FY6GOaSmrKiVCS7Q0zlRY2t033HrhcmfvJTHTlmaaifhNMwJ+4tvhhe3SLoWtMlF1ExlYRJEES56X6+uEZdQ1lfUCOhpJ6qW1+XDEXa3c6Rv3xGmbOHvLHks4Vhsryxp97tscaD7B/0j/ygwmqy408ZopRrNQj3N12spP+BjnZtmHyfgauJYASxpIHWELKRe8AqAr+A+I8zRIXoK6mhuNeqglY2PXtg1fZhEJlLZLmmsAKrrBIFC/6un9+PpoD0x3rj48fKnjDvQb7StzcGwc186UvCmY0sLQ0mT5ksMXhVGo3Ued9hvh4cO5kUI/QVf0sCaR/n3GNR9o+f8SZDVZOuT02UTQZjWxUANW8gZJZNVj4RbTZfnhqr9oNDSZhPklRBMM2hppmB0Wgllih5jqtzr026MVANjY3x1KPbvNa1JtSnhmuDgTZxsBvPRUadMGCYWcRZPniHWuSVlgdr0knT7Y5U5Hm0kZb9C1oJB6UsgsfkBi90HtKhhgravPKeOnpqehy2ZDDdmlmq1b8MXIGxAAJtsSSRbEnlntFyLM58rp8vhr6qbMjMsccUGrlcmRI5eYQbKFLg3uQR0JuLsf22zqkTOEEDmCY21b+q6Hh0nWlZSnDOcyTRqcvzFNMe7GlcdrddPfAz8HZtFKpbKc0mLACWS8gAAJ2C6D5m2+NIz32lVGUcX51kc+TiRKdYocrkjfeqq3jVhA39knWCD0srntgCj9rs36A9/qeGqmWVOG1zuQwH8I6gx0AncKLbt89saR2wz19NtQYRpDoI728iQh0UQYlVE8OVcAeakyLOmdUuHMb+Lb4dJ7/ADw/Bk+cDUj5TmQ6Eg07k/K4Fv240pvaPkkE1PDW0mZUztHTvVMaYlKIztpiErX8JY+V/M2xdRbHLxPlHzLCRxcKGztc3/P4jqmNw9N2xXzpUZFmcYEq5PmTk3AtSSFhvfyvhDZRm6gn9DZq9za3ujj7i2Po6wPbHCPTCG+VjFD/AA7fafgocEw818xzI8Tm6MhQlWVl8SkbEG/lhmokqlhZ6elSoe1lRpLXbsLdPriX4kjrYc5zN2o5fFmFRosw3Blax2Pcb4jIpJYpDropTq2BVb2OPtOFxAr0GVTbUAfaJXPc0gxCR71URyCKWknMqqOZygCga24BuAfnhKzzspJoXUX6l1BHz3wqqnnS/Loq6UMRp5dOSLH9mFxLUAnXl9dpvbaEn72xoDgOarSU3HJZFLJJFqYgK5B6fLbtiRynKcxzF39wpJKoxWMoiS+m97bfQ4EMFWo5vuFWQTZTy7Eb7bHf8saX7EIZ0rs2eaEoGSEAswO4L7Wvt162xxe0ubvyrLKuMpAOc2LHa5A5etOoUNboKqa8LcRKLjJKwemjC14Y4iA3ySt37cvG+ADHtIx8h/WvmP8Aks/1fFa/MmdVgT8McQltS5JWj5x9ceHDPEF/+Y67/wAM9cb4xUDewxwFfPBDyq5lvwGf6viqOCp9Vg68L8Qd8mrB/wBzHZOF+IQtxk9WfkmN3JAtj1xva1xifrWzL/JZ/q+KvzKn1WCScN8QDYZLXG//AMI4W3DGfaQDktbvt/Vk7Y3clfMY6COm2C/WtmQ/uGf6viq8yp9SsGXhfP7gfoWsIPYxnbHV4a4gv/zJWW/7PG83A6kY5dfMYn61sy/yGf6vih8wp9VhI4Yz4+I5RW/LlnD8XDOeaC36KrBv0MW5xt9xa+2PAqRsRbE/WtmX+Qz/AFfFX5jT6rFBw7nwQ/8AsirN79EwmPhzOwv/ADRVC2+8RxtoKkdRbHdj3xP1rZl/kM/1fFTzCn1WKDh/PLXGU1YaxF+Vv98dhyHPI2JbJ6ux2N4ib/txtVx2tjlx9cV+tbMtuAz/AFfFTzFnVYgMlzuScxQ8P5lGioGd3isL32C9ybb+WHm4bzuSQ6MqqlFv1oyBjaRpv13x0Feu1sT9auYj+4Z/q+KvzKmeaxpsgzdIlEeU1bFviYwnb5DDRyLP5AAcorAB0LRm/wBsbXa+O2wP61sw/wAln+r4qeY0+qxFuG86Df8ANVY/e5jO2I+EA1Kq2r4rGw3xvjDY4wSAPNVzy01NK9NCW5krKVKsWNgFO5+l8e47F9r8R2gfWbXY1ugCInnPU+Cy4igKREc1IVMNMyIi6m0g3XSOn03w2Q8apGlGjx76yzdB8sDNW0YZTOZWYfEBE223SxthEWYtHfkx1iK11sFIBHqL9Me6BCoA7QjpeQJyIIjo021FNIX0FiceDqv4YI0nyUYYRldgxp5rk3AKdj6DCKqrWKpWIQ1QLA2tTuRt6hbYolCQeiclESklBtbxdMJZjGwOqxHfHuaojNhMCRupjb9+B6nMo4m5Z94DW7U7kDfu1rYtsRKEgzCcaM73YEXvY7/fDVTZnAaRCe+nphCV0Ls4dnjW/V4msflhmbMKVoGb3gOiAsSF3AA9cMsVCITjIy+E33HW3THI76QbLgVZxIi6Vks3drYdRo1YGSVfK/YYpzQqBTrLY+HSbnppvhCQKGBZNj172wlq6kUgc6IuTpF33x736CRv62I22NmGFQiiUIMt11bSytqjabmnRdSbRhNP7Ttb9uCrC1hpC999sKlqqdwYuZEyg+Ia+uBRNTxBmlnZwS3hFvCL7fbBMEFVMp2UhVsjr87YByiZ56JXm0ajJIBZbXUOwU2H+iBhz3ylLIyzQjcFkZrMflfEdleaQRwvSNM4lWeeyshvpMrFdyLWsRbfpjS1LIMqVjvpuDqub30kXx5WbSge7AEgEkYj5M2p+fGkaVLjoSka2+98KmqyzEGB1VTYM7AXwL3Qo1qkLBlHRQR3wDPRiOq96jqqiOQOHaOJgFkI6BrqdvthqStCIHjHNLW1APqIHfbCjUosu7tpA6FTc4gcCN0Om6MyWWpAczSxNIzXZVNyANrb239QLYkXIsQjByb+Eb4ipa1HUO+kuFC6tHiA9TgZ8waKO8U0YYd3ubj6YIEDmppUzOVMWrYEdNsDMl2VzspFgMQtRX5tLKYoDTabXDBGO9vn+7EhzZmkMWliu7ayP3dhiOeOSdTpkCSkutswp5BrBZzCQq7eLcE/Ir+ZxJKhVbX6Hysb/LENXsBE5b4lZZRdbklGDbDbyxOwNBPKJrsY5U1gKL2PbriqTyZCqq0BFUVyPhW6g3G97YOW5CKdMd36HAwVENluQQbi4v8A34WwURxiXwKvTrc+WDKXCK0MFXXIrKWBJ2G2OVCBrsNSr/ZJ6euERFGIYix7lb726bY4dLamWM3Y7KR0wKi6FhghLhSLncWN8MSPGYxbTrkN9v1exuMPyl9LagEA2UargD0PfAFUyxMHOkyuNlHYeZxLqTKSRCsWgoDJ2JOw+mFNIioEQREg23S3164ZVAjc4gve5ue+Go2Rr7Am/Un9+LNkG6cZ3YEEr6ADoMIpISZ2UWJYAkncAD0x3lrGCAvhta43th/K6Yga3VnkbbYabDc3xRNlGjvKTNNGYipQC48QNutuoGB5Fjig91pmAk35si2J37DvfD9d4ILuim4+LzBHmMBl4zGyxxsqquq97n67YppkXTDHJdoYY5nRFpyscA06OoAA8/PCMwp44TYIygnUbgb7dOvpgmlqI41PMkBViL9bHDl0qVOpdMfZzvitUG6stsoRowQzkhQBdRax/jj0XMlkIEeru23TE01NC58UkjX28I6jt17Y4tGkQbQigqpLEmwvicQIYURrAUMLAFrC56nD2pi+pW02N+vXA4YtIsh/qlQrEH8OoXuWII6n93bCGmK7m7XbxC/w4aACpZTd45SzLpY7Ei+5wPIjqwY2FrEDfAsUyxkqjg27MbE+uH4agj4vFcWAY4ohXsn1jHIYEk9wO2OotmupB232w20oIBLAegNsKp5BLFqHTqBqvgFEuTxdVHlfDQLohUjWD2w6HG6kOSfIX+mPOhje5Nz2tiK1DVsckbsqppU+IDDaCUwF2ZzbexsT/jpian0ymzKVP3xE19KYQSHLCwvpOLCGEyyusZBNyetjuPK4wunkLDxajtc3wwHRtIj0302G2x9cKuURQxuQN/XbByonXm3ChSBfv3wqOTUw2t52wmV4URQAbvsPO+ExAakDkJa99XT0xSEhFM+lAbaSWO19xhgqGddesAdT88KqBdh3byt0x2FWZgegN+/XFiEBuUyiutwEW7Nfdgb+XX6/fHnflzurqfDbUy9/v+7C5AVPhIt6HbCBuQ0ahHN7G19j1OCEoyY2TihWYlLBB09cLDIBqtbz8vTDSOqMUJZyVtuOh8/8DC3iQorN4WG9uoG25xRQC5uvK7MNIJ2uLBcJK+FmAuel97C+OsjMumO6hfIdcJ0nSC243sB54EXTAAEuXcqkaWGx8P8AH9+G6xkSimUoxXQxZVAJO3546NTqhZWuOwJwioRjTTI/66kWJI2+mKQc7r6c8H9r8sew3oPnJ+ePYy3W66+dpryRGWdm8PRrfkMCVMloLyAgubKRsSPr1wihd6se8Bi1MbaF6G3fbCK+r0ytGNOlbhARuMdgWXIPigJidK6AwN9I7hm/uwdHUpEBGGUvGlhZe/7MAsXYrHp1gsNgLWHc9cSEcASHmxSARptvvYeuAJmyIAhERFRSExsA7WupA6n174ClK/CwVWF9l6Y47Fl0XVwm6EG9v8eWBmMhjIlhU3N7qvXC3FNZskTTTLLpa3KI63thyVo9JLdFsSCLi2GjIdQ8AsD4fDhMsnLsdK6bb9RhRKaO8lDRaMRiNoyfh8u+OCmmMrmTTygoZVHkbEY7SFI2DSKU1bhbXwutkjCxfGXHT0viCFC2LpX9HCrbQwA+IjY/34FzDM1hkWmSN5ZBbXp/V62B8sCZ9mq5dTSyaXmmQeIIhbTtqufLbf7YrivWTQs1dJEZJ5GYlW1jdNQBA7dCB274z1a5HoptKlqs5SlVWV1QoqJWMCAfixgEhQAetvO2I+kkVqS3LCFn03dbFvELFBvYb+p267YMMMld7u0kLcsOVCM1vB5t36gbef1Bfgo4mEi6eZI7ElT8fxmx1fqiwUBfIEYyNJqGVpDQ0QlGBFEUamlaNY3AultiRc36A6bD7YHWnaR5HDSLTltZjvZm6b/Ui9j59MHVZA1WjUNazAEEgDa22EKbm4YNtvsNsU+xhaGi1klIYJEQyx8rS6vdr31C9r+f7L4dB0ABAVS+zN1J72BxzWgKbAsDa7dDfA89cERbqzSDcKo3+W/TsPrhZCiIKp4ruQpNgdVif7sCzVMTOwhZWCmzNfa4HQnucNVKzVhkE8htfaMMNNj0uQQW+RFvnh8U8aVEcpd25cYQWPU2Fz023A+gGK4ZQmoAkUkTVJ0BypaMOAw+HfoTv8sKqQI5C6yCWS+7G5ttsAOn5YGziGuSCKbLZAskLeGANZHXyI8/I4coJYq6LXrJKsVKD41YHoR2OHilAsszq5Nk7Tu8i31gOT4rDqfM+WJOnhjP4kd9djq7r9PXAyppYnU2w+t8KqqpctpveKyeOnjVb63bSfpbrg+AVYrAIlhSRIdUQOo/AB4mGIatzINUR0VPB71mRe0FLTsS0a26sRsg8z645WS5xm9N71l4/R2WSDfNKoWBH/wo7XY9Re2LNlPDiZaDSUkL5XTWR5HRtdZVMRfxE7Rjfe92vcaRg20Q26VVxLnGG7IbhvKqwRilmgWrzIO2oRVDe7Uqm9g8h2BsOvU9sW/KKZqKokemq/eK5zeSsK6FiXcaYh+qLGxb4m3vtthqlTl0yUixpTUiEsIIh4bnqWJ3J9W3w3XZisUVkKhVB+HYD1JwZcfRCU0CZKkTLSQRF9Zdl2JIsW+WK9nmbLBAzzSLFESN/UmwHmSb9BhpZa/M6U1OWwcyDUEFTMeXATYm+s/EBbfRqPTBVFSQUc4lgY1VeTYVsiWEQPUQpvyx/pbsfPtiw0kKOchaHLiP6TxKuu9niypX2tbrUnv1/qhtceInpgwyVeblZzWvSZdDIInquWACAL8uIWtYdz0HQAnbDGaUvKjketki5cZ1JSJIb1JuN2dQdK36qRc9DtcGBzDNMzzgwR5ZpipgBG00UX4UY2BWIHZj28hbc7WwxU4wpHMM6hgZOHcmlWPklefKGMjqv9slvicjoCR1J6DEdIaeGJissjuReSSQ6pHO+5bv19AO2DstymCjhaOFVQ8xjICQXc23Yk7k9LnERPAOaztqRFOw8z9L4FwhQSRKQrRMx1y+Nuh2xyzXGnWSQQptsPMjHTSurbhTqG2odfl0w5GVXUdR2XYeXmPngYUSo4rrzL/iaAuknb54FrYZPczMApmVy0Y03GsC6k/W2CIqpjuLkHrtgoSSRxNyY9T6fCBa469L98EBKtu10iqqYpKuNqWnAhnp9cbhQpsCF3sSDsV8X6x1HCKppndBErGyqSGFgel7YC4fp6oRJEtOY6VaalBJsA/9Ehv06nV1PqcHZjOKChblI7zD4EXqw9PnjHiW6nyteEdppJCqseqSbw3tZdz1Pf74bkrqCBAJqiEhSTY9sVufJ83zqRpa7MZYowSukN4ACLEAKd9jbfzOCE4PoyV95r6qW4LPpawP8MUKUBGazjspUcQZYJXCVTsyj4dJsR88XT2JZ9l/8/45Kif3WFaSX8SokVEHw2FyeuKXScM5XDHZYpyLXJMlyfrbBP6GoW1CMVCkDrrBv+WMOZZc3HYSphnOjWCJ6Soyq8OlfU68U8NEf/SLKbelbH/HCv5z8N6bjP8AKrHv75H/ABx8sU+R0iIpdqlmY7BiBt6G2ETUFBTZhTUtTWPE1WG5KlQL6dzY+e+Pm48leGP+IPsHxTXYpw5L6Wz+s4Ozv9Hms4gy7+gV0ddDorox+Kl9N99xudsQY4d9nxz+XO/5wQPJJNPMYjmMfK1TxlJdhvuD3Ox6W6Y+XK/jHh+lSpSmiq3qIZNOmXSivvYkEX/PB+S8RRz5jR0OaZLWZWKs6IZZWtGzHpuRjoUfJ1UoM00sW8CIsBsd+aQcaxxGoXX0YOEfZ61MYDxGXvTQQBmzRCw5D64ZN/106A/2diDiUyXLeC8kraLMU4iilqKOKpjWWozBG1id0eQt0HWNbAWAGwGPmnjXiOi4XzWOhmop6hniEoIkGwJI8vQ4oGYvxNxXDU5rBQVNRl9I6wtyU8ERa5QH1Nj9sNf2Br1mljsa8g+A5iOvSyB2May+lfctD/NGLNMyzeLOKCaXMXiklLVUbKrJHy1KeW2IWPhLgOOhp6IZ1/R4sofJnT9IqBNSsGGl7dSNRIIsQcfG/D2Q8erOrZNkmfM/lBSyNf6Ab41nhzJvafTsicTcE5olGYiRVrRkaLb3cC9h5mwtgW+TmtS9HGOG3LoIHPkNlTcex/zVt03C3BM9RFNUcSSShUp1qY3zJNFZ7u2qEzf2ip8rXtvfFmyatyDLYZoo+I4qkTTyVBapr1kKl21FVJOyjoB2GMIrMqh0iwII+K1v4YDhyhWcjXpB/wBEYzV/J67Es01cW4j+UfH82Wxr4Mhq+kRn2SFdX6Yy+3n7yn8cKOdZOL3zShHzqF/jj59joqWmomvY+d+/5Yhnk5lXJAgaQpII9IO/whup9GGMY8ldGf7Qfsj4onV3N3Cnc9no3znM5EK719QwI3DAytY7dQeuAVqYrFy6hegsCMRdQzwgI0Ltc6fCdVj62OHKTVquARt4rnH1XD4VtGkynPogD2BYnVXhSrVcZs3MjK36km+Fc6nEZkDi997HEcaZJnLaTtsLtYftw7yJLMLWudrb4Y5jUTXPci3kprfEGJHTyxofsQaN67NWQG/Li37dXxndHl0lTOkMVOWP6xva3ljTvY3lb0E2YzMo0yLGoI8wWuPzx5HtywfIVf6v+YLTQqEugrShj1xjhO+OA4/NsLWsz/lCQPVZXwvSrRR13O4ip4zSyTmJJ7xy+BnANgfOxxToaafgLiPguXiuvpMronzfNKiKE1jSQ0ML04EcPMe1wG9LXbbG8yxxyaOYiPobUuoXsfMeuE1NNT1OkVFPFMF6cxA1vvj0+Bz/AM2wzMK5ksGqb3OoEW5CJ6E+MJD6EuLpusG9r/E8+d8SRVvC0Vfm1Nw3QxZnFPl41w+8PKkg5hBHh5ETjv8A1nTDea55Wrx1xXxvkU09ZlUeXZe1dSpvro56djzlH9uMhW/1deN8ipqaJWWKnhRXFmCoACLWF/pjkVLSxBlip4UVlCsFQC4GwB9MbKPaXD0KTaTaEho0iTuCWl025kGIiJQnDuJmV8+0XD2TZhQex+asoxM+awCGuJkcc9EomKhrHsQD9MP13EstN7VhxdDSZmMly7M48gapVT7mlJpMTktf4hUOpvbog33xvnu1MBEBBDaH+q8A8G1tvL6Y8KWl5DU/u8XKY3aPQNJN79OnXF/pS1zialMuBDmwXbBziTFt4IaD4fUp5seRWVe1j+bEntS4dh4wqqSDKWyetYmpqTChkEkGmzXHisWtbfrjPa6pkqcn4LXM6ujlyts0zdMvmz+qkhiko1/qDI/xHb4Sevhx9L1FJSVDK1RTQSsnwl0DW+V8eqKSkqFVaimhlCfCHQNp+V+mAwPaZmFo0qZpk6RHpW+fcWse9c+Cj6BcSZ3/AKLL+MIsrPsFhgTOcvyuhYwgVWXmSooweeCVLA6jCzXVjcWBOKDFmUK8I5nltKsdLlFPxFl6Z5U5NXyTUDUcgvKYW6xrsokAJtc+ePpBYIFp/d1hjWG1uWFGm3lbphMNNSwU/Ihp4ood/wANEAXf0G2BwXaRmGY5jqZdL9Vz4g9Lm3ORzjrbsOXGZ5L514rmyemoeP6bgOsSThlOGVkn91nMlNHXc06dDXIDGPchT2F98TuXZ1wxw3xzJmHBVes+QUmRVNVnsdNUtNSo6aTDYklRK3jFhuR2xtcdFRR0xpo6SnSBusaxgKfp0x1KKjWBoFpYFhb4oxGAp+Yw13aei+lwnU3ERBl3pS0Nl1ruESDaDHRV5u4GZWMewrM8xyviybJs7o8zpJeIKL9Kg1yaRJWBr1Ai3N10yRkdNkO2I7O+Kqn/AJWZOMkps0OU5VmsWSvVKv8AQ1piGjnLG/xCeVDe3SMb43xooWkSRo0Lx30MVF1v1t5YQaWlMDU5p4TC5JaMoNLG99x33wv9JKBxb8U6hd7dJE2jnFty2B7SpwHBoaDsvnrKKnL8u9qtLUrmFJnc1ZxJNAJYauaDM6ZmLgxzwNs8CdLgAWAIw3RT8Wj2S5azPlP6B/nGoD8yX3v/AJ1IIP6ltVx1+HH0OKKiWrNWtJTipIsZhGNZHlq64UKOkEAgFLByQ2oJoGm973t533+eNLu1lLuHgzBbuZ9GdrAA35D6lXmx6olce744MduMeGO61LjWAOPmmR8yE0y69WqaTuTtqOPpVjcY+dmRVqJ9rDmydN/1jj655J6eutifU37ysGNqFhb9f4KO59fGQpZAT3s1v2Y8Jq8QMXd9N/1bj937sGEKdixH1xwxwE7Ob9hj7UaMbLKK5QYevYFlq5FubaW1L+7HU/SKt4pyd731dRh+wV/Azae1zthbFbBdamw898QUZ3U45QMgmfd5yGB6b7YZYVJ1ASjYWO5ODmREY6Xe1uxBwkxoBa4b1Nr4JtEBCariVHch2GmWaM39Dt6dcDSUYVtPPhSS3xAEG33xKNFEykcxTbrcDDRpoXNhJv62wzQG3QlzkGtNCrANUR3sQLdcNTQws5vUg2NwCNsEiiifaNiQDbbCGyunKsvOlW/qDY+fpgbEypqcmDDRg6mq6cMflb9uEmkoSAHrIgB10qD+/DoyuiaK7u7EbFitr+uOjKqAHdtQPQaTt8sECCoXOCCmpaaJXYTl1TckAbfn/fhCe6yR71ZVSLL4D4vOwvfEkcmy23i1XbbcHf7HA36Ey6IOy1kxY/CzKGKDyG37d8SbqByCmgy9DqeuBF/7Frfc468VC9mjrla297gW+pIws5JQtfXVzMCbnwdcDzZHlsYFq2oNhvcD+GLULjKYir6OKtlgNRIxjjWQtFupuSALee2PVVZQxx60mkqCRdVAG/54ap8jpUqZZXqCWlBX0t22tfp62w9Lk1K7qPfXUL/8MNc/lgCJ3U1u6oeLNo1TmCmIIA6SD7dMFU2b08k5aoinjQjZtYY38rADDq5LQFB+M+2xDJYYefIsscBWMxubmz7H0O2KFPwVa3Jh86yaCdQZqsqf+l02VfmOv2GCEzrJJgR78AiD4pAUH/mGGZeHMsKMsEjLfZte9h6dMA/zUyxnctVzrr6EQ3A/8+L0DooHuU/RVmXztGaesgm7qRbf5eeJAVLxnUQCDsG6jFXTh3LKc7SMmomxQWYn/b2+eJTL8uqqKPwSSso8IRn1qw8/MfniaSLpge6IU3UBZiG+Fxv64eyBdQ9yi0ryjZBva3UD7WwLG1lCDrp7jtjuX1Bp89p43lQJMDG19rP1H33+3rhrLBA+XC6s0Jk5wVlJIB77/L5YWJnBBvZgSF7aNwf3YZkllXTpVri4udgfmRhHNDPrOhmPxEDYfwwQEoWmAjzKuizk7WIsL7emOwzcpwJVAJG62wNTz2YKbopupJ2AwsMnNsJgwNrgj9+K2ULk7UNIw1BRcajpIuLDr1wMFWQNK7qNVwLGx+3bDlTIV0JHGXIIuoO3yuevyGOxRiONVKgCNb2sfltiTCFBTU87aCqNbt4ScKlhk1EuAOlxax+wxIBoo2VwWuRso2G+G4aohyUVb22Y9hiiVaYaFlUiwKrsWG63tg2nWKOANdQWB2B6jHIQ0Y8QXxdW2udyTv574ROxkAF/CNvCoW4H5YiiZrKnnJpU3W4uunbbtgVz4WIOxFh/DD7MYUYqWKOxG/bA8hIuriwO5scEqhcBvcMg1Dfpg6iaFYwspe+oFSBYDASEoFa+pSO2FapFsDsT2IwLhIVgwVLsIXOtbnSPES1u+2GaieAqYYgshLC5J6i/niPDOGLF26XsPPDwRtI1sd/IdMLDUZdIhFNS07uQAgGrq29h5YFqsuhYODaxF73F/wAv2Y5VIW8EfwnewPbHYGYHlhwGBFid729MOFkKhZo1jgeSNo+fqITUN2BYFhfzsD1wTDzGYSFGETbX/wBLuMENA01UC8VMlnLDxANa37zf7DCZhJTygyMNTAMAFG3rgvUoTASGAFrm5639MOUbCKUXCqCDvpJ3wNzUBOplb0AucPwyKbjUVZe1txgSEAcpWjKsTocMQOhN8LkAaTmFyLHpgGnlZJC1gBbbvgqOpVkJFr4EiFS8ypYaR64alijeMOVUm2m2FwTRtqUyAbW3xxxqIj1bDob9cQBECoCtgEcxfppPfa4whNDxlgAxI6K1yPXE40TSJpkUsBt/i+BFy1QVePUirtpI/LbBgooKDWFowHC7HbcYS7FwUNum+/fEhJDIQoLHb9QjAjKAnxhm2uPLF7oU3fSQGYlmPiF77YcLENpRlBFynYXx5hIF1bXO1h1P1wmPUreJDYsNW+1r73xQCB26JTSWbuzeXQYZKAkhTruR0Ft/LDSSSKpanYkEkDa354fplc+Mra17sdvphgREwuFWW412Qm49cKfpvsNR746ptGdLK3r5YaaQ9De46bYEpZMldnN00J4j1ve2EQNeQKxK72Xw2wogMoDEabWIJwuH8JDosATaxN74EhE0pCtypNJHXzO+PV7p7rIR10nv6YWQpJswuMCVMeillJBuyNbSN+mKVc19G+8t6/lj2BeYP7Z+392PYyyulCw6taKB0pkIspCkAdNtr26YjKiN5ah5EAQLfrtYDCsxq1qJA0Nv6wKG9AP8WwJUNKrhZE1xObm1weu+OoT0XIDSTJXYaYSSM8UnMZ7My3tZr9ftgyqOlFjVtwN7d8MLMojYQxNpYC5B/wAb4S8hPR91Hfa+KRALjSc0FQSAd8IViSRcMLeWB1aPnc1ioa1iy4eWzoGDFl1Wv2OFuuUcQEiQPq6qLX+VvPDbpI6qP1etjtg2mjaWUrZ1G2wF74YzKupqCSSnnlhR0ILKDc2IFtvrgbDdE2Zsh5iIyGJJY+uxxH5tmkSosSTprclb7XWwvt5m9tsQ2b1skwq6hdRiddCushHJUjdmA7gnp6+mGo1Z1iibnhYbjToFh4uxG5sNPzCnGN9cGQ1a2UiBLl6KnNWySjmrDUhRHFq6otrm/XSfM9du1hiRaaKmp3mcmNVLGNYrbDT5ept072+WE0lPMkJMrPaSFNYOxChQNNtwLm5NvMjtfHkjUI04SNlLAgh9S7dBt0G3T5+ZxkLTyWhro8ETkqRWZY4/GzNqRgWZLtckk287+l8deYxtJGruuphqRBYEkDv87m/rgtKlqVnUwCVmU3cggMPO19tvngdY2Ntw4030A2c+g+n1xAdNgoWzcoeN5jBHEZDKwYmw3/PDkjLa5XTbv2OFZjTxU+azwxMhCsOjdbqDhlm0uGqJNILWsO+LqekmUx3ZXCjVk3JKmMAAAg27+Yw5PRczkoxaIWPMEe2vsN73tufXHIkWKaUnUdU1yCB0sLC/a1sGFjK6sCUCpbcX27YtgB3QPcQE37ry90QKnkosMLp4laX4rkdfLHGjkW+tr38z0x5QwjuSCybDT1+uHAk2CykdU9PPCiGONrnoSev0xXczR8qlOZ00QKM6muSOygCx8V/O9vsMTOZT0tDAKmrqIUYkFSzC9/IeZwAtBmedUsc1XNFl9JJssMq+KW9wTa/lta39xsZp33S3Gea68tdnGmi4eZVjk/ra5nASHzsDux+WJXL+EcmhzNWkM+f5hCq2arbTT0y6tnftfrsbk7gKce4L4aloMwmaonq5oqhNcTAaCfFY31C6ki5vYkjf1xcqifJ8po7VNRS0VPBdxGG0qDtdrHdm3ALG5O2D1RdC1vVM0dI0s8VWZHmrFXw1LrpWIkb8lP1fLUfF/q9MFTSQ0a6FGuy+Ig9/U4reZcXxlkgo4Z3WRdnSB3J7A2X18yvfAb5hmqH3psjQaFBE2b1capG3mIlBUel7n1xLuuVRMKc98rsxRky2mNTArWMxYRU6EHe8jbH/ALuo+mPJQUsbczMJYc2qEIZIVQrSxHrcgm8vb4gB123xFVc/EVU4kqcxyuJESxGl5NAHkSVAHpa2B6OkmzASM+cVlTHFYO0IFPGxO4AYbk2HQNggAEBk7KezLMVjf3vNa2OJFsNUzhQoA2ABtYbdBgOpzGpkBGU08tQjC61EqNTxW/7wuevYEHzwmno8so3EtLRRB1FmmfxOL7m7sSx3v38sdaZPdkWOdmbUWLnqd9ge3li5kyjAI2UZNQTVi83M6uoqm0lTGIwsIv8A6A+Lr+sT07YlKRitJyZJlVbDwMNJFthsD0sMCzzTBSwslwOrWv8ATDdVVGKAMrRkGwBB8V8EDCosi6ekmcWfQ17gfCTv5nA9bChCswjJuGu2979ThME85QLIyIB132b54S6yPKLTKAdrdvzxUqu8h1YvdXsQvS/Q/fDKqBUldQ38hffBMojZgQ4t0Jv3x13CxqouASSW88LTGiRdd5CqBo69xbBlJyo2tIb2Uk4C5p0khlCm257Y5I3L5jGVCSpCsTt064MItkDwtItTRGulLCRZZYYBc6eWkVMjem5VT98GVLySHQ8fhYq6sR0+XzJ/LD2RU6xcB5VNq5cbNVAp13Mim/5fnhsSGXLYwTYarAlb4VUbL79EVAxTEFN5fdo9TlrBntqItsSP3YO5DvGzawNXQm2GaMJyFUwp4HbSRvb1wfBcDSwUnVttjNViStVMWXGh5EQBcyEm24wRBGnLJcKSV88NwIZnHwsB3JxLwZezFUkkC3AtZbg4RumiVGrHCoAcad/Dc329MQHF9SuXZhlmaAF1hjqYzHoJJJj1jp0/q7X7XxZqqIxuVkU6Vawv0uMQfG8X/urXuElkZYCqqh0nfa+/YXv8r4JhAddBVaS0ws04EoKbNc4zBJaRhJK6ErMgJGpnO1xttp+2J6szKB+MouGXokkioHUiaU6tVowLWIsPi6+mJKGizKi9o1TNUUkySSUtJMyhSwDhAWAPfSTY9eoxH5nkNdRe0t8xEbtS1cJdZFXZW8IKnyO1/rj0tGo1tOmxpsSZXjMRSqOfWe4GWgEb7wis+ljoPbnk71CLVRRURZkZAyuNM21u+HvZHXU83A3FkC1QWWozmjmjiK7tGomufQDUPviJ44eWP2s0VRHHzHjy8uLWINlm39bfTD3sZomPDWf1qU94Yq2njLg/AWWSw+tvyxhwkOxDQeq6uYjRh3R+dlu3sgT+mxkNvqtjc+Lat6XhSvnQ3eOgmMYtcF9B0j74w32RrJHVxOytYPfocbVxA8b5WipqqTGVlKOgvZTckbdQL4349oNQLnZaXCkV825RXS1hlirKWSizCn0pV0sxGqJyA3Xobgg3HY4kYIQJtTWKnrvfEjxjJBV8Q1tZSGTl1EqnlsLaAFUdNRG5BJsB637Dae6r164828BryGmy9bRLjTaXCCgczCKoDqNJ2tbr5YqfEcr0VTDLbUsjaQbf1Zuo+e4Y4tWcuFhBZlj03JJawGKBxZWxvSJXU1SKmngUyuadldrh12uO2knDdB06wlOqDXpKLoXy559psxh0hY2DxgqvmRpBvYeeJYwUsbPypqpVW5XSosy39fTfGVVHtBiFS8sEFaRquFMiLq8wTYm3yOL7w1XQZtlLZpTl5ozEUaJZCdLC5K36+nTDn02RKzCsSYUzl88kk4iJZrG2oKAT8x1xJKqMgvG4OAMqymOCqWqLgnWWs2rv5A98WAcpkKlRt02wg6S7u7LXT16e9umqcV8VMz0lM55p0qSwG/le+L/w3FmuX0oK1XJEkYLL4SQfPcYrOS6/c9KRE/jLYab7/wAMXCkaSV2hAeK8Ytdfn374GphKeIpmnWaHNPIiQfqKBz4NlUPbZx1xFwt7PqnMctzYQVxkijgfkI5BL2bqCOgbtj58Ht79q1gf50sPnRU/+5i9/wAqnOI04cocliJEk1ZzZNh4hGpBufm67emPnQeZ2xhd2eykWGGp/Yb8FldWqE+kVp59vntVsf8A3pb/APUqf/cx1fb37Ve/FLb/AP3Gn/3MUQ5DmN3GiMlOo1YQuS5g2gtGiBjsS4sMPHZLA/7m37DfgsnylTH9771fx7e/ate54pYDsPcaf/cx4+3v2rAn/wB6T/8AqNP/ALmKO3D9eJRH+CWIuPH1/LCHyLMNhpQ32Pj6fPF/ofgv9zZ9hvwQjNKf+b71ex7fPav1PFB//Uqf/cx4+3v2rg//AEoNv/wGn/3MUaTh6vR41LRDUNrNe+HDw5WFbXMbqL2dhYj0tfBDsbgj/g2fYb8FRzakP733q6/8vftW6fzpN/8A8Bp/9zHf+Xv2rAj/AN6Dt/8Acqf/AHMUeLh2sZvHLCqA2JBJ/dgyPhiEozNWSE32tHb9+GU+xODqbYNn2W/BLqZzRp71T7yrafb77Ven86Pr7jT/AO5hP/L57Vrgfzp//sqf/cxTpeG6hRdKuMr21Ag/bDcnD1TGsbrUwHWL2AO35YH9CsGP8Gz7LVYzmidqv3q7D29+1Ujbir/+xp/9zHj7fPasWP8A70Wt/wDcqf8A3MUluHa7SDHLBIe6gkEflhn9AZiSzBYtK9Tq6YE9jcG3fBs+w34Ixm1J39771fR7fPapcH+dB26/0Kn/ANzHD7fPasD/APSkfL3GD/cxRpeH8wUJcxHV0Gr+7Hhw5mJkEdoQbXPj2xP0NwX+5M+w34KDNaX+b71eR7fPaqTvxTb/APIYP9zF+9lntw4ozipfK88zv+mbvBIKaJVkXqVNl6jt5jGCnIcw0cz8LSDawbCocnzKnkjmiYCRSHTS9iDiv0OwI3wbfsN+CIZpTn97719i0/GWfMSDmzP6iKOw/wDLh3+dvEJIIzVwfLkx2/8ATjJ+AM4qc0yyOathdZ47xSuBZSfTt9MXSGYKh1EbDbbGJ/ZzLGOh2Fp/Yb8F06dd1Rupp96ti8T8QSOoXMpBqBK/hx7kdQfD9sQUNJFJPJzLbklyFvY7+vy7d8CpUXDq4dRe6qrWAa23064coqphHK6XDkgLdQxO+9zjbgcuwmCk4ak1k76QBPsVPJee8vClDgxxRgknbbc+mFDKZnJ005JA7deu/wA7YkaGeKCN2LJI8js5KqFCAnwqB5gWGGYKuWaV5XVuUHe9reG67D76R8746HEch4YKj6vLZKUnmRrcG1gb4QtHGya3RiqmxANjiSzKtgqHJdGVerC9ug2P3JwDDUrI2ooWGonY9Db/AAcDxDKrShJKWmVyq6jYXN22Bw09NGVuTqI6+mDIpCwfzO3nbDcrlSoexN77YLiSqDUyKFVBAsFJ64aejAIAO1ugGH3nUOVF79beWG5pZB8LnfuO2ANQlMFJx3THuhN0PS+5BtthMmXFSGUg/M3w7rdN1u3f0w9JKZBqKFb/AGwO5UgiyAlo4mAXUNS9QGtho0caarMG898HsCCDYkXud7YH4hkqVyuUUBWOpZSI3KBtLfqki2++GB0GEDwQLIf3ZQLKu/ck4YenfWxYkgi4HZRiM4KzjNcyySKfM1pxM+4dALOvnYHY4lf0qtKVNVNTwam0LzLbtYm3zsMVqAMq2sJEymhDqNrAWG5vgR6dDIS1yLW2xKVGawCmM+qAAprVgoIZe9iO3TfAOV5smaq80VPOioiNqlpTGGVhdSCRY3FyPTF65uFWkkxKYijjjO1OrAA2a9r4eSFGGloyPmAcSUGaiJlTlrqVjpdRcj526jBwzNSLtpIPQWAxOM1G6iYUK1CoQsoZRbcFcMzQObhIt79T3xOyVyA3fk79tAPTA1TXLe4hQA72IH7sWK3JK4RUIaeRnAEZJ+txbHZYHCkHVGwGx0i+JU5myA6IINRGwIw1+lpnPwwBm22UC/5Yovko20yFFxMNBEl5HWxN1sDiQpFSUloGaMg7hNr/ADwuURzhWIHNHwm+wviPqI5qWQzwaFINvCpN/PDTCjpR9SxjkCFCfF17HAdVrIM+rQYJopLg22V1JufLTcYlH1SQrddiL27fbAFXAJKKopjdTMjC4Nu2KdACg2VrqpAw1OrAA/EDY9O+I2p5/ILxHSS9hY7H0+eF1EuulToAYxdb98AzCdBzFYCw6FxZvT5+Rw1s6Ur5ykqVnRA0ykne6+QwfrgEYKUwZje7ByLfQ4j4xKH0DSQOhudye/5DBULMZQxI0i+x/diFXCI5qCOKJW1kHWzDcXIH7MLjlIheoZXa4AAJ/ZgMkgl7D5kHp5YVJIWSOxYRjaxG2BIRBPGQlEOnStzuewPbCqOFWs0pCKdhfvgZnICpDc2O5674bjaUyK0rXF7fFgdkMqSaYCMFSLkm/a/bzw1q1OdYLALYb4QpOkboSTsRbY4ZlaQbajta9vI/twQUTmtVj3Y3FrAfxwPH4pCNt+h649qCX3APw2OOxWJZtrAdAd8XeFUXT8lomvYsLWsOmEFmLXC2H9o/swpi3OAVdwR4gbi3l1wwrMpOw3vfSb4oq0rUVliNwNJ627YIedpCSoILbfTywDJYbMRa+9u2FxyBBuqje/XtiAISSEQWZpFHiVbWtbrhYUcwMqlhqB2uLfK+B1c8wlTuLgX8v34cjkJYAgWAN9rA/XFqSU9yxqVpFsp2j0jqL2tgWtRmriIxe7AFe4sfP74KlbSFuoYW2N+pwmgeGMB5VJYMSCva/niiYCMeKCeGXbTAW67A3sN+tzgWONoEQ2DKpILEWB2F/mb/ALRixzJzmZgxdiD33t1PofzxET0MqpGWQhWOrUDe3Tr9cUHAoXCDZKimURar/F3wppfDpuNPz3wApYEtcKgNrAfD974dU8y7rfT54OFThARUbJzSxtbYYJaX8S+xAG+/fEegu9gbKT0w+0gEBt8Q7H9uJCoImozCMag9rj+ybYXHUoy6v1ALDUd8QUkuweNwSW8RwVTtJIfCx0P5/sxcWULipNwGUOpVb9wMMvGApvGACbm3fAle7rTrYkKd8E5XqMB1rYgXxNlYJSHQMVfRdb998C1VKAVADKpNr3sfniRrDy12sAT5YFBu1tR2F98SEJN0BEJY1JZHbT0S4NvO5+mHzJo1FxZQ1wU6DDbrLMrSEEC9yo6fPCXTQpiZbX3NzsfLBwidsnYmFm6bna2ONpcbr898Mhroo1aQNumOMw59xsbWItviilwidFvCSLW6+WGSpJNjc3+mFhtVkLAk9ibXwvl6QFsDbqMUrSRI4Yg7KwsTawwxLJpjnn38MZGwuLW7eWOVJJmVGLH+yB0GG51aOmlJdraGYraxWw7XwCsbre/fl/0ftj2Bdbf5qT/bP8cexkkLprBJCQqlWXwPe17C/qcJlcRpYS6WJ2UtcnvhTQanZKnUEuGAA64fk0NZoxqN++xtjqEyuXskKypCgW4AFySbXOGBKS506WRtgThirfmiRTGAmqyW8sF170rkCmh5YHZsC10mFEHLFqXSpsQdyBe2FQ85UGp3kC3uDhxI3sD4rdSRgPNqmRYnotZ5s3hJjI1aR1G/bAvc1okowC4wnpOIaOihZ5SHlKsqrHuR/pG+wAtbFZqJ6mSnepnKxyzTMbyDSCwKnc+puABvYdLXxydahVtGsErx6pDCi2uqqQwBPqwBN+oOAZokncCZUAMwssa9gjabDcd2F98c+vWDhC3UKIbdIeSFi1OGIifSjm9yWMlrb7FbE2Fu+4xK0kbQCyDQsd2dibu8h3tfpYC3TDFDToayGcw81Sfw0UrcAb3Jv/or0B6L64lDEY+drA8TFk1bkj9lvLGQ2C0je668LsnMA8WlQw/tW6Y5TkGl5IlaNCbRwE2BIO1um2HZkMZUeIFze4H5746IZFUktYMh8TPaw/j/AAxbHEGQo5rSEpdIfUIgGFr7npbpjzojIbMdQ3ITqNsJpwV5sTlJbsNOnr8O5sb7fbDiapDGIQSzAjSOwHa/Tr3xGtJNkDnt6r1QEM07A73ADOCRsAOv0wwEaWSxVBHfqB29PXBFPFK1QEIjS+wtbr5YJhjKqLlCrGwHb73wypdylO7YSYoQI21xAEjDB3kCJJYgW0jv8zhc1QWLRX0L0uLeXbEVmOaUeXVEdIHZp3F1jj8cjdLCw7kn7A4cykYus1SoAVIOjROZZGVEtcgnYWxHfpKpzVJqThqIuY2tJXSACCP/AL36x9AO+ER5XV1MIruLKgxwtYJlcJ632HMYbm/WwJ6fTFmpssrqxYYK2I5fQw35VBEoVyLWGqwso67dd+2DJa1JgvMqL4c4fp55nqMvc1VTCRHPmdf4jE1+iKdh0Nu++Lzl1FBRyFoS7zPu9TPu587X+FfQb9cKSOOlgRWCqka6Y40ACr6Afvxn3F3FSZkaijgqGpssh2nqlYAzH/Np5g3F2/cdxkkoiGtCm874nMlbPRZPNAEp9qivkty4Ta9lH6x64rNJnfDqVxWGOTNq8traWRTKxa1rjqL+o++IeippuKbc5jl+Q07aI6eIWaYj13v88XGiWly2NYqKlgpIBsbKBf5nqT8ycMkNQNBcURBntfJGJHyOsQdtYCkD/vEYRXVRqqKaCqoKgwyxFWVbarHsLnriToKSorVvTwSyAmwIjYgn54kJeGc403fLyoB2LSL/ABxYfKPQSIWa0VJXQ1KZBm+bRRRykVFJG/4jEB9TRuSBdrfTyxcqvMlp6WOFEWqlgH4Sy+FBtYWGmwPqOnbEHxvkOb0MIqo4ImaeWKKEPKQYn1MQQe19Tb38h5WFyPMQ1HUQZqyLmFGrNUXB02S13G3SxH39cQWMJURZSsdRm7K7VEtDEr3CJFE5Iv5ljufphuMSq5Mkkl7WsVG5t5YXldTRZll71VJWw1CJfUIyTYje1ut/pivT8X0q8xafK6+sVCQZEi8IIO++Abcko9hJ2VgnrXYFeXqPQXO18NxPEtNJzRoa2xte+/bFaTi2IkSJw/m+5sCYxZj5Yssd8yy2GqWlqaO5IKSoVZSOo6b7EdPPDCb3Qgh+yQ7o0+jdrN17Ww40h06S5UeVxbCVg1ITbXboT3w0VJUAoo87LiFGQQn4NrB22O+29zjxaC4jPMI2vY/swmBQo30gdja5+WEcrfW4UMD2XqPPAF0Kw2U6kUbnQshBHZhscJqMvaouJ54BBoI0gkb4YflwqCSSWYKNOxYnoMEMZI4letQRgbt02JNgPL6+eK4hKI07L2TmmbJo5YiyCtknblFvDGIyieEdr3F8PmMLTJEGO12A88RHCEj1WXQwI5JoZaxXVhYASPGVt/sHE+8X9FJC+IWIF9+u+JVdsqoMKaot5NAChepAGD47h2AX8sAU0RWYgXFhe/niXpILkk76l64yOcCtjRAXUQB40AYeLa3niYapeeqiSRhcKQxvspJxHcoCQaQNjcgHDkbx80SkKQwuDfvhYTGmEUkfOEkkpcrzdrH7m2GGCP0GpCDcG1jhdLK8Mb6Co1He+/nhgaj4VNhi0DnAmCl1iRvZrXe17g/lgZwJk0aSnkCOmCZCWi2B779bYajigPi6EduuL1ERB2Qd24jdVnO+FIa/N6PNhqE1OrROq9JEKtp9Nmbp3FxiT4UySHhrgziWiaoF66uo6mBAhGyc0Ou3lrT7+mJuIMSoCmx/WsLftxE8WzikWnEkSmNywuw/0e2I3EuoniDkl1MHTrMNM7FXLgGupqOiSucSzQgixiXc7221W/wDjVaXi/hxctYGaSKR4HXS0ZPUEdr/ALcYFwZmDvlhopmJMYPKttcXNhsL26YnJdZ0an3tY7Yc/M6tYajCVTyqjQs0lenn5k7y6bXN7euHBMxIBAA2uLYGVbkkqQB+eF6GLAqp9B2xlbstxuhuJsspM3p4KWtp/eKL3iKSoh1leaiuGK3G4va3yP1xVZuE6akzpKnKKWhy6BYXjaOGRmEt5SyltQvcKQv/AHb9zi6MpZrkG2OGAMR4Re++2NAfNgkvpNc7Ud1n9JwNSRCVv0PlFSdV/wAVnuwuTt2viVy7KKuio+RSUVBCsQ/BVSwBDSM5U+VtZAI7AeWLT7ukYW2k288OaTrA/VI6YY+q9zdJ2SWYSjTdqAv1/pshdMhZCALXsfMDD5PLB2Bt2wprAFFC7bXvgcJz6pbTcsA2YX2OABlPI0hWGmfTlcSJeJmYAPIxXfSGPT0OJR4qsSVNJNO/4a7BDYX6AfLe+Cf0c0/uscix8tJFZRfUVsLYOqqQTNK5e2tQVYHr3v8AljUDAWRxBJXzN/KZyqSmiyauEcuiSWdGdj38BUehNm+2MRYHp6d8fTf8qOniqPZ3l88LkCmrl1Kf1iQ4Py64+ZWFwQAcC4LK7crRyqnMGPMAUgdO+2BWjU0bgsBpfa/lhUmpq7dhvuN/TCXDe6yEONn6Y96yNH1L5+Qde/RPMUWopmVibLY4VGwFTUre9xsfLHDBCtJFVJGRK8qgtrPS3l07dcIs3vU1wLleoOGtBm/5slQCLH8ylSt/Qom21K1vpjwZZKuMqRa3fthtr+4gqTs2488PmCKEUTQxiMyKxc6ydRv1sTt9MXJJH1KQAPamJFJV9DmwfceeHnbQI5FawPXCEVilQLWtjlSpOXxmxI6HBAESVNyAniAZJLHwkXGGtxAsi26/D1H92CJ4o6aqhjgBRGiUbsW3t6k4FVSIXiAJKnc4ioRyTr/g1MbI2lXG4HnjoGqKpUHc2I3w1WKkkUEbteNvCxU9PrgioRIayeCO/LAsoLX2+eC+dHJQ7Dr8CmHBaniN9wcKk8FUrHYEW3wzqJpjfqrflhTqs9VTLKLoWBtqK3tv1GKJtbeyINvfa65Ych1BJsbnC5WUtEVJG3lhmNmbmLYX8gb4SWJhRvW2K1CLItBn89FqHs9cPwXHTKDePMJpNR6HUqf7v54uNQgQxjr01b9rYrHsviVuEybE3qWJB6Dpi4aCzK7ld7b32vjwuZADFPjr+C+gZVLsIyen4pMsYY7A7i4IHXCVjF9Wq4ADC2+2C420xOO5+tvXGJ+1jIc2zX2kZbRUzCF8wp1igeTUsepS5NyAe1u3fGIGbBbqncbJWwJBOV1qtiLkXPXHJGeNwG23N+t/r2xT+C+Dfahwui08NbktXRvKHeCSd7t2IViotf8Abi4NWQT00U2oXcrpYXIYHbr9Rv6YY5jm7hBTqipsU0wlk72BNid7YILSLGselG0jY3/hh6bnRQyRrGxuNYAKkn8+mBIRI5VZRpkf4LxEC3zuQMASFZB5JdORdix8Xl2wPVHeyrfuD9MOOroSptc7YZOsCxYC3QkWwtxlFTBmSmDGLhtALd9Q64V4LtzFMe4A9cLUPZm1Jv12PXDKxyGbUdNrnqcWzdNeRCfEYYarjbfr598OMwVNN9VuljbCUUd1uR2Jx6S3LPhFz5YsjvJU2ldCgldYJUnpiM4sq1y/Ja2sa0QggZ1YeYBt174kDG5Cg3AJHfFI9s5kbIaZUQvarAYAXFjHIMGN0txhqXkdDXRcD5XSNnEVDNyuaJRCCEVpOZICSxUjxWFrfngCngohlUlZUXrquaZnjzE0bRrGmsEOuk6ZAwQDxDa+19jiC9l0+aPEUq5oqjKpA0RimJdgw6BQeguRi7Z9Uc2ikqpWYOsiszlu1wLHzB6WwJJokBw3UZorAlrhax9ae4PqMsrqGTJXzPLWZIpoqdVmCMBcMFZQb9QCel7dMeywTRQsKuc1IZQkGmoMgiiQsEjA0KEADGwHmfXFG4EyLL6+ozfPKqmvHl9bLIY0keJiq2fwuDZSB0uDi2wVklLW5hGsrVEcelIEEF9FxqFyN/1gN+wXe5OJoM6W80tjgJceSK0stQZF1Fe+25w4zPCxlAI/0bdfniDb9O1PGWUwMWoolyueqq4mksgMSF32B3ttbvuMH0lbWzJUrUvT3iJsYlI8rXuTgOGQtLqwdYIiFpHkLHZr3JvthTIxLOsukIdiRe33wLR1Uy3s6Fj2O1r4JSo+HmFSLXJ7fLpgWjmq2N14JNp1ySKx7WsD9BjogiA8KKHPngh5hKByyeg3sN/z7Y4yXcNe916jphghW26W7xIos1gR/j/Bw3mkrrls0kK/icshCouxNuw88JaLYdDq3G2CaVObyiApVGGxNhfEBmyqpcyna/8AotNTRBmcLGqBn+I2A3wzA6SlQWV2JPfzwTnqEzBJNKmw29MQ8TvBA1TEgKx+Ijzt5Ys7qgLKcow8eX08MiM+iFVLN+tYWvfzw9MqlUVUuBv1IscJotL00W5bmIHBO+x7YeYtewbbt2xraSGgJAbJJXUVwg135RPi0nf54cgX8MaidLEjW2wwmLTpIZj53/4Ydd2KabhhgS8IwLpuUuGAYqbHa3cee+Fa0MQuT6KDt8sJOoG5Go7b3x5g+sfCL9LHpiB0hR266Lg7r4Sbixwo3JNittVh6YbfUCt/EV2IPbHEMmtg19+3a+JCFOshjLLqIa91s17fXHiPAqkXsd9+uGSD8bKNm647ISVFjvfpbbF81JXpwS52NtW3X88djUKuolfQHfHZVXQTvq63wv4I1DqodU8IK9cFKpKDGNCrBVvbYi4/44aW6jUu1h36knCDq2cgElRe3W+HG16gzFD5kjEIsrSSG5YGklSbi3y74SFKBg9txtbyw4TuDpsCLgA4bLsG7t1IA64EWQuElLkI0HWpvawN9sOKpQf2vLY74HRmKXa1j2v1w+3MDELvfzxaHkus93RGWwAva2FySXKRofCLXUrt9sNnXrQhj4dhbvjycwKzsp+pucQhGHQjUMscmtZAoB3U7bW9MPy1IkaMzByXXRYWsOov9L4jneQhEsGLjqRa3rj3ilKqq6SlyL7D53wEIgZXcxpHeWN1Q6yQHVNz/fgGpPuxFMNDSEksQvT5+uJGKulMusBrBSSAerd7nyxxKeKWkP4itI5Juq9+97dt8FKowRZAQTq24A27Nh+WYTQLGiWUG5NhgKWFonZiWDkdRuDa+1unfHSXCAg3W3XywYQFdaF7A3FrG5J3Pl9MEUerQCNQK7kHe5wMZSoPgDaR3wXFLcEv4Rp7fsxcKkpn1EX7bHyOCKeTTFsoW3UXviN1sWIBBv1wnXNzQVJBPXffEIVAqWaWNgbkm29sCVLiMK4Ui++2G6d2kJ62O2+H2hLrovdgNiNgPXFbKFCIJBsBa9rqcdkksblE8sEOjwRFQo63JuOuBpkNiSxuR5Wt8vPBISh+WxU2QAevXDwUKLAE+Z748NlFwNVut9xjgZjpI03sbfPFFWF6OMbsQLkj6YU0ov3sMNtcxnmbb22PXDbhh8J2Ha++KKtOMl11Kl++Gq0SHLqjlEcxYmYa2spNvtjhLBBrbSb9D2xyoCrSymwYsCLA+mKIUC179JP/ANQl/wDGX+OPYa5k3+ci+x/jj2MULoyVl1RBqDOHMjdxuLYBmpwJNAZhbYdyPlg2tMkbusRKlfLvgQvYByzmUHcsb3Bx0z4rmtJTbxmylwCBbxW6466apTHG2sMBpB2OEaH1llKlm23Nh1w1NJyYmJOooCbE9LYElrRKPSSRCZzTNkypDQyKDWMCYka+xAJsbdNhiBlleCGOGCXn1NhJrLgXkNhqI6EDVcjp3wxUN7zItTNGzTr0kDByCTu1z12PbbDcZjNS8vOqLSBZAxIXVY2Nl/sgBSTe3r2xyatYvPgunSpBounees8s9VEUiXlnRZiwRiLgAeZCAm++pyLEi2CIbSzqIZmkMb+IvpKGwO49L288M0FOXgRjIod42dCUYqj72IB9DYDYbHzvg+FBTRKikuejkm1/4DCpWkNT1KsdOHuysW6uf1j3wpSGl+F1Mh8ItuwxG1UgL8vbmgKwjGwC7G97dOn9+Hsudp9kVpUisTIVve4HS4v1vvhZJVwEc1Rrlamh2maTSCV+BB3P2w5FzJUJJVilx12Pr64YyRaqGLME0xqLyOQR4vFMNA77m4Hp64Plkpo8x0yxzCVdSlXWwDKdxb0vb+GNlJjYlYarnTZVWKWr4e4nE5hE+W1BDs2o2iaxBIt1PS9x9jti0y1cc8UUsBGg+IHYlr97+WAc0hp6uKWm5S8hh4kYfHfr8jfFdyqoky+sXKaySRIRYQTu20m5AUDqOgFvp5XfJ2bZYyNNyrXLMqqdGgubar9Pl5YeWoikgjkSJC5jJ1Kdt8Vyrz+go0nhhjkqKoWVIkBOtreYvtiUytp8tyWnXO3CVTKWZFJNtTNpT52A+4wmo0Agp9Ks6IChY6LOuJKwQ0NPJlWVQThZ6uR9DMqkXEYt2se3z8sT/D2X0mXw1GXcORi/M/peY1JDENuVUtYEtbfQLdd+xwfSZdXVIRszV6SjR9SUisRLID/a/wA2L7/2iCBtibpqQCNAIY4oIgOXCg0ony/j1OLdVsqDLyUNluWwURkaGSWZ5D+JVTn8Rxc2AHRAL9B9b4PM8UAIIGq1gBub4aqarlDTGNRI2J+22ILNKhhHKquQAjGeS48AAuQL7E2+3XCWtLyjB0qD4w4hnqar3CnquVCBqqZlO+m9rKfW1h5m/wDZN6dluU/pnP2asY+40gTXBGbBVJskYLdb2Nz3se+JGtpZpLCSJvepmEzqtvBsAqBuwVbKPkcSvCtIaXIDKGld6iQykM1yOwF/l+3GgwwJIBe5TEWkyiKlCaE8Ki1gPpiXpIsupwGqpTqJuxa2/wArnb7Yh6J1SURvdS1yC19rdr+eFVVflKkSSVsW2wNg29/I7ffEZDjJTj3bK5x5qsNIRRTOiML7yA38ug/fiEWteUM0tS5brp17LimZ3xDVNIfdnSGJQNAYW7/bEGa+e5knna7C5LHYC/bDIG6DWVo+ZPQPQzc1vedStdD4ri3T/h6Yh+HchabL80iqtCrPTtSpV8y8jowa21r3UGxJ6m3XFTqK+JZ4tM6xSg+I8yzW8wOnliSp89mLpT++tHBoH4YsCB3NwPlgXXEhW0Am6H4G4czHKKlZ6vnUUQYmeQTAcxVJKkg32tbtiXeto46yTNoWnoGzKllppFl1Rq7JKrI3gG4K+IN32364qAz9oV5hqZHqJARobUbDbYb7n54Ml4mr1aizDMKaCwflxxtCGaQKOjXF+4N8FUxGoNhsQs1HDcMv786ipDKJa+ikNS1I0tTpKxVAV5NaKN7uXIFyFAABO/XbeRymeriyqopoKKsmjXNqyqdhTv8AioWjCEXXobsb/wCjiFq+NoJ9KT0QijXYLawU+QAtgmh4pp5oSYAdI2JsPnv3wD6pIATKVFjX6gbqwtmLaADluYwqi3saZzf5WH7cR9RnNNGVMtHXxgj4vd2N/kOuHKXOcvnIvIBYAXJweKmi0eGZACbEk9Tiy4BPBJ5oFMypiis0lr2IuGHX5jHZcwpYYWZp1Cqtzc6R9TiWj91coXcELuLEC+F5jQoMqqJW1EGnZjY9ABcfswrWCYVhjomV2ipI44jKza5WHxeV+y+QwnU+siVFNt7NYi2FU9Wk+U0lVoCNLErlfIkYSCWUjXqJFtjhcmVpAEKMyOSblRRQzqtNNEska8srzS6I+vcX2tJ163xMuhmpQAfHHuLHe+G6bl0dBw7A5vLUZLETI3nGqIB9Rv8AQ4XFKvjBaNTfT4ja+NFVvejwWOg7uApillUXV1tINip88SsQYKSD0FrLt+eAhBI7aboNgSVY39RgynsrD4iCLbjGFwg3WxtxZPxDUQVG52uDgqGnIBZzt0se2A4HFzGdF772YnB8RdtuYijsCvX88UAjYOq6ILINtiNiD0wgUxD3El/pheoAgCQjzsBjsmh3sZmsDYXb+GLsgIuhpBUKxWKK97EeRGHIoOYytIyoCb2sBvghUp2hHiFj0BbEVm+d0eXy+7UsbVNXYnwjwqfU4jjAkqg1XbgfhuHPZKz+mGA0+kDSga97+vpgrNvY7LmbaqjimbYmwFGtgCLEfFgH+TbVVtVU8QNWmQke76dZv15noB9sbN2x8Q7V9qM0wmaVaFCqWsEQIHQHmFuFNoiFlmT+yGHLprjiKVxpsFFMosPucTC+z1FVQ2byED/4I/jii1tPJSe2X3uly6vr6ibOgXWooJ4p6ePlaTLFUITG9Pb/AKN7dSLXwGufcX53keYLmyZrVUifouu0tl7K8RFcDOoVYlN1RRdAZLWPiO+NtPEZ69rHMxndIaTLWgjUYsIvHUxPJZzXFwWrSB7PY7D/ANqyf+CP44Q3s8U3tnD3PS8N/wD5sQvD/E3FVZ7WBQu9ccmepq4Xhmo9CRIiKYmB5e2ogkEyNqB+FemIifMuJq72h0DVD5hPV0Gc5kIaBqApTQxrSTrTtzQo1awVO7G5O1rYz08f2i1lj8WB3NewPWB6PUXOwkKcVkSG81b/APk61Ag5y5+UH/6WOr7OUFz+mJbn/wCCN/zxE+yPP+NM0izNs6dqySOghmjjlp2iMVUQ+uK5hjFrhfDdyv8AaNxiGy/irjF+GTUtm2cSzsaH9KO+RhDlZeRhUCLwfiaBbYq+keK5vbE+UO0grPpedtlpaNh87b5v3/VKnEYQDCuX/J5HpsuauPXkg/vwzN7NY5bk5u9+14dv/VisDPePq/LqZKbNcwpVFDm80dX+jY+ZVCCWMUrsjJZC6kmwA1C9gNrC/pfjKnr84rY83zaOrro8nlggehDwRpK8SzlAUsNOpgRe+5J3AIc3HdozIONbPqH0tN+5bmfCLquKz6Ktn/JfF0/S7Wt/1f8A/Sw4ns0iW2nNjba4FON//NiG/TnGVJx3HlDV9fU8rNIKRaR8vXl1NCYFMlW0qoLOHLdGCgqF074s/sSeQezTKKaojniqaVGgqI5o2R1kVzcEMAfr0xhxue9oMHhvODiwZ0xDW7OBM3aOnqvvZGCx50wpOPh1aeUVPvTNy08K6BtYYCrZI46QzBrXAFjsLWvbFprz/Qp7Gx5bfsxmvFgrxlsC0yPK0rEsY11W6eWPbeTrPMfmza5xlTVp0xYDeZ2A6LJi6bWRpCzb+UGjVXs/zKOBPBCsUgQ28Ol1uR9L4+WWJIsCfXH2FxnkKV2U1WW1E7mGrjaJ2XqCRa4+WPlDijJ63Ic4qMsrVtJEbKV6OvZh6EY+jAiVhqtO60aSigSskvEh0bXtbthp6Sn5ZHJXc73GDq5r19SR/a2A8rYGZgYASb3OPqVFjdAtyXyl7nB5unUo6b3NTy1OmQC1+m222GuVCZGIjHzthWvSyqoUat2OkXPlvjlxrkt17YYGBKE7kqPzyroctpIzURBhI1gFUYCk4kyMrThVkGkHrGTbEFxxVmfNFpxukC2/7x3P7sV/VfzBx5XHZxUpV3MpgQLbL1mBySlVoNfVJk3sequ8Ge5ZNUGONWBfoSLA4mGSEwIoQaG64zIEhlbe/bFpyziCMUCRVT/iIbXte488Py/ORVJbXgJWYZMaQDqEnqrfNDTvOgUK1owQQb/nhqGmi5TS8tdZOIZOIctWReXKqIsegAIQLD6Y7DxJl/L0tNYk/wBg/wAMdcYvDHd49oXE8wxQFmn2FSs1PBeJXUKuq52tg7MKKlXMHAjB9b4jqxo6gxgFXRl3HUHBjSNJVBvCBpuQMaGQXGAsrg6xnqmlpKXTJ+CpIwtaaiMkH4aAatyTbCNQCSkGxv1x6KwaFmVX0sGsyhhcdNjhhaOQU7x3JXPdKMa9MCgnuMNmmhi5cqxpt2ZQR9jhYJGu23XHnOqFQT3xRa0jZEC8Hdad7MKaOLhgMp0hp2J79/4WxbMystSCumzWIBBAH0xTfZxVU65DHSCRfeVcyOl9whJCn7qw+mLfVsrAOWZiRYd8fOs1/tlQeP4L6hkt8DSJ3j8SkSNYgI23kf4YFVadq2GappY5ZIJBJEzHdGAIBH3OHHIbSQQSfTvgPMaqCihlrJy6Qx6QWAJN2IUdPUgY566RuIUpnFdV1sQiirpIY1YMFUKLkEHc2v1HYjAIiLizsCCb9Om+H0gRiAhN7bjUNz3x2VBpKgWIF7DFue7YoW0x81eDqGD22Gw1b2GB5XJFlIuL2uNr488cge2qy/LrhuSPxliSGvYbXuMAYBlDBXVEr7M7se/mcVjN+JEyeN5c6ijoIXYLExqS8ki28RVQp6G17bbjfFW4948zih4oTJchiSpkQLG6GIuzyn9VQDv1At54gK7K+M+KeJMufiHIc0oYVkWMSDLpFVVLC5Ovbpc/IYc2mCY5rM+uWiy0qDizKXy6GrjqIpTVSCOlIlIEjdwSyDSb7WOLMsamnjkOhWkQMQCNtvpfFA4u4MyafM43zTOq55ohAtLBFQiOKVGlvIpMahUtqY39RiJWjo8l4wT3PiaSij5jxQQx6qzdYgAshBBF32AN7G3kcHUomnukUcWKhstX0KIwTc7b2w3u4uAbAefTD5mLRAk6rLa4FsDyyErdbC+w2wn5y3crLwa1mdri+KL7Y6arqeHV9zp3leKrVykalmsEcXAG53YYuTyOYwq7liBax88cq0juQ6A26npgy4BDpLhZVLKOHYqWPMhFKYEqZlmgiQEaAVXXuOhuCQLYIrKRaikakqLtG1tRvcn8sTM7IIhoTSQvW5t16YjKi7ykB91F7jphVWs98AnZaKOHZSBLRE7oDK8po8ryfNMvgqnX9INIxJS4TXGEsB3sRffbtbCanLKOq5oqS0gkMbiMudnUKPPoQo2/usdIGFi7hiOgwC8VREwlbl3IBHmR2wDXkwUDmAEjqp3mq1pRFFq5ZiUFA1lIAKi42Bt/i2IUAU4tGgLE3bTt9vPbCoZn6MLEE3AucK1QzxlpHN491I6H07YsuJCnNNe8IsmvQ223l+ffD0U6tbUWFzbpt/fhKRmaMBBGW1bsxAsL/ngfdDqQgC56bgjFNJCpwB3RaSxzgQs8gNr99reWCaFGKXV2UHqD3thiCWN5FcuzJtba257HB0TJ4lQ3Iv374YN1GWXGDB7upAB2Kjrh+ltLUxKSP63UQbC9sMO6liwZdKC2zdSOow/lqM+YQXSyqCR64jd1RlLz+oRa5WUEruhFuptgCJpKbLJZVpzI6Rt4Ct7m+22D87eN3aoYEBCbCx69L4j51M609Pq1f0mPWB+tpOu35XPywXzlPmqZVmQpGXDFV0k9r98dEhtqNjvhouQW29Nj+/Dd7BdN2+ZxqCQEYZQBZRf1OHo5L7BdsBxLuXLdNgvngsAhSTsbdDiaRuiT3OUhlChdugGEsLWKqSPMb2xwRgIXubk28O4Hzw2ruD4umq9+hGIBCEp6QrYbtuB26Y7DGzH9dlx0MGQWdbsTe4vcfLCAzareR2AxFERVromKutvAuwuLbD8++B9i5DDYeXU4W2uRr723Hi3wlUuunbwtuQNreeJaFN0lioCqF37knrj0jDmamDbg777Y7Iq2LJc22G198KqPEi2C6rXNz19cWFSQATESDewHUbE4cD3iLnRe9um+B6dUZiJLjfoT1w/NEvJjY23GwBvtiFXCaW+mx1b9PLCjpa1zYjY+RwgkFFsLafiI3wpFa9yBfsAN8ColhhGilbH6Y6PCCxBt18r44QbXYBfS3rjzAEBb3HT1xYKqAko41JYkgbkG9sEyEMN9J31W3OBN1kNxa4IsemH6EgO7HwAdbD9mCKkBdAOtfCSDdtsJTxnXYBl8zhWpnYhF1XNgRj0to9RUqu19m6+mAViySZC66Tfa9m07HD1MCJE1sUXyYWH02wJS2csZDa4tf+OH+aOUYdV97+YxChBT2acqUNHEoZgTcrYWHngMUgZQRIrXG++ww62iN7WVr9COmGZC9mu0dgSCNV+2LCIwmSiEtbcgdxa2OI620sfoNt8P0qHU9gApW+3c+mBpUYtq2sTbc22wxLSku8qqqk37g4WkQE6SG7AdVJ8P1wqNiifhqNRB6jbHKYhXBkPhPS2IqT6qVkI1nb7YdIGon03wzEQAz+I9rLh5XJRgt7gb3FtsVzUT8kYeG7kb9bncAdNsRdQQisrC2/U74kQt/Hq8XQk+VsNzxBzqLA+V8EhBJUcDpXTZmPTz++GWRksdLgjbocGvGSpYMtjsdrdMNSxix0KLeh64pRMNq5lr2HU9rDzw27kOV3ZbX1DBNgzhbCx6gj/F8NzxqGZG0gdNvL5YpFKFnkAk23B223GGppS1GXRQFN/CTqv1tv2vhUyusXLVAbXvpsevpgSbTHSSxqjGS1iVO7bbbnAKwFsvu1L5J9z/AAx7A/Nj/wA63/h49jPC6ELMnqWqZ9mIB2Jwl0Ou9zYG2DVpWgoUrjHpjvpR2G5bvbzxGVFRI7szFhqPxeeNh8VzoiyRLIVYnob269sVDMKurrS0gdhErPpVPCdhtcnuSNsS2Z1glY0ME4SWUlGfV8Ph1fexv8r4DeCSIxU0MkbvGviQn4vEPiNttv2YxYmpbSF0MLTES4KNqGlZKeqcpE0IEblbkG217lTv0vbY4KpYHqXtM7Mzag9xsd/Dt9/Xp8sOUUE9QYzrMaxAoVXlmx9LCwsPPEmqQUVIZGZwAfCnW57fMnGAdCtuq6SirTxWkAjj028XSwAH3sBiMra1zHI2kcoFSrhvCfELgkelz5bW+SauuknkEfhdlcMVBuvh3N7C1hte5v6Y7BTNNULO8zMRZbbm+knceQuTtvg2Uy4wEL6oaJKRQ5fJWm0tQVp3RQxuC8g6k9LAdR033xOBaSmQGkjVQAEZQBpK4HBso2Bu21vnhBBDgC7bb742ebtjxWE1jMhH5lTo0Qkiuqleg6FhZlJA9QPzw9rnraPnyS3EQIJItq6kn6knAOV5vR1Ez0CStMYxqvGCVjYEWDMBYX+eCpy0lPNFLdop1KPGNkF79PPbGf0TCa5wcJUJW5481S1Fw/lz5pWqPEY1/DTbqW6DDebcJZ/mNEBmGYUEUi7skcYCoO4uep+w2xOZVNPUUv6P4VyxWpoBollFoqWFu5d/1m72F2Nxiay3I6WndJK2obOa1bMDImmliYf2Ij8VuzPc97A40apWQsuo7hzh1cnyiF8vmgkSQWNU4szm120m123NhYeR8wJqjy2niMNRTxvz4/8A61MdUjb7nyB6jYCw2weIpJJDNWuZWPdt74VV1MVPGXYX0jew6YU9xmFopsAC8sMUMILA3G4PfA1ZUsygAWB7npgfMMyjp4WeaWyE6R+tc+QA3JwJTUM2aFFrlZIZV1RUzsBzRYG8hG6gH9Ude/lhcE7poEpunqWqXZoNTw3KrL2dgbWW/Uddwe1sQ/EtODBDlMUYMsswhf8AE2VbGRwTuTcLYnrYnFklmoqWd5UjjqHjGiN2TwqpXogt4Nyb9SbDftisZRVGfiGVeaxCJUMVYA9eUikeXST7nGmiIF1nqwLKpU2VZk+dVIrs8i5UQVp1jTULMSFXe25It3xc8ugH6PSPm601sfhKm1wBcbev2wbFBlORJPUQ0ytU1HidiSSxA2uSdv2YboamOoyyGpVFCSKXsQRtckYCsAEVBkFIhTxusUGo2IVmJN/z2w9yYJOXLVU1M5U+HVCPD98QVaZK/NqfJKaokp55WUI8bt8bHwoQD388XfKvZDm7Rr+nOJqfLJHIcRF3mkF731Bdl+/l0w+lgy7crLiMxaxxAbMKtVWX5RNOdeX0ZcWJ/CXbyvvjyQZUruY6bLyWFrclDe29rHFyl9lXDaGR34nrpWFhqGXanI+kl7fbCE9mHCcUmo8RZiAF3AyU2PqLPtjQMAfH3rKM2bOw9oVKajyhpHAoKCRm6jSD5Ed+mOVFBljs2uGlj1jSfABcH0vi1ycAcGw1P4fFNajEfGMje4+ofCzwLwY0t34trGJtucnYW+uvE+T+k+9X8qjoPaFTXghoaKb9H0dLJKLaIgAmokgEaiTtYXwVmGXUMgp5KtI2qILkElSEJAva2x6D7YsZ4J4Do51c8U1csiPcN+hXb6Hx2++OHhXgSbLmjHElVIQdg2UPsfq1sUMC0c/vQ/Kp+iPaFWGiykspeOhN/OJBf8sPMuSmFoXpcvKMbHwqL4kX4H4Lk024ikW2wMmSMTcepPTCf5iez55GRuKqiNm7/oFyAfP4sGMGIVOzMdB7Qoqsy7hSYjk0tGpQdVk0/sOGRkmSOpWFQG6qY52JHr1tfFgPs84JVEGX8YQSVHT+k5Q8SAAG9zc74q3EPDeWZJS/pClzWlzCOGdIpjSI8dpHvpF2sLeFjcXtbC34Vg2KYzGvd8y31KSp+HcsSYTRUskchIs3vEh379WxLVVNLS8F5jJPI0hFNOASoBtpb74FyCZqrKKaSVQXeMMb9R/gYP4pa3CNfFfxLQzNb/8AFtjnVW6Xrr0IdT1DmmJ6ZIoYqeJVCIi2B7bDpit+0PMqrJuHHq6Aqk4kVCzLewN72xaZ21RJIWB1KAD07YqXtXHN4LqmYWKtH17eIYsWdCB8ikSFmtVxtxDVUtDTNWhEo4BDC0caqwUAjc9b2J3xBzVtXUSmSWeWR73LM5JwPex2644T3F740uJNyuaIGylaLiPO6FbU2aVcYHQcwkD6HEonH/FQUf8AtFWA2F4EN/yxVrm2439cev5k/LAETuiDiNltXsn4zbOayXLM1EKVJUtCwFg47qRfqOvyvjTnokeHSmkN1uPPHyfQVlRRVkNXSymOaJgyMD0Ix9G8C8V0nEWSLOGCVMahaiIH4G8wPI9jhFSmIlb8LXHoFWGCFTGqu9ib23B3+uOxQxpKSJVYADfa5PfAlRMYqXnBGOprKQ2nf1w6lRejSaMSNuVcMR8Vr7YU1siVpdAMFSkCQrS2B7bna+PLBYJy1YuVvqsL4RR/iwXG1xcAHc4VLJNToWdZNCJcMi3PnYWN74jhCrdTfCHEb5Eake7Co5xW4MmjTa/ofPB8ntWmMzx0+RQyBe7Vukn6aMY7nHE9fU5jKkcaU6rbSshGojzbvvh3LazN8wpDURLSohv4kazkjbe4x5DMezWVZhiHYitTlx3MuGwjkVtYDputQb2zBZVilyBAxNjprL28ifB0xM5d7SGq6ZZv0VCNQ6LVFwProx82V/ILuZp+bKHGt0kv8x6n64k6riV+H+GedSMkSPNpDuNZDaRYBe4Nhc3Fhe24xmodickc6DR/1O+Kqq4sZqVzzj+U9JlebVWWz8FXkppWiYjMepBtt+HhiH+VQZ5Vhg4FkeRuijMRv/8A88fO3FkrZtxjmVVTxkmed5Qqg997AWv+WIQ3RiBdWHXfHWb5P8gsTQ/1O+K4bsdWuA66+sT/ACk8xBQD2c1RLjUtswG4/wDD9Rjx/lJZooJb2a1gABJ/p/Ydf+jx86cKZwDLHS1TD8NWEbHc722/LForJnNNKVR2HLKjwE7kHHew3ks7M4ilxG0z9p3xXAxXaLH4erw3AezxW3ZP/KJra7MKemm9n9XTxzMAZjXAhAf1j4BtuLnsN8X9faQGiEgynY7/AOUdv9nHyxldSDVxJpYE6V2BBvcdMbA7S+4TRwJGZWDaNRst97A+QxgzTyZdnsJpLaUz/E74rqZNnOIx2sVBcQtYg4urysxqMngjMSu5VKzX4R3+AbkW2wEfaHLoZhkh8PX+k/8A6OKflPFOYVdLURTZfHSnlrGwbUR4h27H/HlhvLoJKiKYF/AWVSTv3vfHJd2C7PGNNH/U74rrtrVx6SulNx+a53pGynlFvBf3i/U2/s+uGI392zZMsPMaOGIAtfe5Abr9fyxkXH/GUfA8FPMtI1U09UFC8zSdKKxJHXuy/fFcrP5QQqcx97OQSoSLafeR0/2cd3Kez2AycP8AM2adUTcnbbcnqkVMQXWctk4wQRGBVQ2LMRc3t0xmPtO4Oi4kyotCgGZQITTsD8XfSfQ/livZn7doa/lmTIHVl6sKgb/+XAx9tMOhQuRPqBv4qgWP/lx0g1wMqjVY5paVKS8L8QtWVEkeS1Tq7DS1gNXbpfEJxPHJw7MtJnSLRTNHzViZhqK3IBsPUHElTe3uaJNK5IRbyqf3acV3PvaNl/EOc5hmWcZVLJFPlIoVgEguW5wfVqt4bbnp2x6Sn2hxNhpC8pU7OYUS4OJK5DnGXTNHorIALX8UgBH3wVRympqdFLomLHSNEim5sbdD3sfth32D8E0Ofyz5nWZi9MsMvJhWOWMOHK6tVmB1ADtbzPbEj7ReNKnJJBw8E10s0KSpNHIUsG0m6gKAbMpHQX3HfGsZ7X06nALO/s5TAlpKzLjrJ83y3NXnzHLamlSoN4zLGVDWAB64ismVJMyhSRQyNe4I/wBE4vfHHHmX8UUFNl89A8cFHHNymL/FIVUIwHa1jsb7YpGQxtLm8EagEkNte36pxxdYq4sOHM/iuzpNPCFp5D8FOQU1BJTh1gQ6X0k27457pSVcdQtJHFfkkIbCwNwb/a+JmsWWXL6RYmTTGmkbgAb79OuHYqaaaRDEi6I6bSviFrD/AAceuOFa7ukCOdvuXlhjNPfDjM2v96qMGR1cjFC8YsPIm/5Y7JkdSD/WjTuPh6/ni1xU9UGOlA1hbZhbDUlFWiCwhF77HUMZjk2HjY+9PGcVy70h7kxw6larGjfVKEAMe1iQfPFspcnzio1mnyuolsv6q7dbdcL9n2TU+ZcTcjMiGpxT65Qkljsw2J9d8Fe1LMYOD3p48lWZ6apmc8mWU/Cvy+Ywb8xo4H9iTcJbcsr46a7AIm/rTFLwrxHUvFDHlEgaVrDW6qD9ScOScJcTrVinTJKh2HUqVK7G3W+I/KfbXW0dEtO2T08h5jyeCVowLgC217/COvrh8+3GrK75BSdevvDX/Zjk1O0dcE6GiF16XZzDFo1vMoqr4P4mp05jZYpDb2E6X+xOA5+H+Ioow0uTyWvYeNNyegG+5xxvbnmyRaIcmy6MHyJv+7Aje27PXmQvQ0YXUNVtROnvbfrhY7R4oi7QnHs3gvpn3K/8GZLUZfl/vtQkdNWMRA9PJADKYwWYEPfwrc9ut8WhW0QDUCCDc+mIvIc5o87oYswpZhLHMtwL7qe4PqMGR04QMztqLm1ix6fPHFxWLdiKhe8QSvQ4PBsw1IU2GQOqKi0uwlVgwB3JHX1xXfallkuZ8FV8NI0iVKMs6FNizKANJ+Y28uhxYlKoNSLt3FsPThGWKPklg7AGyXFz0v8A34yhwWtzCRC+fstRslzilqoOIazMplEM0fIXmRyF4kdoiCx6FmQ+qnG80MktRl9NUSw+7SSxK0kY/UYjcfTGHUXDecVGeVdRNQ0kNHHFGyxPaAOsioyqNIGt1Vgdz2PkRjY8rmU0FNFHNrCKFYnc3GxB9cU8ylYZpbKPtZNWok4S+7DbrvcY70BKk2PpsMNLIeZYjp3Jxeq0JjmrBcyp6uP22GONpKeoNeJIXQ2ZSVDKwI79LY2Woj4tqKYQ1fEGd1KgghZat3sfPc7Yqtbkol9v3CE+giKvqo4WOnbUtx1+RH2x9dQ8HUWgB416b7DHYwdanTadQXncfg31nyCvmFsjzWRfx6ytZz/akJOL3T8PSUfs0LHNa5Kf9HVEs0LSvywpDs42333NhvvjZm4Oy5FLPAtu5t0xlntP4hy7Jq7MeEpqpGZqdoY6ZKRyzrIg8KsfCSFkBJvt+WNJxFOoLWWNuDfSIi6qFIUNLGI2Vo9JsVBswvsRfsRv9ccnC8k6QGIPXrhNOEWPQiSR8sBCD38iPSxGPSPddNgD23xw+cr05FoTaggpcn4t772xyo8DMdRIPUd74W9k5ZQqAW3Y72wI8rSykqQVG5PXUMU5wKZSEJmodWmSOd9AN7bXBwG6FWZb3Ub7dTiUhkjmhW2nSO972wJJCfE19RB29MKhOCFlV3QqFVlYWO1sMTII4ox4AqDcad7/ADwdKLxWDKjX3wDPFIQQw1AjYaRhjdkl4vKBqEeQ80T6WvtYWucCFqmR7AFRfYKLgnucHTJIijULMb9BgeBpUlUBSXt2ubX9emLICCZCUgcr4pBqAH36dMP08DOTy2uoHwjC4KZ5LSKDzFPVhv8Abphox1EEshIcW3PrgSLK2sBK88EsbPGVdbWKMVt+3BdNDFTKzS2MhH61/D6Y7LJLUxxHT03Yb3PbCFVAVWRW+Lfc/ngIITRTaiKZSbMt3Bv1/dg/J2mOagIhZEjLO/8AZNxpH5N9sBrcQnQGa520gn8sH5GPDOJIysgaxJttt0+f8caLC8pEQu1waRXVyQ5c3/diPq0K5plqEbDmSgjY61XRv6WkP5YU0lRy0DNo06tW1/1sIklZs3iIe4SkcMOwLOmk/UBvtiMIJUOyPQOSxJ1k4Rcqq3dd/M4aJZiCSTbv1w27uGGxW/e1x/HGjZZ4UjGF1KQ4Y+h6YPjtJAzMWb9/zxEU7FWDDcegxJQShVuHtfrhg2U2unVQhLr0b88IdbOFG5Jt9cOrMpDaQAw6Ad8NsAQPxAD1OKIVAykSrfTtpPyx6Egtp6sT12H1wgu41W3N+tsOoiMLt/WdLnAkQiSpGMexB1eYttjoLGFrlQOthjjLcMNr+ovfCSPwwhto/wBEb4XrEwpECQuxkrEWubgnVcbWw7qQhCNLC1n2PXsdsIaNkuNWsbbgXGHeXZSdYCsLtYd+2HBA4pHLubiQXGwNvTHWsQJXLFQLC3nhKP4jqfxaPCdyPyw6HCRqWSwvYaunTrfEKsFDKpbSW2DdTewOHZLAJIsgVhtpJuRhpG6oqNYkW27jr88ExtFzHiZRe3W21v44qESaYtoDuNWu/bDXxEsAbA4eZ1MAU6VIPlgeG2tibDruOmIFS6CTvpN++FvMET4bMd174bTeUXB036+ePOA811B1Hp2xatLpmZJ9am5C9Dh6ZlMZUFfCR+ra+BoyoFzvbDtkVDuAwHXscUVa8G0kIvhJG+kdfnhVP4QyCMNY/ER3wkOxchWBLDTa1tsKTTHGy6yJOpUdMRAWr0hQRC6EyX8NztbA5fTTsNICg76cdnJMoAFyx6YS5YRlfXcWwSgCTSEMyLZwlzfywqUmKQqSukHc9/mMegfQCAAdut7YekZdILWNlvf6YhRoA1F2azk9bE98K1vGF0Lq9MJdYtnDAbCw8/vhN1Qk6vS2LCWWp9ZpALhxbuMehrCSNT2uegGGuV3BIv5C18JjAJGm3Wx9MQBAVIRtqQXY2B2FunzxySW8nUdLAYZv4NiFCm5OGEn/ABPFff0wSqUU2kREEXv2GGmkWNWKqD2BIwuZDpUqLC2GpX5cBZrFR9sENkv5yDnnaxKkKQdhbCC0tgzSgkncDpbCqMJUAkAWvfC3hMchAAKg9CMATKcmo3YuyEbdQcDVH4cbs5UqqFiR6DBT2jhLG1yLegxVuMK1o8rkoYCJKqtHKjW/9rbCHGFe5Wqfp2l/zDfn/DHsDfoCb/7Q/wDL/fj2M8lb5d0TGb5mmaZe0EKzSCJVQ2Xwqq3sRvYdd9hc+WKZntVPBks86kq2grHpUElu3T88WER1EtG+XRAIL82Vr/HYG32viocQVq01RFRRq3JBZmJa51sjW289Wm19t79jjTXq6WrNTpa3XTGVu3uJglWEuqk3C7uoGkkEEEEknz2ZfXHBLNIkoJs7AAS9AdwCB1P/AB8ujLQpSxwqpCVBKoW6K1yDt1uBY7dflgylgKhqkJYklUOo2A77EdMcs95dQCAi4zFHCsBksIx47m5IGI5apayVWnlVkDhzGP1F3sDsfQEdza22PZvJJriguoeNbnoCxbt02HT6kdMRMU6mtVo38csg5ZC2INu/W5uo9BpxUCbKExupPLoiZo5Sp5QpwxIutnIsR5E2G5AH54JC6IrJeONdgAeg+uORstPOlFHzKibSSRGAWI8zawHzNhthcXDgrJjPndRNIv6lJFNpiQAg+Nh8R9Om/fGymA0eKxVXFx8EgVYYimpqaauqDayQWbT6knZR8zh79DVdcXTNDHBTkaTBTudZB/tPtbp0X74lafMaWKoXJuH8u99rAoCU1Ilyo7E/2fO5wRFkM0sgl4ordYJv+h6KbZ/SaYfmqffDHPgbpESVH0FWkjDIOF8raueBSWSAhIILdWlkOw3vuxv28sSScNUr0zR53mT5nVnrBRM0NLHuDpJHjk6W6gfPE6BK1OlMkNPQUEZ/Do6ZQka736dz6m5w9EgUgIAi9z5fXCIko03BTqY44xDBBDEtkghQRog9FGw+fXC5DHCLbar7BcNVlZDHcIAFF7m354iKOXMc4aU5JAstMjFZa+VtFOm9iNXVyLHwoCflgoVaoUlmebmOjImlWGGAF7sB4R3JOIaN8xzZAaaP3SlclUqZk8UpK3HKjO7Dp4msvqbYk4qDJqashMrzZ1mKuJEnlASnprG4cRXtYWPick7WA3thnM68zPpqKjmKTdyt9BIawuOrDSL7mxvuoPSgyTbdNEi5TNLT0qySV2XsJptK2qqtiyqT/m1HW2ntYG5F9jhdVNR0sz1SyTTVM0ehZXPwoOoAGy3O58/oMMZxm5qljWjoqamjjp44AkCFQSqKpYncm+kHcm2IRqiYqTIrEn4VJFh64PhdVRqmbKTiEZVigXSx2JO59cRXDCleKM3R1C8uGMgaLbF5CbedzbCYMySKUQO9OkhNkGoXJ9Mcy8GHigvrZ5KmilLHsNLppFu3xHDQYAASXGSi83BklvKGK9vQYcywc3LYw+sMLoQ4tcgnf5YZrWDNebmWsF8J8j02w/lbtJSFivK8clgwtYXsD9sKqC0rTTN4R/sSpo6rj/iPOahFl/RdJz4FcXAlJRE+1ybemNQjYgPNKdUs8hkdj3Y9TiqfyfaeJYOPpmhj1iSlRWI8QVgSfuRi1OSUuRY32x6DAAFq8NnbnNcIP5n4JmSqk1FQbAHHo6iV72dhbrbDLi7DoLnHo0CKbtf0x1C1sLyzalQO3TNRSiolLO8vTrqsL4DqMriW5VdR6bnEsdXKsLXJ2wxJGzagzfngdA5rSK7gFCmgkaS5ATsNycSUVOGpGhKAAg6iqgb+mHxCzOTqvth2NGClQ173vgHUwAmsxJJUBT5WBY1DO+97E2wTURUMSjVCm2/TEsUXQQxFvUYq/ET6Fa8iroBJJBIt9N8KctdN5JQWcSUmnVFGi2Jvc9cULiOoX9B8QQGJSsc1BNGbWIZnkQ/MWxMVceaiR2Sll5ILBiQQAQAT8Xbe233xR8xzaM02cUUjHnVM9GxW1wBGZbgntuy4wYk2EhdvLy4E35K/8LOz5FQEA7wKCbkHpg2t1VeT1mXpCY2eneFCwtrJBFx59euK7w9n2Q02WwQS5vRRFFClTIARb64Nm4l4fB1HPKBiNwGmHXt8sefrCahK9fQe0UwJ5Ib2g5zPl2Ux5nljsUgrWR1KEiRVZlsT2BIH3xmPEPG2Z5zlj0FRDTKj6dRVTfY389sXqOrpuIUmpYqiKnpCrs0b1oVHARm+Hl7amsb3uN8ZhXZNWJmTUccOqcu4WNSWLWPbbDQyYKyVap2myiTe4IAItj22m+2CHpKqPeSnlXw6rFT0te/ytv8ALA+5NgMEQQs4M7LhO3njtxsAd8eGqxHljo3PXAkq14/F1viV4azyvyDMBW0TWuLPGb6XHkRiKCm+4x0dNib4HdWDGy0eX2sZxLEsT0NCFU6lCxkXPr4t8Mr7UM2iSONaOjdlHiZozcn6NihQugYczcHZja9h5/PEimWze9yD3WRoQCSQpFtgR5+Y++DaDsFbqpmSVfqf2wZ8kS/0OgbSDayObf8An/djtF7XeI5Z41SgpXkVhZUjexJIA/W9f2Yr+Y5BTrl0c0DzGwZ+TM1lY3A1R/2l2IO99rdcKyrIKCthaAtJFWNIxibmBYWAKgqTY9zpBv1PTa2KLTzCtr3HY+9Tje1XOK9v6Tw9l1SiMEcrFIWsTbTfXsTvbEfF7SMwjjFPT5XRhuhLCRmJ9btvviNy3KaGGvRMyFZSIwISSG5ZXVhuNuvmO2xwNmOWxic01HzqymDiUTRgFzfY32uPkeh874UabeiaKlQD0kmt4ozF69pmijjDHU0ag6SfvhGacS5hnNFFQzU8bLG94whbUG8wL+mH8+yWkiUHK5aiqlC3Kl1YqiqWa4Ci4AF79Nj6HETT5dmIqbilqEeN7N+GRoYb9LYttEC4Cp9d0Q5yHoquoy7MVqlAMqHo98MVUrTTNKQqs5LEDzJxPcQUlFJSLV0kszVAIWZCosAFG4sAP29MVw3BBtbDhUfo0nZZjTYH6ua7GTHIHRiCDcEHpiywcS1VRTGB4FZ2BUOpN722xWiClidrjy64eoamWlnE0YQsFIsyhhuLdD38j54dRxVahPDcRKTWwlGvBqNmFZIuKcwpZuYtGgMdiD4tvI4npPaZxWNoqenSMkkAREm1gevfY3+uAJEy2ojo6ylkkEM1zODEgZNjq7abXGx6DfzGH58tVTKlLPU1umECBngQvGqqCEcEXDDlAAd11Adxiq+Jq4gDiGYR4bC08NPCETujE9qnFEahCtOwF9SPFZR8rYlsp9smf0eVTwNleXMs97E6xpIANxj6p4D4XyI8E5GKrI8skm/R0HMdqNLs3LW5N1v1xNfzX4bHTh/Kh/8Akcf8MfIa3lRo0Kzqfm5MEj0hy+pdo4BxHpr4H9ovFmY8TTwJmMVLG1PqI5Grctpve56iw/PFRA73xtX8rvLYab2rrHl1EkSfo2FmWCIKNReTc2HU4xqelqY15k0MsY2F2Ww36Y+k5XjxmODp4oNgPAMetcyoAx5aSmjsel8eJuNu3nhJ/PHR33ONqpeY/LCrgXG/3wkgH0PnjwvvvviKlM8M8R5pw/VNUZZNGjHe0kYdQwvZwDtqFzY+uGeIM9zTPqxavNat6idU0BioUBblrAAADdifriMBF+18eBF72wYJiFF1b6dzvgvKKsUOYQ1ToXCBgQO91I/fgOx6npfC9PfUMRj3McHDcIHsFRpY7Yq2txRDNHBCaeSy+FQFVQtzft1698WHM85paDK4MxkmMsc7SUKQxoA0YRVbXbYWOpbb73PljMgCCGt06YKfmS5cHGwWWxCrYbj+7HWZnGKi7vcFy3ZPhSRb3qyLxbArMRSy2PYsMLj4qp5njgannjTV4nFiVHnbvipAD9XST64Up0tsBe21u2K+XcZ9L3BX8hYMGdPvK2fgbiDIcjzGXNTmBnE1K0BjakmtGTYk3CkEixxUvbNndDnWZUElDUGZIYnR/Ay6W1eRAPbFGilYSBAbFjuScSecZLmVNSU001NVsrxBwWgZQikm258+v1xjr4ipiXa6hkro0KTMNT4dOw9aiFuAdiMJZkG4BvhLAghSLHHD5HzxmgSmApbMO1z+eOdT1++EA+LtbywonYbi3bEsrV99jXFUWSZ+uXZnKY8sq3sz32hfoG+R2B+h7Y+kHpKZoxyaosCLjbp5Hpj4xuel7jG3+xr2n5ktHFw1XVSlol00skhuWUfqXPUgdPT5YEtYd06lVcO6tSqZKaOd4zIzgWvdbG+HqWppnUMk5DWN1JtbbttjGPbVxdxLBn9G1LmdTSK9JqKxNpBOtt/tbGe/zu4nc3bPK64/+KcTQwckTq9QGF9L1kNHKKkJVTwSTvrMiJGWOwG5077C3yw5AkXJjR6nTpWxkltdj5mwtfHy8/EnED7PnFd17TEfswwc9zttjm9cR5e8N/HFGmwqhiHhfWiJSIEEtWjb/qv/ABGHBTUMkoK1QS3XU4/hj5Ijz7PEbUM4rwR0/pLfxx7+cWfD4c5zAef9Jf8AjiNpsBVnEPIX1lJlFFNWUVWK2FZaCtirIHVwCrxte3yIuD6E4vkvGmb1aKI8xp6ZhIXOiQeIEEW3Y7b3Fu+PhAZ7nl9RzfMLnr/SX3/PClz/ADvqc4rx/wDlL/xwwQ3ZJc8vMuX21mma8S1cgNJxNVRm1tMFSPv1xVp8qzietnrquaGpqJNAkaSNHkfSSQxl+K/QWPQXA6m/yjHxDnqHWudZiBftUuP34fi4u4lhf8PPK8f60xY/niydVnFLa1rHami6+p2y+oDswRYw3dnsL4U+TuHju8hJNjpZdvrj5cTjri9LheIMwAPUc3Gvez32h1mcZZDBJSrVVtJEEnuBd+wc/bf1+eKIYtFN7ibrQ6nJHmUrBIFXzZgxB+mIebIqmjcf0vUhuAu1v2DCHz6tDn/2TZeuyD92A6jN5NyaOVT6Jf8AK2EudTWlgeiHpDHHZajR1sAV2+mGVhcR3/SDXG99v4YjqjMzpYTUtSTe5Aj3+98MPmMSttHUoPM4HSHbJocQbqRnSbwyCq73OpBviJlFaKiR/eNKnxkKoudrWvfCJc3pmsWFWwXrpA2w377QSAky1d773UH9+CDYSn1ATZOMKyaLUa6WDSCblRY+mHKWadIbCcntc9Thn3yn5ZiWplSMjYaBsPvgEVmXLMYJppoyN4303DDuB69D9cEyFXJTCZhUFljWS4/1RgmN6iSNkEjWJ3BVSDiElzHKrIWzEp21cs74Jos7yWO//thhbb4ME7TCoOMqRkSo0LaR0N7aSt7j88LiirDLqDLqPUsCCPlgQcScPowvm8THtt3+gwmo4iyHZmzJN+tgd8IcLrRTKs2VSS0qXWKMt2B3wFm2fVVLA9EMvpo1kfXrVLbnbVcd8QsXGGQ08m1eNA3N0ONx4e4D4ezzhzLs1M1aVraWOcWdbWdQ211uOuOJnWf4TJabKmKmHGBAlGKZqbLJqaCSWiRix1Xv579cNU9E8tdPUs6qRTpFGe5IYlvnfw/Kx88banswyBE0JU5iBfYc1dv/AC4Hg9k/DsUssi1ualpSC2qZbbCwA8OPPU/KRkoO7o/l/qo7CvIgLJZl5SqXRiCNiPP6YQunUGADahcE9SMbG3swyBgVarzFlO5BkS3/AKcei9lvDyKVFTmJHa8qbf8Alw8+UzJf4/s/1SfMaix5VYklEAAwuLqPFa3l3xsKezPIUYstXmIv25if7uOn2aZCX1e8V979pF/3cX+s3JOr/s/1QnA1fBZNGdwCCLbknvh+PS6EhbEjt23xo0nsk4bbNP0itbm8UrKquqVACPbpcae1ziQX2eZIltFRXqR3Ei7/APlxP1n5L/H9n+qjcDU8FlvLUAKDve523x5UbmalUk41VvZ/krLYz1t79eYt/wD049/yfZLa3vNdb/XX/dxX6zck6u+z/VF5jU6hZa0fi8KEEj5bYajupBIYeXkMav8A8n2TcwMamvNjexkUgn/Zwv8AmFlDAgz1pFrW1rt/5cCfKZkn8f2f6qDB1eqytmZQ+lQQpFybH8sdn1RFWWQso629e2NTfgHKTGy+9ZhZvJ1/3cZ7xLAuXZ7V5bTN+FDKBGXF2N0Vjc9+uO1kPbDLc5xBw+FLtQE3EWsPxSa9B9MAuUcI/wAULoQEgE3va21v24dlDyxiOOMeAEEkbfnhFQ2qQFySSSdxY/uGHBGZVvDy47qQ7M3Ujft8xj1hWfZMcslUCQJIpGpWHhBHlv1GPLTByVIuWJYaRfb+GCeTOrcoo91sAb9/T+GOSXbeMBVUbHSdsS6IIIx6VOvxC+kWG+2GW0NdU3W9/XB6I0jpzGUMFbcDdj64GqkCTuRF4D8Ntuw8sEAomo1BiBIAKnpe2OWGq1iSdwLYXJ4UOnSzdWPb5Wx3lM7rawZlvYdx/jtiGFLJAUIxuPCBht2JtfV028OFNe4YjSDcgdMckjJjDk/T0xWyqUgNqBa9m7WwQraPEyWI8+5wOFIJYagALA+ZwohzGAGb64kq0ppHuWVQO427fPDRLNYo127gthW4B8VwV3t2xymjGpnuCVBN/LF8kN5S4tCtup2F9umOyOCptGFHzx6NmeJpBq22OoY9SqocsLlfptiK3SkFAXurAkDbawwwwkDEhPD0Bv0+mDJpH0qQGJAA6Xt8sdYvMqFd1Iub7f8ADEkqboaU2Y6iSFW3TAhltYkFhbYX74JnUCVlIblsQRY7fI98CuxZWHLBZTstjb0xYKAiEt5dURGpVIF7d8BQyPzvF+r02OHlC2OuwJ6ix6+nbCYgsh2AAbtbp64s3QqShdpIdIFrYHnjmKFFv6dreuFq8gsFszd+1xhwVbHdha/a2LmyqBKTBCIYVUgEjvjsgVjuSotvbDIm5jkBtwen0w1POyJubj9bAyiuozOqowRSSl1AQ7Kx3P0xXuH4lzfNJM8mIEcStDToDt3Be/3F8DcW1VRmOYRZJSsedO2mRlPwRnqftfE/WrHS5Q9HSxxwxLA0QUdQLC5+1/vjLUdzTabSSIWne+5b/wBdj/8AD/8A0sexSOe/9tP9hv4Y9hGsrp8ZDzV7c0s5JJ2vboMVeeaOqqZ5pFYO8nKeR4tZEbHwFRcA3IHXcb9t8SGZVYpoJ6l3fSiMV0C5Y+VsVrk51HDHpempFRywcO7vq6C6/CSBf0wyuS6yx0nhpupilZEhV3RJNrWYBSpYgRHSNwd2vf8As3wfTRgpIk2hSwOlSbjETl2RVGkVMtfPO7kShSoVb9jax3HX6DEkIJYvidbI1yFPluD54yx0WppIuVH5vO9JArTa1VY9gbaiT1/Pb5YjshoFrm9+LVMOWgKqlDZyFUgqHIJFt79ziVqMvSaaIS/jx1UyKIi9tfiFlUd/Xa+LdTZFVtBGc6m/QdIFutFTi9SwI6b7JsbXO9wdiOr2MDSSUmo/VsouOajpjFQ5ZS+8zTgGGlpVMk0rbgknqR31H1wZUZJUukYznMBTzE+PL6A3bTfYPP0U+iqfniaowlPE8GTZeMsp3+N7lppP9eQ+I/Lp6YfjghpbMRqkG7WFzbBPqwEAp6t0zRQLS0ZosupKbKaIgBooB45f+0f4nO53Jw/BTRw/AAD1uf44GqMxRQ2kEDYen3xFVuej3iOgpopKuvm2goaZdUsh+XYep2wsHWZOyF4DbBWCapQWIYMB18vviITN6vOK/wDRfD9G+a1lyGWHaOEecj7gW+p9MPpw3IFer49zNKGkCav0RSTgMf8AtpRaw63VbfPErDxHUnJ1oMpoYuHsnC2g0RBGcX3Cxnc3F/E9t97HDwQ3ZLAJPeQGZZHSULrJnk02e1tgVyujBNPGbdHPWTe/XSvTbCMyqq2uihllQUrxpoiphYLELghQq7Dp2wNT5h7qGjh1IrqNbu5Z3IvuT5m/5AdABhC1Ek07VM0kjoTYG298AWFybLRskT5nKY46Lliki5ekpEdjsOvn0+t8RlnZCXOlwfCBg+VqaerXmNdVddR6WF++FZytG9Yf0ZKJIFUFiezfvGHU2EblLJCBVVju2ogFSW88BVSRzMtPPEJ0HQFjb57Yeaokd3jQED4bnbfDFTUMotAq6iLFji3bKmiSq9wRkNHWcSVlaVmjjpCwXWSUMjA7A2tsN+vcYnjHys6oKlfFzOZTubdLoWH5phzLpDTUJo0mZUJLOxNmdj32tv8AuwxXlk9ykEiKI6uPYdTqDR2v/wB/AgQFHAC6lK6JJGTXIy2UgEMbA4byQcvL12lk0tIXaS9yST022GHpyz0/xEaTcnSSLd+mE5UQaExgu2t3N2UjYkWtftvhVR1k+ncq/ewAA0XHxsQObRMLjzQ4sNUvLAA3uT1xD/yfozJS8dwRuqGaahiD6dWgsCoJF98aBUcHVM3TM41//E//AKWDxPajK8lhuPq6C6SLEzt0BXk8xy7E4100GTHiOp6kKkgsTp2Nj1GFAMFHTpi1fzEqwbjNYh8oD/vY7/Marsf/AGrF8+Sf97Cv1k9mY/tY+y//ALVwj2czMmeCfa34qqs1l7AjA0k7aviA27DFuk4DrH2/S8Qv/wDAP+9hoez2sB1fpeH/AMA/72GN8o/Zjnix9l//AGpL+zubn0aJ9o+KqvvJA3sLjCopdwTLYX8sWY+zusPXOIf/AAD/AL2Oj2e1wI/9sQ9enIP+9g/1j9lo/tY9jv8AtQDs7nIP7k+0fFQUciswRyBc9SMRmc5UlYrBJympSvhFifT5YuS8AVwO2bw3vcWhP8cELwPWjrmcRHf8E/xxmreUPswdsWPY7/tXRw2TZsw96ifaPisryqLI8rqKur4iehkSFBZDHZpD4iulrEMxINxsR1PYYx/iLJ6KbOswzvKsyqIopVWazINURd2AiNuuyne4uBjffaH7JqqfLa3NpM2p2WlgeZk5JBYKpNgb+mMW4uWOj4NqaYbyrUUjXB/VPP8ACfkQfvh2Bz/A5y1xwVQPa03MER7QF6HC4N1A/tRpdEwoul4JzCahWqHEV1ZBJyzEw2PoGwzXcFy09HUVM1bSzCGFpSOS97AXP61r40DhuSRuH6RSGJKDqN8HyUXOy+SAxqUmQqyk2uCCCPzwD6hDyAu9TwrHUwTvCz3g4ZfxPWvBR1f6OlpwJBeNWITWgOi99wttrjv54lM3yOjyVhnVURVCgczNHE/K50e+w2Nib2727Ww5wnwceH88o690paiKnqhLy3Zg80Ya+hiBaxG3XpiS9pvNHCWaPy+WXjFlBvYFht9Bh9Vpp6C1wMpFH9ox+tkQstzTiXh2cTx0fCppw6FELVmrTtYXGjtt0tioEd7Y7qNuxtjwLHv+eKdUc7dZWMaywXPCOgtjwtbYX+uPKRfckjE5kFNDPSvzFQnV1IB8sOwmGdianDaYS8TXFBmshQd17k3w5Hp1DmdPQ74tBy6mDBkSIbWtpGCKPLqdZwywwnbuoOOoMhrk7hc12c0gJ0lU4lG2CFSD1vi1cP8AE+V0GWRU+Z8PDM5YnJWY1rR+EgDSRpPl+zEmMuQdIaYt1tyl3/LCzQxpE3MpKPQw8X4QvbvbbbD25DXZcOCzOzui+AW+9XFq/Icq4SpOJWyWWoRYYI1o/fdKQrJNU30uAdW8XSw+LriSp+L+C6bh6DMv0PSwQyE6EFfKsgIYkqUUEtYkm9+98QfF2U0w9mk3u7vFSQUFNV07NYhlNTUaLkd2Vx8j9cE8TeyPI8o9jLcRrnEk2dxIsrxpOpha7gMgFrggHz6i2OPW4kwu3h2MLSQNkBD7SeHDmK24bEAWVnFRJVyMVG1guhVIv1N9Xlgmo4+4eoYlqKTIhVyain4VTKqW22II6bDv+sMY4wKdTYeV8S9H+kqLJ46tKScQipJ1vCwjYaRcFhsQfLCNbxzTiGuW2S5lwxW8PGvpOHnlqmjaRojVSpcKfGrbkXFiPt0BGKTlfEFBVVYnm4USCEpV1CTLWuFKrDITGdupsF1dBe9sVehqKozNHSRVZp5rvE3M8QawVyTcCzEi4P8Ao+mJfgt8sky2KnzTnq9RLXRs8aBtES0viJBINxrNhcDrftimPqaolXV4TmyRdTXDXEfB2bUtXFU8Kz08MEMkzQjMydYWN3OkmPrZGFj3ItjKc1ajfMZ5KCnkp6VpCYopJNbKt9gSALn6YunA/CsueQVM+WV45UUjxOJS0d00nSSEvYkX8P8ArC572rJ/YtWZrB+k6yuiyuianeRyKeRyJQwUBRpF0a+q43AB22xZc99kNOgAZCxckjY48vXYb388HZ5lddkuaT5ZmEJiqIGs6n8jgJLlunfrgLjdFur37Ms6myCWSfMcparySayytLCSiEsBrv0O9tsb7SZrw9mQLplaNKFR3kjnsrHT4GHhIOxO4Pn3GMB4YyTiitpBEMpq/ck5geR4tKR6L6tTMNrEEWPQ41D2YZPDBk7tR64YaGKokzKarZQEVNLBwRc6AGcbAbg3wyoysxocFMNVouqFjl9ScOcoZDQclBHEKdAiA30jSLDCc5zzKcoMS5lmFPTSTMFhjZ/xJSSAAqDxMbkDYHrilcTcVZjw5w3kQyv9EyrV0R5VRUVCiNpFVGRVBddQcagLdyt7C+GPagDPxJk0a0UMk6hJGkmp5GSICQHwywjmq91uBcLsCcflv5GNfGa6x7jy/Yie6TPqXpKtUNJ08lkP8qzP8py32jRU1Vks9VU+4xSioSu5VheQKoXQbWO53328sYZxHxDS5nQilp8unp25iuXlrDNsAdh4Rbc3vjUv5Z4t7WKfbrlMP/5yXGHHztcY+/dknkZJhgPoBcDE028Zx8Ug+I3OO7kX2+px49SRsPLDlPDJUVEcESanc2AvbHfEkwlmAJTai/rbHu1yN/LBv6OrAJAKc+A7kHHhltdYf0YkHpvh/m1b6J9iR5zR+kPag7Hcm33xwDvg8ZZWhtqckjrvhf6IzCwPu1tR28Qxfmtc/MPsVHFUB88e1R7WttjwIABF/pg05bVrtySLbGx74WMqrwNqbcnY6hiDD1/on2Kec0Bu4e1AtI5AW+w8zj1+m/8AfgwZXXtq/o526nUMK/RVetgYRuNvEMX5tiPoH2KvOsP9Me0IErfc2BthO3c74kBlWY2JMGw/0hhRyivvb3fci48Y/jivNcRPoH2Ked4f6Y9qjQLm+4PrizZPxBldNlhpcwyearmGySxVKINO3xK8T3PqCNu2IkZVXWJ92JA6nUMJbLqxTb3Y79LsME2hiWmzT7FRr4d9i4e1NZnUpWZhLOkRhjZyY49QOhb7LewvbztgS1z0wbPQ1UFMJ5YyEY2B1DrgUXAboD5YQ8Oa7vbp7C1w7uy4ot6Y4fLriSyKliqpWSWMMARvc7XB/hiYGT0QhuYt7/2jjfh8trYhmtuyw4jMqWHfocDKqq2Avf6Yeglmp5Y54HMcsbBkZTYqR0IOLMcmofBaLa241H+OEfomkAP4AJB28Rw85JiOoSRnNA7A/n60DxZn9RxDVU9VUjS8MCxEX2uCSSPne+Ia977YtAyemYsSgFopCBqIuQhI/MYrDKwuOh/bjHicBVw5GrmtmGxlPEglvJcubX6dsca2LjV5DQRT1SRwFo453VCzkHSGIF/XDDZPRqy3phZh/nG2xq+Rq5AMj8/UsfyzQBiD7viqqAD1NseHTpi0Jk9HzXUwhgOhLn+OONk9GafmJT2e+51t/HE+Rq8bj8/Ur+WKHQ/n61WNttjbHCSALHbFpiyijE2l6e6kXA1tjq5PQtOV93AS3QOf44oZPX6hV8s0OhVXH5E98ccjVi0fomjEcgaEk32Oo48+UUZRCsAvfe7nf88QZNX8FPlih0Kq9rb2GDsizSpyfNIa+icpJGd/Jh3B9MTf6JoBPvBZLbKXP7b4akymjKtpjsb/ANo7fnizk+I6hEM4oTsVd4PavRKqLJl9W9hvYr/HBie1rJQQhyqr023OtTb88UTLsjpaqp5UdPLUME1FUc3Av1xPJwI3LLvlLaLat6obD/avhTsoeN3NH1pzc7aLBrj9X9Vo3D3EuWZ3RCqpXYKTp0ufEp8iPtgurljKsobcHFE4cyBsjzJKlcvMQkPLJWoD7WJ6XJG4H3xcaeL3mON20o80EcwTowR1DD9tvnjn4iiaLtJIPqK7OExQxFPiQR600X5qlY5QpvY3Fgfrjs0caRySSFFVV1sTsABuTfCp6fStlXV267DA0lV7vMsLqq6wRoI+Id/nhQKMhN5fUU+aI7QqJEjNuYE8JHYqT1Hy8j5YcqYkplEzyR3VtgVG/wAsPNLA0Y0JENCafANNhe4H3JP1xC11W1VUs19KAfhp267+WGNIUIgwjjNzVDmCJV6A8sE3wl5LjSFh0g9dFhbEfTMWS5JAGx3wZAA7CKRtuqn+PphD3mbJrGtMJ8qqpqVKY9ypA39cdpponLDkKFXqbbflhthcATRKTvd+m2PFUi0HTcEE77EYoPPNPFNnRGR08Ml1idBYdL3x9N8DbcG5MBbahhG3T4Bj5kpWYpcAXHQnv6Y+i+DM7yqHhPKIajM6GKZaKJXjaoUFWCC4Ivj5l5TqdWthKAYCe8dhPJNoBrCQrSb4zL2dcTZfRzcbpnGeQh6biKtaOKepBdIURGsqk30izbDbri8PxFkCmzZ3lq996pB+/EJNR+zWqqpq6Wl4VmqJSxlmaOBncsCG1N1NwTe/W+PluX0zSp1aVek8h4GwuIM80x9yC0hZNwNxdmk9PxH+kZ85hfiHIqrNKUVkMsSQToZWMdOXADKIXhN0uPAT3xEcN8S8QZNkebyZbX1T1s/Bq19LAuaNmI5y6eZUMGN4Xs4IQC2x32tjfKmr4KqY4FqKnIZlpkZIA8kTCJWXQyrf4QV8JA6jbDOUS8A5NrbKn4cy8yizmmMMWseum18eoOeUjrIwR70Wi1tuXSOXL61n4Bt31nmSy5dlHG3BUPCXE9fmyZxTVBzSOXMnqhLEsBdahgzHlsJLC4tfURit+x6vzSjzfgbMczqs2oKLMMurHqKirzZ6mHM2SO4HLZiISoDPfa+kjG0ZZLwDlEstTlr8N0Ek/wDWyU5hjMnfci18J949n0lDT0XO4aakptQgh1QmOLUCrBV6C4ZgbdifPCznILHU3Yd5DhBJF/n+EW1ACZsORiL4N5Dh+YWVezbiWfPeMZoM4mz1cv4ypqiSFJ1ngjgeORjCkEhAFmpyGJjPVcDQNW5R7Ms8zikzfN4ZW4kfKqqtkr5pTS0IrAjOutiFZU21gX73xs4q+DniolFTkbR0JBpAHiIgIGkFP7NhttbbHUr+DqahqKdKzI4KWZmedBLEqOznxFhexLHqT1wD87JqaqeFcBLZbFoaT4c2wNuV5V8G0F3X3rJuIKmPIa/i7J+E8/rqrLf5mVddMDmL1PulStxG6yMxKMyljYH9UHEBUVmeZXkfFMclTnOUmTg1aymp6jNpKszyawGqY3LHlkXVSo38QONxy+LgWhoqjLaCPh2mpqpSs9PCIVSYEWIZRswIJG/ngvMqThWRUjzClydlWA06CdIyBEbXjF/1TYbdNsHSzvhw12Gc7aSR3jEXPKbRebc5QmjN9Sy+tkoc6424qpeLuJa/KIcpoqVsqSPM3pAkTQ6nqBpYcxtdxc3tpAxf/Y1mea5z7MMgzPOy719RSBpXddLSC5CuR5soDfXDucDgPMeRJm44cqzTj8Fqrkvyx/olugxJfzj4dhgDnO8sjiAFmNUgW3bv0xyM0xT8ZhhSZQcDLeVhAIMQPnG7vEJtNoa6S5S5xh/HmZ5fQ8ZZuKqeKJ+epAlIUt+Cnw9zjWl4l4edA6Z7ljKehFWhH7cYbx+uVy+0DNcyiEE888kYSpQLL+GIl8KncDxAfnj0vkww9almzy9hHcO4I5tSMc5paIPNOx1cMtKsiA6NNwANmv3AOPc6wV01LqF/7W3e47d8BR1AdFMIZmUWF8PK+kBOUYxYCyx26Y+/ArmyjI54zd31SLfw2+Xc+eENUJcLpdvBvc98R5qo3Z1QLrJ0lQOlvTth4vKumMIYlO+odb4iHe4RZnhmdF+FQdIYqQN/78N1DBJT8LEixJb0x6LwKNUTHa46G+HIy4XVEEsLgnQn8MWXKkIkysjqWjcXH64uD3w7FywoUMr60JXvt3P0w6OautuQqKdwbg6j3PTbqcIgV5yTqWJUvpUWB+WICpCbMY5fhcgWGog9e9v7sDNYpqXTsdrnrgyohjEPMjpGDgkS6VCluguT1OAG8Is0Ysu7ML/cemC9asGF5pRYIYgT1sN8Lidiygiy39N8IiKg6fExIsw9ccnUpIEJ0jqtj+2wwKuU5JfQNN732t3wqNGAY3Zu5UdgOuBWknRTzBcKNwMKWpY2carEADboPLFjZECltIoh5Woqb+fUY9GFRQdRJtjmpuXrdQSd+nTCUfwAsFU7kk98FMqJ2GdxbVsWbSoPn1wrntpKqzFtVgSOp9PTDHOVkOoKzA3G1++OJE7SarXZiSL72/hiko7p6EK6hZQNTHT1sf7sCzWgkMccmvbe29vnbDsqkRhUcmdm3se30wFWSJTpy0KvLa9vXEVEJTyAOEADWUlm1bD5/fHqcBY1W4JJPQdf8eWBSycsJJKFJIN9vE3lhcYcQqG3YMNRB6HzxJuqKIkLMxsN77drY67yRyW3W2+BTIy2CtcE7m+CWmWQ+EqVUHob4m6vdDTU6vWw1CSOJVYElT8a7+A/f74GzLMI6eB6uU2VVLFSwuCP1Tba4O2JFFBNza/XY9MZ77Sq4tUpldEyiSpe0gUdT0ucA86RKIDku8ERvVZhX5/Osh55MUIJ673JP5fniWzOZajxKp06WFr9B/fhyKKPL8rip4CpSGEKSAb6rdfqAb/PAWrmRMzodTLcLfGJ91qYwASrHoXyH+1j2E89fM/7Rx7ArUq5nDiTN6eBTqFN+JIoP6/6gNtr9TY+Q8xeRpKFpJDEAbovit0T6nv0wHkuUHlxRmTY+KpqP15WJuT5+Y37HFggZnrDleTUz1NUW3A3WJfF4pGOyi479e2IZebJFMBgk7pBWky6MS1M5vtYNclj3t59umJNcpzKoojVS0kOSUx8KTVhvNJuN0iAvYjV1sOlj1IOyvLaCglWXlJmGb20vM5LwxN30KdiLjYkfbEnJCZZ/eMxlepk1E2kPTBBrWKi59Sw2UPklDRUAWbKKSVq+15M0rW1TLcb8sdIx1G29u+DEpVjbW5aokO5djck4JrK6JVswDDayKtlHyGIqur4ooGnrKhKeBBvc6QPr3wRdIsrbSINzCJqauEPpBu1uij8r4gsyztKaVYGZzPJ/U00SGSaQ+QUXJxwJX50kUtEgyzLWY2rqiK7y9f6pDu3fc2GCcvqqXI2amyOjqa3MJWHMqmXmVT/AOtJayKP7IsAOvfC7FW4kGAk1XD+dPHBJm9fBw/BIdUtOpE1aVte1h4I/Lc3HlgnJazL6KmmouBMpeKJ1ZavN6lwdTDazykEsSSo0oCBfsLkDVNDBVi2e1LTte81FCdifJnvuPP8ulzyvqzyEpqZFpIUXRBBHYIgvsAOm2HNEbJJF0801NSgSNN+kapW1NPMnhDHclE7C+4vc374jq+rnqpHmlm1sLAsxG+GC7mIvMwsCSLN1wK9WqBo1Ift074INDbqOBKIkqNPicN4vMYcpJmYsWUWG63/AG4aRlUK0oe/XbcE44x5k2oMbd18sHqlLNk8xjaHmLb4bPc/Ee+ApiyanVAFU2222OHKxjHATGe/YX3tbHXKSRDUpVbWPU3xchEAEO00TJpj1XPdsIVu/cN98c0xXtstjsQLYHn10oWR4xyNdpXHYddY9PP6+WFklG1oCKRrltJPUnrh2tiaXK5hTqJZ4wssYG5ZlOoAfO1sMRSIyhoyHDf2Gv8As64kqIunjhdzKVOjcDSex88XIVFspbyAUz8lSzWIVQSLnyw1kMZiy2Fai5qJDre76vE3r32sPpbEelVJTUawcyYtFEOazoUCWG92+EH64Vmc9blUGVvR8mVdSxSJqIVrre6ta43Btt+4YTUEkBGyBda7/JsYpXcZxkAvJPQOFA2FmN/2424dOmMK/k1KhzPiaR/C00dFKyk30sZOnS3bt5jG6jHxzyuf2jDep33hHlYPf/PVUf2s/pl/5vUuT++lqjMzHOtNVvTFk93mIDSKCVXUFPTqAMU6TMeK8gzCjiz/ADXMJxR1mVpXTwLI8bKYKjm2CjxAsE1bbkAkDG1W88JNhfbHz/A50MPRFB1EOF787md4MWt/6XQfS1O1SsfyjMazOfanmtJLm+aRZfUTWoUSpqYfwXoY2uicvR8Rc3ZlIYHa4F4Z8y4zpOD8lXKm4jzPMyajNa4mcFtEDlEhYyEWRza6rdjpNhvjSE48pG4d4bzn3GYR58CYo9YvFankn3PfaIjbuRiFova9k9ZS0EkGV1xlqKKarqImsDSaIXlVH/0nWNitu1j0Ix6KjUxjiHU8HqayBBIIlgcDNhuRJ8W+KQQzYuuf6Jr2gz8TZ5nmTQcJ1VbAtZktTUxOtS1PHHLrg5TyeE6rB28BG+/liObM8/TioxSZnnTZ8ufxwJQLE4pHy7UoaTTp0W5ep9d7h/Dftiwxe1DKnrK6iNBVJU0slEiqxFphU8gXQ9Dy/eI9Q6i48xg3IuO4M2zulo48qqIqKtkmjo6xpovxTEWBYx6taqSjaWtvbtfCWnGYaiGvwgDWg3MSfnXJ3EHYXIBvYooa4yHLNKLMeI6nJeKoP01xBPV0uUz1sNXDNLHpnQtpR4nQGKQ/2EYqQvQbY2/hmkaiyKkp2qquqYR6jLVSGSRi2+7H52+QGDwB5YVjz+aZuMc0MbTDRM8ukdB8E6nS0c1E8ZLq4SzhfOhmH/kOPkr2r0nu/C9XUADlyVFIpsOrgz3v57W++PrjiwX4YzRT0NHKP/IcfKftomjXgUwqFWRs3iuN+gjkbf74+v8Akg/sGLP8TfuK89mxjG0/5SiOGG05DSAajdF3vtiYq6l1IRDYj0tiC4bu+QUQcKt4x074kHkXmawT16EX+2PolW1QwuxQjhhSESFw0r1KKEAa7C+I3ixIM/ytsjhdIjUgu0urayIzmw6k+Ebd8JrpZzEwUgq1uvU4aZaSSaBDCJJhGZkJBIXa1/z/ADwTIkEoaglpaOawTN8vqMrrnpKhbOoufIjzGAfD6/O+No4z4bjz3LHanRBXQjVF/pj+yT6/kcY3UQyQSvDKjLIjFXU9iOow0EQuXVpFhhNqdxYdPzxP8N8s0khKt8XUfTEAAQV2AGJ/hoOaV7ADxHvjq5N/avqXKzX+z/WEe+k2UEr4gCWvYDz23xIwRUiS6UrOZ52Rh+0YCl1oupyqgkDBVHzBIBpF+1+mPXUx315asZZYoqIxhjqbqbCzE7etwLYe5UJZUlkd7ncKe2EReJn0hGZWsdPY4Ig5jTrGIyp/WIGNYHdWBxMqd9pKUy+yXLWp4amNKmCi0Kza0VRJVbFioN7+e2+21sU7LK3iniCky/hCCrmqtcpWClL7MTYhT6LpuPK5xqPEnDmYcUez7J8pyakbmw0cMscay73BqWAYDYg6SLkgXYetq/8AyZYYqH2oVGa5ogj/AEVQSzHnADQ5sgFiOviI7dTj59VaS8Bu5Xvqbw1hc82C03IfZrwD7M6KKr415WZZlINVmcFU8IvYDcb3tiA9ofGmXZpn1G/A1BRS00sDrURSUcaGJtQu5PVdv1tsCe14ZrxxnL12XQuICwjUu1lPqB/ZA3OIT2a5ZJQ55U0DU/NPJE8+pm/EWxs6INjptezdeu1xjYadLDwCJK4dPE1seXBhhs25fXKby/gmLN8mJp66VJ62dnmj5erVIJCGVABcGwUg36qL7YpueLmORRwZtV646tp62BhIEJaUoY3YqLWsNN7jcm+Pq3gLImyzKsszFaVIKtoGV0WEAMpe6Wv8Nha57kk4+dv5S8M8PtQzCKrqNdKSksMOsnlFwS509ANanYdrYwYmm4EVGiAuxgajYNKoSSBZZjkmd5pkzSS5TX1dFJKul2glKXHrbrjbMj9pVTxT7NZMgerjgzajaNjTL+GMxtJfUXt4WJKXNxvc9DtU+JOGcny/Jq2T3+mp6VjS1UEM0688JIhuNCAm97bW6G5PQH2T8N8IfzWqsxzriGEhY0qIEoqsaxcHXEYyLlr6Ow3JucJomHLe5p0qgZvl9XTZzUwZnDUwTrOecJh+Jub3Pnfrfviep8jiy/hjM6+WrymVwqosUhLzaWYWaMAWBI/WbYA9L2xC11UmZcQRrCzJAVhpoi6bhEVY1LAE76VF7dTfGiwcKwUPDf6aEgj1ZeaWqWQGX+sFg/mLA7i3h097XDDpmCEDWPIkFE+yL2qZblrPkPE+TT5pRVAaKGRXvKA91KOCQHB1He9xjTM4raCqoOLfd6NqSH9AV2XQQIzylZfguzeRZmAuPhF+2/zxlGSwScQQ06okqSt4eaHUR9Tc6TcWt598bn7JKbNKSfM452nrGRmnMtKgld0RoX0MJOxUFTrPyJ02Mq06w0l2ymFNBxdpF1c+IKmig9mPCrS++LVz8Pe7xtDV00SyRvDFqjk5zAlWIS5TxC2xF97Px+nDqVOW1+dZs9FUU0KvBT0ZRKibxA6Vk+PSTtpUi/cnBVfw9S8ZcNZNVZZXVeTZbJlv4NLEukLHIiNHcAixQqpFvIjoTiK9rUkQzTLKaTNaumeJY5kp2nAhrSJ4wY+XcGZ7X8N+ltvFcfm9tWlWxVOix51B1TVG4BJ6gi/h7QvQuBAJjosK/lq02j2hZdWWNpMvEd+11dj/APNjAzbzO/bH0d/LVjBzqgl3vHGg6bWfX3+aDHziD2uN8fZOyBnJcN/KFycSP2pXD8FiOmD8g0nOqbrYE3/2TgA9TbwjyODchdVzWAkeEXvbr0OPUUjNRvrWSqO471K3ajeQDewx5WukenrbDYNy5AFsX32BZbQZt7TMty/M6OnrKSSKYvDOgdSQhINj5HHrsyx7cvwVXFvBIptLiBzi68bh8McRWbSBgkwqTGly9mPTvhxQ2iMlyd7bDH2mvs64F6jhPJxf/wC6r/DCj7POB9IX+amT2HQe6rt+WPkB8t+W/wC7P9rfivQnsjXP94PYV8WMt+de/hGG5F2iKk9cfax9nnA1j/7qZP4uv9FXf8sVjiTh3hDLeNeFchi4OyN4M3eqErtTDVHyoeYNPbc7G+H4fy0ZfiX6KeGfME7t2AJPPoEJ7JV234g9hXycQbyWNreeFkLqjuwF/wA8fSuc5JQVvF2c5DwjwFwpMcmghetlzCMrzpJVLpFGFG3hG7Huw2xccj4C4Pr8mo6uv4EyzL6uWJXlpXhRzC9t11DY2PcdcacR5X8Fh6Taj8O+8GNTZvcSJm4VDspWcYFQewr45GpTINR+WPMWDRnUdxj7UHs54GubcK5Vv1/o4xw+zjgVjduFMqNtv8nGMP678t/3ap7W/FF+iNf/ADB718Ubssq6gbH74SQx0G/bbH2dmns44Fhy6rki4UypXETEMKcXBt1x8YQXMaFj269Me37G9tsL2p4poUnM4cTMXmenqXKzPKamXFupwM9ELmtG8+TLPKWEEMzBtPbVbc+nb64qlTHynZbhgDsR0IxreQUy1PCOawAC8mpSdr2sL9bAYyyaDTHqspBXULNcje2/ljTmtQcYt036rs5XTPBD5seSP4WUe8O17WZRv6hv4Ynz/VMR2OI/gqgSqgrZ5GdeTLCAB31CTr/s4sooKfxCzffHfyak5+FBHivP5xVazFEHwUNBKsqCVSStyt7bEj9vXDhOzWN/TB1LlEaEJzJOXfUif2SetsPNl0Ad7tJt5NjpspVS3vbrmvxFIO7pUWCSsjA2008x2/7NsVHkGSAyEnbGi0+VwvFUk82wpKkjcbWhcg/K4xTYYwtG5KgXjJ3x5rPdTajQei9JkbmupOcOqs2ZEJU5goJsKmQA/wDeOGHIKU9jcHY4OzSiq2zCtU07Amdzbv8AEcBilqdKryWup79sd5jHaR6l51xZqN+a8lHIstTM0icoMFQDqbj+788NIF9xNjdg+98HLT1fuEh5Bs0g8XnthqkoKsssZpySzghTbcXwTqZbsOqjXh5gkbhNlGM0Q1CzL1vjiU8kJBmdW1lgoW+w7X++PtCH2dcCMkbNwlk5IFx/RV2/LDr+zzgVgAeFMoIB2/oq7flj4g/y15a1xBw77eLfivWjsjXiOIPYV8TmO8bjVvfYYbdSNKgnf1x9gjh32WySRx/zcycPMZgoaiA/qiRJfbaxB69bbYVDwv7LqiaGnThrKDJNSisRWorHl+e42Pp1w53lhwn+51Pd8Vp/QrGtuXe4r5Clp2gdFnaN5CgY6b2F8MkR6DZB188fXX6I9lVRCZ/5uUL6Visv6ObWwe4jIXTdgbGxHlh4cK+y8w1MjcLZagp6X3qZZKAqyRktuQRf9Rtuu2Dd5YMKwXwdT648PHxU/QjHN9J3uK+TMglmp6iaWkmankaKxZH07XHfG8+yfhnLOJOEo8zzesq6msMrpqaulXYGw2DAfli9w8O+y6Kq92i4cytSzpEzihOgO1iqFrWBNxsT3GJXLG4Iy3LYHoaOipKWUTPHopyq2jJ5h6bWsfnjDjfKtg8TSDPM6gPqC3YDsrjcLX4hMgjofzyVKrPZ1l8SxyqzgBiGVcyns3r8X5YFzHhfTPTwU4hKJTqigrfwLYKLm56Ad/XGjy5zw2I4y9PKWaQxLH7lIZNQUMfDpv0sb4mIaDLpljqFo08SDTqjsQDvax3HyxxanlKwVEA1MM8ez4rrVMsxNMXtPgsffg1GjErRQI3YIWW2/ngR+Bst1ap2lkYXNxJIR9y3ljbjleX7n3KD/YGKbWUFRNU1CQwryxI6pZrW8R7Y9F2Y7X4LP676NKk5paJvHWORWDFUatEAkzKxXi2hpctzCpp6CBo4owu5dn+JQTuTfviqBHLsRHbbY4vntHM1Jn9VBp/EQxlxa+4UW/K2KcIWiy1J7FneVgy22UD+N8epjvuATGGWNlNwPFAA7GxDdLbHEtRxrKitGBaQEgkjxeYGIqMNpIsGAubEbHfBkcqpSq7FSijSoIvbubfYflhLgJKa0bIhdMdoxewN9xtfHqm7IZCdwvwt2F8F5cYppDIQCttRuNlPbDNf+MroqqT0uT1OA0pwKFppFW4ZrEPb6du2JemnYiWwYaF2LWGrbqPTfEPTxTFnBVFKm5N73xOIsLQs6lLSxlQfMEdsHEBC4SZKDr9MsxXWAEUg+eKvmXElQmavluW5XJWGAeMjEjVztT1lVKsbSC0a6RYsATuftiO4TjlfiLM5RJrQMFLXuxuTuT+VvTBUz1SKjb2QuaZ/m9LlE1VHkbR6FuWlvZdxfpb9uO8K8bRZlAsM1JI9Xe8kcSlgR5i5uB9cSnHLI/CmZAXuICLdgRjJMrh05fJWQySxTLqAkRypAtuNsacLTdXMN8fcsWMrDDGT6vap3OuKs5oc/rYRIREtQ2iCWIXVL3A8xsRjQMgzjLc0oRNBKbMPErA3Q23B7E/LGMzstZWQuZJqiSTSJWcb3GxF7m4AA3xp/s/y+kpDmYIlEaxa4YVk1B5NQCg36bE77dMWMM+pTc4bDdKGJYys2m65crdCJQpNOQGIsdR2H0x5ZYZozT1iXDKY2Qm4e/kcKjZhEoA22vZd8C5hSloRJEl2J3IH2xjaegXRLNNirBkshizWh0wiyyAsNgLLubjv0xZ+Ic1gr5WbmBm7i4tfGbZfmEsEgSUMQG3uNwMMZjltTUtqy7MqeME6glRdfzwTbmSFC3uQrhK8DOF5kANxa8tv34VJCHjYNpkRrqyXuCD1GM3fIuKkZDA1LUG/WOYdb+pGOrR8cRtG8mWVLqA2lVlVrn/aw3ijaEvhlT/DscvDdVFkWZWdKh2NHUadj08DeRFxa/XFoemUWtFGpttcdcZbmf8AOiojnp6rKq/QhBJ0G+odNJsbkHyw3lvE/FrSJl0kEyV0TWcTp4ulzsbbkdvl54tjhsl6C1ajMZU0iNlBt0C7dMGUVZOCiyuGAGyntvc79cZpT8Y8RJIVbLHnTf8A6BiOvW4w5Wcc5lT5hVUUWXovIqDCSAzHUNm3+d7YheN5V6SbQtGq2hLrJGZdZPiUfq/XBdOpYgPNHfyZ/LGO13F2ZVgMZSRCCERkeTUdiOzD+/HosyzaoWREp6x3lPhEaMevcm1sFxAq0kbBbFV5nT0UsTTTqpudLQuXsbdOlxiTo86L0oqEqrMzeIhQSPIWOMLTMKxZEjiknLIQQNFwT0t09b/TDddnVa10Jrbrb4YHF/yxZeEWiFuFVnEdVLynqgt9KtK7qm19rlcQ3Ev4+XSigzJTVwsssQjkGpyhvpJJA0m1j+3GQHMszMB1GsCFgbNA4Fxvvthtc3EScxpI0klB1M9rut+t/K98TiBUWLWeEM6OaUCx1zCLM4AgqKdSVCkqCRYk9Lne9jbEvUpRyECOUgE7jY7eQHfGB1WbVDV0OZU9SvNp1IYLIU5kexCm3XcD7DFhpuLqjk8w1iuCVIBI289/vta+CFQQqDOS0+ZYWQqYw+5sSRcC1uh74FkAjfRBfSVBNz0PS25+V/8AFq3lXEokkGvlkP69fv8ATFloBFWKGUITa+g/u+wwQIKhZCaJlCWOqKTcWItb54SkjhwR4gdiLb388SMdOXHhU2Atue+BauMLHsXBAFgBvb5fxwWyGY2Xmj0IruCWPYd8JkaTSVsAR0HmMJDtILMWKr4dW9x9LbYdKAkEEE9r9cT1KwU1DGxjLMOnQAWAOHdbkfgq5cg6ivQD5466SP4CJALC+lb7X+2+H5kljcxrojbVZACPGPXyxAVRHNR87mnhZdSmV9yV6/TDUdPHKoaS+lSL2AO/1wmqltqkeVS3p08jgjLk5lOrqCVViBc3LYtAZQOZJTK0pbURew138u3rYY8jE6kBJvdhfe1sFVyI84lnFz1A6+n8cCBzqZbNoDbE2uBiGyohMz3OoA6V09RuAPP9uEGTQytH0OxFtreVv34dqEUC6ad+qkXJx6klEPNn0gst7AkbnywJRAXS62oanpzLILeG1tQ2xnmSxJnHE9Vm8w1U9M2hLnbV/cP24l+Oc3mGV1BBId/BGnqf8dMKyChXKsmhpGj0yKgkmN7eNt9/P+7C3uBTWs1EJzMWVUWliOlYzuC17Hy/IYbcqsBL7bHa/T5Y463kYkahbr5nzx5lvEVK6+wAOMWqStWmLKR1yeT/AGx7D36Mrf8AMn/Zx7BLVpUtlGX1FZK0NOopqRG0TVLEELb4lQfrN+V+uJ6GCClpFy7LoGhplHiJa7zN3Z26n69Og2GDqpo5Ejp6OlWlpYF0RRLsqjDEYSNfhLn1/bhmxssYbq3S4ylJHpAsSva2o/LyxGZhmGhGLaY0RSxJOw9ScRvEeeLQhVCy1FRMSIYolLPIw7D+PQYhZaFGBzLimqglEbajRBx7vD2Gr/ONf6eWKFzCPV9EI/8AS1TmcGrJqOSssCPeXGiBO19X6xv2W/0x6hyuCOtgnzSpbNq9HBjCITFGxBA0oNgLi+pvqbYJppM2z8Rijhky/L1+GeSELIwudooiLC/9pu29jiz5PkdNlFAK5cpWON9w8jMXllufE7dT37fIYKnTc60IC5szKrtXlmY5tpnzCeXLcv21wQPeeVugR3HhjXY7Lf5+ViyrJp4IYoIqOOGBWuIENgAL3LEHf5k9zg5Uqan3mYLM0NQ+uKJodLKASQNxYG2kb9z2w8tWtPTxw1EZCagZDIpLgEta7Xtaxt6Y1sogXSar4hRuY8NpUUKSJRvBXKXVhGyhVIY2vsLkk779hikTxSwzPBUqUmjYhr2FiO18WnM+IYRQLBRxo8DHxFSVYPqYObX7tY3vvc4rmZSc8iaQmSYrZrn4vK5v1xVSAgYCboKp0sh0HWLb72sfTAywBQp0kbbE4Ukg0gE7D/R6YcU6jb088KhMTYI1N0Hzx1XYMSbb7bDDvLUrrsLA9x3wiRSj3JH26YAShiSm6mVzHpXp3FsNs7mAJqtp6b4fRUIJOrdb3ttgV3EcYGlyT+sF2GCRNbC4ouQum5J73wtqGGojkgqFJjZbGM9D8/7t8cjnAOol17A23OHamqeOkmkB3RSQLWuTiuSMBRCW4doJ4njmnh03pgWJZ3IsI7Dpc2P1OIin4wzqCUxSZbSRuttaOkoKg+fl88SmXcNiup1r85qJppXBKRpJ4Vub3uDf1vjv6BgqIzlNVy1rAxZHc2NQh2Vz8h18iPK2I0wElzTNrKTzWspswejy2kC1M086PNpGtEiG5JtcHof7tsE8ZO8lCzJURRySVcIjaU+A/ipsbbgW3232wbkWTHJ8ujinamesaNVkkga4IFyB0FgCT0v169g3xMkppIEjhjZJKyFZPF4kBkXxC/fCHmXgytTGHhkrRv5PoeHibiCmnZVaOlpmfcWGmYA7+Wxxtgr6ErcVlP8A+KuMO9hM6VGacXu0sKsmXFAiEtqGrZtRG2+2k774PFvdQmkXO+wtji9ouwjO076dR1Y09EizZmYPULg4rOn5a46WapHWOq2P3+h/65T/APijHvf6E7e+U9/+0GMOlni1aAiajsNsKUJGPFGLk7eHHBHkRoE2xh+wP+5cx3bqq3+4H2v6K9ZbwJw9Q0OX0Jz3Mammy5r0cc9YjLCOTJDpXwjbRK31A8sFycHcLSZTl2XLOUioKVqVHSVA8iGm93u5t4mEffzA+WM3Ogktyhbz0jDJMZvpQD6Y6n6pMS46jmLvsDnP8Xifasru3mm3m4+1/RaLNwDwjOicypkZ48wp8wjk94XUksMUUagED4SsKagdid/KzuVcE8MZfxBT5xFWzSGklmlo6aSoVoaV5dWsoLat9b7EkDUbAYzQaB4dABP+jhaLEQBoFz1NsW7yS4pzC12YuIIj0OREfS6W/wDaFvb0k/2cfa/ot496pT/9Yh/2xj3vdN/1iL/bGMKUR3+ADb+yMFJyrIAq2tv4ccp3kOpf74fsf+S1N7fPP9wPtf0WucSVEEmQZiiTRszUsgADC5Ok4+VvbxGq8E0b8qzPm4ZiOp/AYD8sajTMokGlegN7DpjMfb4OVwdRo/MaVc38W3hH4L2t9seu7OdjG9lsNVptq8TWQdoiPrKbh85Oa4kVCzTAI3n4LnDSf+71IxGkaBpH7cHlgAR1J6emAeH5QcggkZLAREsLDb7YKWX4JFRmDdNu2NNT94V7Cn6ASiCIz1u46emEQTN78+mkXlRw6WmvvckeEDv3P0whHJYqVNwdsN0XvDxyTVEgCc5kiiG2kAixPe5269sE1W4oyOyxkKLEt1t64o3tV4RecNnuXx63UA1MaLuR/bsO/n98XyANoYuQBvYdcSNGyimJdFdtP+Bia0DqOoL5ct4/TE9w0gNNITcHVa+JP2ocOPk2ZvWU6r7lVSF4yu2hjclPQdbenyxH8LljSyqFB8W9z32x2MmOrEj1FedzcFmHPrCkFQIVPMcsDcHywfTGVpCTVPe3djgeCGaomVBRvOFIZljY3IB9MTNNHUSVl1yOZAwNoxqIHpcjf+7Hqm12MfDjC8nWD3NkCU0qFtzM7Mep89sKKRLLeSWRQBuVJwY9NNEhP6OqYrsWcuCVHoNthtgcNIfDFAha9wSe/qcbm1WPadJWAhzXXEK8ZFxXXcN8FUGc0lNJXyR5fR0zQPqJCrU1a6iQO1kFt7XA9cdNHHxDxmvEU2XJlkdTAkcgpHkEVTcDS7hkBHiUKRYb2PkDS+K5qxfZjBSyB2qY5KeCcQyrIilZKt0N1JG4b0N998M+xjPM2gzmhyZ6mo/RD1yPPHpvHcgrYm22rwj++1vAGoadQAi4XvxQFek4TZwX0DkOR1NcsEdFEq0vNKyTKuoWUkOg3tbt373ubYu2ScLZRl1VFVcp6iSHand0swFuhI+I7979cA5lnkGR5plOVplsklNUxSOkyMFKojxp1N9X9ctulu/9rFpreIKYZV7jT5dmdVXctClN7uFMd9iS5IXsbm9t9jjQ+s0kFy59LAPpNhmy6wkWsR5ysVLEokmkZvgQH+PXyx8ne2Csy32i5rUVGU1FJLnIzSpELq1uZSCEyIu2xA0GzebEY3yrqM94ho4oc4o56LJlkfRTSkAl7LYzb+IKS1gNgRfxWXHz17POHEOS5xVJWha/K3qqaGgDI/vUrQsuobXsFc38tJO2m2E1H8Qwt1ChwW6juojingieokfMafMaVqIxxBq2US8uNliAfUQvQMpA0qwsRv1sI3s4rky9aiOuywAv4qnVMIwgRmZySuw2AvptuN+toLRmGRu2Xy5olI+mzxFpPD5jwqQPlgGWuqYKGSnp85LQHblKz3PiB7gWFwDjA1jgZXRLmkJ2iySuzWNauGupFsxjs7OTcH0U7Y0Tg/h3iIZ9lhrOJaSelp4+S9PC0wZ4vEShBiAb4j1PQmxxUODkonydTNNIk3OIOnp1GNV4DSmm43yum94dTLULGGG9749JSy2lUw/FO8SvLVc2xDMXwW7TGyt3DmVSUNFBktFU0UsTSyRioYlNaNG7uqjoHU3I1FtgevTFj9nK0OX8TvSrPysup4JFWWWcnWWKnf8AV28XnfV2xDyJHFWO6opkR7m479L4ZllaKhqDAyxz6G0EpqAbsSo6i/bHmKmOqPDZ2C9rSy6nTc7TuVu1BTUVFQwUeWqq0UEax06q2oCMCygE7na2HHRTYsoJBuNumKtw7xLltPw9QJX1ccdUlNGJ1jV2VXCjUAbbi98PnjfhkGxzOx/7GT/dx+Xcbk2ZOxVRzaD7uPzT19S7MACFhP8AK2y+bMMwqlpo5JpIaOmlKqOih5b28+v5Y+XSpDeJbY+wvabmGX5rxi1dQTNNEaOOEkIV3DOT1AP6wx88+1Dhb9G1RzShitSTG8ibfhuetvQ/l0x9/wCygqUcpoUqjS0huxsVyMXSkl4VCI3sR9sG5GjfpWnUAk3Jt9DgP6i4xL8JIJOJ6BHTmq0mkg7Aix8sepon9o31rl1bU3E9FPqtw9trdDjSf5No/wD5tZXqIY8qff8A/FtipDL6P3QOYiPFvZjjRv5P1DSRe1OikiQgpTylSSTa62x6LtbQeMhxh/8A1P8A+Urx2UYpjswpAfSC+pFvjuODoMdPTH4XIK+uL1/PFU454RqOIszyfNKHPqnJq7KWmaCeGCOW/NQI1xICOl+3fGcZhxzxcON82octzSernpOIIqGDKFynXHJTERF3adV8BUO5uW/VHni/ezviKrr+Hc7zLO6pCtDm1fBr0BRHDDKyre3kq9cemdk2OytjcWxwm21z3h0IgiDHMbhIFVlTulBVfs8zE1s2ZUHGmZ0GZV1JHS5rUxU0B99CAhZNJW0bgMRdbbW22xc+Hsrp8kyWjymkeZ4KSFYkaVy7sAOrMepPUnGLZJ7T88rOCeMq6bMqR8xp8q/TWVmJI25EEgfTEwFwWjKrfVv4xfEpkfHXEed8Q5BwvLWxZXnDU9dT5sqQq+meOJGhnQMN0YNrHY3I7Y6eMyPN6rHMrvbpZv4Q3VNhJAHdHQ2AAS2VaQMtG62THjjOPY5XcUZtPntTnvEf6Qhy/NarK0g9yjiuYnUCXUoBuRfw9N8aOMeWzDBOwVY0HOBIjafxAWhj9YlCZyP/AGTV/wDYP/6Tj4GXaKI9iOl74++s1F8tqQf8y37Dj4dFDSKjjS3gUW8ROPvHkJY5wxsf/r/614jtlUDDSnx/BF5PCJOB86F2GnUdumyjqenbFOy/LjV1ApoYYmLppHNJsDrvtbpfp9catktPR0XAc0ySVEU1VV8gGME21aAG6G1rnfpvjTaH+TvkFLMsqcQZuSu9jy7E3uP1fO23pj1/a7tDgskxjRjHEapiATt6vWteS0X18G1zF838JUM1HFnVK4s8NVArixA25ouPTE4qlWYel8a3xH7M+Fsk4prpK/MeJKhamKKqrZ4lhMNOpcojvfxkAg9A1hcnEVFwrwLL75JJm/EFHBRyolRPKtO0cKu8kYdypOldcZWzWYXBtbcbsl8oOUswgIDyPBh57f8ApcXNez+MxGJL2RHrWcE3WPfdTjp3d9x0640xuDeBg1LSTV/EtFPUSckGaGK6PyIpgrgfCSJkUX/XOnyw9kfAHBdfXU1IlfxJKlWyRrVEQLGkjQrMIyPiJCut2ClQWAvfHSf5ScoY3UWvAiZ0GFzv0Vxx+j7Vm+V071Ly0sZ1PNR1UcYva7GCQKPuRjNK6UiB0MZRlQg4+mp+E+A8gzGuqBm3EIkySqgimdER1czNo1KLeJVOoNbcFTscQXEfs49n+aZvV1TVnE0UKlzVz08dPyYVWd4DKR8RUshN1B23Nsedzftvl+Nc2pRD4iPQI3gj3H3r0mTZTXwlJzKpG87/AFKrcQIRxJXhblRUyb26+I4jIVbXIWGpe4v1x9HVnsUyGqr5qx83zQNK7Oyq0drk3/s4yr2vcHUHBeZ01Fl9RVTpUUxmZ5ypIOoiwsBj0WQeUXJM5xbMFhHuNQg2LSNhJufALyuYdnsbhKbq9QDT6+pVO1o2VmNSOZHKpceQN7fsOOUjMMwRr7B1t9xhqZmSGJSx03BtfqccDFKjWLgHpYdMe+IJC4TIBkL7chb8JfK2Fki3XHxknEOePEV/TWZ3v/1p/wCOFz59nymIJnGZG3nVyfxx+aqnkJxbnl3nbb/wn4r6L+m1GP3R9oX0zm3BK1tRmUqZgYfe5FeICO/IuCJQN99YZvlfD8/ChkzNswjzFo5ueGjUINCxCPl6LfItvfqemPl1+IeIOawXOsyPh/63J/HDI4g4gMRH6YzMEf8A3uT+ON/6nc2Ajz5u0ehyt4+H5lbR5RnxGg9Nx4eHh+ZX00eBWSjjp4MwjFo6YSCWEuskkRPi+IEAg/CDt2wRV8Fw1iVRqpoGkly5aOIpEQsTDmeMDUf7Y2J7dcfL03EufAD/ANtZmuw/+tyfxwkcR56S988zQbbf0uT+OCPkgzef7e2d/QM/f4I/1kVDfQfaPDw8F9T/AM1ZhqplzFBQS1cVXNHyfGXTQbBr7KSgPS/XfDdNwhPE0cRzKI01MlStKnuwLKZjc69RIYC9rWF8fOPCNfnucSVVM/EOZxlYNYc1UhsQyjpqB79sSpFckWl+JeITOf6zlVrIlwLGynV97745GO8mWYYSxxrSemjwid94supgO1+JxgJbTIb1t6um8Wlbm3Bc7RRj3+mvHUNKsTQMYVBjCaVXXcf2utrnFypEaKmiicqWRApKiwJA7Dewx8oh82iiDLxNxCCb211hbt16DErNmPEFMqzJn+ZGdlia4kYkKYYyL3NrlixNh3xx8V5OsfiQBVxLTH8P9V0a2bVMQ0a7wvp0ttiuHMaeN6kPeyO+qy/6RxgJ4h4ta3/vFXautyQP2HCafiDimn1XzuR9d7iaFWv+ePQ9j+yL8gxL61Srr1CLCOYPVcrFVDXaABCO9pVSlTxHmUgLEaxf9UsQtt/yxVIKyqSjEAkm93YljFe63Nrn8sTFTPV1dSZcweOSaQEsyppB+m+Iurk0SPELLy2tfy7Y98XHU4tVU2AMAcmvdlDKdzcHxX6DDNQhCWRidIsFNtsO0s780HWikNcB7Wv9dvvh+eaFkRQUlBW7Mv6pPY+uECZKe2ItySaSslonJicSo7AFXUeLbyx58wQSKAgC9GYdF9eu+PSxRSUjSBy2g9GNjf5XwEqKJjIXjAGxBAJt+/BgFQOEqXWZTMIk1FGPgdhbUL3DW672waqsltErKFUApYWO/na+IRKmWU00RmDBFJCCP4b+ZGJWF2N4nGrTYAE7/fvinGLI45qJlQivlYu5EjR+G3S2378e4cplXNc5nXUn4igrYADqe3oRhyr1e+jUehBsOx2w9k0YDZpVMyhJ6tioBHwqqp29QcC02hAW3lBcdxhOEczKkHVTk9fXGWcOqHyuQeOxkN7fTGncasH4UzYl72hJGnuPLGZcN6f0a+qOQ3kNivyGOzkn776ivP8AaEnhW6hWGgg5cUIRpLIzabtawIsfoRi6cLs0+XGKny5ZaxqoRM0RsxQi41C4UBbE3AFz1xS6EXjQcuTqbb4t3AdOBFPVQry5pJeW7DrpHT9px3M4Y1uDkDmF5/Ii+pjtJPIqzcoXKeK97EXFvywFl+vTPGdir7egxJmMIZNzqv33viHpi9NmVRE9yXF1NtvPHjmOIdK94TI0ofMEdyyldBbxax0OBoqgoBFLaVdQAYLuD5Ykcwd5KcAGwuVJ04AmFoVQAXD3B6YaZQAWTsVRJTsvIcOzHYFe33vh2WvrHsWFrLtZj0+ROAC7qNTRjmMwZWY2tvj0QMigbNvuD1wtzJura6LKQp66TUVcM1yDZdzb1GPEZbLmcNa0xjqopEMnMh1rJGpPgNyLHxH/AAAQLCxVxqBDXsL7i2FTBpJtKm4BuSO+2FtcWI3NFRTr53pflrJrjF10DYbfXEblb02VTTNQzyJBVVPvEkTxCSzkAGzNvvp/bgIQPHJq1X8OxsDa/rjy06qgXnNsNiBa5xQqE2ROpjcKfgz6aORi+gqTbUIADb5jDkHFRhmJWSXw30lnKgDyxXGVQoKt4yACSp/bjk8Mp8Ekjkm42Nx98NElBCsx4jgYrI8jFg29jsPzwNUcUtJLcVFQVv1MnQYrEcNuquSTa9r4bKBWJLuN/IYmooS2VbI87mqAQMymtf4TJ2w/zoJIfxJIanax5qq4I+oxUkiSztIBLqtaxsfywqSGlaFuT7xCyi1i4IOC40WhVoU9U0NFUO9svy5rjSb067i3S4G2Kzl/CaLm80EVBzqSdvwRbwRydSCey7Drc77A2N+kzAMYak3IuVIsfvh+mzesjYNI7KyjYgj679u2I2qDuELmdEDUcPV1NIz0KrOqtdk5tiCSbkYfy/Ma7LZkEyTQk+ICRCpA9PPEkK3nneJdwCSrbH6YN96pcwgWCuUyad1BNmX5YMGTZEdrqXyniGGrhAey2trcMBsO9sECqpqoXpapZCRvqG/yxQM6aTI0pxRlpJqqYxwBhYIACWZgOvyxZMgzGnlCM1OgmVbdLX+WHB0mCs7mRsplYUge9OqlbAsDsNvTD0bqxL8u9+wGwOFKVb8QXK23W97DDEltyBbbYjfDFTU/C8KSvNKzA38IU7H/ABthrNQzpHMGueaFNvrhtHYut5NAUWK3wRNKGQazaNviVTpub9tvLFQVZFlHVRVToRFBY22UXva9/wAsOhbQBdZUncEtYD546VSaNngU7E+Bh4x6b2++BtMkrAGPSmm6nqV8136j1xEuJXV/qwWfSBYIBYkjzOBqhSHGtSwva4v088PRuAPHBpeRdJ8XW2GpItSIjsQ2olnUbH74hlWWwhZlDf1mkom4NrbfPzwNNMI4WUqNNj07Hy/PBLKtyZBpFiNxt88RWfTLTRs62souzMbADobn93fAElCq3J/7X4qp6dgGpqJubJY7Ej4Qfrbrifr3OkRgjUTqffqx/gAMRXDJMOWyZo28lXJzrEX8I2jX6nex7Xw+sjyStI92FgBt1PfGKo+CttJkCUuwj+OxsPPB/DlFUZhmJYCyr8Plfz+nXALXdxGouzGwti88PU/uOVST2OsjSnqe59f7sLpiXI3HSFNfzbof+sv/AOL/AH49hnk1n/XH+xx7DrLoSpCcLTxMtwW6W7DFfz7OIstiJKSPO91WOMankNibAfIXJ6AYKz3MYaeKWoeZViiUsXJFgO5PyGIjh+n96dM2zJXSorUtTqRqEEB3A0/2jsWPyHbEMkwuewwCoWmizhq60UKJmNaQrVkt9MKFdQSEW8Sgb3Frm5Jxa+E+Gk93auqqZqiuiO8lYyhU3ILouygnc3NzYD1xYqLLK2AQyxlJKlFLPIviMi6vFYm36rGw6nbyw9ltQryTRxnUtWdYZje+4AvsN7E9PnjSzDzd26SKk2CZqoKSnEcqzxzmVdUdQH6bdSo6KfDt69cSsr5fRGlgrI6gwSIFijdfDduqgnruTte4/LHs7yqkyukqZqzkSQwQKgHMLLYtcjrcjpsQLb7m+KtmXHCyQzTZfBpqeaHiZgCF2sT53/4402aLIN3KZbM8ryA1Xvk3NlWZi8Ou5N22sDfa3l9sVT2gcRPWy1VHTNLoLqxYsLEbkC1uwI3xXqqomqqqSeZk1yHU1hsD12wwqBg9ja+9+5wp1SdlTmG0pnKwwkA8QuR9MTpoTyGVh16k9L+eIyn8FgxHnbEhS1UiNpZ7I21iOmMtWRcJ1IDmofSbsLC9zvhyJW7atzbfD+ZxFKtgxt0II8jhqnfmVEkCKQYbFmI2ub7fPbpgmukKEXRVOjDfawPY9ceLBJCVBs4vceE7f4/LCn5YI0EWUeL0++Gpwz2McYII/rCQE2PTzOL1dFNK9Gq3LuDfQx0sSx+ZY/stiLljnkSN2hVgCbBzpB+g64NZdFy0jGQrYt5egH+DhJef3aUKG1lSFK9b+l++IT1UgJsQg6Ty2Av0ZQN/4YNnpr0M6BQGaI6COxw3DHOtHEtRG3NF7hW1gXO2/wArYeqHWKknklKhVQi7OBv5XOAe+NkbBe4Q+SwsmS0SuWLCMA72uRiR5FIamOoqFL1KIUita2kgg+tzt9sAcJa04agmqUdDOzSrqa50k7G/kRvh6Yw+86gWubW32FsQEBsjdURNiic0rjSZdU1awqfdYWlMYGldlJt+WKzkvFlDxRI1GUeknG6RB92A31KQOoI6HFjrY+bRuhZJFIsQw2YEWIxVeGuFjlmec/3OJqRJHkjmIRpFBXa197g3sMBh6IrOIcY8SgxeJdQALW6htAWr+wzKFXiTiFKDnqxyzU+iR9Ml3UtcC4JsxI264l6Zo4mME80oZQwsIzc+ex8vPbFB4T4+m4Lzaesjy2oroKqIwxxsCrsl7lmG2/h6dsSWTe0XIP5yRVX6AzZIAsi6JEEi6m63bmAqRYW+Lqdu47WErii2CV5zM8J504FojrKqnE65xFn9YoFZDMszmkhDPrmAVSmlRt1O573IG4301FkalhZtmMalr9bkb4quccQz1tflVRWtIFpKFYI3Sge58WsuSLkOSVve4uCbC+J/Jaumq8pfNocwkmpoqn3V0eIoyuVDAm4G1jjVg6ruK57nSDsFx83y8DDtZTZEbnquzOy2XqDb54bLXDbYIkCy2KSBix6E2OErSTKoJA9N8d9tVhC8NUwtVriITauQwXfp1x1XINjtY9scljOsguBbvfpjmg38Uq7D+0MEajeqVwavIJ6JxcA9/PCy1kYAXOGIztpDIT3uww8oYEa2S/8ArjAOe1OZRq9EblSc2aNDcXIJ/vxnX8oLSnC0EYB1DOL79LclsapkEbQQNmNRII43ICAm9wNydsZz/KDngq+COZBpZo8zhOsjoxjlBsfIhR9scfHvD2mF7Ps/QdSI1DdQ2TK0XCUYRQzLAx6bHe+JIXWJCQPEALWwzw/Hr4egRhYPCdWCF0+5xWB1afqe2PKP/eFfSGWYEyody90tqOxAvhOX09qGIe8JUStcyyLY7hulx5Wt9MKVBLG8bNywYmVmJtpB74dyxUNFTCGJY15aDTqF+nXbbBBQmUQ6BkTTew3JA7k4Ogvyr6gy2NjbAOuO+tW2AIAU3tY4KhmDgMCfEvTFQFYPVAZ3lMGbZNU0FYmuJ1Nze5UjcEHsRjHcioUpqurpKmYRiGZlJK2ZgOht5H9+N4y1NUZckE9xq3xV+JOCUq6+qqY0jBqjDqLMFClUmFrnfSzNFfb+GNOExQw9UPWHNMEcVRIG6m/5M1LDB7T5BDLzEbLZSdQsQdcdvn3x9PhV/sjGVezT2cUXCfF0WZUMtZMHoZY5ecVYLd0KkGymxAP6p6dcW3j3jnKeD0hWviqZqidGaGKIKofTa/jcqoO42vfyBx8T8qYqZj2iazCguJY3/qTcipOw2CDau4JRftFH/uDn5Ub/AKNqLb2/6NsfI9VkmaUjQlomIdA4Me9x+3vj6140WSt9necLPThXnyuYSQh9Vi0RuuodetrjGdR5VklNXUcdFmlbVpHHHyVnsSkZJVwJLLYN4dth4e9rY9P5Hce7BYXFN56m+4Fc/tBlnn9Rkch96w7NGOScDe+cyhr1r6yFpKUDm8lhFPy9ag/F4tQ7dL9cTXBnD/BdLw9WZvnUTrSU9THC782VZJFEKsSunZmMuoAC217+sTxdw1DmntNyPJEqZ6VM1qijl5RIya3vqAsPDY9D0II6AY0IfyaqMTPCOMKxZFmWPU9GoBumq48XTtfH0zEYgYmqakyl4XDOwtMU4hd9jPFtTxz7R6HK83y9JqHL8pqRTGpLSy7vB/WO2ztZRc2F8fQeaRxRUVRVKiiWOnZVPSwFyB8r4xz2I+yY8F8Zpno4hOYLNlbWhNNosHde+s7jR+eNahzbKc5jrMvosxglnUSwyxKw5kZVijXXrYNtfp088fCO3+PxtLORwajgwNbMEgbneF38HTHB7wE3Xy83tV9p70Hu0uYcHcpwCySV9OreY6zX8sSfsEhZuJc1q89OW1SZw1SxWmkWZRMtPK7gNGzBQVci172v54sWQZMeBcsyvIJs2jzQVSzvTu9AqFOXp1LffUDfYG9t9+mJ2nrZUDiNqWMy3uY6SJGXUADYhQegx9nbi2ABYWYKpUvKxOi4KjoOOcwhyWrgrqKanWal93PPMV33ja24YWPzBHrifn4WzWSYwvTSRJORAHekYIpY6Qbn542Xg8Qfp6hIlJkGqwBuD4Tiz+0lzHwlUOLgiWK1jY7yKP3483j+3j8qzGjl4w4dxI7xMES4jaOSCvkHGcHioQvkel4ZzHIpnypcqzCpiVwxd6NlIYgXU7WNvMG24xbOGqPN6PiWiqWyas5KTbOIGAU6TY3ttvbH1ZT2MKE9dIxH5/JnSRxHJIqGR7tzPeSdhpOm1iO9v8bjz9Dyz4gN83GFaOUl5+Cy1OylN1fjGod5iFkE8NVqeVqWRbjpoP8ADA0RqFGl6WYubn+pP78acZ/aDdT7nw9YHxDXISRcdPW2r6keuPRVHtAvd8vyLSAP13DHcX2uQNtVt+tsc53bqs7ZlP8A/p/RemadPI+xUanSrNOre7OFYf2DiKrKaRmN6KYhSSS8Ww/LyxtfD8mbyUkhzqClinEpCCAmxSwsTcmxvfv2w5nZVMnrGOwWByT/AN04536xKgq8PgA3iQ6R9yYRqC+dKneqvG3gGzm19xbDGa0NPW07U80ETxyqysWFhY+X0wfTqgp4ypVtw2sg3KW/4YmOFo4J+LcvE9PDLEJG1KyAj4Gtj62wzCzOaACvlrjfh9sgzqSlEgmp2JaGRSDqXyNu46Y5wIL8Y5Zfcc3cEeSnH0N/Ki4YtwzPX5blNOYoqqM2p4AHjvrJY6RuLG2/kMYHwdT1FHx7QUtbBLTzRysskToVZfAdiDuMdGj+8b61wMSIpv8AUVdF2gUm1gemLT7OeI4OF+LIc4qKaWdI4XTlxsAbsLd8VbwtAbA79LYWSRJc36Wtj6PjcJSx2FfhqwljwWn1EQV8uoVn4eqKtOzmmQt+T28ZSFB/QVd1t/Wphf8Ay65URtkVd8uamPn4GyEWub364c1qGVrEC2PnX6oOy3Oi77bviu6e12afTHsC2fJ/axw1k1VmdVR8O5ksuaVZq6nVUK15CqpceQsi7YrEnEnArT5gP0TxUkOYmoapp0zhxA5nDCT8PVp31nttjPtetNrnfHCUDq177WO+N9LyZ5DSeXsa8ExP7R/LbnyQfpVmREFw+yFpeacYezqvXT/M2tpA9DNl0gpJEhDwShQwbT1+EEE9DiRPtD4J/nbQ8TDhnMRmmX0hpIpxMo1REWswv4iLmx9cZENh0As3T0wprmVj6b2wR8mmRubpIfEEfvH7Hfnz5qfpVmM7j7IWz8M+1fhbh+KvGW5BmirmFfLXT650a8shBci52G3TE0vt0yYsy/oTMRbe+pP44+fVYLCe3i22w/zFErknquMlXyTdmqr9b6biTz1u+KI9rs0aIDh7AtzrPblk8tFLH+hswUyKUuXTa4+eMCe41KLbKMLO8AA7HCXBLPZbeHfHouznY/K+znE+T2FuuJkk7TG/rK52YZvicxLfOCDpmLR0U1SZq1HkuX0TMWjrpJLqd9LK8IUjy+Lf6eWPsRdwMfDudT6KDhmZv6uCqndgOoAeEk277Y+kab268BTziGKozB5G6KKNrnp/EY+J+V3K8bj8bR82pufGqYExML6T2bqsZgmhxhWHjqk4PjzWmr+Ja2SmeWMRhDUSJDMkT6/xFXwlVLXOrbffbAFXlXs9ipM8pJquKOOGFaHM1Wc6gJXMqK1ty5aQ6ep8Vh5YZ4ureFc9zijObNVvHRpLE9MFZVl5oS4Yqw1LZbFdwb74z2fPOB6TiHMU/nNnS5pGzV1QZMuLIJIpzUI4Xa6IGkGm+6t1FhjwuByTHcBjKhrNI3AaYBm0W6D2kdCF2nuAOoAELSK2n9nueVNRUz1cVU+Zkq6a28bTBYLBeoa9MBbqpjJ2sThuny/2dZfNDnUVddctrBSIi1Dugq4ohFbli+qUIoBsOi3PS4zyHjfgmbiGgzh+J8zkzVAzQ1LZSUicI87TWXYadErra9xpU3Jvc3IqfhmKkEuV8TZuJInjcSVNLrHO5DxSP4SpIeOUfCwIKAhuoxr+Q8U1mgvrC0RDo9Xo7RF4+q1wFQOMgD3K60eVezUZnSVdLU03vdZOjRyrOb1ckj+9JqPRyTGSL9tQHW2BaDJvZeKnLqGjrLJUorQUyVUqwypJK06I4B02LMxVG6jYAjFeNLwcYIEqs8zOoFOYt5ISZmdIJIhIHvcMDJrB7FRjmTJw5QT5YkPEea8mj90eoRKYxmolp41jViVawVlRdSEMNtrXOLdlGKDDpqV55d13j4fh9fQhM+iFuItjGfbzk8WY8SZdI9QYmWjYAaNQ+O/mPPF0f2j8Np8UlV6WhJvjO/abxRT5xnlHUZakhgigaJzKmg6iwIt5jGfsJlmaZbndLEvpOaAHXItdpHNTHUKWJomm8SCqjJwjTzUiaaqI6epEO9/XxYFl4OCP4cx8RJ25Pb74tuSz5dGlO+ZVIimmB1IVPLiINrMw336g9Nt8RuYyxxTzciVqheZ4Cii1idt7m+P0IM6xrRq1fcuCez+XuIBp/f8AFQ0HCcY8LzqxAJ1aACPzwUnCNA0Zd5LbDz6+fXBSCaSMsdQN7WvuO9zhGucR3EjhuwLdr/LCznOLeZLk0ZFgmWFOfWgf5pUBfxzu21iA1v2HCZuDspERUSzx7G5Et/24Pcz8nVrDehN7YGjaW5VmkbvY9MCc3xQ+eVQyLB/5YUXU8GU/LtFXSrcbFwH/AGWwPBwZA0mmozpoU661pNZH01jE9CXViHRwW6m9h+zAFcF5pCq5Nu7XwwZ5iwLO+5Aez2CJks95TdBlS5HmErUVaalZI+XraLl3GoHpc91H2w6plMrO06qSbnUepxHmM2ZjGOnn0wnmrHDq5St2vq6/S+MWIx9Su/XU3W/D4CnhqfDpbKWjkja4ZlItYAdziSefmFVMyLsBdmv26YrNNKY0D8oI25BG9/zwoVLjxOkYZTfYWv8ALGfiLQKUKxFqNTeapiuRudsMtUUoYfjxlR0O3T64g5Z4Z1DWIAFip63v16W6euGYHV6mKE2KFu7WH1Pbf54rVdVw29FNrUx1EjmOWIBbhVuOmI3MjKqmQmIgMxAtdeoNjbBU+WoHj5E8M8XLDtNIwAt8huvlvhNNDSz81VrIYnRSSDIoVgPI28saWNeNillzSLhLigocwypwojjqoUv8IHiJud+42AtiJdBA2kLqK76LC1z39cS+Uwuv4qmjk5jEBOcqyWHcarX+hwiWnqvel1SLFoIa7BkNidj0sR64aWyBKBpg2UdUWdFkMelkJDqNh23HYjDWYzUx5Pu8WgBFEiXDXbubDtidnSiERepzORlUlQIQzbnc72wJX02TwCnqKWpac3u6tYFQOxA63wp1O3dIRNdDu9KjzNHUVqyQqE0oFvfZrdPtfEnlEcn4rlNOnzN7nEZWpTRNGlM0zWJDAqLfQ3ue2JLLNbUrBwgQnc3sfrjMXd+60GdNkBWs36XnaS19EVhfYbHy+mC6Ck0U4jV73difK5JY/twLVBGqFksoJZWD+gPXEjJJUwRQSRUzSIZBrKi2gN+sR33+174JkEmVUuDVHcaQaeEM1Nv+gPTfGV8OELlrqxcNzT06DYY2fN6jk5LWTXBaKJnAKagbKTuO4xiFHm0UKSqYtDPKz2isEUHsLm+OplVZlKpqceRXEz6i+rT0sEmytGXBpLXLmzXG2Lj7OZ4+ZmFMxbVHy3VNYsdRcXta/wCqO5GM8jzyOmip3khciZS6EODtqK7i/W6nFs4Wpat8yynN6WOJqWrLqtQCCRpUkxt/ZIve3exx3swxNDEYYsDxNl5vK6GJw2LFRzDCuWU1JIqadKiSflzkK8ltSqWPhv1IFutu9uoICs4pBKS8S3ljFwSL3wdFCzyCVo1VQx1aR1PzwVOQsfM2W+wNseUptDhde2cdJVSqqmR1WF4iNO9gNhhpGkkkVWURqu+oL3xbDRLLfUUu3mceny6EjTzEO21zhujmlcayrEoBYFiPCLdNr/PHfdbKJlcln8Xy+mJuXK1YKupbXvt0w0mUS67K6jfob4EscrbVEbKJeWRgbBtiO18e8Ur8xlIPodh/DEu2UyAkkxtY3wHPTMsjAghewHTCnUpTWVh0QThwwPUDzBIwoyXXUui/QHv9sGVFLHoVo3XUN2UG4PywkwB2XUDbv2wvTpTC+VHc+QSX0kb9dPXCnmjsVaM2BN7N0+mDHp47sge4UeWGKqictp16j3HW2LDiqlR5nRl8JBN9ri2GHQoyE7k79DgyajdPgcDbp6YaeGdWBK6lJ62xACqLrWSojcF0RkX53OH1gd4mIvYDa69ccpm5YPRD6HBVI7mYpZHJFxfocWWoZKj5VULpZUYkHy2w00fhsmok7en2vicKSmVgwCgb6SMCVFIGBY3XT3tf9mBghFIKimW9roFbfcHCoyRZWJDKdjbDk942Vke62+f5YZd/EWUGx69LHF03HVdWRayJSdHkjaqQO8QYKT1swsR9sOTU4hC1lBJqhG/h6ofLADKUj5oClSf7W6n9+C6GUJIzR2Mj7PC3wyD09fTGkEJcK1ZBmYkIa43Xxi/n+zEzWQKyLJGng77bA2+eKJBMgqgsCShCAUlIAVr9h8hb74teUVSSKYqhuWp+IBr2t6YY145oC2F6SVBqJj1AAAEG2+O0xWddbSIUHwgm378Jq4LSK6u1k7g7dvLYnYb2x2J44zZrjUTfcC/z264ZKBdRuVIb7a2uTcWwuRY5FNrhvTyx6WoMkRkZNK9WublQPl2w3qYNoLKCo6qb3viwQqiEHUNZ1j5dtHwnT0wlYjocMUu5JOhvXb649m1Gc2oWhjkMVUh1wSgmyOAQG/PceWERZtqy/nVUIp6uEhJo3Fyr+YHkRuD5YolBJTNa3IoBI4CksFFxufX0xReI53zbMocjpWI5rcypZT/VoDfCeMs4zfM+Iosvy3SNQFhYWU2JJ+2JLh7LDl0Eq1bCSonYtPIotZQL2H2+5xmqVEymwuIS5eXGiwxG8PVQBYKAAEW3oLn5k4aBZGUC2kDfDlOGa008Y5rb6R0F+2EurNKsageM3AtjE8zdbgbR0RvD1I9XmSadrm2q3w+uLjJVc6uWmiXTBTrawPU9BhjJ6ePLcr55I5snhAt2wjLA0kr9Ea56HYDGmgy0lZazjMBTV4f7A+2PYkLR/wCbm+wx7BQF0dTlR8zppM6z2LKka1PEBU1Z1WUqL6E9dRFz6KPMY0fIcpijpppa4NHKWXZ1G62Xdd9gOhv3vir8PU8+VZdAlREWzSpYVNUbgEN8RW57LsoH+jtbE9xzXw1OWUry1clFUmGypG+19ja3rtv3w6i0blcyo48lP5lNTy8n8cRaacMGQaVZDYkm+xPQ/wB+ILOs7ybLqpq2jmirWkAZowPia1rC29tr36jyxn7V1VOmk1s9iLNeQkj69sDkjSWjBIPVrk3w01eiW1hmZU7xPxXV5rRx0ghWGO4Z7Pck23Fz264rylVOhVW2xsOmFgsyFSBv2x2FCHvpO2EufJT2sSHIszuAq23PkMOxwMzbgqAbkNth+I72AIt2vfDo+LUbsO1jfCnPhOFMHdAut5Otu3p9MKAAF2W5t1v03wuV15hMhREPmbY4sNRMAscciowtrdNIt6Xtc4F1RpCoMcCnJjzaVZGALxkKLb2BxyMThNIiEa2vd9t/24ciRlBSMBNvFqkuxN/MD9mFs1QAQVj1nuWJ/dhYqwIRilqumCkSXkZRKWuC0nT6DCSdZ8TEk7eLDq0qtMXq6uY3NwAxQAf93f63wvXlUV4I4+cy+IoFZrn5nbEY89EL2gboXXCY2bxSFTuI/EfK22FwJMjhoqWxv+vYWvcdr9Pl3x2avlQhaWgIDbEyOFF7+S3Pf0wvm1LKHd44tDAtpBYMvzuLfni3EndUITvKlfeVh5nSNrYYfLaUTJK8QkcEMrublTe+KpxDxZXQVZpMoPvMk5smtOgPS1uvzvgCOh47zNS01RHFBKt7GQaGB/1QT9cXww5UKwFleTLSvAwWeGV0IDFWuQfI9sIhQfEpS/XfEVkuVT5VkaU07JI4fU+kkjffqd8SBrGhUOKQuFAUAEAnbrvYYCYsnHviUVO8ltzq6DHn1MyubelxbHpKoIoNTA0JtceNSb3It18rH647JV5ZD+HLVIGuLG4uP3YgJ6JZaOq9yUnngkmVWWFy6o4uCSrLY97WY9MdrqNKhhJpVVUbxQ2s/odd/wBuELmuUamAqo2Cnexvvh+POMphIvUgdwFGDLiBCAU2i6FmkeKCZaXKDFUThVllefUQoFrAXIAPe2I85+MnSuyyryqespa2eKc8qUo5eNHW9wD3kNxY9BibOe5SvjEo0k28W1unY/MYTJnWShyZ5YeYvmt9PzNsHRqmm7UgxNFtZmmyjIOOOHUhiVuHc4hEYK3WRXP1JA3x2P2g8IlytblmfTR9AupAR9QcSa1WWvKp5sMYfdQxADDzF7Xwr3zJVsr1NIT5FlGN3ylU5Bcb5HokmY9ih5vaJwYJlSnyXPkhAOoGoS/5qdsLpvaVwygCpkGdk3+Lnxm48vgxO01TkMzhEmomkJvpDrf7YJRsubdEgYjqFC7fTFnM3xsnMyalsD7lFU/HHD1QWkj4ezyIA3b8aMhR2F9OHTxdkfLM0XDmasB/nWRgf/L3xLwxxyRkKmhR5JbVhiRpjI0UYjU22uL3wl2bumIT2ZFTImfcFH517UkeEBeHM0kjK8sx8xIkRbWspWMkdBigcTZ82dUkuVJllVQU5qPeI0lnMpeS2kamKjaxNu1ye5vjT54NCfjiMknfwf3Yb91o3uWjha+48AJwJzEO3CYMpc30TzUPQBqPLqaCWNgyIFZT1Bt64eS8iBLm9jv0wuvWPUxA2A2vfEXDMwfSH0L1O/TGVpl0hbrgaYT0wikoqiKqYiPlsHI3IHfEkKmmWFKilp3ijMa6dYsR/DtgQRtK+oAaSu4tt/fgpI2dAshJuvf9mDduqAlCZToleRb2Oq2kdtzc/LE81MFgBVNQRL2xB5crUWZFJEIR303I6DfcYs0MZdmtJ4SLA+YwDO8YUqjS1dyjS9KHIC9CD0xIJGrTwcy7KHQkjpYMD3x7K4YYqfSQT2va4wSpjjvpK3A+n7MIcLwtIEMBWn8N5hNV5tI0tZHIrU6qIkQKFKnt98VL22V9RTTxU9RmiUmTzUchqlIYfCwbry2XxFVS5I0hnNjcWd9nchfiYXP/AEL9+vTfDftgVZuIcoWpj4gangXmxnL4mliWTVbW6cqRSygArfcX2x8m7RHV2nZImGeH8XXmiLYoetXbIqaOTgego6swzxNlsccpRNEcg5YBsthZT5WFh5YzvjaOKHjFGhi0U8eVxLHcEqv4kuwJvbre1+/yxo+VTy1PBtJU1SypNLl6SSrL8YYxgkNsN79dh8sZV7Wa2WDJKmTLRrmhVSgAv1fe477dMbfJyXDz2fpD/qUqtaSwnkq3Jlq1/HHCeaKFD02bw6mI3SMk38tr2J9AcC8e+3XNJf03lXD+V09DV01Zy0zGSsWQNofTdVZbWZVPQ7BvrgnK5Jv0ZTmslj5zRBj4Larje1+l9/ocVTMuHslhbMeRQmrgmq4qgU0KtaN0WRSAV/1wwsdrEdDj6dTqaTZZMRS194K2/wAnDjbiziT2g1sXEFZHU00WWsUEIQKrB1I2Qbkgnv2xevZzXCt9o2Z1K1EjRTxVTrzaOaEzAToLASQoEMV9DAOxYm5AIxWvYfBDH7R6yrhyiWgStoQxjan08pl0gjUNmLbEt1JDE2vizcGRzw+1HOZ5Icq5ckVYwanKagRUINI8ZNyLGSygaiL74+VdrtLsyxM78JvTqfz/AFhOoDTTZ61ZeO6TJm4ZSUU9A9fTl/dSI11IWdb6Tbw+EfUYzQatNyCG03IDA74dzrOEqo/fZ0SkhCCRt/Cu25xAV+dwCX3bL0kr6oEBo4jcx36M3kL4+qCjrAceiGlV4TTzVz4Ae3FFEruWYlhc/wCocWDizOpM74FzKZKOWkMGYpTrzGB16JUu4t2O/rtil+zTK6ubjDLczzOp1zxyPIkMR/DivG62/wBI2a9+1reuE+1FnpfZHxc8EtRCP5xq2tHCyWMkJazA287b+h32HzPtFh6dTtBhzzHCj/8AoZWh1R3Cn1rN/bhxRxPllJl8lDxDm9NqmcNyq2RCdtuh6YyxfaFxzu/878//AP2lNt/5sMcaS1TtCk9dLU2JLq8xcg2BDDc+Eqwt8j5Yrh2Ppj6UMvwv+W32Bcd9VznapWwezT2m8Q1M36GzbiTNZXkP9HmetkB7kqTfv279vLGqRZ9nEMKySZzmBY7b1Uh+vXHyYkjRuHRyCDcEbWxOZVmOZzrIv6VqI9IuS05AsTYE3YdyMA/LMI4Rwm+wfBOpYlzd19P5fn2aszOc3zSQ7eE1T23J7XwZW5vmMtOVOZVrRupV1NQ5+YtfHzNRZjXJBzUzmoDafhE5ABsGH6972LfXbrh+HPc3kqeR+lqpomR2ANQb7IWBK6+u33NsI+SMNypt+yPgtbcbG4W4S3hVGC/hAhfivt0H/DDUGayQ1EVRQ1XKkgYlGABvcEWKsPXGJzZlmfLkRc/qPBdgnP2YAX6F73sUAHU+IdRhg5jmIqAWzuqsLgnndWABsLP0vbe+4F/TGxtIg7qnY1p3C2zNn4izOiqoKrifNJKer2lTVGAw8tk9B0tiu5/w/JUcQZfnUKSVVdFM0tTVPKC9TqABBW+1m1EWFvFbyxmTZ1ny1U1J+mJuashTUag9QbbHV0wumzXiNSIBnjbGNlL1gsLnz1XuCN/K2+NFOWuDgdljqGnUaWkbrRqbhXPZaQTihaKHWFZmIuD5Wvg2fgfNIFEj2INrgTRkg/IG+KBDnHE8aKY+IpYg+kRh6zTsy3B3aw7g+XfDX84+JpIIlkz5pQ6kjVVDUotfexuLjt1x6T9IKzjYLzI7N4fqVocnClbHBzDpve2guNQ9beWOLwrWOl41uwIU3cbE3tcfvxQznXECQBmzyZlKPrLTg7hb2Hi37W+eER5hxIuqb9LSuFcrpWcWuCO1+m+3yPkcW3Pnn0rJb+zdIXEn61d6vhqsojpq5E1DqsbjbA5yWoEmmNdYZrDxLb7/AMcVqnzbNZ5351fEZAyqW5pe4JI66hsOp9MO/pCvFMJoqqmLeHYuBa5A/t9iTf03xrGfUWganX9S57+zuIc7uMt6wp2oyuqhWVpEj0xpqciVSAo+uAIZaWVQwmR0J06o2uCd77/Q4jZMzzGZJFFXACwJVC9iRuTfxbGwO3y88RtNmdctfRxr7ukEcqsGV7ABrFlG5tcEA4s9oaPJ3uQt7N4g2c33hXmkyGrrIiaNJJgLsF0X2Hrguj4Rz+reQjL0isNI5kqrf5XxHcf1tbQVcdPl80lNFFLPqp+aY2S7agxta2xt5XFrXxUYcx4medmNfVEquyPXSr2HQagSdwcZ6vaKHQwCFsodlgWTUJlXqo4Tz6Bo6d6Iu7knUjqygDzIwy/DmeqZHfLZ0Xdbki1/4ny64p1RX8RiEzNmtSbXtorWJAGx213t3v3G/TAn6U4pWWycRV8QL6SyZi4F+nZug8/UYX+kdQH0Qm/otR0+kZ+pT/G1JUUOVUGX1tPJBVQ+8s6OPEmrlEXF9rhDb1wiryymjpcqzeIKkM0cokdALI6Xe3S4uqjfoNyOgxT6+TM6mpepq83eeVgt2kq9T2O4BuxO3l2OH+Fqyt9/MDVM0kApal1jZy6ArTuR4SfS23bHFrVRiMRxCIldyjR82oaAZgL6RknSqh95mKwTpoCo3VgFtcem32xW8zoYKriCHNGYxzihqaRygGphKhQG5uPDqYi9xvjJPaDX5vS8Z5xBHmNTEI52RIo6goqBVFgBcdvS5PriuDiHO9S6s0rSR01TMcY6rJeS1bqdYBgDlt8VFRtVQLWGSMU0ss0DwzMea7oUOvmXt4drKQPywaMuioKZTQZqsXOMbSKqBgNChbbm24G5Fr288YAeIs8ebmfpat1A3F5mI+2CP52cSBSjZ1WMLWtzNsTSSIKgqsBkBfQ3v1O0hJJKAkWB367b4eVYGiLROshAvbe/XvfHzpTcUZ4GOrMJnJ/tufK2OJm2dIyyJmtQrX/60b7j54DhEpwxQ6L6GlkDut4w1gbnVa30thFZOeSoLxCxvfWTj59mzXPJ6gB8yl1aSf67SCLX6jbCJa/O1gZ2zCfSp0kCq/vucB5ueqZ563ovoSF6aoDI9QFYW2UEn9mG6qmpY9H9Mn5YbsFFj9sfO4q80jDP+kJ1KqelT62t8X+OvTHTVZjIA0mYz6bE7VF+ht/a8/44o0SrGNHRbrUSwPUl4pCXHQ3DavywS2aNHy40Tyuttr+mPn+KfMm2SunA2A/pHY/97HjNmAfSMwqHZgDYTny+fYYOnTLUl+K1GwW+pWQioB5chfVZgoU4RJJMlXrFJKQNjdlFtupA374wgVubuCGzGoKp/aqCO3z3x2StzJLxrmFQw1G6rMxv+eDDGlV5yVvEVVJOuhUkRxsVCm9/+G+Bq1cxddQp2IYfE7qPp1xj0PEmego0WZTgqO73v9zhS5/n0btozeewsCWZR127/wCO+D4bUPnDlqslPWxkK5RJEsW8adLbb364chSqLbOim9jqkXcYyc8R8RmLnJnEwDECwdQfLp2wZw/xdnEOYxGoqpatWcFkZgQfMfUX+tsLfSaibiDstUaORUfUyjTsbOTt9MC1EM4F7qdJ6iQ6sF5fX5RXUQeOeR0cXDKgGnfoR2P+NsDZjUUNMjTLKz6AW8adAPUHCIO0LZIAmUJIKjSRcuFAtduow28NTpPIpn3U3u2wOID+ftAtYIUUmAmzPyj9x4v3YvFJUxNCJknWUMt1ZVsCO1t98A5hBuoyox9gUOuWz1JDQzAyctVZXbTfa9iTYbHAgo3Kqj8+Jg24LXHa3yxK12ZUtJRvOxdGUFiS1l6d/TFRT2l5WKgRMKtozYMzICo+Qve2NNN4N4Sag0QCrxQt7rACkWqUganEhDX8rKQO3lgWtaonqFkllqJCB4lLk3t063w5S5pldVTR1dJKkyyjwt0BwzNUx676QCd7XFvnviVKjnd1GxgA1JpVWzsCUcXOkKf34ZMaOEMqNqJsCD0+e2H0ko2Uc+QEXt4Rv9d8ciqcqjmLBHdx0JbAaDzRNdOyTPBTBlanR1VTvrfqbfsw80sa0bKpB8QLKPLDcldRDVog3t3tc4HqWMunlU8sUjWG42J+eB4ekyoagNk5URglGhGrt16i/wDfibmCS0qQP2ZJFC+aEMAfS6jEQAwpFEZCkNYm3XfElcBlEh/EAO36tvPFgQiDpCA4liWn4brli1gCjdQCwuPDbr3xk/CDUwyuQyFNfPOxhDbWHf741zP4zLkuYeNdPu0nhN9rKd7Yyfg1mGTSMBzCZ2uBLp02Vd7Ejr+7HWyRoNYT4rzvaMxRP1KRkeCRUAKgEm9oQD29MX72WQxy0U0SC6S1oBK7bkKDt59MURJGCxl4CCGJ/rR/vY0X2Uq8tIsiKUIzFd2YH+x88dnPmacE4tAB9S4fZxxOOaN9+a1HOPYp+kZaMxcV11HDTyGQxxU6/iHtc37b/fE2PZjHyyj5uzg2/wDq4H/zY0NPhF8RHHGbzZFwbnOd00AnmoKGaojjPRmRCwB9LjH5KpdtM/qVG0mVzJIAs34L6M7D0gS4hUvNPZZNUUYgouJJaGQMDzVpFcgDsATbDeXeympimnet4nkqxI+pF9yVBGP7Is3T574Cy3MuKMkzPgatreKpc8h4mcQ1dLJDEqIzwGVZICighVIsQSbg+eIHhPiTi6r4CzniGpzXihpoMqzCaKeWnpRQ8yMuqaCo1lhpGxFrq2PSDNO0ukvGNGmQJgXMuEXZ1aVm00JjSrs/suhZ7jNnH/5OP446vsyjXpnEg7f1A/3sQ+V5lxPxdxHmeVwcV1GRwZNldDIrwQxM9TNPCZDLIXU+AWtpFu++Bc34u4jg53L4goqzl8D1mYe8ZfpanlqonCrMhIP26dcCc37TcThjGDVuRAsCJF9EXHir0UInSrC/svhdg36Wk1f9gLf+rDUnsnp3Qqc5muf/AII/jivcJcfcQVfE3B3DucVawZkpqBmqqqhK2H3UywVC3F9J72tZlYeWCvZJ7R5eI+Pa6iqc7paykzSGWryymjZNdGsUzpy2tvd4zHJ4v9LExGYdq6DHvdiAQxuqwbcSRbu9Gl3q8bK2jDuIgbomT2Mxk+DiCRf/AMlH+9jh9jfg0niOU733pR/vYs3CGcZjW+0XjjLKqqaSky6aiWkjIAEQemV3tYXN2N98U/hDOOK894Xy7j6o41pcrjq68q2WVcUS0iQicx8nVbXzSBs2r4ja2Mbc97ROmcUAAGfNFy9upos0nbfkjLKP0evuRaexxLjmcQMwAtYUlv8A58BcU+zKkyXhvMs3fNXlFHTPNp5Nr6VJtfUfLGyDfEF7Q11cDZ0pvY0Mt97baT3xzMD20zqpiqbH17FwBs3r6kb6FMMMBfM5UK+wJU9CcNSOupkYXAPY4l6iBEjimQlnt8LAFNxbqD6/lgGWnW2p4SR3Kd8fpALlGxTCLBIAbC9rYIhoTdnUqR1AbcsMCaY4wojlIJ7NfD0FRJGNQckX8rj5YtWiTG7BVkDIT523wNWwvEtrEgeQvgpKwTX5l9QOwBJwmok1Lp1lb26XJwt9PUoCoKoaQ6iUI9ANv7sIYLIo0Ao47X6/liRkhDlyrXN8A1UADEKpv2PngHU4EpgN0qnmRJvxI35NgsgvqUEfrXPT5YEromidxrsba4nVr/I9MeVJ40LJJImtbOFbqPXCZGQwxsqm6LpbfYt5+nyxRd3QFGi8pHCGY1FUlRQ1aMJqaxI1+Jr28Q26XxZqSeeOch32HXyHriqQ5bNJU8+g0RVxYMrKbagD8Lf6JG2Jelnp85y2SdI3injkEc1PJJ0ItqU9Nxgh3o8FThG6tsdXHUR8tCGNr3vsT+zDdQrSINBtIPECbm22AOH4xE2qONOVYEKpNh/fiUdZmMhUAR9rbn7Y1MdqCS8QUmna4+JybtZQL7/M749KmgMQVjTTdQDf91r9sL5asSqgLdR4rnbc7W8t8NOixzO8aBmt4jp3BAH7gPywQshJREHNETSA3ItsbD/hip8Z1Swp+khCzNFaOZFaxZd/uQT37X88WKBnWNmeRCL3B020/fFI44kefLKrZFiZgutjbvvvgah7qqJULwrVVFTmUmavRGOOokMUZ1khVCkk2t/ogX9bYss7Xh5bFiWJBJbdgOt/Qkf+XAGRQmlp4oKSSKSkkg8DBrhZBbUdXlbUfmD0vgpi0k7uUZUGyK3UL5YwVQtVBNS7KDayggGx2xIcNZe2YZkh/wCjufF3A74C5RldUC7Mdv8Ahi85VQDKspDgAVEo2P6yqTsP4+uFsbrctDnQ1O14iDuhA5UY0Lfe9vTDdNHAXtGzW0hiNxpJx54n5Qdgxudib2NuuHYIj7rruJL+guSB3HUdcdVggQuY65lS3Kk/6vN98exFcyb/ADkv+1j2FwF0oPVF1lRE1XUKzLeTTZn3LAHt2H9/XEJ7RM2yumpIaF5jWVsRBRoztpboCOtzt8vXEZmVRmmZEQ0qxwkjU7HZ+vQDpgSnyOGCq94bQ8ym9ybkEi256nrhdKqQIKyVGd6yDy3LveEFXXyTrPJvyFeyLsNttycSkUBQEHUVB0qNdyBfzwTHEyC7DcC9j5/X92HWhViTMoQW30mwOBc4RYpzRfZArpC6jqO5ANwehOFKUDaxsLdSb98OJTifVFRxE2O8zXCDz3tc/TDxyqARWnYzMQNt0T7dT9ThJqEc04MB2CA95Z5OXSQz1Uigk8tLqp9W2UfU4MhoZnAWsqil9zHT73PkX/gPriSRNMCRxoiIp8KIoVV+gx1uYVtGF2O5A2t3wp9QlMZS6oBo4KJWnURU4tp5jt4iPIudz8r4j0z7LjJyoVrKyYnwxwQMWJuDte19mB+WCOIRIlZTUtOC7VJKgDYhtSWuew06z8lOJKF6XKGSOCFS0oHvM9gHmIHfy7C3l3v1luavRNgUDV1NQsa6culec3JjYoukWvdmBIX5E3wLJX6YtVZQS07Fbk6hIv0ZScGVTNUys6holUnSCSCw+tv2YJezALJGNN+63B3HbFWQFjjzTVMsNRB+FIsqAgag17bdPQ44YEsFeMS26X2b6Hv8sLFFADqgp4oSpCiSMaTYdBta4t2wuTmA2ABUHdj1GCa4jZC9oO4QJhhLhoNR0/GWG/oDffth6Sj5lBNFIjSLLGylVNtQtuL9sPyhDJrOnUB1vvgWsYtSKjrqUt8J8sNDpCU5vgoCOmy9sximkhjbVGCUdQQobcb23IsT57YsNPG2iLmMJnKi72Cg/TpiEENbVZtGkUqQ0iIBINI1Od7AGx7HzHfFijVS66dlIPTp8sMNgIS+aYqodUVn02B6E4aghiZvEOm/xWtiQkCLtpBN7Xv6YHRACxChfngIO4Rhw2QtXw3lNTO07xu8rm5JcnfHYeGMoUFTG4c99e+JGMEE+Ek2w6xRFViVH/fxNUbhVAJsoibhTKSyWkqYynULILH7qb4L/m3kQ/6CUnzMhP7sHO7soOodPPfCdbBtRBtfaw64jqh6IwxqjajIMiI1e6yEdANZGGZMpystf3OMN0v0NhiXaYc38XTfoDqtjlNTxTyMyoykdb98UHFA5omwUMMnykTNK8cpdoxHs5+G5P78Il4fyuaERjnBgtiQ9iPuMWN8rVCbkXt3XAslKkY0oQ7267DF6gq4Y6KFXhnL1S0cs+7Enxqepv8A2cNPw2iEGOtcKOqsg3+1sWRFg5hjbSFA6qbd8HPFQql5I2DEWvqtbfrgOJPJNbRuqX+hc1ijtTZmWBO2ssmnDclHxHAvhrOYw6gTnb13Fji1VM1NG7IhJI7HvhkLqGsAhb7jpg2kG6F7YMSqpNU8TQyxIaaaXTdygZb/AHvb74fSt4lazNQNFcHbwGw+d9ziz+6qYzI02lQPEpG+HPcQ+WGuaemWPmmFBpLOzWv8I6C21zhoOpD6Kq3v/EXJTRlpcHqxKqdu3XHaLL69mM+YzJGrfDFEAx892PU38tsWcUwh0yRVUUqmMOSIyNLG91sfTvgWWOWaUuRvYX8NvthPNM5JsiaIqVVljtfc72wYod4UOlhuD5YREJFYiNV+u+DGneONRI177jtfBAyouckywst7EMSpJ3GJeipJKWljepQq5u+9+mwsb/PAdFOsrcuZrcw6UBsLnFqzatjqZpqWoTmPHG9nTpe42+VhfDabQEuoQVW6IkMUE+nV4ip3LAH+/Dz1AJt4AtupO+OxU6xpHqsxA/UBJF/PC4cvZkMyxS6XB3t1+WM7mXWljrAEqyey2qWXieNFa5NO5P5YV7Y6U13F2R0oruSJEWNhLTvJCgedFGvTNGdMhIjIAbYm9hvh32Z5c9JxGksiSoTA4AZdItt2wR7Wq7LIc2pIK/LuE2ZaSSdKnP0BV9LAcmIkfF3Pltse3yDtI59PtK0sF9HKOjuq0VYNIfnmrwIfdOGxBanUQ0gT8FNMYslvCLmy+Qudu+ML9qNUoy/M51qFDR0LvqQ28Wk6flvja0qVPBC1kNEaJTlwkSl06TCOXcJbtbpb0xhHEmVpmerW0oRjdlDdfn5i2319MafJxqa3FB2+ofiqczUQmyddPFoCMhUWUi9vLe/b88LgeQIVjKqFHQ4ISOFKVUeJwSOgubfXCo0hXU1mXWNgRj6PrcDOyZw2xCsnsw5387Gm94idxQyBI9G5N13JvuPTb54M9m4qKniHMc1qcuqqPmU9SVSakqIhEzSqZNOuZ4xrZdRCqL7G/m17LKYfzlJcBlalkUg9CLjtjvs9fLYPaNnmXUlDQxuI6xo5ly2GGZkWoVSpkWUsyAnSNUa3Cjy3+Y9p3F+MxAAvw2+yXSlVWtbpjxWH+03M85m4VGZ5cjwZciInNNi0wba4FrqPXvfAPsR4xp/eGyGvihjkkANPLa3MN91Y9zvt9vLGttTwVWVChrKeKSCWPQ6MosQRuLdsfOftH4ZfhHiRGoZHFLK3Mo3J8YK2JB+RI3/vx9dpkObC5tTVTIe1fR2S1Mj+1bgyFRaEyVZZVvp2gNiQDba/fbfEjx5lBzL2V8R0BWrMdXxChW9MFkA50QsFJ6XFgSRcWOM59gHEOb8Z+0vI56uKi05SkvOKtaRuZERrCntdQDbpq9cbj/KGr6nKfZHnOZUj6Kmn5LxMRcBucguR8icfLu0tY0e0eFpxOrh+6oVqa4VKLndJXzFnvs4CcOzV8ktWiKXljqWoAiALGzOrWNgLKGFt+otjMc6y4ZZMsT1KSSWUsoRhYEA3uQPPErn/ABzxJnVBJQ5lXCWnkILJylHQ7dBisuzMdTkknoScfWy9pFmrjFpndeNg2+/yw7SuqyoZE5iqdwelsE5tSRUk8CJe0lLDKb/2mjVj+ZOJnhvLaWtpIVmZYjJMFeTTqKrcXIHfa/fGjD4N1epoG6zYjFsoUxUOytXCHCdFnUUUmWJPXswLzQQvH4VsduhItcXJ2JxNUvs1nne9PR1koncLzQ0ZRQxspNhceHrt2xHcCVWXcPZtNHU2SmqJUh5jfChIcgntY2tv541mnXWniW/r2xWJZ5rV4bmynYN7cZR4jSQFnEPs2Vy0RgdJW1wkNKAJNFjrXQpuDt2G98BTez9KTTMKKssJSkkkbAqttyQbbjp08rbY1V4lYKZOaoXe6FRsDuLkHbr98Jo6Wgp6yQwExzOCzKXJFiT/ABt8gB0Aws4ln0QtBwLos4rH8w4DFBIK6VHSLSJZGkIKsrMFBGpLC7EDfzHY4lU4BgdIJloDeVZOXGX0G6uVUnbpq67dCDiL9uzSniCipzERaAssuokPvuukDqLLvgngbjbNfc8jybK6cTQUkapWO6Ns7zv0Iudw6WIF732OL1sN4StBa7SSVM/zMp7lYqKamkW2iTWLaifCb29beZ73thFdwfTwZRHMqQRxOWj5kraBtIVIOlSdQKEetr4vNFmuYS0VQXgSjqKaZonTWZEuNB2bSo6ODY2PXyxWvabxW+U5MaVqKSslMYLyoDy421DaS3S9+nr9cAKomNKdwe5qLlX1yHKqeskoXglDaeWbsQZDs1wWXa5AAvbp64sFLwW00NSVpZzBoXUwCDlM4vfoD5bG1t8YEJZkcsJPEdvCxxpfCOY57xjllPwtz8vvR6Z4pKqr5EjaOgRrG7gX7Ha3liy6eSS2BMlWqg4UQRoJnqgzCQwxKhcEI7Ib2W4As/fv5Ww9RcHwmVYYnSeGQNKIoPxGADWQE6NurXvc9Bi0ezFcyq+Gaam1slKC4Weon0EyGecuLsFJIuvRf1rb2OJaoqys6wmSVyp5akG4A9PTCqrmAwQtdCm6oLFZZn/BOU00/MSsqeUihCphaE3udRZmU32t2HQ4Dfg7hiamy6opuIp4maZOcjgM0gIv+HYC5v6H672A/lDGaPimjZDIF9zC3ubXLsSL/n9cUXJqCszOCZoJlQUxVvGx6tfp9sPo02VCGsZJWDEPfQJc58ALXM1y/K+JKkzV2cVEM4gKq1wYxaIsGdtAJuRqJvsCbdBhiu4PykTPNQTVckBLI0rBmMZWwPwqPECPUbWPfGdQ5JmiReKvCxv4WCsx277d8WHhCjny7M5YJFy+upKmM8xK1LhBa5cH9U+Z+/TG+pl9YgvNL3rBSzKiHBnF9ytcnBcDwcmlaoq2Q8plQMlnZdQ36fCpbf7nD6ezRpqSTRl7grN7uQuYKVVtmuTy973v++wGKdndVUZXSRycLZ7VVtOM0kInXXYWiTShLAaiLyAdrb98afwLT8accPBmuXU1HQ0EHMiqY0m1SglCwbQ6nwltG4vsT5G2E1QyAGrpijxJ1PKrEPA0VFNJHHl1SzyBm0ieOR4gABqbw61F2X0Ood8DL7NsxirSyyVMaMk0epqOQgB42RTso38W/nY/LFn4YmzSfN6l6VKgTZhVQzU8bRhY6uKSBF8DNpuQyk2/0T62vGVyGt94ano66CEMseuppZIAxa9tmAvsDhNSsAdWmE6nhZEa5WPcS+z7NKivq6qWikkNRKxiVI5FcamJ06WsbKNVjbt6WxT63gCvoqRqmunelSNNUhkp2VV8tz579u2N742pmOVZrBO6zFaWW5FzYlDfr6HHyY2+kgXHywQq033hLq4d9K2pTUmW5Oqlv5wwtYbKtLJqPy2tiJYRrIwDllvYNptcYLyjLhmUkiLNFAyKCNd99/QYkG4dlWwfMKJT0F9W/wCWHswlas3WynZYn4yjRdpqPv8Anoh8uosrmX8bPBSORuHpXYfK63xP5RlGS5helpsxjqKpVDEiJowAB4r3Q/e488RcPC0nLMhrqZRe3wm37MS+W8NvQzQZhR5vy6yKQOgFP+Hcf6Wq/wD5caWZbiedNZqmaYUbVfd/RSsnAeYzU8lJS5ZJLLHIxWpRWuSraWAGkXF7jrfa/mMOZTwr7zUy5by5TXRBXnjK2ZAbeIgRtb4lBW4PU73xpGX8UZfT5TSy59mdMk0yFlJbSG3INgd7Xw1HnfA9LVV2YJntJ77XpHG34pAshbcjuTcD/ujGKpUbSeWOZBC6dGi6q0Pa8EFUKi9nmYSTPEMkzLko2jUZUCNazEgle97fXfvj0vBFJTZhBQVFPU01SwJlWOqhYwrrQXN1GrZ16X62xqOTVWXZlSNVZfVwVUesqWi8QuOoxintSrqjLPagaymtrpTBKiNfSbBTY2I2Nt8LbiGEkaUdbBvY0OD05JwfmlC6Tfo2o5MhH4Uy6GKjw6tRUJ+t2J3IOJTL+BquOKOtrcvroTNIoRVqIy0mprDqPCNup63vsBfFPpcmzWelhqhXTkyxhgS3QGx63v5YsvDuWV+TVcGYU2bzy1Ckaoni/CZbi4ZtZJHyGOizAVnN1CkVyjmOHa4NNXn0UrV+zmr9/aCpyuuhVFHgSaEnxbA3AsTtuBfse+Bo/Z/UTOOVRz8uR2HMJjsgtc3OmwFr2N7Ag7742aKnp/do5GiC6lBAA6XHT88caj5hGo/D3vb745TsQ1ti1d1uBc64fZYXmPBDZZMs1UjIGVmdA4YtuLaFRDtc2uL9dsTrey6WeGKSJpJdai8iUMhRT56gp1db9LHYXGLF7VBFlkvDuZtGZeTmKxyhB4zE4IdVN7C423728sVr2qcbZnlHENRl+SV1MkUVTKojBEsiqGsqv4bKwsQQCcObVa5pOlZ6lHhv0lyUvs5o/dXMlWwexWNlpWYAju10H/Hbe2K7NwOsE5dXrmiVTJGUppC5QS6SxAXyVt+m4+qaXjzjki6U5lt1tSMdv+7iHz/Ps+zap59fTSQIYPd35cbRgoTexLXtv+WC3+Ykuc3k77lpnCeUAwj8V+XGbzIY2DKmkJGpDAb7De24DdLWxEceUFfQ8HVs9RLC7l44yiixiFwG3HUGwHn1+lpqKmpo8kFRmtXSZTIzK952usm7WDBbkbEkb37eeFDLZOIclqjS5plFZTTDQdJYhrgbXA2IuDjPWZpqSLBa8PUL6cOuV87X63uDi6ezbPGgziDLK6pVaGbwLqHwsegB7XxFcdcN1nDOeyZfVgFSNcMg3WRD3BsPkfXECCRYgEEG4xCxA1+kytw9ptPGnCFZytCEKCTzd7a1AsO4O4+eMPIsSF/PFw4PquIM/kThmGqlqfeV5UUTzKq732uVJNz287YjuMuFc44VzE0GbxRRz9wj6tJsDY7dbEH64FlPSEdatxHAqW9lvElLlmY/o/Nl10VQQqsxJ5T32I7AEnf6HtjbvcoRumW3v2JG354+XXVgw7HGxeyDiLirPVThzK4sqnqqWHUjV1SYmkQHoNjcjb1tgH0i7ZPw+JDBDlfhSIuv+gU5YC5FgSB52ww0Shiwy2FANiVAucTJ4R9rEpGjKeHum/8A7Qb/AHcRmT1klXSVEdYixVdNPLTzosmpRJG7I2k7XF1NsA1hZutHFbUMBNqsR3SnQAdr4YllaYhLIF6mw3UjC8orIsyy6GvEMkXMTVpPVT3BxyrC+8kCO50X1mxvc4hJTNDRuoxjyljTqCdzt54mI43YfEqj18WBJacgRgAGx2B3vviTp7ypqOoqtwbrihZSW8kNn1GVyLMX5jW91kIIH+icYrwO8q5TNZoQOefjW5vYemN5roff8sqqFWF6iFo9wVsSLAnGJ8J0ldl1JWUdRDJHUQ1RWRTIFA8K9j+3HZyIf7QPrXnO0v8AZiY6IzXUsU3p2Fz2t+7F64IeUcG5g8KIKhK34gx+G0d9rdQL/liqLqEYN2uSducpxd/ZdLE9DWQysN5vhLXuNO3+PTHez1mrBkLznZ1w8+HKxX1WjLbrhMyxTRPFKqvG6lWVhcMD1BHljE8wzCaolDTzyzMFCl5PIdBiPqtculgQR9rY/Oh8lDi6fOv9H/kvojsfBjStVyDgDhDIczjzLK8oSGphVkgZppJFgVviEasxWMHp4QNtsH0fC+QUnCkvC1PRhMoljmjen5rm6yszSDUTq3Lt32vtjDahJlBZt/r1GOpUOiKnOUD+ybnG6p5M8ZVMvx5Oxu08tvn8pMespXn7W7MWw517PuEM4anetyxi8FMKRXiqpYnaECwjdkYF19GvgibgzhiRBGcrjSMZW+UhI3ZFFI9tUQAIABsN+vrjII6qba7AA+d8INZOQ4Y3YHbfe2IfJtjoA8/dbax/71Xn7PoLX6/gfhWuqKCpqcsVp8vo3oqWVZ5EdIHQxslwwJ8JIBNyLkix3w/Bwjw5B+hWgy6OFskFsvMbspiHL5drg+IadiGvfGIytUhNZkcb/wBq9/tgUtJzG1zsTsN2IwDvJvjS0NOPMeo85n5/OT7SiGNab6FtlfwDwpW8SycST0M4zOWWKWSWOtmQO0YVUJRXCmwUCxG/fDY9nXBQztc4GTIKgVPvYTnyckT/AOd5Wrl677303vv1xiXLmMjMrs4G4XWdsM1IkEZZmPUAAtc+uL/V3mDWwMeYiNjt09PZWMYw/MX1BqUd8QXtEkUcC52SwFqGY3v/AKBx85IjayBK4J2Xfa2DPczJFqMokDi1i1yPQ4x4byWupVmVfOZ0kH0en/EmOxstiEhqkckhCCh31N8X0wJKRMuzMVJ6Hph+ojjSndNGmdDsQ/hYeWBEL60HhYHoLY+xALCXkpuaBhc6B4egG2FGCIoHTwkbkMwuMEqG1FiGFzhZXwFtJ2NyTtfEKIFBaJCTpLMLX27YH1SIfBcHtte+DYSXl0SytGpYC6i+2OyUtbJETDGzR2sSATikJMKOLzwsCw0776gMPFklQmSNb9t8eEchhbmltRFjqJuMDSQyBwsShpCOgGKMc1JK8RKjfhwiSMjYA4Q1NzY20XhYdR54Iiiq4hpkh0kWv6jzwZHJDIFDqQSLLtii0ckWohQNE81FWXmXWik2IubHvv5YPejE2b0uYZbUxRxVH+XxOwXXp+EjuW20+oIGJF6YFyZhqQ4Fny1qcmopiWVv6xG/d5YQ4PbstADXtunKCqdedAZFIhkW7hQvOOgXYgeu1vTE7RVKOUjs4PxHrb6HFco+W4JSQkfr6r3Xvue/9+JKGq8axA2IB+/a+G0qhBSqlO0qVeysApXWxIFtrnfzw0yMist9wPERt87YTDIJEYMRfrsbFT5jDQeWFnjbpe6+HrcdO9ze+/qMaQQVmTFXOIYrhdVwLX7nz/4Yz3iSabNOIKTJI0KxcwTTEXAUDf17ftxd+IpYqOnEkqKbMPDfdibDFF4emqc5zquqzWtDAg08uJvCwt4tI7keFb/6RI6YXUcIRAEmApuFlS4pmKJLdI0QADTfxN/3rW+Qv3wsw6QqxuWULc7bkYfhRmkLkgJa1guyr2A/LHGDSTaFtqZtPyGMLnSLrcxoGylOCctNZXmrlv7ugDMGFunQffFkrak1VeRccoGwB/LBFLHHlmURU2oB5Eu+9ycB0q6SS4UFt7gAX+eH4dghKxDoFk+5tpUqSVXYjoL/AN+OiwcggaSt9/P/ABfDSuqvyU5evrYXt1w6ZjF4igkC3uQLi56dR1xuBusMlEc5f89F9v7sewzzIv8Aqn/kfHsK1Lt6VBJFLUQRMQ1hGAbkhm89rm3XzOO6Y4EMjhY41F9Tmww9mU4io/d4G3YjxDcj0xHUUs2ZOBTlI6KM2adkDM5HUJfr/rdLja+Mk2WWJdCJWRpQuiJpDa6gKQLed+gGFikj5olqHM4HROka/Tv8z9sEiOOONYwzMFW13e5t6+uEMANFiQvUsdsLLrJ4ZJXVkIa8Z0jphEzOT6HuR0+mHWVVWxJuO4wMWYxmRFZbpsLXNz1wlxTAyEtJXEWgjz1E9scV25QDMxubN2C4RHCURxpbVIAdW9ybW/YPywplHJWIPr0AG43JP1wCNR+rncWI7MqcqkIUNub6mAYetnt6BsG1FMZplIW8ZBD2ubf48rfbEXlTrW8TZm8erTTnki6AEMBH+8N9BiWknmWpYQRsosNVmNwT0Nuh++GE7JbOafWjIJk1IEB8QB/dg1zTm0YcKT0G++B4oLrepZgOgCjc/wB+FI0aSAKOaRtdyQRickYSqkmIDw2UHubC/wAu+IsSMWLkm1rbd/ph+ulSVrsPtgcsQqi2xBsAOmHUmg3Kx16kushefIdRCrzBvubYW1SkjmK+mSwJUnqPP5euOqi6STYX6DDjRQTw2lCSDqofax8/T54IiNkGoxC5TqrAl2s2+6gANgyEsgRZGvYbDqfT0wHSUNTyjolO5uAyll77db27d8Oo9XEVWeiDEDcxPdb/ACNj38sS8KNElE1Kat289hbrgYSWYQiOxXz6dMFoXYgKpTex6YU8Tcy6rcnuO2G67Jek8k0gfSFvv1JscP1FO+kADbTckDtjouWsSQtgLk2w/HGzJ1uvqb4W4gmyOmCLlDIt0HgNh13xyYSO4YAE3O9umCZlINlAIG5v0HzwqFLArbT5bmwwGoc04SeSCNIXJdgRc3vgmniMAJU+Im+7fwx7S7FwJAUXYEdPvhLoZX/EkTQp74U502CMN5p6onZxcvdmG64F/Dco2u7L18INvPCap6eMElmkPY32/biP4pzKmyjh+SuFZHGyDppuHYHZDbz6emIwSoXAbqUdQSJFjJ26ttjoE0g0eG58x/g4pHDnG8lcszzz5QqRMUMMkhjdv9JSWIIxb8uzWmqZTLHJBKFXrHcg7kbH5g4YQAgp1dadkp7NazB/7Q7YaSlCH8SRgwO1m6/lgs/iLJKyg+inYfLEZntRTUmXc6uaRYiQq8vUSbm1gFub/LFl2kIiAZRo5ZLkup1db73w05ESXUquo2BuMVWplyuoyypr6SbMIWp1jZ0WomTQGIv8RN7A977jExLXUtPmdJk081TJOI3kRpBqvYkHxdvlii4xZCI5ojiPiKk4fyX32ovpLBAiW1Ox8r26DFKm9qGWhGZaCsZ2O7alBxFe3DnJmeXLrPI5LaFv4Qbi5+fTGjcKzcPS8LZQlR7O+H6h0oIQah43DytoW7MQwuT13xtw9F1WzVy8XjOAe9YKnJ7UcqZbT5fVkXGwK/xxY+FeLcs4jklgpIpIpYgHKyAXK36i32+uJekocg9+ilk4OyWQowLI1MAjDy27YtSV2TVUc1NTcC8OZTMoFqmipBHIAN+o7G2+HVMHUpiYSKGZNquAndVqJKSq4lp7LG7w00roxAOhwU/Pp08saRSZZTTZ00bMx16mkI8+wv8AInFJnN+K6G0MSD3SVToXSSdSWvbqeu/pi5V2ZpQSvIpLOpuVJ7W2+W2MlIzK6lQQIKg87zCioKyoj96hQU0pik8YXSR1HbsQfrjkNdUoFkbWkZsRqa4t9NsR3DiQ5nxRmWa1tNHO/hWIOLgX6ta3U+Eb3+AYczCp93epgaWadgdBkcrue1rAYRVGkytdJ0tAV09nWaT1/FqiSTmL7u92Y7ggjYbdMM+2ao/958sppOJMvymFaYy8utzWSiSUcxQ9ig8TFQVv1W9xY4g/ZBNIfaPFACwjFDKdz1N0/icX32h5vU5TmkLUFFSy1M1KyGVqOqmkVdWwBgicAX33tvj492lqvb2ja9rZOi3v8D9yfVaOEB+d1MV8kaez+WaJomjXKyytHKZEI5VwQ53Yf6R3PXGJVFZrWMFhYgHwnpt542ubnRezZve5pJ5kyj8aUowZ2EO7aWFwSd7EX8xjGBSwTxqRqVyQSb42eTu3nQBnvD8VZBIQprHXUDUEKR17j8sMLmjI3hm5x8i4P7Bg+WgjKi7ksp2bC6PLqZZlITxDfoLE/PH0ioYCjWuJ3Vl9mNRUPxA7vH4hRynSnxX2IAPngn2Z18NXn0cseZS5nUS5K8lUPfnmNBLzEvDIpGzEEWvZgY5NrHBfs5UDiZQqKB7u42A36d8aVKiCGQgAXBvt1x8e7ZY9tDMKlMtkuY28xHpfGeSbVpGW/nmsGyopNSRtdm8IFuvbrimT5dScV18GW5mzt7xks4aZ4RG6yJPAA4X9Xdb28ifPFwyuMUs0RFgjgDSDa21hiDzJXyziuTPqHLZ8ypVoZKZ4qAB3WXmxg+E2tsu5v+qb+Z+2YJwIErlZg1wpkNQvsJ4CreF84TiNxNUZrSVs9GKaIFIXTlHdpN7BgQRt1039Jj25+0duIeAs2yBuGMwo1kiZ/eaiRQqvE6EgLYk/EPLfBNFx7m0MpFPwRxdCGcOxjpEuSFtv4/l9sZV7Yp6vNMrjqJ8hzfIoopJpWeqp+Uk7vo8N9RuxsSfl0xnx+QZXjMSzGVGzVZGkyRsZFgY3XMZiawboiAd1jxLMQB98OvR1StoNLMNgd0PQi4OEWB1X6Adb40+pSM5jA1gT7rFfz/q1x6LLsB548tmIXOzLHnBNDomVXePKGdsuySuVNUa5XSxvZen4S2JxJ8FKP5v0bKIRIJmsZBsBqF723ta+C6yV6iiqIpFV0WIRKp6AAWAt8sd4eokpMop4TMtwzH4fM9Meow2XcHEh420wfXYLy2KzHjYUsduHSPVc/im+No5n4blSniaRXzCIsI0vsI5ewHr+WJj2c+0CtQUeVZ1TvyQeUtZJq2JPhD/sv98Ny1YgoqmjSVmklkV9QFrAAgj/AM2CfZfklHnPECZRWyLDBPC7uTTLMDo8Q8DbHfGLNssbWc6sXRZb8lzZ9FrKDWzdaTmJYhW1KEKkEdf34dpzAtWkvhGpAFDAG5w/UUDUTikW0oh8POba+/XT2+XbDD7M6qbi2kk9h12x4moQJDTIX0akSQC4QeizD25UdZJVU+YJGI1ioC6ujIzWEsaEEdVN2Bv5NbDP8nKSCfiH3P3aUypRymR0k0LczQaHJG5IPnsCFI3viy8a0a5jxZkuX1FHUV8WYQVFI0MGkNYhHLKW2uNIPpbvj3AVFwrlHGTplGfZfDRT0NQk+qsUNGxki2ZlIJAGorp6FAepN9dESAVyq1qpCvApocwzTiGnoBpQ1buL9ikERY3I7kHFdzJY8yq8wyqd+fTTQxGdAAoYkyLf1NlFvK2JmgrMurxnNXkNUDRyVzLBJEwIcCGNSb97kHELDTmnz6oRyzsKOnLFrXB1z4bRcOPCmMnzQlfP/EmR1OT1kccu8FQnNp3G901EW+YtY43P+T7w1w3X8ESZg1ZFT56tSWFSsgElNGrLtZhpAI1Xvf4hfyMJxJliZvw1w1lsxhgjrc0ipXq+SrNCHkYXud7X7AgYtnsJ9nub5FNVZ+tatNnFFmPuUtFJHdGRXXWdV77rfSbW+E+eGUAC42Sg2COn/pWn2GcTQ0+fHgOeGCkLS1dZHUGoKNJULVyDQAoAdSOoNrgGwNsezWJVqpyGuBPIenXxHthHAXBWW5hn+c1FdT5jFmGV5rmMsVbSctUUlwyoWa51AliNiOhNumCc4saiVwPikZgLg2BJNsZ6+8LXgiR3h0QqcF0HHsFZkNVy4pno5JKSo3BimXTpJt1G5BHcHGCZHw/mGRZnxDlOaQPTVlDPDDNEfM8zceYNgQe4IOPqL2Qlf56ItrE0UhF/ml8E+1vg7Jc04onra2QxSTU8KymIhWkEevTc/wDfb8vLGnAYlmHqtfU2BWDNcJUxLHU6fpEL5kWkIQR3PxXGCKOi5slQFZ+Y1NPYAXuRE5A/LGujgHheadpIpq2NVW1lYW7eYJv174Fg4NyfKa2Gopqiqnljl1KshUgi9rHYbWx6atn+BexzA4zHReRo9nMyZVa4tEA9QsQ4JrhkueVWW5rAZqGoj5dXTBviUjZlP9tbhlPmBj7DyjPODo+FqKnhzuq5NPTiKOSkopEV1tsLKhW2PmLM+AKnivO8xzqOtWUR5tFQVDxJrRS0erUNIsoULYg2t07Y2v2Dfomf2f0FNWZPDWQQ0S+KKE1MkcjKxfUApOltiB1BJFhbHlwQ4Ar15ZUa4yoH2QxtDW5VTVbvERU0ssMDzSFr8pVK6WPg0rPuBsdhYEHG2+1+NI+F4yAwJnAP+y2+MS9kOQVDVHDVdNTCGHLoqeQtcCSeSXQNrj9Upv56gOt7br7YVA4WhLGze8D6+E4yl2qZW1rYLQsG41SVOGcyAmkZhRyNzG6gaCcfJbHcG24x9ecTN7zk1dSiQKZqV0J7gMpF7Y+UM7y6oyyvlpKhSHja3TYjsR6HC2CAixguCEfwrIiVM5cAqEFz5b4tYlp0aNxTqbMHGpNSkjpcd8Vjg+IyVFVCVVlMY1K46jE9RZbnjVPLo6dqj3aFpQ2rYRC1w19vkeuPaZS8swjbSD8V4TNWB+LcAYIjn4KfoaoIgk0QyEsGCvECtx6EWwkmM3bSQxJO3TAMlPm/uyZglO4eNTrpibagfIf2sShWZFs6KHtuD1Hpj0NN+q0LzVVgbcGfrUR7RJFaiyBmsuihcAWsNpDb54pUkVVWAyJFLLY2usZNza5G3kN8bLxbw1UZz7M6Ctp6dZq+luqaNiULbr5bdfvjUv5LuTyRcA57RVcZp5Y83mR42BBQ8mDa2PnOMYH4x48V9NwZLcHTPgvnn2a1XE3DdStUMhzSoyyscxsoiYK7g2ulxYsLEW79MR/tlcSceVbm6/gxEgixF0BsR1B3G2Pu6TJ6WaqK1KQy2qEdRJEDYhRuAf3Y+Pf5RfBGe5d7Sa6o93lqv0lPLNCUVmLJq8Nut7LpHoR2FsKfRDLhObWLhpOyrWUxxnJKQtUst41NrnbYbYsFPFCIU1ObEje98ROVLKMipP6FG7LEoIJsenXE7Ts2iPXT6ht4b4+hYURSHqC+aYxx4pv84rZKiaKlyeihgOoypGhYmxK27fW2HKr3MgJDNIH03k1jZifLfa2OCKOX3dJLlDEmj0vbb92GxGIa125ULWjCqQSW33IN+hx80r1BxHAjmV9hoU/2bZ5R7lmP8oypr67hqlLxx8qGZFZlXe1mAJPzOMW4jzEZpxFX5oqsFqKp5QG2IBYkXt3xsft2o6ysyWkrYTIaalYrVIDYqDbSbd9/tcdjjEamBonClt2UNbfoRcX+mGseYXPxDe/daBkUv6SyyWqgqHVQbSpo3X7Yhs8zeigkFFSyNWQqfGxTSCe4G5JHrtivZfIkEqmeNqimP9ZEJCmr64TWU7QmGQpoinTmRXPVdRXz81I+mOxUzao6mGtAB5rgU8npsqlziSOQV09ofEMWfZFBJBNJJHDVCKMP10qrm5/2uvrbtiH4H4yzbhOplmy8QypKPHDMGKE7WawI32t1xBQaVVw0OtitlsSNJ8/XCNDE7x/345dQh7tTl1aQNMQFp/HFXxdxlk+Uy1+RUfLqZgtBUUyNrZ2F+X8RIuOxt09MZ3lGWVuaZ5TZRSU0ktbUzrBHEL3LsbAffGt+wXO8uqQnBecrUCOabmUjc20bOTcrbs3kQfmPO78GcP5ZT8dcI8dxJHTmkm0ZigHgfmzPAjrcm5DSpe56b9gMMPeEhRtjBXso/k25jkuXJms2bifNaccxaSClEiBuykswv33FrHodr4k+JPZc+aQ5zRVMNNBPNNEmvxGRZEZW5nSxHLZ7W/tkbdvoOmzXLs5kerybMXqTRcxJqeGy80i6gHWBcalbSwIB63IwAhjzaq56QzwxtCjWcAFXbc/Ubg79R9cC8d2eaZTN4X57Z/w3X5NndRk+ZK0WYwvy+SUYlyfhI23B6jzuMfQfs9/k653luURZ/mObzZfn0cXvFLT00Ss0Ug3VCzEKSdha9tyCcaB7R+Ccpr+NMh4oMUJqsproHlR13qIlcFl9St9Q+Vu+NdqDRVtZS1UeYzK1MpJhjc6ZFk0jxoOttrHtc4XSMiUT2gbLO6WpzeGlSkzWjWjzBGWyrINMoI2ZbHp1GnqCD12JxuhSQVmdaI7g51X3BG/+UydcfQXF1dTVlAY2oZI6tahkCMAXTqFbboSAG9PmDjDKZFTMM7FmJbO6/Ue1zUydDgK5iy04XeUunjslyVa4A2HbywDmA0SSLbSSRb5WxIoFQkAFbHY9cQct34hqYlLHlohszHcm9z1t0Ubed/PCdEQVrdUBOydnjs6FdQXXbr64kuH3iqKCaaJWVTPKjXvuyOUJPn8P2tiLiuZgS5Nn6drXxMcKUqfoiaujqmEVS5cR6VGgHcm/W/zOKcDAKjDJlQ2a8a8M5RmUtBWZiIKmK2tRE7abgEC4BHQjFdzPiL2cZiJZpa5veJEKtLHDIjd7fq2uLi1x2HXFJz3LjxD7Xc0oKl3pFZqghguohYYGZevmIx98BRcJUzxy2r5br/8ACG/546eDweIqd+gNvFcPHZnQpHh4jY8oV7oc79nEcTK1a0x1alMsLkj02UD8sSOX8b8FUJYU+YIqkhtIgcC9uvw9cZr/ADQgWmjl9/cF9gOUP44cPB0IqDEa5vCur+qG3p1xtq5fmFX09vWudRzfL6N2QD6v6LVo/aXwgwOvNtHpyJOv+zji+0fhAHw5uNI7GCTb/wAuMiHDCPSu3vukx+C3KHne9/PCm4SgvCRXSnmdfwgN/vhAynFfR96cc8ws+n7lrB9pPCJFjmp09NoZP93Dae0Hg9JB/wC01YdwYH3/APLjL04NhZ5UbMGBj32jH8cNvwnD7qJlrX62tyx/HB/JeLj0feg+W8JNne4rWqj2icFPblZloW/QpJf/ANOGl9oHB+7HNR/4cv8Au4yyXg+NZI0GYN49/wCqG3545DwfFKZSMxI0ecfX88X8l4uPR9/9VPlvCfT9x+C1g+0ThBlsc1ACrYDkyH/5cMNx7wda/wClY7j/AODJf/04yscJxGBZPfmOo2tyh/HDn80IRNyPf5AepPLHT74D5IxX0ff/AFRfLmF+l7j8FqS+0PhMC36UjHryJP8AdwiXj7hN7gZsh8/wJP8Adxly8KoY5ZTWkmPtyhv+eGn4XjVI2NawMh/zfT88V8lYofN96v5cw30vcfgtLl474eaYRU9UKlnGwWNgdXYbjFvpozPSo5QjUo273/bjBP5uqlXpWsfwC+oLY/txuXAM01bwVTVMhcnxBpC2osyuU79yFBPqThGIwlbDN11BbZbsFmFHFv0MN907PRqjC5KnoLi2PQqkUZjiQMxNx1vgiWO+lpGY+QJ3OKrxbxNLw5n+XQz0yDKKlglTULdmjbUL7DyXex639MYRUcTC6ha1gkq0CBJV1BGsOoIIx6SFn06QpA9NsU3Jc0q4/aVXZbQZvNxDlEsfOMsVwlHuT320gbbdbjvi8VZMciqG2UdQOuLcHTuhY9j9gg/djGzl2G52AOOR1EtMNMffsov98EM6tud/mMD1VOpCyKGB6i2wxTXOGxREtHJB1hOlpXfSoB204Yy+IyzmUHVY3W56HDtewWkMclhbcbXxWuOZsyp8mphQ1zU0s9bBTxsnhJ1K++oHYCwuD1uOlsCNTircWtbMK3ywtUoFd2YgWsBhEVAFmuJP1bC4/fgLhn2L+1bPMiy/Pqfi9IKOupkqYmlrHW6ONS39bEbYsGV8GcSZbmk/D2fcUUhro4VquakXMXlsdKrqBTe4Y9D069sNLHgSFmbWY52yh6KSoFSySqCbAAKb3wXKF/6MqR3BN/phvNopaAxsZdbFRYkFe3xW9cBUEUUdVNURRJG7AOwDXvceV/8AFjgG1TMFbSwckHmNP7lULUw6jFIbSIFvb1+WHivMqIZqaRFRwdTb9h1xLzxR1MO4UkDdQMQRHuLsqnmU8jWYAHwYF9rhQHkUYZ+WOoKjqRuDgtpddPqkYlFI2BsfniPr4KiOJWpo0litfUCdX088VnMc9MdNULGT4bqd7Xt+zfB032us1RkGQo/2hZw88poqdy0jlUiA6knt+eJvK8tjyqhjokVWlKKpI22uTv6kkn02HbEFwTl0s+YNxDmSreIWp4mN7E/rH6ftGLYiyNPJVSte97m2F1nXACOgw3cUme0acsG22on0xJ8B0DVlX+kZ4/wEJ03HW2ImVTUERal/EIUkfqgm233xd6DLRluVAGZwJABoa1kHbcEev3wloly1gxdO5hK0tQXsG1HwkDpbr0wNSTLNrGsh1GtVcaTp7E/nsd8ENJBTxRwMzLNIxWMhCbkq1lHa5IvvYWG53x2aAQROUhdGMTKh1qrSk307kW8JI0j++/SYNIsue8y5M0P4oklJChXGxFivp98FwRxSTIZXBuSFPQrbe31/dhungkaGWaFDO3NAYWKKrbDpv37XPU9hstEljk0yMFIc9FI37k9sPaUk7ru/+em/8v8ADHsP+40/9lP8fXHsBIXS1eKqRvV1wiKnlI15Swtq2uFHYjpfBhkCpyo10KBZQo6AYUkSwxBF1AsBcdd7YYdNV01BbWJYG3lcY5pcmNbCcjR/6pRe298LkVgo6swNtwCt/PDhZVQDfftbfHnXXDp8KqexUk4UZT4AQ6GV7Kbi532wSFUQ8pvDIV2v39MMq8Ucyxo4UgCwtbErlkcFRXQQS7cyRVJvuLnthbnhrS48lYEmFGlrSBY31EDex2X0+eGFjV4WZLaEPj8geu/rjZ5/Zfkc4OqsrwD5Mn+7hEnsqyR6aSEZhmS61ILK0dxcWv8AB1GPEfrAygHd32URaYK+fOCHM9JV5i+oNVTM4UNtpLNKhPraVRfyUYsaVEhF1tYEArcWONWyn2OcOZbRLS09dmZRQoBZ0J8Kqo309lUD6YJPsqyM2vXV5+sf+7gz5Qcn6u+ylNpuAWURzSKp0Xueir54Znd3IeUb262H2vjXT7LMmtb9IZha1rXT/dx4+yzKCR/7TzIADprT/dxP1g5PHpO+yr4bljckaatSkdLCxwO8V57NqNhe56Y2R/ZBkTuZGzLMi5Nw2qO4/wDLhw+yXIzpVcwzJQo6Bksf/Lgx5Q8nHznfZKScO4lYqkcZBKvqLbdxh5o2TQqK1iPFt0ONlj9k2RpYjMMxJHQlkP8A8uM99o+QwZBn6ZbTTzPC1KsxeUgsWLOCNgNvCMdDK+2GX5niBh6DjqMm4jZC+iWtlVynmcTMArFl6+HphddmDxQXcByxuFOxwyZ1hjKyKovutu/3wJRk19UlRqJSNiFXqCfn6Y9UHJLFOU42UlWUkeLpthc9xKLIWwpgyQ6lUMbCzDoMIha631MrdSe+IbhPgC66jPZbx7Le9/Pzw57w7+FCxF77YDrHBLMgZpNJ6G4P0w3TsXdCp0na99uo9MLuDCsaTdHTq2u8guBvueo745IZCdfhSM22HU9Tub7Y6AWjZSlyL3LHb88NxzAScooq2AFh5+fyxT0bV6CshhR2BUsL3FrknDMdWZ49NixI8Xa3phdcFiAIKBm21EdPpgOjRhM5LaiRv6YXdEQITebVtJQgPXVq0gkYIpZrC/YDFS9plIsvDslY0iSQiRWZFAQv4vMDc7/txYuMOHl4hgp6KWZo4EnMk2k2JGhgAOv62k79r4zrjTJKXhfIFy1swr6isqplkjAQCnVEuCb31BvH0At38sPaG2A3WKsXCZ2RuU8EcR1GVUvEOU8E1U2VR6JKxaioUwymw0kG6soIe9vUH0xosWSDh6oqcuOWxZc8cpMtPG5YoWAIViXYMQCNwQO4GKrlPHmWLwseDIBnM9U9G0fvP6WnFOJEi6coMqhLoNzewvt5aFmUEUFZNGnIaB21RGCoM8aqd1USGxewstzvtvhlVoDQ5Jwru+QUPSS3Aja4AFvS2A89SWpo44qaqkgtKrNLEAWABvtfbErTRroZg3wpe2m1/wC7Ac7LCWuCL77DthVU2sttMTKqnHLOlNR0bZlUSJIzRVtRWEOAjMNL2Fj4NzYdRiJ4XzyozbjDMJp6d6XVFJFEgOrlR6i2i9rsdjuf7PfFj42pDUcPVkoWAloxGJGYbbgn5dME5XQ5TBnuY0VJWLXPlTrTxVEICxSRabKRa9wbN3PTrvgmehKTUZqqaOSq3try3mcN0tdHdmo5zGRfoj7X+6jH0F7MPZ2mZez3h2rUbT5VSyMfUxKf34zHiHKEzzh6ry2oJhWdVs430sGuPpcDGrcDcaZxkPAuS5XFNkyvR0FPBodJXZdEaqQSNidj02xqwuJNIW5rLisDx3ypsezOkRl7f6Q64E4m4OocnyuszCnqN4wBpYbkMQv8cTx9o+WNC0kU1nN7K0Z2IA3O/Tf8sVrinjEZxw/VU2qmViyAaY2BNmB8Nz3tjS7HRZxkFZW5XI1tER1Wb1Ehbi3L0Cx6zHJe7W22xNcTTRytJymF3Ist+g0jb8sV2So0cbZc4UErE5ufW2CsxkJjBub9tvnjHSkSQulUvAKf4NcoKyUgkRRqtgviLAXI6eot9cAZzOgmkDix1Bj4Sf8ABwTwCszLmYmS4El/DIWDDSLHoLHzH5nENxBIY697so8ZHqRucZazjAlbaLQrh7EaeP8AnytSkTafcpQri4Auy3Bvfc2/I407irhD9OZzSZrDnddlk9KoC+7pEwYq2pSdat0OMu9hsmrjVU0SDRSSliZCQSStvD0HQ9sbxqHnj4j26xFXD5yKlEwdAHI9eqdp1NAcofifXTcG5lzZmnaLL5dcjDSXIjNydPQn0+mMKgqwIImI307EA9Ld8bxxirPwpmyJYu1FMFHmdBxgkOvkqEaNX09dPX1x2vJr+6xBPUfirKfWtSRBplGth0sCR88SEMoW7HQbjYHzxX1y+KyssjarEk3tfB9RDVLRSNTlWkCHQpG2qxtc/O2PpVUyFdMQZKtvCmcUmT14r6wOIREy2QAm9xi4Px3kslKzotUQQQPwxe/yvj5z4B4hzvO8srKHN8pq4mpZWikqY6dvdyQygozdA3jG199vPFxyyWmSl5Aa7l/ESBfHm817JYHMcQK+JnVEWMCB9StlZtYSFK8C0f6Xz+loK4lF5ZZjGRcEC9hcY0Sn9nuTRS1MqT1gaofmP4ltfSF6afID1xQvZjG1Px5TRRvdGSQkBgwHhPU79/XG2jpfHiu2ueZhl+PbSwtUtbpBgesq3tBAVaj4Ny9NxU1Z6ndl/wB3Ff459kXDnFuVLl+YVmZwxLOs4NPIinUFI7oR0Y40bHH3GPIDtbnQM+cuS3UmOsQvzQz2nTL88zChiJeOnqJIVLm5IViAT67Y0OoJbM4rdPdov/za4pXH1FUZfxrndNVRNFKtbMSp7AuSD8iCD9canUl/eYhvb3aE+n9UuP132U/aEknkCvm/aipw6TDHMqvU1Mgy6uqTPKW5gRYzbSB1J88E00MTRCZ5ZFsg5aADSWv1N+m18SkTkUNQTJ4y66LDtve5v8sLjkZVIdrIU2Gm5LXHe+wtfHsxSA/PivGuxJJP55KEhp4J6mukmqpYhBAXjWNA3NfspJI0j13xYvY+pfjKkctoIp5rm/8Ao/xwHHJIJ5L6dGhr6rk3ttbFg9lzsvEWzaU92cuR5alxhx9L/Z6hnkV0ssxE4qkI5haFmYc1JdT4e9++IGqYwmWaKTUdVnRtgLAd9r3vibzGpgM3LBAk03APe3f8xiAatMrzx8hjyt7kA3v0Ppj5hHdJK+vkkuCLoatveomFNvpJ1vpIH0OCqivo1nYyPEXJu2hBux9MbRw7SUsuR5fK9NCz+7J4igJ3UX3xIGhotWo0tPfrfljHy6p5Sm0ajqfm5sY9Lp9Sa7DjVdfO1NWVE9SixIrQID+I1wT6AY9G6Vc8ryRhZgwQsB4tIuQD9z98bZx9RRHhGuWAUlM5VQskg0KPGOpHTGBZhTVkdZJespwyNYiLVdh5g2sdiOuPd9lM/OeUXYplPSWujeeQM7DqsGLY3TwnbL3EsNdDBlcuT0KVElHX00qUj/1cgSZWswHa/XrjVaL2h1Dxw1VT7Po0dkDMRmEKhmXa/mBcbX3tbEJ7M8gbOpXfNSeTytUOjwNcOQSd/QdreuL4OAskvcPVj05g/hhGcds8Dk2KOGxAdqgGw6/Wqbg+I0EFUXI88zGhp8/eWkpof0rX1FWV5okeFZbeHUpt629MR09TzDYra++3ljS5OAskkWzPWWPUCQb/AJYzLjrLcqoM/qKCkmniqqZAscwILoroC3Xr1BsR2GF5T2xwOc4g0MOHaoJuP6nqntw4ogA7Kz+yR4X4lmlCBZkpmVSfiA1C9vLt+WNKraGjrXElXTRzva13W5tjL/ZB7uufvHFI7yJRkMzkam3W5Nu5+WNY3x817e5hiqObubTqOaNLbAkLWxjdIMKLPDuSHrllL/4Yw1Lw1kQJnGVUokVdm5YuMTOET/1L9vCceJZmmNn9877R+KYQFkPs7zdOEq7iKStyavzKnzKeCSGSlEG4WIBi4Z1IbUW674Ly3ifL+FKOU8HcFZvEzyRB4J3p1jCghRusjEWU2vY7KB0GKxT5nzI4+Ssjq+2soQOtupxpHspdK/hl5KiCMuKl0N1v0t0v2x+ic57QfIuAGJLNdwImN/qK54wYe4nUqZw/NLSZblrz3pp9MUsyxsHMbCxI/wBKxuNvLFs9q/F+WZxR0FNldQ0gjd3lDxMoB8IXci1/ixdvdKXp7tD/ALAwmSho2WxpYD84xjw48qLZP+zn7X9E12EaSDOy+a2zMTwToItTsFQbbkttYfXFR4r4JbiHIJGpYLZpT3aPYeLzQn9nrjR/5kV1LWU8QHvsnuj1DxpJpZyGAuARsPEPnfbvaNymipEUSc+WlqShimEcgBurk2PqDffbYkY+p03HQ2oeYSHAVJYvnzg2Hl5pUx1LGB1UAq43uDYg+uLY0cCklau/1xes89n8ub8TS57kqwOZ4gKhCpILA/GNI6kdfM4k4fY5Xv8AHnGUxhrH4Jb9P9XHrcqzjA0MMGVajQ7obFeDznKcZUxZfTYSLbLOaQgRbT3F+xw5qS9+aSfXvjRKD2S15EqvmFLGEayloZLOPMbdMcqvZXmUYZUraaUhdQKQSW+V7dfnjqjtHlkD9s32hcR2R48ukUijeD2jl4MpU8DjSysSOviPUHr5Yo+X8be0r9MZ/l/CucQ0cNPnAEzmFC15Zlp49RZT4RZF9ABjRsmoJMuyKGjMqSPDGEBsd7NfbfbrioVvDOaUfGVVmeUqqUNdLRz1qLKAZGhqFlbY266VPzvjxmIqMfiXvYdzMr6PQo1GYSkx4uALJrgT2u+0Wo4gly9M+p86MatUMZaUJqVPDaMhRe+oHcA+G22Ee1nLs19ocKZj701ZmFNSmppVj1aXjZrMii1uwAtvcWOJXJeEospzSkr8shSmqqekanke7FZiVjs51MbeJG2Fvi6bb33J+Ec8epSs/RsfKmhuJIWVer6ul++ona29/THMx2Y4fCsDq1QNnqQB71po4ckEEbL5xydQuS0ay1wiPJXw3sRt0xPwKOUsizHcAaj0x9HtwfM1NHF+iKe6dNaowv5nzwSvDeYR08sMdFEFk66ERB0HQDpjsUPKFlVNgaardo9IfFeWr9kKtWoXB8CZ2VLepi51LHE7u0dNEzyIOnhAufLc4ZmmWOGaZWvsS736n54ks+pKujhGW11JUUjTq5EplQiWzq19hcW8I6ja/fcV0wxUOVjLYXuS7OxPUm+o28hc9BsMcYYmhiTxaDg5rryLj2r2tNj2U4dyUVxE8EnBnER1FyuWz6tSaSG5Xcb9/ltjJ814GzDNM5zX9DLTyQUKqqIagySOqgqACfiKhAG0kqpIA7DGz1GTz53QVeTf5PLX0nuiOSQqmwVTt1HcjvjWuD+DpcmpGgkehskg5AghKhYwLKDc9bXue+MmOzzAZW4NxdTTquLG/sCzVKD60EBZj/JM9mHDVZwbLxTneXUuaVNTKy06zRa0hRLdFbYsTc3I8sGfyk/Y/lM/C8mf8NZbDBmUVTGxgQCONo3srWHTXqsx7nfyGNayTIRk2dy12XNHBBLq1U6IVjGrRqIUEDUSvW3fDfGOR1+fUy00eZ+7qtVHMGEV2CrICwG/UpqW/wAumMbO3GQgf2gex3wSKmCqkiAsz/k1eyPIKTg6m4h4gyamzDMKkuQtTCJOWA7IAEYWsVs1973v0tip/wAqH2T5Xk1DS55wvl6BjLK1fGgsqKSNLADYBb2IHYA7WJO9UuSVlFz0y+uFGs9VFM/KTTqVHUlPS6Lo9AflgkZQ09HLSZnUe/QOpXlzIXVh2DBidXz6nucV+m+RRHHHsd8FDg6s2ChOAvZRwbkXDWX5bNw9QVFVFGHmq3jDyc1ep5nUG5NrWtbtbFNzvhiky2hrMuy1Zp4KOthaSJIpGZI3qoZnFwCWKjXe3isBt56lQ01ZQ5CaGjNJDKspkjsrlLs5Z773NyzH5m5visrwXVVWYZjV5lWuDVz86P3asqF5RA8NvH2N7gWB222wyn25yIN/tA9jvgl1MFWLh3Vn9NxXNkHCEclJlNdS1EFLGkscsTKk+lUiJf8As6U0EddyQel8VI+1rNMsoGQVmaUitcRJEsMqIoJIA1INt7WttYb42St4NzxI+VR1uV1cRYm1ZTMrKLOtvAbN4ZHW5A2a2JDgzh+q4e4Wpskl4c4brzGX5kryFeZdyRcGJibAgbk9MC/txkOwxA9jvgmea13btXzdS+1LMq/MU994gzOnhacKJvdomCXsCxAQHpboDjfeA14jiyqNq7PhmMNn5fhCl1c3BJUC/fbceuJ6uyGjrJaeSfgPhXVTyrKjJMVYEeog3G/Q7HEXX5Lxl+kqh8pj4ZosvZl93pSJvwQFAsNAUdQTa218LPbXJDEYgex3wRtwtTm1Ref5HkmUpU5vmlTUVFOqoZUmijkjjVNRZrk9wd/Owxk/DmY0eYwZnUUbFo5M1qZYU5egCJ5XYWF723229O1sa1mnC3HtfQTUbVXC6rMhRy8E0nhbYixO4t2OIfJ/ZbxHTLIK2ryWaSSV35kKtEArOWtpC2JuT4ibnvganbXJHf349jvgm0sNUaQCICqckZYrotZTcWPXENDqk4izAlVtG0Orpf4H7dh139PTGtN7NcwI3zOnA8gjHGP5ZUpOubZipYxy1BhgutgRGNFx83LfbHVy7PMBmYcMJU1FsTYjf1gK3tcw95ERsrg6gLA7/wAMWbLKekqOD1jSICnnB/DO3hPl5dMVWkkV5JQD4TKUG3Q2xaOFRMOC8vfWWbSQ5LXJsSN798dBx2RU4krI6/KK+g9sokrNTJV0FWYZza7gUMikm3e4/wAXwGlCyQykPJ3uNQI/Zjaq/LYa2dJp1VnSKVEZjugkjKMVPY6WOMz42oq7Is1pPc6SOTJ6scvms5Z1ksTv2F7bfI49FkeOo0mmnUFyV5HtHleIqPFWlGkC6hvcWejh3dhsRc2t+WHmo5DOfG9tHYjE7w1kWbcRf0LJaP3qdI+aU5irZbgXuxHmMaVQeyGukoqKepimhqjAwqUMikK1zYABgCSLb3x1cy7RZVljxTxVZrXdC4Ax1iV5XBZVjsbLqTbXvCw+CiVIJ9LP13GrDrUeo07HmXB/t9MbjH7KoectOmRZkicw6nkqY9JU79VYn06bHDZ9keYCWNmy+m5KHotYzPb0BUC+/wCWOSO3XZ/bzhn2m/FdP9GcxN4WJijbmTDVJuOuvrhK0je42UPe/TVjfD7K4FkOjKqjRb/PKSfU+LCJvZTG0JWKgqoibbCdbfmcCO3nZ/8A3hv2gi/RXMvBYQ1I/vsFpHAA3IPTClppC0ytI+5NjcfwxuTeyqAvGTlVVdTvpqVF/wBtsUXiv2d+0OmzdoMg4YNXRBQVlkmjuxI3G7j9mLHbrIP94b9ofFU7svmIiwVDFEVoUXU+7XsSP4YWtLfMNYLEhBuTiz/zE9qTt/8AQNUjFvwxVx29d+ZfFs4S9nOd1UBbPeEq6iqksCwqomUi/azb7djiHtzkA/xDftD4qh2ZzEnZZX7o4pZyWbxHoCCDht6RxDTgklQfMbY3oeymEh75bV2Klbe8L9+uIzMvZalLTGaXL8xaGBSzaJ0BsBuTim9u8gcYFdvtCL9F8yF4CxL3FDUTFZpFfT123xpPs+p1puCaW7+IM5vtfeVz2whOEslnqWaH35Sy3GqUAfe2+JShoRl+Ve6Q3ZYnCjW19r6ifXqcFmuZ4XE0NFPeQuxkmT4zC4jXVFoT1SpdlUFgCbm3fGT+1Tky8cZTHSA1NegCvSuLxMoJsD3ud7jytjU6lzHNE4Cm7DXtuBY9PrbGXRezTOuLPaxV5WtZBGKxZq2KaUnSVVh4PCCQwBA+mPP0vSlelxM6IhSWSVdYPa7BJK9Hw1JNl6gU9OuqOrW5Ai8i1gRfYXTz6me0rjVeGc2goky9aqSSHmOWl0aQSQNrHyOLNwp7BOJOFc6OazZhlGYJHA7LThJGdRYkaCQBrv03HXqMZn7XMqzjNjFnz5cKeGACIhj4ypbY28hc/fDi0yB1WdropudOyaHtcnUbZHF/+sn/AHcP5d7VRUV0VPVZTHBHI4V5TUX0AnrbTiuxZHGtPHqghLE3O+2PNkyEHTBADfHaGRVYmVwP0gprXs2CyUkYjZQWB3tfthrMM8pskoaIVZcR1OY0kIlB0rELszsfQqCPr88QfCnNThWlWSYyyM8p1FybASMtvpYYe49yObO+DYY4WHMSugfWzAIq6JAdR/V3Ki/S53sN8cSrSNKuaa9HSrcXDiqOa2qgzKnEKkZvAgVQthWJ0HbriMzjMoazM8uaDOCxnqCjQLWkkqIpD0DbC4XcYVwrSUEfDGWLJQwKgp0CtIik2CixJ3uT1v133wqqq8tinjaAUcWhyegu1tvI2Fj167DEeCHFVTa0gGN1nefyv+kqiWRmbnShT4TpVgLavS9gPn9cR9Axkqp1QkAxxM5tswtIFA+RN/riWz+QBZojFqBJ1E2Yn037YgqMMK8JPOFjWNHW4F3sJCALD4rlT8hhFuS2EnmrDQSOI4ZgyaXAJsTa3brhNcFErK2ko+5Aw1SxLBQxQ9SkdwCu4ubgE+gIGHSWqIzcAMOm/lgmwZlR3IhB0stVSK4iRJoibiM779cDZglHVxtE2WxoXbUzaQzMfUsD5DHakTQRtIrWsSSO+I+WsluJRb54TrLTARNaHC6PWPTFypfBGLkKTYXuT+/7AY7IV5YVOr9du2I6OWWomDNex7Xw9WO0ewX4fD0vgEZHRTHB+XtXZsHDroB8YKX2HTvtc/sxZM8r1idpQQBEfCL33Bvt64heEphQ09VqVwQoCnazeZHfqSPphdeoq6LWo2DXuBcj1GHUt4KXUdDbJ6mzCWdJZAdEyXuGfdyT0tve9vtidp4Z4oi6Th5GUDTYghreOwt5kjY2sBio0kxpii6SfHZyux0kWvft1O98XGOe9lMICxfCA27i25G/n2ON4usUhD0VK881Yonfk6ykvLIj0IVsy9TqJu3iutu24w7MTd6l5fCzH47AdbX+uCI3KQmQhFS4G52a/YD0t+zDcjgqsIk/WJDdSNjtv23xbZCowVC/pWk/6xH/AOKuPYjv0CP/ALOzf/bT/dx7Crroy3oi5nYm6W3FhcdcNqioCSN7X2w/ULzBdVAA88NqshDag1lG2xvjC0yjXnTUoZVJUnvdSPrhasSjJuF874HaphIIU/ELXJAtt64GnzSipXRKuvhhlUCwMyrf6HEIJ2RGoBujtaICwve42Ow/hh/J6lUzmgCHY1EYO2+7jESudZUSb5pRt4d/xk3P3xI5Lm2VNnOXq1ZSM5qYgtpVJJLi1t/PGXFNPBf6j9yKnVBcF9Mp0thWEr8I3wq+Pyseaeq7xNxdQcP5rHR5jDMqSZfVVyzi2i1OFLp1vq0tcdrA4i8p9o2U5jFSS+6VlKkuVz5nUmZVX3OOGTlusgvs2vULC/wN5YL9o/BsXGNHQQNXSUTUtUJWdE1c2FlZJYTuPC6MQT22xAVvsuWoTieFc7kjhz+aMMog3p6bnNLNCpDf9I8kp1dtfQ2x6bB08lfhG8dxFTnvG8T7CD/wxzWZxqh1tkWfaflZ4Ry/PhllcsldXGgSilaKKSOYajpkZ3CJsl923uttzj2Y+0WakGUD+ZfEMj5sTHTIfd425wDs0RDyixCxs2r4SCLE3wI/synpafNKXKs/Bo6+rjq2pc0pff43cR6JBJrfU4ayEWKkFBuRtgvIPZ1+isr4UoVzYyjh+vmrL8iyy8xZl5ajV+Gi83YXawUD1w9zMhaNTL3NjrmC0kC0bOgc5uVQ43NHcNccjPc+nyymyDMI4IKqeleskmpxGHhZlbwCQyWJWwOnvi4jGfcI+z2bIOMKvO1rMnqEqaypqSf0QFq1EzM2j3jWTYFv7O4HbGg44ubNwgrDzP0YHXf6/wD0m0i6O8vfO2MQ9t6n+e6G5/5uitt/8SXG3nGM+2VQeM430aj7jGPl45Mej8n/AP8Acs9TvuQ4j0FmwoFnIMmogi19xiSpKaOJQqRhbfCo2A+mCAdQvpI+mPdXU3JIF7Xx9+iFga0uXpHj07hwwFmRdz9MB+9RpcJqZdIsxsQT6WwYYlmVmBt64BHIhkcOrWDXsQABfrbFtMFaCJCSZdZsZHYMLAW74ZhaaEaRGSf1dYtcYLR1duYgcgA7Aftw9LqeK5iu46emI7dU2whMJLOYHVYlVieo22wp1AZW13N7Eg9NsJi5nOu7eEixv1wis1IVRb8tuqnfbCnpoMJUsZkdGVFcgAkb7fPzwVRiNmJMaq5G+Bo0SKJmJlYsurc/u7YcEnJTSIyNW23UAfTAoxdP1JAc6T1H7sUP2oRVoqMmzCkpzUR0kkhkAXVouBZjsdhYnp2xaZKiNbnmnW7gWZ+n5Y7WRc+N4agIFkFmAPxDuPrg2ODXSs1WkXDTKy2Wio3qaJpeIqfM5a2RqJYlCvyuahQSm1iNJe4Fuq29capxRXx5VxPLltZHRUMtWVmjpqZ/w4wSYwEuB1MeogDq564iTw5w/TGL3TK6YG4OsDSA3p1It5j74zjIsuk4l4yzOXMc/jpjQ65xJWmR9Sq+yi1z3vg3vLzOwCzhnC2uStE4X4gqqzLGaqiK1sYVZlYAX3Kkj6hgR2IticGmSPWjSEtsSR2/fjLoc2yqkzeuZ6iopXhzAVULMpPNp2F2UDtfVrt+8YsVDx/ldVJJFHI8CRuAjvpAf6dQPpinsJuEdKsAIJurjLBFJBymeLS3awsSfQ4XQ0UdMjaBFGhA2jQIL/TAtDVPNRLVJCWWZQ63AAI/tXt0tviP4lz/AC7JuXFmszQiW5U8tm1Cwvuo2tcYsBwEBP7p77lPgoXuG2bfr6eWKfx3BxbW8QZHQcMTVEclafc1VJdCmcnYEnYXHQnCMt454ZfMIYPe2k5rAc14yiqPUtawxoNDrpTBmFLOqPEwkikQhh6MGxGlzd0vQyrZpVOi9k/8oBy8aCdjGdLAZpDsbXt8XkR98T3DfAftYyGlq5uMaI/o+ylZnq45WR9QAHha9jf8sW2r9oXFlLlbTxVhklVtTB5DGdPQkkKSbDtbDEvtBzfNkOV1k7Xajjnkgkl1OpY9GHSwt188OBa7kkMbUY7dUyvZ6bjjKYWkA5sbAi176ep/MDE1WWLKqnSRcde++IPiKnnj9oPDkcyhklilawII0sotf7dOuJ6rO8wYWcyuLW2+I2wyj0UqvkyiODNEVJWFmDAadSHe3h9MQOfRs1YzElY3NjbYi+18TfAxtHmYsBeRAbDf+rW18ezin5xdXdltctYkb4xYkSV08OYaJVSSSWOVVGqNiFJZWZb/ADIO+JUTyx0utZ6h2ks4JZhb0++BpxJHUCB1kKAgK4INx9Rg5paYKsUkcjBkIV9Gx377AXxgcxjruErRpS6OqmkhdnmnszhbFyLeffHoAVXxE3sFAC9OuA4ZoIHcahcPv+Gw26b9R1vh+lZJlMsWsK5B3BHfyPTBU2tZ6IhCWlPwTCRlCkhF3v5fbBuZVWqimjilRJBG2k3vuF/biKDTxsvgYqHsyj9W/qMcleTUdQ5KBGJ5ZvfV5/Te5ww3UcCGFUv2E8R8R5nDnGTzvV1OWyhqyRlQ6RPeIXd1Fh4UHxbXUG2L6aeSKtEgFoUcvqLHULqbjyt0w9/JhyWpy7hv2k5FLICqU8MsEqG9wUc9fUKAfUHyxKWE0SO+y9WC9jh+KE6SsOXyA4HdSvspjl/nXSzysxZomXe1z4epxtAxk3s6khHFFNEouwR7G3bScayp2x8J8ocfKjf5B95XSfsPzzUdxRnVFw7kFXnOYc009MgZliTW7kkBVVe7EkADzOKzTe0rJ6ji+PhwZfmqM04pGrHgAp0qjDzfdy17hwvXa19r4kvaiMm/mFmrZ/NVQZekQeSSm/rlYMpQx/6esLb1tjMOCqfLqbibhzL86ybi7KD77U1tHLmbxNHmNdIhJeYp0lCa9Kmw28xjl5Rl2Er4J9Ws1xcNURtZsgjqRzA2kE2WOo9weAFRf5QHBVLnEVZnWVxEZlTGQ6An+URBzcf6y7kefTyxAVDK0sctgQaSEj/w1xd6nOayPMKxamEOgq51AB3sJG7fK2C8rbI6qqSB8spHmmKxrGaf4jsAL2+Qx+mezebsy2iHVBI0j7l5ztBkTs0a0UnBpB5rNEBMEgO198KlsFiONvHCaEArwdpYX1AwLY/kMIbhJWcH+ZYNj15IH5Y658o+T/5g+034rzX6C4on94PesTXSZJTfY37YtXsf5C8TuajlFBRy3Di4PTYep6D54vdRwk/KJi4Oe3dRAoP/AAxW6xo8nr3h/RseX1SqEkRYwrKDY2NrehxD24y7HMdRouBJHIgrTg+xuJw1dtRzxY+Kl+JZIZM1p4aFiqrSXnDLuTdgbH/Z+2Km9JM8MsNPKyyMpVmYlhsTptfyBt+2+CYM4rn97ACKCpQOFs1j1G/b9mAzU1CBmUSB9BN7G59NjvjzLnw2AZle8A1GYX0zwqsicNZYkrK0i0kQcqLAnQLkYk8RPDEhThfLZJTa1HEX9PAL4Hh4w4cmggmizSJ0nWlaKytcipYrAbWuNZBAv5b2x+XsRh6tXEVCxpdDjsPFa3kNMJj2nVUlFwTX1cLBXi5bAsLj+sXtjCqfOa1J2eOfklm0uY4RqZRYgagL2Fh37Y2r2xlP+TjNjIbIFjJO+34qb7Y+eQEZtcVSzgA+IXAPqRj6/wCTKuWZfVA+n+AWWtS1la77Ic4zDMOIpYamrqKiJaMhTIoAGllHW172ONZGMG9gU+rjiqgu9xQO3xEr8advPfG8jHivKG81M6c7+FqfTZoYG/ndcPTGKe0lqGPjisckrNdNbKDbeNRv62tjbD+WPnb2tVaw+0rM43idlHK6f9kmHeTZ2nN3H+A/eErECWi6unsoeh/nRMsBvKKRgWANnAZfEewO4/PF6znizhzJq+moMxzelgramRYoabXqldmNhZBdrXPW1sZX7DKuCfi6pjiQofcWJG19nXyHriKySl4dyjjrK6zhHM8xp6UZjJAlRmGUGSiqZJG0SRrVAK5Y2YKXLLqFhtjr9qcvpZhnVU1SRpa2IFiYMAm8bdDz6IdRpU2gfm6+hBhupP4D9/Cf2YcHTrhuo2gkP+if2Y+Wgd5aVgOVZY9HSRindRDywPE1rfP1xqXsmSOPhyUR7k1Tlj5mwvjEqLNTLSoXnjZHUatratuoF9sbN7Gpo5+E3eLQFFS4svbZcfbu3ZJyUA/Sb+KlMCCrvjjfCcZ1m3E3GdXmuftw5Dw/HQZDLyZlzB5ObUOIllaxWwjWzgAm97E9MXThzNEzvhzLs4iieFK6kiqVjbqodA1j8r4+PYjLa2HptqviDGxuJEifWLoG1A4wFjtZxLnFXmUNd70aOSnXQjQFxcjYkgmxvbyxW8znOZ5rLVz1DzVLtqkkPgN7ddh1wzJmkLyyqjhzqYB5L6r332wh54kN4pkMlwSCD+WP0th6j3UmA7QPuS+EwXG6032Jo8VVWxPM0towbnpucamRjIfYRPPNm2ZpKbBYUsPI6jjXrbY+FdvD/wDM1PU37gtAs0LtsBZ3ZcnrWLaQIJDfy8JwbiO4nH/u5mW1/wCiS7efgOPJ0JNVs9QqOy+fKnMSYUWKdP6seJV3v87YeirWBV2qkYNHp0kg+Lz6Yg8lq1alSzQxgRgeIW7YkDNTGLkpJCgcm7KemP1O1wDQAssFxkovmSGQ6pkcWvZQBb1xvfCwH83MuANwKaOx8/CMfO8cmmblvURtGP1uhP54+iOEjfhnLGBBBpY7W6fCMfNPKW/VhaA/iP3J1FukH6vxUpbHjj3bFLzz2mcL5QZRPLWzGGpmp5Pd6R5NJhUNM+w+CPUAzDYG46jHybDYSvinaaLC4+ARucG7lRftdjU5rlcpbSUimtt6pim5jlpSjk98g/GntyGcn8NbXJt67fLFu9rcqy1uTyxyRtHJDKVN9mBKW+eM/wAwzOqDIkrswjFh4h5f43x977DuDMmpBw+l/wAxSqjS7YqZy5swnOWZfJWt7vFUwFVCW+GRG3N/9HG5LuN8fOGTZhUpnGXBpVGurhBBN7DWPLH0evTHjPKY7ViKB8D94TKbdLYXSQATvYYrEfH3CclFk9YubryM6qjSZexikHOmDFSlit1OoEeK2+LMemMKzP2e8Sy5xxRBBQ2ocvFRXcOSa1s9VPLDUlQL+HTJEVubbOceLyXBYLFucMVU0RB3G23PnJafVKCq9zY0hanXcccMUOV5vmdXmixUeT1PuldKYnIim8PgFhdj41Hhv1wPn/tC4TyMURzLMpY1roklpWjo5pVlV/gsyIRc9h19MZ+OEOMajI+F8shy+lp6ifMqjPs6erPMhSoZmeOFwrAvYuNxt+EMLpOGOLl4A4W4eq8veWo4f4rpLSq66ZaGGbUswub6QhAsfF4cdb5GytpbqrT3oMOA7txq2PzgT/KR6yri1Oiu03tN4Oiq1pGr6wzmKKYomW1LFVkGpC1ozpuOxsfPDk3tI4NhzxslnzcxVi1YomD00ojE5taPmadGo3HfuMUriDJuIKf2rZ3m8WWcVS0FYtFyZMnrYYUcxowcSq7AkAkdvPEbm/AvE5lz/OI466qWPi+PMoMlZ4xT18AMN5OzBhZiLtb8MeE97pZPlNQNL6hEtafTbuYt6PKTvHiQoatUbD3LVeK+MuHuFpaSHO6yWCSsEjQJHSyzM4TTqNo1awGpevnhzJ+LeHs3rKOjyzMoqqWsomr6flqxV4A4QvqtYWYgWJvftite0PhzPM7474UqspravLYKWCuSorqdY2aHWsWkWcEeLSR07YgI+HMw9n/GOSy8O8N5rnuVUWQz0BaKaIS82SpWYsxdlBuQx2/tYw4fLsBUwrP2n7UtJiQBI1QCSABsIE3nlZG6o8ONrLSKrifIqbiql4WnzBEzirgaogptLXeNb3N7WHwtsTfY4E/nzwvyM+mOapp4fLDNBy31U+kEkkWuwsDutwbG2Muz3hn2g13EmYcb09BTJNTZzTT0dC/+VSU9ODHpV9WhVdZZzY7+LAXGHs84pqsr41zbI8vMebV9fWQCndlAzChnijXfewZHBdb2sQw/Wx0aGRZYSwVMQASBNxAdIkeqDANxIJ2QGtUvDf8A0tQk9pnB0edLkzZhV+/M1liGW1JvuBcER203I8V7b9cXLGe0uRZqvtS4dzZqNxQ0vDktJPKWFkmLxEIRe97Kem22NB744GZ0cLRLPNyTIkyQbyRFgI2n606mXGdS448PXHym80YpFFObQwNZB2CoSR+zH1Y+4tj5WMQjyudnaylpBYnqCxFvXH0TyWXqYkeDfxWbGTYqFyTMoi8piDI19VivW2LhkWYPDRx0BN050jqT08TEkX+ZNvTFVy+iiaqXkxMukdGjADg3B2BuLG48QG4OJWlnSjqUp6gB4JEXqNttjj69Bbus4urekqSp4WBsfthUsdNUx6KuOGoU9Udf44pWY0GZysHybO4Io2kLaJiyqi2FgCqkn6gYjoco9oAui8SZQwI+JhJcfO6XOGM6goKribQtr9lsdBBxK8dJRU8DGnfxRLay6lNvvbGpjGEewvLuJqXjppc4zigrKb3GRRHAjgh9UdjuoFrA43j1x8F8pbi7Orme438VvwbWsogAQvEbY5gavzHL6Fo1ra6mpmlNoxNKqFz5C53xyvzCgoIllr66mpY2NleaVUUnyuTjwHBqGIab7eK0yEUe+PnXiSpyufK/aBm/EOf1VFxblmY1CZUi1zxSwIij3URRBgGV9r7HVqN8fQbVlIGhQ1MIacEwrrF5LC50+e2+2I3ME4ZeobNK5MpaajcI1VMI9UDdQNZ3U7ja/fHYyXHHAvJcwmY2sbGY2NjsfyEqq3WN1j1JlOccTe1HNDXZTDWyQUeVSVST5rUUhpmeImTQkYsxuDsbfCPPG25ZmeX5kakUFZBUmknamqOW4blyrbUjW6MLjb1wMa/h6kl9/atyuCSsRTzzLGpmVdgdX6wF/pfHqbM+HaZ7U+YZXE1U3Osk0amUk2L7HxEkWv6YPNMdUx4b+yLQ0AACYsBP3WiPGVVNgZzUtbbHAPLDPvlKZZYhUw8yFQ8qaxdFPQkdhseuGZs0y2CkjrJ8wpIqeW3LleZQj36WJNjjg8F8gAFOkIzYYDz0KcmrdS3U08lwO40nBcbq6K6MGVhcFTcEYB4iNshzA3IApZN/LwnDKAPFb6woTaVh8aRCGPlIVK2BX6YFqZf64zMgQMFGnbsO/fA6V4FPEQDITGpA1W7YElzAytpMLNfYAN0x+qWi26zl+o7QuVbqQQrGwZRcdMB1FZLR1q1VElZTz6SgqaeUxtv1FwQbbYTLUycogRBPF4CDYHAU9R7yVimVWCtfe4ItttiAwZSiJEKalzviuRFiTOc0EDsXfVWSPc2sFO97bn0wTLBO9IYZTzYiukhiGuLbjffFaizTlx6UfYkrv1AB8sebM6nUGEwAFz8PTDS+bhQNGmICP/Q+Wxax+j4GLC4AjUW+4wPHlGWoVYZYhtGEsVUiw72Pf1wA2azvqDStrPQefphlZ6+VTyYZWv1+Ii3rvtg/PKwPpH2lZxgsP9EewKUWmjhh93ihESm+lAyhRc32A6YcyirmolkgkE4jcFSocFSPXzxHxR1mkB6dlfqOwGHWp5gp5kVyw+djgS8k6uZTRTAADdlYIc2qIlQQ1dREALaVkKgemxxyTPKmx/pdQ177azv898QBppdILROwNrAdThr3aYf/AFd9W9lPfBGoSEsMg3S5nnq6sgtZF3JwbSNBCFluXcfBqIsu1rgedj1wGYpuYNMTJtv2wpKeu2VYWZf1SSOmABIT3NlExVsWuzXbrfbDtDO5cmJVKjt54ATnwStLLSEj08sOR1Y5rNDZVJ8J8sWJJVbNTuYQMpZkZWuDcH5YqsbD3loG0jrpxaWJa0mrcg7364rNWjNmIK3ZlNyttsC8QVGOlSGVxmFAz2AHwk98Nm7VUaOdQeTVbyHrgickUyxGwKWHnc4GyxOZmy6940W577eeCDYQalI5VF7zXVssMjNEQqeIHZgzlgPkCo+d8SKpLT08gjvpIse+OcNpCaU1DrHqdtQKLsWKi/zNwd8S06oyMCtmt1t39fLBEXlUZIUBCrsispD6GUmw3AG/2v8APFryWR6mCm580YqHbSokbwWBtq9eu4+WKzBF43aSZEWNrhg4xPUkkFZS3AKHfRyQApJ3uwFr3N98bqZhoWOIMKYzBo5hEkKxq4sJDq20+lhYC9iDgWlDcy1gLqT4TYev19cR9JRyU0knNqC8UjBlJAsOgsv1H7cG3QKwIRSLAgNtvuRt/j1wZdKq83TfvMv9uT/Zx7AvMh/tS/7K49jOutqCqfG1bmtJMj00yRUypcjSDdtx3/xtjNf0rVK0oiq5xJMSHEZ023JHTtjSvaZG3uMEaDe5UeHbt0xm8GXS1FaIoaUzTOCI13JNgT0+nfGZrAQstZzhsglhWZGLiWR38TAkk4ko0UoF94nZQbkaSVFvLEgcizdI0jNBUuSBqtE1lv13FxbHanK8xp4yoy2YEbKFXVth8BJuEPJFE8Md5n0k/rCx2wfwtE68W8PBpGIGZU269B+Mm31vgIZVnUr3/Rc8hA6Eqtjf1OJXhvL80g4wyQT5bVKBmVMxIAZQOctySL+uMuLaDSfbkfuRsc7UF9yx/CMVj2o1maUPBdZNkuYUtBmBeJIZaiVI1OqVQVVnBUOykquoEaiMWZD4cM5nRUeZUUtDmFLBV00y6ZIZow6OPIg7HH5Vw9RlHEMqPbIBBI3kA+K6zwSCAsUqOIeJpuFXmpuLM7oswoOIKfLKqCoo6Tmos8sC6XZVZJCquSrrpvq3BtgriPP+JqSLjOoTjKSik4RhhFPBNTwH34+7rKZJvDc8xmKAJpAttvjUqThjh2jyxcspcky+CiSZZ1gjp1CCRSGV7AW1AgG/XYY5mnDHDmaZnBmWZZHltZWwW5U89MjyJY3FiRfY749MM8wOv9yIkn0WT80xt4OG9g7ckLPwXxv96yar4y4uPtBMcdbmNJQjOMupZIZKaBqOKOeGJnjd7c0SFnIUja5UEjDuXe1LMH4m4leOc1tG1HWy5PSe7lVWSjFrB9I5nNs77E2C9sa3PkWTzmq52WUkhq5Y5qgtEDzZI9Ohm8yulbHtYeWORZDk0cNDDHldGkWXsWo0WIAQMQQSn9nYkbeZxbs6ytzIOG5RaOgvtM6h64m9yFXCqDZyyGLi7iGgpI3fiuszKfMeGqnMkLUEAgSVIRIGhdACFUnSVcNe43vfF09j1dn2Y5VJV55VZxM00MEkfv60gXxKSTGKfexuPj32HriwUHB/CtA1YaHh3K6Y1sbR1JipUXmo3xK1huD3GH+H+GuH+Huf+gsmoMt5+nne7QLHr03tew3tc/fGfMM1wNbDvp0qWlxi+lo5+ER05zCtlN4cCSpW2MI9uubU1Jx0tPMX1DLYpBpB7ySj92N3x8y/ymcuzOr9paSUlJJJH+h4U1qwAB5s1xufIjfHR8njQc5b/K5TFGGKBm4gpCA0kmmw2DEAAed+uCIM6oZIFmFbTaR1Oskj8sUKTI88eUItExjjsAzTKLne56/4sMKnyHiBVAajcXbc89Gt+ePvulYQ8haFHnuWKVvXQkMLCz9fzx6fM8pfUWracX63lAP7bjGeQcJZvHKJ56P3hlBG9Ygvvt0bfCazh7NJF5ZyqSNCSQsUkZt9ifzwWiyviOV/fNcqi3FdC6X3HN1EfTrj0WfURVjz1dAbbI97+VrYoKZPmqgKmXVihQCAiqT9x1wPm2W5vR5dNJNFXJBEhkcuy9OvngQwbKjVIutETiLLiiMkotIbRkqw1b72uO2CZs6y1dMMtZSRyHfTJMnz/tDGFRZ1GWQmpmUhWBZ01Hdbbb+pGGYp6CPMhVe+t4GDpeAtuDcA7/LFmmDuq86PJbfWZvk011TN8tjZB4WWpUlW7bA26/PHhm1O8sapNHUrICRMh8IHmbbdMY/FUZLS05dMxhkOn4Vp3D39CVsD13vgGLOMvSpSRaSRVI0SBWAuve3ri+CCFPOyDJW5CroIZVd8zpQb7o0y3v8AU4JbNIOatOtTDzGRSlluLNexuOoOPnyHOpYUEal1SxFgbbE9MLOdsyhZFcr1ADkb3wIoBQ4vqvoKdlKPG5SxW39Xtfsd+/1xU+IeFxUibNuHCFrapTFKW8UbKSCzDY7m1un8cZzk3EU5qo4Yqmem1NYHWWG/pfFqgybiFYNa09azXIAKFbW6Wtt54hp6RurFcP5bKbo+Fa/kvIKCkefMMvFHUmdbtEyWAeIkeEsq6fk2BqH2fKixGfLJZZYc4UzLHMoNRSWBO5YBbC4uNyT2tiEaiz0MI2pq9Tq1W5LWv9RhS02dsNRpqgj9UvEfysNvyxC0kQq7pMwtSzCGDLUkio8umgy8EQ0sLyIWCjYXufLzPbFX9oXD755R0nKdWFLLrZFOoyqbAqLdOh/LFUhoc6q6xY0hqw5NjI0bAEHrvbyBwukpM8kiAX3kGNmQalPwgldvTbFtBYN0bnh4LSFvi+xP2aJpMGXgy6ttGZy7C+36+IXjimpOHJJMpyukiekhgMTQ+8NcxrETfe5cm1vMk4xmKgzrkokOTZhUxKx1GOnZQfLxAfXDyni3VBGlJncPLvY6W2+/XGiq7iNgBZMOw0XapJWmo/vOQvBNFKlQ6EagTqQkWLXFrWv12xXOAYhHmD+8U+YvNLSIWrKsKBKYzofSQLlbhbX3NjfFVpxxjYhYc6EmrxPzGQkW2t0H/HDlOnHBTRJFmg1oVYyTatie9z+zCW03CwK1OqBxkhXrNVVuPuHZEmSytIti3U6e3+PPFpzGG1bINviaxPUm+Mu4WyfiNeJ8srcxhlMEM1hc6rkggnqSBuT9MbJmFDza2k1kqZqllba2nxWw9g0kylO70QozgDREudurp+HUquknZbRJt+eHKyRZHJVo3N7k+WI5J4sunz2JlJAq1DC5bUeWgG3l0+2M+zyqX3p4VqJqfmOW5UR0Et5/X54zVKcuW1laGWWiVEfNDKyRuvlcg/ljsLhLHXGR35iG/wB++Mthh1MGeuzBZPgBNQdx5+uELRlSL5hVSAj4TUMCPXYjCHUAdlG4t/Ra2opwh1ine+5LMbj93fDZEfKJ51MFuCSbEXv5+eMwekvBIwrql1A8P4zDudt2sRb9uGfcadpnklmDtYXZ9TFW3Ibrt8PqPPtgDQjYpwxJ5hayJI5goaqp2UEFNDWFwQbm1vtvhymihaK8lZTc0C3gtZj8r7Yy2Gipjy9c7ao5dTAzMQV2sLX9D59cOx0lJG3JiqHMZsWDAHc7i3kfXEFEkXKPzogbLVqAyUPMFNXinWWMxyLE2lWW1gGA6gb9fM49+EilRUxsb/Do2Hp1xjy5ZQvK5NRUVDgksDNbU1tlIvtv3tfD8GQ0Fwju7+L8TUz+Jvo4tthow+qxKznFESQ0Lc+DKuCjz5q11jZ6allmIj2LBUJPfFnb2l0K5PPmH6NqvwMty7MGTUu61jsqKD5qVN/yxjfsayvKKXjtOYqJTCgqecSCC6FPECQb7C+/Xw4kMtyPh2Cny7ias9l8lJwrWS06xVf6eneoiidwIJZIdVtF2U6QSVBx8z7VZbg6+Y/t2kkNZEECbutdzZJAMRJRurVHNBW4e0igy7MuCM1pc3zA5bRcjmyVgIBpyhDrJv8A2WUG3e2Mu4Fzmbi/jfJ4eJOLjVmgZ6zLKT9CSUPvsgjZebqf4tKsx0rbrfoMaL7YaKLMfZxmtJLmFDl8biLVUVzFYEAlQ2cjsbW+uM+y/PK/ijjThoZhxl7Pqn9H15qIocrml94kJhkQoupiCCHNx6Y8XkbZwFQzzfePR7ouO6bu2MEQApVPfCrGY5ZJ+kquRauMStUSm5Y3F3JwXwxlVQOIMtlariYrWxHe5JGsYpslLRR5pWScqF5nrZmYrAoZjzDe58sSnCUQ/nVlBDlXXMIC3h8P9YPI+mPuFSgBl5/k/wClJZiCXj1r6oXbtjtsdHTrj18fl54XQXCPTGAe0vLmn9oGbyFmQNJHZrrb+pQdL+mPoAjHy/7YGI9p+eqZ5oy0kWgi+kfgReXTH0DybM1Zq/8AkP3tWbEu0tC5Fl1SlfyGUtCY9XOaQAE73UDudh98LnyWSWQFFpo9Pe5vbz6YpNfLmCgSQ107SMdAPOcW23IFvT16/UdqKqsWnkAr6gzafiMlmI8v24+5GgCszcRC+ushjCcJUMU0lwtDGruO/gAJxhfAuVQV+c5PDlVPxvOiHL6vRWVVIkAoYpCaaVrLcjaSyDxeE3tscbfwcyNwFk7TsWRssgLkm5I5S3N8Y1wpkmW0Vfw7xX/Mqlosjr6umhoZIs7qnq6cO/8AR3kjLaCpYr4ATbX33x8GyeoKbsWLyXW2ie9a7mybWjotOI7zgVqftjVH9m+bJLIsalI9TEgADmJ54wCkiomukNfTyMm7BZAbA+dsb17b3ZPZfnLg6SEj31ILfip3YFR9cfLizVBmmD1xmklZSpMsbkAAg/1YFhuOuPbeTKiHZdVJ+n+AScTWLHwFt/sOo4ouKp5FnV2NGwAXpbWu+NG4v4pTIJqKigyqvzbMq7mGno6NV1sqW1uzMQqqNSi5PVgMY1/JyOr2hVbvu4yx1B1Em3MjuP2Y0r2l08dXxNw3T5dmOZ5bxDIahaSqo4Y5VjgshmMyv4THfleuorbvjzHbDDUqvaIsqC2mefJpN4vHWLwnteXUgfzurFwhxJT8SUVRKlHV0FVSTmmq6SrQLLBIFDWNiQQVZWBBIIIxkPtRpqefj3MxKoLMIQDqG34a+uNB9kK0yZdnCtPmdRm4zNxm0mYxokzTiNNJsnhCcvl6bdrd74x320QmT2n5uUldCBDqAJ3PJS2HdicLTpdoKlOnYBh/6dvDpN4hA+qW0w4q5exSiipuKql0mSRjSMLA3IGpeuIHJyrDKMkbiriduF4swp0pKqXIIUoZnjnVo0Ew8di6gBzsT33w7/J5I/nxUqWdiMtfUT0/rI/78IoK6kWTIsh/nBndVwjz4KmGmXKUVoIRV6KbnTar8ppEXSQuoqATbrjtZ3TPyxWDZ2pnaYs647ru99HYG90L36mNPr+9fQotYeeBaipgJmphPGZli1tGGGoKeht1tiv+1CllquFXSLij+bemZC1WSQrgm3KJDKwDEgeFg3lipex3J3yGr4qy6eiyZqpKjXNW0VYZnOpQVhkEn4qlQNQ1E31Gx64+Y4bLadXCPxJqXb82PEczA57CT6k91Qh2mFm9Fk9KlKi88GyC1iPL542b2L00VNwi6xSGRXqnbVf0H8MfNuWSSJDD+MSghB8TGx26X/fj6F/k+qi8ETaJuaTXSM3mpIXb7W++Pr3lAphuSz/E38UvC1JJHgoHirJcr4pzjjTM5ciyJqrJTFTxNVzSp7wyRLK5qAjAGMowVLg9CTttjTuFcwjzXhXKs0hpvdY6uihnSC1uUHQEL9L2xintWjySt4/rzmFbwg9XScsJDU5HVVE7XVSsTNG4WZ/ECEsSFN7WBONv4clmn4dy6eoULNJSxtIBAYbMVFxobdP9U7jpj5rnVOMBQJmDETqgd0TE2MkTyjlZXSPfK+dIspppZJEWpj18zoJPh3+fngr9CUahnaSMuPN9sViKtcAB5rXY/Cd23whquaW6+9FTYEAE36+ePvWHpRSb6ggbXB5LavYxQx0mY1rpIjF4VHha/fGpgYxT+Tw7nOs01yMxNOmzPfbW1jb69fTG14+D9vGxnNQeDfuC2B2poK9YjAWeoHySuRujU8gP+ycV72ncXjg+iyyslalipZ6zl1U9STojjWJ5GAt+u3L0qDtdh8jXeEuLOIM9yPiWk4hioI6mlyuKq00auvu/Pikf3eTUTeRAqkkW+MbDHHw2TYh9AYwegCOd94+/83CU6q0HTzWfU2V0MdNHoijuFAJNv34famp4guvlcvcgErb+OKNBWTmjj/HlPhAuW6dscWumZS3MmZiCGs25+dsfpRrICztqLQYaOmY8xkjNhYC+1vTG4cLqE4ey9EGlVp0Cr5DSNsfKUVS5KoZ3vqsSX3H3x9ScDyW4OyXmPdzQwklmBJOgXN++PmvlLZGFofzH7lppVA4FTZ3Bx878TVnCNNV8VtNNVRVNRT50rwDNkXlLDOjSxRApdDVMST1OxtfH0OGVr2YEjqAemPm/inOqCXjXNKLMJcgy6KbMKqKaZuFveZqRYmsr69JEjSm5JINrY8X2RpuNWqL2ANifHoCT12vEc4KcSRAWh+0flOmRMnLgU0rsiFtlX8PYfLbFMeGjckvLGxHm/b74O/lAmMy8LqkmqM0k5EhXSCPwrGwt18sZNXqsiFJHv1VQd1J7XBHTH1jscycopnqXf8xVmrBhaNl1PTSZ3QGB4pCtVCbF7iwcHzx9ArPFb+tT/aGPjdCtNHpCqCAAGUAC3yHbHBHFzuf1v1YCwIHlbA9o+yHy2+m/i6NIPKZn6whGKAtC+yjUQf56P/aGOCpgPSaM/Jhj41k1X1RO0e9gyeEj69cNUlPDDGFCmToxJJa/Xrfqd8eY/VZ//p/0/wDkqOLA5L7R50Vv6xP9rHOfCNjIg/72Pi80dCWUTQI8gUHW9iWA2vv16YV7nSx6mGoE9tQsT54s+S2f8T/p/wDJTzsdF9nc6D/Op/tDHOfAP+lj/wBoY+LYoFZtba2YC2t73tY2/PHEp9tLi7DzJOK/VbP+J/0/+SnnfgvtMVFOdudH/tDHfeIL/wBdH/tDHxfGUDoOXdVv188eqJhLTvAxkVGFm0Erb7YL9Vbf95/0/wDkp514L7P94gHWaP8A2hj3vFOf+miPf4hj4uh0oixRIkUQW5AJ1P6k98Oh0GkKIUAB3Yk2xB5LAP8AE/6P/JV534L7N50PaSP/AGhj3PhtfnRgf6wx8bNdlBGlnfZQq9WO1sXmLLP0blNJRyhHZowZdQDeIjcfLe2BPkuAN8T/AKf/ACTG4ieS+j2qILbzRf7Yx8xOoq6OSn5jR63kUMRbcNfzwfNQwmEvEiszE610AA9PufXAU1PHHG6KLISFIv0vt9Met7L9lRkD6jxV16wOUbT4nqs9arxLJqVjSu71D64eWfGJbFm20jT9Te4w3BBzjG0kjMgaTUSeg1m2K9Mkz5lCJZtahraf7IHz+WLTl8azMutmBvbR01XJbbHq3PLihptAKHzilgDg000i2FgF/WPr54AMFWrc0cy47WJ/di2TUasAEqVja+xXqMCz5fmBF4szhQAX8VPqJ+usfswkMJ3KaXRcBTnsMM/8+7SNKR7lKDqBtfUmN73xi3sagrKfi8JVV8VQXpZSVWLSTum/xHG1bWx8L8ojYzf/AIR+K10zLAsU4p/m9Qe0ri2p9oGUPXUlbQ0yZM8tG1QjRLGwliiIUhJDJvbYm4OBJYsky7ivI6/jLIqyHhz+a8dPlsOZwNViknEhMkcmzWlMfLFzuQpF8btYH544QCNxjlM7QFtNrSw+iGmHEbN0y0RDSdybzfaSkmjfdfNtJw1nOZ5ZwNQU7V2UN+kc3qskaVGD0cQDSUgkB3C7KNJ6qbYb0vm+Q5tX8SUtdktUOMROz/o4VlNTTrRxxsZ4z8UJbWAwv+qdsfQHE3EWRcN0aVee5nTUEMj6I2laxdvJR1J+WGaDivhuvhy6aizikqY8zkaKiaJ9QmdVLMot3AViQelsdVvabFPp6xhyW3gjl3i430kG28ztPMhBwGzGr87LDc1oKbNfYZU1UvCuXwz0mfxRUs0FG4SohNXCXmhSQFo43u11Hh2PY4L9qPCVRWcZ8S0GQZFkXu9PwnFy4qihuE/EqCRT6dkk3uDbqRjX4eOuD5uIP0BFxDlz5lzTCIBKLmQdUB6Fh/ZvfBR4o4eEebuc2prZMCcy8W9KAuolx22BP0OA+XsdRqam0HC5cAZPpFgHLq23i60K+Cwjf83WNUWdZfkWa8QV1WtfHR57wrQDKpHp3keoaOGRDGSq25oLLcG3XHMnoKHIcw4Rf2h5Uz5RBwdTUlOKmjaeGnrFtzlZQpCyFdIBI30kY3umlhqKeOohcPFKodGHRlIuDh2wI3xjf2lnU3hETY969m6RBixjfefBWKHiqp7Jkok4GoRlmVZhlVATI1NS1rEypGZGKnc3VT1VT0BAxM8Tlf5uZlqbQvukt28hoO+JMW9MRnFaczhrNI/7dHMNj5oceedV4+L4ptqdO87nqbn1p8aWwvl2XMckipY5p8yjAMPgBLHVbyt1OKFRcfRw5zJBW0pkpL2jKsQwPYn0OLxS8M5RS0cDvA1xGouXJIBtt2xTeP8Ag+nly6XNsrjImgY8yPfxLfy8x+zH6npMYBZcyrVqRayTmPHsdFnbK2Wc2gNtuZoYAgXt64VxBxtT01PBLl+Wa01bGaW4YE37b4g8sgoeKeGzQkx0+b0K3i3IWVduvXxbYZ4QqKOoin4az13gSW4idiRynG9j5XIw40mys4rPOxVxoOMKDNKaOaPLuW6/1iXvpAB+e2HBnE5hj5WXUwVviYRsxb8rYpHC1QeE+L+VmiXp3UxswJ02JFn6bgWxsAgopmjmhJkjkNy4lsCOx67/ALsU9gYE6nVe4RN1U6ziKvpqc1EFLGrKNyackKB1PTAVRxtmFVly1NLMqLY8zVD4VPc3ta/3xc8ygysF8vqJ3/pimJkaVhe4+e2MfTmcOZvWZFmoZqKe6TW3sL+GVfUbH7jEDQUuo9w8Vf8AKuJM3zeFJ6Mx6GBvpIOk36G5HTB8D8RSnUJUi8iApJ9euM34TzU8K8TyUlayVFC76JHRrqQfhkUj0P7cbbR09DNAk9LCksLjwMshIIPl6YhYByTKdUutKg0jz0x6jVG7P1DxkD74RrzCnDuGVpG8ILyr1Ppi2xUFPENQpydrWvaw/wAeuEvS0uk64pVA6fiG/wBd8CRGycGzclVqDMJoo71lIroosXjbUx6b2B/xbErSVFPLTqsJsum5Y7H64MSnpCrXpiRe/U3JtiGp8uoZaeXla4nM2o6DYsPI+mBumetDZTS8S5jm08zy0FNlSElTzCXZegAsfTqfzw9V08X6QdaYKovcWbB1OgoaZ0hkYIo+FVvcD5998R1VIq1qSKSb6QO1xbvipUhJnikJRdZe5IK32XAIjtMCu5VtrfI4lpAnI1RhRGexG+AGATVL0Ykix7W/44ZErOHkSEzVsiIZFW5tY79T54YywymSo5YZZDHpDhdVi3hFx3te/wAhjtUVNMyE6XPithOXLztEKvGBIyrJ4rEi5aw9bj7XxIkqyYVvyBAuXxixEZu67WPiJNyD069O2C8zlaKjKq/9Yw7DD1JHywqW2A6jEZxDISxRCSwsNu1zgiLWVg2QFISZDOY0Mh2BIuQCOtvkcTmXJHHTtrZbFt9GxtiEoqlUkYFzGok1qY7A27dQb28sTZQzQa4XQqFLWJ036322/LGqnZqzPF0ShtKWZdwvhVh59/LvfDwF7soDMNwO3p1wBz0lqAoACpsSVvc3viTp5Y5IHZfAOh3JJ69Rb8/XBgobqK5P/wB2j/LHsHf0LzT/AMYY9hN10oKqftDjc5cGLBwm4F+tz0+2KNw9RCbNqaEkN7wGVpFfSVUqQbeu9sX32hSOuXRRvpJeViVB30gf8MU/J4VizSnqHNl5iqtwNzeyj0te/wB8Z6aXVCu0zR6WWMaY12Cg7EdsDS35RCjV62FxgiMELput+mGhax7b33740EpYEr0YIiuQGY22wqnmaCup6mMXaGRZV1C4LKQRfz3GOF00gat/M4ZebV4QG29QBhNRwILSmNbBBWiL7VOKjYLBlpt1PIa3/rw8ntQ4oLeOHLSAO0Lf7+Mtp/fHqZo2MZQC6gk2BNwQbdQR5974KnhzERAUtTcjpqFlHqR322t8seYPZTKD/h2+xP1kc1pUntQ4mij1SLlY8gIG/wB/DcXtT4maMsUysgmykQPY+nx4yzMKfMY6RmWdQqKXILaSbdVuRYDvfz9MD0q1oYpTy3jfczdiouQe/Wx38vpgv0SycifN2+xKqV3gq3Zx/KPr8tzKegmpqYSwNofTSlgTYdDzBhgfym5yQORCf/yI/wD73HzfnU7VeaVlW+kGWZ5CF3FyxOx74VPlWYU7tHLAVZWKkFhsR2640s7G5RUHdwrT9RXNqY97D3nwvpA/ymakAHkU+5/6m3/73Hn/AJTc4+Gmg6b3pD/+9x84DLqwtvBbb+0Mc/R1bfaGx/1hi/0Gy3/dB7CgObD/ADB7V9W8P+3PNc7gM1CmWkK+l0eFlZfLbXiI4sz6r4hzb9IVkdOJhCsJMQIUqCxGxJ7scfP3CtdX5Hm8M6pJyWYLUIDsyE+XmOoxu1LTkRBtJsSR9sMw/ZrA5bUFajQDHbTC10cUMSIDpQKh9V5Lg27E7DCw4U6tTG/QX6YMcIzhSpUdye+GJodLaXA36EY6YMpjmwmtarEVI3BubnHCxJ2AU9b72thtrbhmJtubDCj4N3U2P3wW6EG6UkhQ6g5IOwOkYq3tPrDT8K1ManedkhH3uf2HFouAFbSQo8zio+00GThWo21aHRxt8O9r/niwyDKB57pWSK3Z7WG2EkEk2tYdMKubgm2E38JIIwwLEkk79PEPtjviBubY8bt5DHN737DFlReuxG4GOC3TfCiSTcjpjh6fXpigolLdbA9e2Nd9mPFQr6VMozHS1XCtopH3Mi7/AJgYyE9Pl54Jy+pmoquKrppuXLEwZWv0OIWyiY8tMhfSLTi6qp8B8jvjyysztZTYbWGIPhXNkzrKUq4FuL2kFvha24xPUwNizMQb9bYXcLoAtcJSy7MLK1h9sI1Pq+LTY9z2wTHYKS1m8jbA8qMSQEYkG97dMQAlWYS2lcR9WsPI2wgu0hAViB+3CUMltNvocKRHdiQi2A64IPhVp6JUZOr174I06/D074RFDqFgp+YG+H1gcRgkHbr5jDBUHNVoKeyuP+nQkWtrG9vXFkqNbtRKLl4pwzsTvckbn9uK/FCsqLHo1AkeEm1zfbFvoIqZWq4Z41HOK6Be1unhPlbDNkJUTxtlVKpFUYLxzllcqd7qdiPpilVHD2SP45JazVY2IkHfqMaZxVThMkhW7kvIQLi4A+fytik1NKVZLC+2wOMtXUTZaGBumCs9zTKqWkkqKmCrqJIRcgTzRrq26W5Z+mGJhR0tC+biCue0LNEObHHqAUkKwB1C9rbA4ls8oqr9IPUJLZuWQItFgDvvq3P5YVHSOahZYaOeR3Q+NbqgA/VbT1722+eB4g0xF0vhEOkGyrvGFPn8ee0dDk2WwyR1cOuLTZjqAuykk22wwvCvtQni5scMKAXAHOhUiw7b3xdcx4lyjJjTx19QkNfEgkQMt2QlSOtj13GKxPxvmVRVmPJuJoooI4Aw5lAkkpN1B1NKAL3Y/CbbYFry4xF1ixrK7XTSIIS8l4ezjL4nn41SSGFSqx1K1sbFSW0BStm6k9emO5xJTZbQ1lVTSznkVAp6OOaUGWUgLqP9VYqCxA3F+nY4RQ5tmFdks+VVGYLmE0d6oSMFdtQMZGlVY9C7Da3QWwrL1rG9otflFRU+8UdNXyQMI1IEUR1gsLW3sP7QN7b3FwbQdZB5ImueaALjLlGZZwnmXE/HeRcP5bVQZTV17SMHdjdHSPm6nUKpFxYC488WfiyipOFc0zLh7N5Ja/Ost5UlRJDmMkELowQlrldiOYotbz697RlNa83Fsef5DNRSLA00kdXFAqsJm8JGghxYIbC7d+mIHjfL2z7P6yrzatFVXV55lUZkKAxqI7BdAAuBB/5idsEarZhtyro0aoaHVLKwewuhSo4wmo8wipyaunlippoqz3hxC8WojUQLEA2J07mwsLY0PNfZxw1lOTQUmccfcTU+VxGJI6erzdFhOgjQgUpY7qtgPLGU+wSSPL+PqinnLZXBNlk0lPEjNIaXSo1OWO97LqAO9iB54t2W8AZ9R55w/wAWQwZfx1RoxqDVVE0iVlRFLGQjFZ2ZPDqDALpO3QHHy3tWHfKhPH4Y0tIEC570d490HlfrstdD91fvXPsWzcawSVPC9ZBDmdHlkjqoWqq4Fmii8Q3ZGIBv03PUjGXZRPmHD/GvDtPNxNwrxLHmdW1M0NHlcMFRBaJ35qmNm2GmxuOhxf8A2s1lJl/s+zSqrcrps1i0xxikqP6qV3kVU136KGKknyGMuly/ib2b1dJVQcMcF1ldm+uhoTlNAYJoaoxsyAlm8UR0kMbggb+mPEZDRNTDOa6O8XAA6RJ0ie8biJBOw6XWmq6HKHrMnyRMwnT9IZlEefNqW8e51k7X3t1+mD+GchyWTPsrnhnzIiOqhljbmoASHBF7De5A+mKXx1wdX8T8b0tPTJEzSxVSykQvIqsGBZlVAWDHxbm42F8Snssy5csp8jyuPN55K1MyVZ0sdEKicFEHnqTSbnprI7Wx91e0/JxJPzD/AMqx03jiAaV9SK22+FjfGP8AtdpOIqniygrMuo67NcuSnMXu9Fma05icrMkpcF1BLB4rNvp5bdCcaTwXDmFLwjlFNm9QKnMIqKFKqYNq5koQBmv3ub798fm/F5Y3D4WniOKCX/N5j3+2w3tK6QeS4iNlLt0xgftMyHLqvjjNqqWqqUlkePVpjUhbRoNr+gxvZxintEKfzyzFNe+uPYf9muPV+TgkZq7+Q/e1DVbIEqjNwpkIbSazNFAboEQm9vP6YWnDPDSs8LzZqwG5JMY7fLEkdN7m3XfptgWqdLsTcmx6d8fdGm6xmkF9B8PwwwcK0FOitJDHRxxgbamUIB272xgnC68L1/tCyTLMo4Qy3ht0qPeNOaVM5rE5LoQvJLBUZr+E3cbE22tjfeHwTw1QFW0E0kdiR08Ixk3BUfFXFXFctJnfE1Bm0WS1cc4q4MlhlpJSrg6EmNikotuFB0+fbH5/yqpwn4x5dAE31OESSAYAIdc7H4rViG94BaF7WI+bwBmSbbiP8pFOMBq8tpamZGqXqAq3ChHCm5879R6Y+gfaoB/MLMt1G0e7C4/rFxhxaMMAQD3Hlj3nkxMZdVH8f/SEjENBdKt3sNyfLqDi+qqab3vnPRspMkilbF1PQAdLC31xZvbFSZi9Rk1ZktFmxzOFpo4azL6unhaIPovGyznTIHsDYC4KA7YhvY0UPE04BsfdTt/3hhf8orKWq6XLcxkzDh6jpaaCrp9eb1DxiOeYRiOWLQpJlTQ5HcX2xw+0jQ7tU2TA0jlPzTaARM7b+tNiKAj83Vg9jUTx5RmXvVNXR5i9ez10tbWwVE00vLQamMJKpZQqhdrBRt3xQfankuVVvH2ZTVQnWRkiBKubEhFttqFtu+NA9isKQ8GoaeDhiOmke8MmROzxTAKFLuzKCZLgg3v0GKH7T3Ucd5kpVthF0HU8tf7sX2P73aWrH0T7tPr9k22QPkUmo72K5TltFxZVVFLJVM5o2is7gpp1KenW+3ngvKcs9mNNQRzUdVmHu8s+X5joaSZtSNM0dJGQf+iEtyI+gIBO2GfY0wfiaobSy/0RrXP+kuKvkuTZVHWZZm5yfiqPheaupqWlrpM7VgwSo/o7PT22h5p2321XsL43doaQfm9aajm2p+iQJ3tciTbuxfwUPoNgDn960/20vTU/DuX5lNmWV0UlBmcVVTjMiRTzyKGtG5AJFwWINjYqDbEJ7Kc1i4l4mz/iKXMcgaumooKZ6PKZ2mCIjSESSOVXUxLkDbYDFg9rX6Y/R2StklXTUswziESSVJPJCFXW0gDKWUsyDSDuSMI4Dr80rZs3ir894WzP3T8J0yeNleCTxXEl3a3TYbdDjxOHdpygkXJkT0EiRGk7+JHgjd+8WQQcI5ElJThJczTQgAZqgHSbWva3XGv+xeggy3hOSnp555kaqdy0wUEEhdhYdNsZdT63hjJfYqD06+uNa9lJvwy2/Socfsx9K7fz8iX+k38UWHAn6vgs645yispuNMxzKmg9oEUUVd+kBJltJSSU4l93ERkTX4z+HdbHvfbGw8MVENXwzltVT18lfFNSRSJVSgBplKAh2AAAJ6kWHXGT8S5lltdxtmGVZrxtnmURy5o1BJTx5vDDHHEKRZuZpZNSqxOjr1PXtjXcjp8vpMioqXKTH+j4aeOOl5bal5QUBbHuLW3x80zw1PM6DawvAi1tMdevqVUgNToXynzqR4ah4cvz2X3chpzCoZVQvpLXZFba9yApHr1IkcopMnzSnnqaJ80ZNEjREmFOaUYKAAb21HoTa/0x4pmOQ1+YVRyzKpMoOWVM0lTTRvEyIjauW10QSEH9WxPWx2N6jJPT53lVbU5vPXsKZKWbl0T6HUxU6xmyEDqzDoeuo4/Q+HGum3lYfcsDqmkdVuvsa4ezLKarOZ5aikklmo1SEQVXNYNdj10LbqN74C4dzP2mz12VZXWZTmsbpJlC1krZjTsUWFv6W7BX1ESA3O29sR38mrL3pq3PZabKuIaSRqKMCTNSw1PqewUEkdrncdtu+HuFuEOIqJ+FmX2dLlucUVbSy5lna5nC8s66gKpnsdThwXOkk9Rbpj5Z2gdRZm+JDyzZkaosdJMiXNtIExJvstDSXU2m/P7/AFLUfaU/E65DGOFsuoq6pM6idKlQxSKxuyKzKrODpsGYDFR9mj11ZknFqZ9VZ6M3lMoqYq6kEAjhPMWF41UBCWQbkM26gE7DEx7Z6GvzLLslpKOmNdE2Y66mjXMxRNUIIJbAPcE2fQ9h/Yv0GIT2ZUtbl/CeefpnLxR5w+XJLWZj+llrZawlJDzSLnlrfUVHQ3NumPIYFrG5Qbt1EjaNXpc5MxbYDpveGvJ4ngsrq+EcySeppsrjatpKFYw9RUV8FLrLIHNgynpex3O+22BoMmkbMORWxUsCKvNmb9P010QfEwXR4rfMYlK/hTJeJ8vjrMwp4zWzQRl6ogM0jBRqZW0i41XHQDa1rDGR8Y5HTZTxFLRSU7UtI6JHSzzBVDOdJZriwKgFgTbYjH6JommWQRdcmq2sw62mxWx0nDfD1fTNNl2eV1RGTp5kAidflcE7jF09qeWtScEcOyxZ7lmXRQ5ZNlZlzBZC8nPhRLxrGCS9kJsL4zT2GOzcMSxa9axTkqHHTUqsQBYHqxxuWbcP59mcHC2bZDW5dS1eWQSeGugeVCZEVbgKwNwARe/fHz/trWbhqmGe52kBzr/8Jtsd9pgxMwurS79Ekb2SfYiktRQZvncmaZRmAzGqjIbLuYEj5UMcWhhINSt4L2PnikcX55LFmHEb13GHEmW8V0tZOmSZTSq3JmRf8mAiCFZhJ4dRJ6kja2NK9nWSZplFXn1TnOY5bV1uY1iVMq0MTRpGRCke6sxIJCA3vvjOOLk41r6viACozqGGglrammmo66NBPKWSOljjIYaYolu8oaw1XJvjwOAfSr5lWdqGmG3sOQsJbcfNNhO9gqeCGBN/ykYszrs24JooZIqGsqzNHPUTNangGlWdWNid9OxHZTjNMrpckzCCOeozWsoqhQ4fl0TTwqEdlJDAbBipI72xqH8paShTL+DXz6kmzKH9IRmphpD45hoOoIBa977fMYzjOcz904fparKKCRamrRkDVOrnSy8wgyMCSSSxYkX2vj6d2Mh2UUR4u9XpHZJqEte4pMOX8P1lcsL5hm86Ot41NI0Bsq6mfxgbW77jbBTZVwslQrrmeYsim0gadRYfSLf5bW88IHCGa1WTGWt4kqjVTx2lP6i330gbdOmAJeBmQJG2fTjw3FoluT5k98ez0eCQ13JTS5Twm02+Y5gy7sLVAt9wuHRl/CCkLLUVTC17ip3J8yfvtiA/mK5F04hqxc3bw7E79r2x6HgxomDjP6m43BMSmx7d8L4bXbIuIW7qwDKeEXnRkzaqjDX0RmUEj6E79R2wA+VcJmUo2d5o8rOAFUgXFxc7ntfp6YhxwdOjBoc5nZ3kUu/LAIXvY9bbDbBEPCEysSc41bAAiCxHqTqwwNtCEOkoqPKuHxFeLOKpUYi7yaPxBYG62bzI6g46KHh2NCi53XzTHbT+GtmHX9WxHTy64jKjgeo2kPEB5Vxe0Q3/ADwqDgtIy0i51OGYg6uWoGLjwQkkGFK1OV5Jzg0FRmLLoDEcxRfcg9U2+/0GEy0nCqwIY6muV5nCKXqk28zbRviGXgGCVCZOJKtiNtKxgfTBNP7OaRSS2fVoBtv4dh5YotlHcKUy6h4UVo1mqq6aMwiSRTMV1E2sRYDSCSbDr9jgfjip4Syrh96rJ6WWpqo20aZKpiE6gkjr2w03s9VGe3ENTpZepRTuO98OLwBRiDSc0qZlHiCkIRq33Nwbjc4EUwFeoxCd9leQ5vmWY02Y5zEkVJTASrGCLliLpsOlr3+2L7xAqtO0gFh1BwdwbSiMVEqltGkKWFhc2/hgfiVVErKLg22N8ZxOorVA4ahhLqtGCdt+uAMozBM0asVI94nKGwP8PPBSAAHws1upxWuCpXjqs2Ro2VGrtSswNiCbbf7N/rhsyEjTeFyup1jz9rL4WUm5HUnEplStLVVFZdRocQQjrpQAauwsSf2DfthzM0jM4nAYt4d1F9t749lRkOWU/PfmM5dlYpvpLkqDt2Fh9MKBEpoBAUhGDrBDAAjsMHo51qu337YDgJkljXl7gdltgqSVIFaBChPVr9RhbyTstFFpNyrd7Lgh4p5gQAiBwCPmL41i488fPlDXT0rL7jLJFIfCzQvpIB6km48xicizjNQ9v0tVvYf55v44+d9p+x+IzfF+cMeGiALzylaWgEALaLjCS3ljG5M6zGaIrFmdYrja4nYfvxCVuZZ/GxkXPcxA66fe33/PHnf1cYsj9832FQgBXv2hUuY0PtC4c4vjyesznL6GlqqWeCkQPNTvLoKzKhI1bKVNtwDiEpsvzWvz/hXOIuDnyOD+cNbVTRKo5ixvSyIs8wGyuzEXG/bFag4gzpSobOq83a1veXP78OvnWcSa2izrMl8P/WJO31sMd7Ddk8dRpsYHsloLZ72x1ctvnG6zPpi5lJp8l4hf2a5T7NBwlmMWc0eYwu+bNGvuiCOpErVSy3uWZQdrarsRhvjDgviYx+0HPckyuoavrauqo3pipH6RopqaNAyebRyXZbdfGO+I2HibiaKtkhqcxzNoXP4M0dYxW219VyNJvfpiA4nreNJcxjraDiniAQAaJqeKtmFz2caW87A9PPHXw3ZzNBVlj6Y1GTIcQS4tJ32B0i3LfdYa9SnTpF5BML6m4bSSHh7LoplZJEpIldSLFSEFwcSFxj5u4UzriqnyZUr84zJ5bkgy1bswF+5Jv98SycQZq5JbN8wRRtf3l/448livJnjeM4muySeUrXh8S2rTDgCPWt8uMR/ETWyHMCCBamkO/bwnGF1PEebAFEzvMLXPi96cW/PA6ZznU8MiVGeV7KVuFNVIQQRax33G5wDPJri2uB4zbeBTTVBsoxGLUCgoxQKtx3PTEFR5tTDOZ8pa3OYsUW5LAAd9thbcdbjFiy/lPGsQGpSNI9BiK/R2XtxLPVpTu9c9NYOSdEaLIGOw31dSL3G3pj7IwwFjqMkhZH7ReHZshzMZhRalpZXLK4FjG172/fiFqHrM/Ke65ZPUVqITM8ALmQD9YqBtbzxtPtBhgPBlbC8KunJZomKiylR0Hra5+hwV7MPYwKaA57mWfTZbVNKi5aIU5nNEiAjmDYBSrbi97B/K+NjLrnVW6TZYPBVVeY0EeV+7TVdQJB7uyEsyixugFrm5tiy8B8V5tklQcnfK6nMJC/LipQxSRXvbSBpJ69rYvvAnsyyTjLiyfNafPUynK6eJ55hTMGeGeNrMqA9VOzqfInyGJnMfZ5kOae0ueeOuWc5bSe8ZhGH5LySsgMJYhrI5JIIDbMB64I09Q3QNc5pCzTjiOs4loHzxKKro6mgbl1VLMb8sHSVK3AN7EHp3xXM/kOeUMecJqNXAixVyE9OySD0IsD6/PGrZdmWT5Zx3nOSxTvV5bViJqaqYF0YcuNNJPcBhpVj1tv1GM14iyyq4P4hWppmSpo5k1RswuksbfqkdLEf3YXYGEyCW6lUWJYC+9thjTPYxxM1LUnI62Ue7uC9MWPwv3Qeh6j1HriCyPhnKc0zKZ0r2ioWp2mgB3cuOsR22I7+lvPC8l4ZpzndRTTVjzJBFqSSn2Bfbwat7EX3HnitQEyn0sLVdpc3Y2C+g6WXUoZBoWwNz1OHiVceMg7fbFd4YkigyzUteklHCi8uod7gjSL3bps2pb+mJJK6KeejWlAq46qXlmaFwyJsTqJv0uLfMjC/UtAF4O6bpZ1lM0SW5kTFHC3Fj174jqSOQhmsytcAgkG2w2xMVasXBhAuhtYk2YX36YgaCpiUSSh1eJmKlCCdQGxBv3+eITKMCEmeWpNZyQloyDdtQuRe22BswcJmMdoyBtuTewsMLq4IXmWOnnkLHxfCRsTe3mALY7nhC0sTAMp17gC+wwk7ppsE7XKDRsni7W3tgLlmKk17yJ8R88GTkHLCbn4O99sDzBTl8cgJ1XItb64c1YiLqLlZYZgXJZNJ6DcG235kfbBORRGTOKdBGAkUZY3O6m1lI9PiH2wPMyiVF3A2BJFwTfcC+JnhGEPmNRNdrRBYlUmw6Brgf94fbEadNkZbqVrJMdMW6ab3xTpqjn17m+pYhe3mzmw+tg33xaM5cw5cR1eTYC+KrEpjGphp1Mz73Bv8ACvXtt8u/fFtsVHC6cdyPg2KnYDv9cLNSfCmohgbBBuBhipVQ6AWdm7OSRf5YYhmdW0eFgCQQXB027ABf22w4EpUKZpJmDkiQHfoNr7dcSdHUC5VbWPxH0xARMWIZAQQepH8cS1K4MVjfV022wQN1CEzzvUfbHsd5tR/1pfvj2AW/iBc4zhi/RBktcgg6tXcj+7FGy0075pRIyOXMqMCSNrDr/jzxeuMyrZEI1UEkgAdSTbGf0AijzimOrQeeFBZb3AIvbCWFLqwFdmBD2ve/c45JGvK6kW9cPSxFiTqC3P1wqSOyBgTYeffDHutZLYLoaOEdRc37NhUEDXN9x1NsPrGTGSo2PXHII1QMWBC+XTCZumlVHPswzCPinLKDL6p4UkmY1CuQdSrpsCO3xHcdb+YOLtNGoBZTbbbGW8XRTRe0yA0NTyS1MrhyNY6sbW8tsaWgHJgeQtrkA2HTc2xsrUQMKyo0Xkj17LnYauXYypScbQCB03/om6qNJ6JkLurutiUNiLjzwPmkzVmYVMsdLFRwzPqEMPwoLAED52wcItBIfw32t3whkAYjqb2NhfcHHOa8E2XUqM5QvnHNYZKfMKunYBTFIyMB0FjbF84ge+Z5oCALVs4Hy5jYqvHUApeLc1jDagahmPn4vEf24u3EOWcvO8yjM5dvfJrm1gTrO+PV5Cxxc/SvD589rQyeqgorF737YVHoUENv5YMhysxxqskiqQOibj88KXKxzI5VdHKMCRKptb6d8eiFF4Exded41ImNSAG1NJcAgnbG/NCOY4B8BYm/Y79cYrHlscghi1ModlUkC5uTjbq9DTvLE2q8bslvkbY8/wBo2lrKYP53Xqey72vfUI6fBRNYrh3suoA3I6beeBpX5gAIUE/XBVQ7Ptt0+WBZvhA6W8u2PLheoM6kwAmti4B239MNyggg6ma/S5wRGEOrcgeeGSgJFuqnBCeSrSmGVr3csRiH4op5K3h6vp40LOYW0herMDcD92JsqwOoG+42O+ENGQ+siwO98HNroC2bL57cEPpa4IJBFrW9MKpo+ZURRWNpJApt13OLn7TOH3pKgZrSxj3eY2l09Efz+v7b4qOWkLmtIz/qTISfIahhjY1BYKg0ypmfh6FZGVJZiFJALW88cTh2AL+JJKSd7rYYnRKjzO17XJP54QqvLHzhOgWIXKkElje1hj2Yy7CwDoC8Z8o4rYvUN/N+nVyHkmI7Wt/DDU+RxcrmRPIXvbxWtbFjMT1NfGiTKi2Je467YQImm5aRyrEvMGtmW4A7m3ninZbh7jQo3M8QCCXqClyCK4UTSAkXJNsM/oGbQ7xy3VT3HXFiq1dpVjQqJCLAncXwqZEiaWOOYuit4W021D5Yt2VYUmNKJuaYgAHUheBIM/y7PrZaJZVlUtNDGoZpI1BZgBvvYG2x37HpjdOE6Giz3iN8oTN62CKopPe8rnZICtSgU6gbHZgykWFzbcgYzb2X1DUfHmU1MY1sjMxDGw2UnG7TcWUiJCafhvL0kpmeSlmWZzyZGTRcADcaTa17bDyx57NMPRwtRrW816TKa9bF0y48iqfyiFsfE3YjDcjEMz2NgOxxE0jVr5ojkPAmqdXjm1MzjV8S36C9rehGJORWYKttuwA3xzS9puF3GtITkClgHt174Kp3WBiSoIJ64SlOeVYFQwO5vj5yzDPeI4K6eCTOczjaN2Up704sQbWtfCoDuaJ9XhRZfTPvURcBYxvYg6sEo6ObNbp0HTHyzTcScQQ1CTDOK5ijBgHqHZTY33BNiPTG78F8Y5TnWWpO9VFSTgBZY5WC2fyG+48sXok7q2YvVaFbqdNMoKbeV2viyU9NJUVcTaCraVY6l8wOvpitmnMMUckcyzahq8BJ0+h2Fj6b4vlA0cNDRSWLvPGi9NiQtz+w4bPRASCjc6ji/QMaTaRGgAJ62NrYpEhiKgLZgpIv3OC+I8xrK/NI8rWbl0KLLLIV/rJNAACg/qglr36i23XFegkHIDLrBV2Vg1+oJB/MH6b98Z6sg2WmkZF0jNskTMIpIBIVSUWIFrjvsbHETXZDUFqZ46meFqYNoMIUlr2HiBFidsWSOQlfi0m1zY4SJSZAE+lze4wlybpVA4i4IXO82o6iqhWV4vBKWfaZNPhJsB4g1/t8sEL7Msk5VzlNNfcA3b+OLRXTGmk5jM2+xABH2t1w/HXNsLyjyupw6li6lJsNg+sLLXwFGu7U4EHwcq9k/BOW5RUtPRZbFTzMmjmLcm1wSLEny/LExDlFNTAiGKGNium5jF/vbDtRmGnxyA2AJJKnYDucBV+eQU8lOJammQSglS5sTbcdT0Ivv02t3wl9R1Wprd7rJ9JjKFPhs95lHwZbGpOlI0ubkIgUE/IAY7LlNJMbMoN7g+HzFjgT9JTSwiopqiKRAL2VSQfqNr49UZtyYedLWxU6MwVWmjKqWPQAm18WIRyVP8LZBlsOZVUtPDy5pKCeIskeo2KG+w3O+9u+K3k0XsqyaDJaLNaLjOLMHEMCPLHWwpPMAoOlS1gCd9PYYs3s0r6it4neKjkjeSOKVBI0bhBJouD03FyNwcQvJz7g7i79K8RcZcG5pxDVSCNGqKeeerRCf6uGNGtEvXoo9ScfPO0FQuzF9IvcDoaQGkgk97o026n3IXtAa0gLYOPocyrMoqKGHh2l4gy+ppZI6ijkqRC8jFlCgMdgukub9QQtvTOOCMiqsh4zyupzTgmupmeRoKOqzLipKv3XUjXEMbbklQRtva/a+Ne4gizWbJ6iPJKqmpcwYDky1ERkjU3F9Sggna464o380uOsy4jyLMOI+I8kqqXKa01axUuXvE7tynj6lyOkhPTtj5xlWODMO+k9zWtIdN36jYxYHTfb1bplRkun4KpTZbC9dU1EJeN2mk8SuVtdjhWQ5LSpnWXOAE0VscoCCw1BhY7YBTMY556iSGSVozPIEO+9nIPzxIZNmNJDW5XWVDCOOashiUsp3ZpAqj74+7vd/wDHFv8AB/0oKbIeCSiPaF7MIpeLKrN8o4EyXPYcxpSsyT1hpnhqS7s8xNjq1Bl6bjTt1xqfBOWVGS8HZNk9U8T1FDQw08rRCyFkQKSvpcYzv2q5pnVNxvDSJmXF1DlX6OSRDkWVe8Xn5jhuY5RreEJYD1v1F9G4dkmfhOgljkq6idqKNleuTlTSNoFjKtvCx7i2xvtj4RmlbF1MvoCs8FpiPSnYgSTaw6fWjY1oqOICqvEntGnynMczMHDFdW5Nk8yw5pmSVESiBiiO2mMnU4VZFJt62vbFY4+h5nF+ZPywbsm5Fifw174i+Mspq6zi6XMcxyjgr3uNqP3xJc+qIlaZ1UIJYlGhiWUhdW7BVviX49cLxdmADAMSnU2/UXHteyeFw+Fx1M0hc0zJmfoeJ5yZAAgiyjNTyZVVkpiCQSQPQ4akUKCSLKB0vg5m2OlgbdTfA7hXJXffYnyx9QBKmmCt5yIo3DNEZCChpELX6W0C+Ml4R/5M6rOct/m57PM19750VRC/uzRpDCXHLqizPp5ZIJHVjpPhxrWUrGvDVKJTaMUihj6aN8YhwdltPl2ZZFxAuRcU0GQ1k1JTUNY3EjSMYy/9GWWnHSIlgNOo219Nzj4LlrA52KlzgZtDoBPese8C49AJO6vE+mtY9rAL+z/Mwux0x9f+0XGENs5V30/Ib43b2tMF9n2ZsxYALHupsf6xcYWoV72fe+x7/tx7fyaujAVf5vwCGs28q8+xFx/OupVZC49zPX/WXBftfz2DL+LMkzKLM8jo6jK/e4eTnMc4R3kWBhJHy1N7Lcav9Ijrew3sSitxRUPcb0hFu/xLiV/lBZyuScOisp+LavI8ySmqHoaeBEYVsihbK2pG6G3S3xHHJz8g9pw0tkuaB7QRezrR4W35IqgiiPzzRPsLnWry7P8AMFzHJ6v33NmqGXKxIIImMMQK2kUG5K6j/rYpvtQP/v3mmm9xyvT/AKNcaR7KKyhrskqZaHiyq4mUT2eonVQ0TaFPLGlFFu/Tv1xnHtOFuPs0Jv8A9Fbf/wCGuD7HEntLVJEdwj2aRzA+4LPWtSCkvYqSeJ6i5P8Akrf+pcJqMl9lmVZ1l+SRZ3nVdXLVxGnyylzOoqVR0cMuqNSVVVIDHVYADHfYqAOKqqzX/ojWv1+JcQPCdXwK1dkWZcOZ3nHDlaKjUMr90mrKRWlbTIgLR2Gq9tSsAL3xt7SU3nOazmueAGs9DaYdGqJIHiAfWOZT+zbMc9/WtJ9tUY/m5QVa1UlLUUeYRz0zjLpK1eYFcDVFH4rAEkHswU4r/sYo8ihzDNDlGYNIaTKqShNNLl8tJOVUSM08okALtJI0hBGwAt1xcPaDm8NHl8kMXGNFw3VQcuaWaeBJiY3LqqhGIuWZTa1z4bW3xWPZNUfpjPM+ziXNM0zqdYI6L3+oywUUAVCzcuNPiJu9yWHcWx43Bvqtyeox0ho/EjqyOvz5vsjeG8Qfn8fwWewVsRhQx2fSLC24P54172M1AqeFJJNIW1VIuzA9LYyWnhgp4FEaAHRfwL12xp/sGaM8IVXL1kfpCUEsLb2Xp57Wx9F7dPLsk/4moqTNJ35IPjetoBmWZ06ey3MM1rSpVMwTLqWRJXKDS2p31EDYbjti6cFR1sXBeSxZjEIq1MvgFRGEVNEnLXUNK7CxvsNsZZ7TsvyJs7zmep9mxrKltIOZvnENMjOyAIx1SgpuLdP1Ta+NX4Vpq6i4TyqjzKpFTXQUUMdTNq1cyRUAZr97kE374+Y5i1jcupaeZG5k+jy7zoF9obysowy8r5sz3P5s4zObKGyzOM0nzAvRimi5UVPyShZirGOP8QAttcCxHUi+IajqKXhYy1tQ9b77QVK0tVR00kcbrzEMgLMb/CrafCwuy7dbkSmo5BG9Wud1LZlUVq0hqI4tD0QViw5m9ze2kgdr9b7x+YU+aZ3QiqbKWkro8xqTUOtGS8kjA8suSxJUG3pYd+/6Lw4DabQOgXMlxX0L7Cc1qs3ps2q4DmNTSQstPBVVtQG5zqoa+mwZb6wb7333xFwV3H547ySPjp82y+geexTKYV9xao5kYgVpELSFG8YbmaR07XwP/JgfNYcrrqKqpUoouRHPTxNCVKl1AOom2rZV9R5+Xs2zrOM34pyDh/MeMeHM3mhzikq3pcmyuaSQcmZGJdxIyxqNtRboDj5Rng4mfV5a0jS2SQS5o07tgEDrJWwN0Um3/Mq5+2tBHlGU5kmcjK6mizDVT/0Fqtqh5IJYuUsSkFmIkY+lvLFT9m9FyuDuK5HzmWtqKfJYstammyx6KelSCGXlh0Zjcsslw3Q2xonH3D1TnyZTLRZ5+hqzLK01cM/u6TXYxSREaWNukh+2Ktw5RzU+e8cU2a57NnmYHK6UVErUcdOnLK1GhV0E3O7XJHl1x5XL8Wz5MNFr7i5EX9MWnTtefS3MQmvb35/OyxrJMxqBlVFQ5XVyQxsZI/e2pjUMpCB9IULuLlhfcWU7i9hB8aUub8QJFkmd1VM3u8xaGWlo1apY7gEx84EAg9APLFc9olXRUdRLllNDO9QQbX1bAqp1dN/Frb5uT1xAtxPU8qmo5Vapp4Iruk0QusgBAIYjULelsfo6jwdBJFyvOVxiOIA13d6LavZtFQ0KVVLBM8s0KxwTNMuhrItkuh3V7dVJO98bVxJkHEHEHCOTR5Hnxy3lRhqiAl0SrUqAEaSMiRLb7qe+4OPnD2fHi/8AT9G+ZR1lZQzQnnVLMrILi6EsDf5A/v2+geNK+go+GOHFqeJc/wAoqKhNFLT5QgeasbSCVsUYbDe+wF9zj5v2+bUNXDcE31GLavm9L/035Lu4EnzUh/h4J/2FZf8AoTJczyCpyykpMyoKwrWy09TzveXccxXZiA19DKPFvtjNeN46dc/z6gi44ytG15nSijGS1M8/9OKtKiFGtI6hF+EHTfxY0j2FQySZXmueMmcMmaVnMWfNJ4pJ5uWoiJKxABLaLWNztfFMiyx6/wBqktLQ57xPlUT5lmAoKxIqVoBUka6mNQVLgWBALbHSbY8Zg6mjNMVUc64EmL3i+4cd5EC4E9IROE02j8/gnv5QtO6ZNwg8cUkjUkyPcHlsllADG4OmzFT0NiBfpjK8/WWGjhml9/haCokcNPKA7Fm5hbsCrXsCP343j2zMkC5PHUSvKFVtbgDW2lo7kDpc4wriLMIytDCtVSVbxGQFiLBDsbaTa6qqdybaW3GPb9jMRUfl9BoFiXGf+Io30m6XOKsvtLlkj4UqZI5DG8Jj1EsQCNa36b9L9MVBUSqy2Sop4a6CoVUMkZeRkkH9pe9uh+RxZPaHmmXvwxO0VfAzgIwEUoLHxDoAetsZzlmaFWFStTLNIGEcbazqVT2a4byG3TH0/CXHeXl8ycWkaU5l9FmGZ1M0a1tRDyHWMBdTEk9z4hbHa5BDxDLl0kzzgOIy4kYA72O18GViRKsgpq2mhlqY2YIZNNirL3ta5OIAQZvV1pnOWVcuhLmTlsbsB1vbc40B1MGDusGmvAc2YRuXPUyVoR6moEXi+FzcC1+l8ESmqhgilfMpWBl5TKGZbqLHcn6i2IylGcJK5TLK67xtH4KV7rfa+w674Lpqeaqrp1q6XNII2bXzGopNKbg3Nug2Fzvgv2cIR5xN596eqp5gsssNRUCnMxVVEz9AbX6/lgioqpFlnSmqZZIecUSOSVgdI+E9d/XD2eZUsFJRJT1cVZIzMHjokeoIsb7lARvcnEZUU03NlqBQZ0JWkKxf+zpLBCdiPW2298T9lPJU5uI8fenS9X+lIkjmqJ+ZEHIDNuQTtt6AffEmwJrtLw1MSQxjUjVLHxFVJ3B9e2AaeKs5SGHJ84EkbE8s5dLqkG3QhSB3/wAHBtNX0lM1M+Ywz5fTpAVkSeJl8QYgAA7sbFdgNrdcQGj1CtrcSOqArq6GliKPz1uSoYTutyL3/W7HBPszzWer42pKcSv7tplJDSM2o6fM/TFV4oNfMkbyLy1dmMeoaC42JYarXB1YkPZFC68bQ81nDJG7KOg3BF/IjfCazm7ALZhGv1hzj9S+muF6lIaeqbckte3Xt5YB4hPPu4Gm3U2tgjh/QMvldn3L2AB9MZ5xxndXw7xzzKqYSZdmqrHHHYlqeZVsCFAJKsbduuOORqeYXqNQYwTsnZneoqPeaatvTckMjKLq5uPuCPt69MQ/DFfDPmdfSRxhQk6lb/FfcEN2G4JuPTvhHDNcRltG5oKxIpg7M60rctZNWkrfew1I+5+R3wFwVNzM8zmVQFkEqk2G2nxdPLAP1N7sKqZY46gVa84Ajp20li+hggG1zbBENQ09FEY6dVAQre/9kkfuwJnk1o9xbY9r7nHuGpXmyKCRw7MXdWLdbhj+/AJimKU1EdOZGVS5W437YHgHvGac12bkLCUMYvu5IOr5CxH1wTMxip1c2Ft98M5eyxq7bsygnr3xfo3ROhw0yh6SBE4skXmyCnSBZNAXSmo3BJ872HftiTM7PIZWkDuGNhGAoC3IGw9MA5bJzs8rFFKizCjjPMYXvqL7fQg/fDTVEdHGRK7hVGlZC4DMB8sLqEMCNkuUxOBp1pKA9gOuB6aqkRdMzXF+t+nzxCe/tPq5dVJY7HS/TDqyl1URa1CW13BIt5b98K4rTyTQx3VH18H4q1EYFibkD9+ODmSDZbbEW899sF0kkTkxuw8WwI2wRLRBJjKtwhWx3t9cCHCUbqRIUHmFLJs+piB1BP7sNVjTU9K00dNNM4BCiONn8VrrcAbD1xMu9MpJDadurHY/XCopsuK+B4b9yrfwwzi+Kzuwx6KtcQV2aUvDs9dBSPERG8ic9SvwXPiU2I6bfMYksuzWnzEPNStT1VIrNEZkcsC6tp2FvhPUHuLYk6qny6oieNo1dWFrE3GI6hy7KKCOGHLsto6FVi0OKcWM5CoA7b7kFWPzdvTB6hpg7pGl4dPJO1UMZXmHlrGF1OelvvgPNQj0tOVe0SyxNdTc31rYet+n1xJwyXh8V9Xcj9bAGZQCeOJFUuplViFNiNJ1A/7SjFSITWgxdOUSsy80jxX30mwOGqUxNJUMdLOrEKR2vcWB7d8RvEefQ5JSxRyUVdNzvCDToGNwRt1BvvhnhWvqM1qKtIMkziKYlpPHRsAwBtt1sRfcfM4c0SFme6DBRHH8ELcHZgzBWKU7OtwTpNutvP1xC5fx+9dDxDwjxIFslFMuWSRqYyVWJtCEr5puD3uR3GJn2mJNl3BFe+ZQVVJHJHyA7QN8bDwg/P8AfihcPcWcS5zJmEWWZNT1pOVw0MgSllk1rGCsbsFB8QDN17CwxsoAEiTC5uKefmiSn+AuGs14e9plNQZjkhzeknhiaTTCWkMErDTKg3OoEG/yb54m+KPZjxZkXHk6cJ0M9Vl1Sed4k1oq3BKOTfoSLHqLg36nE7TZ1xvkfF2Xz1PB1VmNDl9Gabme5zl50aNNId1QjUpF9hsSw88To9qXGc9TWZnBwXUkQK3ustpyqaRY6yy/i7i9hY7WxZ0hEyTAWb1NDFkHEOaZVxPClBNBRJSOsCArDG/LaKdRuSUk3YA33uLWGKRxrxTJmEs1AI0kp1RUGppDynXZivisRttcdMWbjjL8z4pzqGsrJK6XOcwinnnhlo5UKMjrGLEgahZR0FlHXGa5rSLR1RhWTmKo3/tA9CCOxBBH0xnJaTZGSQNKslVkVVk/Cq14zGOWmqoBI0Or4XJW1h579R5Htgeuy85dwnT1MOYCdaplZotB8D2O6nzAuDi4ZLwPledZdl0tXnFe7T1AgA5sen4SxKAnbSABY21FhpvY4kF9nOSS5TC65tVIeXVSrrmXl2iQNr0kAhTe39oWvYg3w8YaoQIIVOx2HBPcIt7DzKh8qy6qy/hqpo481ilp6+nhGmVm5SmV1S4P6tg1ycG+xRMw5E08GYiOASMslO8LMCFXVqV7gAnVa31ti4VWWRQ8MSpE0j1FFlFPmTrLPHJurxgKwWxY2WMgi4sQCL3tDeySOon4JzBKQ0sVTLM5ibltojbkoSSL7m2/lvhNRjqYgrpYE0cVW7ghoafaFfHlsHI8QUbgYqEEM5Z544w/Mf8AE3vvtuDft+zFjolqPc197FptFmNxubfM4hKJDpeNjpUNYC3T0HphBVOPeEIZohDVcySMtqfSyt1ANh079ccq0ZIGuVlfUwItbe1xfyNsHV8HOWUItpAwMdiT/jphs08j0tYzeFro4v1BI3/ZhYJ1JjvRSXUNk+o7eG3XrgKrqXkpKdFFgBv88OxSF8j2kDHoAfngRkaR0jIUKF3sdzth+wlZTcwm2IWpiMZLXUmVCep3/Ze/0xYuDI2SmndtJ5k8nTzB02+VwT9cV2SMGrQAFdOm5v02xZuFg6ZNSPISzFdRBt3xUgojIT3EEhaaONiLqDte1/riLr3Uvy4zaxtfc/4ODczlVswU2HgBt+3EQS6tI5FwzBvi+mLCWSSV2YXddZNwANVr3xzSzWLKN+5GHBzCqKGAW1wvpv6/LCTpDeFrMBub40QhC8Te0ZXcdxiQoZ2aSzXJ7E4BBVfhJJO5tbD9NYTDSrmwva++LG6olPWb+1H/ALY/hj2A+fF/m2/2MewELTJU5ncMUuWgEanDaQSNwQSL/t++M3oZZUzWmMasD7ympz0bxWta1x88aTmUksuXSc0BHVzpIH6t/D+WM+gp2/S9KA5LCrVgLgKB5fU7/XCWbypWEhXqUMWsSTv1w2xbl72JHQeWCDGx/VXr54S40ruASBfpii5U0uBuvCwQ6ha/54QrASAWLYRqBOpmBv0semPMg1M5YCw2OF6jKcQoPMsto6qu1vHJLXRqWhKqBpUiwU28yDa9h9cUafOOLa/hN/esxKw6hdFg0SppfsygeXftjU0igqZGMsaSak5bN3K3vb79PLDy0+XlTFyYghPjBUdMdLD4lhYadYEtFwByK42LwdTiCrQcGuMAk8x0VI4FhnpK+SKprK+aKZPgq5Ea7eh3PS49b4c4u4izPKeJTDlmVR1cU8XNsl0UMCQQqjwqAbbADri+OizKqSnmxppaNWYkLbpt2tho09MhM6RKrMWN1269cBXxGFqVpa0tbtCbh8LiqWHLHVNTt5j+qzD2h5HSZ3w6OJ6Gnenr0ANVT2uSoUA3ttdRb5i+DeJWH84M06f5dMOu39YcaGYKSctJNTQkOwLLpsGt6fTEjJS5XPQwQVFDTSFGdpHaIMZWZi2pgQRffrjXlua0sBUcQCWnZYsyyapj6bNTgHDfxWOk3kG/bbClfdgNx5Y2A0GUSSJKMroA8Ysre7JcD7Y6KSgjkXTR04AHTkqP2DHbHamkf7srinsjV5VAsfo5oxNEzk6VlUkelxjXMwEnvE+rWxEjHxdT4j19cdNPQwo6wUNKik3P9HUXY972vhqWRyB4tbH4ixN8cXN83ZjdIaIhdvJcnfgC7U6ZQkzlbWUXt5b4DJKqRZ21bm22Caopr/DBLYZA03LLYH0xyAQd13HNM2TGoavCNIttc4QSQ97kXHW2CnBdQFAAA6Y6QAoDMbjFB91YYhjCWXVa7bbjHTECtmc3HbB1LHLUO0NLDJNJpuUjQsbedhvgkZPmOkp+jcwDAXBNM/8ADCauKpUzDnAHxKnDKh5svgq6d4aiFJoXWzq24I/x3xj2fcNtkXG9FTlTJQz1cZhZhcMhcXVvUdP+ON7jyjMWXQcsrjYX/wAmf+GBc84TmzKhip6jLqwPHOk8bmnc6XU3va3zHyOFszCg1wJeI9YS6+H1tOkXWb1MKisdNKgBmAAA8zhqFFS7kditrCxJPXpjQ8p9m+Y1GZg5nLURQSEgvHSP4Ce+4ucEU/svkCkEZhIo/W5ZGr1+HH0I9pcrA/ej2r5w3s7mTvmLOY1jFQrMCAu+lQN/THliUyKWUhAwLBbAn06Y2ng32Q5DUUz1We5rV07q7KIWjN7djcAbb4cj9kOQzcT1kRzqrpssjCGFuVfm3XcA9iCPLvi29pssdP7QIXdnMxb8xYpNFzJfCCha9ioG3rjswEgkcKPsP4Y1uH2R5f8Azj9ym4gqfcBBzFqlhNma4BQgjZtz0uLL64jeL/ZnQZZXomUVuY5hAxsx5DeE+pAscOZ2jy57oFQe1JfkWOptBc1Ujg+qTLc9irxCjvBDI6AoDdtOwsRY4uldWSTK5bSBcmwTSOvSw6YFy7geaiq6fMVkrJU8SmnalkDqwIKtqAG1wDhFHwvVRMVNNmJJN3aOOxO+9/ARv1PrfHlO0OY4bE1mGm8bHmF7Ds1hq2GouFQbldLu1bTMXJJjYFdajsN7Hc4PgjjMniAY9flgOLLaWmqxUROxlCkG5HiG+kbAbC/rgyPc6iBbsR3xyg60LvSdUlPM4tpjjHn03xltLwe+ee1eqV4lqqaCqgmq4CdzFIwDMSCLAEgnfob+eNTEgCksADbrjEOKJa+p9oVblUeaLTR5nUQ0lTJuqaCUA1juq7E/LDqRg6llxBLrLcI/ZH7O8wYVFDPluloy3LSuOkMuzDUGO1we+Jek9j3szoKGvnNGlVURRiamX38yFuh3UHdbkA3HTGLcIZwckzery6mz6LNIZ5WfaB47OLlmJcWANuxN9sahwzWVLvHmU9IsFQkMlNJET2cKSdj12BxqdWDrQs9PDgNkEqaooKSmp4qKkg93pokCxrrZrDp1Yk2xbpJQlDkpBsAwNzsCNJB/binPMiAm6sQLWPT64n84klfL8vF/Csag2PU2G+FUxK0vMCENmDf00yGNzzVlZSB4diu1/rihV+aS0/FZo4YHdqhhMyowAIICg+I+HdX6EDptc4tvvDnNUpSJVSOgZgxB3Jfffp+qMVesy6Ko4hkzOeHeKONIXXrsZNQ236MMJqzJWyke6IUtl0lSwkNSY2Y3+EEXF/n5YckmWO7gsqr1bthugIjaRQtlLEk2679dsB5/mdDQQ8yuqYaaNiVTmSBAxt2v3wgtJ2Tw8DdFisilQCTTY9CR1w5LDTSODFceeljbFEpuNeGxTNSHONKwNyw0rFi1hbVe297dcdj454VWrWngr4+Y5VNSQsoJPm1v24EU38wlmsyLFXN6V5FZRICpUqR3IOxxX+I+FanMq6pqmzSenM6KiuFkd41BvYWYCxNzvfr2xJU9TMwDCTUCfCSL7df2YlRUq7lWNrja/TFttsr0ah31Vzw1M+RU9B+kYZJIAwUvRlkcEfrBid/UH5YM4byx6Oklo8xSlqw8bLqjp1QK5DWYLawZdhfuPIi+LEiD4enqHthGk3Jt0v3wTDKU6np2T3sv4XpIuPTVs9ZN7zTTxky1Dfga1AYRKoCqD6+Qt0wRnHAM3s0ly7irIs2pKiiyuqA93r6Jec3PPJN54gGc/ibawbHviQ4Mlp48yneeWVIlpJjI0ZsyroJJB87dMQ/AOQUXF2YDLc+XjSgWSjgzaijquITVR1NOX8DMoFlYMFOk369dsfOO0r61LMn1XPikGN1iAdQJdO5mYna4Ti1uhoaLlar7UqXN6zgWvgyKGafMS8LRRQyiN3CzIzAMSALqG74zzhuq46zbiulSqyLMKWjTiOTMJ5jmsMqwU5p2RYCquTYPY2tbvbF/9r8eZzezjN4MojrJauaNIglGCZmRpFEmi299BbGb8P0GW0vHfDNRwf7OeJOGjFUPDmE89EYoZaVoXFpCHOoh+WQT3HXHiMkLRl7xDZl8SNu6DfvDewFjeUVX0x9X52VFzLjDhjKc8nyqqr4kkiq5I2UI2iM6zfUQLfP1xMZzWAVvCaRsrw1XEFCgK9Lc3Xf/AMuGeIss4SqJq3J6zIoUm1zsa1KYapJHcyEa135ihfhuNm9cRfufJzXhKOnkqXp4OIqKTTO12VdWnod7AsP9rH2+GnAEj6H4LLqqF2ly0v2iZTxEc5h/SEXEvEFCtI6UbZJVilaGoMsh1SgOt7IYlDbi6Mbb40/hg5jR8H5b+n5w+YQUMXv0oN7yrGOY23Xe52xk3tMyngmHiypy3MMm4ap3iyiSvgqc3q5I1mnknkbQtmF1Da2a2/4i2FsapwBNT1HAuRT0tB+j6eXLqd4qW5PIUxqQlzubDb6Y+F5ydWX0TFpHzQ3cHaHEX5wBe5WynZ5/P4LDeLs94VzSuzqHJ+OaebLOIJ4pqzLoslkmrJZAqR6YZG0gEhFtq+EkkYsfHzqeNcx81ZFIJ6/hqf34a9o6+0BOJsyqaKp4qWmWYRUMGWRxGB7xxmO/hLAFlnV2PwkoemGuPpWbjrOIyoGiSPST3vCn9+PddmgHYujpcCDTcdwT/di8NbeABedkFMkEqHlIjZirKVPYXwOJ73JUbX2JthMzG9g5sOu2GBIuk2BsRub74+kQQIUc8r6LyWRBw1RySAaBSIWB320jGQZBldFR0nCXEstFnxyGszCI0mWTZxzIMvEjgUkoit4hdlOnUdFx1tjX8hv/ADbodK6j7olgT18I2OMfyzhvMMh4iyqvzPhGCiy4ZjEtPBJxRLNTUssj2DRQFNOoFjpHQHpbH5/yuo1tXENLoJdtIEjvCBLm3vYiY+uC3EDvLSfawNXs/wAzGortHvf/AOIuMIEbJKXZiQPU43n2qFV4BzIsuoBY7j/8YuMEDEuzm4A2AA/ux7zyZR5hVBHz/wAAk4idUgrQPYm2riaf4re6E7k/2l9MSvtwreKKObKv5vR56yiCrmJyuiFRqqUEfIjluDaJiX1dCbemIz2JsP5x1K2/+qkg/wDeGJv2xQiSuyR83pM6q+GVFR+kIcs5hYykJyTIIiHMY/E6bXK3xxO0haO1AJaCNOxEz3TaOZ6DmYTDJoD881J+yNqiTIJ5KtuJWqnn1TtncAhcvoW/KQAAR+QHe+M39qZA9oGZAnY8onv/ANGuNA9jCTx5NmYSHNKfJ/0g36IhzIv7xHT8tLghyXC8zmFQ2+kjFF9p8ZbjzMtyTaIgHoPw1wzseI7S1Y+if+m3hGxHLZKrfumo32JkniuqFj/kjb/95cRXCvEubjiei4d4c4+yF8joahYaiWejip0YBh+BTgyF5GN7arBR2JO2Jr2LqBxTVW6+6Htb9ZcQ0mYZo/G+U0Of5BlvBlJLXFfw8oSVnKkGIe9EGP8AENhZQGF+uOhn7RVzjEMLQe403g7A3DSCSQJiPr5KSRTZ+ea032ojKGymip8xyCLO566tipKSBmCHmNdtXM6ppVWa43223xDezzLoOHeMuJ+HqY5rDBojroIaqrFTHKsmpTKjEcxTdCpVyeikdcXTijh/LuJMp/R2ZLLoEiyxyQytHLDIpuro6m6sPMYA4X4Qy3hla2opqjMK6tq1UT1lfVNPM6oDoXU3RRqawFupx86w2PoswDqDnHUZtcjcEHeBEGbTffo9zCXysBy2oqamhpyXU/hKS7DvtfpjWP5PUEtPwnmMcram/Skpve4+CPpsLDGWZKo5McDRsJFjAJBuBtfr0ONp9kMSx8NTECxkqndt+9lH7sfU+3Lpyb/iarpCST4Ku8RezbNcx4qzbM2lyWroZ5JK2npqyFiZao0wgiWc7gxRjWygC4LnGg8L5Y+S8LZZk7VBqGoaOKnMpHx6EC6vrbGTcdZdSNxZn78T8LcU5zWTOpyKpy3mtHFFyVAVCjAROJA5JYDqDcjGs8Kx5lDwtlcOcyc3MkoolrHBvqlCDWb9/FfHzXOKtd2BoipVDh3YEAfN5EEzGxJAv1QUgA8wF8oe0y9A1C2WwctqjMVlniSMtzmQ6r9D3ANrd+nbEbNmDR+z6SLI6mXW+drFKQxWpnZ2LadS+IEhe9zsfMDGlQJE5WczyQzRSkwzIFOkNdTde6lSykDs3mMVrK+FMnyiSZqJpiJpI5Lq+kxul9LIRZlPiPcnH6FwxJotJ6D7lhLSNir77Ac4zCtzLPpM9McKCmEopQ+rkRamIBsoubbE33I6DfEdksuRZdxfw3V8KQcSZPSL7ss8Mjxz0tLDXTARxFJGLoZuWp8B8AKm3UYt3sR0iSroVANPFSxxohYt4RtuT1+uLVlXs54PyxaaOhygRJTVyV8a852HORdKMbsbhR8K9FsLDHyXtJmeHwWb4gVtR1BogbEaYMgmOdrH8Dt4TnMbCB9tJy1eF6eTM+EpeJYlql0wKrFIDpb8aQqGYIBcEhWPi6Yq3sOTLqjIuJsxoIeHqVZXenakykN4BEZFDu7nU2q91uq7dsXL2qZ1U5LSZLMMxbK8tkzSNMzrhEH5MAR2ANwdId1RC3YP264hOCszpc3zrjeqyuuhzTKpIomir44FX8QrLrgEigcxUGggm9uYRfHmsHr+R3CLG8y6LPAiI0g89wVZ/ej88lkGa8M5NnIpnzSIVDRJZfxGUi/my2OGI+BOGlmaVqFpSV3Lzub737nEtEIYhFzSdLIBuRdv8XwVTmJEY6iBa+m++PvNMv0jogcxhfJAlKoUjhiihhRRBCAI1B+EfXF69ofEGV5NwjkcGdZTluY0M9JNNKtcoYNyogViiBG8rsyhR5Bj2xR6eRXcgN+eN14fo6Wq4bytaimhmWKGN4xJGG0sBswv0I88eF7d4mnQbh6tRpcA42Bg+j18N1pDDwyB4Kl+wWGSho89yeWmpaaSgrkSSKjqpZaWJ3iSRo0EjMUKliGUG197C+KLxXkDVHtPq8om4K4eirMxqJqjLp6zN6yP34bs7rygyq1jup0+lxj6Cp6anpkkFPBHCJHMjhFC6mPVjbqT54wjiClipuLeJOKMppON5KSCokjzTNKPMKZOVy95EijdTIyR7iw8ja9seHyjHnF46vXbLdTRzI71gL6hueV52HVZqjNLAEv+UfFXCs4FyqmrBQGqllppTGxIA5a9Cd9rbHqNjjIqRK6HhimgzqeajdKhoSlRMYwEJDg6Tbc3C3B3AtexIxr/APKECScRey2pjleophnMILsTqkVmj3Nutx+3Fc9tOWUuWcVNQ5e868wIqJNM7ySMxU7lrkAC9jcWvsMfVOxAIyakT1dP2istUnikKO45paQcGzvTQwhGSMa44V2XUvS3a2/XGSJmC02XSoIoZCWBRpEDuoIva5uehGNo9pETNwLWing585VQFXqviFyPOwvj50qqioaJ42jJBf4im+3a/wC7Hv6LxTBJXFxtB1R4AU0c3pq1KekrqKR+UvKjeCXS1ie9wb7n0wxkVVQkNFVV9fEqwu45TWGoKSF+pAH1xARtLHIJV1K4a6EbEH0xwqyt9+3XCXQ52ogFM0ENDQYCs8eYZSUjElZm7PY6gGG3lY33+2G3zijjl0xVGbOnmJ9JvfbbFeRJX3CO3mQMOiFyQAjBxvbSb4sAcmhDwyL6irLVZzkk9KiR0ecUlXqu0orebqGm1rELbffvgcZlkwJQyZxpBFrzAnpv2874rcgmLamRh6lThrS97EFT64WIHIJhBPNWgZllAVdc+bzE9bVGnT9xvhibOMsp6/3nL0q2KxldNVIGBYgg9ADbpbfriACyldQUn6YneG8qnq4ap2jVCiatUsdwFsSfyB9cQkdAoxrhzKTW5rLXqlQ0QRV/DUJcKv1Pfc/lizeyEOeLUchARGynvYWO4/Z9cUqakmjcqImex2KqSCMXn2PR1P8AOdS0cioVJ3UgDwm2Hlxi6FlPvCF9DZG/9BljJN73JHy64yH29T5hQ59kwy5I157M3MaFPHILAamIN9idibY17JtJpn1Hxm2wxD8b5Zl2b0UFPXORGkqTAltLAob3v28vkcc5paKne2XZrNeaEM3VbyaDNZuFqLOaimnpqJo/do+UtIkDMNbFAsShiAzSN08++KzwJGY80zq1QbIFUpbbpsb+g2xa8xrKOgyCWOnmiSKFGMcesso1XOx373xUPZxKkuZ51KGGmeNCO21z3xK7mGoRT2CDD03im01PS5q1ZiDKjjQ1iDuB6YXwjrXIaaBwBy5JFPe9nO59ehw1WmSOnv4rOCASfywrhVyMukjdjIsVQyK7kltAA636798JG60AqXzWYLGEK9VJ/LAdGwSMBt2dh18jgvNVV4kUtpOnw2PfAWWRHV4t7sCR0Athj0Ed6UTRSN+na+N45eSSkDWhW2kRSyFrjc9Lb3A8t8V+qraV5Y6hI3ljmGqPWN7EbbDp9cWjJ9UM9ZLMqhveCqkjqNAvfsNr7+mK3mNHKM2qZaWK0QcgAkWJPcXOM1ZpIC10ihIJEWImSAqSTaznb54NppHUBmUiPoCWNiPXA01DIseqSoiEmqxW24Hlc9fpiTpIkWNIVUSu2oC41Wt6dMZ9JG6cCmZs05UEk7yItPENRcg2VR1Nxia4f4hiq4xy6mGoj6BoyGH3viG4sgq/5lT01NRhp5oTFIeYAx1KRq3sBuRt8++Jf2bcEZPQewbLuLKdKn9Mz1rQVhknOhQssi2CDbspufM79sPZR1N1DdZ6mJNOoByU9YVIGkdRsAcCNRvpazSR7bgWYHfDOXVPLqbHUVvfEwlQsy9fDffCi1amvkbqJeOaMDSdS27jAcrKQrM4DAmxYkHyxOVKtc2C2IvfEJPF+MWJAJ+EA7eewxdwlu09ELTCeWiQNIgIYgBewx4o0PiErsQbb7Wv88IjaRobSEeNjp7Na+2H3v4SxuvUm4v98QGQhJgpnOKmnoMrlzRo9SQoHYXsSQ22/oTgjg7NqvMqI5tJNKOcSIlFQz2W5Fz6kj54D4jpmrMpeka4SSO+pT3BBt+X54F9m0KU9FUoKmORSyMqrpGhSCVBC997777jGpptCx1BqdKtHEFGOIMilyfM5JJ6ORfEpkPxA3BB8xtim5V7HIocuh94yd5J9CiSWKqkAe3VtmA362xeGJjc6CBcbYB4r9rGU8PcOzx0ktPmWYUaR08lKJNLIQQpJ26g42Yd5mIXKzTCtqtHeLfUq57UaSelyDLMparraES1LKW55DTWppSupiTquwQd9rYmeCGWjyWKjEYhmhVo6rlhljkk1eIr2KkWuBtct5Yybi/2tVXEP6OIypKY0VSKhdU5cNZWUi1h/b/LGuZHXQ5jk9DVRA8qeMyK3cgm42+uBqFwBIKbhQBpafegs5pYIFhrqJm97gkjipJXkY8tGlQMvUbG1j8zjI/a/lWTUWdyVNHVRCokdjVUqGxV23DgdgfK+Nf45SEcPTVDRLJJE8bKbeJRzEO3l0xkeX5zlj+0TNGzGCkWGorlQvVRiZURCytuwJGqy79hfC6ZBEp2KEP0oLgLOKbKJaZqzMH5LVURkUb8kKH02VjZgdb3NvDta9yMa1keY1DZS1RNMCORm1QWJBi0aEUEv8RiFlHO+O/h+EjHz3HFTrWTwzTqdDsqOBdW3/IHz3xoPAfGdNk7U2WVA955FVIYalwNCI0Ziu1wSw0gKFtaxvsQCN1KsG2K5dSiXLVqClevato5XcSnhODmS+66Tcyhwz32jJN2VQLve7Ab3jOB4Y6Xh2GrpmVFzKaTmwldPJkEOgqPkUNxiQyOpycUMVXks1ZMVozQwvIwIEJYswOxuSxY3O4t132VmL1kNIKmlozUyKw1xCTSSL7le1/Q2+eMuJe2oQGrqZcX4cE9ZHtSZJmTwodRtYnyxF5UXILSNdeZYE9/4YsPJVugCnuRiCy5XdS4IJ5hO4tcEdbeeMxWyO9KIk0pWnSbHVcnsMRtZUkzGmRymsAjSQL2NrHBlYyPrVdnuQp0d/rgRo4pKhHYqwW17LfvhU3TENRHVSOhAJVyNha4x2OMRz3ZQLCx88KgjCidAy7uOnz7YTDE6TMzi/i2BxoiGLI30ymgypJPL4lVE1Bz1W3mMWPKoxS0EMQdWVIxv9MV+qgAy+vNvFJEwsN+x6YnqJmlpowAwUDcOwI2+mAamvuFHzyhqiZyytYEAAWsLjAivTziQGoOpX8QjF7W7EeXyx16hpJWC6DqYnxAKVHlYDCo42PLQIQiq3wqLEswLAA+ffDtMLOLpm6jSVfVfxABrWB2H7/th5HuvgTfztcnHYKQyzvrVUVTdSAOm/7Bt64ZWQxygGOTl69I8IUt5nc9MMa7UoRCKCF00hSOpJI74eoIH5t1dQCPPDbaDFdHVhqIDBr3+VsNowSQDU4+QwQ3SXSLqT5P+kv5fxx7EFeL+3UfZf449gZTtatGayxU1NVanfUzB7AXsPCO+KAhnn4sh0xhAlQrElQTpHXruMaNnNI1Rlde0Sa5miJiQgEFwLp+dsUD3aGnzNpXaItzgWBUsb367EemMrSdS2VWgtV4VSA2m4udttsNuCAwe3TYYfVgI1bz7+eGJ0cgm4UA9T0xFGid0xEVs2lALdQBjqKj+EBgLd++HXp4zqs253uNhj0cRSMEEk9L4jd1VUWsho4yt1A0i9tsKhhYS6mtvg2lppphKyqzLEnMc26C4Fz9SPvjjIRexPXqDiEt5JcE2SOciqwNiOnr9MMVBX4FOkDphx4zqtcH5HCeWXkve2+4OICAqc0u3TMLEShFv07m+JNSyy2axA8sCxQoCSCBb1w6FQyFixvsAB2xToJTqYLAinkNrEKFA3Ix0SqLdLnp03wgkX0eE3FhbHRTXKsnhJ638sZnG8LZTaXNlLkkYgXUuCdgB0wlKWZwOZZWJvbBaIUVgBqKjYrth1VYm7NuOw74knqi4Y5hCjLY2N7A+He/fCno6cRDVYC9hY4InDAXQr074F/HJA0Ahe4GLDlYpg8k0tDFo0qw0ncXGGpcvGq90K2sbi2Frz1Bu97tsNN/vhLvUrYOlwepC4LUUBptG4Vp9lVIKfjOFmtqaBwLfIXxs4A7jGOeytpm4wh1oyryXsCB5Y2VRv2x8R8oRPyq3+QfeURADRpXLDHbAY6QMeO22PDAoUmy49tjL87oqhfbE8C55naUxySXMvdlr5BDzklVR4L202/V6YzPKs74xh4HzKaXNc6iqKrg0ZjSczMDUtPKroXnjPWIqrC6D+16Y9LhezjsVSa9lUXDdxHpT7Yg/wDpZ3V9JghfTgA9Me2xiPHfGdZUZvxHWcK5+8+WU2T5astTRyiWOk5tZIJ5QBcCQQ2N+oAB7Y7nGcQZfluYQ8JcaZjnaxZllAIlrDKkJlqlVkFSLkh1+JN9I3tvbFs7M1zoBdBcQNjadO55ekN97qHEC622w8sdsOwxTvY/Xy5nwo9ZXV1TU5u9VIMyhnYg0dQDZoFS50KmwHmLNvqxc7Dyxw8ZhnYWs6i4yW/n2fgmtfqEpNsIlA5Z+Rw4bYTLtGfljK03RL5iy2ImnQva2mw23wQqgIQALDpvhyCUe76wttQFrbkYontN4tzDh40kNLTwu1SjkPIW8Okjtf1GP1VhwXMb6lz6zmsElXNVS+ovdtticQHFHC1JmecZTmqxoKikq4Xl8IKzRLIpYMP1ja/Xra2MfreOuJ6ltRzFogGuFiAUDf0/fgmk9o3E0EfLkqYqlQLDmRi4+oscaw0jmsJqsPJaynBeVZbX0dbldQlYYPeI5I3pgnNSSOQAMSSCQSAegIN7XG89lCS08MnPCxs0gshkDaVVVUC49FGM64E9ogzTNY8tzGFKZpb8qQSXBbsCD5404BCBq876iNvn8sH60ynpOyUZGe+xXTft1xcaWhasyyHQFBWFSQ1xa4t9cVV4JXjeTwaBtti25VM0WU0qxOLtGpJ7nxHbB0jIsgrNIJlRmeRSUNXywIEkEDIkjgFVIZeu428Q7jEJIX8evqDY3FrYn+I2/wDbtN7xIkJZpww1WdzaIgL5nwn8sVvMBd5GDMBfYHY7YVXPROwzSYMrtJoWYpvc9hiA9oORQZ1lTUFTrQatcTqRdXANj6je1sSUbuJEa97bXxKvHBVx6XCswHcXxka48ltLGu3WM0HsqvUO2ZZhFHFFSiVkilUmRyD4ASdm23BHcb4PpPZjlFPUw687mepQ3blqjx3A8upH8cac1AnM5nJjuTckDr64akpUTcRQbDuMWHvNpR020KYksB9cqtUUMWRUgikqnq4o2tAEhLOq7DSALlrXte2ww7lFfA6T5g9Y7U5USDmq0ahG+CxY26A9O9unebShpjGCY1ULe1u+O02XU1NIZqdDGxAG21rdLeWKAA3QVCXu1NgBM5dmb11RyKJYKk8ppEWGpV5GsQLab2BJPcja+PVNXVsiRTxVeXTPKqsrx721bgPZlBIva97X6HBNTBHMXkkXVI4CszNcsBuLn6477hSS0y0skMTxquwIG27N5ebNv64eRTiW7rEXVC65BHqUt7L6enqvaDV0DzV0lLNSOORNVxScsOpDKSqq/bY2A39MWj/k09m+WcT5XlQo8zSvqIJJKW2ZVJAjhMeoX12ABdPD0xGey2mgp+LoGSNY9MLg6Tta21/P64nanjT2c5jxnkOcx8Y0T1MIqcupoo21LI8rQ31G3h3RACbA6ticfJe19TG/KxFEvDeHfRO8P0zHj7pW4NYGD8fWrP7Rc+n4a4Qq86hFOXp3hW85tGFeZEYk3FgAxPXtin5Z7VKTNeIaTKcurclqZKjiCXLwsNSJHalWnaQTgBu7Lpv0xYfbPVVNF7N8yqaSKKacSU4SOVVZXvPGCpDgruCRcg264hOGF4g/T9H717I8lyeDmHXXQ11O7wCx8QVUBPlse+PGZbh8KcAatWmHOlwBLw02DeRiYnlf3InudrgH3LG87y6pilqpMsgzue1fM8qvWMhbW2s8sdPD0I07nHvZsKIca5I1RWmskpcxkhLTeJXnLLcpvYLpcdAPEgxc62GA1lQwjUXqJbjv8ZOH+G6OlGfUMogQMaqNtVt2YMLEn0sMfdnPJy0z9D/pUZQAqghSntzraCPOaeGr4wpsqmio+bT0U/DyV+p9TWcOynTcgCw/s3740LhOozSv9n2V1VRrTNajK4pJOfEIyJmiBOpQLL4juLbYzf2hVPE2V8Q1FBS8YcXVzSo1W9JlOSUk3uVO7MF1O4B/VYDcsdJONT4PaifhLJ3y6rmrKM0MJgqJ2LSTJoGl2J3LEWJv3OPhuaNFPLaGxvMx4H+Bpv4l0xYqMvUP5/FYSa+kn9neWcJ0VJxMvH1NVx1AWSKcSx15a0szyHwGI3cm506T0xIe0hlj46zZ0W7iSO9z/wDCTFu4o9p9RlGe5zSxZHFPQ5TEzSzNWBJJGRqbmBU07WSpBW5uxWwG4OK37QaZW44zWUht3jI3/wDhIMe47LvxDswa+pT0B7XuHeDvSLCR4AWgHrugYBcA7KoyvqQlnsWPTDShwhNui7XB3xLiki3JFgRvftjxp47Mulxt3FsfTmui6E0ySt44fYfzWoH1MimjjJPcDQMUJuGspliyPN8x9pmd5lQPmFLPQpPUQNFVTCQNEo0xgtdgNgfPyxf8pPJ4ZpiCRopF3te1k/PGGezuDIpuJ8p4jg9o/DJrKx4pFy56CNJEaQgukcYmZYpXuFYoLk+e9/z5l1EmpiaoeWw47N1T6Vp0nT6/cn4g96IWt+1v/wDx7mn+rH1/7RcYHY8zUF+gx9Be1BVbgTMgRcaU72/6RcYRKqXOhWO/nj3fkznzCqB9L8Ak4gXlXn2ItfiSqGgKfdj3/wBIYuntCzTPKGuyahyPNskoqnMJZIVir6WWZpmADDQI2GkABixO269O9O9icaDiCpZQQxpiNz/pLi6cc5DnFdmWWZ9w9mVHRZplyTRKK2EyQSxS6NatpIYEGNCCD2I74872v4Y7QzWIjSNxInSYmxtMck0TwhH5uo32NJLJT8QZnU5tQZlVV2bPJUGkgkhWGRIo4jGUk8QP4d9+oYHocUr2nbcc5ltpuse/n+GuNP4DyCuyeDMavN8wgrs0zWr97q5KeIxxKwjSNVRSSbBY13JuTc4zP2mpfjnMthcrF3/0Fxr7F1mVO0VR7DI0H3aRAsLCIFhPRZ67TwwPzzRvsUIPFFUwHSlIN/8AWXHRnPFnHdRLkVHX8GSLQZhTS1vuVdO0yLFOrkAFLG+gr1te4vhfsajC8RVAHX3U3/2lwfw17MuIchzagraTjnwUlOKMRfoiLxU/MDmMtqvckfF13w/tTXwtHOKz6jg14awskEib9AfBMaHGm0Dbn7VqCjbCajaCQ+SnrhwWGG6of0eTb9U/sx8qaLrSTZfL8VdVxUTze73lCXVVvvtjafYlPNU8FiadGSRp3ujG5XYbXxk9CsbU8WlUdSo3v12xsPslRE4WIjFhz3738sfde31HRkcjq1Y8JVLyQenwVA43y6Wu4s4jlzjhPi/NJxIseTVeXTFIqWMQpZo7SrZuaXJNj2xrHCMmYzcIZRJm6OmYtQwmrVxZhLyxruPPVfGZcQ8O5PxHnnG+ZHh3I3qcreOJJa2pmVppVhSR2lCsAkehlVSO4JO22NK4IqKWr4JyWroaQ0dLPl8EkNOSTyUaMEJc9bA2v6Y+a5y/VgaTQPR0T4dwER3juLmAJIvdMpemT+d18yZ9kuY1FD73RZ/PHXxMtqdVZac6ZPENib3FzuOwHQnC+HOF6Oijr66fNqylrqmd5Up6dAY5Lg2Da7hfFZrgE2NsTCQzhipWxZiCSfU47XytHJDGX1hEvfH6IwoHCYT0H3LmEg8lffYPSzUtfXLWVXvNSYACwAA0CRtIsAACAca6PljJfYg3NzjMXtY8he3+ljWwMfn/AMoAJzypHRv3BdfDiKTfzzSJI0kQo6hlYWYEXBGAc3iFPkNYlJAo008nLjQWF9JsAPniSPpgHPWKZNWuOq08h/8AKceMpOOto8U07L44zuKpz2el/nBk1bSR0sTOkiMUUkgbE9umAJOHlnC1ceXcT0cOkMI6aQzN02JZmtbY9MaPS5nM9IviTxINiSQNsEyVhfl6iCuxZCxsdsfqprKgaAuX+zeZJWfw8RZtkfDc0VPk+b1s9HrEs1RTNphWxYGRhfcAjuNsfWfs5mnqeAsgqapOXPLltPJKliNLNGpIsdxucYZSZjNH7wkMtoqhHhlU7h1ZdJBB2+EkeeN+4Ucy8N5dISSXpo2+6jHzfymNAwdDrqP3LXhnOMibKU3ta+MR4py6mqYOL80oZeLE4fp611zXL6OpiWKsI/yt4ww1qo31AMNRDWxt9rjFFzv2aZZmWYV0v6bz6joMxkaSvy2mrNFNUswAe4tqXVbxaSL3Pnj5rkmNpYWo51V0THjsQTbrGx5HpumVWl2yrnthhoauThOaBV5FK61lJp2A5bRMlvSwt8icVrimGfi/itM1YJC91Ghd1ICkb7/I/TFx9sUSRVmSQRaI40hmVVG1gDGAAMUUeBjIrAEnY3vj7V2KJdk9MtmCXf8AMd1nqBuq+6tK8Ogw8kV0IJN9JJB+Rt0wNNwfRNBoVsuIJ8WpPCCfW2IGStnUXWWO463F/wBuB6jMpnsvvSBh0JGwx7DU+EEUypaPgZWfxUmV6QSFYWP1+H9mHDwFpO8WXWVSPgHz/s4g0zeti8Iq0a5t1Plh5+IK0pf3iQArYqEbfFcV3VVw6aMl4NqIGkQQZMqKbAFzcj5CPDNVwu9MHq6hMohSIrqmkmCqOm5Zk6bgfPDcObVjLqYzNcbEA3/PENxlLJxFw7V5NMlRGs62UsCLON1J33FwNsWKzp3QvoMDZUhUcKQGGJmlyNjJuo5/h0gdQdHT6Y9BwSJnLwDJHK76hK3/AO76Yy6s9m2ZzUHDAq83MIngaOqkkiLrQIpZgCELM297Cw6jtixQ+xHJ0OmT2oxpcDRpyypsSdjfbbfDiXzusoe0C7Vf6bgVWVRIcqYMN1iBax/2Rhyl4GqHWY08lAB/VuGuuofQG49DiscMez3+aKVdVlftCauGnTNSR0csZLq6HwavCe4v0Iv1thNXNxRl3CdVSUdTU1dcl5I5Zo1OpgFtYXP9m/XqSe9sZahfMBbaIp6JIurdHwDJctM2VK1zpNyT69VwqPg+rpWEkcVHqB/UZRcYyDhfiD2hvmkEWdR1ksUkp1GSnKNGdPxE2AAxpYzHNFiF5mlvYliQPyvgn1HN5oadNjhIHtR1dNmOW07pDSwyyX6SPpVj/rAH7YzTijKeNM/nkmra2hpkdvDFE7aQu/Xb5Y0VKqumQFi2mx2vcX+mI+tkqkvpolZQet7YWHibhMLZEErMKbgHP7xvLX02pNSIAGa4PU2uMWPhbhyqysVb1EkIWSywgJpdgADcgG1vi29BieL5rJKrQ01OpANmJP8AC+HUoMzeORqqeNI9DAhEPltck/uxC/VyVNpsDpm6CzqQGjOojTHcG/nYdsc4ZRkowH0uWmve+/wL27fL64HzKmNPlDhmLtfUdQHi/h3wTw0t6UTKfBJUk28tKqtj/s4ETKaYlSWaIQis2wsLfTDeWyPIQuhUA3v0v0scP5trMOxA22wxlblI3uGAA2Plhl9KBoHEuieHJ5KjIqyqe0fNq51QWst0ZowSRuTt+eEzCVn3VAhYN3OO8LJzMhlKTyVETSzTKSlkRWckKvf1N/M4TMQviABJICjzPlgWQTdG6Q0IOrp9IUCIL4rnpYXN72+eFU0NnIWNWfUQvfvh2QiQjcXDWI7A48kcYp2LqbXYuB3G4+fQ4lRoJBVMqaRCpVdxZmRzGryyLh2rzdKWQxuyzhgw3tqATa4wdQcX8TU/CrcJUfB2Y0uSwr7wsceppUmLW16iLAeK1rW6bebHAldHQr7QMxidIDCC8EhI2bSwjIvtuSLD1GJvgZPaLxD7MZs6zKqnky/MCIKOppZ0WoklFTF+Gqi2k2STfYY0BgiAsDqhcZcUPwrmWbZjz1zHI6vLUp0W88x+O99rWG+18WilqisV7ggi/wA8Vzg+kzCt4mrEemzqqpaLNzSV0QEs9QAF6kG4A209Ba974nJo+VTsS5uotvuTjLUpDcLXQqkiCpOKpVaUILAnfrgLQZKpSpA0nY9sA65ApbxGwv6YLp5X3MgRQbbjCHAhaW1A5Me56fE4J8ZIINu+DBAjgCzoOpNrE4dM0HKALLYHYHqcOwyU8pKq6lgAdIxCLJluaj66KSSIxRnxfqFj8LbWJsd8B8GZXVUWTGonkc008ssdMGe+lYZGRgB1UXt9ziVzFVihZzZQBck9vrhHB9QW4JyugIlWWm57yc2OzKZZmkA+ekrf1w9l2krHUBDx0TeZVSUUSzzRyNGXRPw0LG7MANvmRjC8wyNc99rObZVNM9PzKycs2m5BDE2xvtRMiRapCCeouOtsYbwRXJmntglzEMbVMs0xLKAfFv07b40YdskSubm79FF7mm4CjOF+DIM4q81p3rpIvcKjlAiMHWLsL7nb4caPlN+C8siNVmPPoaWB0jSypYl9ZY7ks1gwUC3UYhvZiscmecUhhe9dqt6a5MOe3FYRwxRsnX3tQd9vgbG402mjK8r55iRmLaYd3ZFvqROd+0vhyuyqekikqBLKAoZ4DYbjGT8VVNHWcSZrV0Tu0VRWSSwnTpBVnJ6duoxEkjqRhUb6JA5AJB6EXxjADRAXqHvc8yVa4clopeEpc3pqR53FS0YTn/iIgRfFoHXxMDc9QD0sTgmjyCjm4cbO2gq54o6hDURRyIrrE2oXAsSbFQCbW37YvEOccNQ+wej9zenjrYs0liT3qnPi0uZACVuWujqN7b3wDlVZkT+wecymljr6fPJYwzQmzGSmlMelrEkA22NgLfLBEFDIi6lfZXODlbU1DR/hitkiSNZw7WCBtYFrkbG7dL2Hzv8Ar0R37jY4q3CXFXBeWeyuBVmWneDOZaZZBS6mZOY8qnV8TKU0qbja+JmHNaPMA9bl766SZmanNr3Unb/Bwp5IcLclsw7hcEo6JrW0Ntq3B7YgqdZo4Hc69RfYm+rZQN7gb38sS1OAzE6bH9b52GAxAXyo/ArhyetyPS+EPJC3CCo6WQmCVyF1C5I/eMKpyjhiEUkgXF7d8cqKRnimdJ0DiPwANa9wLddtrHDEcrxSRSLOGTlaGUm2o7H9xwqbotIS6dFFYxAOrVcjSccqAklVJY7X/Wawx6mkYJzkJKkgg23IJuOmHaRFd3kaMNckA2AxocTpAWcAa0uZbwELHqUr273wdRxSDKTKsZ6EeuA3/DSyuostz1B27DErRtDHlpg2E5AJS+ogG+5+xwIsUTgqukb8xnkGzG+4vbbbHI4IzNq/EbSbixH78PzMVd1BDWJvthtNQJRLEHvptjWIhZJIROhZpl3sh2U3HiPqB/djkhiVGeNehtY7hjhmkjkjmeYM7sBdyw8PTp+0+eHA0VQAkaiRSCTrO46bWPzOLAAFle64WB8DqTcsEUf6I8v1Rfa526eeHoTBNIHjjMahQuoTLJdu527dNsNyxGXQY4yNRuyrsTYWAP5fbCqalsTGOVBbewuD+zElQttcKK5kv/WIv9r+/HsK1J5j/wAd8ewF0/SFoE9jHImogsnU9hjKM6mmWdmebSBLZLg3b16Y1WsZiAPMEfnjKc/o6g5gkWiPnc66cx7ILm1+nr6Yys9JOqyG7LRInLQxIUAso3NvLC3F1UamG4vYjfCXAEwKm6DtfphiupZZ0haOoanMcoYsovrG4KkdLWP3AODIhRhT4ABYBbC1uvX54uns94Wy3Pcpnqa+WrDxTmMcuRQBYA91PnjPcsp6v37XU1TFDCV5AjAXWH2fVa+67W/wY32j8R59w7wrHLk2Z1dHrrFik5MjL1VjcC9r+EYjWyR4oKroaT0W8jgHJVaRIq3MY+ZHoa0iE2JBsDo23A6eWEr7O8gYkS1eYC52LSLuf9nGe8EV3HlbwpTZpxHFnFLPNXRwU7h54TLE0Re4VtOojzFxtbFMGee0TN/bDVZLk9ZnVTTU1RBzliSeUQpYXZuXcr8TC529R1xobTAWR1YwFpntE4Ty3h9aKagqJ5WqXZTzSpCgLfayjFRMQjQMQFJO4BxZuIpZKtswIzCsq6OnrQlE8kzSI8RiRtYve+7EXG22K7INlJN/njLU3laqLXRdNiXQLOgYEm1jjr1MEEkQmdEkmbRGpIGs2JsL97A/bCoUJFhut+tsCZfkxjq4K+uq2qqinsYy6ABSBYMBuA1ibn12sLDAFwG60MY51mqQjBM40AA6fLBSeEFmNj3IGww7TKGkvYgW69zh4ABySAB3xlN3SujTbpZCRCXLrcki3ngiJisjjS3mPLCHkgjTmNy1HQXIGHYJqeeJZIJkdLkEobi/kTgx4K5AMFC1dQYwdUZYW7b374Cy6smqaZpGo5Ke0hGhyoYL+qxW91B7XAOxxKVlKsoB2vbwk4FSlCEgjVuLkHfAPDj6JhQbpkTRnYutxv8AF1w7EwkIULt698eMK6iE1bHqcKijOq6lj57DB3VetWv2cmP+dcVgNXJa/wBsaut/PGNey/NMuqOPpMriq4jXUsLGWG/iAKg3t3HiHTzxso64+KeUIEZqJ+gPvKVULTBb+boLMM3oMvraGjrKgRT5hK0NKpUnmOELlbgWB0qx3t0OI2j4z4YrIKSelzeGWOrqpaSBwrWeWIM0g6bBQjEk7WHXAftTyTNc64ZQ5AYBnVBVxVuXmZtKCRDYhiOgKM6/XFIy32U5hBX1uWGpiiyMZLJT0TqbyR1c1NDTyuR5aYdXqZH88cbA4HLK2G4mIrFr729X1c5EepyyOfUDoAsrvlnHXBWbT1L0WbU88lPSvO7GJ1LQL8ToWUa0Hmtx0wzlvHfAdXRS1tHmtIYaJIgx5DK0aSsESylQdLNYXAtipnhHi/N4cthzPLcry5ckySroIDT1hl97llgEQIGheXHZdVjc3t5XxGP7L+I6fJZqCGZq5qvKMthkkqqu8lPNT1KO8UbAC0enUR6r646LMtyknS7EEGR85pgSQbxHo35RMXKAvqdFo8/FnBuUU9Soq6eGKKsNJIkFOzaqjTqZAEU62Ci5C3tbe2H8jzrhPMvdqHKJqGZKmnNfDHFF4WjDhS9rWBDWG+9/lih1PBPEVPwrScPjKaTM4sozJ5aKqjzB6SreFlbTKsqfDOpchtVw4uepwjIfZ1xLVZplFTxJmtdAabJ5aaWfLcxaCUyNUl0RmjC67RkAtYXYXwt+Ay3hud5wQZPzgZsYMb3tuRExcqa3z6Kus3HvBFBRUuZPm9MkeZRmeFo4XZ5kXYyFVUtpHTURbD2a+0Lg7LJaaOuz+li96plqoX8TI0LGyyawCoU26kgYoPB3B3GXB8eUZhR5VQZrUrkS5VVUstdyuSySu6ur6WDKddmGx2B36Y5k/sjzAzUOWZrmtTDlScOLl1YaCVU94kM7u8VmUsI7OQCLG3lg35bk4eS6uS0TcOBJ35RIiB/NNoCgfV+itlR1dQ6EMpFwR0OOSfA3ywimhjpqaKnhUJFEgRFHYAWAwuX+qPyx5GO/bZaeS+ao5C0alFI8N7eeMi9u8herykajsk17/NMavARyUAUAMm3oLYyb27+HM8sAt/VSdfmuP1XhPRHqXMxY/ZrNR8Q8sc3INt98PUiq1XCrAMGcBlN7Wv59cWr9FZayn+hQAk9pJP8Aex2sJgKuKBLIsvP4rHUsMQH81UkLK6shsw3BHbGt+yXPOIOJc6GTe+wwSytBFHM1KJEjVdZJZbjdthq87YrtFlWUCNg+XUjEnYtLLcfZ8WLh6WChrcsNBTZfl1RHMkaVlPEeZuQLtdrP2Nj5Y2jI8UAXEhY/lzDag0A3W0VP+R1EZZSEVtyLgbHtcYtHCMcRyWml2ayaVOm217jbt1xSq9qiLKKvU0U0jWN4gVDdL9enfF14DKDh+mC6FbyXe/r6Y4dPuyF6itDgCOahuMEaPjLLpDsmifYdSSi9/pisVZEk09w2x6kbdcXfiuSGSvhmbSXSdo0v5mNr/kMZvxTDFUZglFJUTxqGVHig0gzCRrC7MOngfYb74CtaSUeFBPdC60iqdAK6iNVtQJt8vLC1q+SmsuoA3JJtiCrcjj99r5loHM9Hl4pKWCOVouSSNvGDclgGPiPQ9O57xAmY5pQ1+WxT09DdWjgBl5kjpcgMV6i67g9b2xma0OsCtLi6nuFZ0zWwWOSZA7khVuLt8sdNTGzXBcHvfGaU/DmbQVa0+Vx5tmtVSPDUECllqVhJQq5so8Vy3hA3Gk36E4sNHmVVS5dE2a0tVBXBLPG0XLLPpuVRXIYne3zweggpbawdY2VnNWqp1BI7Fscra+np6V56iVIkRSSxJtbzPpiFoqyauyumqWppqf32N3pg9rsquU1XF9O/Y26HyxH5bBXjJs0pXmmqVjkdIZ5InCSqbhtAa5IDagNz0+mLgTBQ65FgrDBmEFRAk0E8csbAEMrbN8rYJqK2KnpnkdGKhNRIHQYg+FxJFk8YbS2hmRFCaQiKzKoPrYC/qTiaSWKYaCCGt1OKd3XQrpt1NBKnPZhmENVxVDDJFOaOpoJnao0FY0QCx1MenfA9FV8K1WS5Tw+JuIKfLqmejopOIJ8nKQZhTQSMaaHmdE3ZVEmkX+uzvs1gyzMeL5srEqi+XywTRxMFKRuvcDpe97nvh+m/S+aZFkPCuZcZcJycLT1cdFS11PHKKmv92kFolv8AhqxMYBIJvvpvtj512iePlI94iAz6gNZkd0yRuG/XykHpIaAb7rTfa6M1Ps7zVMlXMGrXWNFFACajSZUD8u2+rQW3xmns6gkfjaM51kvtDkipqwrkjZlr93poeXe83i8T62lF21eHQO2Nf4xiz6fhmti4YqqWlzho7UstUuqNWuLkix7X7He2Me4PyziCm9qlDUcd5dxPVSGDTRVU1WKqjjq7tqdeUESNSmwDIDe+PD5HVBy+tT1NHpHfvmw2FrW67E2R1R3wY+CGzCoeKsqLRXHvElyzf6Z8sR+U8TtHx5k2Sfo15ZJ6mI81AdCKX63I67dsUfPMq4zqOMsyhGa1FMz1kzRLHOnKCK2wFzZviW/yINzi88ELfNcmhqMwiqqmKph1zwrpV21K2wO9iCPnftj7FU7uAIJnufgn0qpe4ECFM+0TN4M/zdKjKsp4gXNpKiry6IZfnK0Xv1LSXMzyGxsiyF1UHckk3F8a5wZPRVXAmT1OQwLTUcuWwvRRSXIjQxgop+QsD8sUriXIfZdkmdZj+lJfccyz6lnacR1ExkMIvLMyBSeWG0MSQBqIPU40TIY8uiyWhiyjlfo5aeNaTlG6crSNGk+Wm1sfFc4xNF+CpNpMeGg2LgYNvWQfD6+RgLptIeZK+d+LM1zFOIc1zjiKj9mEudZHVRQsJaeYVc/hjcNEhku9g4C36lTbtiye0eVk40zEaQQGjsOl/wANTizcV8WcKUnFFRUVnBtTmIyiVIazO0y+KRKJiFa2snX4Q6k6QbXxTvaS7HjvNVB2WSO48vwkx7TszWfWxtIupFkUz6vmWA5DnHQjldSlDSZMqFNfIpIZOvfCGrSELd8MMEkbmO1vQmwvhqVBYspDWHbH0wOtCk3X0nkIWfh2iV91kpEB7bFRhvJOGsjyfLqShoctp0io40ihZkDOAoAXxHcnbr1w7wzb+buW9f8AJYv/AEjEl074/MeKrVGVqjWuMFxJ9pT3gF0qr+1RuXwFmTAXsqf/AJxcYM1Q5suwPyxuvtbP/wDL7NCy6gFjNj/2i4wQTyqtuWNJFwbdMfWfJs/TgKv834BIq0y87rRvYqzHiKoDAX92Jvb/AElxM+2TL/e63I6rM8gzDiDh6n5/v1BR+JuawTkytHqHMVbSC3YsDbFf9h7NJxLVMVK2piCLbfEMTntmpqiqzbh+OoyGsz3JEWpesooKlIg8oEYhLh3XUoBk233sccXtE4ntKHTHd6x807GRB6GReEVRumkB+d1I+xzL6igyfM7ZXWZRlU9e0uVUFW95KeDlxggi50AyCRgt9g2KP7TWYcdZgBcG0X1/DXFz9icFZSZRm1PJldVlWXLmTNltHU1CzNBCYoyVBVmAXmcwgX2BxSPanEf5+Zi6yMp0RXt/qLhnZV3/AOR1CSD3Tt/w73N+tzebpLmzTClvYyxPFFSLf/VT1NyfEuNh+mMa9i0TR8VTu0mq9I19738S42QN1xx+3wnOHH+Fq0sEMASremG6k/gSAf2T+zC7jDVQRyJP9U/sx4oC6s7L5qo5P6JCA2kaBuMbH7HSx4TbVb/KX389hjGYWZoIySACotY3xsvsX24SYDtUuPyGPunb52rI48WrHhGwSfD4LOPbhmORfzslWv4IyoVcMlPCma5yJUgqQ7ILqEAWVU1XIdxYKdsbhkiuuR0ayVMFU4p0DTwIEjkOkeJVBICnqBc7YpfF/HtPlGZ12QZrkD18srBMujhZJI6y6glZC20JU3J17aRqBPQXLh1WThzL0ano6ZlpYwYaQgwxnSPDGR1UdB6Y+XZm5/yfQa5mkDY6tQdYXE3EcwLSY3CdTjWYK+d+aSH6lQzftw0xYEgkXIvjsqpGdQkQarmx2tvgc1cakBituuzXx+i8LUigz1D7lzXMutN9ixIzivXcf0ddtO3xY1gHHz5wdxX/ADdqp6mGBaxpowml5CoABvfYHFnb2vzqd8lgA8zVH/dx8l7XdlsyzLNH18MyWkC8tGw8SujTrNbTAK1y+AM/BbI64DqaaS3+ycZc3tkqFvfI4Lf/AIUf9zA+Ye1+Wpo56ZcjS0qMhPvB2BFr/DjzDOw2dMeDwuf0m/FHxmEbrPYVkWCNHYkiNeg26b4eUSOmxvY7bYE54eMIp02QAeHywpHdItyBexG22PvpqOgLIykzdSEUEyoQzIGvqsR1x9C8GXHCeVA2uKSK9v8AVGPm6mlRGMysS97t/di2U/tD4ho6eGlpqqFI4wEQPAuygbY8V2yybFZzQp06ES0k3Ph6lqpBrAQt8Bx0nGDwe0/iXnES1VLoPbkDD49pfEbMQk9Nsf8AMDHz9vk9zU/R9p+CM1G9VP8AtvcJmWStpJvHOC3kLx4zyWphJKqLLe3T+/BXE/E9Zn81NLmJUmlD6GWMLbVbb5eEYjk5QiRzZFAPe5Jx9X7M5fiMsy1mGrEahO21ySszwxzpRPOijW3Kvt9ThLTxo6/g9enTAIqIUFozcna5732xyScEFXJQg9cd8VHTBSntaBIRslZZbincjv4gLYT+k1vtCRfp4h/HEdz21AGyj13w07tsNXW9yowRMmEDamlSwzH9VaYC3Q6x/HHDmG7f0YEd7sN/zxCO7FwLN87dcLXmEaQBvvt2xZamcS0wptc0cQralCqOgBX+OFnNDqINO2/Xdev3xBvqAVRIPUA2wtioH9YC3Sw6/fFaFQq9QpuPNlD2Me53I1D+OGZs5Jcg0hN+hMg3P3xFc1SlyCDfr0vh0SxhlLN87kYotPIpwqg7hSBrHkUFqFlJ7F9/yx2OoOgF4iCv9luv3wGshLtpY2ve4bDyqx3PVhuS2Ag7lCXCbI0ZlGmotEEAGwJ3OPPmL2JaF332UMB+/DCUrMB4lAUEbHDnu6IynSBuDsL4ouJRAA7p2KtlbdaVl6DSXW/7ccmzgDmQTQvEwG++q3rthJiBlL3AbbpiNdHFbWtpZGazag2rUNNt+w6dMDJRQBsEniKBf0Y0wcbKdh5HDfDkQgy+OIOjj3iRtv8ASa9t9774czIvNlWlZNNxYspsdt9vtgigiCoFRrq0rux0WsT1v5798ND5Q6ZMonMgrBhpawFrnAmU89lkaWIRw8wqvjuxAPW3Yde+C6w3Dpq2tscCS00vu6OtQ0Pu55p0i4cWNwwPUf4Bw0PAbBSHtJfYpr+c8eT8Fq3uFfVSzkoixoHNgeu2w2W+KdV8fSyVEckGT5ssSK3MDU/e3hI3/bjUeHZoaLg3L544IuZJGGVCgZVDXPcX74EfiWoR9MqUdtwLwLY/LbCBUDRcLRUpPN5gLNaT2hzRUMRqeHswNQVAdkjsrNbcjywxP7Rq4UwhpskzBZNQJZo9mW+4sB5bYkfaJxrxRl9fTZrlk1MtBHZZ4fdI9LHUbHcXsQbXHlitSe0vif3qqjWvilirQWonNJEzQnXsLadjsV+xGNTYIXNqPc10Eq5exbLDNw3xB7/RGKKpNhDOm7KEUb+Y38vPFtpc+oIfYnS+z2gUJU1UlRNFIdSqg5ruFFhfZSDf0xSOAOPM04hpJctqqunFaKbkKhiSPmuX2k1CzXsbEDvY7Yp+cQ8VcNz5fn9fUQOhfR7oKq7Lt4lKEkrtcXxQDgSQoXNLA2FNUmZVuSZXncYjqKlnz0S3R25rELqJNhYrYWP+vtha+0XMxANORyKx/VMTsP3YrGR8Y53k9HV5dQZiimpcVMFS8KM0jbhlJYXub2ub7qB03xfPZ/x7mXEMNPS1WYQ01ZTsDKFoo25yDY3G25uNx0Iv0NsEWyAlteOsKDg42rzBKs+WZg4kLGwRvwwT2Nr7dsSNBxTmebsuWUOXViS6ATK4ZfCtiSWt1sDvh2L2jcRVmQVOYxUlGr++R0cKpBFa5vuL7lunUW++J7L81zQ8W8S0OaVHvJpKOmVY4o1C3IJeypsSWPTv06YE2EwmMhx9JUWhruM6zjWmyCgN6mYpUTIhQlgY0J3c2UbXsCLEnGsJkz5VndU8zEzFEVlLajGQu4uDY+e3e/nilcCiSP8AlFoZ6X3Yrl8ZMOhk0Xhj2s2/fvvjU89ngr81zGWiqS6JWSwubWOtHKsPoRbErU28Fr+pR4Sq4Yl9MchKgMqzmgz+hnaiLywxycpi0RVWPexPxdD0wvK5hza9RKWW6CMXvawbVc9f7IA7acRnCVDU5NlCZTUqbR1dRJDItYzho2a6ryjslr9R1JN+xxJUC6zOLEEORYjbT/i/3xmJDTAW4a6jQ59iE3mtfS0tE9TVyiKGMXd7E2HTtjPMxzvhKiiOcZClJLW5cZZtCsVaYTMA1zbopOw7dMaNm0YGTzlYVYCNiUIBB2O1j1xg2Z8OZeKnNJo2zKOCge9QYKdJI4g7eAajIpJsVuLdb400rwVhxolmk7FWX2LTzV2YZ9WBUVpZUlK2uBqLm354K9uWscM0pZlsKxdgtv1H9cK9jVImT5/xRlzyib3Oq9212069DSLe29r26YX7c0V+FaaWJSVWsXXbcKNLgdvzx0Af2C8U9oGaj1/gFi6gswAub4JNDVCxMR8XTxDDUB/pEYB21eeJ6urKam5QsXuLi2CwmGpVWF9R0Qu3isRUpvaxgmUBKczOWx5W5l91ilacQ6/DrYAFreZAA+mEK2YrlhpFMopedzTGXsmu2kNbztcYKGbwGRmCOARhs5pDoty2O99saDh8J/mJAxGKO7E0BmK0HupdxTiQyiLULaiAC1vkAMW72Z8RV+XZnT5NKypR1EwBd0ZjFf8AsgHa5xW0zCCSWNQhXfe+JbKUA4gotVgvvEdz6XGC+T6FRssdMK6eYVqTw17YlblSyOMwRFhZoZgSZNQGnboR1+uHqVVei8Dh0OoqQdrX2+eF0Meqi5VyVKWve2AcrlaalEWsg+IKwuFYX2K/48sedqs0kr1FN9x4pmrUF7EEgBr9rbemACFEUNTDGHIZbLe1lZgt+/QYNqmE6S8oeFWAfmD4rdbfliPo9TZWsYk6AjxC1yPL64ygrcW806xlDOIkaVkIsga21/X0/ZhcdQsNCKhmU08nSYWIG/p29cJoSxkmdbqdTkEjsF64mIp6efhyaBmRBIDpiJFrBvt5Y1tbqC57naHIKS5GzggC+9sGZdORQ1csiKzW63AY+nywBTC9KLC5LbDBWWuJ4qolerbAdPpgIlPBUMq6ixO5O+5/LHJUIJEZZWHe9sF8syyFtrBui79PW2GamIJHYEAA2Bc2Jv5Ye10WWaJXKV2QFSAFY+I9T3w40PMIaIBrk3se/rhhQ4Ua1h0X6r4rn6YWhBTQh0k7CxtfDUMImOdkski2Y9DqGOxSvdyQzINjv1+2B5J4WnSGWP4zYMDt0J64VDA0DBRzGiN9D6uo9MUFc2SdS/2ofs38cewD7xF/nab/AGz/AAx7FLRZXur0tILXxnfFEhgzaslZtFnQxFuh8I8X3JxoVW+khnLDw3v2G+M19oM6rVODIYjfRuTYnSp3079/LGQOAdCY891Xuka0SF7OWUG46dMOdwDcMe2GqaSnWCJrhdQBGm9u3TDztG7rIJFYHoT0vgzUZ1SWTKXZebsouAOp3xSvbVAJuDkkS34dbEzb9iGX9rDFy1xrJZ5EB73PU4KyuLKZ8zpUzREqMu5oaoiI1B1+Xn0IPYjFNeA7UCjqM1MI6rVeP1jHsb4Vaway05Uk3t+A3njOv5L0Sv8AygeMpmB8OXwgHtuE/hiy8c8Q5HmHDUOV5TNWNJFUrK/vBJ2ERTY3Nu231xEcFzcN5WZMxqWqoMyk0iR6WVk5iKbqHsRqsSf2Ycaw1TKycAmlHim+MKJouJcxpKGlKwwyvaGCPwxoP9EbAAYrsmkG6gXBt1xLcS5pT1+d11bDPq94kLAg2Nj5+WISV0EysSB3G+Mj3DXMre30QFTY/aRlK8RyZTVQVFMUnaBpWtywwJF9t7E9+2L6I6hlinlpJoYpl1RGQFdYBtcXG49cUqTgDLq7OK7MebQRpUFXMc8Jfx76iLA2vcE+oxZcopBl2T01DJWS1DR6lLP0F3awTvpsRa++NGIpUeEKgcJPLoseEr1zXdScwwOan6FS1OCGv8he2ETzAagdzv8Atw7RKaelOhgSFsvbfEdKHJLNZT5Dc45wcF6A+ikVOWZxm8VT+i6fnQ01OzzXIIXUCFsuoEnYm+/TpjI/YNmVZTZy1FHMGpau4kibsyi4YHz7dP2Y+h/ZxX0lIeJqKozukyyaoo4Up5p51j0N+KNQuexIxl9JwrDl3Gs2c5dKI6eaqeQQqu1mHY328/rbHToupMpd/mD7VwsS2u/FA0ztE+rmr4syKxDhtIU7Hzx55F3INifLtgcTKwO4JO2ETXBGgXB7eWObNgu/ZPFl06rncdCeuBmqCDoBbc+mEMzlCRt3F8DyRh1MQXQp6kHcX8rd8CSeSoxCrvskrqSj/lSVxrp4otdM0cbyShRrMUdgL9SegGPqHLuIMizCteioM4y+qqkTmPDDUo7qu3iKg3A3G/qMfDXDhqqn260PPeOWePM0ZnjBItGb/P4Vt9MbX7DZqen9t3tT4kqUedsup5yqg9U5jMR9olAx5jP+yNLO8eKj6hadI2AOxPxXEZXNOnO9yt0birhlKc1LcQZUsAteQ1aBRfpve2+GP578Gdf52ZF/+0Iv97GV11JTyezySBF0wRx05QEgmyPGR/6Rh6jynIsy9nPC80uS0DVEud1kE8nuy63WOOrAu1rndEO/ljlYbyX4fEMLxXcIdGwUx+NGDrClEy3UtP8A57cGnpxZkR//ANhF/vY4eNeDrf8A0ryL/wDaEX+9jG+I8vyqDiWijhyyhpojS1V1ip1QOBJB8Qtvsx/PAHGmT5LR5CkkWTZckjVEcZYUydGPbb0wrE+TTDUK/BdXd7B8U7DVnYjAeeN2vb1LeqPirhiscpScQ5RUMNysVZGx/I4LGbZWemYUn/jL/HHzDlr5Xw1lEtd7rSQ6S7tpjVXkvchAe/mAfLBtbx/lNKlL7xTV689b2FPq5Y2O9r9iOl+u9sSr5MqLHQK59g+KzZdmbMZSNSIvC+kTnOUi4OZUVx/8df444c8yfb/2pQ79P6Qv8cYXDmFJVxgrqR+hjkXRIpsDYqbEGxB+o88DViNp0qbC5sQL4X+rejP78/ZHxXSkG4W+/prKP/tSh/8AHX+OEyZzlHLP/tSi6f59f44wJCoAViL7AY5MC0tipW3cHAnyb0Qf359g+KmobJulhWKMq5BsoC9drepxkHt2Q+/5cSCRolAbVfuuNol1W0hbgLuO5GKxxtwvHxJkbUxcLVJd6d2FtLeR9CNjj6dhyWECVnxNIPpkBfPFDY1kBI6SAfni5gqLEQEH9uKpDSVNNmy0s8bRVEcwVlbbSQcW+0xsOYl8e6yAzSefFfPc9gVGjwRdAI3j1e7g79u2JTKuWM1y8JT7+8oNxsLsMAZcJhGQZEDA72wfliy/pOhUOl/eY9tt/EMelf8Auj6l5hp/bD1rXM1jjfLax5lZVKEOU/VHn9sT3BOYKnDlMIImjVYU0Ek7i4APbtiDzV54qCo0sFIjII7YJ4SlZuH6OdAwLHljU17gPbHzOnGty+tV5LB6k9xVUf8AvRl0ahhpzSRiQdgOQ9gfMYpnEAkm4qaZDTCKE05qo5Y9ayRfjqwvuykhzuovttfpiT47zpqLiulTlSOxrJAslrKr6HW35/LrjOs5rq1Kl6mp1yyTxJIikboULAdPVm9NjiqhbKlMw2yls0zPPayVc2oYKSheeyPDLCQW5ancgKG7kLcDbsMWThzhzKKqGpqc4ympzSqrYmglrdKkR21oRFpi6Ak73O6jytjIqzOJTExeoZV3BvuAfLyvgRa9+XG8sj+SrbqB364zvkmyImd7rYVyn3DjPNsqyyaujiqadVpqyjrHFtIvZVWRVFmfTZunLJANyMR3G+WZu+dw0DVoEcMayaq+chDpjVQqOSxbddRJAJYEjvjL0rZWISnnkUEXUamJPfcD6YRNPWyWdmqXRh4WIbxAn+7FtLglvax0XWk0cmdUHCtCa58vWoqSDCaaYkojDU2sfqN9Tc+uH6jM6iOnp/cZ3MkUSh7rdnYsFBsQT/aP2uSL4yuoqaw2deeoIshN9/rhrmV8kixmSW7XFhe9/wB+AYwh0uummr3YaVrqTIkqvGFbWo5skjMZHIB3IAsDf8iOwwnIJvdsyzSSSeaRJ5FeJSjELZBcL279u588ZO8ldAWVudF5LoIOPJV1skx5c9RfoLkjfBvZquhZVLQAV9HezmbLKniTMMvSJaSXMcvmjqKxoF1IoQ7iQ7qBe/UDYX6DDvNzuv4dyn2dZjS8KZdSUUlJHJmsecxENHC6MHihtrWRwotfpqxm/wDJvlkk9psjZhT1FRTjLKrmRMmvWukXXT+sTuLY0/i7MeFswoFPD3s4zKDOXqaUR1M/Cp8CLLGGuzIQLRgi/a22PmvaIv8AlYUw2bMIMWaZdBNxtPQjwWnUHMBWle16PNpvZzm8eSrVtVmNLrSG0zRa15oj/wBMx6wPW2KHlecZHn3HnCQ4EjzgPlkbUuYtJFPHBFQrE4WKXmbM/MKaepuDc40D2rNm68A5kcjNaKy0YvRC9QI+YvNMQ/t8vXp9bYzbhkcV1XHWR57L/OaA5jmFTejqSy01NlcUJROch2WZn0uO5LY8TklMOwD3OIsXxeJJYPSEXESRcQQUdU98fV96pefrl+YtEkNdPHNS180pbmggvzGsACPCoFhYG/U33wJwDS5nTcecPSV2exVVNDMKVotHL8Ky2hYINtRut/tc9TmGd5VnjZ3WvBluZkGrlsyQPYgux22tiQ9nNNmVN7R+G0rIamnAzakLJNdTYyrY2PnY/Ox8sfa61MtwLhPzPwSG1QXNtstz9qGc0ldxs9Vl78W5VmNBHoqTHw971HJHEZ0WVbkWH401m3U3G22Nj4DioYeCcjhywTihTLoFpueLScsRrp1D+1a1/XFW9ouScOzZ1HmNfx3V8KVtRSiml5GYRQe8wqzEAiQH4S72YbjUcXPhily6h4cy6iyiRZMugpYoqR0k1holUBCG/WFgN++PhOa4ynWy6jTZqEciLbHnAnwuYE7LVTaRUcSsf4ppcql48zwZzw1xxJTvWxs0eURTmgzBRFHZpVvZmBGk22IRb33xCe0nMKSL2jZ7FI2iRJor6j1vBGbj0xL+0fL6zNeJM3ThzKeNajM1qo4xUU+cCOgRwIyQy80FF0ncaQd79wcY5/KEo8wr/bdn9PQxySyaoSqRklj+BH2+mPfdk2ipi6Uuvwzadv3d4kxP1beCzvfom3P4q4e80Llm95VbAW3/AHYS9TQSKVFVGq28RN8Y6eGuLkBY5bXqD8V42Fr7ftwJU5XxDEriWirlVBqbwNZVte5PbH0rhBL84I5L754dzTK0yKgT9IUqlaaMWMy3HhHriRObZWP/AOo0g/8Axy/xx+dlNlOcVsK1dJRVNREzbSRoWBsf24LfIeIBINeW1u4sPwmI/Zj5lW8mjKtRz/ONyT6PU+tPON56V9u+1bNsuPAOacqvpXfSmyyqx/rF7Ywd83pb/wCUwAHpc4xL9C55r0jKa/ff+qYfux6HIs+nmtFltUW1WtpJ36fxx63s72bbktB1EP1ajO0cgOp6JLsWSZAhfV/sJrqep4pqYo6iN5Pc2bSvkHUE/mMWv2nUmUZtxjwrkmb5Tk9XFUe9StNmCatCII9UcQvbWxZTv2RtsYh/I7oK6H2kZjU1Eb8sZS6FgwZVYyxEKbdCQD9jjZPbFXTnLv0fmfAtPnVA9SQlVUK81PTqEUiV1iRpFN2ddh+rcsAceA7S0XfpJDN9PIgG4IkSRJG8TdaRV10QT+bo72NR0NJSZ/lVDl2U0gy/NngaTLEKwVH4UbhrXNnCsEbfqh+WM/8AazmdJTce5nHPUxRlViut7H+rXri9+wCpjrOAIqiKoyfQ8zEUeWUwhioTteIrckt+sS251Y+dP5UVHVP7YM1mQWjaOALqIANoFJ6/LB9lWR2krA7hpH190Hf29Olkuo6KQIVpo+KYKFzJRZq9MxXxPHLpYj1IwXHx5WBvFxFXAHv7638cfO8EVXPDUSUyMVhXVKNXwr5/lgrLKHMMwqI4aPlSyODpTWobwi52Ppj6nUy7DVna6jAT4gFK86dEBfQD8eTAEDibMNv/AL438cJXjyrKlRxFWHax1VrH9+MCbLswgaWKaakR4jZl95S4IOnfe/U4GWGeVlPNhuSAQZlXqSO59Pp9RgPkvBg/um/ZHwQHEOO63EZjl5NveYRbp+KMbh7CJ4Kjg2V6d0dRWOpKtffSuPiSkoqiflQpPTcx7MFaoVQNV7AkkDt9O+Prv+SRSTUnsulSZ4XLZlMwaGVZFI0oOqkjqDjy3b9v/wAQf5mp2EcdR9SZzPJcwzTjbieY+zLIOIolzBEiq66SOCTT7tCdO8RLgEt4iT5dsavkUTU+RUUDUEOXtHTohpIWDRwWUDQpAFwOgNh0xIdMJkPhNjj41jcydimMploAaANzyEbExfewC1sp6TK+Np+KMqirJ4J6mJVjYgSFxqJubiw3FvPHP5y8Osh/pgOmxNu4xn+fcPVUOaPLVVFBCKiokCA1Ac/F30X09e9sPUHCDVkamnz3JXayjT70AQT2OP01QYOG2DyC5es8wr6vFPDqDWK7SL2tftjsXE+QSu8fvbk9hbp8zY4qcns5zpFMj1WWMgIJJqbAC3W5GHKPgjO6YmYLDPABf8CUSGxsb7G/TELQE0OeeS0WkqqWohDxxpID0s35460UbsofWvZhb9lsVLK5TQyqipKki/ECRYeV/wAsWujrFq41YyKCLALfqcKDjsmloPJOJTwoQyG9uxuMcaKMam1MqMb362wtysbWOgBjuf8Ajh2nsHtYlT6+H7YuVYDRshY4muxSS9u9tsLWJhcMyhtJFj1wW0YS+lQEv088Ny6SCAih++9rYsq4Qrxquq4udrXOxw65UnZQtxbw9L4THHYkFgW7EjBaU45aOzM2rrYWscQGFWjUo/QVMmkkWAvvc/QYT+KoXXdg37MEzm0tmYlFA8LLbBEAgm38II733xWsC6rhzYKMj1gjShsOhI6YekBZwWi1WJJ22F8FzU0Q0rE7tdvEC3fCo1BRjqkXfz2OLD29EJpnaUAyyE6yhtfbywkbG1mBt2xIzBgpUO4HU2F8dih1E/HqNj5A4PWCh4BCig0gALC3lthSOxBKqLgXJtgNuIcpquI5ckhqHSsgJUqY9IJHUA9z/A4nloLkSBj5288MJi6EAGwUWoRzrbUhsCNsc0gLcar779cSklOtgSAQlrgqbXwl6XSjOI7azc2A6YrUCoG22UaIw8gHjK9dxbf5YXHRSO+s3CWuLd8GRQaPxOx/VvtbBYiUrbxBNtsXKICbJugh0q1z4HAYbdfLBqKGU2W7WtfCIlDLYmy32vh0aWa0dtQ6b4TUJ5JtJoBuuFSlvCb2vthqZQ7KLsSO/wDHHpRLbVouQLXJ2Ix2FDpuW3HxHChJTyGjZOrcW1IdultycQNRNJFxNmURUlTSRuCoNwDcEfkDixAgFASSD0scVesmccVZn4mCPSpFZv7R32+n7cWUMIyqmigykzSWRFUltQ9O+HsoI1Tgm2mplA36WIv9zvj1QqvlLK4FjGQ1++2+PZIumBNKkqZpDe4NxfBsKH1I2oBYsCAF879cC59Cv838xRmtzKSUXY7XKHr6Yfr5IRHKZACEBN+th3xD8U1aVPC9QlMJvx4VRbAoWDEdzhukkSEJLQYO6tMp5PCFA4cN+Cij6J29MYhxNxVX5LxjLSlVno40ReSxt1QdD8zjaOJxFBllHSxHwpHYKP1QLAD7Yy7IqKE/yh+H4KiJZ4amqhMqSAMGBJBBB6jbAU2h9SChxZLWAhVriDPayhzpFq0oquilp11UGtyqo6f2tKkMb32749xaKaiyTJWoiYopVFVpkYM5JRR4TyxpsS3huR8J643H+Udl2XQ/ym+FIY6WOJHyyMlUQBWIeexIA9B9sK/lxUmVU/EnAkENIiTtzQ+hQqlNcQtYDrcnG/RDbfnZcg1JPeWGZtmdPQyUvENFQrT1FUt4o3CtG4TSpZ0KkE3BO2ncg2vvh3ijPMzzjhEZg0VNTQzzcom93bZbgC2wul738xjb/wCWhR0kHA3AbQ08URYTC6IAbcuKw2xYP5QmW5JD/J64U10UcckmY0ohaNAliY5C2ojto1/UjAls7otUbL5TgpslnySSqZZ4hFKqPKFBbUykhVW+6+Ft7jp64n+DBltbNVZxTRSU9Xl1NK4XQqxMBGVR+5BGzEAG5F8a5VUGW/8A8EM9UKSJZhUUhaXlqHLGSLobX/W+xOLv7KMv4ZX+SlmVbmNExpGy+Z6l40AlcBTcBhv6ffDCw9UoPHRfMPCedUs1J/N+pymWrmqMwSqSohm0shBF7AqdzYbnzONB4PGYtxnxBR1GXNDXy01KjiJuYkRRSSWY2tcD73xdP5CNDQ1dRxGaujgmdWj/AKyMMF+V8A5bKsvtc4v92qquanSKBYGncs+jSx3PzLH6nCqklt0+ge/ZCcK5MY/aVR8TVJkWWIe61xkYdRBTWI3P6zOSb9B52Bt+dTRPmFU0DK0ZqHdGAtfUbk2+v1xD0Ejx12aR3DAVhIsbn/J6fD6sHjL2G5+R+uM9aodAp8t1uw1NvENXmbJFRIFiTlxB5Nzbe9gL9vWwt64TkbKY608lUZpbsEWwG5HzvtvhU8MNVCYZ41aM2Y6mFtmBtb6Yey5FSOqGhEJk20nY9R5nCGBdBx6ofPKmoo8rqKqnpWqpIYy6xL1cjsMYpUcW0FZkuf0CtmOXPmZMksARJImdN1S4AZbsq3Pp9cb80QEaliL3tscU7juZuEuEc0zLJkhjqZ6lJtToGVWZkU2B2369OpJxrw72ncwuZjqL9IIVK9kNXLm+c8SZi0SrJV1QqWRTspdpGIH3wd7bIag8Jx7WCVSsw3G2lh0+owD/ACepFSszhpLeLlE3G36+NMzKjos2hqMsqoY5Ip4mQq1gW3Xpv1G5+mNziBQkheHqNHyoCDeR9wXy7SIHq4EBHicC/li1Znw5PMI2j0MQRcenfA2f8PVPDXGEOXVKkrzleGQ9JEJ2P7j6jFyUtrj1A/THZyHDU8RReH9Qjz7FVcLXYWWsVWRwvFZBZNR6gnC/5s0+50x2GLEDeRiFYE7YblLcvQLkA36dcegGX4f6IXn/AJTxR+eVWpOGoAA4dV32tg/LcrmjzSnkaRTokVhYeRxLyG5RGS/0x3L1UZrEpfQruAdr6Rfc9cU7B0abTpEJtPH16j2hzpWq5bLqjjVRsRY9dtsA0qxvSxzMASjuquGsQLjpjlJURKrCGYtGPhcbEjsbfLDNBJL7iGbxqrv4rW7+Xn/DHzaruZ8V9TpRDShqmeNtaBlGkdG2+e+OJIwaZQi6U8Q223ANvzwJXFXeScpI6IOy2DMbgAnoL45TvOXWiuqRqrXIF2a3TcYyMF7ra42UlCnJp5pDqCqXG3z2xTOJuIajh7hjKGpo45nmLK3MvtYKdvyxd6ZT7jFGTdGUj1Y+eKvmfD/85eGqSlhRRVC8kDs1tLaRt8jsMbaAkkFc/FSANKmuGs1hzfJRXQRSopbQ4dCLOBcrfvsb7YpHDvtIqP0iaTMqWIRSsU1QqdV+ii1/8XxeslyipybKTBOKdWknaQaZtRvpUXsNu35Yw7h+LVm06MgkIjY9L28Q3xsZhabsQKTDPiufUx1anhuNUbBHJb3lVXHWQlwlRCSbASrpNsKqKVEUyPctboThnLJHfhygU6WJjOpWBLEjoftiDbPMviqJ7Z1QJKDywlROLxEE3HnfGJ7XMeWi8GFtY8FjXG0iVL6StNqv4Adzaw6YTFq0szqCB0J3xW6jiF2KKlbku4+Jq1Db88HR1OfSosoho5IWFxJDMHDDzBBwbXEbqwQdipaN4YQXkIAA6snc+o/jhX6Ro1Q/iIgv4lY3Hpv2wwsUgQCQXIFzbphiWiimDLIoVW3JU6ScTWN0UBNfp7Kf+u0v/iNj2In+aeTf9TX/APWG/hj2F6x1TrL6C9lo1cWurMGApXuOx8S41U08THeGM/NRjLfZV/8ASyc3293IA7bFcaxfHwjyhE/K2/zW/it7fRCHqvdqallnljAjiQu2mPUQALmwAufkMQFDxjwlW00c9PXRtHKQsWqnkUyEuqAKCoLXZgNh5+RxZj023xRovZ3TQvlE8GZzpV5QhFLNywbEyl2uO4ZWZCPJrixAx5nLxgnB3nb3A8o9R3sece3YwgeX/NU5U8RcOQZyconrKdKwGzKyHSradYVntpDFRqCk3I3tgWl4w4RqVgMOY0459RFTQh4mQySShjGAGAuHCtpbo1jYnDNZwYKjM8ydc3qIctzSY1FbRrGt5JDAIDaQ7hSqoSPNeoBIxHD2a01S8L51mkuZmNYoSDCsQaCKGoiRPCbhv6S7FxY3Ata2NtGllJYDUrOBgbXvF9xG/j9fICXVeQUvT8YcI1E1JDDmdLI1ZI0VOyxsVd1lMRGq1v6wFRc7npfCpOLeFI6OCqNfC0VRTxVEOiBmaSOUkRlVCknUVOwF9jtiBk9l9F+lMvrIcyeGOgqWnjjSBbqDVGoCK1/CATo2G64QnsxRcrNC+c+8JHTU9HSiejR0jggZ2RWX9ZvH8YKnwra293uoZJYtrv8ArHL7Juq1VugU/HxdwjJWCkWthMrRwyj8B7FZSBE2rTazFgBv1v5HDmXcScK5lUCCjraWabniAJyyG1kOQLEdCIpLHodB32xGR8BRiGNJM4qp3Sky6maWVQzye6TmYOxvuWJsfLAGT+zP9FVT1tDnhiqhUrURsaRSgKiceJdXiYioYarj4V2wIo5M5p/buDuXMH/Sr1VRyCtmVZrkWZ11VRULRzTUjtHMBAwVWVirDURYkEEbE4k/doL/ANRF/sDFc4P4U/m/mmaVorUqP0hPJO4930OGeRnN21G4BYgbDbFo+uONjTSZVig8uba/jz5DmmsLiO8oXjGKKPhXNJFiUFaWQghRf4TjEWLv4gOgxuPGZ/8AdPNepHukn/pOMKYQbvoDg72Hnj6X5O3E0K0nmPuWmnshqulgqiPeFDKjpIBfe6kEfmBgpyoCAPaxud+2AKhgwYIjqinewv27YXTpE7rpB0sNixN8fSjJZugBDahgXKLkdQpJI0EEbdcEwHmx6RbbqL9PXATJAsZYptudicB5TM71LBCNIbb6/PApgeZUmGCpcvcWAAI6YZldQxN7kgXAFtsIrAPxFCiw6kbWOIqWeUKEC6N7LffbFgTslve0G6ruXZXTcMe12hz1Eaoy+Yyc99Op4JpVdVso3O5Hntq8sXf2X0VLw/Xe0arrK/nPndCyUlvEZHZZSw+ephbzBxAM0Jmd5jqupBIbSFB6/lfBNNWx0tNDBCiBY1AVW/VAFh5/tw5moVNfOIXNcymRHjKuNZmFPHwRNRx1EUlSzwwcoN4gSyE7fLfEzwokLezLI1kkSOWlzuvnZGYBgrGqUEg72IcffGcvm5WHlM7adRkEf+kwAJ323Cj7YHTNomdtQuw3+IEY2YSr5uwsiZMrHj8M3F1eIXR3dPvlW/PpUn4nyyZG8Joak+HoLvTncjp36+uIvjCV6rLauMRkiDNoEBHloRjt/wB44i2zymG2gt4dhsf24bbPVkRrcwsvQXAv64zYv9viOLtt7lrwhGHy7zOZsb+vwT9XzgkMFPNNEZZAisLBLsjD8Q9kBI387YrtbluYzyTGLL8pr6mSKVRFTTRzyRhkjUNYMzKwZS1/GLt2vvKnPjGA0SM0nS97WH3xxuIZ5EI0yjVbxayLfbBPcaji5c/L8GMHS4czefamKSpzZeIsuZ8pzBEqaTmIJqYARK2kMSYrL4mQC7r0RLWvtY2llbcw9DuWHTEJJxPICWtIRbSGLX1fn0wLPns0kZMSKCAe5sTbvbFsZJutwfoFlOmtVnOgq6LY3Qg3+2HBNJLJGSCIz98S/CHDubcc0OY5lT11FSzJKacgQ6ArctTqAUWHxeWJyD2V8SRLFGuZZZy40CrfmFthbrbHnsd2kyrC13UK1YBzdxB+CfS1+kQq+pcoC4AXoLdxhvQXj2FiTYXxaZvZvxfIY7Zpkyql7jRJvtt2v+f3wPF7MOMVkZjnWUWZtWlUcD9m/wA8ZW9rMmH+IHsPwTSXHksZ9pXBdTW5jBm+W04kmWwniTdpFtYN/rDp8vliAHC+f67rw/mbDzWO+Po6k9mvFcT65M5y1ietlYfuwdD7P+KFZi2e0rAm+nR0Hz0462C8oeVYVhY2s0/UfguLjsjpYyoHuJC+baXh7OogY5sizFXJsFMRvvg/I8lzD9K0TvllZEkU6uzyRkKoB6nH0SOBOIBJc1mXEAWG7D/5cM1Hs+4ilVUFdlgQOCRZzcDt0x0j5Uctcwt4rfY74LmnslQDg4OduqVnuh6GoAmujISWPQWwfwA0M3BdJLTi8YmMdypAJDbkHyviZr/ZfndZS1FLJVZXy50aNypdSVI9B1xNcM8BZjlHDsGVioo1aFmYaNWk3II6i99scVnbDJWn+0D2H4LvVKNRzbDksx45pnrOKYpIZGjakmkAsSFcG+pSB17b+m2IKtajoswr5nodbx5eszAgXKKXNtPS+5vjXH9nGdPmMlW9bl5Dys9vHvfe/TriArvZBxJWz5m8tflEa1NG9LCEeQ2uWsTddvi3tfAv7YZI4k+cD3/BW2lUA2usu4T4vyXPKqSnpaSKlmEZ8EkCfD08PbofniX41mmi4cr5VV4paeAyBolVXUJ4rbgjt5HFdz/2RZv7Oa7KM1zPMMvqnlrgsQpmfYKLkG4GxFxia4pq6F8jmpc0rHgFTE9pQ+ncWt4r9emx6jHcwtelisOzFYZ4cxxIBHhus1Otqc+k8QW/imci4TzPM8tTM4ps6lPvCoppq1hYbXuI6Ug9el779MUM59LRwx0I4Zy6tlHgFSRVFpbNYNbWOtuw7+eLPFxXwxS0oimkrudTopitXRMG0i1hpiAJNh+scA5FJlUUEbng5cxzOZuZrlpXZFvbfSHZWW29rKB2OHuJAkpbGhxgBTfDAOYZNlVNmaVLGpJU2hqJ0iUyWW595AVFBA+HYL364Tx3nEmUcIR5NlM0dVIankqVldXXRId0sxJsygE3It88A5v/ADwy+CShyyCqSCaWpTRQw8uEU7WKFljAuQC1gehwDmXB9fBUy1dHUVFTJMOZTvClopOa12UliXTwlibi/QHrfFmpZE1hk23XeG6ieogilz41YFVPMtRLFVSQSTnloUjuWWOwIDAC5Nz4bYP4V98WCooMnz+dqmaiE0cUlRflzF2Fjq+HSApYWB8e+w2NfhLMajKKmjq8wodMkaGl5MGlaZv1wqg2AIVBqBubX7DBHBHCmXZHmBzPN4TmVcrERskhjSxdrtYdyhUWOwseuEF5I3Twy4EKb/k90tbD7YBJmoqErpaSpkcNOzRvd7akXcAHr17H0x9QAYwf2TUko9rKZtG8bxzZdNTSKdmjUMjJbzHhPXzxvOPiPlHqa81af4B97lswtPhsInmVWPajR5rX8FVdNklM1TX82nkihWYRGQJPG7LrJAF1UjGecMrxzmXFlF7/AJC9PQxcQzZlPMM5hn5MTU7xrCUVr2DEHy9MXz2wx5jN7Oc1gyuKukqZRHGUolLTmNpUEmgDe+gtjOuG8vyqk9oPDNRwh7O+JeGVSaWHMZp6IxQS05gksJDra5EgjIJ798YcjfGXPHd+fEjbujc6hE2Asbyqq/vB9SzSt4q4byrPK2rlWBpJJpB/kckzghj4tTTMEPe+kdNsP0IySu9omUZhURV6Z1LmWXmBKqGSMxqk+knpYhtxv3U79sWieso/eK2Q0EFRNNI0bS1DvMxCs1ra2IA3OwFvTC+C6SCDiWjkXmO8lXETeVgBZhawUgdfMY+11KzW4BxH0D9yS2i5zgD1U17VpuEa3jmuaXivhmmzWjy6BXps2jR41ZXlIW5YHdJJNSDfeJvK+qeztY4+AOHkhqHqYhllMEmaPlmReUtmK/q3G9u2Mq9rOZ5dlftLanlzbhfh558qhmapzDKve5Kk82VdPUBQtgQep1HfbbXcgqZJ+E6CsiqYcykkoY5FniXlR1BKAhlH6obqB2Bx8Rzlz/kzDNvBiJ9XXSB/qJHNaKccRxWO8aiOl4w4gziifjeHKoKyMZzW5dmsUEEMojiBKwkan0oY9R+fW2Ib2hvl2Ve2HM5JcjirK3M5qaOnrdZApGKLGbsV8JN12v0PXe2JPPstzLiLP48zk4ByWlzGrqkhaeXijXTyTRkf1kCWErKE+G1/Dv0wv2iMF45zn8VbGWMafXlR9ce27NOa3GUmz3hTINwdtA5OPMGNrQIsl6dUrPuMDnecOIsjyrOMzFBMZJ5aUsEQC11Fvj2kU7Ha9yCMRtDwxqq6rOs/p3Ebl0pYq8MOXzJRy20k3/XFhYD4txscXaOeRF/DlZP9XYYanVZg8Uqc2OQAMGQMrDysdjj6U2r3Ul1AEzKiG4ehpqzL8nhz4ZFlkcdyMvkEdRNMzt/Wm27WCgem3bEDxrlmV5dmclBRZpnGaVawc+OqryWEUpDaQGCHVcAqBt4yN8XRKWFYoo46ZERG1IFiUBTe9xttvvh6aW6ssjPZfQbYlKppdLiYRPw2tndgFZVmNLJRZMVq6OeuqoWkMzSSSOuvQmk8vSCF8SGz2vota3U08CZzmdZQ09Acrhra+lkrp1NOYIYbKjCIqQRfYkAL+s19ul4qWvI0nJAY21OUBLWNxfzsemHKGSeOf3lauVZbbOGs2/qMW+sSZQtwrWiFMfyWsrjouJKmZqaGOpbLF5klNEyQuGZHFiR42FyCQbC3Tude9oHGEfBzUVbXRQyZXIJBVlZgKiMDTpdIzvIoudQXxC6kA74pnsWZTxJKgG6UWhT5KGGw9N/zxOe1GhzfMOLeHYMu4byjPYRS1rSx5nGOQjXp9B18tyrfFYbXsfLHxjtTTpYjtHprAFum8mNmk78j08Vs0cKiA383Uz7N84Of5fW5t7hl9DHUVRaOGnmSWbToWzTlPCJD103NhpBN8fPP8pDIMwqPaRnFdT0s1RFUQwqVCbahCFU9egZhv2sfLH0Z7P6Cvy/LKiLMOHcjyKRp9QhyltUbjSo1N4E8XUdOgG+M09qU8Z42roGIBUQ21DvoGK7HVWMz6oaQgaTzm3d5yZ9qhp8RgDl87Zhw1nZoXgbL6hwsyiEWaS8X4rC4Vj3I2NzuPPf0/DedRiV3izCpqmeWFViRhZAqFXuT0vYW/gbadO8jVR0nbV2XCiHsWDsLbm+Ps7cQ7dZzgmKiZdlOZxwTH9GTuiVFaGvEHl06oXFzpJud7Hqd7bYk6WkzgwU0S0NUsckVNpcxuCbVckhIKkadiLg7i/Y4tAFS/wAAZWPfV6WxyI5gos2r5Xwt1QkyEYw7QIVVyiHPJZMrlzTL6yGjVaKOXnOjPIqTOWOgqe8hNn6dTcdPoj+TZG1P7Ofd5YuXMldUGYbWLs2o2IABA1WuOtsY5KsqNuj6epuvT5YJ4Y4k4gy2lqaelzCalpRLqjEbWuSN7j7Y8/2iyuvm+COGpkAyDfwRAMo3X1PcYTKfCd7Y+bRxnxYXAOeVYAFtnG+3XCcz4u4qlhjii4lzCmOtGeRGBJUMCy7ja42vj55+rnHg3qs9/wAEfGBFgszr4Vj4heOm59CC1ZWvJBGoVNKOrqpK3IFj8W257747wnkWVVeUR02Z5VCJ4oweagILk3ubix1DpY+WLDlFJ7jQVEVVMa16wyc1H3EayWLouwspI6DB2ZV8tXmT10hHOdQl12VVF7ADp3PTH2VlQBmkclhp4bSQSq7Dwfw9D4yJ3Vrgq0xNvniay11oIhDQ3SnRQAoYiwHlvgOeUu50k/Ebnyw00r87Srm5HUG23niy0kXKc1wZ6LQFNTmmzJW5pEU9v6wi5tiGqb0EirIpXS4KkX6X+m2EiSQEm4NjfcYkpPdq+FKeRgrW1I5FyGwoshCHk7ommnSvRI2ISXr13GJAoaeO51FlP6puTitVdJJROrtqWwAWUbg+t8SOX5uJNNPVDptrVBiweqMWUuzzMgbUSRsRfDUz3tfZrd/446soAsX8JPhsptgeSVXJTwFr2IF/D88WjFwnDqC3YGx6m++CObGkIaN15im4FsAksrqF8Y2+H1xn/FHG+cZFnlRQigpmVGtEZdRJXs2zYgYXGyB1UUhJWj1EvMN3cMx6ACwwkSBFVVHwm5xly+02XZpcoV3ta4nIH02xb+C+IoeI6OVhB7tUI9mi16jbsQbDbAvpOAupSrsc6AqN7UM9zGHjOUUFdPTRxwxqqRyEW8N7/cnFdh4s4ljkDjOq0nyaUm/0wX7Tqeog4wrmqY2TWwZNQ2ZbDcemPqPK8hy6uyihlzHK6eYJTxjVPArWOkeYx1cNQFUQLLiYqu6k6d18wjj/AIuCWGcuSf8AQS/7MOU/tC4shFjmCSDpZ4lP7Bj6eqeG+GqiEQtkuTst9Phpo9/qBtgQez7gwSt/7s5Vckm3IGNZwB6rKMxWZfyXsm4d4y4+4gzLi8JPPTUhrYIRM0TNIH1M66SCdIHTybGjx6WijYql9AB03P7cFUPAvDOWZzT5rlWUQUdZErGOWmLJa6kNcA2IsSLG4N8dSnEXgBUoBse5xhxdPggA811ssqcfURyhASxrpsq7k+XfCZU0QlhGzbbb4LmRAylgb9jfAlWrXBQlfzGMQdIldIsgwgkmkUgNGG9L2w4ZAz/1Om39k480JZ9esbdgMMWlDN8QHW4HTBOchY2Ci+YpUbmzC11F7Ycg5OnXpFh11YCTmW2sbfTHpZRIEjZWNmU2Bt0/dgTsii6LnlEgkIW4AsNOE0yGxB1EE7XPXA6S65R1svysP44fLxjddIA6nrhfrRDeU+6KoGo2sbgeeKjWM0nEuaPHIPw6eNjq3OxP5dftiyBjNImg7AX/AOOK9TSKuc5zzJ+Y5hBKWFo1LED1ucQWEoiZKP5n9FfUygKbar+ag3/PBWQyvNl8NTKu0hZwCCpUMbgEHoRex+WBXiWXLZGVjY2JG3kMF5UrLlkQIIt28+uIwoX90SvTKDFItks53J73vcfbEJxQlQmXrHC0BiiMCSqWIN+bHpUEbbgkWPmMWKWNmVowAp6i4xG1lIJIGilsdfiOsW0kEW+1gfpjQXkN0hJDQ/vHdE8Rz6aemuyklRqZQfL8sReQ5CY+M+FuK6ioigWjqg0okuPw9XxEjckeVsO1cVZWZJEqRolSkZBjkJKar9L9wR0Pr0xSuMnzrNmpIMtjqI4qdVpVdmEYdid97262W5O5274Ww6XSFdXvNgrc/bXNkPEXtR4e4hy5JXlyyJoZqk00wilUk2GrTawubEXvqPlhn+U7X8NcYZhwxQ0hd6qlMtT7+sLWj0BSsRJXcMWBIHTRj5wWmzrn/o+lmaSs0sjx/pCN2aw8QC6jvt0xLUPDnGO02WtHqkjMfONXHq0MBcbm4uPljWa4gArntwur0Qt4/lI5pwhxF7NMiysmSfMKUxNDLDDIUpdQVTrOn4TaxHXp3tg/2p55w1m/sf4c4ZmLVNVJW0YLx08hFGocq8t9PXSGW2x8ePnDNoOJ8rpq6eoqHQxsr1MnvAZ2YsRq3O92B3F9wd8E0lPxDW5JR5rRyKUesF/x9QXSkITULmwv9rHpgDXEFQUDMFblllTw1Q/ydJOC88jkq5pJEhWnhhkWRiCul7EWBBXUB3sMTHs7zvh3h7+T5VZLnNP+lhBE0MlBFTyFp9RJ3Urt0J8tuuPnp6LiuCkkU1jKgX8b+mKzeV/iOwvufL5YepKfiWoDin4hpzy4rsozFTpHTffbFHFAojhN1rH8kqoy/hF86q8xlamgqqeKqs9PJceG7KoAsbE22+2K7wlPV5v7R+MK6jo5YBPTwlOVTMqDwOLhW3AB7HyxS8v4a4qWmeP33lJTqIvFMy7BRbSQLEWPn2OJPJ+FOIKGirI46unM9fGInne7aVBO1u/U9b9cWa7SIKMYZ7TIV2oYaqgqKz9LRrTztIZWLoU1qIIFZlB3IuDtv0wLJxHlwlP4kpRQNQNJKGFybfq/4tigQ8A8T5gWnqM1LTOHIWokZje/QeQNut/LDuSezuoXimCDMMyE2iN5JE3AJFgACTvu1+nbCXw8yCtNMvZYhXpMyy6UNJVkx0oGlveomiRtV7DxgBuhxYHzOnzFmeCjoKcEiRTSAFSG6XOkXbb6Yovtiy2nh4RjEz2igrIWRE7j4SL/ACJN9/lhz2URVdPk1ZFXwRwg1GuLRyyXRlDBiy/Edz13HTAtENlMdUOvSQr3FCoj0oTpJ1WHbz7bYg/aDkq57wvW5WurmlboB11qbgfcDEkKmNIpYZReOVCrBhcEHC6WakiiDCYuRtpYbm3TfAiWXCa8tqCCsZ9i9JPST5xFUQSQujRCzoRe2sHrjSqOopqXPKCealeaUuyQqkYNndCgJ36AMT9MSGbwJXOamkCU9UViiLbANGsuoj1Nmf74WlDR0tQ0xqKuRmKizSqYltY3UAA9fO/8dr8S00NPNeSOSVhmQrtI0yCoXjbhb+cy04URrU0syzRSEfDY+IfIjt6DAUPANQ0NTLU18SSRxloUVbiQj9W5I0k+diMXXJqKtznPJKGgliWR1LqZbhSB13AJxYh7P+JQgUVGWX7nmv8A7mOYe1+GyY8GpWDCbwV1cbkmHxzw+q0mFk3DXAdXnXAEfFC1S0sslfLRyUUy2aFk1blr7/CNrd8M8PcB1+dcCPxCZhS1EOZvl0tHKu4ZVDatfluNrd742qLgripaE0IrqFaVnLtCJX0Fj+tbRa/rjycDcSxUppYa2hWnaXnNCJHCF7W1EadzYAXwTfKTgtX9qasB7LYP6B9qxbIvZ/XZ1wNLxIhamqaavejmoptnjZQDfVexuCNrd+uD6z2aDLuGYuJDxJSySrGJWoeWBIN7Fb6+v06Ysucz5lklVW5DU1alY5g80cTHls5RSG3AudJXf0tirZnUrUIzKwsR2O2O/S7R18RSbUY4FrhItuCh/RzBsdIBB9aCoWmaMW2Ww1EdrYf4XzMHLS08GnROSCCCWGxsfP1wikRm1oLsRuLAYF4Kp2l4dh5qkBXdTfqLW/vxynm0rt0gdUKXzKWizGqVYqempogSzFWsXYbgnV1tfDTU1OsJq1POZQ1wi27bDfrgeeKFTIRMIApUqbEhrEG1wbjpb64dlkZ45UikCtyxpIF/FbrY4QHSVtIUhQokeXwkadIHhtvcdNsAcNCem4ciMsTowibxBD4elt+1+nXB1NCZ8tCysAGY3I2Nv3YUcxrY8ggy9DE9QVeKodgbKL6gV9drb+Y3sLF7XafrWV7dZB6FBxz8yS7nU2wU3vjHpMonynimVQW5bxuY3YWv4h4T69Pyxr1MiJIFlW22oeZA7/fC0aOJ/eBGqVG5EgTxqAdt+uGYXEcCoHxss2MwnndE05iU1RVDx5JRSMjFxHe19ulr/njHOIeG64cRTtLIEjqWedJmUhWJa5X53PTyxtE8sYIDyDUOg6/f1wIk8ImQ6lBF9JvZjf1wIrt4xe4SCdkVTCONEU2GCBErFG4eqABepQ9ein+OLx7NUr6RJqWaeKSFbGIBhqUk2ta/Te9h640SGrp41udJPfxgE/nhr3qmEutIV0A38z+3G6pisIWkNZB9aw0sBi2PBfUBHqXMxQCViHDEqt7C3bA8KgDwqpLbXbt649NUqrDTdgNzcbYZSqQvuLgt0A2HzxyXVBsF2BTJMlRvu6+v5fwx7BP6Zg/6wf8AxDj2Falq0N6LQqHMqvLqtqihq/dpmUqXQAmxINtxYdBiL4j4840o6hlgz+dYggctojNt+m6/4tgarhkp61KyOT8KJSJVBY6l2OwF7sCP24guNph7s80cXNc7qSbdB54VWyzCYl+utSa49SAfvWZznabFX7KuNeKJ8vpZjnc8jSRKzMUQdrk204XPxhxSFLJnVR4jtZE/3cZZW8RVuX5bSxUOXSVwNOqs8codo3HW43uNvT54gH444oFU8SUdArq1mheS7qQCTfcdLb+RwLcgy0/4dn2R8Ek4rSdJW2rxrxXoLNnc9x0tHH/u4bbjjitTvnVRsd7pH0/2cUWl4mlp8npKjN8qr/eZ2KFaWAutxvsb/LpfBHEebZbS5MzVbmFauF1iEgKliV6W6g74v5Fy0uA83Z9lvwVGsdMgq5xcccUTAmLPpX7eFYz/APLj389+K2uDnVQlj2ijO3+zikcCZXlmS8IZTmFNRuKjNEQTuZG0atLHUAdge22JwuCxTaxO5JxmdkmWG7aDPsj4LZTe5salZ6biriyeEuc9mADbFYo/3rgheKeK9QT9NSG27HlJqt/s4rNLVPEixLEGW/6u298EJOUdlEAVieo3A+dsIdkWX7+bs+yPgtrKlMqwwcU8Sl7vnM4UX+KKMXP+zgmPinP9JY5vK+9vgT/dxXUqRIAt9QHxW3/aMeWui98ej0yc2JVYjlkXDXsd/kemBGSZcf7hn2R8E2WhSOb5/wAUVcM0D5hJNSzRFHiMaWYEWIuBfECYKjl2kQEHptpscTCSho7cohrdStiT98DNIVcpJoB/Vut8acPgsPhQRRYGz0AH3IiRyULJDNHJc203va/XCqWKohA1crfoLfPpviTYMQxKRnvcJ1w28hAP4S7dfCca/BIJEyh3V3K21r4dgqYHpIpYnYlSrE3uRa3y88SYc6dIWEDT/Z3+WGC4sbwqttiSMSCgLo2UbWya7iSRlXqQRucRsiLI5CsfCPIjvg2pOtmVlVrtvYYbRABdlvYWAsdvrbGlrYC59WqXGExHTNrJkCjawF77Y5JQKZEPKWQ9D/SClh8tJv8AfBMcZJNgu3c3/hhqpqRS071EsV44/FIdjpA3J37W3w5uyQTC42S0ck4ZzUsB1BnNj9gLYX+g8pSQ6ecL+dSxw/FOnLJVSNe4Hp2xwTPu2k+VguLlRcjyDKma4ecb3tzTbBMfDuUuGRpakbbESfxGGUmutgCD3IFseWXY/HY9P8Xwt0mwTmFoiUUvCWTuo11U4W19pb3/AC64aj4OymKy++1bb3N3AJ+y4QlYUJQEkH74WcwdRpvuT2A/jgQHBG5zDyXm4QyhihWsq10npzQ19+9xhUfCuVlWT3+tN9z4xe/+zhr3yX4FAv1NrY4Jqj49dh52wbtSoGmOS2H2I5VS5VkVelLJJIJqvmOXIJvoVew/0caCcZ97Dpebw7WHVqIqyD/srjQe2Pzj2v8A/ucQT1/ALaIgRsozMc8y+gzvK8mqZWWszQyikUISG5Sa3uRsLDzwNnPFeR5RxDlWQV9XyswzYuKSPQSG0je5Gy9QBfqdhiue0ii4gHGPCfEGR5E+crlZrOfClTHCRzYlRTdyPXp5Yp/F/CXtD4ozDOeJIYKTKaqKGjGWUM+maYtBaeyypIFTVMWUkg3CjpgsvynB120316wa0tOrvCQ7UQBG8RBNtpuOSH1HiQBK0mp464bpspzzNKiteKnyKpNNmBaFtUcg0mwW12vrWxFwb7YIm4vyGHiqh4XlrDHm1dTmoggaNrlAGO5tYGytsd/CcZ9xJwXn2be0Glnjy8Q8P5zJQ12eK0yXhnpQ7BCt/GHPJUkbeDAGfcH8f1nEmacY0yUqzw53T1NFl7KDPNT0141Al5mlA6STEqR+udx220coyp0B1eJbPpCzjAANvpaif4QNplCalTkFep/afwdTZxUZZVV1TTvT1fuc08lDMKdJrgaDNp5YO46t3GCOJPaJwrw/mc2XZjWVJqKaNZar3ejmnWlRujStGpCAgE+K22/TGe5twzxpV5TxrwnTcMA0/EmcTzpmc9ZEIYIZBGuvQCXLAISBbrbErVZRxbw1nXF65TwyOIKfiLRJTz+9Rx8pxAIikwcg6fCCCt9icF8k5WABxLxYa239CTMQ2JdY3MKuLU6e5aBnvFOS5Lw6nENbVlstk5XLmgjabmc1gqaQgJa5YWsO+IqL2k8IS5PmeajMpUhyoKa6OWlljmgDGylomUPY9jbfFbzzg7Osv9iGScL0EDZrmOVvl7OkbqvM5M0bvpLkC1lNrnyxBca8K8Y8W0/FmdycP/oyorMlhyuhoPe43mm01HNaR2U6Ft0Aue+F4TKssqg66ttZE6mjugtAOkibgm+wi6t1SoNhyWrcL8S5bxGs7ZdHXqKcqH96oJqY3N7WEirq6dr4mdvXFN9mFNmFLTV61+X8RUhaRWT9MZmlY7bG+gqzaQLbj1xcr489mNGlRxLmUTLRtcHl1Fk+mSWyVnPtxppKjLssZKOOo5dQzEumrR4DuPI4yDMaKkq8tGXlJFTkcuVjKzM5N7sWJJB37W6Y1/26cr9F5ZrLaveH0he/4bYx1J3Z1RUI1A3PTpj7j2ExH/wdKn0c73lYhgx5w+qT6QHuQrZBka09KgOZwvTKoSelnijYEADVcRX7X3JxJwVEkMKwiaSRV21yNqdj5se5PU4QEJTUWUL6YdDQBiNWx729L49iXlwvyVCgKfo2lJSfxHSxFx2a2FS1DKote59euE3hGnSpI9emGnlOpiItNxaxPTCplGBFl1K2bSwAHXvvhLtM7BybE9R2x1Ucr4iVYbix2Iw2xBYFpCwvawPfC5RBjuqu3sd1jjSE9RyJLkH0GNyGMK9j7g8cxaQbGnkuw3HQdTjcwcfEfKJ/9oJ+iPvK0saA0JXbHGG18eB73xxztjwovdGvl+qsKmoIkuOdJ8/jOJng6NRn+WM5N/e47f7QxENEDNUFlLAzym+q5+M4k+GYR/OXKCOaCayE2I2+MH92P0vWj5PdH0D/AMqzsu4K5+1qp4wh4jD5bX53QZRBRRSNJluXx1JZi8iyXBVmJUmA6R1UyHti98NjMq72f5cKx5oMznyuITPIAJI5miGokAAagxPQDfGQwQ1Oe8fyUNTmXG1HkuaZjW0VJVRZ+NDTQ8xpE5IS6R/hyBdz8I88bllFDTZXldJltEhSmpIEghUm5VEUKov32Ax8UzvThcJQoFo1iDIAuLi5vedx9yqn3nE8lhZyrOav2f5Z7OYvZzX0ObUckNs0YR+6QTI4L1aShrljZmtbUdRGG/aSP/f/ADwEAkTRW2/+BHj6CPTHzv7UqinpePc9mqJY4o1ki1M5sB+DH3+2PTdi80dj8zcCwDuvduTdxbO5PTZWWBgmVX6jUrhQH3xxioIDEgeZYD9+AZ+I8jiOmXM6PVbtIDt9Mcqs7ymKYU8tfTxsStlY+YuPyx9cDCRsk8Vp5o7mKZNKzN89VrYWSqRaQYz3O5N8QlTxRkYhm93zCkmnjUssfOClyB0uemDMjqMyro6maqykUkMCg81KqOZG8RUC6dLkG3nY4nBdur47QYlErNEWKkra+/UAYWJ0YmM6So8r7jCKhoYomaaRIwou5cgAfXEPmWbQrGkGWyUlXU1EnJhRJAQDpJ1GwPSxxYYXWVVKgaJlbB7EHiPE06Rg390Ynxf6S9sbMOmPmf8AkrvXScd1xrqypndMq+FyCikyqCfCbC+nbvYHpj6XGPg/lBZozhw/hatVKpxKYK63yx82e2bi/K8t9qWZZZWTmnkjihbXJtGbxqeo3vv+WPpMnHyJ/KIaY+1DieFaaSSOWlpSdHwExqjEyb9NLfa3ocafJtRFbNnT9A/e1IxlZ1Gnqam6Pi/h6qzSKhhr45ZpQQgjDWJ+Z2wXDxJl1RTNUU9bE0d7X5wG463B+eKPS5BS5lPUUtE2ViWGvhdlo45COUEVWVJiem7HT1uNsDcOLlmV1kyZhRQ1FNFVVMSRvQwyFyjbKHbxEm56nsMfe/Mmk6Wm65bczqBsvbZX+l4hoqmFpEr6ZkUAsfeF2ubC/lvisS+0IIdsrqgglMbSMbLsTcr5kWO22EyDh2WRqc8E1sYZXUtBEurdh00NvY2te9remI3OKBKbgKirVgiQ1uZVixw320xh0O/e1x3O7Dzwt1BlMXRjFVKhsbKw1vHFBTUyvzJpmkGoIEIIGm+5Ow7dCeo88M0PF+TPUP7xVw084OplluAQQLC9rX6jYnpiLqxWZpmgrauDkPyOXUU0a2ZEPIpgqaksZCCjWUk772xT+OjEOJM2SJaoD311UVH9YADazA73HTFsZO6B9dwETK2aGeCqpoaqCUSRyeJXDAgjHJ521AAKQTtbFY4BLRcKFBAEkMt9J1AG4t3Fr+E3A8xiWzytTLcvlrZVE0MSgnQQb3IG3rjO+mQ6FqZUGgFyO5zKH1rcX6g74SSpiV1tvvZgQcUUcfUC3Bpq3Te48K3/AG4neGuIKbPBJ7tzEli3McgF7djggw+tWKzXWBUvUU8bnUrsNxsDtgSSnKDUBcncaTfbyw5eVZmBVyB3C4VeUk6lJGxG+JBUkG0ptFLFWJ36acPszKpKpY38sEKhUjUmo+dxhb6C6gqGLbm46YHRKkwlU0q1SLTVB1Lc7XxGZtQz0a8xLSQnwmRU6d98HtThbHUo22t2w7Tx1UOoxOWFuhG2Be2ytrr7qDy/M5KTZrNvcKxNhhxs+yq1mmOuWQksD8Nz+4bYPzDLaWtDyNGKSYi2pEJW/npva+KzNwJUVU+pM0o2Um9nR1LfliNZKI1SNlZqLiDI44gjZlGulbEsxJOJvhzPOE6HPZM1qZqaprxTGFC8lgIyQxuLG+6j5b4oMPs+zanTmUuYZVPOOgYHr82XphUfAPGAgsJst8YBcKQN7+enpbDWt0GRukveXjS4WWw1fGnCVdGBmOXZTXIpAKyUTH/vAlCAOo9e2A6rPeBDQzih4ey6gmkIVZUpkj02NzY6FPbGWRez3iw7yPQXv8JqPCT/ALOEVHBHFsUd46elqXDKSizKQehIIawIxZqOdYlUKbAZDVN8d5fkvF2XJJSTwLVUz/hSbgXuLox8jj6EpK6hqGVIcwo5Adhy51bt0sDtj5nXhfi4xiP3IUwG40GNgSN7AByR09cTmXTcVQQNGlBNUzqTG2nwqpHe99vlh2HrmilYnDDEEE2X0W4CIGWoRHvclZPLrgVkoapw/J5xvquVUt8+mMEo6ri8tzqykljCtugWRgOvcKR9b4VLxDmlAdstzGrlC+JoonAUHp2uT/DGkZgRuFm+S2kWdC2rPhGMsrag2i93pnZJCukrYdbjFFlzBQU0ssl1O9jt5YoWdcTZxOIlqKLMVE2682JgRb0xW34mziNnBmlYAknTH2HnjNi6wxEQNlrwdA4WbzK1v3hKlRrsD3W53xG0ckk3vJBqNSzsrqzfBfxW9BYjrjMn4orxGyrzkkdbIAtlXfrbuRhdPxfWEOXqZSl7XCbem9/3YytZC2GstXJsCVuT88Ntr1C4Olj18sZs3G8kWpTr2Xwl99R+hGOfz5MsLCaF3HfSzDfyxZYeqsVQtHdgGYKt+wa++ENTndpFHQW1NvbFLouOlZA4pjqt+s5NvXocFfz4iZyJKfcIDdFOm1tzuB9sVoKriCZVnSn/ABioZgFFwL3w6ui5QqQbbWN74rkPHFK0jA0zx6CLk7E/TD68aULsqosiWO5dAB874sM8EQeOqskbgRkMhuq9QpBxUsqkp5cw4iZQuvSFJJuTb6bduuFV3GOWCnkZJwzMCLA/GB64D4FWdsgzuqYsedOxQgjTy7KdrHruevphZaQ1WHy+ArPCNOWBNSowjvbrc6Rh3J5r0ILdCxA23GGob+7oLliUFh5i2F0w5VM6ohJ6/DawOFgpxEiClVMyMZYpCBdLDxWJOGQVZuYxkOhx5gHCmhWQe9sVAVfy7nrhAHNL2kCrtaw64NxMWS2ABSFO0UaNcbC5ALYTUqsgLiPcWYAbAm22Ip7K2vTIQNtxe+Oxl1a8moL0sB/DFNLuaCoQDIXqmgo5QHkjUOCNPhBIO+97Xvv1v2GEUuV1UmZTR0Wcigj91eZjKnMDyKAAg28N/r8sO84qQViaw3BOA481pp5FjeysxYBGJViRsbDb/BwwAm4Sy8DwWqewWkqxPnSZjLFUxlYeUOVa273vfr2xrEdNBHskKKL32UDGVewSXXX51Hy9BSKA9PMyfwxrI8sfAe3FR7c6q/8AD/yhbqcFghI5EN/6tP8AZwgUdKlylPCpI3IQb4fxxumPINqu5lGvnPPacSZnmtK51BqupDC53VpGFr/I4epaUiKMDUxRbBmYknbzHXCs7qdedZmgBU++1A1Xva0rDASV/JpygIcpck22J+mP05gG6sHSn6I+4JLaha5LFXG0ERijVvCANtJ9RbDLLT0+f0eYCXlosUpe5AFrAm/ytfBFJPRyRRfgwxkICLbgfI9/ngiSSlMtxpbawFr40BpButDnNc2xuqVx5xfwzmfPymuHPp9IdpEUEKb7G/n06eeAoOMeEsthByUcuCSMHlCEhi4JubE/4ti3VFPlMhtPllNqB/zam/1thBybIpm50eR0cs9wQzRrf79saQ9gELnOp1HO1SJVWrfaHkRC2q3BAsbRv+e3XCaPjfLJJYyrVLRklbqtr+u+/wB7Yuv6IiTQn6Ko9R2ARBt9xiFqqGlEviy2nilLEu/LEg07fD0t0wUsOyWW1AhMg4nHEFTJFRVEulFIKvGOx64heO+M814YzGnpqSSmqVPiqInjJ02sdN+1wfmPriVmoa2eZYP0hJSUai9qRFidz2u43v1JAABuPLA0fCeXtQ1or4RUVFTqLOx1FL9NJO9wLb9Tb6YjWtm6jnvDYG6vnsK4zoOIePqSCkWSMtSSuVt3AFwflj6It54wf+Tvw7w/kktFyYUTN+S6yuFJeSw3Zje1jckAdNgcblWVC0tJNUupZYo2cgdSAL4+IeUmkBm7WtvLG/eVqwr3PpS7xTw+WOnpimZJ7UOBczyqirW4pyWkkq4Y5fdpq+ISxFlB0MNWzC9iPPFyv26HHg6+Er4V2mswtPiI26J4cHbFfMvtgzWGH2mZ7QyV/KbmxMFs11/AiPkRb+OKjRV9FU3Vc4pu5Kl+h8unTGp+0saeN82L08Lhnj+NBcjkp3++KlVvRJHJE9HQ2aMnwxKbfS3nj9Kdnng5Xh7fMb9wWN1MyVFwZhSUq6HzSnYA3sCOnzvhGX5vlVNTSQSzKqPM0n9Yb2sBffzt089++O8OVGTVacuXL6L3uD8OYPSopLAbsBvYdbYlJa/KUYpFTUms7k8sab/bbHUqm9irYwtuFEz5jlMjFqerSUHqDb/hhqrkaNA0Sq2rcEE9OmJmNqKokYx0UZsfiFMtm+tt8cqaTnSs8cSINh8IAwIATdRRtARNlSyttYAkX+mI+eQRZrp1ELLe4bexxK5fCI45IQp2Um/b6Yi8yeIu6OhDDYHFuNpCCBKTS01IlTLUaI2lbd5W8TAeQPYWGAanMZOfImkKoO1gd7/tx2KW6tJIxUhTcL3wJQoZCAx2ILR6SLse3UWJGx+nXBMhxhA+wsnpjK6obmISmwYrsPI/LA8bs9RJCrFpF7KCf4YnM3yZpWikmrZHJXYkaVNl3O1wLdN7dcRiyGiAmMitewII8R7AdiLYN9OELXlO06z61ZmIUi3w+I/PthM5DPpc9Da5Xp9sFxNDUUzOjsHsW0lT+/AT080dnccosDbU4B+YtjLEOWgullt02HjLgEyhfhDKm38d8KgoKmGYyLCrr5rcMbnfr5AYcWPmwqy88SAnpIb/ADBvcjC6audKhoY4uVbbUW1MAL9z8z98NLW8kpr3RBQGmj/6rXf+JJ/uY9iP/SGZ+UX+238MewGgJ0rQ85nSmoaieQgJHGS3i09vPFDpqGrlypKqZ5B7+Vkjj5DlIEJNrsdjfrvbri6Lwmc3rZqvimpVaEECHLqdv663eRvLpsMO8WlZFNHlmVxRFlCh40uwXuAe22NBIhYiHEqkUuWQrX5g08cE+iOmKlsvaq7up9EBte24PTthjlpBV8QQJRM0xeoWPVZTq5YY3W2w1P5+eLxwrwnm1RU1FQ1XVUSvHBGXpyQzGMuTsbjo4633wRLwFmUmeZrUSxtLDXlz7wYVSckxotyRYabq3hFtm+1VXDRAKBlFxcSRuqt+kKuvq8jmy2CimnWpklp15J0kqiub31CwC9PntfCvaPUx0+VUhzSCmmzGpmvT8uPXEshU7lXvdRexv57AYn6XgKtoUpopDGkFOxdmSDSzExOh+E7m7jc3sAPrW09mPFU+ZQPW18ktKC0ayShpHVWsL2IXsOgJwuiddTvlHVpmmw6Gp/Lq81OYUmUTRiP3NFeCNLRxMugbhbDxXLWtdbYsiRhrKFXzIwJT8KZlT8QUtV+j1CCmeEmONQFUG/Xz1FRYAbEk72tKz5dmKdaF/mSqn7E4MhgsNk6m1+nvISSLupNx2DWw9Ts2oAozAkC+EyUmZ3CrRzKbgnwrv+eFEZnpUigqG6WAA/jhbgTsnUy1puizHEVA5S9eoFsDVNEjTiqhVVqEIs4vcr3X5G5+98MzNmsRUvllZ5nYD9+GZKutDC9DWtb4rKthhIpOC0mswqTjlV0EpffcaC97HuDbvfHXuVsEW/mN8QDZjVpKkooqwRt4XEh+A32ItcfngeXiKeJyrUlWWJsF5Rv+7BcF6A12hWdQEYjWL+hO+HltsC33xS14rl13OV12m+xKD/eOGW45u2lctzC5P+b3PyH3+2LFF0pbqzDdXiR4wL6rgddsBVdRGNwQd+hPTFPk44gRC0lDmihTZhyG2wPJxllDkFpKlNXTmRMBhgovS+OxWZyl21i7HybDTDuitYdTqxD0fEmSyyWXMV8rXvg6nzjKZCCMzpjfa7SKMEAUl5adkbTRoTZgNvnh16eNleN4w0bLpdeoIPXbHIMwy0j/AC2n33FpAb4fWeCXeKoRgN7A4MdEGkESo+OOGnlFNpYAi8QKgDv4dvL5dMFRR2tYKb/PCpo451ddS+d9W6kd8dpWR6VTNZGvpLfCGYbXG/Q9Ri5IQhoKGaE80hjY3+mFrC5F1ZQfO2CXRNQ0sOnY9cIC2X8OXe+5DbYGZKIMQ0lM1gOahv5YHMLIwXUwsO4GJRgoB0vJY272/bhFiQTqfYX33+uLv0V6fFRxjfVcSA3HTCuQ+j4ybXv2xJ6gni0Xv5HChNdblGAPkb/XBS7or0Njdad7BAq8NV1gb++m/r4Exo49cUH2JkHh+stqt72d2/1FxftuuPzf2vaTnWI/m/ALoNs0LhAxSE46q6nNalMs4XzCvymkzD9HVFdDIhYSghXZYviZEY2LehIBAxdze2KP/MGaLNal6DibMqLKKvMf0lUZfCqjVNqDMBL8aozLdl73IuAcc3LfM5f5z0tMxvf0bzG3LqgfqtpRFNx1RTcX5tw+tFMP0fA8iVJYaKh41jaWNe90E0d/9Y+RxB5T7X8rzGj4beHKqwVmdVcdPLSkjVRK5XTJIf7LB4yv9oNt0OH6L2VZfSVVLmcWZ1n6WjqqmoqatnZlqBUCQSKYy2lR41sQL/hrg+H2c5XBlGQ0kDpHVZXJl7zViwgSVgo10xh9/K/na+O6z9H2gTqdsOYuAQT6jYxvysJSf26Eo/aS2ZwZdFkvD1TX5nXe+SLSCpSMRQ005geR3bYAsBYAE+L0x7L/AGrZNVZjw/SPQ1lMucrMpllAtSTRSiIxS2vYmQ6Ab2JK+ePUfs2kyuDLZsl4hnoc0oBWRrV+7JIssNTOZ2jdDsbMVsQR8PrbHaP2WZPFTQUlRVT1dOuXVVHUCUDXO88yTPMWFtLB1uLDa48sET2fkzMX21T86N7R6Mc5mbKft0LUe1ZAuWCmyQSPX009SBPmMNOqLFOYSNUhAJJsbDt8sSOWe0NK/jSfhyDJZyaeqNJPKKmIyRNo1azDfXyt7a/PtbfEOnsleCnydKfiFXmy2imozLWZbHU84STc0sVckBr7X+fniUq/Z3JW8X0OeV+fzVEVDXiup4mpY+cjhbcsTfEIrknR9L4Or8gyQw2h0HvzMmLbbRzHiqHGi/4K/X2x7Hugx7HjIkrWs59uJtl2Vbj/AClup/0DjI5WjiA8R1DpjV/byAaHJ7m39Jf/APNnGKyVP6QzKXLosurpEjSTXUcsiJSq3ILX8sffvJ9SD8oaP4nfeseIqaLhNz8R5NBdmzGlQgkm0i3v5YqUOS1fFXFWdxR51UwU9HEtUqqWkVrqCAPFtseuM6FO4hVliBOm/TrjV/ZNTzx1ma1Q004qMlhSnkZggMgjVWAPZtWPa16Zp0iW7rHRfx6oa/ZT2b5dV0U9TwpG9W8iUStDmIJJ8QsNXTxXVum1vXAPDc9Rl8tRQZpmTV84VJbG10UrZrefivY+Vul8TvGub1UNJBUZczVtQqIsgSbSG8NyWJYAi43O+K1kdfXVVV79mVJl9LK4MavG4Lqt72L3Nx6DGegxz6YK0Vi2nUgKdqMwgjLaEcLbe6m49cDDMkEgUR3DG+plN+mHp6uAXLVgbe4OrVpw0tdTl7mqjYA2OphiNp3umPq2smeHs2zSkzMSRV1dBohEcjp4eYb9VZTt03Ft74mYM9zVq6pqJuJc+WFwpRffZbKR1A36EW+t8QTVVQMw/o+dZcsBlBZJbARRW33BJZuwt57+WCaXPaP3ZDWV9Akm/him1KPqQO3pgK2EoVnansBPiAVn12iVMVef5nKscsPEub0yJJciStns4I6Eh/lgOfiGvaCd4eLs2c6wy6MzmGkGw0ghiLbMd++BZ+Lcmpl5keYwOykXC72/LEbWZnwxUDmz1+qRujanbxfvvjO7AYUW4bfshG08/wAU7l2Yw09TI0lfU1E2pl5bP8B1Ektc3v2wfwTneSVftOoKKqbMTWmvpmhtUvHGtithoB0vc2Jv2vgel4ky+JKcU61DcpSABSOUJLFrHw30m9rDeyje++DOFavh/wD5RMilossrmqPfYIRUy0p3vIPGWYa79QCTaxG22LxQ0Yaqf4Xfcru5wHityrfZB7Pq6vmrqnI5WqJp3qHcV9Qv4jklmAEgAuWPS3XF1yyipsuy6my6jQx01LEkMKlixVFAAFySTsBud8PjpbCsfmfE47E4loZWqOcBsCSYW1rGtNgvHYY+ZPbJFlNTxtxHBXw1blpI0ZlnZVF4I7G269x2x9Nnp1x80+1WPPH9pefJQvl6o0sRj582k25MdxYAnrfHt/JqdOaPP8B+9qViWa2wqNWZTwjmD0RGULFHE0pmNMpvMW0BQd/1dLeXxdMA5plPDFfmBgWjzeJIqUrCYiouqvYagVv+sd/IYn6nL8+eJXqc/wAppgD4kiDP9jhtsrrAgjfiuNWYkXgo9RA69Scfb9ZmZWXgADZQS8N8KpltVBJRTJUMjcqaV9RRiu24sDY79O+ONS8O0nDtVlNLLHHPUoVDSz6lW+k3OwN7opFvL1OJv9DUOphLxLm0rGwYrGqq3yuP34fky3JUhJ96zdpPJ6kBW+2LFQzuo3D8wFCzyZRW5FVUlXmFSZ5G8Ai3jVQqjxNq3F1J+vffAmWycIZPSw0s9dJq8MxWCd2AlUW19PBszdCft0snu+WRoSuXtUSAhiKirkOr08uwwmjiyiYNP+gctuQLMUubX2BPc4MvkXSxQgyFfP5NOYcNzcV1lLkM5aT3FpKm7MxkPMWzEt0tqttYeI37Y+gwQO+PlOKakpH5lJSUdLKqFQ1MuhiuxIuD0uB9sFU2dVUKsRX1YVtyOcT9t8fOe0vYt+b404llYNEAQRO3jK206mloBC+oiwvj5b9ts9DD7Yc1M2RVNfJJBBHrEAMY/CB69Tt+wY7FxFXwo/KzKvYMCADUudOGJs1mlAkqKqpfWN2aZjf6k4Lsv2Qq5JiziHVQ6WkWEbkfBVVcKghQGX5jWwzNUNlObySykF3eFgr+Rsq9umHJqLM56ta2jyOujlNRz2Mc+kMbWK6C4ADWBO1yQMTEWZaCQ0zyHsvMJ2++GZa5WmFrr2tYnb1x9Aa4gyN0tzWuAB2TBqeJ/HyuG+VKrb8yZNr9B4ST54Ny6q4lyvKqWioeHKGP3Yy8mR6hQy81tbi/WxYAnvsMMFoGIflqx7ER9D59LjHXm1WUxkj+0Ad8U9zjuhFJgNl6AcZe+SVxy3JObKLNrcTAbk3F7gG/W3kPLFN4i9n+fZpmdRm1bPlsM08zSztqKi5uSQLbYuImVWNoUW57jphtqt1Y3sbN1ANsWHlrpCF1JrxdZi3D+ctWtHTCWahUrolhkfksbC+kg+dx8xi0cPZZWU3D9bBW00sxldwsUpJFiB574dqXiyWrlrAkj01S95Ro2iJ/WHp5+pviV5iOFZC7Lpvtgn13HYLPTwzGuuUXl8zUlFT0kvDGVBkRYxzqiPU1gAOouThWaUqRMKuh4Vi96aNAXJjQhv1gNN7r5eYwL7xT6SXSbWT/AI6YZNRJcB1kAPS3/HGcWMytkMIiExLmGcLGqR5DVx6d92TxD6kYQM1zctaLJaqR7DpLGfp1wUZpWUrvv0DHHVqJYlKltLAgXTDRVIMpRoApj3/OWjDjIag3vYiaI9/9bDgnzu4K5DVnz0yxbf8Amw9JWsNZeTVYdiNh98J/SNO7A76v1tsXx7bKCiBzSo5c6HXIKwr1s0sRJ+gbDtLmFezukmTZrGo/WNMWUnyFr4YgmpP61URifNAMeGcCKoaiKBCRqUBdmB/fhfE8FOGOqJGbVLhg+RZoN7A+6mxxxc1ZHZpMtzNANgzUbdPtgaGSJiq2HgJKq6nb7YeRoDIrae1xfUDf74MVBKgpIpc9o4HElRzYFWxu0bD9owQeJMvlLtJnVLGDsodhqH52xGx5lMsjotTInYC7G2HXzIsiR1Tc1fJ1BHz3wYe110bWEI+mznKJFCDMKae+9udYnBwkppG1mdAo6DY2/fiDM+XVEbRyUtI6ld7wqL/PbDfu/Dw1FspoXJO55C3OFy1TSSrMsqS6SsjbbgiTY468six295Y38n6YgkiyQIhTKstYr3anW4wkx5UE0plVEhPcJpt+eLDgr0FWFHnshE7n6nrh3m1KMRz5d/Mk2xWGgppmULHJEoFvwpXS/rscPe607EJrrkv0ePMZVP2JIxeoFWKbjsFZlnqggX3h33vpYYGmmrybR1DAf2QvX8sQQy4xgvFnGbKF2t70rf8AqBvhHulcJbQ53mNiNnliSSx+lsVJndTSeYUvJT1LaTIkDEbkvApJP2wh43ClRS0BCbhTAtv2Yi+VnkUh05xTSgbjnUbgn/ZOPCqzh1kMqZQ6gd6iRLj5FDi5PVVpEbJ6rpo6tuXVUGXyC+pQ8KkA28scjy7LVTQ2Q5cwtpYCIC/2OBY6jMBCXbKKdtO40VkZv6WNj+WHFzOriUSzcPV8YvpAVVa/qLG5xZPigAaNwnJMryRy/N4fiRiPFYutx5eE4G/m3wzUIQcueAtfxJKxt8g18FJnVPTpdqHMKcM23MoXQD8sP/pzIpXKtVxQsbC7qVv9DbFXV9wqIm4KySWTTFX18YIsG1L1+2B39ntGAFp86mVivikZAQfqCMWcHL5gEhrY3BubBgP2YeSCIgqhDAiwHX88WCiFMG6q8fs/omdTNnErX+KyoAwPluTiakpaTKMoagy1QIXWzKCCxJAGrc9ABgs0QIBBIB2spIHyx6emhSNXUAael79xhbyTyTGU2h0hNUKlqE3vsNj6gYfLhKRJWADMmyk73wqmjDUA8IBPUkeXf54ZqacmzqVswuLm1v3YUAnJqdHejkDEIGBBt13wlZRTpfTzFAFwNuvpjglnEhjCH+2x1XK+Q/4YHqnmqabmQSgxmxuWI1H09MFBCV3RIlFiVJYyyqQx6egx2MRnV4gLDpiHj98jnBeF1X4SwOq/zwRT1XNQBLoQLnUtvpgp6pRaOSk4AChUMoviOquG8nqcxjr6qi51THblyrM8ZQg3B8JF7euFxSuXuU1eg6jbD8NQ2sEE/LyxYcRshLQd1p/sIo4oM3z2p5eiaaGmVz5hTJb9pxrQxg/BGbZjlXB/HWaUBtW0WVGppiyarSKkzLt0IuOmD4/a6ldnZy7Lc0iczZxltNR/0RwZIJDGKkklbCxZxc2tbbHxntTkmKzHNq1WiLDSD9lvh4+49FpFRlNoC2rHmvbHgbjHCdsfNQei0L5tz404zzNIkD6mzGpPmbmVifpgF4ohr1TKVAIYnY7+o74kc7MMmc5sY0syZhVXutzfmsL+eIqWOnGm8gZjvpJNx89v44/UeWNnCUv5W/cFicYXBWRRsiBVsdySLkYMgkRiSL8vbxYizCiSBdKHb+zf74dicwsEJFmsLA42GytplGALKDrhAJawHYDzwTSSSQruUCjYXuLHCNN1i0zEh1vc9v8AFsGM6RGwRXW3W9rYW5yc2n3k1PmjMgJjcuLqLAEG/rgVa28qM1i0alU0qLjy64B4o4hy/hukiqaqKQrM+gCMBiNr33OI3I+M+Gc5zBKWlnKTspIDoV3Ha52wVPUW7JdbSHwTdTJDvVuVie25LWtY4E97hn1NSzCazshC72K7H7HEu00GkF7C4uTbYYp3FdP+ia5eIaCZo6Y+CsjU3AU9JFX+0D188E0yYKXUZojmtJ9jFxx7TfiyEGnlNjvc2+e2NzzKIz5dUQhdRkiZdOq17gi1+2MS9i/u03FtHUxMTencqQNmBUb43Cs0+6y65TEug3kBsUFut+1sfEfKESM3b/KPvK1UgOGIWB0Ga1eR5jlHBSeyjh/MMzihii0R18VRPHEqhedOwhsuwBuxFz0x9AkA2J38sfNWXUfD+W19PRcAZ17Q83nzarbTUQVcdNS1MqqzuWnkiHMIVGJtq6Hvj6WX4cc3tSxrXUnBsTJM6gSbSS1xMDp9d+isOTeVgntOp6uXjzOdDqsLPH+sbj8GPFU/REcjCSxJfdruQMWj2oVci8e5zTqOkkVjbpeGPFZR5i1pHUi3XcY+0dnL5Xh/5G/cFCLoLO8u93dM1oIObPGmmZFH9bH6f6Q6j6jvgmD3KopY54midJFDqbWBBF74JeQw6YwQxYW69BismRcr4lpkEUjUWYuU5QBskxO2n0a/QdxfvjuBspb3aLqxpIihQqXA2CgYURNIqiS53DabdBj0wCI8kKi1r2Atq26YFYtKwALAdWGrTt06/LywsNm6MbqXoKmOSoKBHGxVlYWtiNzZRTV50wxkMvU9fpjocRSprlcu0QOkMxFtVu/03w9VurULs6+IEfYYst7qH5ygkIaOWMW1te2+9jjqxyU3I8GuEKRpF7rtufL749VvDJMjIgtawddiN+/ph6GMtT6WL6tIY+LY736YUDCZolFw1DogDtzk0mNoz0Zeu/luAdvLEeDKI3p9OsNcixFvW3rhwQV6znlxidJV8RKglRb7jCGKFY4NJWa11jWxe3n8tjg9ZUNMJFK3LuSpHS4ZsKr5Jni1RsRYWPr6YbYs0ZUIVYdS25/ux1YJpI31ta262Fx9cUO8h2Cj21P4SI2AO4OwwTSOfewQqlTGFAKL8XpjopzrQvywSfhBJubb2wZT0YlkgDyOkW9uWx1dD+zbBsbcBA7aUB+hMr/zA/2Wx7HfeKn/AK7Sf+O+PYbw/FXqK1U0qiTUUQr1IIwLC0EFZIq06tKlriQE7HoRvhiWesYeMU19t1BA+xOI3Mmry4elhpGnQWEjORqFwbdOnUb/ADxBEK9cbK2x5hKqKojj23tpJ+++FS5hXFPAEQAdRHimRZrm7x3FFRFCbHTVv2/7vbBL5jXqulaEL3vztv44EtHREKk81MSCsOvXUSeMbDSLrt22xxnkPxOpuf7F7fniJkzSZN5KVmIGw5othqTNZlQD3ZiCNvGOvle2IGt6Ka/FTDtOGuspvaxFyAPkNW2GZZKhiENSCVFrm9/24gmzmrFycpZh2POX94x6DNpZNnyyWMn/AE0OK0DopxR1U6XqQgYzRnzGj+/HBJJfX7yFJ6eX7cRMlczgEU1Qm3Qql/yOBjmhiKq9LWeIWP4ANvrfEsbKy+LqwPJUEKBOB53Ub/nj0SVA8UVYQb91B+mIeGuu4ZiVBGylSDh2WtWMElwpB3BFsW1ltkJqjqpAT1DVHJNfCzJ42V1N7X22BG2EzrVyM5FbHEW3sovY/UnFfrq2MzrUQzoKhVKAtG2kgkEqfPpt5Y6mdQylTFLcH4bxEb9LdOu2LFIpXGJCm44K7QVOaSP6FVsOvp6/ljqCtjZCalZGHUMALnzPhxFvmcSgF5lhJNgHGnf0vha14JJaeMKBcnbE4RF0Jq+CkpDVyBhJJGytcWLbfa2GEpaqnp1jjmiABO9gbXN7dMMe8wsoaKr16j0FiMNz1LK946iG/dSwviiPFWx88k7NTSvAyyRUkxe9yUHTyI74ajyiiK6pcnoJT12hTft5HCFqZyWUMreYBwqGvkQXkcMCbAA2xepyYA08kSmUZWykvw/l+of/AHaPArcP8PTWduHKcX2YhVT7WbB610D2Bt6nUPL54cFZEVP4ka2PmNsVJRhrYUYOHMgYlRkxULe34zAD7NtiAkyThuqzo5bPltalUoOlZK+ZiYwB40sx2ubdvzGLklZGNmmjXv4mAxDZ7SwVs8dfSTwLmNOrGmcsum5HwkeR2/bimud1Q1KbeQQK8CcKFdS0OaK1uq1ctx92wMOBcm50kaTcQPtuxrtNx5bg4n8hzeLMMvSd0aKX4ZYnI1I46g2/wcEPmMSgjUCR+rttg9RQCmw3hV2DgDKYUAiq8/itvYVw2/8ALhZ4Lpu2d5/F1IvU6r9N/CoOJj9KaT8Bex7dBhceYzOxZYHCgdz/AH4gcVZp01ERcLLSx6o+IM7I63ch/wD1HDQyGokKmPiXMSCLXNPbfzxZRUA2EjIlx03wM9S6KQdbANtqFtQxNRV8NnRWr2YZ2/CGS1NHW1MuavPVGYSyfh6BoVdPQj9Un64to9pFJYk0JAHlL1/8uMqFSR0VdQF9jYnDiVWqwIVRpuNJ6jHmMZ2QyrHV3V69Mlzrm5+KLW4CAtJyv2tZHmKMaaJmZSVZCSChBsQQRsfTBx9olCFJaGNCL31T2H/pxhObWybPIc9pY5uRWMIq1Q50rcKFkt3Itbv1xaoqqGoi1qwcEX1A4y/oFkn+UftO+KFtZx3Wkv7RqQC60JkHXUkwtb6jHI/aRTMSDlkqgC/9YMZ1DKrKbMRt0x41EYuGY/bEb2FyUf3R+074o3VHRY3Wmx+0CjZdRoZQvmHBw8eOsuWPWYyPTWD+zGVGqRN1lkBHla9sOLmEbFmIawX4tA3wX6B5If7o/ad8Us1XjmtGk9pGXI2n3Kct2FwL/fBb8bRIkcjUB5bmwPOUn8r4yaeujJUh7X6Xtvj0+Zkw9dVhe1umL/QPJP8ALP2j8UHHqHZafmPtCgo1LJlslQQbEJKP4YAi9rGVsPxqCaJ97oSdQP8As4zuGujmUEbC92ucKemgkYvsTbax6Xwt3YPJOVM/ad8UxlWpN0f7TeJKDjekooIUzKiFJI0wkjK+O6kaTsfPGc5CmRZgkiPU5v71Tty5YfeLWI62K21Dpvi11eX3RlVwDbsMV3N8hraYtmWWw82oWxqIxcc5B1AsOoHT5Y9FluX0ctoDD4YQ0eM7qqkEyQmE4bypWC+61wUX0FKwjrv/AGb/AJ4IjyTISGWeDNVsOpqNQ/PBVJBFV0SvSmRkbYEH6HtscLWg0yFVWQEHcl7Xxrl0wpqYBMKOGScLpGAKGsmIP61Qd/rcYbmyfhg7w5Bcaf8ApKpwbnz8X1xKvSwu/KaRVYjYPMAfzwPW0lPSoHqJ4wgNtRmFgfLfEJIVy08kHT5TwvHERUZBAJQNtNQ5/bh0UXDy6yuRZe2rbe5OGveMrYhUzCj8zapU2PythuWuyWCYRy1sSBeoJc39bhcX3nIZYOSLpabI4SAnD2W67/EFtcde98dl92knVUpMvQE7IaRSflfDPvmUCETLLJNEDYGKKR7n5acN01dR1Ls0MFWTECSr0zKT/tWxUOCovZGykOdBATycvooWB6iMD9+HHzKcMCBTgeYhG2I2jqBPUhP0fUJckfjhEFx5eI47WT1EDgLlwsW3HOUXH1xXeVtc3onTmFbBmEjCpXlTFdIRNonGq53vYG/5YnuEa/NG4rylTUllkzCHmKDcAa1BG/TbyxWquWpelWOnoKaW6nXz5Alr+oU3+mC+HM3ny/OqCeaOk5lHURzSpF4yUDAkAlQL2Fr4zYyi+ph3tFyQQPYmMqgOX1iNxhXyxlK+2WhDAHJKmxBI0zKf3YWPbFRGPUMlqdXdecv8MfAT2Ozo/wBwfaPin8VnVakelsfNXtapmf2jZ7KZBYyxWUEC34EeNAk9stCB4cjqmP8A2y/wxkfGtY3EHFeY5zBLPRx1jIyxHQ2grGqEXK/6Pnj2PYjs/mGX491TE0i1ukjlvI6FKq1QBYoXSCbtUr6gtYX+mG4+XfSSZFB6Br3OGUy8RjUtTUE26843+Z7Y5Jl0TgPLU1bDy5zC32Ix9U0hLFQlFxxoEYiNl7WINz+WOEIsYedVRfJtr/ngGLLoEXU8ssjG1w0rkfYkj74dkpKSnVnSljXfUtoxpHyxAByVlxSpZ6ZaZ2eWCMLa5MwCj574CWpp4gvInpWhlYEiFi9m+SX69vXCPcYZKl3kRSGTRpJsgXe+w23w4ZIGlkoLpd4yQpY6tFxsDboLjpjQBKzknqnFkRi7Rlm0/EvLcn7WGOSVTMgZVnBU7KYGX9pw5lvvBkaOdpCqgaTfqP4i35jHpIObbUsikEm+q9/p9cQs8FNR6pFNVSVB1inmS+41IAD+ZBw5+kAqs80BiS+kEi4JsTf4elgTf0wpbIAo1JpXa9hv8sNiHXMamokG6gBG+EbjoPpfvgDI2RtMIyOpp3R+ZGsZU7OpJv8ATbAVRXzyVAjppacX6K6sT9TqH2x2sN1cJIASt1AB64RSxqoGpF5lrhx19cFRJ5qqhaTASk/TJUIYqBbg/DrsNu4J/ecNPFnicuSSvo0DNbQtLdr+g1dMSD1C08KsurURsQbEH6j92GFkSZtczKlx4mY4a65SSIKZkesjDyTVMTqq3LclV6elunXvh2aSd0v73HCo7JApJ+d8ezGqjpnFNFoK23JFyLjzGImrEFVRvA7bNtdTYg9QRb1wDWTuo98IuagnqEKnNXCFSpHJTcH6YgV/SdBmH6Lmr5goF6ZkS+sW2U2U2P1/dgjIsyqVd6KtkvURnUGKgak7WxaKFKKUcySdkkEiOhJIClSSTsR1Bt1+W+DIAEIBL9lEGmqFhiMlbUiSVSy3QC1rdQR698OCKWBFElY7Pa+prBiB+3Ezmk0VRcwyam1EgjoTiIp259ROZLBUjsHOw/xfa+M4AK0EQmubIICj1MytcANpBAFyNtvTCaqGpkgEsWYTppI1uUjOr/y9sKniikYBG0KhI0qdwOu3136Y7GGSmZWlBQkt0J0m5te+DaACqOyAeOuZz/7TkOxNuTHv5dscVKlWLfpOQ9jeOM3+unDzSlybk74WNQOwuT2tthhAWcFyTFR5g0LSpVCSOPeQsETT37jfCp6Gr301q60+EcoDfyJ746016JqCSp5aTXDIR8a9x9sENZ2WzEAn4iDgWjqjAJTFEs0gLJPC9xveGxG/T4h9sFRUtQXYmemCjcnlldI9TrOG+UkEiCmILyuWm1b+gNvP9uCpJuYjwSOYwxsNMYAI8tsA5m8BOafFRyyVpl8NPTXAv/lO53+Rw+RVaS3uZJUaiomBP02wtWeOQh9Jt0uLDDiyeIbAD/VwbGgBAXOndMxPVtYzZZOqqNiJo2P7Rh1TOx5LUFcWsSAojbbubhsPCqToRffocORVDqzaVAB64hYI2RayLyg0lLBbJVyKwuDyTa2HI6jRZGaXwncmlkO3+ycdkqFMhtqNugDbHDck7u/w2BO+K0DorFQ8k0mZQpVrorY+XNYorMy6W3vfbbEj77FpLipgPkVqVG/13wMeXNDokQMpBvtj0EaBzDLDGbC6NoBDDy+YwBpg7q+I5qJp6ieYaklawO5EgIJw9HUViMqg2IG9+/nbEb7tRaWUUMChxYnkAH8hh+gyrLZHaJKfQD/YZla/mCDe/wBcBwovKMViTCOq6uZIdcas7L4mTlm9u9t+uFQ10rRqyxMVZbm4ItjzZLQrC8ay1iEA3ZahySfO7E/bABjoaILE2YV6wkAJK+lyrHoLabAfbfFaQEYcVIitRnDsofexGxsPth0NTMW5iaARcaRv+WBKWOnWnIE81UTsjsFGk/IAA/fAVdFmCeKnrI0W+kiSnJLE+i9v8euL0og4BTkfKZS6SyAjoRJ+7DWX5otTNJAJ6hJ4iRLDKVLDfZiN9iOnzxGf+0owy8qklCL4QGZNW3nv+zELMub0ta2bQZfLznP9IWOZWVox0Ud9QHcDzwQYb3S31WiLK3y0tI8WmXL6WRQ3Q06i/rtbCUoKJJbpQiEnf8KoeMA+YANsRdDnUdaF5YqRGw1IDH8W3XY+h+u2CqbPIWmWN3KyG1lkgZTbsenT1wEkbpjdG6OfLYTGQtXmS3upU1RkDA+hw/HSGmyhKeMyukOhFeS+oqLAE7b7DDSzCRm0iN/Um1j5YfNYyxaEi1BdzpNxfBapEKogyEVGAII1C6gQCCT0Hn548kY1lCiyaBqBI2H39cCQV5RfDd0JuC1j/diQjqqaYFZQoNrf4IxUBWCVDZjUU9PRtTZkGp4ZbhpA2m4byPY4fqajLqu1XHNC8RUDQ0vhFj1Fj62OJh4ad4dMhQhel8AV+XUEysZlilLKFu8as1h/pWv+eGcQsGmJSzSDn6phInEdWoknhIJ6SJIOvl0wJNSRq4ETyi2x1oDYfMHHBkeXncxqGG+tXZbfIA6fyw5T5TEpDJVVlyxJYzlgB5AEWwBjmqIdNtkytI2ohHV2Pa1j88PCIiMgKARYbE49Jl+YWUtWJYXvzIA7W7bi1vscNSU2YQrFJG9BKHv4dUiMDcgA7Wtte9++KBHVAZGwV79nuaPkPDfGueQRrLNQZX7yiN0YxpMwB+ownKeIZssruHswg9q9LxLW19fS0uY5Xzqd4ys7hGMSp4kKMwI67A3xJexKUrDxB+m6Wnp6GOnj5sksiNE6fiayx/sgWvq88StLxfwDVcSUGVcNcOJnFZNKClRRZcghiVWUPLzWABCalJ0X6jHynPXhua4hgoF/okkBsRoFiSLdZBB9ie4SGmYWoDHGPYYj+Is6y3h7Jp82zepFNRwAa3KljckKoAFySSQAB1JxEUfHfDVZxdJwtT1sjZkjPGRyHEbSIod4w9tJdVIJF9sfMKWBxVZhqUmEtEyQDFt/ZN1oL2gwSsJzmoCZ3nIv4/0jVEmw6c58CNIKnTJuRe18G55S1cmf5wsfLZRmNSbCxP8AXPgCGKeEm4s29hp7Y/TuWwcJSj6I+4LBebr0wdH1GMsALk3tbzwiCoE0irDzBYgfACR8t8PapLFDGW7XFsNRKsLq6wBdJve37DjS8CEynupxKlVXSxtc3vf4bC1un+L4BmqwKkbrsN7G+rHElEy2jKqR5j+OB4og+kI9mvYsAb9cI3Wqbqqe3GdpeHKNo6RHQzEO7LvGbbWt9cUXKouE4oKmeWrzGjroqnXR/hNYKCbBiOjDbF49uKyQcO0cbTyR3kYCNSbSDYm9vLbrihGrzebK6qZ/cuUKhpJDdleRmNyQL2I9MdLDwaa4uMJNdX2PijPKifJcoyPLY8zratAAsjbyMbWCkkXJv38vLfEnxNnUeRLT0XF9A2WVlVEZPdv60cvWyH4em6sLHy8iMQfDfFNI/HBoKqhpIirtDTVO4JAPhG/cjp9MVb2wZS1FxMa4M7R1y8wauzD4gPTcffAQzVBRlzwzUCtZ/ky8S01R7UUyOkkeakWmlemkKlTsLkEfv9PXH1VKI+U3N0hLHVq6W73x8X/ySsvrab2v5fUzUVVHC9FPpkeIhTdARY2sdsfZde6RUM8kicxEjZmU9wBuMfDPKSyM4YBzYPvct+DqF9KSsX9oeWezzJ8rfiXhTNqXLs3oZVmp4cpzNEV3YiM/heJPhdr2W5BONwta+PnT2a53wZU5rTDimH2d+75jlq18Ip6OGJqCUsv9Gcm+prOD2N1bH0V2xxe01OphhSw9UucW6u868gxYHoCD7eiOgQZIWHe0unY8b5rKCLlotIsP80nXFNzCqZC6WTUASNIuOmxHni9e0tI5OLczV1aw5Z26N+EuKVmFPEkJkbcgHSo6dO/rvj7P2cJ+SsP/ACN+5HYJiCZXV2YltC3cFvhv5YoHtH4nzbK85pKDJ6hoX5QLKsauWZmIUC4Jvt09cT+cZ5Dk2Xy108E0kXNUaEWxZiTYk+lsZtmmZvxNxnDVwxVEiA3SKNQJFVAWNuxIsTfHp8OCXeC5uPqBjPEI6Ti/ixPdxW++0iJtrip9Bbpe4I0tt22xc+B+LYc2qpcrrJiahRqjkqIBE8o7jStxcfPcfLEJDVCJVNUleS229Pcp/si33wbw48NZntPDBmCztrtIjRctwPQ+Y37Y7dfKKTWFwevOYPPsQ6s1jmWJ/PJX/OUj97WJY2MscKBXts6knYfK2G6uOR6YJe7k269sFTyrPXM0EpliCauo03Jtse+PQkMJgxBWNrffHnBdeydE2Vdr0Sk1uiq3KBZ/M2Qm2DqaB44oVbQWCgM25B87emI/OQfdq07cuNGY+HZtjt9TtiSzIyR1ukR/ghyoA/VHQbYW4ABEDJSXRdY1RwOCb6iPFcHY+WOKGmVhDGjKttWo7A9TthMjioURqbydbD9wwqrpatKSMKwifT4XVh4vRsC3e6J4tZIKwJ+GsTAmMGRt7X+XlgOokGphC+9tyLgEenfEpG/NoWnqJIlmSygabsPt19cC5bTxyzMswJg13Co29/lbDhRNkgVBdByyOnKBtpVugO97d++DYZainBqY2ZSRYlAGt87bjB0saqYJHQj4r3A1Hb6/s+2BBOj1CuzkEsASj+EgbXNxufng+GGmZQzNlX+bL/ZP+z/dj2JXlR/9aoP9s/xx7Ff8SdA6K2yu4YAnUO1gfzxyoZe9iLdjgiusW1B4lCna9j+3AzaSjFHF+12G+Ik6VHcuSGYSU4OmW+tQo3b+1ufvh+NyTZ7gr5+WEc8oQC6qR69cJRyxMj3bfBkSEsGEqePVaTof1rdsDuQVIkDNuB089sEnZCNWwNyLb46dMaKLWFrjVuTgZiyONQQaGN7hZDpBsRYGxsDvv64WeSAFvv2IFhgmNeYVLMAR5jCnp4yWmezMT0tiakBYRsgtTxG2trdwOhP3x1ihjXTIVYHYeWHZaaJiCWF/la354SQsNwPFgzEWCoB3Nelb8MOQGANjfr9MDs5WRja6G1gV6YNhkpHjEbAA9SNsNTQ0zhVYE+IHr0t03xYdAhXokyg3mQzrCWtLIC4UXGoC1zcdumBprwzSTKh5bf1q6WYgdiAB527Ylbt8OkWv0B/uw3zHicOEO3QAXvgtaAtQ0UsZDBo9w5VtUZW5G3cbjCmMMilXgAB23HUY6/Mlcs6SXJuNzhMZ1GzQlvmcUXzZCQkSU0HKAWmj0dhYYZaCkUErTRKb/wCbGx+eCLKGutJHe5IN7m9rX9NsLYa0syKtj11bHA25obzYoN6ejkCFoVJG1yL2GEjKspDOJaSBg92N0vvba1x6YMELhwxRB5XJ3wqQgJqdYgq9bsbDETGzO6Eip6CNNENBSKi7hRCLX+2GzS5XLK/My2KVm35c2t49XnovpOCXWOw0SQMPQk4RJyoyBqj8xa98VKskiy7KlIITCaSiaJRfky0MbJuR0sAR08z0wRSR0qtaTh/LX83ghTT8tJ8X7cNRkLdQykkk/CcKnmV1udQI8ltbFEA7K+IVEZ3SUMGbw5xl9HTrDtFVUy0ygEEgBgCBYgkd7W+RxMtw/kxK++ZPT00kyLJG8YCuynobqfn3w7Q5jPBKWifSwS2pgPDt2wiFnqAsjM7WRUVxsAFAA2tbt2xewVhNycPURCikNSrgbEVco/PVjoySAqY5mqyANI1TOx6+dzh9alopgt9KnzNhbB8bCTo5B8ieuAJjdMAHIqLgymOnp5Ep6iSnVu8gVrHzFxhE1O5mf/2rWsjE2VdChRv08OJOdm0spgDIRbxC4+2GoYdQ0mHSrjYgWt9sTUjCBSiiUb1eYO17gtUE39PTCUg0sZRNWISFsTOTtv5/niSjjhjFi0l1NgB3PqThpEhSPkA6Re3hK2X7DFAwokLTx1kci1DVTq66SvPazDe4P3xXeG5qiizI5J+mJHQzOIC0qaQosdB1Drv0/iL2eOnCOGSpK+e18Reb5elBmCcRQutSadRz4JI1IkVSDrtYnUukH1Ate22CBlKiLqXamn1yH35yA5tdF+gG2B6ilqOWrHOJ41Xawjj3/wDLgikr4MwooaumqEeCUB1dVAFj8/34RUFdWkJzRuQMVMFFEiUMdRJVJ5lUjqFS/wCzHY6fWvKepdkXfxBb/kP3YICxIkbLGFvuwwwzGSXSsdj5+mCDkD2JUWXx8yN6nMK2VUOrQrKiH0IVQSPrgyTllRJG0tyNruQB6WwC6xJIC9uYdgovc/LC4xWSyuSY4YLlRq8TEW62Gw+/0GJqlKjTdPrFNIb+N3J8OkbXw4EnWYLI6K2xYWJI+gGOPHJHToY5BYCzMBucR8TSjOp6WyrGKZJVkvuxLMCLemn88TRN0YqKSmncs93OsDwlUA+4N8R9VVTiNwZZnaxskbBSx8trdcLlaQMSLE9LhsR0jB+5V73vfpgxTnmhdUUXw/QtLnRWlarNHPKiT07VpvBNIWADBCSASvf+AMu2TZZTzuv6OgBDFSV3uR33O+IaqSpyKtfPct52uR1NRHEwFx/aHkRYHEwlaKqMSrKX1C4fr9cEWBLa7eV2WGhRw60cAkA2fQARhqefnLolOokeE32FsKfWTqMrNYeW2AnSRpC1wFH9nAGlN0QcRYIuGGnlUI9PDJbqDbBM9HJHDdUFgwNtZBHTtiKViCzBXvewIGHp5jFAjl2cGVENib+JgP34ABM1J2SnZWctpQ36MbW9ccF0AYMp67XwqpF4lBvv8LN164HjGlwCL+W+IVUg8l6WqYghUPqQTthhJZAdyxB6A9Dh+eNmOtbhQbHywNVMieBrlvIYEXQkkbLlSWsRGzhiNrEH5Dr0wrLJo5KeB6qnYVCDxL0sbbjr/jbDBBZVCXva5BGGZtdO3M1M6mwkUL+YPXBK1LzVCFrKrG+979MNKQrEqzWK76j3wGutm0Krbjqt74U0BjFpuYrEXBYEH88UQrBT7VAW2lA5J+IN3wmVuW6llNmsbdbfbHkpgq6ux364f1Frkjp8JFtz88CTCsCV4S3K+AW7Ag9Mc51hfT9G6YbZgGJYkdjbzwkhitlu1x574CJR6osvTTE2Frb72w4Z10jwiwFjvgcINI1fD98LSyoQSBcb4kKtdkMik1DSlpNCre+rHCscskcu/MS5Qh7deoPocejmJaqiZheNU3B874SOvQb9xh7TCSe8jqeaJqcLoAIPiHe/lhbuhPhaxXzxFsxUPNFdyQAV7kX3PzxIQcmxZ2IYrfSXtv8AQHBF9ijbaE4TCwJtclenrgdrI2pluAPnthayF9tFl8sK1AWUMbnCQ6U1zZ2Q45XMLFVG1h4cPw8vmWC3Ntr4YWRWYqw2X1x4uAV0hlBO5B64a26TEFdqHPPVW8RA3YDHjLHzAnJjK9btf+OGHdS7NCQyg/I4G6k3JGoX2bcYYBAQu3RUwhaVy5LtsVN9ht0wmMROw/DUuB26YakV2K628QAv5nD0CkSIqOiFgSb74EvHJDpEyUJnNIJkWpjjVKiH+qbzJ2sR5H/jtheW1Iq6S+iWJ08MqMBdT5YJmk0wvKXkIRSWJAA269+nXBeW1dBR5fmCPQq81bGhLlVIUqdQI/8AL9u2BLgVCC24QdOwMgQEWHYi31wsmKV15kanSdyNrnz9cIR1UmQAlm9OuGw5uNNrk32GI1oOyYCeqKMIit7uYio21EEHfDc9PUOixrLCovc9TY/fDM0sxQ6ZGjJ6HDZim8DCqe21zbe+I03QlOHLnZvHIwv/AGFXf7k4U1G8a6ec5I87fsthCiQAl6mYkdza2PCoqdzFOWt3Fjg1QTcmWVNTNEqOQA4Zbp3H1w+5r4lZNQeoXZYdYJUk7H5W369MK588cKTSszSbjTe354cOcVrNEgMShFCEhbsV+h7XNsCHAqwhoaOqGpWshuSxAv8A34eFNOum0yPvttvhLvNJI0hkAZuyJa+HI3mVCzuwVe5XFkwETRJSkpZSoleeFCNtLMSThD0dTr8M6gAb6jcHAtTLPdFBuCLkk74SZZQSQ7kH64jTKFwgqSipV0qs1UpcnfRGbD673wU1AgAdq5QnY6CT9tsQoqqhQCkkp8wtv4Y8KutddDzTAdSLruftggUF0f7oZC4iqXsTZTydrffHaaBJGZTzXeKVkkZFvewG9r+vrgGJ6uawEku3lbCZYZBOrxtJzB8aM3xDFSmDZTMUUccpjlSqIPSTlqqi/S41XP0wg09TIIzERFokDXkVrdbdB3tcYCpRC8Ykj1shH6rbffHUYq24ksP/AIh3wL1YujWpF1a2zSlS+9jTtt9dX7sPmGKGJZjmFOxH+bRr/YXOIyZmADFmIJ6BrHHI5ZgjlRq07AbYXE2RtN1IHMNcgjDBF6kkf1h7bdhhuqaKSIwlgVYEENe2/bbtiKN3LF3Ia/QYUFfcrKflgxSHNVxbomkeSnGp1RmU6A8bE3F9rjriVgrKWaRqcvG06qHMZ7DsfyP2xApzlHh1de22AczpK6eWCeAWlhbWG3BcD9UnywOmVBV8FcZmjLAr4Ta1rDDevSCAAO9/LEJl2Z+9wtID8BKyIx3Vh1GJWlkWRb/rDtfEjSd1CA5RRj/RmcrLEwGX1kgV7kBaeVmAv6Kxt6A4sciVVDVCOaORWVrMhO6i17/LAySR6gk6kxFlZ0ABDlSCNQOxAIGHK2saWolkL62e4OwG3yH0wTy12ylNjgbo45jUJRqkIDAlvE3fEWMroqqMvVUyNLcFmUm5v16EbemFQSbqCrdQAOwH1wTHOskTOSoYfq3AI8r4zEQZT3XQ2VZJHRNJDTy1MEdwyiGUlQx6qA1xbv364kYqGpaLkwZjHr1eIy0uq48rqVH/AAw7JUM8aqFdWA8tvvgWqqJo4rR3VXPxJuen5YuxuqAheneqp5SshoKgowUe7yOCt/MBTb13OFrmkgJ10VWljpuLPe3fa5t8wMCRhlhASkfSOyIST9B1OFKJ2UukMwAPRkIt98GLqpRMua0bm0lQsDMdhKojJ9AGG+CYnRowBIpI9LfmMANHKVBaFyW2vpvhl8sglNxTKjE/ElkO3qCDiFkqtZUuJXH62x/WFzbCEldV8Q16jsDiLkyiqVvwaqeMatdxUA9unjvjtFT5zFP46iFkLHd5UVvTobH7DAOpwqDzMELReBq+no+FOM66soEr6emypp5qV2AWdFSUshNjYEAjoeuBMhrvZvHxRkdZwpxzmOWViSJTxZc6TVNMUlZNcKCVbx6tKi6kAWBtgv2XZZmWZ5NxdllTHl8vvuXCCKBpnAJZZVIkZQCFNwLrv1wdw3xjx1WcXU/DlDlPDGZU1LKiZjV0EkzxUaAjUpkcANJa9lGo362x8wzvV5/idBMiCe8GjToAuHBwPMbDpKZU+bP3K/8AtJy6kzLhWWOrziPJRTzQ1UVfJpKU8sUiujEMQpGpQLHrfGZcBfoKp404eyzKM7zLiAZa9XmFZVRUGmCSrqBLeoklJHhILoqrcXtv4TjTPab/ADc/mVXHiumkqsq/D5kMYYySPrXlqgUgly+kCx62xnXA70eXe156HL874jj1xCjmps3p1mjn5cJmWGKe+pXjE2oq1yRfc2x5fIX1DltYAusHEWGn0QDeC6YO1gLFSrHEH1Kl8TZbmjcRZs1PmUkaHMaklJabUP61tgwN7fQ4C5ud00YCw0tQw2ASYBm9BrAt88WXOJJjn+bEhv8ALqkA9dhKwwGbEWkJY38rY+1ZZUPmlKfot+4ICxs2UG+YV4kBqMnrIxfxFAJAPql8Io6rJoQYUqp3qWZi4ll2UlixFu1r287YmQyx9Y1Jv1QWP1sRjkkjMgjaLUo20SKHAHpcY3cTcK+DJDp2QSC4ZojGb9fFe+F0l45zzow/W9j/AAwufK6CZrCjSmcdDA7ID9F2/LDa5XIsycrNKpFFyUZUkHyBIBFsLJBTND+SivaXla8QZL7vDGDPD44GuLBj1BHb54zqk4HzWDLWnaKt98WoXRFGY3iKdyx1XvftY7Y2Ganr1tyKqmkGoXWdWQn6i4wzM+Z07sWyapmjDBdcDLID9Bv+WNNOqGNgLJWoio/UTCiW4YoJ8799ly+GcR0oiZjEhVyDcGx/W7X+XlhriPIos/zzhylraVPdYMxho51AYNypnCEi3cXH2Hliw0+c0zuI5hNTsSbLKmkm3W1+v0xIxy0ySR1MJiZ4zdGW119bjocW7FPiLKhg6bnE3utS9nLZRQRUnD0de09Tl1OKWO6EF2jRUcnsPgOwPpi7zmMQsZbFADqv0t3vjIfZqwk42glbUz8qTxl2N7jy6Y1nMtJy+o16tJibVp62t29cfD/KTUNbOKZP0Gj3la8JRFKlAWWUXG/A1ckK5RwFX1tXOBLTU6ZKitLAdlqATYCInbUTf0xrfbHzxllBlWWZTS55R8ZcaxLQ8N0VTEirTaloZntFCPD1BTf9px9D9vljgdpMJRw5ZwZiSO8XTNuoA2I2+CKg4umVjXtD34vzQsWIUx2BFgPwk798UupYThl0DQBvc9T9sW72oVUcHFmZKZBc8vYnvy1xUogWp9emyk77bD1x9q7Nn/4vD/yN+5GWzdZJ7VM5lWeTIDBoiRkmWUMbvt0t5b2+mB+H8sip6KjrjltXDV8nUJoXG6uCQ3xd1a1rY0biLhSi4hqIKqSMGemB03A0yCxsrel7G9jjw4FrFCxUTxgWFuVUaEUW3Fjtcb9Bj3OWVcOO9UMeC8vnOGxlS1FpPUhUYySwAqmc1Ue4Le8Qbfmo2xonDuWwU0YzA/0jRCLygeF3awGknpe/54ZouD61puS1USAhuSVYL/pHYEgfPEpWVNLMUy6gC+40vWQjeckdQewF7fI9tjhuaY9ukNpOmUvIspc1xrV2RG35hJopdaiqYJaQKQQunw2+Z74RQSLPT1gB8UhLDt02x3MXaOhiUnTI2yi+w73OA8sE8cxqJHBphIERRuQOjE/964+mOAHAQIXqKgXK+KWBKURw62qaqBQoF7gyDUSflfHswqUNQ0MijqSSO48sS+aBo/cpGXSnvUYNgDa7WH0uRiOipEFTM6hrubWItbzti37KmbpmljaOOSSYBSV0gnobd8PVE9S4CyOJFZe2/wBsNVEZNMSNIeLwsP1iL9sDxzFGMbzWERuSO4PTbywmE8FLqBNHHpSR01P8KHoPXBBNTGuuMBPU2vf54TA6VMIKxnwtY9iPLHq2nIRwGIRFvuev9+D1kABL0C5XKgtLGqyRl5E2a5/ZvgLkuIiZdmAsLWBOCnLRQEzoBcCxa4a/a+GKDkzTMsjA6Y2UW69Ce/e/niwHEoZaLoXSn+d/MY9hX6Jf/qVX/sj+OPYPSUXEardUNC7aWC7dAV2w0Vg1WKKPS22FyRgOxkUWIv4T9+2GgAPFfe+298PLD0WJr/FIZYgurR4V2G1hhmN4gu6kC/UdTghEiQa1ABLbkknAlWulg0XTVYhj2+eDYJN1RNrJzUoB+JRfvgoSLIttidN1JNh9MBEkHmRAMwG29scqJURUVnFiQF6C5JsB98CWXsEbX2RmiRQ3LhZ3tsC1gD87Wthp5JVjQn3fWdrFyQPqAcDK7WuxPWxUN0+eHdUZWNGDEKd7bXwJpq+LKVDLM4LN7npUkDS7m5H/AHBhMssT0+mKSJjqsbLffuL45E8YUpNOGdLamG3ruOgwqBFQlw9w1z4VAt07Dr3N8EG9ULnyN0EhKORIyDyKKT+/BUQkbxRsGAFxZNyfvthiRRpLBCST4ixsB64TBMygqrOpO11PTEI6IJRUVQ8kV2EsTk9G03H2vhKS60JWViQSoa/r8scpZruS4A3I2Pb1x5mD33L7bntfBNbO6ouTDrIkob3qbSNyGINz67dMInVZhpYS6RuSkrKfuCMGLEpGyqTbr2wPJGBGqyuCzNs19Fzvbvv8sHpCWSUMG0RALcgdLkt+Z3w0VhZxrQbkb9Nz2wY0YAANix+uOKFRGRowStmHgBHU2O/fbAvb0VNckyBigiYKY1AsNIP+Bho3UFY9AQdlWwwtri4IdgrFgSb39LDHitn3jZlYbMu1j/pfliECLog4pKwk7qwuAbgIN8cMbaV1LYkCwJ3/AG7HDygIoAu1mNyfy+Yx4sYvxBAqFSASStiD+tue1++FmJsrldjicIWsW0jYX/jj0oIiJi0M4IvzGI2v5gHDzSLy2Vo7qxAINuvngaS1yYy5UjYM23z+eKiUWoJxm0sTy10jp5keex/bh6WuhEAVCqKBtcXA374DkUiIl20eG5JBta3p1xxG3XS5U9xp67bEYrZECg+GeCqhs3n4gzHPGK3LRw39dhe9relsS0ho0lKpFyWLX1ROUv8ATp+WAmDN0AKtbfzPyw4iWLKzEbAi67Hfpf6YuJUa6Nk+rzCUiCtqVHbmRqw+m64XFU12opKKaYxm2oExsdt20m4AvcdT064GVoQ6q8SWA2IUE7m/U4XI8rraKMso6jptbzOBACbri6MSom70sqkgm5dHH3Ukfvx56mFT+KJI2/04yv7RvgFuVuBFrO43N1IOxBHfAuY++NQzRUb8ud18Fn2Xfrb88XoCnFU4rxudSMFUnZj3+/TDIlhCgNIvXY6r4Y4U4fzegytKrOsxgq7rcIXLW222I/PDs1RewEcKoxJKr19LW6dsLi9keqN7KCgeTJM6NDFGWyus8cZJ2hkN7r6A2v8A8CcWKY6FDMhsw09RY4HVaOpUQ18HOgZvEL7jr0Pnvg7Oq2GqqnamEkFOAojjBICgKBa3bDA0FK1EWCY56iMAU8sj2NtCkjC1UcpJJAI9W4VV3Hzv0wPE0l9Jay723ucLaTZm6uvwi3T6dftinMg2VcSRdLMgsQiqPI6TcnC4JNUh16jpHcYBVtZvYi3djaw/dglX93hDddQ2G2LISQSUTMTpKC4B3t9cRrHVnxBkUsKZU0ldJJ1MQV+Q1A/TzwZDKhIZ11G3S+A85qbzUiwSHWs5SUIBcjQ1gfTdfrbFhFyXKu6kjof9be1sClCFFmsbd8OsbuASF8PQ74blum7G21r9sMaQhIJSKjXywz6WA2sLHAvD9DmQzuTKYqY1EUpMlMQLaVsLqT5C/Xt9RgosugIBcdcSeS53XZYZVpCqcxbNqXVe3TFlwVc5KEqKZRsl2c7aRvv9MCNG66kcaGZfhYG/2th0ELCbXUDbb+/DY/GG1wB1N73wouTSExoUXBJFvXC1A931fEqsptfqQQR+eEScpAfiG+/zw1WSj3GZFYodI0gedx0xHQhEqSqbPCgJ3jYqfvhipjRZAzN+eFVBupZvFZ2APrqOBHnLldSlmA2OARBOVEqhPCpIA332OBGMb3JuO9r9cKqXJXQyuSdzc4GBAbSzOL9utsULIXbp6E6tgoHa9sORnkpJr2J76bnCITEEIZtCj9Ym1zjkgCnUtiOttrnFSiCIpVipkexJQ20i9io9SMemnM1Ty1SMR2GwUX+/XAx1EKChGrqLYUmq5dr3AAsPLFhXZP8Av08ZaBVCoRZhY3/bhuJmO4WwXrfc4Hd7tcayb73PXCxLoGu1vIYByJpuiBYKCCet7Ww3MwJBS4v1GOGYauh3sSLYad1YeHax3wMFE8jkls5tYC+/l0wouBue223XA+oF7jp3x6qaxW5sLXtfri4KXISYiOTWWBBMg3Pywyjqqkm4A9MKDhqK6FgHY3wFK7KdBvuL3vhgQnwRcEwEpJA288Eq+uQtpAAG2AIGYr+GLWG/rgiBnBDN07b4pyJpKJDsl3Gwt0Jwh3GxVbbeWEs7G9xe/Xvhtix8Oi1gevbCRPJOlPI6677i/U4SZLsPEOu2Gwn4lwWucdU6pAVFr7D0wxhMqiAkhtIfWCqgkYTEYy4kkIYA9r4YeVrlB0JPfD1Mh02/V6nDnGGpLLuuiEC6GfTsW2JOOUoLSuy2tpsdumOySRhAotftvbHqRhGHDggm9jbb74QfBMIuvZiGiowqMAXkSMG17Fmt08t8JExeV3KAKx2sNh9MdrtLtAOYxYuWsp2NgeuEEGEKxZQb9O5wY2S6m6aWXVI7ggE7H0wqAfiMQTthMjvJNYndTuAbWwqlADt4tx2v+3BSdNlQKcZ01XuQV63wmSoRiANh2wzMt31EbnrvjhjUgmxOwttvihKpxS0kdyfELdxbrgumjeVfBHYKPS1/44CVNEZDhvMb4ejk/CCx3Dd7dj88GwwYchS8yKJZVkDso6jpgSnke6uw07+eEy7RlVtoI3Fv8DC4EOgDUSR6DBQrkqRo4nnLMSIwLkdgdu2GqOr94iU2JV1DLdSCLj+GBwGCtpkKtY3sBcY7ANCAuSqqLEk74F5tATWG6ZqCyu6gXVeljjkTlpLDUT1It4R8zgZ5dLswKgHqcN+9uiWjAOx3B3viNsEsm6lMwiamCkP8QuABgXnSF9RP0wvMsznqoaaAUkgRQSajSR+f06YCR2VAOc5BNt7XGBYTzUeRNlJQz6FBBIx1pWMgI3v62wDzNLC7sB+Zx5XW7aRYr3KgnDQUsu8VJxMVdiJNmFrE/renlhydxEdDnx+Wk4Dp6qZLFpGK7XCtbD+bTU9Q8Rp5BIUj/EA6qb4U/UXJrCNK7JI0t1HMI/PHqmRIV92jDaQbset8MRTMsBddiNt8ChmLHU1yTvhsAIdR3RqzQ6bEFPmMeV1LWJsB3vgYvZLadx5YTG576r+RwSomUfEYNQswIvbY4ILCMWjkGntviPQqyM5K6h12x5SLEKV+2KgJggIevSahrBW0wXkzG1QoPzOv59MTdOj8sSKPA0YZSe4NrHA1EYhIvPSOVL7o3Rh5EYk6ypSdYhDGY0hjCqtwb9ydgMC+IKgBmQUxdtizXa/S++Fr4m0rqJ+eB76p9SOSG+IX7/uwYmhEvq8Z8sZhZaGEmZRcUSwopLB3IuQf2YYkMVNJNWqty5XmDrYDYHp674REFVidYba4x3W0qOmo2PY7g/TFESilFisV0CFVBbY7bWwXJVPDHeMBmk8IuOnr88Q8CzK4jILxjcNb4R5d7289sESTSJGWI1A7A6cWBCkopZ6hiQ8m5F7kYfWVdFyh8JtbpfESk7pfaxI6kEjDqVTOviB2/s98WCZQFSD1WwCRXvt4ugwQjRsANIsRuAbYio3Y2Ntuwt0warSqgLAkt0uOgwYlCDBRT8tTZY2svQG+OQPCF1syrY7llvbAwJa7NKwU9b46rEkLG637+G+BcCiJEyr97N5csly/i2TNnmhy/wDRo97kXUGWLTJrI073Av03xB8KUmQzcQ5BQ8CZLxGlBNKaqCprc6lgpmiikQyssIYlvjGzBbk798WT2JzUdJJxDU1lQqU8VNHJUPNYIqDmEknyAvfAvFdR7MYc4ybN+Ea7J488kzijgvlVeImdJJ0WTUkZs4t1BBuOuPl2aVnDOa9HS86tMROmdFtQBGx5zty5oqoBaHSFqfG2Q/zj4eky6OseiqFliqKapRQxhmikWSNrHYgMouO4uMVnI+D+Jp+KqHPOKs0yiVcueSeCDLaMxCad4+VzZWYkkhCQAPPrtjQscYbY+W4XNcTh6Ro0yIM8gSJEGDuJC0mm1xkr5pz2Kl/nLmsnuqCUZjU+f+eff5nrgMwVRmcU9VIhZTbUdQWxHbvtfvgvPtK8R5uEYj/2jVX+fOfDUMqtdmupGP0llw/2Sl/KPuCyA3Mpnk5+alg0WXNSl7hmlYELba407m/rhMNRXLG0s+XsFBIPKe42Nr2bSd+23TErCyhS2m6nz3w3WRB4pIneTQSNtugIONkXRhxQJzGi1j3iQx22JliKi/8ArEAG3oTg6lFLUASU8yuB0aNwRgbk6kJExBItc9PlhtsrpzUIxpIi3XWo0sT8x29MFoBV8QhSo1ogB7m1+5OOsIjqGkKxG9tsRXImhQrT1NYBe/iIkA3JPUX7267YVzK9ShZ4ZV3vdTGSSeg+L7/lgNJ5Kg4ndRFJxnR1GcfoKtWSGoCnVHVBDGGHbVffsR5jE/Ll9DIFflz0pDXtTSWDbdwcNVNLRS5lR5pV5CK56RybaUkEiFWBRg/UeK/TqMA5HU0dFUSqsQy6nlsyU8hdY1cbOUMhJ0/DYXIGCc5p5KMa5u91oHsrpxFxnCRXyyLyntFIm+6/2rb42Osj5lLLGACWQgAmw6YyP2YGOXiqmmQqQyPYra3wnyxrtTp93kDKzDSbgdSLdsfEPKGYzZn8o+8rWGw0L5oky/g/h2rgyjjfI84oZZFipVNBxG1ZE6KfAhi1iQIDew0EC+Ppzt54+feEuGOJqHNY39n2Tz8LUjuJJUz6eCRmU7kCNUaYH/WcY+ge3ljB2trCo6l39RvPeJjb5su0/ad9Sy4YRNlgXtUgDcf5tLI1k/BsLX/6NAcRKMktOUiYNp2suLP7SsqqKjjbMZo62ogEnKsAiOthGoPke3nitxUGbQ1BjZaeWMfC7FlPpsAf24+x9m2g5VhjPzG/ctFxyTtHEIhdr37C2D4II1ieoqphS08e7yyyeEfTucBsdEsMNQwcMTr5Wq6gD5C+9u+EZlJLWMiCERJHuiki/W4JJufLpbpjvF3MJIbySM5zOKpjjy2hpZoYGe81TIQrTAC4svUIdtja/wAsR9MirLojUljud/XvgwZcSPDICepHXfBVPlppyZGZnBFrW6fngWtLiiJAaqtn9Uz1AVWDOp0AAWAJ2H0wZS00dFEY4ZJCvLS2pSQxDNfrt1JP1xKS5XIal6mCdoXZbeEDAlXR5yKZ1d1nFtjcqftbDOGRySDVB3RNdEazJRZ7yrZxuDYqQf2jCHiHMEguQwub9Dth7JQ7UrRyB0YGxB2IxyVSgendLGM6oyovdT+/EcJCpr77qMnZKepKyqTHMth6EYBqEAqysQQlRYAnZ18j6jEvIsNQvLcX19DaxH5YhpllgzB4Ht4RqQ3/AFT9cIKaSTcI3LpEtJIiFUffSL7eeF1cHNWwnEYc9LXJ9N8coAW5irIC172H9+CpNDIqOVB7dLnF2VDUVCV0s4klhWoNnOkvsPLsMNUhhNcBMrNE7BpbEA+pv0BwZndK1hIzKFGwG4J/j2wqjjEaGoRHBYAWdetyB27C/wCWGskpToG6jtdR/nv/AO5//Rx7Afv039mX/H0x7F6gi0hXmtSXmI11SIXMgYb9OgPzwM6RJJJUA2UrdmLkrYd99h164mqyJeUwsoFjsehvtY4j0XSjI0SWAsVFrDGkPtss/DvKYj0vGjxAOjjUrXBB7g/I4GMpDhTpQs1gDtdvL8sPo6KRHpbZQRZSFA6Wva30whowWClAQDcFt7H08sMayLoJTVOet0KAMQVYWNgSL/W32xwGQKVCoBq20m9x2OFTLrEkd26bEGx+mPIvhQBA9tgSASO1ziclYXA2tWVkXXa9m/fhnS40m4uWtYnbD0qxI0bqiAnUQqm3lc2wmQhXvpAubX7nAtOpWEpRI6HcWU7G+EyMygBhqNt2U9MdkdmQJcgt4hsT0w3IyndEQNp2DbH5X64twkKBLTQyXK732AHX54HVuYWVb3U2Ia+ofO++HVY20c25BNrjbCgGaQKWBv0PTFAQFEI6utn0sLHoN8KjkYudLH1GnbDwjs6k6iQCCNtLb9fn9cdnViVcuAo6gAbj92CalOELmtOTzLEr8RVb+H6Y68oMQYadIGon0t1x6Mf5o7X07mxwofhIZNLq/Y3BBxRtdELhMoVC3KspDG4Ppff5d8JkFwjL4hffxdvlh0DUj3RQXBJG29++G3Vbhl1WIAJvt9jgOIgLYKbAt4iSFG/Q3x0gEsxVgNRNlNh0t/f88JkFz4Ha5F7AdMIi8To3McBlJKMlj269xinOkQjAS2lEaquvTqNlGvr3sPPHWEDEqYg6ybMCNjtbp0w/VIvLUFECrcjoSPPzOBREFlILizAFbDcjAQrSnKrFZ/AqjudgB32OOhwq7RiS43uMJbSp0A6nIJ0jqQPK+HEBAV9tBUtctuPmO2GAQqAlMBeUAqBVRbAKB0GEo1n0rHbte5wsCSae5YFrDa2332w24DEMY4i0TEoXFrGxH02PXELOakxZKjW0gfneBhYJsLH59/thMzghw0YkDDSwPceWFmbWul1QEnqw6fLDEurmeEFhbtiw2FYT1OTZbEAWt0O2HJGAX4j/AKrHp898Cxo24uFNj13w4OY+mxV1BYHUdx6ADbrtitCIHklAkghTue2rrjyxDVeTpex7WOOPGQWV1khZdw2gbG9u+2HuY6NeELvta/bvgSY7qic95KqITNJMli4Dt8Nz0+WEyR+I7Mn9nfYg4QUqHgkaMkOAx3Phc22v3xy8gQeFjqIBNtl233P+OmFgQikndcBJBNgANxY/nvhwPqPbfbphhHGkaizGw3XYE49JYtqYBB87WwSmycZyXOpiW26knDgk8V7kgt07DDGlWbUE32IKi98dcfiBSSXt8LeXyxd0s2T7sGDh1R43Fip7g45PPqclugFththpgkKDfYdj2+uESMDtcjbt3xFDZFwSAFQSCLXwFRJz8/zW7KeVKuldO6kxR77+g/bj2tkYHVcd9sEZfTpDmFTXmZi1ayEoNraVC3/LEUCaljKS6A4VyCFJW9tvnvjlQRuTvpG/kcF18BkZtjYA2OA5gwjWNNWtjbfpiFEmwwWMhLEt1xF53U5jTMv6PSKSVkLAPq7ddxiYkZooEEiKCOw645TVMlPKJUjPpfcfbAFX60jJqbNUySOfNI+TJJ+owsST2Ax1QRJZVK97d8EV9XW1roZnLBeg02tgXW2pfDc332xSIuSZ0bQ11A7+eBHC8h9Y363AOD2QlTcML9dWA51Bp5L6NZQ20nfFAygCIexpmIvu7de1zfDCQkp4b3OFRuGj0E6Gudr3v0woGRRtpv2ucQmExrZQyIZptJJ2NvO+GqiOQXBUWHph9lmuSjflhp0qLbkC5wLXShc1N06OHKOCPDsbYdWG17O7yb/EwKj998NJBMSRIW0qeu+FCCRXvqup8jiOEhRoM3TrhgN7X7kNhtCzE738h1wp0soAaxvhESOijxHfc7jrgQYRQkSABLArfe+EEoqlXHTceuHLa2JsR9gcDuoUXa57XJxPSVXCQ7nXfWPCbi5vhwsr9CTtfpbDL6RZuWfUg4Up/FNrabdLdMMaLICSuiTTYE3+WFTMWXYm9t/lgeT8OS9wL+mPSBqjUQJCiLvp264qVJTkAb3WIOtmudr+uEVMd2uCoIx1GCouhgwtbbc45E5aS1xtvucQ9UUpuEDVctZm63OHoiTKACCoW+xx3SLEjRv3BtgeWIoS2ok22vtgI1XVklsFHPIqGxb6YZ1fr6dPa2A3kZVNgzGw3x3XIw8StfzxYEK+ITyRvMCqdj8x1w3GwsGvYgX3OBXklY6Fj6WG5tjliXDM24NrKtyfrhjGQZVl9l1wskgI7dsEK7BbXN7ffAvLkAVh1B3uBvhyM2ezKBte4OJUvZLba6bllLP3DA736jEnTKfdgRIRdtrm/lfEbcNLtqJ74LdwiRqsQIF2cgg3O/rt2xTmK2uIKXOGauDAoEEWkr6lgb/kccOmNWmYrsRba5+mB9RaeSqVOWrqqBAbtte3p3wqSplSUoKe/huN9/vivRCFxkoTNDUyU9RHBtK1rNqscE5FktdQZUlVWzEyTMbqZNW2GAG+Jmcb9O+DWqZWjVGkuB8N+2GBktQCCdlyxDWO5646oP6p+Yw0W1Dz8zjjm7gWIA64trdKKbp6ckEFyQnztj0DroLoiiw2N+vrhhpCCRqAuOnlhM0gjpiwspI/WGKcLygJMpZYyoCdhe422w5AOpXxEix9cDI4McasoO3UbYUCpcgMbddtsQmEwSjUWxOxK26dzhireMLr1DQLXXt+3HllYtywCEO532wLVTsU0DQABvbAzJVhMSnUS17AjYHocMxBdWrcN5WwlmuWa66LdzhKv1INh3FsEluBCSZLKUFipa+m2ww5Hcpqa4INt9sCCzSGykWI2wuaSwMZvv69MSECJkmU+EnxXw/BIOWNrDucBQiK6gjU67n0w8oY73F+wxatFh2J8BuAe5wZSVEUDXeCMnqdCDU3lc4jGkIS4Iv0vjkTv4iNJHmBigUQlGySF7/qmRi3LG2nfbr6YRdkYknc7YH13ewYtt9MKV/GrNp67XGCUMnklyyEi6E+RvjnM8YtbphMg1AsTdbdLY5T2JsTbffE2QyjKeUrKPCWuel7Xx4atZsTt64YmcsdUbEgfnhMbgKWvYjFymAFHrWUyVENI6vJNIb+EbAeuJDwlSVYNEO/r5Yh6SAVU5cCRXYaboQCQcS9WkVO8dGgBKruBvb5274U47lG1KgF9TBCB2sMPx7hCzWJ3IwxqaIaQBa1yfLAlVzKoiKOpekgsebOYwR/qgXBvhY7ycXaRJUrzllB0sAo32HpjrOIKdpS9kW2piN7kgfvxHUzLylig1COMBQGABa3fBsbCzkgFSOwvgiICprwTZHxwShrSK6mwKk9DhqeMCUqAGNtzfqfPHKWfl05jjGkAk2G392OLNM7DVNfyDAYAJhQObz1lFRSVVPECsK63Um11HUA+eCsvqYqyjiqIw+mQXU6h/ww4jvIjI9ihvcdQfTEXGf0NXmnexop2/AN78qRv1PQGxI8umHgArPcOnkrBE/L3CKSOtzfCjKztcuo9OlsCwBWjLkn6dThQcCTw6rdLXwJsUW6eV3UlFksx37Y5E0ztsb9ugwlEuGLqCx3Ft9vljojiC7yoGIuAVP54oqK+ezitfLsh43zCKmSaSlyYzLFKmpJCiTEAjuCRuMd9lvFfDFNxHLlWdZ3lmc85KR6Cr/QywOtS7OHiGhBsCIyCbW1dcSn8ntbZrnN2VlNPDbS1x8T3xsyop/VF8fGO1ua0sLmOIoPpzrDbggEd0fwn3R4rQKZcGuB2Shjjbg4706Y43TfHy8laV8yZ1BzOJ85JkP/ADnVbHa34z4YEEgFkOu5tYbE4N4hmQcR5wsaL4cyqQbje/Oa+EQz2hbRqRiLagRj9SZaf9jpfyt+4LC5qRH0Ja4bddnscDmVlVhdrdh1wpll5LBmEjFu2+3zw3eQKLIoF+o/fjfyUnknoiovqJsNxt3w5E4WQK0rE9djhVMjFNTKpN+tsL5cTSHVtpXrbFwoBKcp21P4gTY+EX6YLjAsQoW172axxHTvBRxNWStohjXU7aCbDzsBg6BGljDgEofhIN7+uITARBPx6WiIMaBCf1Ba+BMxpOdVUolkQRI7OwO4PhIAt9eu/f0Ifkvp3awUnpfHoTJIHUAlexJwO900CWwp32X0NLHxlQvGukxLKQF2BJTcnz+uNmqhK1NIICFlKEIT0DW2/PGTezON4+J6YuLXV+3+icazUFxBIYxdwpIt52x8R8of/wBsw/wj7ynAHQF82U9DkP8AN+ioKTh7P09qKSRGSqeGfmis1jmTNN8BhJ1HqRpPTH0uL23tjA/Zzm3EvD2bUldX8Pe0GuFZlts2jqKcyqa7Uh5kYLWRLGQbW207Y3ztjndrnv4jGEyJJnVNzAIH0WiO6OhWfDRBKyX2hi/FVZpiDWWPxK248A64qWaV1TTZfNJSU/vE0S6uXrIL26j5+V/TE37UqBJuNK2VNSSFI/Eh0sCUUX29ANzitKK6MFHkWaO24nFz9xY4+w9mrZVh/wCUfcnvk2lBcPcSZbxDLP7oza4H0SI7WYG/UDyxK1Ei0kTS01KZ2LBdKWJNzbucV+PJ8lFfUV9RRe6VFS0OmWJtDRBWJkKuv6xDWswIuN7jpJ8MxZg2TJFWMr1ERKk8wMSo2BJ63Nr/AFGO69wcbJVMPaIcuR1tdLraH3WDRIFa8nMYncWst7dQev2wbHJIRVxyyuzIQASAB0vth1qSUyI9/wARSCGB32/dgOjjk94qYagBnc3OoDphjHQQFT22UdQvmOYZpV09PmC0MNIFE82jmHU3wxot7XtuSb2FvPE1H7zAhL1TVkQt/WqA1/mP4YYp8uNLXVk8WnRVskjAdQ6rpP5BcIerlqXWKkV20TxmV2jZVVQ+/wAXU2BFhvvg3u0pIZqR9NPBUOSA0ZI+EjHKyHVpdVBYeXfEVmMgjq4o1jCuZfCFI2UA/lc4PkqStWsbggMpOoKT0tc/mMU14JgqOplolCy061AckNGuq1xsR8sBSUzM1mNpVWwPUMMTKKGBZbG++1vF/A4YdY0VhLGYkLWUkDZj5HAuZJRsdAUCNVNOsqPZt9aoLk7dsSalJIuZERY79L/T54CzSm1gyKxZFO7hj4Thyiq0SIoymJlUBBYnV6g4W5h5JjakbpmtmhkVonil/De3mOl+4wRF7zHFIzNCvLKrcC+pfK3TpthiplaQloI2D6SzORey9Ow9ccpq6aOwjWwGkMV7/wDHGvDw3cLDWl7rFB+40f8An6r/AMRsexPe9Sf21/2Mew7S1TvKYMiF/Gh9LYTNErRtZSxI/V7fLHJGKSm53Pphcch1HQegNwcZrpzSIUXy2DksjhUUG+1sNKloyNQaxFi3Ui/p6YsEIViXjkMcrCxI22wDU0pZwdAAJ+JDuPpgm1HHdQ0xyUSaZmEok5ckZ2UXuCvr5HDUBUxxulljIBXR0K2xIz0skT2W7C1wQNsA1EatJFMI2coSLhulwd7X3w0OskkEFeUAvfWrMo6soBtf0wjUHkHM03O4W29sKeQXU6CCLi1hvjrodjrux2UiwIxGgjdVKRoBjLB2Y3Jtf16YalKyKUVLkEb97Yd0MCAdOmxub7/TCHBT+rkRWLKbuva+/Q9SL4IqiUyt1YaStg1gcPuZSUU+LU3QKDcW6Enp/dhmolDSWJJUdPMYdikB6hQCNyN9sUpqSgAZAUGtWGxWxAwgpAkJMumJrAMzOTbvYm/S/rhUJ/FYqWNwBufD/j+7DkRbSVkiTcEEXv8AmcU4HkpYpptNrRMG8Nxb1w2HIBXxEHpbCw83JBeBYpFLAIxG4307jzFvlfCAq+FFjRRa225Hfr164IBLJ6LgKF9xYrsDfphMNNFFJI6szmVrkk+Q9OmHBInMMZI67gi1tu2Oky6fwjHqK2FySCN7bfO2Ae2dlbZ5ppUMUhDR/hhb3L7/ACI6/XDNMjww8qGRiUtbmkljc98PEeKyhmKnex/bhBLXLeHxDt/f1wBaUwEBLeQERvGSr30lgL7dSD9sd1xGIoHdWJuBqNr+mGiCzEki/QgmwvjqoA2l9dwSDbYj74qCEJd0SSyO5sU1r1sAWF+nyvY45MxABRhpPxBun0w6jCV9BjfY3G5v026YTI6xDQ2ggDbXY/XBtB5qgYTLExFWKhdvi8vp54ZLrIDvdlP1+uH5FQgMxIAW5F8c0CQs6K2ph8LDr5E/QWxRkFQFMojlNaBnNx8JF7XHnjzsF2BUXsBc9Th6Z441Isq6R4jfp+W2OEXdVUxqT1BYg37WwQcFCU0tw3gA26i2wwRERY73DEkhFtue+EaSFIlIjJ6hrH9nXHh+GH0vqWwsLW+vXBqwU2JNR1lCrdGOjSWNgLkdu3XHnGmMvrCFbdSbbna9sKTlKDqRze5sL4WQDy5FiUMgJUkWIJFuvyvhT90c2XA2iQyLOWHQqB4SMIJi0AK7FdRJA2sfpjyRsulfHK3Q72Pzw3CYzKY2jKsgF2ZbHfoL98DBKqUoNGAX1OY1bYKNz9sdDhVJOrzF+1/nhTIljpZm1HxIx2t06/nbC0UoESIMtrDT6dsWAVWpIUybg2BHltt+84cZ10/CpB63OPM2nqwGo28iO1r9cekRTKo5YBHla5xZBCG5XTJoUNpH7sNxETTaGjuOt7YKhVBGwKgFrC7W7d9sdbTHC8nhXawsLXxStBPq5ulI0ZCreFTve4tt8r9/LDs68uqV7Mg5CrYHvc3wZFJGLLcWUjruemBqZdeXs92dhVTp9A5t+3FjdXyThlBTqb26HpgYkh20tufPHigC6TsWv9MIXUWW56eZxb45KgSo7OKnMxIsNFFS8x2veRtrD6dcAnLuLJheWro4SN7Ien8cTwhBlMm22267YkIY2WMeK7f6gwkuRNbJ3VbGR54DeTiWJGtf+oJA+uOmhzaNRp4jhceYgsQPPc4srxO/WZgLb7DA1NS5hLSfpCoyqsSnaEuqIVknJC6mblruEUAgsf1iB12xbQXCQFHEM3KiPcKsKplz55GuDp5YAb52OGZKKvkNhmLRbdowQPl0J+oxYWp0B25gsPPp9CMIPLUWJlJ/7QD92Bc0hE2CoU0teketq33yQrbSyqgXfHo4ap4t4FFwdgehxNyJq3QELe5ub2+2G3Kpvso9B1wt0jcprfBQZy+WQ2eeogcnYxODb5gg3w2+UVo6ZxVHr1jXEdmGc1wziakRAYxYKdYCjYHckX6dv2YEgq8+jrqikaZJI1fSmg67G1yL7dNQve1rW2vi2tIEpLqjZUuaHMgbR187bbnlpjiUGYEknNZkXpvCvh+2BqurzSCJU92aV1jjeZ0msFZhuqi3iI8gd8BpNxEpepnenSnivrDSgAsBawJNyL/t74gDgbmyPiNItKmpKCrRQf0zITfp7uDhPu9fzd6xVVV2PJ3JviApKjOJc3jCVavEC0jqGL6V1AEEgdydvpe2LS9JVVWZUlBT1ZgllQyB3BOyvGG9AArsxJ6AE9sFGowFQc0tm6FFPmLqb1qlr/5i23+1gN48w5wXmo5NtyOv+BhVbSiCpqEy7iCXMqqglaCoSalMCgq2kEMWuxL2sBvpuWt0w1RZZnuZy18FPOKWSijZy8qjTUaOYP1jdb6ALAE77+puw7+SQK7OaRPHmJcD3ilAHW6EkG3ocKejqmS75jGGP9in3t874WmXVlUwqqXNqGnephGhJKhY5HK6gxCkkAXWSxPUaT3tiRqshqVDCjzuoYPoKS1NEVUKTYkupIGyyMNvhUEnfBCjUhUazAVCTUcqxALXylvMoL/tw5TRAqye9vLMNmuw2v6DBtPkebSV5hhzo1tLQwF6uZabTDzFhaQoG3ZtkcarAEqOl8R/D2TZlmNFUVkOYQiuIUqtQY1jK6wnik1jSQXXqOit5AETRdKLjgiwT5pGjYMqQhr3U8tQR9hhl1l1G+oaTuLY68dZRV1X+mqlKaCnkMENRFTyPTzuG3dXAOtLK3Qg7gjEfw8+Vyvmxrs+qxmFNFJJAtgkEzol9P6xJLbAWAPnvbBGi480HnLRyKktJWMxmx3uNhhvSrklrove1rHDdDU86KLUt53RmvGdQsp3v5fI2P0Iw3PJGHZXkVCDudQ6+VsBpIOyM1AeadkCICI31AfIXx1ZNcQIZh53O4wMNErERSxM9rkBgdvlhQjdD4tKm19gbYp8BWDIsnQ4Ui5Nx3IO+HAwYAi9x3wzK+y6SqEjqRhxL6RcdB1BuPlhrdldynAFDlipsdwSceQEPe5a3YYG/GL2BS536b4egmKXDpuOhvbEsrKUugOW5T9d9iPyw7FIhW46Mbk3uQPLDfvKsCjKbEbWOOmVDGqKq2UW8I74pw6IZulmRrFeovu1uuPRWlkYmwF7bkDHEPg1PpB6b4WVUBiGsC1x4rYXF1cpq6yOxBFr9BhsuqzHTdgPDa3fHi7LGbHodrb4abWhFkLdO25OHKhunOYOZa2/2xwuEF1LX62wkvLp0pGukdbdRhiQ/iXkRh0vfbAkwiTomTUzPZSRsB1w3JPzYFjRl3ax77YbqGRyAGK3Fg1u2FwwxKV1Etq8tsAboTungx8JB1WFreWHl0rYsSf8dMDyXDKyKGA2G9r4eiIX+sUk97HBRKMLlTUKEtG4GgHrfASvrhDEsxYXvbBFmdW0lS4FgzbDDFUjGymYEWF9+mKLQqJiyHdlQDYkHDUjsb6bLfa1/wB2HWEZFgxAFhvhMynQWV9PYbb2xAgJPNMq43JF2tYnHYgN5Hv4dwL4VCE0EDoB4t7Ww3bmnYEKDsb4kIJRWu6KeWwUd8chYMljtqPcY8oQoNJNx2wsXW3iO/TbFogV4uNaoLm4+ePNIIbrcj0tbHgU1+IByR3/ALjhLOzFw7BiR0OKhFNkuOViA2m6kdcOjlsbS6itui9R64FQBXARiSe47YdiY7l7WGDChTjF2SxItfuMdQBjYEAeg64UoGggA3GFRWU6iAfO2BJuh2TYUowKPqud+2FxjXcnqOuPMCxDhdQB8O2FIdLBm289uuLJgJoKdLFVCoN/liWpIOWhMoKv+sfTywzSQokfOkVVYfCBuQPl545U1JKe7wKWZtmOo7H1PYYTUIOyYzu3K9WVJeUQRSMAAS7BbkDzwytnCRqDpTaxa/54bhGkCMMZAu5Y9SfO2CopURD3N+uGsaALpT36kuOyEAagTgkT6FOg+LpsMC67MDsAcOlyWOjwsdtWD0iEAJCJSZvdpnVSzqpZFvbUQCbfXDFDm8NZHIVh0yR2WRdFtLW6X6X6Ydp5QEbWTdT98QVQhyiskzFFkmhqJPx1JvoJ6MPlb9mENYYNk99SYgqzQPGWQEN0vbHa2lpqyBqaaLmxyeFh337/AHw3StqOpGuGGxvsRg6EC63Uardt8GRCjXSCnsjyDOKbh6aozActKeTlo0qFDOnZh872+nrhl9a2DIVHcHBdTWZg9EaGWomelUiQRl7gNfYj5YinBkvIslyNiQb74m6hkFGJJr2Cqw9BhRlZpQALW+e2BYIJI4dRbdhuSO2EanH6xW/U9cTSUOqFsP8AJ6t+lM4sB/UQ9D/pPjZh0xiH8m+XVnOdKWJApoD/AOZ8beMfnrt+0/LdT1N+4Lo0jLAui+ONj18cOPFI180Z6uviPOnDXH6TqdIt35zg4cp8saaLmWQKosQTa5x7PnpY+Ic7JLGVs0qduoA5rfY46ksUdJpNQw1joTYE+fpj9S5cP9jpfyj7gsYNzKak93U6ZkDW2NmwgrRBQRzV32B32x5CmkEGy/62HlMbAAOFJHnjYJVEiVyJIdIK1Jttj06tI4HN1j5bY4vL0lQBa9zhcMQL3AG+KkhWEXBAFpXDDwvsQRcH6dMQ9FBFktdyZ5Zlop6gilu2qOMlCOUW6gE7i+19vIGy0qDkcxxpF7XH5C2Ba+joamCWCrT8GQXNrBlINw3mCCLj5YoEymObaQm2y1iiCFubG12A6n645FCsajc6dfiAIP0xFZTXVKyHLq6cmtg1KJFUgTRgbSbdDYjV5G/YjEojgMAJFYm4uTcE4NwLd1Kb2u9auHs3df51U6CTxaX28/Cca2MY37M2WHiKCeeoSONQ5JchQLqfPGrjN8r6DMqM/wD45f44+J+UOhVqZkxzGkjQNh4laRsEcNseOAf0vlf/ANpUn/jr/HCxmWXsmpa6mK3tcSra/wB8eBGFr82H2FRY77Sqgxcd5gFj2CxXPn+GMQKzE3LXta4G2LB7RqemqOLq2ojcyCZI/wARGBAAQDY/TFMyNJM0innpIMypoVlMUctVTlI5SpIOhv1rEEWBuLbjH6S7NUyMpw4I+aPuWWo6HQpB1SQKBIBY+W37MIWMpJ4UCf6Snp9Rh4UlTGQkhMgHcX3+eDaemUn+qt5k7Xx29IKsEpuCpVkKNIrg9HHQ49KYGl2sJR5jtjsuXNE/NpgbC90PYYa5aPHsG1W6dGGKMtRAgi6eqZEjj5jsiItrljYbm37bDA0ySyXkQGQj4QzXAPy8sQnEmVy5tldTl0sjPDIVu67lSpBFwPlh7g/LKPKclNM+c1czxRhUp5Y7LrMrksDvpURlBa5uQdhhgc14vYpBL6bpaJCejhl9/LT9TsOgA2/LBkSpHTnnmNA+7MJRt2sL9/p1w9HHCZPCUW5FtQ74bzHL0qKWanliSRJFKuNA7/IYWGwZBRHURcJlHkEj8pCV03INrEX7+uH6eZJA3gJUbWPb+Ixn+S8LcZ5Rn2YTZXPBVUS0zTsa2VlR1XTcdwGF+ptspOwxasirKvMKeaSSiNHUU87U88AcOEdT2YddiDiF0O0goWHVuCEfUR6mISO5Jvse3p/DEVXUh3miDPyzbQo228x2xYIhqukw029Nj64aljchirszqNmsNQ9P9IYZKjlXZKuIwf8ASuoUqRf4V8h6YTl0Esk+mGrWIdQSbk7bXseuJc06PIZK2nhXYkSL0/u+Rw9Bl1IJgeWq2W+tV3v88W0lrp3SyJHRA/og/wD2zSf+b+GPYmPdE/zv5Y9hvFHRXoPVPVaj4tQLD1HTDSMpF7jbywxEQjMUYILbKQSp+nb6Ww5BK5OmopSUJtzIvGp37j4h9iPXEqNgoKbpCIhmGwZDHbyN7jCnqCQSD0PTDLKj/ixtqQ9GBuGHoRhkl1udgL/bAMF0x5tKIjOstdum+5scNS00NQyNIWVlY20Na+3Qjv8AXCRIsq+OwPQYWEbQF1gkC2o9Tt1vhlkATTZZIzMIpUdV/UviOhE7Jd6ZoWX+0Bv9Lm2JKnMwd1mIbyBIO30GCTPCVXUVbbYEdsRuoCyh0lQBYpMdQ+HY7ix9fp+/A9RIvNIBvcbDtibeihZWI5qC9wR4x9cBTZRKVLRPHK1twrWP2OCLuqU5joUapFwGIPe98eLLGjs8iJEgLMxAUepPlgiShljsHQr4e98JRXUWQAqy7E2Nx5efTE1oNLoSonRFCsNQJ6jphZdCfESgAuTfphkaQRfzJAvbDlgsezA7/wDSXNwTvhkKArwCSeKMiQXB2IN8c5Z5pIcqDYWuLDz7YdSSOKzAbKNh5YQ8IcI5TfVcXbVYm/niKc5Tc12TQdRNj0S+49emEILqxU/EL6WNvp1/ZhcpkMRBsCV3sxFvtvhiGNREyBokIBKqq2HXyviQq1QlkMwW66nBBAVjYnp9euGo1C6dARYwfAFNgP8ABwuOyeFZAGuA1vzFie+FpqkB+EvfbW1r+fnhRfCYRKWutmIBPTa7YYkVxURsEdr6rtr2sAOo79cEyRuquwDeFdkt5C+22GIpCxR9LWdQ24sd9+mBNzKFNRxiCyRAIqCyqOw9MPjW8EjRtHzOgD3t+WEyImgzqSNW1gf23wwskciOyyRyBWKMt72I6g4MOkqkuZYnRopY1lVlAvf1vjsiI0BuxBOxHTbyBBw0bCYEhQPT9+2PQmEIVX8Rb7E3J6373xb/AEVAuXBZo0jV7eF9Y+Ha42PX549zXQBAslyOgXy7HCpJEe2oE2A3Y4Zk138BdgLE7ae/UXwnZFulyayS7AgE+EMB/DHCqupTSZCxBte243wl2Isos2rcX+IYSXEbaS5MvLLiMDcgEXI39QPrhxMCVQCUuoxKJInhcsQyMynpsDsT88dkkRRuWUDqX6fTyGGiHJ1aAhsGDSDbfqvW9/44cMkkVxcKxPQdvrgS3VdXqiy8qSO0jRTRvdPCt7AHz1dbHbCVYrcMLW+2OqSj6zfzPi64QBH0RG03J3JNrm53OCDSFQKcUayCJQN/h0X1fW+2F2KhpGHKG92Z+nr9cN6nTYuhiNiLbljve/l23x4uNRSM8trBtJBNx+zAj0lXJOREq/UaiTY2vh1IkJJLsWHkbC2Ai0hluxvYkfLB1JpdSZQWN9rMcR9yr5IgNDGBbcnZbnv54RJGZCikDZtRN8JeZWdiOWANgdOFl2Oy2sB1JwsGSpK8sJvqcKL974Gy7UculjVgWSsnawG9i5w/rsBcjr088M+6rBUmojYaHisVHZtbsT93wSIFddHNxp0G29/PHUjUgE+K/mcPxqGUH4iO5xxHYbKzar9wdsCTCidggDOgVguo223wloXao0g23ttvf1w7SSuJIwGDnUNNxhyFuVUySOCxJtf1vgaYDnK32EhG1OT1EcevnKqBQQSOu1/P1GCKXhdPdUmqq8QnT8PKa4Pb92+BGrJahI4kChYvEwA6m5P3xJjO6d8skcoTVvKXEbi4AtY98bWtDdln3MlATZTRu4SPMJwCLD8PUt++/bAU2VwISRUO41Bf6u1x364DrM0qWVkMixAdNIscBe8ShTI7s1jvc37YzVCJhameiqxxZV8XUXEZpcoo3qaWwdOXT6iVIAIa17WO3btiUyOszWRHTN8tnpqlJQjAWC7i4Nib9L/bD9RDE+iV9DFW1LtcqR0Ix1KaMyPOlryEtISo8R8/8d74SYMSiggSCVX3y+0VdncbiOOPMtLMb6WiICgqV3B5kfQWvbEZS5yRmgkpjrikuHR3EbAs2rYkg6rg7m9r4Njo6morMxyZs7lipZqiJ3pVXSjhrNrO29j4dt7jy6SeS5dWz59V1kc1fTw6JHozOeVG1PEQol36bEMSbC5Nr41aQ5llzg9zHwg48+MFHJMaSF+Vc299j1vc3HhJ1HqOgNsQmY59VClAqVR4KpNaJDpuhDWJJBv0w7QvJlsUYqKKrlllaSomLITojKbsV7C2tr7fMjBy0sVdlkdU9PDBM9O/Lj2DrCLqj26gEoAD0v8AWyW0gRICa6qeqK4OSCbKlmRlYGRg5MVmUhmIDNbc6WH3+WD83WKaSmblFlVtGgmxZTYmx7fCpv6YqAiraegZ6XNXpBHAJ3j5qkTSGYxs1iwIPw9iSL9gTixcK8Sy/oSop6uB8wrIoXfTHTNJpQEDW9iABd7X6bAnrusU36tTU9lZhZocgshysZzm5yqjpYudK88zAzFFaHSHsSzDT0C+e4+Ii2JnPjzcnoJcryud6pYfeKuVWdXLmPcnUB4SUkBsbARO2w3xG5ZFUQyV60A5VbJlkUUKrFeaJzU0wDNe4DFA5+/TFj4eyqhGYyZS9ZEmaTQClk5dXLaCRkmjkuQgPRlGnWRqJLEjbGxmrTBWMwHWXszhzXK8nzHI6ioqI66meLmq9MwH4EKKoDAEgEKSvTVYG25GAaOGoUvLHJO2WyqWMclakk0ugqWmKqy28JAvpO2oG2kWIaupKHhTMoc5y+aujmdUpImmZmEzrIFmkkU9bKz2BtpsLnqBZazLoMvmy/KRlXvNUzU7yNdpBGiKGlCAJGiNZiLByDq8RHQ9UBASn5gtZVBY6aoKWJeMugZZ7rCTJ4bteQEKgOwA2FjimUElVBlNfMslOsdVohETlUaRGkjlNmuAbeEkkG2ntif4rpqnKKaTK1q/eKmGsipeaabUoQR6zNqJ1HWzkhdO1tzc4r3DVHls2dZXR1tBFLE4mMqms5aziONiqgk2jLEDc37WwtxBdZHJAhSOa5Pl8FFU1MD19Tkq1DSR6HULdFUKgZ7XI5guwU7MNt8V2n59ZnVQi0SSGSUyyK2p721Xu6Dp6gWxZc0z6aKbLcojSnKZXmDSwTM7kBfEy/hncgE3DEAt4bjzgMwlq6uvbPKGeaGYhqma0Lx8nXISpFr7HULEnqbYombIT1U1SStypqWOS6xvaRllBEpsq2IHUeAEHuLYi8yoaNq/nPTcwSRvqtNp8Qtv0NzudhubYlJ6imqs2qKvL6Ojghl0+CEkhQoCgE/2iFDHpu3zw3WKGqqZZXIiMoDaSR4TsbkG9vPcbYTUfFloYNQCGTLYMsFDJPFBT1DcxAdd2LAL1Fz3LA9Oi9L4MUa5NZOrbbBud11RUw5RTtFTF1ilYxikjhZI7WUs6gE+ExkrfYj1BwBE4eVFNgG67Yz1NwnA2hILBXEZIub2tbcYUirpuR1HXDuZ06xzqFN1ZbjyJxxBewVdwLjbrhtN0iEV16nVX5mtlFgNsKEiFSARcCwBx5yLMNlIF/IfbA4AUFyytceI/wBnDN9lRREMXMA0sNX+ltfHihhUhtieunHIAsOzEeZBJ2w8z6hqVu24YmwxRMKATsm4lVibM3pc9PTfCZdJksfELWFseey2N2uT0U/tw2HVHKsQD2DADCwZMq3NgSlBWHwhkBPphTixs5BHYHe+OGUW089DbsSMJW5fVsfIXuSPTDroZXIjInwi/ewOOySyAeOIP3036/fHtTqLgaiD02w2+oTBi1tQ3W3bCiZspKbeZ1OoQFVt02NsOBmIDBrWHTDYeMMwQrv4iD0GOgoQWUrci1xv+WLAIEK0+oVIRZ9RPfDsDiF+ZIfCTsPIYZUx6UkLHYDt0w3PKltQNhggoF4NGWe73A3Bv1w3LIks+xUnbr2OBVNg7sys312w05lDqWO+xuPK2KurlemUq2lrEX3AOFFlKAG6t02GwGPCnqZ5GZIS+kjoDthUcU7rpAQWO4ZgLffEKW667JKOWI3ZiLC1iMIVSHNla/fV5eeH4adt21xXtY/ijbDi0dTd9Nqjz5R1/wDpviFDCELlTpuAe2HQ+tbEqB5g/sw9DQ1TSiNaWYMD3iYYdOWVcrkaCpvveyj8zbFSrQUrGOwUhbjrhtmYrqsLW3I3xIfour1KphhAP6zTLb7gnCzlU+yyVFEsY8pibfZcFKKSEHHe+5sbeEYWqFR4VW5wcMuIsVraTc/6Z/dhw5dAEBOZAG36tMWH5sMQoroNC0aKCfE3S+O2YsPTy2wXDR06TFaitkdeo0U+kfmxwp6PLy4YVtVb/SjRcDqlVpLroBWYMzE2t+WCqVV0GaRDr6KCOnqfXD70dCxDK9TMF6iWUFD9Ao/bha8s3ClUYD4rYE1JTBTdzTNTmACpBG4dzsEU7lj+71w2rpCGjaeOV+rFSAL/AMBiVpoqGlivHDTvIdzI0YZyfmb2+mHpaqRrsrW2tew/cMUyAFbpKiolIUaQTtsb3vh7lzbKFLFtgqxEm/zwalXUKpC1cgP+vYD6Y9z5iGvK9vPzwTnJcIUU0+gsYZQVHiJjYD74S7TKfht3uB0wStRUEEiRrdiWOFrNUm+lz13Jvvgw+yoNBKYjaULZXUKzeIdSdj0+uHlCvE6soZfhYHob9QR3wSxjbS9TCjkXBNrNbBLUFK0Ub0Ukkpa94nFyp7fPF60fDKiOHcurYauaghjkqILGVCAPw1/s72uf7vPaXZHRU36qGHa/0PTBeXVNRl5eMyKhkBLalFxse9gcRjEC6idnUMSNR6C97fuwXdc2yWQ9rkJxDmOe1sZochy+tLh9LVHJ2G19IPQdRvg+PLp8jyWko6uoaasYFpX1E7nt8hhEGY1cFlhndQNgAb4RNK9S5eeYs9ticQNQ3Jld57nSoJAFt+mOmW51Eq1u1r4aLKoBI1E4VGFO53Hyxe6seKsPCvE2acPSzzZRKlPLUKqykxqxIUmw3Fu5xaE9pvGBS/6SiG2xNPH/AAxn2qNraS48wd74fZ10iwYD/SFscXF5NgMXUNSvRa5x5kAlOFVwEBXs+03jBkcLmMQa+xNMn8MO03tL4smAb9JRWHxD3eO/1264oCaSx0up7274fggmaVWpw2ljaRRvbybp9MZD2Zyr/d2fZCviv3lSkxNdmUs0smqSomeWRgADqYlm2+ZOGZ0lmndguqMdLm+3bCsvLJWWZrhASxI2FhgIytzRpCE6jbw47FKm1rQ1ogBWXWlFRxyq5Qrv2Gq+CgrlwStz5AYEpwUjMjDURvZUu32wdEbgoQpHXWdr/bFkQiBBXYlbmW0kDy74KRWV0uRsO3XHFKqg0IGcjtucGe7PH4mVtYAItgXC6a1tk6pZYQgcAsbkeWKbV102RcQz004qP0ZUHVFKSWWKTclbnz+vYfKys66nWMDbqQ198Jr8no8zy16KsRmV1JUqbMp8wfPFMABUe8kWReVZuZcnYRNCY6izxkgFlJFtvK4tfAlMZ2eRGRCA7W2JAQDYnyOKfwtVT5LxAeHMyLuT4o5LnSIUjVEt89JJt3vi9Us68rTGAVbqBexPpg3yBBuqpw8gixQyTR2szFiN/Ath+eDY5Y/C7yswPisRuvngWqhvJoIshGsG25+nXAckxSWNI3PMA8QWQhQvkR2P3wkMTHukwp1s8ngpzCjrIGNgBGP2nv8ALBlPXnRHBY3Y6tRb4cVSqF5QGChWbUUeVnBAPbcfsx2SrVWGhdJC3AtbYnfBhqWHkK71piqo0CKjXBBINwxHniHFZVpF7sJnSFTflo1lBPW3ffEfTVLPRSNrN76djYdvQeeIqnk5MhLzSFQ1rrYg/YYaAQIBREhxlWVKsBwH1AHYC4sfzw9T1tNUlmgZTZiNhisCtlkXTHJp1Abb3HzsdsF00ksVMZZJo0jAZyzOD9gd/wBuLFipKtsJjCnUpd26WOE1VJBURs3hDLtq3DDfELBXElY5WKgi5su3niQppfLUxIuDbEjmoIKYqcrrYT+EizqB1B6/PzxGVr1FOAr5bUyF9gIB1vtspxaaaUBdgNPe5vj0y3VnUlbm9v4YGAi0kc1VKJM5nk1pklTAjRqdVY8cA3HXc6m+i4PfLKvSHqM4ooHIBKQRvOB5gE6B9d8SEqnYkow9cMOukWBABO2K0wl6yvCiy0E2zDN2XSBpQxIAfP4SfpfDOgK5MFbMGe1xPEr3sRe5XSdwLeh+2FuvdgAbYQGIPisFxCxFrXWgm0a5Vj03+OFi6/I3sR9rY4kQfdCtr9RuDjysUYvESB3woCKSoEjvJG+m10Nh8yOh/bi2mEBuENUUyyo0d9DsLXv1H7xiPipDRSho2NOCNyCXiPzFrr9NsS1UJIoGHhkup0MD4SfXyxW+Ds9zqs5sGe5FUUrQSpCaiKFmgLtchdXQGw88MaC/ZA4tbZyk9cn+epPuf93HsSF6H+wv/hjHsXw3qaqChFUDmlRLHzWJKs9ypt23IHyGH45mVQqA6QLEHqfnhrx7FrBwouFOwPphsTG7LHo5yhSyatwCev5H7Y1xKygwjmqFvrLFWPh1kWYYbvUXZoXE+ojwSeHb0I/hgUTPplWWNgL2BLA6thvt0w6kugktuCNh5YEgAJjal0zzVeqaO5pnYnTFKti3e6t0Yb9sPhJRYxEuL2Gg3++OyxwVaFaqKN0CaQ53IHXr2wNS0tVFIJqKt94QgERS7EjsFcf/ADX69cYyHAytYLX7IxKp0GiZVa/0PTDQMR0sfhudJI7YcSshlfl1cTwTbWSUW39D3+hw9NDENJkYoeo9cG2p4pbmSdk9SOY0CK+oHYb3w1VtIkrOqswK7jz+XrgJEn1lodyu9x0w808isVdrt0NzbbE4gVEHZEpOwRCZGC26G2GpZqaTwskJUjZ7Wvb1GB1lS4IBsSfCTexwqSCSQHTOyC1rdRa3SxxevVsq2XDR0z/iNzVuu2khhbz/AMHA0tESx0VETC3cFbefngyJSjALoXY3Iv17fTHHcI51K1uthvhg1HZKJaEBJSViadUTaCNiia7fMre31thpY5mUWOoXuSN9r/uxKGpkiN7gb9RtbHHmLBfeAZFttzFDEj64ICpsgOjqolVdTaRSV6DSDbp8tsMciWKSOOByUVDqDXZzvt4ienXzxMymlZGVqdhc7srkdvLpjhoaZ4AY6iVZCD4ttv8AHniODwqaxp5qEaUTQq/4itb4XUhh8/LC6fdrKddhYgCxt3GJGSioirhKlpGZfFePqehNr4HOXGMAR1kJU7szKRb7XwuY3TA0pLzEghXCA9D0t9scmdWp1jMmoMNxuD8wRb88ORZfO0pX+jysCCGV77fWxH2x6qy6qQlhGNjtokBBNvK++D1NVFhhCrIjrp2kDXBNtt8KmQugbUxUdSCDjhp6x7XjNu2mxP5Y61PUCIssc7Lp+ExMBe/W+JraNkvQ5MeBXaznV+qym2/8MdVtSxyLUbXvp0ag3p12x1IZWXVHDLYdfCbYHkUR1B8PLOrUVHQk98RzgQr0nonWlGlTqVS2x3sT6b7dsdJkcFioJuBte/2A/wAeeE6JGLJqKgDckWuMeIhEtgRv13uLjAKQksxK+IlCG+EqCL/PCZWhc82SAFiCL449hplkaIEAghEN2O243/jhIOgMukgnpc+IfT+OCa690ULxkJQSKzGx3UHYj7E/bCXJ1kaQPPe5x0Mo3k/EA2LkXthTaVfTYP8A6QI2O2354LUFUGUnUSDZbgizALfHbJIQoQi3WxOExAtL5A7HthyNZDJIjSxPHzbqDD0XsL36g9/yxeoFVsveBVsLhhvftbythCu9rs58wL45okVWMrxly7BQl76b7Xv3t19cOwGyWIBvfr0+2KBEqguWB6ktvtY2/bgiAGysALA7AHDAQCMWYHy2w9HGzWLOLjyxTrlXyTyoNIUWA7XOFsViYBnVh6YRGoFgrgm9zdbjCJCwk8bIRe+wtgB4KyITyyC12wxEwetqrIdcbrHI39r8NWH5PhUToXXqL7HDmXUphnrJZXVufUawo/VUIqgEdj4cWUTQnEXSwPiF8O6SfivhwLuCF88IZbjWCLDrc4EtcmHSEqljJniKRhxqGxNh188R4zSljrvdTPG9RLq0qDcnSSCfoVP2wZHOyGPUAFDArvbf54e4Zqpzk1bRihqgzUU8zTNC4eUBBHoh0tZR4WVZCDbUTYknB0qJJ3hIrVO6AAga3NaTK8vnzKqMklKg3aOxIvYCw+ZGEJxTkdXlEFTTV1Mkj31JK4RvsfLHvajlEEfD1XRU2V1OXy1lPTtS0886qsSRxIvJjAOmQ+DWXBHxDqWAw1m3D+W5nM1ZV0KpuUp1lN9IANrgdmO/nY/TEqv4Ig80NOXmwXJH95dX1bONQ0kEW+eFKxphzHUFG+LxXscR5jOXCXLqKgmqOW4HLiVgN7atGq5IUEXt+W+HeTm9XV0lHRZBXTtUyuqmR1iFktqZdXxCx7HCA6TJTibQuVT/AIqaFU6epGwx52mNgWBA6YakklpahKZ6Ood5bcgJpYS3Nhax3vubeQJ6YGWqzGUZfGmUSCprYjIkbSKFVQSAzN+qCVNri+3TpiyqDwF6pp5JFYKsS3sQxHQg3v8AfAPuEtO8tQnKkMyxxsWJAIVwTe2xGkFbG/UHtg4VNTHJNT19EKc09lnm1Awo7AMAGv3Bv0+uGFrq+VYzBlTTQyEW0zrrIKhh4TYXsb2vihU07IHMabwl0iUstPPBOZFdoWRZI4isZAGysFIsDcdBbr5HETNT0lPNV02XwTiSniV5KoVBh1x6kDIilSeuq2/ZifCMWRI1kgWWTmRpIobS3YEd8VuseZa/MaOCRYw7JIzsoLt4CLIW7g722+YucPbW191LfR0iQE9UUeYZdFBmkVRl9PA4DU0AdHl0SuwCbqDdQCTtcEE9DbExwVn+UU2a5dBm9FKjUcjRt7vDqkqyVmj5aqDvdmXc3FwAVtipPRSSUquyqlRMZmSY6lIKIWCjc2L6bXO12+ZM5lH6aizWirnpKOsVGE6w6tSy6CIokfTYHxmx3/WDWsDhgc5sABZg2ZKnMwmpWnzhcyy9JZn5CRxmKWJZNdSC2zPqLKCxB02Ogmx03xD5jHRJXVBFPNDSyCnppVSMpFUyRrrdr+G36pFhY61JHW5tQ1TWVcxpM1kozQ2jgiYCIJNHEkL6nkC6C/KYAnVYW2G5BNdBmeW5vma11TRirVobyU86tDDLZj8KxtbowFlsNvECuxuN0bBa6qNbmk8SVKwSJLM8006sxKSIw0WUA730krcjcXUHELl9c6v4I4oJonUKjSJtchGuz+IAhh3tsewxYcwoQI8vWnqJJ6htKSSROFLS6hy0ffSQAgJI3uzNdgDaY40iGVZPlca5dUU09RSipQRNZGZ2AfWjgFGIv0O+np5KqudPdVsaDuq1mXv+hBVGR6Sd3McxBkEjRobLc/ENagFrne/lh6WNcqyOKpiZI8wzRNDRrGPDG2rVqLWKt4gAADtbfHcvnrDmtEIIUCJUwsCiyKWl0jwANtpBvckb+oxOcYSR5lxHmtXNJGIjVJFTBGWS1njVrXa5BjUeK/UdRgAIRm4MqrVlfBFQTUlZTpPm7wrEsmiNVhRDZVWwNyGRb7Albi++Dsoyhhl9NSKrR5jNTzzSgS6BpUMbsNgQuknSQTYixOwCcyynNKiqWqagkrGlSVAQjDlkqHRQyXDWD2Gw6EdsF+zumlp6nKpC0TrpqoGMrAKpkhkUJG17Xs2u/bV18nEta8SkQ4tK5NRikoKWMUvu1UyOaxH+IS86Trtf4OX13xH5ndk5J0qxVlQrfY2NjtviRzoCLiGsMUyzRjlMZI1Gl25Satx13uL97YiamNpXRBIiOSWBZQCdugJxirHvwt1BsMUtR0cMcrxVE75Xoi5ohqIbyTmSNTcGxsP+9ucAUay80ENcL8VhiQjzB6vh+kyqhy5jV0ssx94gQIohZiRGfMbXBY7XIwzTZVV6Qal1RgejyBbfx/PCx4pmmAjYiWVboCLb+C5wwzxx3PcX6KDbHGpKCnXxVjyOBcBCzgedzsPzOHIfdDfS9TIQNj4VF/kL4YLKy6AmzO0jXtqBHlhDzrYBdhfuB9umHQiayxpS+2+tyR+WEyCO4001OLDYiM3H3wQJ5JZM7ppp9RIKLa12sN8OQSwSIpjBZiCRZb2/LCjUToQVKJYW2sDhE1bXOf8AKH8uv3xTiUTIBSpHhj/q4Wbe/hi3GOye8KoiWn0k+QQH88DhHkYtI0gbzYn9+EzwamCrKCOovigSidcJ+GEKbiiiLE2Jd1JtjphjQ2WCnv5if92BoY9JIYsVHcYf5dOsJYTEN5EDBazCprJXnWygxilJBA+I4YaBJCeZNCSNwAGOk/lj0CIzX1lRfbe18FiKC4Z7M19vH+7FSReEfCCEemijYAzRhiN7Qv8Al4sMpTRGS71LAekXT88SJSnea3Rh1F8dMEO4Wwvva4scVxHHkrNJAyRUkdxFNPqB31Rj9t8cMEBbVK1Q9xccsqv7jh6qgtazqSOw7Y8pWNBqO+9rWwPFlDoCGjShja3uzyEkk6pSf2AY6zQgrpooTY7a7vb532OFfhmQkne3XDsKJrDPcHftgDUdyKsMCaNZNbSukLuSqoOuGJKiUkMiL18hg3lpZmDAd+uEaYNIWyjzOBNRwVimEMKmvEmnmTIosCA+x+2HJKqsaJknqZXjLXClyd8OM0drCVB63w0JUIIDqy/Lp9sQPceSBzI2TUssoAIkKDqLPhCTsAEZmJ2JBY3OHZIWNiG8IPUIcdiRgSQJJR0HgP5bY0U2gtuhG6aadmNluR20m+HvErAFDv8APCkcrH4Y57+XKb79MEIs1iVhlO3mv8cW6n0CgsmFSZ7FFsO5OCBZRZvCx7b4bYyGbQ8Ei7bFpFt+3CxqJs8ITzZpR0+mFmyYHCEzOugEGVj/AMcIDrsFQufnth2olWNCiTUsl+3Ma/1GnDUUz6TqMSAC9xcgfkMQJbiE7HI5UoW26EA7HCiCwtoa1rbDa/3x6KPnprWrpjbexJF/kbdcOxU9YzhYqdH73M6rf74hFrKwSUuOJ1HYi3TBkS3jKlwo7jphpHnpFY1GW1Cg3GoAOPyxyWtpwtxIFNh4WsD+eKEhGCNinowokuCp7/TCgi2Zy9/S9sDmqpXXUamNfMahhv3inIuKpBc3vqGLkoXaRspmFaJ4EEzCNh+t1wNyx4uUzae1xgRKmkuFesjPnd164Np1STSY3VkJ/VIOICVA4G0L0sMnh8Rb+1fDqa4dLJqUg/EDhxzpWyg3v3wh20xkMov6DDERCflqRKESoLSKAbEdR8/PAmZ0jUuiSFlkhkGzi5APkfLHacEMNR2J+mDkXQhBUtHILFSe2ITBsg0zMqGi1bknfsL4cs48TNt8sO11M1PMI2Q6CupHJG/phkeMADpbqRfDQ4FKiLFdUlyVW17dMExxMUJ2BvgeJbFWVdx0IXf74MQHUBKPCRsQRv6YPZGAlBBovqa3exwqNb3a7G3mcIZgPDrJUHe43+uHEcEAq1l8wRhbuqtdAaPUShAvbcYJjn5UTuzAAC5JHQDvgTcl0vFs229yMFU5KatLAW/Wv8WFlRPxhjSTHUoZjsxB6D1+uGKdWN/xC+kk6iALX6DBU8LKkQBIVl8XiB0kC9yL336bY5JpC9FZVUnSCQR9uuKaFJK9TMxKrr79b4maCIzoSGCWNj4b/byxHUVM8zR8q4Lk+IiwA/j6YsIYQUw5LgqpN7jqelul+uAeb2TaQ6rjKlIIjHLqbfsPF5X8sAVdVUwytDFPqnKXJcagvi69r99sLNerLK6Fm07E7G7W6fTEexJRXaQMbWYdN8CPFOcQEbFImmzEM3qO+DqWUGBmL9fhHpiD17gsCbbYOiqOWpCsQL2O2I4IWESkZnkmU5qBHmNT7u0atNHKYy5RhYr07XHQ7YlYUgBZ6AaoS1/EjC9+4BNwPTtiC4qzhMu4eqMxhUSVRXwMVJUW7kX6fvw7kFPxIMllzfPadaaKfSIUQG7KfhYjsO/1xRaXASjD2MJgbo9auGoWRqeQpHGxQFRqtv2v1HrgF2flSalSRTINWrf7YbsoidmkYuBsNRt9gcMzl4oVcCNkYGz8zow9AL9MW2QgLpT1aze7xyEKzE2I07367YEEz+ONVLehY/s/dh6NwwMRVlD2KhTYm3fAUpWGN5i62W4drg99ydhvjQxoMJElF09SQ+giQtLcDTe67dx0HTvjtVZZ10JfUTcabgHY3t9BgTLpIFrwiTE+AO6lTcqSQN+nY/bBdTFGrQqbxh1tYjVc9ev32xVQQbBG02XOYbCSZkDnZidiBbfbtiWoZHVByHdS/W4vt6YjaRQFdQQz6d7DcDztiRoCisYnkkSQ9G03LW7eXcb9r4Eo2lFUlHGkEcRDOq/5w6yfnfEhGzw+IOxFx4RuMMRAlQNAQDqAbjBMZBB/DK29f34ICyJPwT67Fup7H+GDFczKqgWCk3I264iHKiVSYt7XuFvb64OoHuZFmAiQEGJluSRYdQem98BBVklPOFCt2XtcXufpgZwdZJHa+DNaODdiT3JGGpk0Ob2YHr2xCqQTOpXxbdR6ffDekgHQA21yCf2YJkiIS+kKPXvgdnW5Vl26YqJUCaCXYk+E4Xbw/wB2FcsadyQO2/TCXBS4Vmv59cW0KktGeIa0AYDt1H2w5TSVMc7VGXyvDM1/w9ZCm4sbeR8v3YYVtKargm9jvhROpPDsQb4I2uFCARBCB5U//wBmVn/hnHsHaX/zrfbHsVqqdVOCz6IUFqjtdmI8sCFLsv8ASJdekLqXTc2PU7YSyoju8aKNTamKgDUx7n+OOsyMLOlyDqBIBsfMY3wudqT0/MEJaMIbsA2t9O19yLDrbtjifhUwXc23YlifzwnVcaCOp626j1xxmMb8tlZr7qAOv7sU5shE1wXmllaL8EoGNirE7Wvv09MO09SyMHIZlBAXT164FkX8ZgQBcCxBvvhCOASQp1A77bA2ws0pCMVIKsUUtNVCSmq0Ri6kWcC5Hp98JkoKmNi1G/NQjaJ0uoFrWv27dPtiMppySFViGB1D+HyxNUOYtoLF7Anex3B+WMTmkLose1wUbLXPRhIaqnkpNbDfTqS9r/EOn1A64TKDVKZoGEmtRZwwII69e+LGaiCanUyBZkvsCBbtviNq8jy2acz0TNTTMo8S7G4sLX77ADxbYoCd0NRukSoXluAAw8J6gDoRh+PeM6CAb7hdzgkcOcQQq0qSx10YchNcYawFgbsukDe29jfsDjwo86opCazJHeQIWfkyXBA7jUB233thzWCEgkc0MOeqksWv6YRJM1wHQ37bdcGGqoDDpljq6eQ72aI6el+q3BwmOGOpiSaCshmOmw0OG/ZgnSNksNBQqTopF1bfvhbPG6aup6b9fXCpqUs3LDXO3ba/lhqSF1JDC2nrtvihq3VloSgqMpYbNa3X/hhIGlgQT4ttj8OPBGAvYbjbfvhllbYsRuf1cMDyBdLc1spdhqdgWuRbYYGMYTwj9bc2Ng2HEDiRowS1t/PDl3uwCMfLa4xU9QqjouRgCMWBUnZ+9/4Y6sKsgN7E3Km9vocO6jp1SJYEbYRzYx8W9/W+Da5otCqD1TcsSrFoki5iHoPL1w3NHCEuscYBFgqi2/8AwwTqQ7o5J9cJkKhCNtXazWwcCdlRJF5Q1Hzo3VYUncO1m5TkHTbckDywXAtcshXmSpGAWHNk6H/Rv+3AplRZlXcFrahfCuZclHUqALeHtiODCIVAuXZKsFCKuWNSHKFWKtv03G++E+9UJhIvABf9anT5ddOHGERAZ2c23APXHAoGq0hvtsWsP24UWsabFEHEhBP7s5JEsdwujVHbVv2x5qSOA82W9rgm4sPS+/rgqVY5EO+kCwIU9sNVPKeIo8ZZelg5BPr1xXd6qwSuuIVAJghW9/E0e1vnjipTSFiI4TcCwWMj8sPQTLAgWEBIweoA/bhp44pZCFZ2Y9bOVv8Angw1h5qtZXWpaQaD7tHsdrSNufvhEcFMNzCATceGVh9r98I5SrLdUa/a5v8Avw6utXJAdABYBLC2L0s5lWHuGyGnp6DVZjWI4IJCSL0+bKcJFJRsngqcxVVOwEsf1Hwb4KIDEltbuR8zht7OrKUfcWPYD88Tht5FCXu5pl4aONiFfMNh8TGMj62QYd00bwgrU1akAWIKG5+2GZGRr/g6V+G9uox6LlRKoS8YuLqF2xBSVF67LT6DrWuqFS1wSim3z2GORCGeMiB5ZpFNmsoIPz22O4w9JU7/AAF9trj88JjWx6FdR32NrHBBjWlUHlI5dEsYEqyFtW+hQSPnbpgR6CjmkNTJJUFiNKJpYaQN7ka7b+e1zg0qoLKLqWPRRa/zw480aOV5D6WFrnbELRKLUo1Mpo53uBMWDEWCAWHY7vbHTklPATIRWNIBtojiI/8AXg0ljHpjvFva4GGZ6plmCszEgWvpG+KcQFA0O3UeuUqk6SB6q4JA/o0ZYD56/wA8cFDNBUqCKyzG+tadNQPzMnpiUppZT8TkqNgD5fbCpnC6Sgdr9bgYTxJ2RaU1Lly1iCrrM5zD3qSyyGUIXAsANyx7ADv0GI2fIoLtprMxv11PJG9z5D4cSDT+Le5AHTSMJafwX0te+1u+FPcTumNEBRC5NqLxJU5ktxa3KiAI9W12wfRZDpqIaiTPKxniU6AUVSlwdrht7X+W2FpK7s2lWO1jYYIRiYtXiJtbfAg9UQaBugqPI81qJU5Gc1MJhV3j5QjidFUWLDS4sNJt1Gxthhshrlq1H6UqpW1aRKYkNgepvzL26fbEixIBsCCeuG7voNyNx54vXHJQtBUfW8LVEsTJJmNRKXJFuWmk2sRe79L3+5wk8MVMjcv9IyMsKlIysESAG3obkdvpg2Rrsbt9SMNEan3kBF+liMXxfBCacqHm4XqaiqHPzR0kKiUGWNQelrNa4HTDy8HTxx6DnkBmJJK6QbH0JwdeWMERtIABc2YjDBleTVzNbqNzc4gfKEthBTcJVbSJHPm3NjWzAsI7X9AWw/BkE+XsksXESRSaSCYogpUdbbeoB64cje1gIrj1w8mkeEqR+eLkqg0ISPK5EkMrcSVLaLBG5G4+R1dBjk0K6mEua5hK7+NzEFTUdJW5NiTsSBvsCcLqpCJREgsSelsOAqC2o6bbkE9MTV0KtrWnkk0ssUbJOs2YvOrgiV5rOLXtuB6n74TNFFUVrVtRUZi02vVrkqCx1HvuMdeoghiM8kqJFbeR7AfTz+mI9uIcsWXQfe5Ba+qKG4P3I/ZggHuFlTnsp7lEJTUylmRZSxYtqadrm4F9wdxe32wM2XokhEBlhBUAqshUFRvY77/XA9RnkayApSVL3AIvpBI+V9sdjz1JBqNPV26nZPy3xXAqnkh84o9VJ8+rflJLV1BjhVkgT3xxywws2nfYEbEC2GYaWmiaMNTyvoFlJma4FgAAwN+gtgH9MxOpRKCsZwb2JW4+l8JGd65ykdFWNYbDw3NvTyxBQqcxKs4mlyKlaaPL43YR0JUqb2MrkH7nHZDR60d6CNrE21sSPtfETHnyvIB7lUWN73sLenXCYc6SbrRTx76VNgf34rzep0U86pAbqwzZxUadKRwrGf1UWwP0GGP0hI670dKp6g8rfEPPmyK+laGol0d9Ki/54T+mr25WWzEntpB38uuIcPU6KvO6XVSzZjMbho4l89MQ6YVDUyG4VIWHW/LAOIRs1Koswy+UxXtqNhv5YX/OWnV/Fl8sY7nY7ffBcGpEQoMRRmZUrLKySAlUG+3THEnZStnjIG9ioP7cRv8AOOiuSaeoYncbDDa8Q0JI101TqG5AQbehxYo1ByVOrU3XBUlIwZyxVQx6gKMMSOQukhQRvcjAEnEtKEX+iVYNug02/bhv+cdGGW9PUoTte40j6XxRpv6KuNT6qSFiCLglxuw7Y6YYfCbhtPmzC2BJM8y6SP4ZZJB0Xl2+tybDDUWd0jSaJKeZEPYeWJwX9FRrUz85HhadR8OoDykb+OE2jAv7qpB85m/jho5lloW5SqN97BSf34bqMzotFxS1co2B2sBfE4NTor41PqniaQJf3UMb9Oc/+9hLCByStO0ZtbaZjf7k4HhzSkkU6svqgE6kuMIfOKVZtCUNSAV+JpPy2GGClU5hVx6P0k9HFEJ9TmQW/wDiHD6oHU2EqqOpMvUYEGdU8ZutBVuo+LcWA6DvjzcRUVtAy+r0kbgEfxxXCf0UFel1Rvuy3B1VBYC9zL1wPUUSqLwy1MLFtzrDj8xhleIaMnSMvqSoHdhjhzyBoGPuEyqp3/EW/wBsBwHD5qhq0ilLTso3qpipv8RAsfnbDVRSuSvLrJkBO5LXv+zHRntEdOigqBYbeJcJ/TtM/wAVBMgHSxGJwXT6Khqs6pPJljW3v+q/XmRXA+5wkQVwbWtVCW7EQLjy5xSN+Gadl67kE79scOaQWI0Ou2xuft0xOE8bNVcWmeae5WbMdC1kIOkMQIl8J+gwkwZwXsa66g2/qhbAxzqP4WopRbyIvhyDNaZo2daeexa5soxG03T6KEvZ1TzQ51psMzW1+nKGECnzqS7fpO3kdG2HocxpZNAWlnduumwuB998ceuhZ7cifTe48Aw4U3dEHEZ1TYpM8vf3+M2AsRffCpaLPSNC18bX+PwgAemEfpFBMrNTzqD20qT122wupzqGOZbRysRuVCD94xXDfyCLiU+qaiy7OTIQ2YQIB1JF/wB2HnynOtV0raYkD4uht9sOLn1Lo1RU1S7nb4Rtjr57Ad1oapvUaTicN45KcSn1TTZNm6xs4zWIv2vGRq+tsDTZVnjXJzWApborNf7AYklzyIRt/QqkXX4bAk4H/Tia2L0M7AgWK2v+zAuZUHJW19I81HjhvMpDZq6AjbfxG3p0x48K1wk1JmFGf9v9mnEoM9RpAq5XUFibDZb/AEw5Hny8wWy6pIbz0jA6avRFrpdVHnhXMTYjNKNT85B/8uOycJ5iWs2aUrhuvif94xKHPk5ZC5ZPqUFiGIGw6nHI86knUr+i5tZ2UKRa3zwWip0U10fpKOfhCrC/850xXvbWP3YbfhOtjNvfaQgmwAZv4Ymf0xIqqFymQL2JcDt6YcXNYzGSaCRG1aVGsEsbX/ZicOp0VcSjzcoFuEqzblVULt18QZbfljhos7ypjKRIADfXGSwxNJnyRO3vOX1CoAd1sxJ8uoxI5XmtHXKOVKNR6oRuuw6jELH8wipvpk9111H5LntS3hrCJUH65NiP44tEbxtGHQq4axve+Iusy6OaQSwhY3b42AuD9MOZZEIPw1vdeo09dzuPthUzstQBUpoUnci3YY6BZtXUdsKWM8q6jqb4UNNyrC5v2OAJRhdqYfeKXl3/ABACUv28xiJSNtaFTfxeL5WP77YloJFv8JsDgeqBiqdwOuoWG2+CYYKBwBTF1ZFGkWB+EqD+0Yegsbu63vsDbA6pIjkiFCqgnRHu1/IeYw7MgKpqj1C5BVt7fQd+mNKWnfEX1mMwksdutx2PXC1dVuLddjhKkBFI1HtYgjDqRM0hkZbAgm7KAoPlt3wDtlaT4ebeXmAJ0J/WBwTClNMyh0JBa8dmNxbpv3wK6sZCosqBrn5ffBtOYo6eQgkIp2NrEnrhUKC5ScwlBJicsVtoIBIsMPcwMgkdwADe4IFh+wYCTmvOpII1kgnaw/fiQolU31EMF2J89rWscG4aQhBJMKZyqJwOY8UoCtoADdQQDqI+pw1MecsgdJgNZILed/Ly7+mDKRTpiRLspJLkuRcW9P2Yjc5qEWV4wJW0Sm6Wsrix2Jt0ub7eWElsmVpHdamDIpjQ3IZb2CPcX6Hp1H8MNshYrvcMDcE4Z16qhXtpS/QdPU4k+SpUBhta5a3n0xUIdWpCoAFUWXfqMFmRYaTmMxXcAfM7dcCyMI92VwL2sF6/vwqEtzmjEYkjlXrudLD8hgwxLLk7T/iSG6RzG4/rFDD88SGe57XV6NBXPO4QgjULKbgfCBtbcjEbRRCWqWNHcFb/AAtZT8xhNXLE8ghdxblKFkX4WtY39b4hEWRtNpTMksDKupVG21l+HtsOgxx5nRAdV1ubBbEj6eXrgWprDHFqRH0KdThEue+1vL5Ww/UyxmGJS6Lv4dAAuNza/f5YmjxQmoUys/NaVolIAu0fhubWvYdvPHS6aXkI1M4uxKhh8xt98IhZQDHZBbwoytcW6Wv0x6UJHbQ0g2B1AAdfLbDmNhDMoijV6qWmjkYEX1Ovwgjr+Z64MrLKFRdQ7ahvpt127nA2XvzMwj8N1RWBt1Jtf+OC6+QSmN7OjWuW0dfpbFO3TGpFM6y1XMRTr06bna62HliXpVQSDw6b9gevXAlCWVY1a1mGxt388SccbsqKRe5IJvbCyU1u6KhW8a72Ftt+uFuWCm4ZSBfvfCoo0WMo6kNsdI3AOOKw1hIZ44pWBISZdvK9hvb7YIK02GlBEjOWJ3JB26W6dO+FXJ2ER6bG38MNwRFiQ4j1M15DGuxbud+3zwuNY1W6OLXPQAnAkwqlEQyqltTE9yLYfWVGPK1K4I2F++IpxZmIQ+rdDhcEqRyLoFztc2ufvgiJCHVJUmEjCFSGjZTYhu5/dhmaJCStgQbEW64eSq5kZI2Ym9+9sIsrJbSGN7BgO+B2siQphKnUDcDe3fDJmAAEisDe48hg2QFwG3KjYm3TAdQnhO9yDYG3bBQoTC8yRMNcb3Pe5xy4F/DsT8zgTpt54IQAtuV9CR1xSXxLr3MHmPvj2FaG/sj7HHsVCbrVbbwDUAxW3Qja/wA8DxKEESQMyxoCGBOoEW8zvsbYkmy+vkTUHoFAFz+K7XPl8It3x2TKWMZZ8xpEIHXlOwH5jGzWAsBpOJ2QccsjqVkjQeMhdDFtgdj0Fj6Yc1XUra6+TDDrZZULaQVlFYHYXfp5/DvhceWVD7xT0rEDTcE3+9r/AJYo1Gom0X9EDHfddIJ6Ai+2PMfErMLkjqQQMSlNklXI5ElTRqxFwAW39Nxthz9CTxySJVVkSN8VkiLXFu1zhZqDVumcJ0bKJQLrQAHY3uu1/T5YMox8cybBTYA9NR6fv+2HmymMOkb5hf8AWRvdNwe362HGi918MqszIdbFdgT2v/jvgXua4QEykC25R0ciNJGkkZdU3Xcgg9wCDcYY4lgr6eMyGoSJXUlVcAD0sw7XBG/W4374EqZgY9EZGkfGQQbegwFTRLK8gKoRp2uLjbpthOmBBTHOk2RP6bzRCkVJUyorEDoet9jcDpsevnhcudZq7Fp80lLqSS7AkqRvsfLphoeCOORIuW17MtrgW+WEzQrKCVOvfYbAdN8WA6ICA6dyj8vz6uzPLletKTnURqljVmVgSD13HTrfzwKcroHqEneAzHTy2UyG1v8AvX/LEblM70tXDFIRyqhdKqemobW+fTE2t41sUtb13xbZc1C86XISSjy8NIKasrKAqbkRksq+lm2A27YRJSZuZUeDN8uqYtWopPTOj6ewuGILfTBrOWUo1irD9bcDHoyqm7eFgLg6bg/T+/FmRdDIcoyR+KUgaR8i95AP/QToAB57nUflbDKScQVDBaXhDM52AGpm0RJ/tMw6YlZJ5FkKx6woN7jcHff5bYahl5s3OnVmlW4QMbiP1HrixVPNQUhKHqFz0RqV4PzZ2a9xGEbpb+yxuNxgFK+ZJJ48wyjMMvEJs7zRnQD5X+mLD79PDAC0pKgAEk6W28iO2O0WaVUwiqIzWQoh8IWo3Y36m9ycHxQeSDzd3VVqXP8ALggk5kugdyjAD1uR02OBY+IqAmSo92qkVWK86RUCXHUAk7226Y0JK2sljdZc4zAamDNzIYWK+gJj2/x547DYcxGz+vlkkIbXIkbMDc33I6G9vpicRvRVwKnVZ9FxRlcztpnAKqG8Wg3B73BttYnr0GEtxFlAkUPOsQlF1ZmSx8rb40l+WsUlPSV5aKR2eVZoEfWx6k7AX9TfEbPGJnYGsouZqBBShjDC3QdMTijoq4D+ZVGbN8one8WaU8l+6uCQfpjkeeZDrKPmkNksrEN8J+fnjRRyZI3p6h6GWAf1Yko0Zh0JB9L4GNM/O1xtRKxBO9MrAE+gA/b54o1WcwrbQeNyqTVZvw8VQfpeBQ5K6iSbntY49FW5ZI7vBmFNOFUB2Y7Yu0OV06JaV43uB4UpURRt2A6D0ucAVHC2VTxPGcvy5jrDAmkACgG4sAeva/qcCalPoiFFx5qoNmmXxOE9+iUkaiFU7i/XbDU2c5QG/ErGZ4z4vwpPDfp+ri/13D3D9TTwomQZVHLGCpl5AJYGx8h3Hr1xHJwhkMfOjTKMvWORlZfw+lgLj/Hni3Ob9FRtB87qoHPOHwxEmaop67wSbD7YdGfZCy3gzESyHbVypfsNsXGXh+gi0rl9HlkCatUimlBEhtYDzAFhtvhZyKh57k0mXrEsgaNYqXln5nff8vrgOI0ckRoOndUJ+IaGlYGZ6nTe+owPb7kYcj4myt4g4mZA1yHeJgLfO2L1TZbLTzy+4z0wj1gMrU7FgbX2YtYdugwzBlVY0rirjyupBjIDPEVvfquk329fywTag6KhRVSTP8tcDRXRsT5Bv4Y9JnNKVLGc7mwHLcX287fPFi/mzHIytLkWQ6Ta+kAEDz+DD7ezjI5hHPJSZct/1UqGW3rhmobquE481UGzSi1H8Y2tf4Gt+zCFzaiUi8wF+5jew/LFy/5McmZCFyvJWBHxCrFxv373t3xyf2bZTAYk/RWSul9IKzkkW7nbfF6gg4Luqppz3LYzaarhXqLsGt+zHv5yZGLhczgGm1z4saInAmVrp93oskjlTwgsznwg7nYC5wZHwpT09J7tTR5HBqbWWFLqvfqO3Ww3wJqt6IhQd1WXS8QZQZFc5vSrfcqev2NsOU/EPDUzugzOO6tbdLX9R6Y06k4RoIpC0oyRiw1ExUAW7W73JHnvh2q4Yo5tCSihmFw5HuaOF3/M4rijkEQw7jeVlk3EOQKrMlareKxdQbKPtgRs9yGaUKmYBmG/mcbEOFMl16pBTPdri9OI163Gw22I62w5/Nnh5p5gZoFjVRb8AAMRe9ha/l1PbFF+rkpwS3mFjkfEWQKSDXaiovZYyf3Y5JxXkBWy1rlR1Igew/LGvSZDwwHjSBUe62LcixB8vUdcM0/D2WSSSHkUS06C7PJtJq7ALa1rYDU0clNDjaVjrcWcOXJaumUW7Qt/DHpOLcg5aiOrcqeh5TXONgky3K4GWU06ooJtpRenbqMQ1TLlKVEdM7or6TIF/Dtta/bYbj74mpnRXw3AbrNZOJsqhGsGpC+bU7Afsw0vFeUO50VLgW1MBGemNNnp6WRfDmJN+n4aMqi3TpgVqCKA885lAVP/ANzQAC3nfFS3ooKb+qpFPn2UzxqwrLljYArbfDxzPLhJo98jRrdGNsWuLKsoqykVauXVKsfEDAgBHnr7bHHK/hng+LRDJSZYC6Bwiush67C6sbdOmKLm9Eehx5qqJW5bIrGOuge3xWkBthsZhlizaYswpxbrdxixRcN8EzMFOWwqkRuSGYBj677gevlhp6D2e06o4y/LmV2GjUQ2o+hvvgNTeivQ/qq77/RvPoStppnO2hJAf2Y9ELszWI33BO2J6voOFGjMVJlFJHJawaOBVb6sAD+eK/Q06UrNR0tDNGqLqk1vq0li1rb9PD1OJAPoqjqAhy64VSpsbeuOrKoawLG/W+EyhwTtq7bjCIDeQAJb5jFiSEsiCusi80FtyLk37euPNNHBA88rhY41LNte9gdv+OOVJcSWBU6jY4iOIp5fd4I9TBHl3ttYjz++Cos1PAVV3CmwuV74Z4cyvNMqp87zWikrGmlaGVI5gnu6gDSFUXFyASetzfFd474SoMpr5GoZZZ6SUF6dmFnAturC3UWOA8lzuooKhDBKVRypkBBsxHf+/GrTpR5vl9PJFSGvV2QkLMFVS3xEgi5IH6ovf0649WKVOmyAF4njVMQ/e6wSWOHw+BltbYG1j0wiS5bTfX6k9/PG45h7G6UV81DFXVgaGLW7NDtsASV38Qtf7G4wJF7HonpIamKomkWRVbaJbn5DWBjBqgzC6wwjyIkW8Fjck80lXJNMS8tyjyMd9ha3bCBJGjhk16iOxxq9b7MVpY2nKVqwEnS4UeI37WPbvhql9nEBMQcm0gJP4dmI03G+LD9VgEBwjmXLh7FlytEpssRBsLC98OMY1DCOFgFBOzfCP+ONUofZbDLmaCVmhp225mrwk7BR0tck9LjB8/scqzHUJDVxmRGIQRjaTYdfLv59MFJ6FJNKT6Q9ixxB1cBgAL/4OOKvMmjijV9bm6jV1ONPq/ZdmNExeVRMikKQsg2uDuRbtbf5jEbHwHLT65J3UlDcqt9gftt64ovjcFEMG7k4KguYwSpBWxO19zgOQxMbhG6nctbGp1/BeTiJ7ZtSRThCVieUK7NbbvsPmMRzcCwMVeOonqEkDaHjW4GnSGU2HW7YhqDorbhKgO4WcqU1G4t1N7447Q7khiD640N+CaFI2lLzyKq+KyG48jbTvjsfBeVxNG0vvRZgV0sCAW7jpheodFoFF8bhZwpjZRYXwspCLFlYXF741T+Y6VYpqeKhdkQECVoymomxsW0i/TpvbfCpvZzWJE6x5aspVrACdSNjY9P2Yhd4KCg7kQsrCxGwK7eeHdEC7APb/WxpdFwBJWglaEAKxWwl2uDYi/oQR9MEN7Larkkmm0kDwgA6mP26Yk84UNB+0hZfGI0F11b9r3thQZSl1VrDZjqv1xqyez+o90p9OWUrmRQY3IID2AP7scrPZfmUxV1y9KUkfqMqL9rYJrvAoHYd30gspXlFgWVr6huXxwpGnVWN/XvjVaj2S5oJoiiozzC6xhhewZVLXIttqv8AIHDh9kFeqNL7yttNgSw/hg2vvsUp2GdF3BZM8a6dKiS1v7WAy6A20X+uNkb2P5jfVHUq6qPxCUvb9mE1XsbqEmiRJ4SGa7PzLWFtttPe3bAvc6bBHToQLkFY1YMblbeROOghvCRf1B6Y2I+xnMHlISpgVUNiDKx3tt+phUnsXr9TtHVUoYNpOu4289lOAOvomhreqxtArOCEPh6EnCg6uR4ArdNjjYp/YtWCoUrmFMIzbVdTcDuRtvhyT2NR2T/2j4lks94rfnfFAP6K4DrSsY25ZclgRfucehIU6gq3PUjGzt7IRK+k1IRAt9duu52tf5Y9H7INbPGKtUEZFzo63ANx98XD+imiLSsdQRxgkgb+Q64WsTEsbBbnxY15vZJOpsMxQAkBDp/K30wmL2Q1jznmZsDuQ1k2vbFgPB2SyxpsCsnRUUgb6fqP346sRJK6tNz33/bjXZPZBW8vScyhvckDQb27YZHslqhVLAMziLaNY/Cvte2LdqnZU2geqyaeMKAoLHyJ3wEzAuQCR88bgnsilYj3iuQBrnwxnYY7D7DoJm0/pcg73tF1Nh6j1xXf6JraXisNUAtuC9/nhyPStwo8Xkb43in9hFOCqnNpbad9MG4t13v0w9/yCwFwf0zIsd+hphcj53xYLuiE0RO6wJXaMgqTc91J3GOrKCx2bfqTjfv+QOnDKDncoGmxBhHXzFjj0n8n+nNOzR55LzbWVjANN/lfELjOynAKwSJ5QFVHPhN+nTDkVtQaRSxttfG9P7BlEKac6kBtcn3YG/5463sHbUpizoghf+r2vt9cSSq4BusNQnUulAAD1OCUsqeFLC+wBJP3xuMHsJAkiY5wzBhuRD8O3z6Xwt/YYA+n9O2st7Cm9bb+LBcSOSDzMm6xCLWDcMdYF79P2Ybl5rRMjan8Vwd8fQMfsSpEAEmbyuLWBEQ2/PCqj2MZTPMsENfVqUP4psLH0sdsG2p1CQ/BOJsV821ZZbBtQAFxY7YUJq2HLIqlIIoI2kKo7hbuSLdCOg8xtjauLfYnWUVE1Vl1X7+sZvJHoCsF8xcEbYw/iuvM9WsSxGKGnQQxRmxCKPp3O/1wt7yLwnMw8ADmFPcM5zJLVGhr5FEotyigvrt136YtlJIDUqW7C3bv3/PGRUNYYaqGYFdaOLWA6Y1GnnZUhmsNQYHST645eKpnVqau/gqxLNLtwpqSN01RfUXwGLq51t4hiZziJYlikUFbrYg9sQrN+OAFvf74yQttk7dtF7XuNrjCK2TmaJFbZBYgDcnt9MKDajbY7WN8NzIzcs3VAL3279tsFZA7ZNxvqUhyULi23XCyDpsbq9xupG9sJWWNWIKoxbuw62/ZhZmLaUVUudh0vbDWk80mU6pla4KvIeoJNh98PohcO9rAKCp1WBPkdjthiNyq6iptfdQdrefXHTN+GzEO7AEhE62HzwUK5leWzvywNIHVidlHW59MPtqjy7QTrIfdhsrHzAOBZZWimPisX31KtgOnfBdWVWlpom0tqUt8798C5t0thsU3C5PxEAefliaoqUNEpJ1oXUgqbb3BB28sReXortplsdZ0hdAIb0tibQGKBVpz1tHpVVIAvY9TsPlgKhujpC0o1ajlwPqV7gkLe122te/1xXXeSWTlzL+Io8VzcnEhm8ghiKqwWSRgdIB6Dt8v44BVToR7Fm6MTgWAzdNeV5YywuQoCjvfEiZKabL2iYBm/WVh26DbyvbAMlzGxAJYdLL19Rbp9ceppAs2tn0jodNsNfEpYkJ+ofVFHJHIgUoCdI6beffDMMoTl+O4v0v03w/WBZEiWWxcAgiwIPcHb0wHIyrGC6oHU3Okn7+eB2VOEmyOikWCSaQPywSbsy3C9wf+GAKsxRSmQz61kkJUKTdbm59cdqJptSzRvckhn1WN1032HztiJqqhoa2RHHLc3Yui2UXJuSAPMG/zviG6gkKQq1W7oLKQDpYgG5t0IBF/PAMwZlJdl0KLkEDbe/8AZOA2qJGcMY7lirBmFgSdreV7HoMeirhMhMKLAdAj1aBfVvclT2B9cG1vUIXHopCJkDWNyZFvfYC/039L2w3IlOZXKsVa+p9KtqBsu3T0HS/7cDI4aS6y30vcgrbUQoHy8zgmnkkEhDqoYjstj17/ACAwewUupPIw5qxJJHy0sbkKQbWsNvn64lamBWrkam0NYhZF28CkXuR17AfXDHDEYdpqgEtvpGpLA99t8SazBqoqN9Z28O33HfAPunslOQ08qGB3mDR2to0kHVfsbj1xKJGY3SNo2VXuwcnUt79Ot/pa2GkGsICvjU33a3ph9ZFYgFQbXvpW9tvT6YUnQOSdV/C8TK7OCdgCLfTAczKiaoz+JoJ2Bc/UDr16d8P/AICyhm5ZU7FlB8O3l1w37ukok1KXYkfGVboev+Di1aZVVZw7MN9wAlv22/PBDCJTsNKW6FD+7bDRiKKilHsBbaO359MJLOfChO1xpC3xIVLzvGrFkUEdiNrYYMz6rG25BNjfDjKVsLKSTY+Yw3JBEH8Stbz88G0EpDjC4pZXNhoQ7kXsPpgqOuMN43a6mxYdLetzgV1ijuGfbyw3I6OjWGra17dBgSPBWHeKnIqmCVfw3AJFyOuG5dJVjcXv2xCxvIilGCEKALjY4kEZJkuo0N5kdcWFeolceJJPEmxG1jhUcap8RPqB0wyzuoLEA6evzw5zYmQGS1r/ABAbjEUtun9a+bfnj2GfB/ax7EsjugnYQqSFVbdcPKYZ4GKqisBvcdMByqQSuqO47A4SLxqxUgX822/LFkSlaynDy2gDNGjsBYdRt8t8chl93F4o3DE3trP8MNOxDKQTcgBvFq2+uErMzaii6h0KkAA/lig1FrUxRZiHUBomL/q3cdf8eeDpKmBYFk5b6r+Ilgx9NvL74r9NI7qwAA0iyhje5+lsTSEimjKKu9rkC/3wt7bp9IyEmtMBmWpkBewGlCx+mI/OKnSGSUMCwG56kdj5jEisqTFgj9PCQf3YjM6p49aOEb4bNfzwDd1dSdNkBTuGLhQyre4B7nDkIAcm+kX8r4GQOi3ttqwSrubhB0HQ9Dh5WUSivCrEA7PYN88MckwvKTc3AIJO2PRlplN7C/QjbfBCMs8JlLKwsQST07EYNljKFx1NhR1fSy8vTGqB/wCtiKm3iHY4Ipp0rIBVxTOpdfh2IH9+F1Ck0YOk2j3+nW37MB0rtHmdXQlkCELPTjTbwMN/rqDflgI01D4qnPDqcjkiWaoiZi0Qlj6ll7fMdfth+Fy1zE4ZStsOU8etNQa1j1wmqphILsSslrcxdm/x88Wd4VNFpXhES+qwG2FvFrRPHa31wLashUjQtTGO62V/sdj9MPJVQSABGAYDdWBDD6HFaAi4mmy7Ak8THlSG4IsemH3Mq/iLeQn5XvhtW3/Vw7q7bkYmgJgf1SxUtGg0x2t12GEGrdm6Jv0uMefSFBBJv54aVQZOlgcFwiRKE1gCneaTuq7H54bLm1wpB/1jh+njdpOxHTDklNoIVmUX7YTMbpjZeEFJM6krc2+V8JEtmBvYE28jh6oiVSAoBIbYBsCySzIVAkdVBv1v6YY0tKz1Q9pTjVbqoGqw8rnHTWOuoCXTci56nr0wLKGcMzuSSdjhrQbaLg3NsFAQanDmpMVLk+EhvPfpjxnkJUA/O+IkBkYgbEbHHi7XtcfbFEIxUIUrJPMjAagL9BbCPeZSd5id+lsR8Zud91PkcdVXJ1J0HS+IAFRqOJ3UilTMFIEjAHrh+GokU+KRiD3JxGoNQvbc9cOpO5h0gKAd9xuMEGhDxHdVICrOrdS1ug1YcFTIPE8Lhbbk33JxFo5IFl+uDIpZgujVdANxfbDAWjkrJdG6nKeGmq6JnhRzIqDYb+LHaejUrd5RfRq736eu1sB5TAk4kKQTxqqXeRBqsbjpbphFfLGsHLSpeUo5N/ED6C2NDWDZKLypVKGNYo3EzszC4uuHVo4izoWDvqFjfYdiLXxGZbnc8JRGe8Z+Inc9LYkaXMIS06II7Ot9QJuD9fW2I5g6IC4jmjIYY+TpaEeElWIUixvhcUdKRzmQKNxv/f3wKubxRLyyiygm97m2q+9xiRp6iERBjSJYjewv9r4DhjmFNb+RTQSG6hKcEnYEqLDBFPl9PywOVCpudN0VgDYn+z6YTK5WFZ4QylpAQt+x7W7YkqWQNEHQLYG9m3374mgDkqD3O5oBaGmMOiGjhcBjrYwIB59SPPETUpCawxxwwRlDf+qT+GLRUyEROoaNkIOm46epxUszRXqgwZiwO1sAaYPJM4jkNW1dUspjWRmCi11QBRf6YSIamSnkKl9Tr4rFRqHzwzOgRWE7OjldQU98MGdkpJL1BUEWAAwGkDkoHk7obOJKmmo7rI/MA3vJ0+uKnJmVWQFkeRyb/rnElm05WjdOadxsTb7YraJJJ4X2U/l9cW5oKEPI2T8PvtZVWSeUqx3AdhhfuuXQTkT0c8gW4KmslIO4/wBLfpiUWkFLDDLFHIhA6kbH5YEzGF5F5ocbjcAWJxTmw1G0yVFVMdFKXKUMawnYIWY6Rq36k3v54EqtKwRQQoESOxVQAbH64fq1KJdB339MAzhw69CCe9+mMpaCny4JwySLNzPEZGt8TbE+dsVbPsxkh46pXWVld4kjcjbTckWt8rYsbatajYkm3Xf5YoXHYduJ4ljiWNxEgQoLF2uTf53Nvpi2MugqOOmVf6hZROEHc+LfywPISz3UEN2tiSzinjSodECjSQRbpuMR+nS3QWHW2L0wYU1FwCUS7IjtHZm26YrnFWqHNYSpOgxB7W2BDNixu/gFtRse56YrHEhBzddRKnkeHyuTjTh2w8LNi3HhEFD0sskrkzuQCbX7+mPo/wBnNDLW8D5ZmtKjKss85mEe3KkR7Id99tGnrvf1x8yQI6yrZu/QDp/j9+Prf+TqqS+zChjNiI3lRh18XMLH9uOu8ulcLDMbfwVhyKd85oYoK9SmaCFrSsN33IvpO3kDidy2gngpwlakEjhQqlYyuw7W7DEXxCpoJaesVU1qx5IJNu2wbopNu+xxMcPZ7SZxQrPCzaibMpHiU+WFlsAQuiyoS43grrwJHpR6BDEbkabEDe/TELX5TRmeKWmjaIG0b+G5W5HbFrqSnLXUG0k7kYEaOEyl1HiHhNwRti23SnO0oJcrptDK4Ur3BXY4TLTQFBJq0i9zbucSKoGO4B26YZddUWltKqPLBrMd5ChqmgppZhI8Cu1iFZh2PXANbktDfUaNEIXSulB08v8AHniwEIGCg2XsTjkyg6xbYjvizCjHO5FVKbJ6FDY0sQ8vD0wuOhhMbRrALG/ba/8AHE1PABKQAWFu/Q4aMcTRrGTazXDX6edsSx5Ig9wMyot6CAROQi7C2wHywxLS00tOYZaNmZm+ILfSb7W+W/boTg+qQpUySRSsyFECpbYEE3PnvcX+WCaeJJIpNURtL8QU2APW4+uK0CEfGcgpY4QIqU0+ksCwCrYErbp2HUfngeSCFHkZYFD3LN4QLte5O3riaqoA08LMWGlrhSxCjrvbubG2+Haeihk5TMiKRe9u98Q6RyV0y8kiVD0scWkotPyxu5Cpa5O5PXzv98MVSssyySRlUktHGQvxHci5+QOLPHTQxNGNSaVY9Rc2Pkev0w17s82Wx09SIw7adYU9LEEgHv0wvUJ2TnB5ESoWBIXeGnaEuYVujEbLtb9hwb+CGCy3NiNjtfew/PBcUISZrynla2AUAdD9O2E1sNLP4mZC8djqYbgqwZd/IEH74KVQaeqaaMSTLGg0lOq9txgM8mWjsiuiRjQQ2xTSd8TNJCqzSVAcMskusWA2GkC3r0wFVRRPJVpVaGik7KSCRaxv69cRrhKqsDpQiWhaTSlllWzWW+2ET2cCLSSjCwB7ev5D7YkRIIokZLXFjtgZnQR3DlioNjfe563OGtI6LIQeqYSKIyrIbqDZbX6kYTU00SO5NlUte5OH5WhsNaprEisA17DsSPWxOOz6Sw1KjoW1W8jbtiKbIGdARdQB4tj54S9NGqtIqrolN7gbb98FVWg9GJQHZbbdLYHUCKmSIMANRG22JqR6IbqlMNCBIl18IFiLb3w+Uh1KzMdTk2B7+WEyEMpXwsoNySN74SzreMgAFD4bYuFTahBuuIik2ZBdD5X0/wCL4cjpY3F1cbkm4HXDJlF3DFgGPnfCY3tcXKqB0HTFESqadJlGpEpADWDAjcY6KWIzJKVvIqlRpPbY/uw3G6ButxbxHEhRSwO41kgAkjbf54B1lqpnUJTYhVihMe6iwJ7YKpqeAKrrp1AnphumIRUiaS6hD1G5O2+FxvGYwo8Cgm9uth0vgd1bn6UXoU2v18htfDg0swugIvYjA8SJKVd76bliPM4LiVdQBB8LEgW88AQmtdIldFOhN9gQbb98LaJVHQAegx2NEJOolhfe9rDrhRBKrp6euAO6tNCJShDG9+mEvECRsCQDY2t+eHxbTa4+flhIjXWzq9i+173Gw8sDsrCRDGq2Fxt5YXyoZe/iB77Y9ZS7M+s2II0j0wuIRtp1M9xuLtiFEHWhclpQwK2AG9yMJTWLo8aoR0IPUeeHVQaGCyEHfcnfDKqoDK6syqoVSx6+uLBhCumMcoK5Av1/4Y+IfbvSR0PtMzihhstPFKCi3vYEA/vx9wy+GNnYFWIuLb2x8PfygZWqPaxnc7X8coVduoVQt/8Ay4jj3UDh3lRLtqUahZbWAxqkEksmUmoVApEd1BHTbqRjKUTXYBFB1bnGq5ZefKorkLqjs2gW7WxlqXC1YUw4q552uukiLSEgC3TuMV99XyO+48sS1XMJcgy+pZUXmwrIVXpdhc/mTiGexUG469L4wRddFz4siYAGVdIJv1LDe+EVSmyoApYtfpfbCaPwxa3uG6gWxys02SQqobezW3A8r4r50KEmE2ZCqW1argb6bfTCxzSvgLF7XZQL7evfDK26tIwLW06et799jgkvIiLEBct0Fxc/ww/Qs5lcp+ayhiLGwut7W88Oc1pGIYjsCwG/54bjjCqWBCi9iG8/TthSQMbEzAJY/qdfL0wQsFaUxYSjQNTK1xvt98P5nIXqI3Mup9Hb064YCqGuAAgOyp3OF1N7rTqWlC7amA1H52/dinO5IRzUhlDy87QqsfATuuJSOohSi96Z78kaSwNhcm1vnfEZk5c0zSRrpcrYLILN13Fuo/4YNzQRx5ZKpYF2ZCLb3J3N/t+WEO9JPZ6KjyzS1Dcwhi1mZmbqbDa/5YMvLEmuQR6VUFh12738vniPsHRbNYqRexFz/dg2lnaJRpGq37cMAiyWTKYltqSVBcWvYt19BhCMglDKSUtsrD77YfrEYSaoxeJ7nbsD0GBTpWUklGtsL/LocWGyoSQi5SaqNWjZtaELy7gC1+o9QN7eQwEjrFNYxhlIIYl7deu/+OmG5nURFTYAHUNxs3mL9P7sJnLz+KRtTBehN7j1OC4aHWvRTGzEiVgZCw1nUBtYWHl6YEq5YdPPmLOqddUdmBLdz02AvhJkk1LqClRIRr28BJJBv8x0wyI5kk1q0TMDZtYKhgW3ufO1yLjyxNCk2Tsaye9RxHZnsUYuSCLj8yPO/Q4ZsqVglI1SvtpMoJ6b2G9u2CQvPEA1s8cM4MpcDVGQLgfssRhBCl1ugGg/hCUFyOo2J6G2CVSm1RkqGkE3Llc6ha5IFvW49fniTooZamaNPGxLDeMnY2vYm4thqBuY6yrGhjayhmG/918TmRU8iiaomvG0eyspuLnrY/IeWKcLIhdSVMpo6ZIVBdyLXZjuep3w/SBQutlKm1xYm+AnmLVAH4i2W6sR4SPL5+mC40dUuJEDq48T7kr3HX88AngypSBHckansdhtuThbmN5NDDSyjY23P0F/vgWnErt+IiSKwDjxCysO++97YfcLbS0gEhsVVmFtu/X918LKen0m5dQfwpo17MrG5J63FrfmfphisG7R0yaQRvIza2vfsNh088OLVSGOOFKkqNO5MjReuzE2wwzsSBLHUXYWuzA3+uLCpNU0yRoClTKQT00AdfkLYK97LWUsGP8ApCx+lsNFFClgUAA7jDQjHL1vLYeSiwOCiEBeV2UgsWsV+owMZkMjjnBRHYuSDa3W9z12w8667KGOgb2I6YbmF2UXDALYAna2DBlIdYSk+EAHcre+wxxSulmABN9t+2Gxs2979/LHPGrHTuDvYYMNsl6rpasShbp5XNr79MFRyuvXSCdgv8cCqNSglAvlfzw60gBRT0O3ywshOBlEiYTEqwKMpt6fTHUic6rLqAG4GAXEoqA4kVI+pAHxfU9Ppg+jrD4VaxXre/XE2VrnJi/zcn3OPYM97h/zP/mx7FagjhR0irJMVZdNuhBGGGp9OtTIRfzXDBrJNa6gx1ffD/vkVjqaxA3BGD0mUDXCLoeSmnK6Sym24N7YHFJU6iR3BGzYN97S4sQy9Rv0x1WOsMZAFbtbFlzhuEsgOtKAgMsEukm8h6E9BiwZPKdDLMvNHUlz+/EdM4Lg6g2HYndXDwOQQLEHocAb3TaUsO6NgEayvIoA1NcBkO2HKuNJonhkJGpdVh2wEZPxW5rMVbcX6ffBEcumQMSLBfrbthLgtTXCFBT6oKhoWVzY2BvhMYV5w12W/e+Dc9pnKR1KyalZ9DEedr4jqfSJSCHFx3GG7hYn2cjoYJA3i0vc3GCIbLzo2CqDdlAFtu/54GacggaWI8wMd5qCVJGZ9Sm246g4oSERLU6bKhGo+LZu97+eILMb0lVSV7AhaQmGfYHVGxsp+hsfviaqRpbUCzKGsoHfANfSLVxvHKGKTI8TAb2BBF8PrMluoJFJ0Og7JypSoCmaiq9El9YDXKOPL0v5j8+mOjPYF/DrWFO4NtTMNJNr7N9D1scI4dnkqMphE20sQMT9rEG37LYjM+IqKcwU8QzB2YXgUjYfXpfEdBbqCtri0kFWhWRo9zt1vhiqghmFpY1ItZWtv9CNxhvJcuzCKlWTMYjSmw/o5bUU9L+eDZ4SliY/UX74SQTsmjxUZJFUwANBJzbEWjkbe1uxG/3w/FUsLiaJoja2piNN/pv98EU5CEuQAfQdsEGGNwsgYEn8sDdEQhrtIVNwVHS29/rh0FPhOxHmcNmPRI1h8RuSOl8Idz+tt288Ma4wkvAmyk8vtzgFS9wd8GT00ZfW7m5sbHpiDSoAYHxqQeo2Jw5JmZjAjsW03sWN7HCHMJK1UqwaLoyZIElkYOCQLAX6fTELU2Z9zsNgL9sLllkmYvtcncgYHkBGzEgE7Wxoo0oElIr1tRgJwAqlr9/PCZCACbE2PUDHYkkkFhuBvb0x6UOo8S6Rh+kLOmSO53N8JuSbAj5nDhDEbAAXuT54SQBYrcE9jjKUbUoIpIuLAeQxyP4rJfT3ucKCM3w/PphbQkAOGQEbkdzfE1Ig3mnUUJHZiACfnhFw1xcYVCxa9lFxhDAq2/frbzw9gB3SyUtSQVCvYd9sE0sxWZVkeyGwJG+2A9S7A3J7Yfpk1Toist2w0NCEnqrh+klossjWhq08WzsRZgp9O5xV5SXlYsSQSdJ7HBWYxhYAFjXwqLyBbH5YDhCM4EjtpB3tvhrNkDjde2A07AYNomVCCHIv8xfAojAlFgWQnwkriazijioDHKm5kW5QdFH7cQmbIfEoNU/G5fhTVcgk3+WHRmdckfKWZgoGnTquPt2xxamOSLSdLWFt0GGiEncKqaAP7OBm90ZiFIxZvUe7aX0tKCLdQQAPO+JnKp3FCKhtXu7A6iPEVPrbFeekjip4tDmWola6hfLyPriSp3rIcvIdGUooDIT8QvffBFtkkOgqXNdAiteTmDTvdsQlVOJqs6SLnZLNtfDcPMlALxRxKQV1Dz/ZiKq6qpoZBZi2lrq3UbYCEeq6kM1eVZHM0S+AlBvfpiBqqmMoSbovYXxM5xnEdZlSVKxKtQD+I4QHW3nip11Vr8ATcje/XAHxRakid4Km8Y03Pme+Amp2TaMWDGw2BwXSU5jZXREOrr0viXpcvldguhUvext374GJVBRcCzvGsMkh0qb7jf74bqBpVk1FgnUdO+LOMsgCn+kqrdTqXofLbAtflSiMiFyxY7kX/LFPEthMba6pNdoQspNrtsO9sRktmk1Mu1tgMWHOMuMUplKOoWwDXviHqISCDqLbXu1gd/ljKQRuntIdso6R1HLlW5sw7Yp/Fhc8XUNlSSTlooD2tcuf44uqxtfwkG72sD+WKjxLTu/HNCmzXjjfr0UMepJ9MRhkoaghl+qu+YEJKU0hSFANjffEYpYSE7d9rYkK5xzG2svUEbi1sR+sKb6tz12wwCVTiEuY2ABsATewIxW8+tFnmoqCDFsGG3XE5K5ZwwF7HriLz8xnMoGs/MaA2Ftgt9rfW+H0fTWPEu/ZOlQqlHnS1lNvERfrj60/kwwr/wAnFPEkzTaKiYtfqLtsMfK0URZeZqYNe53Ax9a/yaY1X2YwMhBZ6iXUb9SDb92OhUJlcrBkGVoeZUnvULU5hjZClrHe+K1llA+U0kkSpynWUupIFyvkSMWtmOpjoa/S98ROaTLLKYQpIRd/nhtMWgoqzryCpGKQT0QZj1AvbthTDxMVBB2F/PETlUxhCqLhDY4k5WYgkAsLdBiBhBUdUBavKzE7t8sDSM7y2B8I6jC2e6EKSLYGkZg+pyNI7jY4hcAYQhpcAV14o7hL27gk4ddktoBJa2BpXBRXXXcHckYQj65CwYlVUX9TiijBAXZVuPGCGv57WwJMQG8AFhh2aYSOQW6XthICuSupwCu5G359sQWQkoQx3lVUGkN28sLvIgYorAjbbD0dNpC6S22wJNz9cGMinTe1/TBSoAgELOl5dSntfBATUyr4RdRh1oi5UIpG5vcdsNS0rGouCTe2AN0yn3SnBTapVdXGpRvZcEvCtgzkbdARhp5HUKzKRY2OFxSK6nzPY9vlgIKfqalSLDybObfTANRDR1NDPSMWCzIUJBswuLbHscdnWQSo13ZSfFv08sLZEe55e4PXpbE0+KriBOahybBdAA8umI4zQsp5gbSg3IGDKqlHhOvwkbjVYHEYaWNJPCFaxvs9xggAo4krxfwIrAlxs4X1wONFPdlQ6NwPCRvh8xvrJsysykiy9SP7sNLTAjWzIYxdtJIuMFKDSISUlg0/iqyrpIuRe1++OtInIK6D4TcEb2GPSQRDcIzqTvZR5Xwn3V3A1ABBfyuB5YOUpwDrNTYj2KlrgkaQu25GG5VNgNJ03IvfBDJo8aWchQVIA+X78JKEEhbWvrNuoPfAzdNDe7BQp0rHdxdlG4ve5wxKfFcGwJ7kXw8IIzrUMLn4Rq3O+PJQfE0jaSrHUG7eWCDgElzCTYJhtIQagSTte218djZeWA2JCOgj1uqzX8O/hJ3wNHRKKgRsQCT4bt0/hitSbojkuiSN/GiFF07g9bg73wZAwMw0AhtwdhgBacRyvEY3WTpuPth8AFvGZI5I1upt+WIVbZCkEuW3Lbeg3x6MDXpCkAjrfrhjUNMcrXN9iewwU8Fk50T8w91HW3nihCF4JFkWunTyo7AgXPpgynRnVnLHWBtbptiMhMjtIQwF9imrt54kqN3SwaRCehtgXQjphwN9kttZZLEWA3w4rAeEoxucKhjWEMt/CGJsN77k497xYhdJPfphZaSnagN0wT42BDqOg2647C4UnVGxF9jpx2LmNHrkWwvuPTHgApIVj8N/7sCQrCVJIBuoPTyxxOWwuNIG5HUHHIuZ7yzO4sdhc7fbD8qgReEiS/YdMAUQukG4C3Qb7Wth0WaIXQFu3bHKZDuWkDDqLDYfLHQJG0jTpI6sW/YP44igXJI45AUXwi3QA4+Ffbzb/lVzwKCAs4Xpa9lAx90ExglSzNfffHwz7f5oqj2s5/PTgBPeCp3JFwAD+YOLI7qW5w1AKhITcIQPivfvjUOGyBk1OQbeAE39RjL49JljsSbtubevbGl5U8qcOIVBlk5fw+eMtSwWrDekVbHJHDGUgqbiji7f6IxGSJ1UJYBhuTiXrjFBllIsj2WOFFN73vYADEQ7EuQpIFtyB1xgbuug9Ow+FFHn5Y7WqEKKzSA2UFACQpvsTYXx6giZ2ADldJuT6Y7WyRzZnM0cgXwqr+KwW1+vzBH5YICShJhkpqJkDktqC37nTf5jqN/THkYupLOLje97ftx0FVUMxZt/g88djlEhAWyKLLp02II6XB69sPhJle1AOysLN+qvQn+OCIEuGPguDuCx3Hp67YQWY6lkMKm5AHbY7H9hw5qjsoEhbazDzxLK+S6EvKlpQBc6lAuThKKUmR1Om5Nzff7HHFOiNhdwDsDtfCqdUaVSqEORpZrncX+2AcCSqBCncrihlhEhAbl3JFrlr22v5YZzZ6hZljMCGNFI/Ca2nfpv6W64Lo5BR0ttDSeJbsgAI6kk3PTp0xFPOvMdpKnlxyXkL2PhPr59cA4XlGDaF0BA6rY2U303FrHvjzu8cr6VF/1zp8J89/P1wRBTa1YRkmTSCR2bb54FHN8SGTSw2F+5+uGgiEN5XWaOSIKZWQp8DBd18wD2ww28bC2ytuRY3P1vbtjlQAboXMjKoLRKLW636i/bA8jAqFRXYEE+IX37fsOLEclXrTE0sjVQCMkiRghypBOoHcWtjrzCVFZEJZW16ug6dCCMeSeFJS3N1EtcMymxPyP7cKXTqk1KdZOxUkK1zuLdzglUpmWRHsFXTqFnjJDKB3uO/ljzCN5FSUaXXxK4bUsgLG31sBucKLGOWQypGkhtYbBj6ef1GEsFWSMtGwFrFS1yu/Xy3t88Uoh5UYaY4Z5GZkOu4UMR22G5IAsPlgtFWblS1JcLsyqvhLEDz8t8JhCy1DSuQXYKi/hKxQAne/179MFRQRu6GomWNTYtc2uBfYr/AMMQBWnXeU8plQHWxYJdRewFrdvPfbFilqY/c/dor6mNhcbC3Un+HpiMyiEvSqbuyooVXt1H7+n54IUCaUqQdAYNqDEbj9U3Pr5dsFFkYIREBfbnRDSi3LW328hb9mDkKQ6CAnKdhpOpmvfpfsPkMD0ukapEI1aSBqHT12xILLLspjCuTa97379AT/HC3AAprTZFxBEW7FGVRckuAL9gTcY6j0yhQkiKzt0LKL/Ig2P54GhUqkavK69yii/13H78OzBjK5SEk6NKSO5AB7bKQT27i+ElNBRValHyZIJK5fFuiIzBm22tsGP0wzTTj3RI0qKgAWAj5Ra+3e42+uOxpNy3IkiCxeHTHEEtba3S/wBzhu4VL80kkX2uSPp0GLCubJUkxOxplXsDa5wwZ313Kb9L+eCBIzLu7EWsdS7+m+Bna5IJ0+WoXBwbbhZ3kpMkpOopcHyt1wMzs0igkLsNlH78EuSPCAQbb2bA7TASBSmx9TfDWgSklxhJldwgXTYeZ3OGUV3lXVsCbX26ftwtmYrqANv9bCotYe95GYdNsMgQlzdOtteMsXHW9+mOrZbt0UkEWx5nW4DRHWSPCn7xgaSaWdtCHlRow7jxdO/bythESE/UAiKuS8AViRvsxFsIoiJZDGAxuNjcDfDlVUoiBOUsmk2Go4YpKiWdrGJIt9lA1X+ZwJaVNXijvcK/+xF/4649gX3lP7Uf/mx7AWRyeqH1Ksh1MBY9Tcj5YZmdGdtL39NhvjlUr8zbQFB3F8M8tyfE63A2FyMbCOaTrIsiAumMki7W2364TdgnxSDy6fsw5GsiwFGKkW2NyCPyww0UzWLWuNr3xC0HdUiaeodj44wAo6jBUE0bPdTv5HviOaF45LAte3c4cRWDDYNcbjXYYB1Pmia8iyk1lVjpLWt2I2wQiMzBlNrDax2IxHwa9YO1r9Ab4PimAI1AgdsZ3CFop1BzTdVC00ADhi6NdbdLYjZIgHU2ZR5nviY13sy/I98MVFMsnUhB3Pf5WxGmLK6jNVwgQqgERuAT5re2OLGoUozFrixFtjh4U8iMWUq6g28Ox+2G/wANj8ekqd16H7YO3JJggwU/YmAoVBPTY4SiDQwJuVNwCO+F0jrr0hgS24u2ERKTUkFhZul8Nb3mQlO7rpQKLDFmlVDOUWCqTmotrAWFmF/Pv9cBcLVNHnAlpcseNoOZpUwK6CIDuWkYlm36gdSMSFdTDmrMQspWe3L6/huArW+tj9MD5ZK8WcaWyhIMqilSJCAFE2q+om30Fz1vthLXGITyIIPVWiokkqqDmyCRYQ3LhaRgDMALFwOpF7C/2wy5Jy9CbcxWKknpa22CKzXHS0omkiLPCLRL0hX9UX+W+BHVgmlCxF7kdcaW0gRKQ6qQ6ENDqIYCVQR1BO5wfSNTTIRLNoPTcd+2G6emhqpXlIReTGXMXQMB2viTy6uoo6bk+5xx6zsunWN+9zvfAmiFbapm6j6mMKbREMt9yOmApIiSATsOuJXNF5VY2hUSORQygbg9vpiMqmjQNdTrO9vMYU5sWRl0pi34liCFHlbfC3I0kkRi9xsb3w3HONNyF8XQHrjnLXQFAW4PW1sCoU4OY0euMEKuxwww2tYD1OOgsNSqRpI++EFiSQMaBsEpSNNHEYT/AEjQwGx8/TA7GRxo1Ar12G+ENExU23+uHYSS2lri22D3QEwkSRMimNwB33Iw2I9Sh9gBtta98GVUas17i69+1sDXVUIW4I7jCH042TdRXItcZ6nr5/twsNJJqN1Fz8QAwkM1r6rkHceeFSN4bgC42K4Vp6pgdaEqMlXbwnp9MMMzC97b4dW530kHywhRqGzX+ffDqfNLeF7SfDbyw5TSPHMrRizAWsDhBUabG4I6744rguo1bjpthwSnKelrKabLjFURnnx+IEXBP8cRKHyNje/zw5I/OjjXSxKDdgMGZasReOJkZnD+I6hb5WwYshN0VKzTmihZCNIABtuATub4nszo6eupjMDeeBBa/cHzwBO1PLnIg1M2gBU0rsuGM1qpRzqeFQkYbS9je5W4Hr3xQ3Qu2UfDTzSgin8Vydut7b47l4Qu5qoQyhS1zt22w/ksvLmKTSBImQlrDrfb6HBeae5ssxidUVY1iN7nv6d8MmTCESgKOmnhMdUhA1hiqaL2Hnv5+eDM/NYKRdL+AKA21rnAsmYu8cdEjBkSQBBuCR2x7OMzjmnSOK4dLg2Xa5Av027YNUfBP5NTyzwyVMjoqw2JVh8R9O18V/NpY/epHQmRNRK36jEhWZ4Icrjp4nKyMCrAgWC/PqTiBmkVmGqM/M7YTuUUoqV3kygM0GlXksshfxbD9m4wzTUiu4MjAEjuL4k6taePL6YxMQf7BFxc4do44ZJAGKEEdDsL4otUm6eoaSHk2CRkAi7A3t6je+DhCIyHbxAdCDgNIHkdjbSU8tsFxaNIRidFr3tcjAkQjaZRMRgW2sIlxsT0GOoKV3bU2ogFh5DbbAk84W0d2Hi8uvl1wNJM631R6g2176bj54Etuj1QgcxpnlqWppLFpP7W1/virV1EKdmRpBqDEC/cYtVewI1qvbzv+eBqqCKaMM6AbdxgXs1KMfpMqkSRKLOqsCpO4AJOKDxzoqOMMuSSPUrQpGU2Goa2uPzxqlXRLGEZCQochgB18sZh7SEii4zoKaSMctYlIXVbVdj1PbpjMynpJTK1QPAVqq/BMUEQQAAaQbgYDJDDUAdj3GDKsAygqBcqOhwJAGL6dm6kA9MMS5XgCqEXG/e2ITPU15vBa7aYLAADsxxMMCRZrAG42xB5qddZAXsLIwAJ2Nj1vhtG9QBZ8Z3aLiuQi1tWm6sLAjfe/S/fb88fU38noGP2cUz0yKsj1EzMhNr+M7/a33x8tw2E1m0kWFzbH1P/ACcUVfZxDJ51UrKD2BsPttjq1Gd4Li4OpYjotHep/BN0IYDfa9z88R/JSQvUfrEatXUfTEhGIwGdjIWffSTcbYb1RWIOx7WwbRCbUOq6jqSVTHZGU22sLX+uJSGogWAFtiVsBfAQhgE4lG5b9bTguniDQ6iFJJ6suCdEJbAZhNNZgzI6nUAdje488DSuwTYlTciw9O+O1lE76lsmogKbbXUdicOUsDRQrGXuBcb74B0bpjSdkPFJrhUWIU7nvfCmDqGVBtboBfHQrI3hIJ6WwYkZ0gtb5YUTC0tph+xURUQlpFdL3tbbtjzCXkqhZ9ju2kbYkCsZY2su++EAIDpDNftYYMFJLIKaQvoFkB7XwoxsLagCR2W/78PrCE/XPXbCHMlRPHBG1i99z2sL4CrVbTYXuMACSfAIg2SAmxq5hLO6oRbbbHqaILMWJJsLEMd74MTLKjSLzKLeRO+FmkenjklaeFbga3YkbDzPf648v+muREwMS33/AAWsYGvzaVG1DKI9OjcncA7Ww1azBkfx3tscScsEUkfNSopjGTpDFri/lfA/6OkkB0TUz6Ws4LXCkdunXBfpnkfPEt9/wVHBV/opmNXZiAAd97+WHXVbMyruD/auMdijhNz77QgCxOiXt0/eMPiGN1Z4qmmYILyWe4HzxD2yyQXOJb7/AIKxgq30UJMGZRYEWHUG488CzIELC5fw7qCNvXEoaCRvxVeEK25YMbWwOKB52Zoqmnk2sSr3t9hih2zyLfzlvv8AgjOGr/QUdPr8Ll7Dob72Pyw0oddRkYarWA1WBxJQUBaUIlbSFmBuqvqv9MeXLS0xgqK+nMotZUezdO4wf6Z5GP8AEt9/wUGDrn5qh3R0U6V1KzdlO378ckhlcuYmBIW+ncftxLy0klP+HLX0aN1Ad9JsceNBMsywPWUQmYXVC51MPQeWKPbPJDcYhvv+CFmArNN2qHZSyanU3UlWsO3bDjiNYbRaidtRI7d8SKUQM3JbNaV5SdJj5u9/K2EHJtU4pvfqRptN3i1ENbzt1wP6ZZJ/vDff8E3zSt9FQ6yQqSFNyb6tgbeRwMWlACoxJsOpJuMT75K3IklNdR6F8Lkk2Ug9CcJnyPlyiOavpYzKLIrNYk+QxB20yM/4hvv+CrzKt9FQsdW4Yg+HfY6Th+CtjD2ZC7D06YlzwxUiQNHNAd72a+PScMVbksZoDfqASo/Zgf03yGf7S33/AAV+Z1/oqJavjk0hlBbqLre3zOHAihVMqHTcKrDoPT0w0KeOMvFoUlHZLk/2TbBUcSkBWcCx6A49TSqsqsD2GQRI+tZS0gwUyo0y6LRvpt1H3wRTvTNKrmFRYb7EY6wAOgxo6E2INjjhijkJ5ai67i5FsGAhc7SiY9LJcKQRt8RF8OxzEuQl2s3iv2/PDVOBpbWhe4O991P/AAxyBzYOFDHVZum2Ki6vVaVJwzFUFxYAeffCzdZGmjFw1u/Q4bhkDKH0gdhbDxkVmADDcW+RxIQkylOwAJBFwfPCTIGmYMhCXt8/z6YakACOX0gX8J8sJp5Y2j8Uq2JNrHfAEJrHyYRrFDso2x6yr4dxttvhsoGC3bUqn4iMLFNTSkO8Ubn1QYS5NEryyQqdJlALbbvbHhV08katFMsgZtIKXYE/MXt9cdSnihJEcUagG9lQDCo+pJ6k+eJZQSgKs1JLLCjAg2uQAOl+uPhr2zo0HtLz1HlSVlq5NRAFrliT07gk4+7ZnvYFCQQRubjHwn7advabxCqqFC5hMCAvm5I6YImWpBZDpVKQ6WUhv1h2xqlEI46CIqAqGMEW37XOMrGkIX5i6tQ8O9yPPGqZW8Zy6AKCx5a3t0vYYyVNlsw57yseZmneGlaMnxQht/hbYHpiJt4jsbEdR1vgqbVHk+XLdDaljvY99I6YBcuSh6Yw6YK3F8hG0klpdMhOm3Ybn54GqERqqeRdfLeQ6vMEAL07DwjD9KOboMbbkWNumAZ6gtI4tYklSbbmzED9mCbcqnu7gCehkVHs1iAbjD6yRtIS11YbC1xuR+eBabXdnVgWvffaww41zY8xpL37bD698N9STrRgMegGy7d/PHOYFYbD074aBKodZABsLq17D92PRllaxO3qO2ChW0zZOTyFo97tpbffYXw/RvedALHfcdfywK5jCWMkavs1nY/DfsAetx1w9S6mZSrhD8rnFFCd1OZno9wkF1j5v4dm6Fe4t64CgjVPCutLEamUCxPlfewwdmsTQwwa5Iy4B3PwgeRHfEdTrLZogwYLv4yQGN79vLCw2UwmESjvHJup0b21dvX0wmaWjaMGUmWWOx1KAWsfIWH8cMTTcxWV7kgeGy7g28+lsDOlvFrLqbi6Hc7+nTB6FQciJkEqRyNUU8tlAARtNiOo36fPAhjYktVamW3W3QDe/XxD95w3MFE7hiq3UrpUjYH6Xv8APHVaoAiEMZYWKuwYq6ix6eZuPzxbWwrJlOQRpq5ZEjFT16lNv1j269b98NnVDJEwZ3iudLA6w3nv5DD1SpErPTSqJQ97B/Ce1n+gF73wzVSo0Kxio1tLdW5TEi59O/8ADBRCA3Tc1O0zkRx9SS/LA1qTbpcXGxP7sJjjn0hEvqbUCTbbe4Fu+3XDkZ8CKsrpGVAMAX8S1rDt4emOJySIVjW5Efhc+IsBa5PY9t8UrSIYI5FEkb6xfdCnlcWHfqeuCIJNJK3cyk2XUF/duPLvhqOlfls5kmjGu68sHWSTex67YksppgJ1d1kOgczUSd99hc7dumLUUvTwvSUIjMmkIp8RIAJPfp26YaogHKrrGtrkal6iw8unUYHqWFS9opLHVcKzGxNr2sNsScCFQUDFhtdbE/Tfti1YF0XSqxKXkOm1jZRY3/ZguOFG0bMraragfv1wPTlbWEXLH9llscFLaSQ6Y7oG3JHf9+Fv3Whuy6jy6ivLCKhtqBvqG29zbbqNx/EkXkZUdWi06j4CxJJ+hxzSGkcakBW1iRp+237cPOkTkvqQCNtwyk3+2FuajlAPI/6RYrHTQu11BXVrBBtY63NrYISWp5a+8mRGF9pFW/3U/vx3lwlWEEiLudOqMj+PX54ZBKLpQqwvvynvv32xUKpgLz1CBd3Nuu2BfFITolD33vfBTEMRdHcAdSLfswHOYkcRaG2vbfDGdEh7pSSrRAHSR5EdPvhp3/EDqNiO56YeM2lFNunTphqeZ2kB6AD54Y0Q5LfELkbMfCpsL9AMKksB428JPnjolZYiSLre5LHAbapJbSR2BOwQjYfT0w0pYRRkVV5SsAWsfC3iAw2XSNtJkHcC/nb8zhDEs4J3VdgNx0874SVUHTfqxIUnrve+3e+Aa2EYKcJAQDWFPQbdMK5nusbyA3AB8VrfIYEiAVQdWogWDXwNm9WzmOkiBMri+kk9fP7YGoYaVdMaih/09J/Zj+5/hj2G/wBHUvp9zj2MV1s4Sn5o4ZHuzRrcXuZLY60MUa6hLFa3nf8APEXNDFzNZhUOTc6QQrD5A2xHUxo6tefFF4S7LcuSG0kgkb9CQcbnB02WTU3mrPeFgG56eVtVsMiSJF/r4dI8jgHLTFTsjpRU8gANldLrftt9cdfMpPd5YloaF1Z7/iwAld97Aiw+31xIcpqAR/vNIy7yKT02NsdMtIWHjW4HXEJ70Y4ZX91ikVRcrHEin6XA/bh/3xdY0U9Mtu3LWx/K98UQ5WHtlS61VJq8MygDcm+PCrpi9lqogT5tgX9II81pKOhZWsP6ntsbWvvuPnhzOKGgSlpZo8vo0cMWL8m+q9rhreQv188LILRdN9L0VJxTAJtIjE9Cp64eTxpc9AehOKtVUFH7iJqNI8uZrNLJCukab77dOl+lsWREMWqElza3xemAcAmMJsksy8w72Vem5x6SKGZDrCuLHHasct1YEgW874TpBGoFgx6YpoVPeQYQNTaCop9JteQAd7dvr1wqZn97HYKSCfPCK9ZU0Tsqs0LeEWuDcjc9MPzgsSdRvv8A3Y00hustU8kiVQ4e7HfYDcWxB/pOtnr6miSJlgjGiJmW2trXuD1I7ee2J3UocFzqXqAOxwEwigqysih1vqTqNG2567nbBOaAVbXSFZaJOZl0UrKsZ5ahlBvpFun+PLCIUKxsVYgm4sO+GuGmabJVEtVdpP1L9LE27DpsBtgqlDQOHbTY9AehwdNwOyCo0jkvIZoaKpiQ6ddiXt28sKpKqCaIQVScvTHoVwt9Ive/zxJZiZZqJhcgW3W4AI7bYgZYwkKsXAZgSQD64e4WSWkhN1Maq0lqskC2jrc/TtgWokeQC73A9MLltzNrm+174SIihvpbc+WMtQRunMJKGjiZzdXXbcC2HlbUfEulx1FtjhxhokVWFiO/fHJGR5dS3G3W+EhsppMJuTci3bHL9QSB57YcmILXDCwtt0wizPYKV9cPbtCS6+y8R3VtgeowWszaSQfFsDYdRhdDSxTRSFpIlZLXRmtqwlqfQpZLqOlw1xiwEJBSGaRksWuB9NsNMnfffbphzSxJ3B+eHESRYy1hp+fX6YMgIGzzQyvZr26dfXCllLMxtf0Yd8JcgXe6i5x1I2nOiKIuw8t8AWjmmhxhdsuxYkX7HHgQvW/2wyouQAx27eWH0kVl+Jum9+uJACmqd02d99V8JUG4A2J733w7qUkXtp7C9v3YXTxCWojW4AJF7euLalkosSPFSLGxiViwA0nxYb94Yz69R1A9bbnBvEaRxzKEVCo6ADc4jox3IvhzQEDjdWPhtaxY6maBDJG+zOf1fUYiKgjVLeQHxX1364lslqJaWkc/BA6HU1xe+IadgWIA1XPU98UBuqJ2SoJ0WQbsQQFsN8FNM9MzIqloJHuyso1W9RgbL0VqiSeRTyohqJAvp8jg7NbI6zTyGRzALPpBGo2tfsOmJN1bdlFV6mOTWjlCJSFKsANvzwJOJ4udNq1eG2/UEnrgmtm1RQxMojKvckKDq264CrasGecQajHILWvYXxHFUT0QVNC9W0jElgik/M4IccyvjjLIFFgLnw4KyJCYajWWBMZKnRff+G2E1B1ycvma2TwKdNrgdNsRoCs9VNQUwfLVdHBZDsPmf2YOFKsaBAkZYHew6k+mA8oWSOkvzHWTy7E+WCpW5YW0t3bcgDZfriyEtrpUXnE00LGKh1CRgA5C6iNxew+V98AtW5nBJEpMRVt3kKdFsBa97XJPl87dcSs8bRvrs5dvhu3b0xHV0U88TxSSxoFIdWC9CLWPU/34U5pCYHKQFQlRAtRT1CywyDUrIdQPkb4EeTU4RdZA6/PDtBSulTJMx/Am0ssbXY3IJJuT8trefph4wGK9gXI367L88XEhGShVRiNRAB8sC1c7BSGuR2F7YNqGCopJdiQAoA33xF17BeZAlpZ1AvGWHhv3PpsftgdgqBlDB1DhX0HUbgE3JIxlHtUmVON6CdkDf0ZGCkduY9umNglChHE8nLvqKbKL7dvljIfaG2r2qwKSZFjWDSqlTbYNpF9huThZFkZurFLqWZla2223mMBwvonFvETcb9sGVkmuoc3Nibm+BFjKPqDeInCoTAbJBL3CsNvltiHzmIGeGOHSxCsTudhtiXfZBa2osRbEXmAhjaGKWQu1ywTcenXth2HjiArLiz+xchCZEYIpUA2IBXe/zOPqH+TfXR/zEjp3lUaZ3t4SACe1+hx8vJLEltSa1uLhiBf643v+TdmtPBQTZYUDSzTmViT20rZR6bE/U46bzEErj4QS5wC3WZn8IG4+fXywHFWWqWikiAcLfSO12t36/TD+luS7RrYg9FHXfCHQyODK6shW2lk6fIjDGEEIqgLTCehdG1G9t+jC2H0YBdItcdLYApo1jrbxu+gL0O+HmsGsgNr3semCQByXIbyNqJFutjhJnjZCL9PTHCjSwMFUjp23OBIoQj6uhAOoX64EgI2kyEsSIH/WNjh+MsVBjc7jod8CkMHDWKgffD0MotcEMR59sIIJK0NIavMVDFT1PfCoVUTEnrbDJlCtYqSfMYdiRpCGAP3tigCjLmEW3S5dLKxFwB2vhGWxE5hESwOkE/LwnDkrwROBIdJPTbDuXywe+Ko1arNbw27HGDNSfMK/8jvuKulBqt9YUlviifygBf2QcQi1x7utx5/iJi948wB6i4x+RMFifNcVTrxOggx6jK9c5uoELAuIeDsx4ey2szSppcoyumzHPclEWWZU7tTxmOqXVKSyr4m1C9lAsovfFg9n3E2Q5fnfGnD1ZmUMOaVvEtY8FMwOqRTFGARtb9VvtjXCAdiL45oU/qjHoK3ag4rDuo4inM8wY+jG4MxpvzMpIw+l0gr58/k+ZdwzmuTUmW1sPB1ZLLk6GopoaFhWnS0bfjMxKtZgpOw8Wk4H4egpKP2DcOZZlWTmfM+K6r3SsSkVVqKmCOaV5dyVBtErKCSLa+uPosKoN7DA2YV1Bl0UctdUw0ySSpDG0rBQzubKov3JNgO+NDu1b62Ic9tMnU5rg3VN26rWAkSQf+H2D5uAN1nHsvzCeT2Z51w7W089JXZB7xQtT1BBlSHQXg1aSR/VMq3BIup3xQ/Zx/N55PZqODI0HEKRQ/p5qNGCin93POFQR4S2vTa++rpje8+zvJchpBWZ1mVHl8DMEElRKEDN5C/U+mF0ea5VV5R+l6TMKSfL+WZPeYpA0ekdTqG1hY3+WF0s9ewVawoGKpmxOnVpIINrgkkxPRWaQMCdlhn8nw5Kua0C87hP9J8yt8C0zjMwebL8T6tNtPp8Nhhvjr+Zx4s9o8WeRRS565p/0MsUTNWc73NNHJKjUDr097X642Wg4y4QrqWrqqHiTKKiKjiM1S8VUjCGMdXax2X1OJWrr6Gly6TM6qqhho44ua87sAipa+ok9rd8PrZ/Wbjn4h9F4LgGxMEHUHWOnnEREwd1QojRpBC+eOJ+TFx1COMJ+FY60cM5aKk5/SNOrT/i83RpIsb3v9MPcenKKH2iZpWxpR5tXSZnRt+i62mkir9QEYVqKdDcx9DY7bMD67zLnmSR0+X1MmZ0aw5kyJQyNKoWoZxdQh/WJG4t1wLnHFvC2USzJmme5dRvBIsUommVSjspdVN+hKgkegw+l2jruqNLcO4w0iAbbiY7si4gwZHIjnRoNjdY/wADHKV9qmcLVz8LLWniatMKVFA7V5JY6dEoOlRfpt2PnhjgwcIRZfk9FnuXV03tCTOS1V7vGwr+fz2JkZza8GixO+nRjdMgznKM/ojXZLmNNX0wcxmWnkDrqFri477j74kNIJuAMZMT2mIe5r6TmmG21QRpBEejMGbjewuiFC1ivm7iGm4kPsy9oL0tXlq5MM/rufTvSuahm95X4XD2t8J+HEtxUOFI+JOPU9oFKJcyqdIyTmws7vTcgCNaUgbOJNV9O+o743sgeWO2B7XtgB2smQacSSZadJ+Zzj+C/UHwU838fzf4qvezSPN4/Z/kEefcz9KLl8IquabvzNAvq/0vP1viwEYUMc9ceVxOIOIrvqxGokwNhJlaGiBCz2oqCaqpAU6RUyBhbydhhIqGLkRlrXF7AG2ET0sq19a1ro1RKfCbn4zhMSRtpdJQCvxX2vvj9i5R/YKMfQb9wXlKh7xlPc1+YBoABawJt1w7TTE6lCnbe7dB0wyqoZRJrRhcNpAt9MPMVQl9je91IsL46KU4Sn4KvlMRcgHc7bYeWVXCgMwJN7gdvPAdNyFARpRvc6Swtb64cWFGJMUo8K9bi5xICozEBFxSppIB1WNrlrYM95RCpElg231xE0iMivzQSrGw6efzwiRY2En48g1Hv1Hpi7FJOpoUws+qQLoZtTAedvXBMdSskWpbqB207/bERlhTZhfv8VgcSFPGzJ/WkKRawAG2BcAiYSE9z1LhXYKxG3b8r4eDzRHqrW7Bf34agEN9JddwbAHY2w6pVkjI29ACLYXATA53VcackHUT9BjpcBAA9jfp54Q4AYHUquP9G+PS3JCs5Num/fE0hVrckzSWA3NzuABc/bHwz7bbJ7Tc7YvzXFW92FhvqPkN9rY+5ZYZTAQgAYDYFe/0OPgz2s3HH+crI7SOKly29rNqJIwDhDCrBJqAeHwVW1Ekl12vsfM41DI9D5ZGyIVDxrYdbGwvjLuYBIhRWA2DAtsf4Y0/JC/u0UxBAYEKoN9I7D9mMdS4XQw/pKZqzpoaJABdKdARv2AwOHG19jbt2wtpxLS08i3sI9JHyJH7sMxC76gAdul74ylpWwESiMvtHqOpvDcnfbAdOkkUh5zRStpFrr0Pcjt3wYsioG8PVSLHARiAcKGGq2xta2KY0h0oKrtk8sMLSAk6GUEh7nb9x+uHY7XKq7XuTcr0FyQDb02w1CmlgJHLC9roLH5b9cOvM58J2CnSLMD97YcAlbrruLsO3bDsToBp0EHtfDQCcsgIZHb/ABbDkZWEg+Iy+u2kH088XyRgRdO1LlkHQqDawO2F0QDVarcDUwHawt1wP4ZlKyMrEXButwfpiUyaKNK1GiLgBrglrkfLyHpgYVhSueMVhp1aS6EG1gL9Ba+/+L4iYSykAbHqAD+eJ/N4pCkMsehZGFjbq308sQy0+okMkpdbgFRZD9MC2AVbrmyRKWWaPVynRj4gzBXA72AG9tsMo8LiOSmcAuNrAEeoBwQ9PMqrKrTRyp4eWx8JBYAnbe9ht88J/Fai0RyvpHTUt7b3tc32OJPeUAMKOdYzKdLRujHWSjajfbba9hhJURzBo2SUhgSO5t2F9vywVTRTvQuUQOQNlRSewNvpfCZ4WjjLKXR2a9nPU+WwNsFZUUHFeONnYOwLAbxjxFj1IHTc9RbCkH9FkIkWVzsyh7jyI7fTbBjRhaYhZSzODtoC2Xt8zvhl4o4kSVkJdRYbm4H7sFCqVzSG0RvqHkbajt0sD0OFxqvuyRBZV0qLBRuTaxJB2H54UsWomRnXYEroHpub/vx6JREULzMF0m5I3JvsPPe/5Yil17Qt3RlUhl3VmOkC29+3niSo4oxAtl0k9CTawHp5b4iVZVZYoZVDbEKV07Hr29MTuVkyLe+rQNzY2OLsoCUNG96yRF/VI1Ajz6HbExTAgM2kykKSPMjy/dgWmSMyMyxnQp1AdSD8sScMIeH8FijHxXK2JIxREBG3e6XBHIdRUxaVsAwB6nrglBKoBVg4Xe/U/wAMN8lwqNfTq6jV0w5EjwOZA8iC5a+oE/nhaePBEU7qVaSpjLFj4dZtqPUDa22EUszmRVnpo9S6iYoB8J8w+ont0t5748hM86yNZ3BDLH4FUHpvsThUwnaQp+i2JQ6SSPCNz0Awt0o2wlGdAXAp6lUfe5jU3+rgH7Y4vIUG0Srf/QAP1thmYKsjMmXNBG29td7ne5ws6Xi0pIUNtwyYtu6F9wULUKxL3eI9xoBG3r54DkD6t2DAeuC5IJU+KS/qTfDGklQOYGse4w0QszpSNF/jUfVjhOtN9QkOkbAsLemHpFBhPwAr57W9cCl0j0GVla46ML3+mLkzZARATUwllcFo9KA3ABw4UYJ4GKySWKkC9h9cdqJhpDavFt4V3xxHsLqoBJ/WNwMMaOqFcjp6hQ/NKgjf4uuGaTMKeaiE8SiVCbKb7HfqLYQskt2YuzlmJJv2vsMOuxCheaFTsABYfbF8lYKXGyyJr5TIRv12xGU0QmzM1zahYkKCegwVKrJC7KxA+K56ftwiIi6ENZhu6A3HphDgXOumNOkEovXH/b/Zj2B7r6f7GPYLhNV8R6FzAVBpZTSypz2UmLmKQoNu9sBUcb0lLFSo4tEgUMdzt3xIysimRR4yxBsTtYAbWtt0wCEhgEUCaUGg6Ev2FvX1GHJMIpNRi3Om5629MMCYLIyPIjFludJ3te1x3wRAyRweJdYvZr9Rt5YFaBY9diqoWJSw3VTvp63tcnbp0xaEptHIAS2pbafFvf53wQqvoDbi/mLAYREkaSEseZck3A7fvw+kTNE9PZ2XYmSRBpK33HztilYHNcjjB0yKbEEDriYrpRNkxjGhpImOnyIPn8sQ8TILRMAQLBv3X6YPCEOY3YmORNVrnYi/5YCo3VAT6by0FM1crQrSAMX5zRxhCotuyr9rn9uLBHsq3RhddW/Xr0xE0dhOkr2d0kXlXW4Sx/biTqJhCY0e9+VdvMYz1LGE+mbSUvMGsYyBcOLEAd8RcM0gXUpYxg+JWB2F+3piQSdZlGknUBcKRY2/jgWRdRGiwt1J64puyp4m4XqxhUUDiFt5IyARvY+eOx63gEl9igO4t2wo06BQ2px0BsTYfQ4ShcUqgA7FlPe+/XGiieSRVYdyuQhi5FhYHbDWYKqmCa7BY3OoXsACCLnD0D+Ij13AxyoiSoilgZBZ174Y42SmC6XwnVx1OTPLT0gijE8lm0/GA58Q9N7/ADOLBVIDQl1fxgggdbg4hOD4o48mjoo4IYmTUrILgbMdt+5/fiY5zQfhFABbS22FURzWh7hsUbSIk+WlNNpDbxE9N/4YjczgCoqoykpdTYdxhNNOeaNNwL2sBbE1Urry+N0LIhBSQ9SL9/XGwussRMWVUTmI6sAxA6gC5xK0MSVtRyllSN3PgZhtiMqC3PAsAu99rYfpaho51lgXePcbdcJe3UIRU3AG6droZ6WblVIRiCSCpuDY2xGzG3iJAHqcHVlbz53mkjJEjk2PQHCI420eAKxkBGltrffAMYQbpj38gowv4iNivlffDgDA9d/LHcwUQyKCFU7X0m+EwDYMDceeDIhCCn4BY3CqQeoPfBcXLjVvHbyUHa+A1uFC73vth7Q+hiwIHW5Nr4sKEpSsdYIKkXwuaQyfFZbdhtgUKxGlVYjzv0watDO6nShBUeK5GG6QlAlRzgayqkBSevliQyiU0k6yC3W2o9sDxUsrsbKACbXNuuJLL6GnR+ZVSFlX4UUi7HE0qaiLoPNo40rVmhkWRJgGv037i2BbnTcAWv54P4gpVWSMRq6ixtr2O/YjtiLhp3jSxLM3lgHNRgynbArqNhcb+uFRsRIBf6DCUdtDeEkj0w0FYENcgg3t0tiggdspyOSkmp1kmmYzrfZz9sR17MWW5+t7YQAGa9jfyOCqABaoCWRFQ/FcYYDCDcoh+ZHlq2lsrG+kjAWtj4uvltg7NGh8IhLaevXa3pgKnKtIB8PzGJKnNPxVTw5U0KICJGs+OTvPUhmc2ViLnSdNgMPzwlINIKyb3Nhe2GUKmMB9RUG9hfrgUSbqqamhq6OWXTJAyaiLWJ6+fTAstFCkpZ1lDbixH7PPE3U0qCaExoT+Hq/EG1v8bYDMMlfORFGCwuWJ/aMWAonadYooY4nSdBKhF9QUddiPPp+eFHL5CxllkViLWPm1tr4PoYZY4pIzHzLCyrIliPlhyrgdHAfVE0r7RkWAG25/PBICSUilomSohilm0gR8zpcE4VWqky87mAt0ULa1scrCKmrYxmyL4UB729e2CZqaJowCysVG6KbD6HBASga2BZMCGCWIMGVmABIHQC2ELTxBtRj8QHQ9QMOU1PeVxdtKEkBu4thxw8kZKLcgWudr+mI5qY02QVQgBUbG/QeXrhNS6mk0mwYrb1w/NAVjDuVQkAsNX5Yh8zgkqEIWpkgU/Fy18RW4vYnptffAFsXUJQVUaj3lSZ1SEEWBbc+gxH1gjaKSasSyy6QDGrX0gAgWG+q97dAL4frY6ie6JORHZfjiudQ/WFmFsRHEWYSUlNIYSzpEpYsVNgbbDvsD13Hzwlx5omkAI+IFkjc6B+HYhrWv36/LzxlfFKRt7T8ucqHMkqmW3RrSMNvSwGFZ7xKc7NLRZXVz0gWbl6zIVLXIAZgOvyH1wBQ0NTJnqyxVq1cVEEAla3iZiXCjruSG89r4WSIRTeyt+YaDVSgdA/bywMoPUbA9zgvMFkjYJONDHey4DPxAm7DsMAbJnqSZCNATurXOIPPvHURor3k0ksAL7bf34nFAZNVhcsbjEXVxsubxaiusxHpa1yzHv3tYYZRHeCz4r90VDGOUBQI2INj8HTG1/wAn2iqHro6qycpIgdTLY9TtcdfruMZbJSVAsXC2Y76SB+w4272KKq8PPUwqOdBV2mU9GBjJ+X6w6eWOrUaDAXEwVQsJIC2KqzFOWAduxtthFFVyVEpXdYlG7HYdfPAb0xqooyLI1izADYXxNQKGQLGuhLb/AOl03wZhjbJkuqPunKZAfEjE7kehx1ysZFlJueg2/bhOhtISM6d97HfAdJNLJmc1PJG8aKoZHY3DC9j8iD59cUDIVOEWRbSuF1oGJHYC+PBNU3MYEgrc44qlJVK3MbDvjsmoLpQlWtY4uJCsGDdA1U/NvchFOyi2+EUynaMC7dBc4JWCNBqkAPoRh6BYxGChKlDtpPX52wBCPV0SKanK6mJBa+18EPKV8KbkHsptjqPGHJBv+7CGZiCzDw33IxTzAR0W6imHTmsAy3AHW3Q4fy2ArXo1wAA3/pOO+AIGOpbC9sM+9umYchI2DcrU0ikWF7iwB3OMOOpOxGGqUm7uaR7RC1sptY4OKnb7Y9iNR5WVdM7na5Nh/DDsbSFgGlY36dMfAv1U5vzqU/a7/tXb+U6XQo3HsCFZR/0r/lhJ5u1pHse9hiv1UZt/mM9rv+1X8pU+hRuKH7bYqqThjLZaWirK00ud0FTJFSwNLJy451ZyFUEmwBxbpectlSZyxAO4GwwDHWVJkWOYvGSNuhJPe1h0G33xswHkxzjCYhlcPpnSZiXf9qW/MqTmkEFZ/n+ZPXca8OcbS8M59U5LR09XRyU8mWSe8Usz8srNySNRUqGXUAeuJjIYfePZxxPJR8JT5Atc1bJBSMDzajUhAlaO142c/qfLzxaJ5qpSCtRIb9BpH8MDVOYVKlhHUubDoQt7/bHTq9gc2dSZSboAbAHffsHFwtpgmTufZN0sZhSBJM+wfFZP/wAn+et7LJs5zCYVGZpwa+XUeW0uWGCVOZEhKSeJmkcFQtrDe+1zid4ozXNuKuE8p4aybhfOdctZTRVq5nRS0kLQxKZXDOVNlYxhL2Px274tkubZggDGrkVb2vZP4YbfNcxCsf0g9x8IKpv+WNzuxueVXipWNNxa4ubdwDT0gMvBv1lA3HUY7s/n61m0nC2c5vkHD/BeeZNXwxZXxFPCZqdXZIqd6eZ4JY5dIuI2dFDWAugva9sP0mU8WVfDXtUg4gyeolzSbL46eCSKBmWueOmeMSxbblrKbC9i1sX4ZvmilmWtMqqASNK3B8umEHNc3YgpXzIL7Axrf7Ww13ZPPXNLf2QuDu7fUHEju/OIuPgp53Q8fyIVp4RhNPwvlUTQmCQUcIkRk0sraBcEdj54lb7Yo/6WzAFQ1ZMLn9VFvb7YWM2rgpD1kwc/DZVI/Ztjy2I8lWcVqrqhqU7md3c/+FaRmdICIKupO2PXxSo81zIqWarltvpPg3HythQzTMNW9XOVPSyp/DCf1SZv/m0/a7/tV/KlLoVc748TtikjNM0KhjV1AXzCL1+2PR5vmIBEtZMSBfaNfp2xP1SZuP7yn7Xf9qEZtSPI/n60ioeVaqp0kAmokG47azhEvwhy5u9rEDbbDJqNWpiZGdjqY7CxJuTa2FyMeSQQVe1wvW4x+gMBQdh8LTou3a0A/UIXDe6XEpXLcqi8w2Bt4V6XwpY3Eli0xQC1rgfffHoZEvG0kRBuL6lBwSHjUqFiaxPQrcY07KhtKYeE9l8R2uUU9sLWE2C8tg6WswsMeNSDYRIAAfEStr/LBEUiujX5gA2vo7Ys7IGul0JqVAhVo9bnZrdzhMry6LvDLaTfSSR+zClRWUhpb2P6oGofPbCkp4UcM5Zw2+7C6/YYEOARvZqMpiKTlohiDlb9Ge5GDFnZWN0cgjYlrYGKRmyqjENazK249cERRqLcxmZQBsu5IwZ2WYIg1k8QsqixtYarWw/TVDPY6iO9r3viMtrlCDQEUXJtvfDkMbpdol1sb2Yta2BICIOKlrjS93sSQbk3x1U1RqPI7HpgOATrGLSMPSw2+uDoG6FmL9sCUwXSASzsBe4/0sfCXttjiT2lZ20aWjapdgLWv4jvbH3c7FJioVt++22PhP25E/8AKhn6glkWumCljfbmNhdQ90omjvg+tUdVHMALDT8+mNay9UWCBY1j8KqdjtbSPzxksYbmr3uRjWMvQrSqxYhjGhue/hBxiqbLoYf0k+jF8up4w2mwe59dbbYah0Rrv8R8u+HJpnbwtci118xhDtpUarXwpplsrS5sPT+k6OoIPUHrgdgI5G0MZFB6qL2x2m8drA2t174dqEYp+CFF9iCfviuSojVZcijLEtrRR6gm+OrHIdLJFGULWYrYaduvTf8AvwmOOyWW772uBuMKjlayEBgbfFp3tiAwq0orQsc1w5kYoLIbWU2O3rhlnvpadtBZrWbbfHuaBq3UMosV2uL9/PHkPMS4W56kkbX9cWFZsE6HIkYyKyxldSOD+tfcHb5fn5YNop0ikad3Kxj4zIfCNt+uI4MzEgOWROl8SVNIskRRbsBsRew3/fiEiVTZhWWdaiojp006iAbsDp0iw6YQaarhi1WVgdirqT9TbDUcvOy5AWZXU6XKncEdL+mOpWz6FVZQV1dgdxhYKIhJq6DTSuvPnW7hQFfVpAW/YbdCfrbA1VG8AaMyvG7gHSo02/jgyeomkZotCOrWNrbgdNxfvv07YjKppXlXnwgbAcuTcCx2Nr7b2PXEMEqxsmMzrJZgbve0VraiArW9Ot+m/mMc1VYVXgnYsV0abBvqSx2+WOiomEZjEFKkgWwZVZlT1sbbbYZnJblOXnUrfUyyFVI+QNvXpgwIQlPaqmN0LSTuguWJQWNul/MjCAjkmSOcMeXqYsmyjzO4+2GnURxJIZq03j0qkRubfs36XwyqyLHemq6plezBdZ6m50m9r99gftgpVRKcWERNZGcxlmGl2JKb72Y72vbb92O8pfekba7NpU6r3v2tfc7Dftf1wgVMzAsaepiSRwCxa5HYHv3sOnqcS+RU/NhknEhskimC+xcW8Ulha1z4fW197YtRD5flgI1VAAdibLZho23G/e9x0GJukovd02DFbWsT09MKgi11VmV76fivg6PeaxXYjpqvfFgBVdR9JDsbLy2G1h2wZGiGneOQsi6QLoSCPkRvhmYn3wxkEMOlhfbffpbBdOpDG8gC6dl7E/PFE2TGhOwOXYjUo0krY/34fEYD6ZHkA69CR+3CIhIdIJIUjYFbi3ztfBplj93sUJAGxTcH0sT1Pz+mFlNao4cxVPLaxaynoD5bX2w3MJZFRpYpCO12Q6j+w/fEkYDJI+mViw3BYXW1/oR9cCVdGEkKsqXJ8Vrb/UYEiURMIWKJFkDSJLBJurprUb36HST5ftw6zkIFaVCOl7sfzOE8tQ+1ksd1IvvhmU00AZbGEX1Cw2O5uLW64trBKW55SmCtcu6huw1XucNMqL4mIHfbA71V2CR2CnsNrj1w245g0zctgGDC4HY3GHaEgunZF1DxqgZWVj0AAvtiPdZDJcrq2w83MJuLDDTuwGkzl9rW6D5YgbBlC90hKF0t4jc9sIrHLwiE77+Y3HythKldRbSb2+2FR7uDbr3tvgkIcmgHK7CwA2GH4Y2kBY2CA/U4VHyYhz2N7djtc4beV5pQqX0g3J2GKRtC5NPrljgCKVuSyrY3AH/DDEEQWpZtIUKu4I2Hl0w9GSsigoDc36/Cv9+2FwoJJ5ZGCHUw6HrtiKSk7eY/PHsO2H+bx7Fo0BU8wyoY+VbUGk1Fvhtbwnpe46Yakcs4bw9LA6bG2HWEjiRFCq24VgoIP2wnQVSy7svYdx6b4KEqU7IpaADUNrkdfLAjgKmtYXcMwva50ki1+vT/ABbCqhiEQEurXFxfYXHpjw5JGjUVDLYWG4bz9cRXIXAq8ks8LShX2CqCfK4wTTxyRusJbwgAXIFjsdvTe2GSjlH0smsAhWaPvbvbtfCSzyxaQ2iQMGBC7XBufy2xSkp6QSOquFMd/wBU2uPnbBEsyRxxabGRV33G4w1GQsrOE3/WPz9e2FPIoptLXVQep7DviijBRohZJlEdwG8d+vrhWYSAVrsX1EIBc+R7Ycy2OKsJBBEqCytfqCOuGq6NnklMcYZgQbMPLtjPs660fNBC771qKMpsWFyLbg9MFsVYh9Nha3pfEVSRU7qVLtHL8bqjAMCSTc/W+JOGlY30ys1gAGYd++2KcRNldNxCdgZix76uoPp6YSADzroF0tqHXocDSR1ivpEmlgOq388Oz60qImd7tINLgHFscAVTwXNTbqVnH3NtsORv47Dc/tw27p0JYkHa98djZeYfCCxI3ONBEhZRZJo3ejzZ51ZIYlAn1s17G/iuLb9B6b4nakxPDqiKlQQQVtbcDFczyBpMqmkSLmSIrKosN1P+O/lieyYqcsREAaIJo0qNQUAbHUNt7bD54GlzRPmUwlwbjb64sWWuJKMoxJVRrHUnbrits+997L2GD8rq2jl+Ky7g+gPXGoLMZXM7pQa4yUqfhsLjY2t0/diOo5YYWl5iMWKkLt0xaIygiYIFJtYC+532/bisTKsVUxbqvT5nFwIQixTKSNrUMCx09LbYKMoWERyJudwRhMzRy2kSNYmCgaen1wizOgK2J6lfTAtCavVAdaRGKINZJF9zbDcQTTYAD0GHzIsqmN1YAggWHQ4FkAUALf0IwLgqYb3T5TqWFj2J7YkKSlE1MGM363S17YjE5jLoI3YbnEzlNRyKcHSvxf4GLaFHFLeiipmKzGLUV8KDb6+uErUe72ABMgJvvcC/bDUrrUzAyFV7dOm/pgx6PllJSodNRso7/Xth1uSEm1kOGpWkPNpGTvZAfub4dFLEhWphjJAvYEXv9MOQTPMxEqlY9JtqW1yO2rviVpGUq0ccC6raSCpba25xSCCVA5rLJmU1442DLs9h9r/niKkQxqbuWN7fLE5UUktGZ5InlVtidfh1Dcbee2I2tKyqrpCEc7OATv6/PFVL7KAxuo7fbStr48SwF1tfBKRhWBtsf1fLCZqZmYdNN9hbCgFNSaiYhQ7ow+WCdast7kH1w0EIGjW1u22EaXQ7m47C/TBXVJxtbN8W1vPC4lNwTYWHUYdgs66kUlt9S6Nl9b4bZiHv26Wxe6uERJLpTRG5K6fIA4GSRlU+Ib+eOtpZVt2w3pVQdW9xixCueqk3rJJAoMigMAoudxjhglpFeQTxsPhYBiCL4GpYDKq7HYX2PXHKuRmiRGQKQdjff64IIS7op3hvmSU8x5put9LG59O+FTyAVLuzNIsbspve5sbdOtumA8lZ1gR4HZmJKlR+eBq7Mm9/kurRorlSqm99+v1NsGBKrVEKXjVjMdICoCW0WBHpbDc1YuhEMJddtTA2vv54EObqWSKkiDEp+ttfzvhqSXnxHnqqgNqS3QYuIQyOSOSqRdceh1a5FwL/ALDgtFtTIfEHIv12OIVrJUwWLltmcgbi57f34loKmSV1ZVKrp6Ha/wDi2JCvVG6arC5jtYeZHniIqS0kZG488StTMkl9IGoEBgfLEdKimUi6A3IsNj+zAlvVUXg7FRb0vOtGdZF+wvfEXn/DtPPRV0CTSRLVR6ZGcFgoPZR+qL9dh1xZtBQhtgQbghumAcxLSQuuiKU3uBISdxhLwETdlmP83sgyeqeKBTLWrH4Q5DFQABq22FzfbrvgClWrp42jy5oY7xrtIwVVYAi5J6WDtc+owXxLSiLiBq28QmryiMFlD6SrC4BG1tv+GIyv5E0EQk94aIvYrGt3l33Cjbbb17Yy2lPiArHWJooKcyNESkYF4SNJ7XBGAAxaZbqStuoxJ5isa0VKIYZIFEfhiltqH22xFRAliDfUenzxbhdE0iFynb8NvF4QTbzGI2tik94gnAUixG3U9P44k1TTGAQoLaj1OwwBWoPeYGNx4HKbHqLXw3DRxBKzY6eA6FyQgppMdnUCx/d+zG2+wfmS8L5kVBhYz6PPUdAu3r8SjfyxieuOTTd2Dg2u3SxGNu9hNRT0uWV1PzVKT1BClyAWk5akhTvtboPQ46Z9ILh0IIdC2HhyCZcqhNSNMrQjWuq9jbcffBcTEsbEbbH0wuj0iliFuijbHWVk1vHCCGPiHfFG5WpvdASFYgsxAVF3IPfHo1RCzKyb7j9v78ImKMioARr2sOuEVL6FAjU9gB9QL4JoVPclu0rQyrzAGQeAkbYTNr5chjGpgmpB2LWNhhioeOGnkLCUbnUe6+RGGIzNUJo1TcyLSSQNKv5j1uPXbBbJO9kVIqrEH8Ktpv4ibX+3TA3OKTPHE1g4DLffY/TbHZDatkiDgxKPiO/ite2/pfpgGGCT3eepDhegTwA29PXA6mjdHwnu2CkEhcVDvq8Mhv8AkP33wRI8qkxqoIC3G/U+WAIKgiN0mBuLHw9Cehw+lWpgHMRmlBPRew6HC6pkhacMNMyjJBzUA7EWI9MI5EWoO0aa7bnTvgZKjwMbmNgdJYjZfocOCq1NGwB0gHUCDsPO+EEHktRc0bosmyjSp3/LC8uSRYgrOWKJYk+Y74ZDswUtqA5hQg7keux6HrgpRyW2Db7k3xTQZuqcQbhLmK2BJHToO+G4yYrDUSNO298eawO7g9wMceRI0Z2Kqqi53vYYJCSvSblJbnUDYi/Y4jaGGGinkS62aRnuR3Y9Op+2HZaoKNS3u3Sw7f488ATlGq3EkiAqFa5Nt/XBgJTndEuWRuTI7E6OYw8x8RxE10y09WHBCrIpB7dL27+uHa2sjSCXUCrAqDptZiTY2xGPODWsjEhFJ5blRax7G+DDUDXXuuyzwBxzCLahfS2wv3xyseAskqEBVVgwBB2sNyD549LTxmJ5eW0hGln02A679dhYb/TCTSU0Tiw0aktIi7m9xZr+Vr4kFHrCTGQ7oaaZNC6g7G9mW2x+mGGrVjo41lnQSCxZk6g2/YDtfD5pVgkcRo5BNl09N+hwPSTQe8mN6fRpl5Sk7ixUk9Oo2GCAQF46pqqzJArJI9mjY3U9SAPP5kYWlarKl38JtoLDY38vzwZPSUtRGkk0CyFGsWOxI6djhyWlg/Ry0yQ8yONtTIEDXF+m+KhEHjqk0c0RCpzlkQXBPkPLbCqWoaWKN2gZEfxAX+Hbp98M1FFHSoohjVQiksF2UjsQew7W9cegqdawvBEjRch5JCCPC4K6R/6vtioRh3REC3Msnobb9bbi+PPPIpViy3sb7m9vLA4YhFcTKFl7N8KsLD94w8xTmAMoK6b2Jt9DfpgYuiEBenMpqEUuxDbRhegI9fW4+2OiKQxhIgxa3Uve9u2DqeJTTxMytG4IYk7EHuNuuG6yWEVCaWSLxEWbZgPP1BG/1xYupMJujlZiYnLKy+EkdrYLgPNigG8baehHT69sR9XUUiLDGtTH42DDSC5bYi1+2+n7HD8FRC9W5RZmAuFGgi++3iODhDqgwn3Vw0nhU3AIPrhtuZFdIwQ1j0O1uv3wRO5Ucwom29gPyHngaaRpCGIKgbsig/a/niQgLhsN0REzzFtal9SghozYqQMLMdQjHTI4OgDY73xHxFoJGkQsga+kaRvft6YVHM5D3MpJ6KPFv2+mKLb2RNdAhyOg1vEb6NVrhStyR2388NU1R4DDKpa9zpJ22xGvJWpEQHWxIKkrue1zbD+XKT/Wro0j4rXufTBEJDHXRyc0MqKFsQSDa9h9MG0T6i4IKi48Vut/LEbU08702qGoMUzDQdNrJfuBbtgml5nOUqdkW+oWGrsP34AjmiJujHmjQgpexF798PR1DKAbKt+3nhM4imB2FtNrW6m3ngeUvCVsQy7HUBciw3GIIKuYRcYkeddRYHqbH1x8O+20PJ7Rc5qikio1bOqsw2YiRr2PcX74+4aea7o4WQM9iL3sB64+Ivb2F/5VM+iiuY46t1VS99HiJP5kn64VV9FMZ6Yjx/BUSFS0qLvqLC2NbgJNOmlixEcdwehsoGMnjsZk8ZAvtjV6WMRpFGt2UItz5XUE/txhqDurfQPeKbnEg0anYsBcD0w1FDPIwYXXc2I7bi/XDtTyuaCXuRtfz9MEFV0oYlI2sSO5wmFoBMyV6mj0RqqXbSLEsbnp39ccLK95GZtItb0wlLEhWJPKc3BGm+3r169sIdi48KqoBG5Hb9uIrLk9zXHwqygG5YkEtfCXkZzsWv064SGYM111WG2k3uP3Y5RkOBqurfFZhYj5jtiQpKeezWFluBv2v88OKbR6VuD3thgKA45gJckqCHOkg97eeHI95Qtrjt2xSk2RMSOwa0RYKLnTuSPQYkYlVSUjKxkmxIHUed+/ywDTXDMWkVAT4Vsdhbpc9d/liRo0Mbtr0yAsdDBeg7A+Z9Rgyo0o7Vy8qEDMytJIoV41ALEG5O99jb6YXy5kVNK3uBYqLg4arpJ0ozNRjU/NAGsHSB38/K2HMpdGPIQFUvcgMQPpbe2+FNEozaxTlnWNA/NuW03UXttffyG2ApqGnBkWQaBzeYNTGRpDe97EbG42sNtsTVPGgADaUt4bgbAD54RLToYnLSNYn4/i+n92IRCuDChKholW6RSFh/0rkK+/bp6k4b5ZILosy6wdTGaxBv0Fxc4kKyOdo5BBHHEpVwjMCxZux62I+x2thyeCVIDNGIEboFbYeV7C5NvTzwbT1QR0UM7c9g8LBl+NgbgeR+XXHVgDJzJSj7alKkEDfoPXBNRHocSMzkvIEXUvhO/QXt1O3zIwtkVE5jLHAw+AkbLfqL4NAEHT04DXifU+qzAvsB8sTWXlDTaAijc7qLC3UYiY4xE78tYEjIvqDXN7k2J6Hvvg6hkldQmu2og+JVFxfcWBv3xBCJSw0oSFchmFgL7j9+OUyhU0+LlqD1JJv33x6NBINJHXoL74TWhYIeUFDF/M/ng1E1TteRvAxJszA3vYg+e2DUkUgKy31DoxFwMC0xPKUh3K2JB3N8PwszS65HJYg+E2Fx9OvXFQrBUhSyQKHZgqyIbBUKgm/mbj9mCVzWjjLiWZ4LMQUd1ZSLel7/UfXESs4PjWFyCDZlsNz264QJJZCpmnYFRZIzIBqNx8Vh+04W6nJkJzagAhSpzLK5fFSz1bygp+LFSlVbcm4uPEARba+GTUqlMyNXpEsbaR7zGwla1vEC4uetr4jo4pVhSV5YD472V1JU2sQBuQMchpQsrPCkWp2Bdwly/zPfA6HK3PCJaaqqTphqUksCfCVUkeosDfAvu8hktLCFk1dQdR++HZnPMtojuvQ2sb/TCo6moAtIyz23UFRcfXrgw0gyklwIQ5pgslwqi/XDcUKx2QGRgotqc6r/U4k0komW0iGNutg37jhRWjIADsSd7XwYfPJKiFFOAfika3phmWCmK3DurBgRptv3sb4l2y+N42kiAc7DzIwJPTcpwrKqt3Ba1zggQVDcISMrsFbfvfbCpljica2VHsWAuLkdz549MsEEsf4jzS32RSQPqcN1spDcsW36D1Pn8hfFbBUAmnckjw2C9C25F8PU4YQ3Q3Zj1JtYYbijaRhrUab4TNUAqVCCx8KDa1vPFBWCuVMwZnaMsApA2HQf4/bjlA1xqJOnSbW7k9P34Gnkc3jQDQOw6k+eHy6RWhVVBOx36W/wCOB5qpul3Tzx7A+pfI/wCzj2LlHKbdtEo8Di+w1LsCL7j545UF5t+ZHz9JMYcnxC4vsO3T8selTUNRVpATv6YcXVqBcWUXXqfn0vhgSkK/jkZTZb+txjsQfmyK6ppX4GDX1C3lbbe/nhSRBpS1wQbkYd0fiJpIKk6vh2+RxIKuU00i+7nSJiBJpJN9XW3S1yCR+d8OqAmqB9a2QNqHre4B87D9mGUcgqNtPS6i2HdLiLQ012JJDINtO9hb69fTEiVUwnW0+BpTGZFDaCoudOq3ffyvh+NggF41ZZb6t+lvME4GDNJG6ho2cC9mtYi4v54fRbvCCdiCLAX74qEYci8okdXcIDrW7AA/F6YLma1S5FgSdrDbAlAzR3MaBtKm3QH6nDlRIWk5aX1owV7diQG/YRjM4S5PDu5CXyi7GR0DWBAJHnh1WUW0alB6/XDRkMcItsbntgmjldqbT4C3QNbcnAEJrei9YqoLF2sLHfHcwSMwosR8W7C5va3lhTl0p5GZFAIF7d8NHsSQFBOk2xOas2CbfRa42U77jCIgQ4ubi/bCUqY2qZaUkCWEKdNrCzdD+3DwAI03sQbjbY41tuFjdAKWVjmilgcBkkQqR5g4F4MikWoKO0pWxRo7EeNTYEbnt2+3lh1brMW8LX2uNsey+VkzZ6dS9n0szXFkuRub9Nxb6/eNZpVl+pSVSEdmeMq0ZN1K9CDj1M6pJsRc7Hyw7WsHWytcqTqYb38vvgaFbMGPnh42lZ3FWCjSRmSUONLRk7duxxG5hScydrMNgLdsFUlfLEiK0cZVb3uOxG+FTV6PIhZUQKrANubnoB8hgxcKuaiKmn0QoxOolQTbvhnLlAnGoGwvc4KrKgN+Be4JJJtsTjsHKKarKpG4AHQYVUdwxKNjdZgJowMToNid7NsScDTwshCsdIvt8sFc7VKEKkEnbbqMN1HkTawsLjcYUKutEaRZ6SFLGxtdcOQykGy3PY3whV1GwJv5YeoZDTz3Gg26i3XDQYSnFFUqB5N1tftY74NiqolDKZXi0ja7alJ+o2wzSzxTBlswYbqdXiv9sAywy8xyjB11bjbr6jDmhKLzyUtPmLNEsJ5em4JZOpGJfJZ6SSLUJ41JcFg67i3riCpcvkNJzYnBJNr32+WC8u98lEyyU6hUW5bvfEIRAnmiquaB6qMvViSJWIVXbYedtsC5vTRl0ejgtddTEEkH5YWgNWVgqfwwb2ZSAPlv0wRTu1KCskqOlujLqJHlioV7hR8FGZ6Jp0hW6nxal2I874jrINnKkC9rX64sstZJHQlIqaFUddABU9/LfAC0LkGJ/A4WwI3BPr5YoNS9iolnpGp1CvaW3iuOmGTA7sNDLp69cTcOX0cUzidkk8NlBH6x+WGZKalSRIopJTJzQrA/CbnttiEIhfdRIZoxoXYMLdMEUtNNNIqJGWkfcbXvidSgpop0SeojFiSCLAdehHW+E1EhokBSFE5lyGG9gL23vfy6HAz0Rz1UVmWUVdCRz4nUWBuBYfLDGkNGxJvYfXBmY53VVcMUFQ5lRbqSDYdTv64FC04p2dJgJY7eE38QN+nqNvviwFTjIslRVDhNAYotuxw3MULjUTcdScJCG+sKDYXNxjjwyGzG1yL2vuRhrRKQXELuXVUcFcup25YO5F8ORwmpeZ5JNKEkgXuThMCQiVTIhYDfcXt9MJhlZ1dSg0jyX1wcKgSVKUFJGjh4VacCPxA2Fr3GG4MuMUTGpcxpt87XvbAksc0EAqVdlEmwsQB63thfvM8sQZm1G9lUfXfFIpHROtLTrKXgklVh5kWPa2HoKqwMbayy9NPTEYyGZWAW736au2C4KKBWQSzOt/iXUB+eLageTplSEUbgvKpIUjz64CRZEkcvJcsbAAAnBFSGWnJopDp6kCxAHlgNqeeaXnO8ihbHSp6m3XFkSktcBBXpneO92sO4IwHM8V7a/I2A6DELxPxTS5LPFDPRVUhnvdo0BVR9ep9BiAy/iXP85qJVyvKkWmiWxqakMiuwaxt5npt2xnfAMLSx8iUnjWOF4qSWOWNIYavkx6bs9h8Q3PUXPyxSM0cvRQpTnkNMUUqI7mx9fPFwz/L5BlWZV+Y1Uc2Y8gqrCyxwIx3KC177WJPXTiAy6nhkmoJ5lZpDIgVE+JiFtcnsB1+ZxkqNutTXS1TlYpQLCy3MagAg9sARKWcm46bYk87C+/Oq3soA673A88Cxx6aa9z8W5xRVjZMMbDdtyrWJNsCZrAoqqQENZ0lJsf8AV3xISKXEakXG/QbnA+ZlmraCMtZlExGoXuDot+w4bh/3gCz42OCUJ7mhVJWUaCwGkg3xr/sSy+JIWreRDKtHOyaip5gJQHbbya33xlUkMxZbSOyE2G5698bd7AqMLw7NIkw5ktUUeFkBHhFr9Nrg46jjcFcLDj0gtXy+uppA8CP+NEiO0ZsGVWvYketj9jgtp3B1CPUB1GoXGK/QCM5vWoHK1QEaF1YaXRQSotbaxZvviXqqadEuAWT0G4+fnbzwZaAbpweSLBIqa2IxnXGwdSLC4BN/nhn32BahozNErqL+NrYdCO8JkWNKobXXSLjyO/XA8MFBUQzPHDCXW6MDFYMRsdiLd8SQqMpt89y2K6y1cEZJBs0i2+XXAtNxHlc80cUWY0heQ6EjVxcn077+mI05Dw9VwipTJ6IaX1hxTqSrdzuL98eq+F8uqaF1bK6RBIgQskSqwI+E3G+3bEMKtRhWITxzABSFNuzdf4YS8kDKIw4UIRsCN8Q1DltTl9GkFNJzFCAFjfVYbW3Jw6UkADGQMxsL6Oh9RiaG81G1HzAUx/RymtWC3NwL2wmVUkUkGO1wLja2ImjlZJY46gEMf7KixxJLKqLrkGmMmy9bA+uE1G3ELXReQDKLjOmMIv3wTDGFCg2O2+BYJL2WyrId7D9YemCIZLkDSwJ23HTC4Ti8OMowhFcXCi/kOuOS6Vtq6dBgdWLOUWyuP7WG5HMxA0+MfECdwcDCmqyJKksLHZb7g4ZmVCbMLg7C5vhlZmXUpBZkNmXAyyoHYagVIOpQ3wEdcWAqLhCFklmeV1SONQHHik74EmhYKDFtIwBIDXB+WJF4ZXHMjkD6FuADYt6A3tgKojmlRNLmKZSTdQLMOhthzUk2QSU8GgmZCFJvYj9wwiqp6Z4XL6Fi1Ab7Ent+eCWpqlEMgmPwj4kBCsenTt8sAyLO1LIk004WM6pBpUgnqAPPcYMJZtZEKAi/EpCtoOltj2OOVUCCRWDr2Fr9d+mA6cTrpQyKElJYEbeIjxAGwBN7nrh1XmaWSOOu1rHAr+FFLG5Nrb79MSIVTKcgdS0qrEQyEhQwsfl8sKjjjgA5QSON7vdACA1/78e930SAJJq5kpfXtZjvfYDocIrFkNMojlUctwsoZVJjFrX/AGfTEhQSF2KZhIgZFCyNJcsbEFL9vI22+mHWn/Dp05Z0zMSXUbp4dQ1D8sDvG8dRTPPW6xzGRNhe7Kb3tfy8++CSdMVtQCHquwv5fPFmFGzcLx57hXKK0bkgkknbELTxJlzJFDSvUpPIVMgIsLsLE4lppUSRYDI1ywVVtYDYn6dDgKIP7sYmERVJHCiMahpJIHT0titM7o21ISainqeRURNAppxGQLSAk7A3UW36nv8Aq4ayx4KmJSZbwSkEc3e2o+BT03NwO2C6cTU9IArAKN7FvuAThgU+t5LJYEhiSAVutitr/O/2wEdU+b2RNVLWtTGKFZadgpsGO7XGkXtcDcjDlWlDMFati1o0e8tgxuLCwPnc9PIYbFPJTQCZ6nQqjdjGB8xcbkb9MDNHVpIX95Z00hQrJYAA77A9cWBZA90OhNuFy6KYe7LHHIVCWa5AW43Ha/XqbXx6mmkSlWVALBwFT0vbe1/8DBCw/gv7zMpQG2nQFuO2FQxKyoDpBbxa1b4flbEBhSqLJ2Cs5upHQLJdgU62ANr+gIth9ZGYFWIY9ASbXHlgcXVixmd9YNwfTC9K7M5N+q288XEpQdpK8zyHXGLhuihSf347HE+pBbxAW8O9iOuOaG1Jdjv26W9MOwDlsRGWZwCQL7t52v8APFgK3O1XXnF3spcMDYgDYH+OCqdJU0iULqJuwFvyx3Q4jiNyCTqcX8K+pPfDpUKGZ1+M7bEG31xRKgELhjUoWbrfw2G/13/PC4w0ZDxhPMsAd8Kp4rSOVbSt9rb7HzwSYoliUgtpJ2+eALk4U5bKRIztGC8iXub7dBgKdNcJWPQWDCwc9j3vg4xwhNBkFydKi1yx/hj3uwWnV2jRW2IWwJAt0+4xAlkSkU1NFHAJZ0Vrb6gD57dDj4m9uAH/ACoZ6I10qauQncb+M74+2JB4AZpG1EE2DFevYemPh72rjVxrmE7RiITyvNy/7OqRjbqf2nvhdX0EdOA8DwVU0gThVYHcb41ekmSWmQbbKguD3C2xk6gBwps2+NKyqOSKlcldPLsb33sUB/fjnVXaQunh26nIo8t61V5SqqEXZ9ySR19MP10saylIdwvn39cStBwxU1dMtUKhY4yo8R3LC3X5YgZaiCerqoIXEscUhTWUAuR3AP7cKBndaHCAvEmX+tGtuwLY5T1KuIhqsXXVvs32whSxk0KVLHyWxH1w7ZkiHNRQ4PRXDb/u7YsWQLx1GYlmbc7i3TBPLV1Aj5XOX4ZXXfftthlIS34oCKxFizeXl1w8CEva3hOLVyvOhVgzMvhJAIvftgiji1GxAJ2wxHrltsrFrC1+mCIFkVA8gCSdb6CLb9LG2KhSUapi5ThWkVkH4q3CkCw33Ha/b1wbE596hSIGR7i4vso8ycAxVghlVkjbmup0yGMkfLVawt5dd8E5WoIkJkuR4Lxnf1sdsEVG9VJZxCVigi0KWLa1u2wO+/niJjJScShdMouNQFjbyviSr545KcNLDPy0v40jJbbqLevUY7HStKVMcizADe9tWApi8K6lzKIoswK/hs9yRqte7DfzwRVVpELSFdAJsGW91HS/zuemI6enKuupXDR3tZrDcW3GEyvJGOZZ9GwcKpJ9Dt1wZbdWKhiE/wA0vKFMcrM+w0s1owOxbz38hjtNpSNl1ytIw2DMSw32P2tgUtLG2lXF2B2Ym32PfDegjW+hUeQDdVUaretr/ngdCmtPCk5yG8zTJG5+DwMSBsNjbufLoMBy0dHFJd6dSVXSyKt2A6gEgmw36YV741RTSQNNOkZUxh43sV6qbbdb3w1zIIA4TU0YbW7WtvYC3r0774NCTCejEYIcR9W2DqCQD8V9ycG0SmOniNPSaNG1gAOttr9LX8sR0E8s06CMqI3+Eag3T1P3viYihWel5ZldLIAzKbbEdBbvY9fX0xFYMqVy2F5JdTKoIuAWF9ttuvXA2YiJ61lBN0Ftj5YJo5y+XK8lJJHMzlmRXB77dDa32PnhmRVkqHYFUZjclVG+DF1DZMJqbQTIAu+rbr6YKXS0ah1Bbov+O+FBBsWCntYjD6KCoABAI7YgCkodIoBGqJEqlbKoC2C/TC3gKoHkVgl76gQBgiSMRkaHLDTuNNtJ773+WEkRaLFWXVuGI2bfpt0wSkwhlKJY3VEBIuwH/wAuOmPU7GMRtc7WYqfsbH8sOMIy7KsdrKTq6r2sCfPc9RjkppQDGkkKtq06Gtc22O4NrX9cK1JhCEMsqjl/i237jb6nfHYJZUJ1XI62Y6j+7D8oq4UBmZJIyDYXuqgnte9h9cMFF96VzKvLCFTEvQn+157fvwwCRdJJhLRonblyqAdRN73IH3wzI8KyaVc39BbCJIk53NCgsRpJtvYYRLGgvpQHvscRrYQkrrVbxyNpnf5A/tw2JGdiWcFu++/ywmRQItRRQewv0GOQrd7KoJvbY3+uLhCCnyoVdUhsT1OGGVnIUsx3uTfrjst+dpIPTsMKVgqa5CNI63GBKMFMysIY44jIA79Qx30jr0wFGzSzvLKvLQMeWNrhP4mwP28scMUlZU65CEDWdja3hv4UHlfqf78OytcaQgHp54pQlch8VQNRv32wUImkPjBFgOptvhqki/EBKkbYMgCc0AKN+pJ2GBhCm/cV/wA8v+Prj2C9P/xU/wBnHsWrlQ5DrIt7Mobe+xtfboMdMbKNIkMgJuLqAB2sP44XUf12PL8Q/wBT+OGBCmYYyGJNrjYEm2FNGWd4mi1xPY2sbW+fcgg/cfPCf+jj+uOVHw/91v2HFhRcSHRpJs4HU3sT64dhQMbBbtba59LfT6YYf4k/1R+zC4/iPy/+XAzdWEfTIZhJDOodWFzcWAPbHqWGapZpaenkeEMyh0jNiVNiLj1BH0x2n/6D/HfD/Bn/ADa3/wCF1P8A+ffFHdOa0FKMrhI6aeJ4xI1jfZtI6sPPdlH1wyHBclFl3csCRfqT+7Ehmn/N8v8A2M3/AKRgL9VcARdMJsuyACMAqet9/wB2FxSgFQrEm/yBw4vSPAZ/rfri4EJWoypJppvdtNyVb4vK2BIgRGASoCkjrfBn/wBXn/1f4YCPwH/WxUBESUlkBq45QPiXSdzh9NrdSb7k4Azz/m36fvwfD1H+qP2YNqW9JK3cdzfYEkfPA2bRuvu9S0vLiif8UEgBh2J+RsenW3lguT+r++E5r/zRN9P2jBFCCpShmWelKBkc9DpsbMNuuEudLkbAjqPPCsn/AKqD/tD+7Har+vb/AFjhjEFTZKjYupVutsFSInKZilxYBTq++Bsr/wAof5HDw/q2/wBc4YlgwgliZn3X63wuHmIx0G+1m+WHYf8ApMIp/wCtPyxTxIRN3T1Ot4mPLLOmwbTtgKpWb3ssoSQ36HYfbEmn+TL9cR3/ANZPzOM9JoTatQmxTMkM0UpVoytjuOw+WFRgAgi1h54mKr/JoP8AHYYipP66X/WGGOSCkBmjcSoy6+1sSOVtT85eeu7+Fm6k3xDDqn+sf2YMi/r4P9bDAUJsrHmE8dFEIoVZfDZbAbDzwRlszLEDIBIGQklT1GI7Of8AnN/+yxIcJfBJ/qj9uGgKA96EPD+BKFqoQqAatTkHT36DB+YLA8KywQa0YX28jgbPOh+WHct/5uH+qv7TgXKA8kIeWzRlmk0R9Re5wmscEPLyyy3t5fn3x1fgk/7TCKj/ACVMEBKBzyGyuUpjhmZuWJByxqAF7G37MDTVEZJVFcSE3ta4X5HBNF/UyfI4Dj/yj/8AFn9mIAkOqmwUdVNNJM2liXAvvucOy1ASMxRiYx2B8T7A2sf+GDqD/Iar/XX9mIio/qG/x3wMBObMpYD6L8y62/xbCigeIabX9Thmm+BcPH4W+YxSYimlkaEB1BsuGLliAV2vfDrfr/LHIv6r64YzZZKhMhJkVpLk9hvbcD7Y9HJJEzaCQDsQTY/TCIekmFVf+SjBFCwnUnXnlqKdadypVSSov0vh2FYDQvA5PM1Art1HlgCl/rh/q/vw5H/WN8zhbjdaWpyBQjspiszC197jCpoiCBywTfwg9/W+FQ/5UPnh6L/LD/rjEBMqOAhACSSnmdFYoCLEqMM1lVULDoUiZgh0jV122GDu9V8/44qvGP8A9G6z/sH/AGYMmAklo2UZltPXTVi5zWpTUMcAOlknLu2oFQRYEDc7dOm/rM19fUyRx0+tqeKnZDO1QW02kVgkYvfe7R9P7WKxwd/k9F/+A0//AK2xd8z/AMmg+Sf+kYS0yJRaQDCrHERppMtzJVRTJDCFfVuSNJKgjvtex+eK1wnR1Ffm1AvJnlCRc15kkCqCxvp3BuAPI4suZf8ANmd/9kv/AOZGAvZf8UH/AODL+wYQ67lq2Cazlg1ZMygDxEemB4200/LUEm/Ud8EZv/lMv+sf24Ap/wCuPywiU0qQiQOARezAi9+n5dcA1q2rKHlKfCkx7XtdbYkqT+rH1xHSf88UX/Yvh2HP7QLLjP3BREqT8vw3UFdYBA3xsn8n6lkfLatqjQ9OsqsImUEg2N8ZNWf810v+sf2jGwfyfP8AmfMv/wAMX/0DHSqbBcrCAFxsrrXZYaetmnyyClglcEXPfy8P8LYLo66sSGGKdi7lbtohIFx6b2++Ov8A86p/rYbn/rpPrh2+6r0SYRMdUAGkEcmonsvxfMDDkskjvI0cQ3Fy3S/TrgWm/rv+6cFT9ftgVBdRlQ2YJEEijjJI332Ax6JqorqmQqdVm2xJt8eGpviPyw6BCU6Z3SUn1CxAsBYYFle89wWUAi9gMOt/UjA/Y4AiVA8gylzcs76bm5IPfHIbvG8UqcxS21x2wxB/0fyxJ0v9UMCQITadQ6pRFKgIRSi3UXB8sOwW1MdWok3+XphqD4X+Rwum/rD8hhJELY18p+oQmz6gtje+OuArh7BjpuTbCp/hX5nCZfiH+qMKKZKYmgWWVZRs2xHzH/HA5pgHMkaxlm6npfscH/8ASH54j4fhb54IIXQkGBIPAsQAQbEbgbX64RVU4kViBpLAFvF1PzwW/wDkbf6uA6n+pX/WH7MGN0BsEPJzUGtImay21XH7yMDyqtbE0ZjQsig2Dg362w5H8E/yH7MN0/V/9T9+CBuhIndIrKNVZdMSkf6L2JwNTU8cLtIkKr+GIhYC5AuR39Tg2p/ysfM/sw3Sf1cv/afuwbTZZqlnJiN3kib8Fo9ABHhUnrbscOrCsxF4+xTdbGw898If+t+p/aMdqP8ALY/mf/ScQXQaoTFTCNEAABkRtSmw2sCPPyOFSKVj1rDG6MQWLWuLeWO1X/OFJ85v24RmH+Sy/TBOEI6bpS55AUtHDGrA3XULH74Fp4XFIIIEEaRqQAOpsPP/AI4RV/1R/wC0X9mOR/DJ/jtgJTepTVXGagRwlSVsdRA3tbf87YPpIFRUci4G1ztf1tgaPq2JSn/yam+X78U/ZNomXJiojimpjTuCUe4Zb77jCKkWk6liPM2x5vglwhv6lsLkpsA3SLAsSCFJvqDG4thMauLJqvGd10i1sen/AKuP5Y7R/wBQn+vgmoavJKjA1Ptba5Pl9MOKCDqG9h1a++PP8A+Qw9/06/6v7sMKzlchRnGt0BYDa97YcSIBFcRFdQO5few6gY9/0GHX/wAnj/1/3DFSpCdpAZ9KENygbtfv/HBciKRpUBtS7b3OBqb+tX/VODaX4IfmcAUxl0mFSupO4tuT0wrW5RwdOk9fEBbbzwPmXwz/AEw+nww/6owJF0QcdktAixlhbxDqDc3Nun5YaQxvYya2dDvaTv62x6s6R4Qn9XF8/wB2DaEJTVfEh1SRhoNNiXubi3Tz9cfEntKRzxAzeGzRgMVFhqub/XH3Jmfxt/q4+Jfan/zzH/2I/wDU2F1vRVUT+0VKSwZQBY3tjUYFBiYqWYMALMf9ED92M1p/j+mNKpv3/wDyjHOq+iupQMOKsXEGa1lFwtFDECryU6AAHe9gOoxTMrjK07ao1EjEm5Xft3O46Yt/FP8AzDRf9iv7BiByP/J5/wDs2/bjOxbanRNwrJ1ZQpHfzGFxRDmsWZpLsWFgdh2GEJ/Wv/2o/wDSMEp/XH/VH7TgwkGZTzBV2LAqSAbC/wDwwllZn/EsQBt32xyl6SY5T/5cf+zOLlGL2T1LFdyWp3jRTtJ2O17+nlgwRqSwLOVRNWo7XJNrX/u8sdg+AfP92E03+Rv/ANgMVKpemYMEUqAFNxY7DbE3ksCe7yyMihpB4ydyNtsQKf13/eGJ7KP8hH/ZLinbK27r0DaioM+h1W4sba/Q4XHDrJeEbOPEoJU29LYFof8AK4/mP2HDsPSP/Wb9pwxqoqVcKblomkIjJWM7N9zthnMqUR8uVGEYLLeXYEqDfTuO52+ROJHJus30x7O/1f8AtE/dgXHvIhcKGloprFkibmEeEFjpJsept0v/AMMD1VGHmsIkUjZQAbqwvv27YNzX+ql/7H/5hgGq/wAik/7dcWbBAuyUbRyNquZXUFrvp0W74ESCBmMn4LOd2cMOnoRhFf8A85L/APi//UcdTq3+u/7WxFaKSnDzbRB10gqG3sfM2G/n26Yl2V6eNQW5ii1mZ92b+zt3NjviIyn/AChfnJ/6WxLZj/l9P/rSf/mWxasI8gRqBGbkAi7t8W+5t5eQ7dNsckSRtDQFVKuNQK6rr3t5HD0n9XD/ANn+/DcH+UnFhQp0xg3UdCPO2CB4Fvp1sbEBSCT98NyfBhb/AAD6YJUiDCVUkKGbQD10nv8AY7d7dsNmUKqqjXQi340Ytc+RBP3wUPiX5D9mAv8A+nL/ANqcWjGyFl5kbHmR3AsFKHr98MlFjdgsSxa2udEdgCTc3AHn3wWv+QwfI/twxSf10uKKElJIVSTrUgDpvhErKbMIwD/om98Ibv8ALDT/ABj54gKArkzeK22+wFiN/nhttjpWJVt1OCJfgXAdX1X/AFxgkK87nmnTGm36x3v8sdMr+FmRQDsdJs2HJ/iHywzJ/V/9/E2UC8QsS8xVJS3mS1/K2B62Q8tRKpEKNdrgX1W2Hrg6L+si+n7cRGadJv8Ato//AFjC3IgLJdIGNKpkYiSRi7qCBv5fTp9MLhh95qljiRmkcgKvmTsBjj/5Qf8AXP78S/Bf/wBJqH/8Jj/bi1SNyrhPiKvzCOjgySsMzA21RFRsL9TsMTMvs44wgi5k+RVAjD6AqsHIPnYHp64+pYfg+uO1P9Ufp+3Ayl6yvmD/AJM+Lv8A7Im+6fxx7H1Bj2JKnE8F/9k="
    st.markdown(
        f'<img src="data:image/png;base64,{PLANTA_IMG}" '
        f'style="width:100%;border-radius:8px;border:1px solid #222;display:block">',
        unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px;color:#444;margin-top:4px;margin-bottom:12px">'
                'CP2 — SAL DE ORO · Salar del Hombre Muerto</div>',
                unsafe_allow_html=True)

    # Grid de botones de area debajo del mapa
    st.markdown(f"**{t('seleccionar_area')}**")
    areas_filtradas = [a for a in AREAS_MAPA]
    cols_per_row = 6
    for row_start in range(0, len(areas_filtradas), cols_per_row):
        row_areas = areas_filtradas[row_start:row_start+cols_per_row]
        btncols = st.columns(cols_per_row)
        for j, area_data in enumerate(row_areas):
            left, top, codigo, nm_es, nm_en, nm_ko = area_data
            nombre = [nm_es, nm_en, nm_ko][idx_n-3]
            n_eq = eq_por_area.get(codigo, 0)
            is_sel = area_sel == codigo and codigo != "0000"
            label = f"{nombre} ({n_eq})" if n_eq > 0 else nombre
            with btncols[j]:
                if st.button(label, key=f"mapa_btn_{codigo}_{left}_{top}",
                             use_container_width=True,
                             type="primary" if is_sel else "secondary"):
                    if codigo != "0000":
                        st.session_state["mapa_area_sel"] = codigo
                        st.rerun()

    st.markdown("---")

    if area_sel and area_sel != "0000":
        area_info = next((a for a in AREAS_MAPA if a[2]==area_sel), None)
        if area_info:
            nombre_area = [area_info[3],area_info[4],area_info[5]][idx_n-3]
            st.markdown(f"### Área {area_sel} — {nombre_area}")

        eq_area = qdf("""
            SELECT tag, tipo_codigo, tipo_descripcion,
                   spec_nombre_equipo, motor_descripcion,
                   mec_fabricante, motor_kw
            FROM equipos WHERE area_codigo=?
            ORDER BY tipo_codigo, tag
        """, (area_sel,))

        if len(eq_area) == 0:
            st.info("Sin equipos registrados para esta área")
        else:
            c1,c2,c3,c4 = st.columns(4)
            with c1: st.metric(t("equipos_lbl"), len(eq_area))
            with c2: st.metric(t("con_motor_m"), int(pd.to_numeric(eq_area["motor_kw"],errors="coerce").notna().sum()))
            with c3: st.metric(t("con_specs_m"), int(eq_area["spec_nombre_equipo"].notna().sum()))
            with c4: st.metric(t("con_fab_m"),
                int((eq_area["mec_fabricante"].notna()&(eq_area["mec_fabricante"]!="")).sum()))

            for tipo, grupo in eq_area.groupby("tipo_codigo"):
                desc_tipo = grupo["tipo_descripcion"].iloc[0] or tipo
                with st.expander(f"{tipo}  —  {desc_tipo}  ({len(grupo)})"):
                    for _,row in grupo.iterrows():
                        nm  = safe_str(row.get("spec_nombre_equipo") or row.get("motor_descripcion"))
                        fab = row.get("mec_fabricante") or ""
                        kw  = f"  {row['motor_kw']} kW" if row.get("motor_kw") else ""
                        fab_h = f"<span style='color:#555'>  &middot;  {fab}</span>" if fab else ""
                        col1,col2 = st.columns([1,3])
                        with col1:
                            if st.button(row["tag"],
                                         key=f"mapa_eq_{row['tag']}",
                                         use_container_width=True):
                                st.session_state["eq_sel"] = row["tag"]
                                st.session_state["pagina"] = "equipos"
                                st.rerun()
                        with col2:
                            st.markdown(
                                f"<div style='padding:6px 0;font-size:13px'>"
                                f"<span style='color:#ccc'>{nm}</span>{fab_h}"
                                f"<span style='color:#3b82f6'>{kw}</span>"
                                f"</div>", unsafe_allow_html=True)
    else:
        st.markdown(
            f"<div style='color:#444;font-size:13px;text-align:center;padding:20px'>"
            f"{t('sel_area')}"
            f"</div>", unsafe_allow_html=True)


st.markdown("</div>",unsafe_allow_html=True)